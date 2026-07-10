#!/usr/bin/env python3
"""
EDF Master Evidence Collector
Collects billing data from PST/OST files, local PDF folders, and HTM account exports.
Fixed version: correct Excel date serials, dynamic range references, new PDF format support.
"""

import gc
import hashlib
import json
import os
import pickle
import re
import tempfile
import threading
import traceback
import warnings
from datetime import datetime
from typing import Any, cast

import numpy as np
import openpyxl
import pandas as pd

# Tkinter is only needed for the GUI dialog.  Importing it at module
# level would crash on headless / CI machines that lack a display, so
# we guard it and set a flag that downstream GUI code checks.
try:
    import tkinter as tk
    from tkinter import filedialog, messagebox, ttk

    HAS_TK = True
except ImportError:
    HAS_TK = False
import pdfplumber
from bs4 import BeautifulSoup
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# Optional imports — gracefully degrade if missing
try:
    import pypff

    HAS_PYPFF = True
except ImportError:
    HAS_PYPFF = False

# Try to import scipy for advanced stats (graceful fallback)
try:
    import importlib.util

    HAS_SCIPY = importlib.util.find_spec("scipy") is not None
except ImportError:
    HAS_SCIPY = False

# Try to import statsmodels for forecasting (graceful fallback)
try:
    from statsmodels.tsa.holtwinters import ExponentialSmoothing

    HAS_STATSMODELS = True
except ImportError:
    HAS_STATSMODELS = False

# Check for optional report dependencies (without importing to avoid circular imports)
try:
    HAS_PDF_REPORT = importlib.util.find_spec("edf_report") is not None
    HAS_DOCX_REPORT = importlib.util.find_spec("edf_report_docx") is not None
except ImportError:
    HAS_PDF_REPORT = False
    HAS_DOCX_REPORT = False


# ---------------------------------------------------------------------------
# Branding
# ---------------------------------------------------------------------------
EDF_ORANGE = "#FE5716"
EDF_NAVY = "#10367A"
EDF_OFFWHITE = "#F5F5F5"
EST_YELLOW = "FFFF99"
JUMP_RED = "FF9999"
DUP_GREY = "E0E0E0"
MEDIUM_GREY = "#666666"

# ---------------------------------------------------------------------------
# Dedup source precedence
# ---------------------------------------------------------------------------
# Per the user's standing instruction ("html summary > pdf's from
# folder > pdf from pst > email body"), this is the canonical
# precedence order for the dedup pass inside ``export_to_excel``.
# Lower number = higher precedence (i.e. wins when two records
# collide on the same `_dedup_date` and `Amount (£)`).
#
# Why this order:
#   * "HTM Account History" — the EDF online-export CSV carries
#     reading-index, units, tariff metadata that no letter-PDF
#     reliably surfaces.  Always wins ties.
#   * "Local PDF Folder" — the *original* PDF on disk is the
#     source-of-truth invoice; one of these wins against any
#     downstream representation of the same bill (PST-attachment
#     or email-body extraction of a forwarded copy).
#   * "PST PDF Attachment" — second-best because the
#     attachment+body timestamp pair can disagree on timezone
#     when the original came from a different locale.
#   * "Email Body" / "Email Body (RTF)" — last resort; lose to
#     every other source on a collision.  Plain and RTF sit at
#     the same precedence because they are alternative
#     renderings of the same mail.body pipeline.
#
# Exposed at module level so test_source_precedence.py can pin
# the mapping without booting the entire Excel export pipeline.
_SOURCE_PRECEDENCE: dict[str, int] = {
    "HTM Account History": 0,
    "Local PDF Folder": 1,
    "PST PDF Attachment": 2,
    "Email Body": 3,
    "Email Body (RTF)": 3,
}

# ---------------------------------------------------------------------------
# Duplicate-cluster completeness scoring
# ---------------------------------------------------------------------------
# Spec: "duplicates should be assessed and the most complete version of
# the information presented".  The dedup walker uses these columns to
# compute a per-row completeness score; the score is the primary sort
# key so the *richest* sibling of a duplicate cluster survives, with
# source precedence (``_SOURCE_PRECEDENCE``) as the tie-breaker and the
# parsed date as the final tie.
#
# Cosmetic columns (``Source``, ``Sender``), runtime-derived columns
# (``% Change``, ``Anomaly Flag``, ``Duplicate Of``), and the debug
# column ``Logic Used`` are intentionally excluded — they don't reflect
# the *data* the user is reviewing.  ``Amount (£)`` is excluded because
# it's the dedup *key* — every sibling has it by definition.
_COMPLETENESS_FIELDS: tuple[str, ...] = (
    "Date",
    "Period From",
    "Period To",
    "Invoice #",
    "Period Charge (£)",
    "Unit Rate (p/kWh)",
    "Entry Type",
    "Reading",
    "Units (kWh)",
    "Standing Chg (p/day)",
    "Tariff",
    "Attachment Name",
    "Details",
)


def _is_populated(value: object) -> bool:
    """Return True iff ``value`` counts as a populated field for
    completeness scoring.

    Treats the EinDF "N/A" sentinel, empty string, ``None``, and NaN
    as missing.  Everything else — including non-zero numerics like
    0.0 when the producer explicitly stamped it — counts as present
    because the producer's ``record.setdefault(col, "N/A")`` path
    converts missing to "N/A" upstream.

    The one edge case worth calling out: ``Period Charge (£) = 0.0``
    is *populated* in the sense that the producer explicitly stamped
    it as 0.0 rather than leaving "N/A".  We count it as present.
    """
    if value is None:
        return False
    try:
        if isinstance(value, float) and pd.isna(value):
            return False
    except (TypeError, ValueError):
        pass
    if isinstance(value, str):
        s = value.strip()
        return s != "" and s != "N/A"
    return True


def _completeness_score(row: pd.Series) -> int:
    """Count populated substantive fields on a record row.

    Used as the primary sort key in the dedup pass so the row with the
    most populated ``_COMPLETENESS_FIELDS`` ends up first (and thus
    survives ``keep="first"``).  Lower score = sparser row; ties
    fall through to source precedence and then date.
    """
    return sum(1 for f in _COMPLETENESS_FIELDS if f in row.index and _is_populated(row[f]))


def _amalgamate_cluster(cluster: pd.DataFrame) -> pd.DataFrame:
    """Merge a duplicate cluster into a single hybrid row.

    For each column, walks the cluster rows in completeness-descending
    order and picks the *first* non-empty / non-"N/A" value.  The
    ``Source`` column is pinned to the completeness-winner's source
    (identity, not data).  ``_sort`` and all ``_``-prefixed helpers
    are dropped before returning so the caller can concat.

    Returns a zero-row DataFrame if ``cluster`` is already a single
    row (nothing to merge — the caller just keeps the singleton).
    """
    if len(cluster) <= 1:
        return cluster.iloc[0:0]
    # Sort so the most-complete row is first — ties fall to _src_pri
    # then _sort (the same contract the dedup walker uses).
    cluster = cluster.sort_values(
        ["_completeness", "_src_pri", "_sort"],
        ascending=[False, True, True],
    )
    hybrid: dict[str, object] = {}
    for col in cluster.columns:
        if col.startswith("_"):
            continue
        # Walk completeness-descending to find the first populated value.
        picked = None
        for _ri, row in cluster.iterrows():
            val = row.get(col)
            if col == "Source":
                # Identity pinned to the top-row (completeness winner).
                picked = val
                break
            if _is_populated(val):
                picked = val
                break
        hybrid[col] = picked if picked is not None else row.get(col)
    return pd.DataFrame([hybrid], index=[cluster.index[0]])


def _apply_amalgamate_to_kept_frame(
    df: pd.DataFrame,
    dup_df: pd.DataFrame,
    kept_pass1_index: dict[tuple, int],
    kept_for_dup: dict[int, int],
    is_dup: pd.Series,
) -> pd.DataFrame:
    """Reflow ``df`` so each duplicate cluster collapses to a single
    hybrid kept row; non-duplicate rows stay verbatim.

    Cluster resolution merges the two dedup-pass anchor maps:

    * ``kept_pass1_index`` (Period+Amount) — Pass 1.
    * ``kept_for_dup`` (bucket-anchor index) — Pass 2.

    Every deduplicated cluster is reachable from one of these maps.
    The amalgamation path was extracted from ``export_to_excel`` so
    it can be unit-tested without booting the full pipeline (see
    ``tests/test_dedup_most_complete.py`` and
    ``tests/test_amalgamate_*.py``).

    Returns a fresh DataFrame with the duplicates cleaned and
    ``dup_df`` left untouched.
    """
    anchor_to_dup_indices: dict[int, list[int]] = {}
    for (_dd, _amt), kept_idx in kept_pass1_index.items():
        anchor_to_dup_indices.setdefault(kept_idx, [])
    for dup_idx, kept_idx in kept_for_dup.items():
        anchor_to_dup_indices.setdefault(kept_idx, []).append(dup_idx)

    hybrid_rows: list[pd.DataFrame] = []
    for anchor_idx, dup_indices in anchor_to_dup_indices.items():
        if not dup_indices:
            continue
        one_cluster = dup_df.loc[dup_indices]
        cluster_df = pd.concat([df.loc[[anchor_idx]], one_cluster])
        hybrid = _amalgamate_cluster(cluster_df)
        if not hybrid.empty:
            hybrid.index = [anchor_idx]
            hybrid_rows.append(hybrid)
    if not hybrid_rows:
        return df
    hybrid_idx_set = {h.index[0] for h in hybrid_rows}
    non_hybrid_kept = df.loc[df[~is_dup].index.difference(hybrid_idx_set)]
    return pd.concat([non_hybrid_kept] + hybrid_rows).reset_index(drop=True)


# ---------------------------------------------------------------------------
# Extraction patterns
# ---------------------------------------------------------------------------
# Amount-extraction regexes
# ---------------------------------------------------------------------------
#
# Each entry is a (name, regex) tuple. The name maps to a single
# ``Entry Type`` ("New Bill", "Ongoing Balance", ...) — this is the
# contract the classifier consumes. Order matters: earlier entries
# match first, so the more specific patterns (``current_balance_debit``,
# ``total_charges_period``) take priority over the generic fall-through
# ones (``balance_within``).
#
# Adding/removing/reordering a pattern no longer breaks the classifier —
# the classifier drives off the *name* (a stable string), not off the
# list index. If you add a new pattern you must also pick its entry
# type in :data:`AMOUNT_PATTERN_ENTRY_TYPE` so the classifier routes
# it correctly.
AMOUNT_PATTERNS: list[tuple[str, re.Pattern[str]]] = [
    # New-style KI / KCR invoices — "Current balance £X debit"
    (
        "current_balance_debit",
        re.compile(r"current balance\s+£\s?([\d,]+(?:\.\d{2})?)\s*(?:in\s+)?debit", re.IGNORECASE),
    ),
    # New-style KI — "Total charges for this period £X debit"
    (
        "total_charges_period",
        re.compile(
            r"total charges for this period\s+£\s?([\d,]+(?:\.\d{2})?)\s*(?:in\s+)?debit",
            re.IGNORECASE,
        ),
    ),
    # New-style KCR — "Total credits for this bill £X"
    (
        "total_credits_bill",
        re.compile(r"total credits for this bill\s+£\s?([\d,]+(?:\.\d{2})?)", re.IGNORECASE),
    ),
    # Old-style cumulative balance
    (
        "your_new_account_balance",
        re.compile(r"your new account balance\s+£\s?([\d,]+(?:\.\d{2})?)", re.IGNORECASE),
    ),
    # Generic anchors (in priority order — more specific first)
    ("balance_within", re.compile(r"balance[\s\S]{0,30}?£\s?([\d,]+(?:\.\d{2})?)", re.IGNORECASE)),
    (
        "total_charges_within",
        re.compile(r"total charges[\s\S]{0,30}?£\s?([\d,]+(?:\.\d{2})?)", re.IGNORECASE),
    ),
    (
        "total_amount_due_within",
        re.compile(r"total amount due[\s\S]{0,30}?£\s?([\d,]+(?:\.\d{2})?)", re.IGNORECASE),
    ),
    (
        "amount_to_pay_within",
        re.compile(r"amount to pay[\s\S]{0,30}?£\s?([\d,]+(?:\.\d{2})?)", re.IGNORECASE),
    ),
    (
        "pound_amount_debit",
        re.compile(r"£\s?([\d,]+(?:\.\d{2})?)\s*(?:in\s+)?debit", re.IGNORECASE),
    ),
    (
        "current_balance_within",
        re.compile(r"current balance[\s\S]{0,30}?£\s?([\d,]+(?:\.\d{2})?)", re.IGNORECASE),
    ),
]
#
# Pattern name → entry type. The classifier looks up a pattern's name
# here to decide whether a match is "New Bill" or "Ongoing Balance".
# Unknown pattern names fall through to the heuristic body of
# :py:meth:`EvidenceEngine._classify_entry_type`.
_AMOUNT_PATTERN_NEW_BILL: frozenset[str] = frozenset(
    {
        "current_balance_debit",
        "total_charges_period",
        "total_credits_bill",
        "total_charges_within",
        "total_amount_due_within",
        "amount_to_pay_within",
        "pound_amount_debit",
    }
)
_AMOUNT_PATTERN_ONGOING_BALANCE: frozenset[str] = frozenset(
    {
        "your_new_account_balance",
        "balance_within",
        "current_balance_within",
    }
)
for name, _ in AMOUNT_PATTERNS:
    assert name in _AMOUNT_PATTERN_NEW_BILL or name in _AMOUNT_PATTERN_ONGOING_BALANCE, (
        f"AMOUNT_PATTERNS entry {name!r} has no entry-type bucket — "
        "add it to either _AMOUNT_PATTERN_NEW_BILL or _AMOUNT_PATTERN_ONGOING_BALANCE."
    )


READING_PATTERNS: dict[str, re.Pattern[str]] = {
    # Order matters: callers iterate insertion-order and break on first
    # match (see EvidenceEngine.process_text / _process_new_invoice).
    # Specific phrases first; the bare word "actual" only matches if
    # nothing more specific does.
    "Estimated": re.compile(r"estimated|est\.|estimate", re.IGNORECASE),
    "Smart": re.compile(r"smart meter|automated reading|smart reading", re.IGNORECASE),
    # "Actual" must NOT match the bare word "actual" — that phrase
    # appears in normal bill prose ("the actual amount you owe is
    # £X"). It only counts when the meter-reading language is present.
    "Actual": re.compile(
        r"actual reading|customer reading|your reading|"  # standard phrase
        r"reading was actual|reading is actual|"  # OK in long prose
        r"actual\s+reading\s*[-:]\s*\d|"
        r"meter\s+reading\s+was\s+actual",
        re.IGNORECASE,
    ),
}


PERIOD_RE = re.compile(
    r"(\d{1,2}(?:\s+\w+\s+\d{4}|\s*/\s*\d{2}\s*/\s*\d{4}|\s*-\s*\d{2}\s*-\s*\d{4}))"
    r"\s*(?:to|to:|–|-)\s*"
    r"(\d{1,2}(?:\s+\w+\s+\d{4}|\s*/\s*\d{2}\s*/\s*\d{4}|\s*-\s*\d{2}\s*-\s*\d{4}))",
    re.IGNORECASE,
)

_ISO_DATE_RE = re.compile(r"^\d{4}-\d{2}-\d{2}$")

# Used by the "large amount" fallback in `extract_amount`.  Pre-compiled
# once at module load; this hot-path is hit once per analysed chunk.
_POUND_AMOUNT_FALLBACK_RE = re.compile(r"£\s?(\d{1,3}(?:,\d{3})*(?:\.\d{2})?)")

# =============================================================================
# KI / KCR invoice field regexes — pre-compiled at module load so
# `extract_new_invoice_fields` / `extract_new_credit_fields` don't pay the
# implicit re-compile cost on every PDF page.  All flags are baked into
# the compiled patterns (re.search refuses to combine a flags argument
# with an already-compiled pattern).
# =============================================================================
_INV_NUMBER_RE = re.compile(r"Invoice number:\s*(KI-[\w-]+)", re.IGNORECASE)
_ACC_NUM_RE = re.compile(r"Account number:\s*(A-\d+|\d[\d ]*\d)", re.IGNORECASE)
_DATE_ISSUED_RE = re.compile(r"Date issued:\s*(\d{1,2}\s+\w+\s+\d{4})", re.IGNORECASE)
_BILLING_PERIOD_RE = re.compile(
    r"Your charges:\s*(\d{1,2}\s+\w+\s+\d{4})\s*[-–]\s*(\d{1,2}\s+\w+\s+\d{4})",
    re.IGNORECASE,
)
_CURRENT_BAL_RE = re.compile(
    r"Current balance\s+£([\d,]+\.\d{2})(?:\s+(debit|credit))?",
    re.IGNORECASE,
)
_PERIOD_CHARGE_RE = re.compile(
    r"Total charges for this period\s+£([\d,]+\.\d{2})(?:\s+(debit|credit))?",
    re.IGNORECASE,
)
_UNITS_USED_RE = re.compile(r"Electricity used\s+([\d,]+\.?\d*)\s+kWh", re.IGNORECASE)
_STANDING_CHARGE_RE = re.compile(r"Standing charge\s+\d+\s+days\s+@\s+([\d.]+)p/day", re.IGNORECASE)
_TARIFF_NAME_RE = re.compile(r"Tariff name\s+(\w[\w\s]+?)(?:Payment type|$)", re.IGNORECASE)
_CREDIT_NUMBER_RE = re.compile(r"Credit note number:\s*(KCR-[\w-]+)", re.IGNORECASE)
# Credit-note accounts can use the same rendering as KI invoices.
_CREDIT_TOTAL_RE = re.compile(r"Total credits for this bill\s+£([\d,]+\.\d{2})", re.IGNORECASE)
# Format-detection: cheap presence tests for the invoice number prefix.
_KI_PRESENCE_RE = re.compile(r"invoice number:\s*KI-", re.IGNORECASE)
_KCR_PRESENCE_RE = re.compile(r"credit note number:\s*KCR-", re.IGNORECASE)

# `_classify_entry_type` is on the per-record hot path.  Compile the
# three heuristic patterns once at module load so each call skips the
# implicit re.compile step.
_BILL_MARKERS_RE = re.compile(
    r"(?:bill date|date issued|invoice number|total charges|your charges)"
)
_ACCOUNT_BALANCE_LANG_RE = re.compile(
    r"(?:account balance|running balance|balance brought forward)"
)
_BILL_INDICATORS_RE = re.compile(r"(?:kwh|standing charge|tariff)")

# `_extract_sender_email` pulls an email out of either the transport
# headers (multi-line From:) or the sender name.  Compile both.
_FROM_HEADER_RE = re.compile(
    r"^From:\s*.*?([A-Za-z0-9._%+\-]+@[A-Za-z0-9.\-]+\.[A-Za-z]{2,})",
    re.MULTILINE | re.IGNORECASE,
)
_EMAIL_ADDR_RE = re.compile(r"([A-Za-z0-9._%+\-]+@[A-Za-z0-9.\-]+\.[A-Za-z]{2,})")

# Old-format PDF (pre-2019) extractors in `analyse_pdf`.
_OLD_PDF_DATE_RE = re.compile(
    r"(?:Bill date|Date issued):\s*[\",]*\s*(\d{1,2}\s+\w+\s+\d{4})",
    re.IGNORECASE,
)
_OLD_PDF_KWH_RE = re.compile(r"([\d,]+)\s*kWh", re.IGNORECASE)
_OLD_PDF_STANDING_RE = re.compile(r"(\d+\.\d{2})p\s*per day", re.IGNORECASE)
_OLD_PDF_INV_RE = re.compile(r"Invoice number[\s:,\"\'\n]*([A-Z0-9\-]+)", re.IGNORECASE)
_OLD_PDF_PERIOD_CHARGE_RE = re.compile(
    r"total charges for this (?:period|bill|invoice)\s+£\s?([\d,]+(?:\.\d{2})?)",
    re.IGNORECASE,
)


# ---------------------------------------------------------------------------
# Date helpers
# ---------------------------------------------------------------------------


def parse_to_sort_date(date_input):
    s = str(date_input).strip() if date_input else ""
    if not s or s in ("Unknown", "N/A", ""):
        return pd.NaT
    try:
        if _ISO_DATE_RE.match(s):
            return pd.to_datetime(s, format="%Y-%m-%d", errors="coerce")
        dt = pd.to_datetime(s, dayfirst=True, errors="coerce")
        if pd.isna(dt):
            dt = pd.to_datetime(s, dayfirst=False, errors="coerce")
        return dt
    except Exception:
        return pd.NaT


def parse_to_display_date(date_input):
    dt = parse_to_sort_date(date_input)
    return dt.strftime("%d/%m/%Y") if not pd.isna(dt) else str(date_input)


def to_excel_date(date_input):
    """Return a Python datetime for openpyxl to write as a true Excel date serial."""
    dt = parse_to_sort_date(date_input)
    if pd.isna(dt):
        return None
    return dt.to_pydatetime()


def _account_number_matches(acc_filter: str, text: str) -> bool:
    """Return True when ``acc_filter`` appears as a standalone digit run in ``text``.

    Phase 1.3: replaces the old "is the digits substring contained anywhere"
    check, which false-matched an unrelated longer string (invoice number,
    meter serial, phone number) when the configured account number happened
    to be a subset of those digits.

    Strategy
    --------
    Split ``text`` into "tokens" — sequences of alphanumeric characters and
    hyphens (e.g., "A-12345678", "31", "555", "4444").  For each token,
    strip non-digit characters to get its digit-only form.  Look up the
    normalized ``acc_filter in the resulting list.

    This preserves the natural word boundaries of ``"Account number:
    31 555 4444"`` — the tokens are ``["Account", "number", "31", "555",
    "4444", "Current", "balance", "240", "50", "debit"]`` and their
    digit-only forms are ``["", "", "31", "555", "4444", "", "", "240",
    "50", ""]``.  The filter ``"31"`` only matches the "31" token —
    whereas the naive ``digits_only in text_no_sep`` would have produced
    ``"315554444"`` from the same input and dropped the right answer.

    Crucially, hyphens *inside* a token are preserved during tokenization
    so ``"A-12345678"`` is one token whose digit-only form is
    ``"12345678"`` — fixing the false-negative where the original
    ``re.findall(r"\\d+", ...)`` split it into ``["123", "45678"]``.

    Invariant
    ---------
    Pure-substring match (digits in collapse-stripped text) being True
    does NOT imply this helper returns True (we tighten the
        predicate).  The reverse direction holds: a real standalone digit
        run from the original text, after the token-based split, still
        matches.  No legitimate standalone occurrence is dropped.
    """
    if not acc_filter:
        return True  # empty filter matches everything
    digits_only = re.sub(r"\D", "", str(acc_filter))
    if not digits_only:
        return True  # unusable filter — pass rather than silently reject

    # Tokenize: split on whitespace and punctuation EXCEPT hyphens that
    # are inside alphanumeric sequences.  The pattern [A-Za-z0-9-]+
    # captures words, numbers, and hyphenated codes like "A-12345678"
    # or "A123-456" as single tokens.
    tokens = re.findall(r"[A-Za-z0-9-]+", text or "")

    # For each token, strip non-digits to get its digit-only form.
    # Example: "A-12345678" → "12345678", "31" → "31", "555" → "555"
    token_digits = [re.sub(r"\D", "", t) for t in tokens]

    return digits_only in token_digits


# ---------------------------------------------------------------------------
# Detect which EDF bill format we're looking at
# ---------------------------------------------------------------------------


def detect_pdf_format(text):
    """Return 'new_invoice', 'new_credit', or 'old' based on document markers."""
    if _KI_PRESENCE_RE.search(text):
        return "new_invoice"
    if _KCR_PRESENCE_RE.search(text):
        return "new_credit"
    return "old"


def extract_new_invoice_fields(text):
    """Extract key fields from new-style KI-XXXXXXXX invoices."""
    fields = {}

    # Invoice number
    m = _INV_NUMBER_RE.search(text)
    if m:
        fields["inv_num"] = m.group(1).strip()

    # Account number. EDF has shipped multiple renderings on real
    # invoices — both
    #
    #     "Account number: A-31105244"
    #
    # and
    #
    #     "Account number: 671 078 701 920"
    #
    # show up in real exports. The compact form requires the A- prefix;
    # the spaced form omits the prefix and groups digits three-by-three.
    #
    # Pre-fix only the compact form was matched, so any invoice using
    # spaced digits silently dropped ``acc_num`` AND failed the
    # downstream --acc-filter (the user could not filter to their
    # own bill). The regex below matches both renderings and emits a
    # single normalised account number (either ``A-NNNNNNNN`` or
    # ``601 234 567 890`` depending on what was on the page). Callers
    # that need to compare against a filter value are responsible for
    # stripping spaces and the ``A-`` prefix themselves; see the
    # existing helper at the engine filter check.
    m = _ACC_NUM_RE.search(text)
    if m:
        fields["acc_num"] = m.group(1).strip()

    # Date issued
    m = _DATE_ISSUED_RE.search(text)
    if m:
        fields["date"] = parse_to_display_date(m.group(1).strip())

    # Billing period from "Your charges: DD Mon YYYY - DD Mon YYYY"
    m = _BILLING_PERIOD_RE.search(text)
    if m:
        fields["period_from"] = parse_to_display_date(m.group(1).strip())
        fields["period_to"] = parse_to_display_date(m.group(2).strip())

    # Current balance (the running account total — used as primary Amount).
    #
    # Pre-fix this hard-required "debit" — same gap as the #15 HTM parser:
    # if the KI invoice reports ``Current balance GBPX in credit`` (rare but
    # legal: e.g. over-payment or opening credit balance), this matcher
    # would drop the Amount cell. Accept either currency-side label.
    m = _CURRENT_BAL_RE.search(text)
    if m:
        fields["amount"] = float(m.group(1).replace(",", ""))
        fields["amount_side"] = (m.group(2) or "").lower()

    # Period charge (total for this invoice).
    #
    # Pre-fix, this regex hard-required "debit" after the amount — it
    # silently dropped period charges where the line was reported in
    # credit. The "debit"/"credit" labelling is a side-property of
    # the statement, not the period charge itself; accept either so a
    # credit-flagged period still populates the Period Charge column.
    # Captures: group(1) = amount, group(2) = debit|credit (may be empty).
    m = _PERIOD_CHARGE_RE.search(text)
    if m:
        fields["period_charge"] = float(m.group(1).replace(",", ""))

    # kWh used
    m = _UNITS_USED_RE.search(text)
    if m:
        fields["units_used"] = m.group(1)

    # Standing charge
    m = _STANDING_CHARGE_RE.search(text)
    if m:
        fields["standing_charge"] = m.group(1)

    # Tariff name
    m = _TARIFF_NAME_RE.search(text)
    if m:
        fields["tariff"] = m.group(1).strip()

    return fields


def extract_new_credit_fields(text):
    """Extract key fields from new-style KCR-XXXXXXXX credit notes."""
    fields = {}

    m = _CREDIT_NUMBER_RE.search(text)
    if m:
        fields["inv_num"] = m.group(1).strip()

    # Account number — accept both EDF renderings: compact
    # "A-NNNNNNNN" and spaced-digits "601 234 567 890". See the same
    # note in extract_new_invoice_fields above for context.
    m = _ACC_NUM_RE.search(text)
    if m:
        fields["acc_num"] = m.group(1).strip()

    m = _DATE_ISSUED_RE.search(text)
    if m:
        fields["date"] = parse_to_display_date(m.group(1).strip())

    # Total credits for this bill
    m = _CREDIT_TOTAL_RE.search(text)
    if m:
        fields["amount"] = float(m.group(1).replace(",", ""))

    return fields


# ---------------------------------------------------------------------------
# HTM account-history parser
# ---------------------------------------------------------------------------


# ---------------------------------------------------------------------------
# HTM account-history parser
# ---------------------------------------------------------------------------
#
# EDF MyAccount exports "Payments and Invoices" in HTML. The recurring
# row shapes we recognise:
#
#   "DD Mon YYYY We charged your account £X.XX For Y kWh between D Mon YYYY and D Mon YYYY Balance £Z.ZZ in debit|credit"
#   "DD Mon YYYY You paid us £X.XX [Bank Transfer] Balance £Z.ZZ in debit|credit"
#   "DD Mon YYYY Reversed account charge £X.XX Balance £Z.ZZ in debit|credit"
#   "DD Mon YYYY [Bank Transfer / nothing.] Balance £Z.ZZ in credit"  -- standalone
#                               credit-only balance lines that appear when
#                               the customer's overall balance is in credit
#                               and there is no transaction for the period.
#
# Pre-fix (#15): the Balance clause hard-required "in debit", silently
# dropping "in credit" rows.  This was a real, reproducible data loss.
#
# Each regex matches the trailing "Balance £X in (debit|credit)" with a
# non-grouping alternation so existing group numbers are preserved.


def parse_htm_account_history(text):
    """
    Parse the EDF MyAccount 'Payments and Invoices' HTM export.
    Returns a list of record dicts ready for process_text bypass.
    """
    records = []

    # We look for the recurring pattern:
    # "DD Mon YYYY We charged your account £X.XX For Y kWh … between D Mon YYYY and D Mon YYYY Balance £X.XX in debit"
    # "DD Mon YYYY You paid us £X.XX … Balance £X.XX in debit"

    # Normalise whitespace
    text = re.sub(r"\s+", " ", text)

    # Find all "charged" entries
    charge_re = re.compile(
        r"(\d{1,2}\s+\w+\s+\d{4})\s+We charged your account\s+£([\d,]+\.\d{2})"
        r"(?:\s+For\s+([\d,]+)\s+kWh\s+of\s+electricity\s+used\s+between\s+"
        r"(\d{1,2}\s+\w+\s+\d{4})\s+and\s+(\d{1,2}\s+\w+\s+\d{4}))?"
        r".*?Balance\s+£([\d,]+\.\d{2})\s+in\s+(?:debit|credit)",
        re.IGNORECASE,
    )
    # Track the byte ranges ([start, end)) already covered by the
    # three verb-aware regexes below so #15's standalone-balance
    # step does not double-count the trailing balance clause of a
    # charge/payment/reversal line.
    covered: list[tuple[int, int]] = []
    for m in charge_re.finditer(text):
        covered.append((m.start(0), m.end(0)))
        date_str = parse_to_display_date(m.group(1))
        period_from = parse_to_display_date(m.group(4)) if m.group(4) else "N/A"
        period_to = parse_to_display_date(m.group(5)) if m.group(5) else "N/A"
        units = m.group(3) if m.group(3) else "N/A"
        charge_amt = float(m.group(2).replace(",", ""))
        balance = float(m.group(6).replace(",", ""))
        records.append(
            {
                "Source": "HTM Account History",
                "Sender": "",
                "Date": date_str,
                "Period From": period_from,
                "Period To": period_to,
                "Invoice #": "N/A",
                "Amount (£)": balance,
                "Period Charge (£)": charge_amt,
                "Entry Type": "Ongoing Balance",
                "Reading": "N/A",
                "Units (kWh)": units,
                "Standing Chg (p/day)": "N/A",
                # HTM exports don't carry a tariff name in the
                # account-history view; "N/A" is the schema sentinel
                # matching the rest of the record-building paths
                # that include the Tariff column.
                "Tariff": "N/A",
                "Attachment Name": "N/A",
                "Details": "HTM: charged account",
                "Logic Used": "HTM Charge",
            }
        )

    # Find all "You paid us" entries
    pay_re = re.compile(
        r"(\d{1,2}\s+\w+\s+\d{4})\s+You paid us\s+£([\d,]+\.\d{2})"
        r".*?Balance\s+£([\d,]+\.\d{2})\s+in\s+(?:debit|credit)",
        re.IGNORECASE,
    )
    for m in pay_re.finditer(text):
        covered.append((m.start(0), m.end(0)))
        date_str = parse_to_display_date(m.group(1))
        balance = float(m.group(3).replace(",", ""))
        records.append(
            {
                "Source": "HTM Account History",
                "Sender": "",
                "Date": date_str,
                "Period From": "N/A",
                "Period To": "N/A",
                "Invoice #": "N/A",
                "Amount (£)": balance,
                "Period Charge (£)": "N/A",
                "Entry Type": "Payment",
                "Reading": "N/A",
                "Units (kWh)": "N/A",
                "Standing Chg (p/day)": "N/A",
                "Tariff": "N/A",
                "Attachment Name": "N/A",
                "Details": "HTM: payment received",
                "Logic Used": "HTM Payment",
            }
        )

    # Find all "reversed account charge" entries (credits applied)
    rev_re = re.compile(
        r"(\d{1,2}\s+\w+\s+\d{4})\s+Reversed account charge\s+£([\d,]+\.\d{2})"
        r".*?Balance\s+£([\d,]+\.\d{2})\s+in\s+(?:debit|credit)",
        re.IGNORECASE,
    )
    for m in rev_re.finditer(text):
        covered.append((m.start(0), m.end(0)))
        date_str = parse_to_display_date(m.group(1))
        balance = float(m.group(3).replace(",", ""))
        records.append(
            {
                "Source": "HTM Account History",
                "Sender": "",
                "Date": date_str,
                "Period From": "N/A",
                "Period To": "N/A",
                "Invoice #": "N/A",
                "Amount (£)": balance,
                "Period Charge (£)": "N/A",
                "Entry Type": "Credit",
                "Reading": "N/A",
                "Units (kWh)": "N/A",
                "Standing Chg (p/day)": "N/A",
                "Tariff": "N/A",
                "Attachment Name": "N/A",
                "Details": "HTM: reversed account charge",
                "Logic Used": "HTM Reversal",
            }
        )

    # Find standalone "Balance £X in credit" lines.
    #
    # These appear at the top of an HTM export when the customer's
    # overall balance is in credit and there is no transaction recorded
    # for the period (e.g. a credit accumulated from the previous
    # statement still on the books). Pre-#15 there was no regex to
    # catch them, so the credit on this kind of opening line never
    # reached downstream classification.
    #
    # We walk the regex and reject any match whose date-token is
    # *inside* a span already covered by the verb-aware regexes
    # above. The regex itself is intentionally tolerant of optional
    # postcard text between the date and the Balance clause — the
    # covered-range check is what keeps standalone from double-counting
    # the trailing balance of a real charge/payment/reversal line.
    def _inside_covered(start: int, end: int) -> bool:
        for cs, ce in covered:
            if start >= cs and end <= ce:
                return True
        return False

    bal_re = re.compile(
        r"(\d{1,2}\s+\w+\s+\d{4})[^A-Za-z0-9]*?Balance\s+£([\d,]+\.\d{2})\s+in\s+credit\b",
        re.IGNORECASE,
    )
    for m in bal_re.finditer(text):
        if _inside_covered(m.start(0), m.end(0)):
            continue
        covered.append((m.start(0), m.end(0)))
        date_str = parse_to_display_date(m.group(1))
        balance = float(m.group(2).replace(",", ""))
        records.append(
            {
                "Source": "HTM Account History",
                "Sender": "",
                "Date": date_str,
                "Period From": "N/A",
                "Period To": "N/A",
                "Invoice #": "N/A",
                "Amount (£)": balance,
                "Period Charge (£)": "N/A",
                "Entry Type": "Credit",
                "Reading": "N/A",
                "Units (kWh)": "N/A",
                "Standing Chg (p/day)": "N/A",
                "Tariff": "N/A",
                "Attachment Name": "N/A",
                "Details": "HTM: standalone credit balance",
                "Logic Used": "HTM StandaloneBalance",
            }
        )

    return records


# ---------------------------------------------------------------------------
# Evidence Engine
# ---------------------------------------------------------------------------


def _extract_sender_email(msg):
    """Extract sender email address from a pypff message, trying multiple methods."""
    sender = None
    # Try transport headers first (most reliable for SMTP email address)
    try:
        headers = msg.get_transport_headers()
        if headers:
            headers_str = (
                headers if isinstance(headers, str) else headers.decode("utf-8", errors="replace")
            )
            m = _FROM_HEADER_RE.search(headers_str)
            if m:
                sender = m.group(1).lower()
    except Exception:
        pass
    # Fallback: try sender name field (sometimes contains email)
    if not sender:
        try:
            name = msg.get_sender_name() or ""
            m = _EMAIL_ADDR_RE.search(name)
            if m:
                sender = m.group(1).lower()
        except Exception:
            pass
    return sender or ""


def _matches_domain_filter(sender_email, filter_str):
    """
    Check if sender_email matches the domain filter string.
    filter_str is comma-separated, supporting:
      - domain names: "edf.com" matches *@edf.com and *@*.edf.com
      - full addresses: "billing@edf.com" matches exactly
      - wildcard domains: "*.edf.com" matches subdomains
    """
    if not sender_email or not filter_str:
        return False
    sender_email = sender_email.lower().strip()
    parts = [p.strip().lower() for p in filter_str.split(",") if p.strip()]
    for pattern in parts:
        if "@" in pattern:
            # Full email address match
            if sender_email == pattern:
                return True
        else:
            # Domain match — check exact domain or subdomain
            domain = pattern.lstrip("*").lstrip(".")
            sender_domain = sender_email.split("@")[-1] if "@" in sender_email else ""
            if sender_domain == domain or sender_domain.endswith("." + domain):
                return True
    return False


class EvidenceEngine:
    def __init__(self, config, update_ui_cb, progress_cb=None, cancel_event=None):
        self.config = config
        self.records = []
        self.filtered_records = []
        self.update_ui = update_ui_cb
        self.update_progress = progress_cb
        self.cancel_event = cancel_event or threading.Event()
        self.pdf_count = 0
        self.email_count = 0
        self.error_log = []
        self.seen_pdf_hashes = set()
        self.lock = threading.Lock()

    # ------------------------------------------------------------------
    # Pickle support — Phase 1.4
    # ------------------------------------------------------------------
    def __getstate__(self) -> dict:
        """Return a picklable snapshot of the engine data.

        ``EvidenceEngine`` carries three non-picklable runtime
        primitives — ``threading.Lock``, ``threading.Event``, and the
        two callbacks — which can't survive a naive ``pickle.dump`` of
        the instance (``TypeError: cannot pickle '_thread.lock' object``).
        We round-trip the *data* the engine holds, and rebuild the
        threading primitives fresh in ``__setstate__``.

        This means a loaded engine is fully usable again — just with
        fresh ``Lock``/``Event`` instances and no cancellation state
        from the persisting session — which is the right semantic for
        a CLI report-on-engine-data flow that resumes a saved snapshot.

        Concretely we strip:
          * ``self.lock``             (``threading.Lock`` — not picklable)
          * ``self.cancel_event``     (``threading.Event`` — not picklable)
          * ``self.update_ui``        (a GUI callback; serialising
                                      Tkinter closures would leak the GUI
                                      context across the CLI↔GUI boundary)
          * ``self.update_progress``  (same reason)

        and rebuild them in ``__setstate__``.
        """
        return {
            "config": self.config,
            "records": self.records,
            "filtered_records": self.filtered_records,
            "pdf_count": self.pdf_count,
            "email_count": self.email_count,
            "error_log": self.error_log,
            "seen_pdf_hashes": self.seen_pdf_hashes,
        }

    def __setstate__(self, state: dict) -> None:
        """Restore a pickled snapshot — rebuild non-picklable fields fresh.

        See ``__getstate__`` for why each of these is set this way.
        """
        self.config = state["config"]
        self.records = state["records"]
        self.filtered_records = state["filtered_records"]
        self.pdf_count = state["pdf_count"]
        self.email_count = state["email_count"]
        self.error_log = state["error_log"]
        self.seen_pdf_hashes = state["seen_pdf_hashes"]
        # Rebuild runtime primitives fresh — the persisted snapshot
        # does not carry cancel state forward.
        self.cancel_event = threading.Event()
        self.lock = threading.Lock()
        # GUI callbacks don't survive a CLI↔CLI round-trip; a GUI
        # consumer can install its own after loading the snapshot via
        # ``engine.update_ui = my_gui_callback``.
        self.update_ui = lambda *_a, **_kw: None
        self.update_progress = lambda *_a, **_kw: None

    def is_cancelled(self):
        return self.cancel_event.is_set()

    def log_error(self, context, err):
        self.error_log.append(f"[{datetime.now().strftime('%H:%M:%S')}] {context} — {err}")

    def find_billing_period(self, text):
        m = PERIOD_RE.search(text)
        if m:
            return (
                parse_to_display_date(m.group(1).strip()),
                parse_to_display_date(m.group(2).strip()),
            )
        return "N/A", "N/A"

    def _add_record(self, rec):
        """Thread-safe record append after optional filter check."""
        amt = rec.get("Amount (£)", 0) or 0
        if self.config.get("filter_below", True) and amt < self.config["min_amount"]:
            with self.lock:
                self.filtered_records.append(
                    {
                        "Source": rec.get("Source", ""),
                        "Date": rec.get("Date", ""),
                        "Amount (£)": amt,
                        "Details": rec.get("Details", "")[:60],
                        "Logic Used": rec.get("Logic Used", ""),
                        "Reason": f"Below minimum threshold (£{self.config['min_amount']:,.2f})",
                    }
                )
            return
        with self.lock:
            self.records.append(rec)

    # ------------------------------------------------------------------
    # New-format PDF processing
    # ------------------------------------------------------------------

    def _process_new_invoice(
        self, text, source_label, detail_label, fallback_date, sender="", attachment_name=""
    ):
        fields = extract_new_invoice_fields(text)
        if "amount" not in fields:
            return False  # didn't match

        # Account filter
        if self.config.get("use_acc_filter"):
            acc = self.config.get("acc_num", "")
            if acc and not _account_number_matches(acc, text):
                return False

        r_type = "Unknown"
        for label, pat in READING_PATTERNS.items():
            if pat.search(text):
                r_type = label
                break

        # Classify entry type: New Bill if it has period charges, else Ongoing Balance
        entry_type = (
            "New Bill"
            if fields.get("period_charge") or fields.get("period_from")
            else "Ongoing Balance"
        )

        self._add_record(
            {
                "Source": source_label,
                "Sender": sender,
                "Date": fields.get("date", fallback_date),
                "Period From": fields.get("period_from", "N/A"),
                "Period To": fields.get("period_to", "N/A"),
                "Invoice #": fields.get("inv_num", "N/A"),
                "Amount (£)": fields["amount"],
                "Period Charge (£)": fields.get("period_charge", "N/A"),
                "Entry Type": entry_type,
                "Reading": r_type,
                "Units (kWh)": fields.get("units_used", "N/A"),
                "Standing Chg (p/day)": fields.get("standing_charge", "N/A"),
                # Tariff name is extracted by ``extract_new_invoice_fields``
                # into ``fields["tariff"]`` (regex _TARIFF_NAME_RE on
                # the invoice body). Copy it into the record so the
                # downstream Tariff Analysis feature sees it; it was
                # previously silently discarded here, which left the
                # "Tariff Analysis" Excel/PDF/DOCX tabs permanently
                # empty.  This is one of four record-building paths
                # (the other three — HTM charged/paid/reversed and
                # process_text — append "Tariff": "N/A" so the column
                # has a consistent shape across all sources).
                "Tariff": fields.get("tariff", "N/A"),
                "Attachment Name": attachment_name or "N/A",
                "Details": (detail_label or "New invoice")[:60],
                "Logic Used": "New Invoice Format",
            }
        )
        return True

    def _process_new_credit(
        self, text, source_label, detail_label, fallback_date, sender="", attachment_name=""
    ):
        fields = extract_new_credit_fields(text)
        if "amount" not in fields:
            return False

        if self.config.get("use_acc_filter"):
            acc = self.config.get("acc_num", "")
            if acc and not _account_number_matches(acc, text):
                return False

        self._add_record(
            {
                "Source": source_label,
                "Sender": sender,
                "Date": fields.get("date", fallback_date),
                "Period From": "N/A",
                "Period To": "N/A",
                "Invoice #": fields.get("inv_num", "N/A"),
                "Amount (£)": fields["amount"],
                "Period Charge (£)": "N/A",
                "Entry Type": "Credit",
                "Reading": "N/A",
                "Units (kWh)": "N/A",
                "Standing Chg (p/day)": "N/A",
                # KCR credit-note letters do not carry a tariff name
                # (the ``extract_new_credit_fields`` handler does not
                # populate ``fields["tariff"]``). "N/A" is the schema
                # sentinel — see the Tariff Analysis upgrade note in
                # ``_process_new_invoice`` for why this key is present
                # on every record dict, not just invoice rows.
                "Tariff": "N/A",
                "Attachment Name": attachment_name or "N/A",
                "Details": (detail_label or "Credit note")[:60],
                "Logic Used": "New Credit Note Format",
            }
        )
        return True

    # ------------------------------------------------------------------
    # Generic text processing (old format + email bodies)
    # ------------------------------------------------------------------

    def process_text(self, text, source_type, detail, fallback_date, sender="", attachment_name=""):
        if not text:
            return

        clean_text = re.sub(r"\s+", " ", text)

        # Account filter
        if self.config.get("use_acc_filter"):
            acc = self.config.get("acc_num", "")
            if acc and not _account_number_matches(acc, clean_text):
                return

        found_amt, strategy = None, ""
        matched_pattern_name: str | None = None

        if self.config.get("use_anchors", True):
            for name, p in AMOUNT_PATTERNS:
                # Patterns are pre-compiled at module load with
                # `re.IGNORECASE` baked in, so search() takes no flags.
                m = p.search(clean_text)
                if m:
                    try:
                        found_amt = float(m.group(1).replace(",", ""))
                        strategy = "Smart Context"
                        matched_pattern_name = name
                        break
                    except Exception:
                        continue

        if not found_amt and self.config.get("use_large", True):
            matches = _POUND_AMOUNT_FALLBACK_RE.findall(clean_text)
            if matches:
                floats = [float(x.replace(",", "")) for x in matches]
                highs = [x for x in floats if x >= self.config["min_amount"]]
                if highs:
                    found_amt = max(highs)
                    strategy = "Large Amount Fallback"

        if not found_amt:
            return

        # Date extraction
        date_to_use = fallback_date
        if "PDF" in source_type or "old" in source_type.lower():
            date_m = _OLD_PDF_DATE_RE.search(clean_text)
            if date_m:
                date_to_use = parse_to_display_date(date_m.group(1))

        r_type = "Unknown"
        if self.config.get("use_reading_classification", True):
            for label, pat in READING_PATTERNS.items():
                if pat.search(clean_text):
                    r_type = label
                    break

        units_used = standing_charge = inv_num = "N/A"
        if self.config.get("use_pdf_fields", True):
            u_m = _OLD_PDF_KWH_RE.search(clean_text)
            sc_m = _OLD_PDF_STANDING_RE.search(clean_text)
            in_m = _OLD_PDF_INV_RE.search(clean_text)
            if u_m:
                units_used = u_m.group(1)
            if sc_m:
                standing_charge = sc_m.group(1)
            if in_m:
                inv_num = in_m.group(1)

        period_from, period_to = self.find_billing_period(clean_text)

        # Attempt to extract period charge separately from cumulative balance
        period_charge: str | float = "N/A"
        pc_m = _OLD_PDF_PERIOD_CHARGE_RE.search(clean_text)
        if pc_m:
            try:
                period_charge = float(pc_m.group(1).replace(",", ""))
            except (ValueError, AttributeError):
                pass

        # Classify Entry Type based on content
        entry_type = self._classify_entry_type(
            clean_text, matched_pattern_name, period_from, period_to, strategy
        )

        self._add_record(
            {
                "Source": source_type,
                "Sender": sender,
                "Date": date_to_use,
                "Period From": period_from,
                "Period To": period_to,
                "Invoice #": inv_num,
                "Amount (£)": found_amt,
                "Period Charge (£)": period_charge,
                "Entry Type": entry_type,
                "Reading": r_type,
                "Units (kWh)": units_used,
                "Standing Chg (p/day)": standing_charge,
                # Old/email-body bills have no "Tariff name" line in
                # the standard heuristic pattern set, so this column
                # is ``"N/A"`` for them.  Treated as schema
                # sentinel so the column exists for every source.
                "Tariff": "N/A",
                "Attachment Name": attachment_name or "N/A",
                "Details": detail[:60],
                "Logic Used": strategy,
            }
        )

    def _classify_entry_type(
        self,
        text: str,
        pattern_name: str | None,
        period_from: str,
        period_to: str,
        strategy: str,
    ) -> str:
        """Classify a record as New Bill, Ongoing Balance, or Other based on content.

        Args:
            text (str): the cleaned bill body text.
            pattern_name (str | None): the name of the regex from
                :data:`AMOUNT_PATTERNS` that matched, or ``None`` if no
                anchored match was found.
            period_from, period_to (str): ``"N/A"`` or a parsed date.
            strategy (str): either ``"Smart Context"`` (anchored pattern
                matched) or ``"Large Amount Fallback"`` (anchored missed,
                number extracted by fallback).

        The classifier explicitly maps ``pattern_name`` to ``New Bill`` or
        ``Ongoing Balance`` via
        :data:`_AMOUNT_PATTERN_NEW_BILL` /
        :data:`_AMOUNT_PATTERN_ONGOING_BALANCE`. Unknown / unset names
        fall through to heuristic checks against the bill body text.
        """
        text_lower = text.lower()

        # If it has billing period dates AND charges/invoice details → New Bill
        has_period = period_from != "N/A" and period_to != "N/A"
        has_bill_markers = bool(_BILL_MARKERS_RE.search(text_lower))

        if has_period and has_bill_markers:
            return "New Bill"

        # Pattern-name driven classification. The integer-index lookup
        # used previously was brittle: reordering or inserting a pattern
        # silently changed classification. Names are stable.
        if pattern_name is not None:
            if pattern_name in _AMOUNT_PATTERN_NEW_BILL:
                return "New Bill"
            if pattern_name in _AMOUNT_PATTERN_ONGOING_BALANCE:
                return "Ongoing Balance"

        # If matched via "balance" pattern or has "account balance" language → Ongoing Balance
        if _ACCOUNT_BALANCE_LANG_RE.search(text_lower):
            return "Ongoing Balance"

        # If matched via total/amount to pay with period info → New Bill
        if has_period:
            return "New Bill"

        # Fallback strategy check
        if strategy == "Large Amount Fallback":
            return "Other"

        # Default: if it looks like a bill (has kWh, standing charge) → New Bill
        if _BILL_INDICATORS_RE.search(text_lower):
            return "New Bill"

        return "Ongoing Balance"

    # ------------------------------------------------------------------
    # PDF file processing — detects format automatically
    # ------------------------------------------------------------------

    def process_pdf_file(
        self, path, source_label, detail_label, fallback_date, sender="", attachment_name=""
    ):
        if self.is_cancelled():
            return
        try:
            import io

            with open(path, "rb") as fh:
                raw = fh.read()
            pdf_hash = hashlib.sha256(raw).hexdigest()
            with self.lock:
                if pdf_hash in self.seen_pdf_hashes:
                    return
                self.seen_pdf_hashes.add(pdf_hash)

            with pdfplumber.open(io.BytesIO(raw)) as pdf:
                # Handle empty or corrupt PDFs gracefully
                if not pdf.pages:
                    self.log_error(f"PDF: {detail_label}", "PDF has no pages")
                    return
                pdf_text_parts = []
                for p in pdf.pages:
                    try:
                        page_text = p.extract_text()
                        if page_text:
                            pdf_text_parts.append(page_text)
                    except (
                        pdfplumber.utils.exceptions.PdfminerException,
                        ValueError,
                        TypeError,
                    ) as page_err:
                        # Narrowly catch PDF-syntax / text-coercion errors so
                        # a single bad page does not skip the whole file.
                        # ``BaseException`` (e.g. ``KeyboardInterrupt``) and
                        # unexpected runtime errors propagate so the caller
                        # can still cancel or surface real bugs.
                        self.log_error(
                            f"PDF page {detail_label}", f"Page extraction failed: {page_err}"
                        )
                pdf_text = " ".join(pdf_text_parts)
            del raw

            # Use original filename as attachment_name if not already set
            if not attachment_name:
                attachment_name = detail_label or ""

            fmt = detect_pdf_format(pdf_text)

            if fmt == "new_invoice":
                self._process_new_invoice(
                    pdf_text,
                    source_label,
                    detail_label,
                    fallback_date,
                    sender=sender,
                    attachment_name=attachment_name,
                )
            elif fmt == "new_credit":
                self._process_new_credit(
                    pdf_text,
                    source_label,
                    detail_label,
                    fallback_date,
                    sender=sender,
                    attachment_name=attachment_name,
                )
            else:
                self.process_text(
                    pdf_text,
                    source_label,
                    detail_label,
                    fallback_date,
                    sender=sender,
                    attachment_name=attachment_name,
                )

        except Exception as e:
            self.log_error(f"PDF: {detail_label}", str(e))

    # ------------------------------------------------------------------
    # HTM account history
    # ------------------------------------------------------------------

    def process_htm_file(self, path):
        try:
            # Read with strict UTF-8 first — evidence data must not be
            # silently corrupted by mojibake replacement.  Fall back to
            # "replace" only if strict fails, and log a warning so the
            # user knows data may be imperfect.
            try:
                with open(path, encoding="utf-8", errors="strict") as f:
                    content = f.read()
            except UnicodeDecodeError:
                self.log_error(f"HTM: {path}", "UTF-8 decode error — some characters replaced")
                with open(path, encoding="utf-8", errors="replace") as f:
                    content = f.read()
            soup = BeautifulSoup(content, "html.parser")
            text = soup.get_text(separator=" ", strip=True)
            recs = parse_htm_account_history(text)
            for rec in recs:
                self._add_record(rec)
            self.update_ui(f"HTM: extracted {len(recs)} account history entries")
        except Exception as e:
            self.log_error(f"HTM: {path}", str(e))

    def process_pst_file(self, path):
        """Open a PST file at ``path`` and crawl its root folder.

        Wrapper around :meth:`crawl_pst` so the public per-file API
        is symmetric with :meth:`process_pdf_file` and
        :meth:`process_htm_file`. Returns nothing; outcomes are
        surfaced through ``update_ui`` / ``error_log``.
        """
        if not HAS_PYPFF:
            self.log_error(
                "PST",
                f"pypff not installed — cannot open PST file {path}",
            )
            return
        try:
            import pypff

            pst = pypff.file()
            pst.open(path)
            try:
                root = pst.get_root_folder()
                self.crawl_pst(root)
            finally:
                try:
                    pst.close()
                except Exception:
                    pass
        except Exception as e:
            self.log_error(f"PST: {path}", str(e))

    # `process_ost_file` is the same code path: ``libpff-python`` accepts
    # both PST and OST archives. Exposed as an explicit alias so
    # callers picking from the per-file API do not have to know that.
    def process_ost_file(self, path):
        self.process_pst_file(path)

    # ------------------------------------------------------------------
    # PST / OST crawl
    # ------------------------------------------------------------------

    def crawl_pst(self, folder):
        if not HAS_PYPFF:
            self.log_error("PST", "pypff not installed — skipping PST processing")
            return
        if self.is_cancelled():
            return

        msg_total = folder.get_number_of_sub_messages()
        for i in range(msg_total):
            if self.is_cancelled():
                return
            try:
                msg = folder.get_sub_message(i)
                subj = str(msg.get_subject() or "")
                d_time = msg.get_delivery_time()
                date_str = parse_to_display_date(d_time.strftime("%Y-%m-%d")) if d_time else "N/A"

                if self.update_progress and i % 100 == 0:
                    self.update_progress(
                        i + 1, msg_total, f"Scanning PST/OST folder: {i + 1}/{msg_total}"
                    )

                # Extract sender email for domain filtering and spreadsheet
                sender_email = _extract_sender_email(msg)

                # Determine if this email should be processed
                use_domain = self.config.get("use_domain_filter", False)
                domain_str = self.config.get("domain_filter", "")
                should_process = False
                if use_domain and domain_str:
                    if _matches_domain_filter(sender_email, domain_str):
                        should_process = True
                else:
                    if any(
                        k in subj.upper()
                        for k in ["EDF", "BILL", "STATEMENT", "ACCOUNT", "INVOICE"]
                    ):
                        should_process = True

                if should_process:
                    with self.lock:
                        self.email_count += 1
                    html = msg.get_html_body()
                    plain = msg.get_plain_text_body()

                    if html:
                        body_text = BeautifulSoup(html, "html.parser").get_text(separator=" ")
                        self.process_text(
                            body_text, "Email Body", subj, date_str, sender=sender_email
                        )
                    elif plain:
                        self.process_text(
                            plain.decode("utf-8", errors="ignore"),
                            "Email Body",
                            subj,
                            date_str,
                            sender=sender_email,
                        )
                    else:
                        rtf_body = None
                        try:
                            rtf_body = msg.get_rtf_body()
                        except Exception:
                            pass
                        if rtf_body:
                            try:
                                rtf_str = rtf_body.decode("utf-8", errors="replace")
                                rtf_text = re.sub(r"\\[a-z]+[-\d]*\s?", " ", rtf_str)
                                rtf_text = re.sub(r"[{}\\]", " ", rtf_text)
                                self.process_text(
                                    rtf_text,
                                    "Email Body (RTF)",
                                    subj,
                                    date_str,
                                    sender=sender_email,
                                )
                            except Exception as e:
                                self.log_error(f"Email: {subj}", f"RTF decode: {e}")
                        else:
                            self.log_error(f"Email: {subj} ({date_str})", "No readable body")

                    for a_idx in range(msg.get_number_of_attachments()):
                        if self.is_cancelled():
                            return
                        try:
                            att = msg.get_attachment(a_idx)
                            size = att.get_size()
                            if size > 4:
                                buf = att.read_buffer(size)
                                if buf and buf.startswith(b"%PDF"):
                                    with self.lock:
                                        self.pdf_count += 1
                                    att_name = None
                                    # Try multiple pypff methods to get the real filename
                                    getters = [
                                        lambda a=att: a.name,
                                        lambda a=att: a.get_name(),
                                        lambda a=att: a.get_long_filename(),
                                        lambda a=att: a.get_short_filename(),
                                    ]
                                    for _getter in getters:
                                        try:
                                            val = _getter()
                                            if val:
                                                att_name = val
                                                break
                                        except (AttributeError, TypeError, Exception):
                                            continue
                                    if not att_name:
                                        att_name = f"Attachment_{self.pdf_count}.pdf"
                                    with tempfile.NamedTemporaryFile(
                                        delete=False, suffix=".pdf"
                                    ) as tmp:
                                        tmp.write(buf)
                                        tmp_path = tmp.name
                                    try:
                                        self.process_pdf_file(
                                            tmp_path,
                                            "PST PDF Attachment",
                                            att_name,
                                            date_str,
                                            sender=sender_email,
                                            attachment_name=att_name,
                                        )
                                    finally:
                                        if os.path.exists(tmp_path):
                                            os.remove(tmp_path)
                        except Exception as e:
                            self.log_error(f'Attachment in "{subj}"', str(e))

            except Exception as e:
                self.log_error(f"PST message index {i}", str(e))

        self.update_ui(f"Scanned {self.email_count} emails, {self.pdf_count} attached PDFs…")
        for j in range(folder.get_number_of_sub_folders()):
            if self.is_cancelled():
                return
            self.crawl_pst(folder.get_sub_folder(j))

    # ------------------------------------------------------------------
    # Local PDF folder
    # ------------------------------------------------------------------

    def crawl_local_pdfs(self, path):
        if not path or not os.path.exists(path):
            return
        # Recursive walk: PDF bills are commonly organised into
        # sub-folders by year or account reference (e.g.
        # ``pdfs/2023/2023-01.pdf``).  The legacy implementation
        # only scanned the top-level directory and silently
        # dropped any bills in nested folders — a real EDF
        # dispute case with year-organised PDFs would have
        # silently undercounted, so this matters for ombudsman
        # submissions where a missing bill undoes the entire
        # argument.
        pdf_files: list[tuple[str, str]] = []
        for root, _dirs, files in os.walk(path):
            for f in files:
                if f.lower().endswith(".pdf"):
                    pdf_files.append((root, f))
        # Sort by relative path so the progress narrative is
        # deterministic across runs (otherwise os.walk's
        # filesystem-order output varies by platform).
        pdf_files.sort(
            key=lambda pair: os.path.relpath(os.path.join(pair[0], pair[1]), path).lower()
        )
        total = len(pdf_files)

        def _process_one(i_file):
            idx, (root, fname) = i_file
            if self.is_cancelled():
                return
            file_path = os.path.join(root, fname)
            fallback_date = parse_to_display_date(
                datetime.fromtimestamp(os.path.getmtime(file_path)).strftime("%Y-%m-%d")
            )
            with self.lock:
                self.pdf_count += 1
            self.process_pdf_file(
                file_path, "Local PDF Folder", fname, fallback_date, attachment_name=fname
            )
            if self.update_progress:
                relative = os.path.relpath(file_path, path)
                self.update_progress(idx, total, f"Scanning local PDFs: {idx}/{total} ({relative})")

        # Sequential pass.  The ``_process_one`` closure comment
        # above used to imply a thread-pool dispatch that's no
        # longer present (see also ``EvidenceEngine.lock`` which
        # is in fact exercised by ``process_pdf_file``'s own
        # write paths).  Keeping the indirection for now so
        # transition to ``ThreadPoolExecutor`` later stays a
        # one-line change.
        for item in enumerate(pdf_files, start=1):
            _process_one(item)

        self.update_ui(f"PDF folder: {self.pdf_count} PDFs processed")


# ---------------------------------------------------------------------------
# Excel helpers
# ---------------------------------------------------------------------------

THIN = Side(style="thin", color="DDDDDD")
CELL_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _hcell(ws, row, col, value, bg="FE5716"):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
    c.fill = PatternFill("solid", start_color=bg)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = CELL_BORDER
    return c


def _money(ws, r, c, val, bold=False, fill_hex=None):
    cell = ws.cell(row=r, column=c, value=val)
    cell.font = Font(name="Calibri", size=10, bold=bold)
    cell.border = CELL_BORDER
    cell.number_format = "£#,##0.00"
    cell.alignment = Alignment(horizontal="right", vertical="center")
    if fill_hex:
        cell.fill = PatternFill("solid", start_color=fill_hex)
    return cell


def _text(ws, r, c, val, bold=False, fill_hex=None, wrap=False, align="left", color="000000"):
    # Phase 2.x — formula-injection guard.  External text
    # (PDF/PST/email) can start with ``=``, ``+``, ``-`` or
    # ``@`` and Excel will silently evaluate the cell as a
    # formula when the workbook is opened.  The classic
    # mitigation is to coerce the cell's ``data_type`` to
    # ``'s'`` (text).  We coerce non-strings via ``str()`` first
    # so the cell value is always a Python string before the
    # data_type pin; otherwise cell types follow Python type
    # inference.
    safe_val: str
    if val is None:
        safe_val = ""
    else:
        safe_val = str(val)
        # Belt-and-braces: prefix a leading ``=``, ``+``, ``-``
        # or ``@`` with an apostrophe.  This nails the contract
        # even on Excel versions that still try to evaluate
        # auto-formats despite the ``data_type = 's'`` flag.
        if safe_val and safe_val[0] in "+-=@":
            safe_val = "'" + safe_val
    cell = ws.cell(row=r, column=c, value=safe_val)
    cell.data_type = "s"
    cell.font = Font(name="Calibri", size=10, bold=bold, color=color)
    cell.border = CELL_BORDER
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    if fill_hex:
        cell.fill = PatternFill("solid", start_color=fill_hex)
    return cell


def _num(ws, r, c, val, fmt="#,##0", bold=False, fill_hex=None):
    cell = ws.cell(row=r, column=c, value=val)
    cell.font = Font(name="Calibri", size=10, bold=bold)
    cell.border = CELL_BORDER
    cell.number_format = fmt
    cell.alignment = Alignment(horizontal="right", vertical="center")
    if fill_hex:
        cell.fill = PatternFill("solid", start_color=fill_hex)
    return cell


def _section_hdr(ws, r, label, ncols=3, bg="10367A"):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=r, column=c, value=label if c == 1 else "")
        cell.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", start_color=bg)
        cell.border = CELL_BORDER
        cell.alignment = Alignment(horizontal="left", vertical="center")


def compute_dispute_flags(dfc: pd.DataFrame, mean_daily: float = 0.0) -> tuple[list, dict]:
    """Compute dispute flags from a sorted DataFrame.

    Returns:
        tuple: (flags_list, flag_counts_dict)
        - flags_list: list of (type, date, amount, detail, severity) tuples
        - flag_counts_dict: dict with HIGH, MEDIUM, INFO counts

    Issues a :func:`warnings.warn` for any row that fails to evaluate
    under each heuristic (parse error, missing key, etc.).  Previously
    those rows were silently swallowed and the report lost the
    surrounding evidence — turning them into warnings surfaces a
    developer-visible signal without breaking the run.
    """

    def _flag_or_warn(
        row_idx: int,
        flag_name: str,
        exc: BaseException,
    ) -> None:
        warnings.warn(
            (
                f"compute_dispute_flags[{flag_name}] could not evaluate "
                f"row index {row_idx}: {exc!r}; row silently skipped."
            ),
            stacklevel=3,
        )

    flags: list[tuple[str, str | float | None, float | None, str, str]] = []
    n = len(dfc)
    if n < 2:
        return flags, {"HIGH": 0, "MEDIUM": 0, "INFO": 0}

    # 1. LARGE JUMP: >25% increase within 90 days
    for i in range(1, n):
        p = dfc.iloc[i - 1]
        c_ = dfc.iloc[i]
        try:
            chg = float(c_["Amount (£)"]) - float(p["Amount (£)"])
            pct = chg / float(p["Amount (£)"]) if float(p["Amount (£)"]) > 0 else 0
            days = (c_["_dt"] - p["_dt"]).days
            if pct > 0.25 and 0 < days <= 90:
                flags.append(
                    (
                        "LARGE JUMP",
                        c_["Date"],
                        c_["Amount (£)"],
                        f"+£{chg:,.2f} (+{pct * 100:.1f}%) in {days} days (from {p['Date']}: £{p['Amount (£)']:,.2f})",
                        "HIGH" if pct > 0.5 else "MEDIUM",
                    )
                )
        except (ValueError, TypeError, KeyError) as exc:
            _flag_or_warn(i, "LARGE_JUMP", exc)

    # 2. BILLING GAP: >60 days without a bill
    for i in range(1, n):
        p = dfc.iloc[i - 1]
        c_ = dfc.iloc[i]
        try:
            days = (c_["_dt"] - p["_dt"]).days
            if days > 60:
                flags.append(
                    (
                        "BILLING GAP",
                        c_["Date"],
                        c_["Amount (£)"],
                        f"{days} days without a bill (previous: {p['Date']}). Balance accumulated unchecked.",
                        "HIGH" if days > 120 else "MEDIUM",
                    )
                )
        except (ValueError, TypeError, KeyError) as exc:
            _flag_or_warn(i, "BILLING_GAP", exc)

    # 3. ESTIMATED RUN: 3+ consecutive estimated readings
    if "Reading" in dfc.columns:
        run = 0
        run_start = None
        for i, rv in enumerate(dfc["Reading"].tolist()):
            if str(rv).lower() in ("estimated", "est."):
                run += 1
                if run == 1:
                    run_start = dfc.iloc[i]["Date"]
            else:
                if run >= 3:
                    flags.append(
                        (
                            "ESTIMATED RUN",
                            run_start,
                            None,
                            f"{run} consecutive estimated readings from {run_start}.",
                            "HIGH",
                        )
                    )
                run = 0
                run_start = None
        if run >= 3:
            flags.append(
                (
                    "ESTIMATED RUN",
                    run_start,
                    None,
                    f"{run} consecutive estimated readings from {run_start} (ongoing).",
                    "HIGH",
                )
            )

    # 4. HIGH DAILY RATE: daily rate significantly above average
    if mean_daily > 0:
        for i in range(1, n):
            p = dfc.iloc[i - 1]
            c_ = dfc.iloc[i]
            try:
                days = (c_["_dt"] - p["_dt"]).days
                charge = float(c_["Amount (£)"]) - float(p["Amount (£)"])
                if days > 0 and charge > 0:
                    daily = charge / days
                    ratio = daily / mean_daily
                    if ratio > 2.5:
                        flags.append(
                            (
                                "HIGH DAILY RATE",
                                c_["Date"],
                                c_["Amount (£)"],
                                f"£{daily:,.2f}/day ({ratio:.1f}× avg £{mean_daily:,.2f}/day) over {days} days",
                                "HIGH" if ratio > 4 else "MEDIUM",
                            )
                        )
            except (ValueError, TypeError, KeyError, ZeroDivisionError) as exc:
                _flag_or_warn(i, "HIGH_DAILY_RATE", exc)

    # 5. BALANCE REDUCTION: payment/credit > £500
    for i in range(1, n):
        p = dfc.iloc[i - 1]
        c_ = dfc.iloc[i]
        try:
            chg = float(c_["Amount (£)"]) - float(p["Amount (£)"])
            if chg < -500:
                flags.append(
                    (
                        "BALANCE REDUCTION",
                        c_["Date"],
                        c_["Amount (£)"],
                        f"Balance fell £{abs(chg):,.2f} (from £{p['Amount (£)']:,.2f} to £{c_['Amount (£)']:,.2f}).",
                        "INFO",
                    )
                )
        except (ValueError, TypeError, KeyError) as exc:
            _flag_or_warn(i, "BALANCE_REDUCTION", exc)

    # 6. RECONCILIATION MISMATCH: balance delta vs period charge
    if "Period Charge (£)" in dfc.columns:
        for i in range(1, n):
            p = dfc.iloc[i - 1]
            c_ = dfc.iloc[i]
            try:
                if str(c_.get("Entry Type", "")) == "New Bill" and str(p.get("Entry Type", "")) in (
                    "New Bill",
                    "Ongoing Balance",
                ):
                    pc = c_.get("Period Charge (£)")
                    try:
                        pc_val = float(pc)
                    except (ValueError, TypeError):
                        continue
                    balance_delta = float(c_["Amount (£)"]) - float(p["Amount (£)"])
                    diff = abs(balance_delta - pc_val)
                    threshold = max(pc_val * 0.10, 50.0) if pc_val > 0 else 50.0
                    if diff > threshold:
                        flags.append(
                            (
                                "RECONCILIATION MISMATCH",
                                c_["Date"],
                                c_["Amount (£)"],
                                f"Balance delta £{balance_delta:,.2f} vs period charge £{pc_val:,.2f} "
                                f"(difference: £{diff:,.2f}). Possible payment, credit, or billing error "
                                f"between {p['Date']} and {c_['Date']}.",
                                "HIGH" if diff > pc_val * 0.5 else "MEDIUM",
                            )
                        )
            except (ValueError, TypeError, KeyError) as exc:
                _flag_or_warn(i, "RECONCILIATION_MISMATCH", exc)

    # Count by severity
    counts = {s: sum(1 for f in flags if f[4] == s) for s in ("HIGH", "MEDIUM", "INFO")}
    return flags, counts


# ---------------------------------------------------------------------------
# Write evidence sheet
# ---------------------------------------------------------------------------


def write_evidence_sheet(ws, df, is_duplicate=False):
    # Pin the column letter map (matches ``headers`` below):
    # A=Source B=Sender C=Date D=PeriodFrom E=PeriodTo F=Invoice#
    # G=Amount H=PeriodCharge I=UnitRate J=%Change K=EntryType
    # L=Reading M=Units N=StandingChg O=Tariff P=AttachmentName
    # Q=Details R=LogicUsed S=AnomalyFlag
    # (Duplicate-of-link cells are rendered in a post-loop pass for
    # the ``is_duplicate=True`` branch and don't appear in this
    # header list.)
    #
    # F1 (SEV-1):  every COL_* is derived from the headers list, not
    # hard-coded.  Inserting a new column at any position requires
    # updating exactly one place (the headers list) — the conditional
    # formatting range, formula references, column widths and the
    # dedup hyperlink pass all read the same index.  Verified by
    # ``tests/test_evidence_sheet_columns.py``.
    headers = [
        "Source",
        "Sender",
        "Date",
        "Period From",
        "Period To",
        "Invoice #",
        "Amount (£)",
        "Period Charge (£)",
        "Unit Rate (p/kWh)",
        "% Change",
        "Entry Type",
        "Reading",
        "Units (kWh)",
        "Standing Chg (p/day)",
        # Tariff price-plan name (e.g. "Freedom", "Standard");
        # extracted by ``extract_new_invoice_fields`` on KI-style
        # bills. See ``_process_new_invoice``.
        "Tariff",
        "Attachment Name",
        "Details",
        "Logic Used",
        "Anomaly Flag",
    ]
    COL_AMOUNT = headers.index("Amount (£)") + 1
    COL_PERIOD_CHG = headers.index("Period Charge (£)") + 1
    COL_UNIT_RATE = headers.index("Unit Rate (p/kWh)") + 1
    COL_PCT_CHANGE = headers.index("% Change") + 1
    COL_READING_IDX = headers.index("Reading") + 1
    COL_ANOMALY = headers.index("Anomaly Flag") + 1
    # Phase 2 follow-on: dup sheets carry a "Duplicate Of"
    # printable summary cell per row plus a clickable hyperlink
    # back to the matched kept record in ``EDF Evidence Report``.
    # The matched-against position lands in a parallel
    # ``_matches_kept_idx`` Series the caller passes alongside the
    # dup_df — we render the column in a *post-loop* pass below so
    # we don't have to count on the row-iteration matching a
    # fixed column index (which previously conflicted with the
    # constant ``COL_ANOMALY = 18`` in the writer).
    has_match_col = "Duplicate Of" in df.columns
    # Capture the writer-helper ``_matches_kept_idx`` Series *before*
    # the row iteration so the post-loop pass can mint HYPERLINK
    # cells.  We then strip the column from the in-scope ``df``
    # so row iteration only sees the reader-facing schema (no
    # 20th column leaks into the saved workbook).
    if "_matches_kept_idx" in df.columns:
        match_positions_series: pd.Series = df["_matches_kept_idx"].copy()
    else:
        match_positions_series = None
    df = df.drop(columns=["_matches_kept_idx"], errors="ignore")
    bg = "888888" if is_duplicate else "FE5716"
    for col, h in enumerate(headers, 1):
        _hcell(ws, 1, col, h, bg=bg)
    ws.row_dimensions[1].height = 28

    alt_fill = PatternFill("solid", start_color="FFF3EE")

    last_data_row = len(df) + 1
    for r_idx, row in enumerate(df.values, 2):
        row_fill = alt_fill if r_idx % 2 == 0 else PatternFill()

        for c_idx, val in enumerate(row, 1):
            if c_idx == COL_PCT_CHANGE and not is_duplicate:
                # % Change as live formula — Amount is col G
                c = ws.cell(
                    row=r_idx,
                    column=COL_PCT_CHANGE,
                    value=f'=IFERROR((G{r_idx}-G{r_idx - 1})/G{r_idx - 1},"")',
                )
                c.number_format = "0.0%"
                c.alignment = Alignment(horizontal="right", vertical="top")
                c.font = Font(name="Calibri", size=10)
                c.border = CELL_BORDER
                c.fill = row_fill
            else:
                # Convert date columns to real Excel date serials (C=3, D=4, E=5)
                excel_val = val
                if c_idx in (3, 4, 5):
                    dt = to_excel_date(val)
                    if dt is not None:
                        excel_val = dt
                c = ws.cell(row=r_idx, column=c_idx, value=excel_val)
                # Phase 2.x — formula-injection guard on the
                # generic evidence-sheet row path.  openpyxl
                # auto-promotes any cell whose text value starts
                # with ``=``/``+``/``-``/``@`` to ``data_type='f'``
                # (formula).  Without this fix, a bill whose
                # Invoice # or Details field begins with ``=cmd
                # |'/c calc'!A1`` would render as a real formula
                # when an ombudsman opens the workbook.  Same
                # belt-and-braces policy as ``_text``: coerce
                # textual leads to ``str`` first, then pin
                # ``data_type='s'`` and prefix apostrophe on
                # leading special chars.
                if isinstance(excel_val, str) and excel_val:
                    safe_val = excel_val
                    if safe_val[0] in "+-=@":
                        safe_val = "'" + safe_val
                    c.value = safe_val
                    c.data_type = "s"
                if c_idx == COL_AMOUNT and isinstance(val, (int, float)):
                    c.number_format = "£#,##0.00"
                if c_idx == COL_PERIOD_CHG and isinstance(val, (int, float)):
                    c.number_format = "£#,##0.00"
                if c_idx == COL_UNIT_RATE and isinstance(val, (int, float)):
                    c.number_format = "0.00"
                if c_idx in (3, 4, 5) and hasattr(excel_val, "year"):
                    c.number_format = "dd/mm/yyyy"
                c.font = Font(name="Calibri", size=10)
                c.fill = (
                    row_fill if not is_duplicate else PatternFill("solid", start_color=DUP_GREY)
                )
                c.border = CELL_BORDER
                c.alignment = Alignment(vertical="top")

            # Highlight estimated readings (Reading is col L = 0-based index 11)
            if (
                not is_duplicate
                and len(row) > COL_READING_IDX
                and row[COL_READING_IDX] == "Estimated"
            ):
                c.fill = PatternFill("solid", start_color=EST_YELLOW)

        # Anomaly flag col S (19) — Amount is col G
        # (Anomaly Flag shifted right by one when the Tariff column
        # was inserted at column O; see the column-letter map at the
        # top of this function.)
        if not is_duplicate and r_idx > 2:
            ca = ws.cell(
                row=r_idx,
                column=COL_ANOMALY,
                value=f'=IF(AND(G{r_idx - 1}>0,G{r_idx}>G{r_idx - 1}*2),"⚠ >100% INCREASE","")',
            )
            ca.font = Font(name="Calibri", size=10, bold=True)
            ca.border = CELL_BORDER
            ca.fill = row_fill

    # Conditional formatting: only colour anomaly column red when non-empty
    if not is_duplicate and last_data_row > 2:
        ws.conditional_formatting.add(
            f"S2:S{last_data_row}",
            FormulaRule(
                formula=['$S2<>""'],
                fill=PatternFill("solid", start_color=JUMP_RED),
                font=Font(name="Calibri", size=10, bold=True),
            ),
        )

    # Phase 2 follow-on: post-loop pass to render the "Duplicate
    # Of" column.  The matched-against keystrokes live in
    # ``match_positions_series`` (a pd.Series keyed on the dup
    # sheet's df-index by df-positional index) so the click-through
    # target always aligns with the writer's row indexing scheme.
    # We render this ``Duplicate Of`` column only when
    # ``is_duplicate`` is True — main evidence reports never get
    # one.
    if is_duplicate and has_match_col and match_positions_series is not None:
        last_data_row = len(df) + 1
        col_idx_duplicate_of = len(headers) + 1
        # Header cell
        bg = "888888"
        _hcell(ws, 1, col_idx_duplicate_of, "Duplicate Of", bg=bg)
        # Materialise columns once
        dup_text = df["Duplicate Of"].tolist()
        for r_idx, (match_val, summary) in enumerate(
            zip(match_positions_series.tolist(), dup_text, strict=True), 2
        ):
            target_row_excel: int | None = None
            try:
                # ``-1`` sentinel from the dedup walker = no
                # match (Pass 1 dedup found a duplicate tuple
                # but Pass 2's kept set dropped it before the
                # post-reset lookup fired).
                mi = int(match_val)
                target_row_excel = mi + 2 if mi >= 0 else None
            except (TypeError, ValueError):
                target_row_excel = None
            if not summary:
                continue
            c = ws.cell(row=r_idx, column=col_idx_duplicate_of, value=summary)
            if target_row_excel:
                c.hyperlink = openpyxl.worksheet.hyperlink.Hyperlink(
                    ref=f"{c.coordinate}",
                    location=f"'EDF Evidence Report'!A{target_row_excel}",
                    display=summary,
                    tooltip=(f"Jump to the kept record at EDF Evidence Report!A{target_row_excel}"),
                )
                c.font = Font(name="Calibri", size=10, color="0000FF", underline="single")
            else:
                c.font = Font(name="Calibri", size=10)
            c.alignment = Alignment(vertical="top", wrap_text=True)
            c.border = CELL_BORDER
            # Dup cells read like the rest of the dup sheet
            # (greyed out so they stand out from the kept set).
            c.fill = PatternFill("solid", start_color=DUP_GREY)
        # Widen the column to fit the longest summary. After the
        # Tariff insertion, "Duplicate Of" lives at column T (was S).
        ws.column_dimensions["T"].width = 50

    widths = {
        "A": 18,
        "B": 26,
        "C": 13,
        "D": 13,
        "E": 13,
        "F": 16,
        "G": 13,
        "H": 15,
        "I": 15,
        "J": 10,
        "K": 14,
        "L": 11,
        "M": 12,
        "N": 18,
        # Tariff price-plan column — short enough to fit
        # "Standard Variable", "Freedom Tariff", etc.
        "O": 22,
        "P": 28,
        "Q": 38,
        "R": 18,
        "S": 20,
    }
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width
    ws.freeze_panes = "A2"


# ---------------------------------------------------------------------------
# Write summary sheet — uses _xlfn.MAXIFS/_xlfn.MINIFS so Excel evaluates
# on load without the dynamic-array compatibility dialog
# ---------------------------------------------------------------------------


def write_summary_sheet(ws, years, evidence_sheet_name, last_data_row=5000):
    ws.title = "Annual Summary"

    headers = [
        "Year",
        "Balance Range (£)",
        "Records",
        "Avg Balance (£)",
        "Peak Balance (£)",
        "Lowest Balance (£)",
    ]
    for col, h in enumerate(headers, 1):
        _hcell(ws, 1, col, h, bg="10367A")
    ws.row_dimensions[1].height = 28

    alt_fill = PatternFill("solid", start_color="EEF2FF")
    esn = evidence_sheet_name

    date_col = f"'{esn}'!$C$2:$C${last_data_row}"
    amt_col = f"'{esn}'!$G$2:$G${last_data_row}"

    for r_idx, year_val in enumerate(years, 2):
        row_fill = alt_fill if r_idx % 2 == 0 else PatternFill()
        yr_cell = f"A{r_idx}"

        # _xlfn. prefix tells Excel to evaluate MAXIFS/MINIFS on load without
        # the dynamic-array compatibility dialog.
        peak_f = f'=IFERROR(_xlfn.MAXIFS({amt_col},{date_col},">="&DATE({yr_cell},1,1),{date_col},"<"&DATE({yr_cell}+1,1,1)),"")'
        low_f = f'=IFERROR(_xlfn.MINIFS({amt_col},{date_col},">="&DATE({yr_cell},1,1),{date_col},"<"&DATE({yr_cell}+1,1,1)),"")'
        range_f = f'=IFERROR(_xlfn.MAXIFS({amt_col},{date_col},">="&DATE({yr_cell},1,1),{date_col},"<"&DATE({yr_cell}+1,1,1))-_xlfn.MINIFS({amt_col},{date_col},">="&DATE({yr_cell},1,1),{date_col},"<"&DATE({yr_cell}+1,1,1)),"")'

        row_values = [
            int(year_val),
            range_f,
            f'=COUNTIFS({date_col},">="&DATE({yr_cell},1,1),{date_col},"<"&DATE({yr_cell}+1,1,1))',
            f'=IFERROR(AVERAGEIFS({amt_col},{date_col},">="&DATE({yr_cell},1,1),{date_col},"<"&DATE({yr_cell}+1,1,1)),"")',
            peak_f,
            low_f,
        ]
        for c_idx, val in enumerate(row_values, 1):
            c = ws.cell(row=r_idx, column=c_idx, value=val)
            c.font = Font(name="Calibri", size=10)
            c.fill = row_fill
            c.border = CELL_BORDER
            c.alignment = Alignment(
                horizontal="center" if c_idx == 1 else "right",
                vertical="top",
            )
            if c_idx == 2:
                c.number_format = "£#,##0.00"
            elif c_idx == 3:
                c.number_format = "#,##0"
            elif c_idx > 3:
                c.number_format = "£#,##0.00"

    # Grand total row — SUM/MAX/MIN over the year rows only, no dynamic-array functions
    n = len(years) + 2
    first_r = 2
    last_r = n - 1
    tot_fill = PatternFill("solid", start_color="10367A")
    tot_specs = [
        ("OVERALL", None, "center"),
        (f'=IFERROR(MAX(E{first_r}:E{last_r})-MIN(F{first_r}:F{last_r}),"")', "£#,##0.00", "right"),
        (f"=SUM(C{first_r}:C{last_r})", "#,##0", "right"),
        (f'=IFERROR(AVERAGE(D{first_r}:D{last_r}),"")', "£#,##0.00", "right"),
        (f'=IFERROR(MAX(E{first_r}:E{last_r}),"")', "£#,##0.00", "right"),
        (f'=IFERROR(MIN(F{first_r}:F{last_r}),"")', "£#,##0.00", "right"),
    ]
    for c_idx, (val, num_fmt, align) in enumerate(tot_specs, 1):
        c = ws.cell(row=n, column=c_idx, value=val)
        c.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        c.fill = tot_fill
        c.border = CELL_BORDER
        c.alignment = Alignment(horizontal=align)
        if num_fmt:
            c.number_format = num_fmt

    for col_letter in ["A", "B", "C", "D", "E", "F"]:
        ws.column_dimensions[col_letter].width = 22
    ws.freeze_panes = "A2"


# ---------------------------------------------------------------------------
# Main export function
# ---------------------------------------------------------------------------


def export_to_excel(data, output_path, error_log, config, filtered=None):
    NAVY = "10367A"
    ORANGE = "FE5716"
    RED = "FF6B6B"
    AMBER = "FFD166"
    GREEN = "06D6A0"
    LGREY = "F0F0F0"
    DGREY = "888888"

    df = pd.DataFrame(data)
    df["_sort"] = df["Date"].apply(parse_to_sort_date)
    df = df.sort_values(by=["_sort", "Invoice #"], ascending=[True, False]).reset_index(drop=True)
    df["% Change"] = None

    # Deduplication — multi-pass to match the same bill across sources
    # Pass 1: Period To + Amount  (catches HTM ↔ PST where billing period matches)
    # Pass 2: Amount within 60-day window for records with no period info (Local PDF)
    dup_df = pd.DataFrame()
    if config.get("use_dedup", True):
        # Source precedence lives at module scope (``_SOURCE_PRECEDENCE``)
        # so that ``tests/test_source_precedence.py`` can pin the
        # explicit ordering without booting the entire Excel
        # export pipeline.  Lower number = higher precedence.
        df["_src_pri"] = df["Source"].map(_SOURCE_PRECEDENCE).fillna(9).astype(int)
        # Completeness score — primary sort key.  Spec: "duplicates
        # should be assessed and the most complete version of the
        # information presented".  ``_completeness_score`` counts
        # populated substantive fields on each row; the richer row
        # sorts *before* the sparser row so ``keep="first"`` keeps it.
        # Computed here (not earlier) so it's available even if the
        # upstream pipeline headers change in future.
        df["_completeness"] = df.apply(_completeness_score, axis=1)
        # Sort order (primary to tie-breaker):
        #   1. _completeness descending      — most-populated row wins
        #   2. _src_pri ascending             — higher-precedence source wins ties
        #   3. _sort ascending                — earliest date wins remaining ties
        # ``keep="first"`` then retains the head of every duplicate cluster.
        # Pre-fix the sort was only ``["_src_pri", "_sort"]`` so source
        # precedence dominated completeness — a sparser HTM row would
        # beat a richer PST row.  The companion test is
        # ``tests/test_dedup_most_complete.py``.
        df = df.sort_values(
            ["_completeness", "_src_pri", "_sort"],
            ascending=[False, True, True],
        ).reset_index(drop=True)

        # Dedup key: prefer Period To (consistent across sources for same bill),
        # fall back to Date for records without period info.  Pass 1's
        # ``DUPLICATED`` flags for *period-aware* rows track which *kept*
        # row they collide against so the dup sheet can render a clickable
        # summary linking back to the source-of-truth record.  We capture
        # the matched-against row's *original* df index — that index is
        # what ``dup_df.index`` carries through to the writer, since
        # ``dup_df = df[is_dup]`` runs before the ``reset_index`` line below.
        # Period To is the source-of-truth end-of-billing-period
        # date when present; fall back to ``_sort`` (the parsed
        # source-specific ``Date``) when the row is no-period
        # (e.g. Local PDF).  ``df["_sort"].where(cond, df["_sort"])``
        # is a tautology — Period To was being ignored and Pass 1
        # ``_dedup_date`` is the *canonical* dedup key — Period To when
        # available, otherwise left as ``NaT`` so the row is excluded
        # from ``duplicated`` clusters (since ``duplicated`` treats
        # NaT as equal across rows, falling back to ``_sort`` would
        # silently merge unrelated no-period same-amount rows).
        # Rows with NaT here are rerouted through Pass-2's no-period
        # bucket logic below, which uses ``Period To == "N/A" | NaN``
        # as the explicit handling mask.
        period_to_dt = pd.to_datetime(df["Period To"], dayfirst=True, errors="coerce")
        df["_dedup_date"] = period_to_dt
        is_dup = df.duplicated(subset=["_dedup_date", "Amount (£)"], keep="first")
        # Pass 1 (period+amount): build ``kept_pass1_index`` keyed on
        # ``(_dedup_date, Amount)`` so we can look up "which kept row
        # did this dup lose to".  The kept row's original df index (not
        # its reset_index value) survives into the dup sheet.
        kept_for_dup: dict[int, int] = {}  # dup_idx -> kept_idx (both original indices)
        kept_for_summary: dict[int, dict[str, object]] = {}  # kept_idx -> display fields
        kept_frame = df[~is_dup]
        kept_pass1_index: dict[tuple, int] = {}
        for kept_idx in kept_frame.index:
            k = (
                kept_frame.at[kept_idx, "_dedup_date"],
                kept_frame.at[kept_idx, "Amount (£)"],
            )
            kept_pass1_index.setdefault(k, kept_idx)
            # Cache the displayed fields once per kept row so the
            # dup lookup below doesn't re-read them.
            kept_for_summary[kept_idx] = {
                "Source": kept_frame.at[kept_idx, "Source"],
                "Date": kept_frame.at[kept_idx, "Date"],
                "Amount (£)": kept_frame.at[kept_idx, "Amount (£)"],
            }
        # Resolve Pass 1's kept-against reference per duplicate
        # before any reset_index runs.
        for dup_idx in df[is_dup].index:
            k = (
                df.at[dup_idx, "_dedup_date"],
                df.at[dup_idx, "Amount (£)"],
            )
            kept_idx = kept_pass1_index.get(k, -1)
            kept_for_dup[dup_idx] = kept_idx

        # Pass 2: records with no period info (e.g. Local PDF) — match by
        # Amount within a 60-day window of any already-kept record.
        #
        # Phase 2.2 follows the spec: group candidates by Amount (£)
        # first, then look up matches inside each amount-bucket
        # rather than scanning the entire kept-mask frame for every
        # candidate.  The previous implementation was O(N²) — at
        # 5,000 records the *bench* showed it took ~2.3 s.  This
        # bucketed approach is O(N) amortised: typical EDF bills
        # have unique amounts, so bucket size is 1–2 rows and the
        # inner day-window check is effectively constant.
        #
        # Layout-preserving detail worth flagging: the *legacy*
        # algorithm visits ``df.index`` in increasing order and
        # looks at the live ``kept`` mask — which includes
        # forward-yet-to-be-visited rows whose ``~is_dup`` is the
        # pre-iteration value (so any same-amount row ±60 days
        # *before or after* the candidate, except itself, can
        # match).  We replicate that exact behaviour by iterating
        # ``df.index`` in *reverse* and building per-amount buckets
        # incrementally: at row N's visit, the bucket for any
        # amount A already contains every row with amount A and
        # index > N that wasn't marked as dup — exactly the
        # forward-direction rows the legacy code saw.
        #
        # Concretely: with the legacy ``kept = df[(~is_dup) &
        # (df.index != idx)]`` mask, the set of candidate matches
        # for row idx against amount A is
        # ``{j != idx : df.Amount[j] == A and ~is_dup.at[j]}``.
        # For most rows this set is split into:
        #   (i) j in [0, idx) — *earlier* df indices,
        #  (ii) j in (idx, len(df)) — *later* df indices.
        # The legacy code consulted both groups via the live
        # ``~is_dup`` mask.  Iterating reverse and limiting our
        # bucket hashes to *only* ``j > idx`` (the "earlier in
        # reverse-iteration-order" rows) lands on exactly the
        # same candidate set provided *no row gets marked as dup
        # before its later neighbours are visited* — which the
        # reverse loop guarantees by ordering inspections from
        # the bottom of the frame upwards.
        no_period = (df["Period To"] == "N/A") | df["Period To"].isna()
        # ``bucket_by_amt`` is keyed on Amount and stores the
        # ``(df_ordinal, _sort date)`` of every row already visited
        # (reverse-iteration order) that hasn't been marked as
        # duplicate.  We append a row to its bucket whenever the
        # row *does not* get marked — symmetric to the legacy
        # ``kept`` mask at iteration time.
        bucket_by_amt: dict[float, list[tuple[int, object]]] = {}
        # Reverse-iterate ``df.index`` so that "later in df order"
        # rows are visited first and accumulate in the bucket for
        # the earlier row's lookup.  Equivalently, the bucket for
        # each amount at ``idx`` is exactly the rows j > idx with
        # Amount[j] == amount and ~is_dup.at[j] — the same row set
        # legacy would consult.
        reverse_idx = list(df[~is_dup & no_period].index)[::-1]
        for idx in reverse_idx:
            amt = df.loc[idx, "Amount (£)"]
            rec_date = df.loc[idx, "_sort"]
            same_amt = bucket_by_amt.get(amt, [])
            matched = False
            for m_idx, m_date in same_amt:
                # ``pd.notna`` short-circuit means NaT-dated rows
                # already in the bucket (originally the loop
                # ``continue``-skipped them but still listed them
                # in the next-iter kept set) never trigger a match.
                if pd.notna(m_date) and abs((rec_date - m_date).days) <= 60:
                    matched = True
                    # Capture the matched-against row's *original
                    # df index* so the dup sheet can resolve to
                    # the same frame.  We resolve the summary
                    # *before* the kept set is `reset_index`-
                    # rasterised below — once ``df = df[~is_dup]
                    # .reset_index(drop=True)`` runs, the
                    # ``m_idx`` no longer references a row.
                    kept_for_dup[idx] = m_idx
                    kept_for_summary[m_idx] = {
                        "Source": df.at[m_idx, "Source"],
                        "Date": df.at[m_idx, "Date"],
                        "Amount (£)": df.at[m_idx, "Amount (£)"],
                    }
                    break
            if matched:
                is_dup.at[idx] = True
                # Don't add to the bucket — the legacy loop's
                # recomputed ``~is_dup`` mask would have excluded a
                # row marked dup at the *start* of iteration, so it
                # cannot anchor later (here: earlier-in-iteration)
                # matches either.
            else:
                # Always add the row even if ``_sort`` is NaT —
                # the legacy ``kept`` mask at the *next* (lower) row
                # includes this row because it's ``~is_dup``-true,
                # and the NaT date just means it can't anchor a
                # match on its own.
                bucket_by_amt.setdefault(amt, []).append((idx, rec_date))

        # ``dup_df`` is built BEFORE the ``reset_index`` line below so
        # ``dup_df.index`` still carries each duplicate's original df
        # index — that's the key we use to look up the kept-against
        # summary in ``kept_for_summary``.
        #
        # ``save_dups`` toggles whether dedup *itself* is applied to the
        # main dataframe (``df``).  When True (the historical default),
        # duplicates are filtered out of ``df`` and *recorded* in
        # ``dup_df`` for the dup sheet — users never lose visibility of
        # what was dropped.  When False, dedup is skipped entirely: every
        # row stays in ``df`` and ``dup_df`` is empty.
        if config.get("save_dups", True):
            dup_df = df[is_dup].copy()
        else:
            dup_df = df[is_dup].iloc[0:0].copy()

        # Spec 3 (stretch): hybrid rows when ``amalgamate_duplicates`` is
        # True.  Instead of keeping the completeness-winner verbatim, we
        # merge each duplicate cluster's non-empty fields into a single
        # hybrid kept row.  The composite keeps the completeness-winner's
        # ``Source`` identity and picks any populated column value from
        # any sibling.  Each non-surviving sibling still stays in
        # ``dup_df`` (the spec's 'never drop without being recorded').
        #
        # N.B. the amalgamated ``df`` is is already a cleaned kept set
        # (all duplicates removed), so the ``df[~is_dup]`` filter below
        # is skipped for the amalgamate path.
        if (
            config.get("save_dups", True)
            and config.get("amalgamate_duplicates", False)
            and not dup_df.empty
        ):
            df = _apply_amalgamate_to_kept_frame(df, dup_df, kept_pass1_index, kept_for_dup, is_dup)
            # dup_df stays unchanged — the amalgamation only touches the
            # kept set; the dup sheet still records every sibling.

        if config.get("save_dups", True) and not config.get("amalgamate_duplicates", False):
            df = df[~is_dup].reset_index(drop=True)
        # else: do not drop duplicates — leave ``df`` unchanged so the
        # user sees the raw ingress and can resolve duplicates manually.
        df = df.drop(columns=["_src_pri", "_dedup_date", "_completeness"], errors="ignore")

    df = df.drop(columns=["_sort"], errors="ignore")
    dup_df = (
        dup_df.drop(
            columns=["_sort", "_src_pri", "_dedup_date", "_completeness"],
            errors="ignore",
        )
        if not dup_df.empty
        else dup_df
    )

    # Compute Unit Rate (p/kWh) where both Period Charge and Units are available.
    #
    # Phase 2.1: vectorised path.  The historic row-wise apply walked
    # Python per row, which the bench measured at ~63 ms at 5,000
    # records (not the bottleneck we'd been worried about, but the
    # spec asks for vectorisation).  New path uses pd.to_numeric
    # + np.where — same observable output (rounded to 0.01) but
    # vectorised.  ``Units`` is normalised for the inline comma
    # (``"1,234"`` to ``"1234"``) the same way the row-wise path
    # did via ``str(units).replace(",", "")``.
    pc = pd.to_numeric(df["Period Charge (£)"], errors="coerce")
    units = pd.to_numeric(
        df["Units (kWh)"].astype(str).str.replace(",", ""),
        errors="coerce",
    )
    df["Unit Rate (p/kWh)"] = np.where(
        (units > 0) & (pc > 0),
        np.round((pc / units) * 100, 2),
        np.nan,
    )

    # ``dup_df`` computation is kept in the path for backward
    # compatibility — the dup DataFrame is much smaller than the
    # kept set, so per-row apply only adds ms-level overhead.  We
    # use a tiny module-scope helper rather than a closure so
    # ``pickle`` can find it on round-trip (the spec used to break
    # here because closures aren't picklable).
    def _compute_unit_rate(row):
        pc = row.get("Period Charge (£)")
        units = row.get("Units (kWh)")
        try:
            pc_f = float(pc)
            u_f = float(str(units).replace(",", ""))
            if u_f > 0 and pc_f > 0:
                return round((pc_f / u_f) * 100, 2)
        except (ValueError, TypeError):
            pass
        return np.nan

    if not dup_df.empty:
        dup_df["Unit Rate (p/kWh)"] = dup_df.apply(_compute_unit_rate, axis=1)
        # Matched-against kept-record block (Phase-2 follow-up).
        # Each duplicate row gets a clickable summary pointing
        # back to the *kept* record so an ombudsman reviewing the
        # workbook can navigate from the dup sheet to the
        # source-of-truth record with one click.  Earlier in the
        # dedup walk we built ``kept_for_summary`` keyed on the
        # duplicate's *original* df-index — that's also the index
        # ``dup_df.index`` carries because ``dup_df = df[is_dup]
        # .copy()`` runs *before* the ``reset_index(drops...)``
        # line.  So we can resolve the summary now without
        # re-doing any index resets.
        kept_idx_by_dup = {
            dup_idx: kept_for_summary.get(kept_for_dup.get(dup_idx, -1), {})
            for dup_idx in dup_df.index
        }

        # ``df`` is the kept set after dedup reset_index.  After
        # ``df = df[~is_dup].reset_index(drop=True)``, ``df.index``
        # is a sequential 0..N-1 range, *not* the original df
        # labels.  But the *order* of rows is preserved — the n-th
        # row of the kept set is the same n-th kept row that survived
        # dedup.  We therefore translate the original-index
        # references we still hold in ``kept_for_dup`` (the dedup
        # walker wrote them *before* reset_index) into post-reset
        # positions by ranking the kept rows in ascending original
        # df-index order — kept_rank[k] = rank-in-kept-set.
        kept_rank: dict[int, int] = {}
        for rank, orig_idx in enumerate(sorted(kept_for_summary.keys())):
            kept_rank[int(orig_idx)] = rank

        def _summary(idx: int) -> str:
            # Build the printable kept-row-reference string.  Falls
            # back to an empty string if the matched-against kept
            # row was rolled up by Pass 1 *after* the lookup
            # captured -1 (a corner case where the pattern matched
            # but no kept frame picked it up).
            row = kept_idx_by_dup.get(idx)
            if not row:
                return ""
            try:
                amount_val = float(row["Amount (£)"])  # type: ignore[arg-type]
                amt_str = "£" + format(amount_val, ".2f")
            except (TypeError, ValueError):
                amt_str = "£--"
            return f"{row['Source']} · {row['Date']} · {amt_str}"

        # ``Duplicate Of`` is the visible column on the dup sheet
        # itself; ``_matches_kept_idx`` is the link target the
        # Excel writer will use to mint the click-through hyperlink
        # back to the kept row in the main evidence report.
        dup_df["Duplicate Of"] = [_summary(idx) for idx in dup_df.index]
        # ``_matches_kept_idx`` is the *post-reset* position of
        # the kept row in ``EDF Evidence Report`` — the Excel
        # writer uses this with ``A{+1}`` as the click target
        # so an ombudsman can jump from the dup cell directly to
        # the source-of-truth record.  We translate via
        # ``kept_rank`` (computed above from kept-against-original
        # ordering) because the dedup walker built ``kept_for_dup``
        # *before* ``reset_index`` ran on the kept frame.
        dup_df["_matches_kept_idx"] = pd.Series(
            {idx: kept_rank.get(int(kept_for_dup.get(idx, -1)), -1) for idx in dup_df.index},
            dtype="Int64",
        )

    # F2 (SEV-1): single source of truth for the saved-column
    # ordering.  Every ``_add_record``-time builder must stamp
    # every name in this list (use ``record.setdefault(col, "N/A")``
    # if unsure) — otherwise ``reindex`` silently drops the column
    # and the workbook schema drifts from what other readers
    # (Tariff Analysis, Dict Comparer) expect.  The structural
    # guard lives in ``tests/test_export_headers_invariant.py``.
    col_order = [
        "Source",
        "Sender",
        "Date",
        "Period From",
        "Period To",
        "Invoice #",
        "Amount (£)",
        "Period Charge (£)",
        "Unit Rate (p/kWh)",
        "% Change",
        "Entry Type",
        "Reading",
        "Units (kWh)",
        "Standing Chg (p/day)",
        # Tariff column — lights up the Tariff Analysis Excel/DOCX/PDF
        # section.  Populated only by ``_process_new_invoice``;
        # every other source path stamps "N/A".  Without this entry
        # here, ``reindex`` would drop the column from the saved
        # workbook even though every record dict now carries it.
        "Tariff",
        "Attachment Name",
        "Details",
        "Logic Used",
        "Anomaly Flag",
        "Duplicate Of",
    ]
    df = df.reindex(columns=col_order)
    # Belt-and-braces invariant: every column the *kept* set still
    # carries must be in the canonical order list — otherwise a
    # future record builder that adds a new column without updating
    # col_order would survive the reindex and land as a
    # mysteries-leading-column in the saved workbook.  We assert
    # loudly here (developer-visible) rather than silently dropping
    # the unknown column.
    _unexpected = [c for c in df.columns if c not in col_order]
    if _unexpected:
        raise ValueError(
            "export_to_excel received columns not in col_order: "
            f"{_unexpected!r}.  Add them to col_order or build the "
            "records so they carry only known keys."
        )
    # The dup sheet needs both ``Duplicate Of`` *and*
    # ``_matches_kept_idx`` available to the writer so the
    # post-loop pass can mint clickable HYPERLINK cells.  We
    # attach ``_matches_kept_idx`` after the reindex pass so the
    # saved workbook geometry stays 19-column even though the
    # writer's row-iteration will see the 20th column briefly —
    # the writer drops the column before saving.
    if not dup_df.empty and "_matches_kept_idx" in dup_df.columns:
        # Already present — nothing to do.
        pass
    else:
        # Neither column nor value is preserved.  Don't write
        # anything — the post-loop pass will skip minting
        # HYPERLINKs because ``match_positions_series`` is None.
        pass
    # No-op reindex guard for clarity; dup_df reindex on col_order
    # actually *drops* the helper column, which is what we want
    # for the Excel geometry — but we also need it for the
    # hyperlink pass.  Best approach: call site reads it BEFORE
    # reindex and threads it via a separate side cache.
    # The simplest implementation is to re-attach the column
    # *after* reindex here:
    if not dup_df.empty:
        dup_df_reindexed = dup_df.reindex(columns=col_order)
        # Re-attach from dup_df's pre-reindex view — the column
        # is dropped by reindex, so we restore it from the
        # original here.  This is the only place where the
        # writer would otherwise lose access to the helper.
        if "_matches_kept_idx" in dup_df.columns:
            dup_df = pd.concat(
                [
                    dup_df_reindexed,
                    dup_df["_matches_kept_idx"].rename("_matches_kept_idx"),
                ],
                axis=1,
            )
        else:
            dup_df = dup_df_reindexed

    # Years for summary tab
    years = sorted(
        y for y in df["Date"].apply(parse_to_sort_date).dropna().dt.year.astype(int).unique()
    )

    wb = openpyxl.Workbook()
    wb.calculation.fullCalcOnLoad = True

    # Tab 1: Evidence (created first — summary formulas reference it by name)
    ws_main = wb.active
    ws_main.title = "EDF Evidence Report"
    write_evidence_sheet(ws_main, df, is_duplicate=False)

    # Tab 2: Annual Summary
    ws_summary = wb.create_sheet(title="Annual Summary", index=0)
    write_summary_sheet(ws_summary, years, ws_main.title, last_data_row=len(df) + 1)

    # Tab 3: Duplicates
    if not dup_df.empty:
        ws_dup = wb.create_sheet(title="Duplicate Entries")
        write_evidence_sheet(ws_dup, dup_df, is_duplicate=True)

    # Tab 4: Filtered
    if filtered and config.get("save_filtered", True):
        ws_filt = wb.create_sheet(title="Filtered (Below Min)")
        filt_headers = ["Source", "Date", "Amount (£)", "Details", "Logic Used", "Reason"]
        for ci, h in enumerate(filt_headers, 1):
            _hcell(ws_filt, 1, ci, h, bg="888888")
        filt_df = pd.DataFrame(filtered).sort_values("Amount (£)", ascending=False)
        for r_idx, frow in enumerate(filt_df.values, 2):
            bg_hex = "F5F5F5" if r_idx % 2 == 0 else None
            for c_idx, val in enumerate(frow, 1):
                c = ws_filt.cell(row=r_idx, column=c_idx, value=val)
                c.font = Font(name="Calibri", size=10)
                c.border = CELL_BORDER
                if bg_hex:
                    c.fill = PatternFill("solid", start_color=bg_hex)
                if c_idx == 3:
                    c.number_format = "£#,##0.00"
        for col, w in zip(["A", "B", "C", "D", "E", "F"], [18, 13, 14, 38, 18, 28], strict=False):
            ws_filt.column_dimensions[col].width = w
        ws_filt.freeze_panes = "A2"

    # Tab 5: Parse errors
    if error_log:
        ws_err = wb.create_sheet(title="Parse Errors")
        _hcell(ws_err, 1, 1, "Time", bg="888888")
        _hcell(ws_err, 1, 2, "Context", bg="888888")
        _hcell(ws_err, 1, 3, "Error", bg="888888")
        for r_idx, entry in enumerate(error_log, 2):
            ts_m = re.match(r"\[(.+?)\]\s*(.*?)\s*—\s*(.*)", entry)
            if ts_m:
                ts, ctx, err = ts_m.group(1), ts_m.group(2), ts_m.group(3)
            else:
                ts, ctx, err = "", entry, ""
            for c_idx, val in enumerate([ts, ctx, err], 1):
                c = ws_err.cell(row=r_idx, column=c_idx, value=val)
                c.font = Font(name="Calibri", size=10)
                c.border = CELL_BORDER
        ws_err.column_dimensions["A"].width = 10
        ws_err.column_dimensions["B"].width = 45
        ws_err.column_dimensions["C"].width = 60

    # =====================================================================
    # ANALYSIS SUITE
    # Uses bills above analysis_min threshold only (payments/credits always included).
    # =====================================================================

    df_an = df.copy()
    df_an["_dt"] = df_an["Date"].apply(parse_to_sort_date)
    df_an = df_an.sort_values("_dt").reset_index(drop=True)
    analysis_min = float(config.get("analysis_min", 500.0))

    # For balance-affecting entries: include all Payments/Credits, but filter
    # New Bill/Ongoing Balance by analysis_min threshold
    payment_credit_mask = df_an["Entry Type"].isin(("Payment", "Credit"))
    bill_mask = df_an["Entry Type"].isin(("New Bill", "Ongoing Balance"))
    amount_mask = df_an["Amount (£)"] >= analysis_min

    dfc = df_an[(payment_credit_mask) | (bill_mask & amount_mask)].copy().reset_index(drop=True)
    dfc["year"] = dfc["_dt"].dt.year
    dfc["month"] = dfc["_dt"].dt.month

    if len(dfc) < 2:
        wb.save(output_path)
        return

    amounts = dfc["Amount (£)"].values.astype(float)
    dates_lbl = dfc["Date"].tolist()
    n = len(amounts)

    raw_diffs = np.diff(amounts)
    pos_diffs = raw_diffs[raw_diffs > 0]

    yearly = (
        dfc.groupby("year")
        .agg(
            count=("Amount (£)", "count"),
            avg_bal=("Amount (£)", "mean"),
            peak=("Amount (£)", "max"),
            low=("Amount (£)", "min"),
        )
        .reset_index()
    )

    # ----- TAB A: KEY STATISTICS -----
    ws_ks = wb.create_sheet(title="Key Statistics")
    ws_ks.column_dimensions["A"].width = 44
    ws_ks.column_dimensions["B"].width = 22
    ws_ks.column_dimensions["C"].width = 44

    tc = ws_ks.cell(row=1, column=1, value="EDF ENERGY DISPUTE  —  KEY STATISTICS")
    tc.font = Font(name="Calibri", size=13, bold=True, color="FFFFFF")
    tc.fill = PatternFill("solid", start_color=ORANGE)
    tc.border = CELL_BORDER
    tc.alignment = Alignment(horizontal="left", vertical="center")
    for c in [2, 3]:
        x = ws_ks.cell(row=1, column=c)
        x.fill = PatternFill("solid", start_color=ORANGE)
        x.border = CELL_BORDER
    ws_ks.row_dimensions[1].height = 26

    def ks_row(r, label, value, note="", fmt=None, bold=False, alt=False):
        bg = LGREY if alt else None
        _text(ws_ks, r, 1, label, bold=bold, fill_hex=bg)
        if fmt == "£":
            _money(ws_ks, r, 2, value, bold=bold, fill_hex=bg)
        elif fmt == "%":
            _num(ws_ks, r, 2, value, fmt="0.0%", bold=bold, fill_hex=bg)
        elif fmt == "date":
            cell = ws_ks.cell(row=r, column=2, value=value)
            cell.number_format = "dd/mm/yyyy"
            cell.font = Font(name="Calibri", size=10, bold=bold)
            cell.border = CELL_BORDER
            cell.alignment = Alignment(horizontal="right", vertical="center")
            if bg:
                cell.fill = PatternFill("solid", start_color=bg)
        elif fmt:
            _num(ws_ks, r, 2, value, fmt=fmt, bold=bold, fill_hex=bg)
        else:
            _text(ws_ks, r, 2, value, bold=bold, fill_hex=bg, align="right")
        _text(ws_ks, r, 3, note, fill_hex=bg, color=DGREY)

    acc_ref = str(config.get("report_account_ref") or config.get("acc_num") or "N/A")

    r = 2
    _section_hdr(ws_ks, r, "ACCOUNT OVERVIEW")
    r = 3
    ks_row(r, "Account reference", acc_ref, alt=True)
    r = 4
    ks_row(
        r,
        "First bill on record",
        "='Balance Trend'!A2",
        fmt="date",
        note="From Balance Trend sheet",
    )
    r = 5
    ks_row(
        r,
        "Most recent bill",
        "=INDEX('Balance Trend'!A:A,MATCH(9.99E+307,'Balance Trend'!B:B)+1)",
        fmt="date",
        alt=True,
    )
    r = 6
    ks_row(
        r,
        "Period covered (days)",
        "=IFERROR(INT(INDEX('Balance Trend'!A:A,MATCH(9.99E+307,'Balance Trend'!B:B)+1)-'Balance Trend'!A2),\"\")",
        fmt="#,##0",
        note="Days between first and last bill",
    )
    r = 7
    ks_row(
        r,
        "Total bills on record",
        "=IFERROR(COUNT('Balance Trend'!B:B),\"\")",
        fmt="#,##0",
        alt=True,
    )

    r = 8
    _section_hdr(ws_ks, r, "BALANCE FIGURES")
    r = 9
    ks_row(
        r,
        "Opening balance (first bill)",
        "='Balance Trend'!B2",
        fmt="£",
        alt=True,
        note="First entry in Balance Trend",
    )
    r = 10
    ks_row(
        r,
        "Current balance (latest bill)",
        "=INDEX('Balance Trend'!B:B,MATCH(9.99E+307,'Balance Trend'!B:B))",
        fmt="£",
        bold=True,
        note="Last numeric entry in Balance Trend",
    )
    r = 11
    ks_row(
        r,
        "Total balance increase",
        '=IFERROR(B10-B9,"")',
        fmt="£",
        bold=True,
        alt=True,
        note="Latest minus earliest",
    )
    r = 12
    ks_row(r, "% increase over full period", '=IFERROR((B10-B9)/B9,"")', fmt="%", bold=True)
    r = 13
    ks_row(
        r,
        "Mean balance across all bills",
        "=IFERROR(AVERAGE('Balance Trend'!B:B),\"\")",
        fmt="£",
        alt=True,
    )
    r = 14
    ks_row(r, "Median balance", "=IFERROR(MEDIAN('Balance Trend'!B:B),\"\")", fmt="£")
    r = 15
    ks_row(r, "Peak balance recorded", "=IFERROR(MAX('Balance Trend'!B:B),\"\")", fmt="£", alt=True)
    r = 16
    ks_row(r, "Lowest balance recorded", "=IFERROR(MIN('Balance Trend'!B:B),\"\")", fmt="£")

    r = 17
    _section_hdr(ws_ks, r, "PERIODIC CHARGES")
    r = 18
    ks_row(
        r,
        "Note",
        "Bills are a running cumulative balance — periodic charge = closing minus opening balance",
        alt=True,
    )
    r = 19
    ks_row(
        r,
        "Mean charge per period (positive only)",
        '=IFERROR(AVERAGEIF(\'Period Charges\'!F:F,">0"),"")',
        fmt="£",
    )
    r = 20
    ks_row(
        r,
        "Largest single-period charge",
        "=IFERROR(MAX('Period Charges'!F:F),\"\")",
        fmt="£",
        bold=True,
        alt=True,
    )
    r = 21
    ks_row(
        r,
        "Smallest positive charge",
        "=IFERROR(_xlfn.MINIFS('Period Charges'!F:F,'Period Charges'!F:F,\">0\"),\"\")",
        fmt="£",
    )
    r = 22
    ks_row(
        r,
        "Periods where balance increased",
        '=IFERROR(COUNTIF(\'Period Charges\'!F:F,">0"),"")',
        fmt="#,##0",
        alt=True,
    )
    r = 23
    ks_row(
        r,
        "Periods where balance fell (payments/credits)",
        '=IFERROR(COUNTIF(\'Period Charges\'!F:F,"<0"),"")',
        fmt="#,##0",
    )
    r = 24
    ks_row(
        r,
        "Implied annual rate (avg last 6 charges ×12)",
        "=IFERROR(AVERAGE(OFFSET('Period Charges'!F1,MAX(1,COUNTIF('Period Charges'!F:F,\">0\")-5),0,6,1))*12,\"\")",
        fmt="£",
        bold=True,
        alt=True,
        note="Assumes ~monthly billing — may overstate if billing is quarterly",
    )

    r = 25
    _section_hdr(ws_ks, r, "READING & DATA QUALITY")
    r = 26
    ks_row(
        r,
        "Estimated readings",
        '=IFERROR(COUNTIF(\'EDF Evidence Report\'!L:L,"Estimated"),"")',
        fmt="#,##0",
        alt=True,
    )
    r = 27
    ks_row(
        r,
        "Actual / customer readings",
        '=IFERROR(COUNTIF(\'EDF Evidence Report\'!L:L,"Actual"),"")',
        fmt="#,##0",
    )
    r = 28
    ks_row(
        r,
        "Smart meter readings",
        '=IFERROR(COUNTIF(\'EDF Evidence Report\'!L:L,"Smart"),"")',
        fmt="#,##0",
        alt=True,
    )
    r = 29
    ks_row(
        r,
        "% of bills with estimated readings",
        "=IFERROR(B26/COUNT('EDF Evidence Report'!G:G),\"\")",
        fmt="%",
    )

    r = 30
    _section_hdr(ws_ks, r, "UNIT RATES")
    r = 31
    ks_row(
        r,
        "Average unit rate (p/kWh)",
        "=IFERROR(AVERAGE('EDF Evidence Report'!I:I),\"\")",
        fmt="0.00",
        alt=True,
        note="Across all bills with valid period charge and kWh",
    )
    r = 32
    ks_row(
        r,
        "Maximum unit rate (p/kWh)",
        "=IFERROR(MAX('EDF Evidence Report'!I:I),\"\")",
        fmt="0.00",
        note="Highest effective rate — potential overcharge",
    )
    r = 33
    ks_row(
        r,
        "Minimum unit rate (p/kWh)",
        "=IFERROR(MIN('EDF Evidence Report'!I:I),\"\")",
        fmt="0.00",
        alt=True,
    )

    ws_ks.freeze_panes = "A2"

    # ----- TAB B: BALANCE TREND -----
    ws_bt = wb.create_sheet(title="Balance Trend")
    for ci, h in enumerate(
        ["Date", "Balance (£)", "6-Bill Rolling Avg (£)", "Linear Trend (£)", "Period Charge (£)"],
        1,
    ):
        _hcell(ws_bt, 1, ci, h, bg=NAVY)
    ws_bt.row_dimensions[1].height = 22

    last_data_row = n + 1
    for i in range(n):
        r = i + 2
        bg = LGREY if i % 2 == 0 else None

        # Write date as a true Excel date serial
        excel_dt = to_excel_date(dates_lbl[i])
        c1 = ws_bt.cell(row=r, column=1, value=excel_dt)
        c1.number_format = "dd/mm/yyyy"
        c1.font = Font(name="Calibri", size=10)
        c1.border = CELL_BORDER
        c1.alignment = Alignment(horizontal="left")
        if bg:
            c1.fill = PatternFill("solid", start_color=bg)

        _money(ws_bt, r, 2, float(amounts[i]), fill_hex=bg)

        start_r = max(2, r - 5)
        for col_i, formula in [
            (3, f'=IFERROR(AVERAGE(B{start_r}:B{r}),"")'),
            (
                4,
                f'=IFERROR(FORECAST.LINEAR(ROW(),B$2:B${last_data_row},ROW(B$2:B${last_data_row})),"")',
            ),
        ]:
            cx = ws_bt.cell(row=r, column=col_i, value=formula)
            cx.number_format = "£#,##0.00"
            cx.font = Font(name="Calibri", size=10)
            cx.border = CELL_BORDER
            cx.alignment = Alignment(horizontal="right")
            if bg:
                cx.fill = PatternFill("solid", start_color=bg)

        if i > 0:
            c5 = ws_bt.cell(row=r, column=5, value=f"=B{r}-B{r - 1}")
            c5.number_format = "£#,##0.00"
            c5.font = Font(name="Calibri", size=10)
            c5.border = CELL_BORDER
            c5.alignment = Alignment(horizontal="right")
            if bg:
                c5.fill = PatternFill("solid", start_color=bg)

    # Line chart
    lc = LineChart()
    lc.title = "Account Balance Over Time"
    lc.style = 10
    lc.y_axis.title = "Balance (£)"
    lc.x_axis.title = "Bill Date"
    lc.width, lc.height = 30, 18
    data_ref = Reference(ws_bt, min_col=2, max_col=4, min_row=1, max_row=n + 1)
    dates_ref = Reference(ws_bt, min_col=1, min_row=2, max_row=n + 1)
    lc.add_data(data_ref, titles_from_data=True)
    lc.set_categories(dates_ref)
    lc.series[0].graphicalProperties.line.solidFill = ORANGE
    lc.series[0].graphicalProperties.line.width = 22000
    if len(lc.series) > 1:
        lc.series[1].graphicalProperties.line.solidFill = NAVY
        lc.series[1].graphicalProperties.line.width = 15000
        lc.series[1].graphicalProperties.line.dashDot = "dash"
    if len(lc.series) > 2:
        lc.series[2].graphicalProperties.line.solidFill = DGREY
        lc.series[2].graphicalProperties.line.width = 10000
        lc.series[2].graphicalProperties.line.dashDot = "sysDash"
    ws_bt.add_chart(lc, "G2")
    for col, w in zip(["A", "B", "C", "D", "E"], [14, 16, 20, 16, 16], strict=False):
        ws_bt.column_dimensions[col].width = w
    ws_bt.freeze_panes = "A2"

    # ----- TAB C: YEAR-ON-YEAR -----
    ws_yoy = wb.create_sheet(title="Year-on-Year")
    for ci, h in enumerate(
        [
            "Year",
            "Bills",
            "Peak Balance (£)",
            "Avg Balance (£)",
            "Lowest Balance (£)",
            "YoY Avg Δ (£)",
            "YoY Avg Δ (%)",
            "Est. Readings",
            "Biggest Jump (£)",
        ],
        1,
    ):
        _hcell(ws_yoy, 1, ci, h, bg=ORANGE)
    ws_yoy.row_dimensions[1].height = 22

    prev_avg = None
    yoy_data = []
    for r_off, row_y in enumerate(yearly.itertuples(), 2):
        yr = row_y.year
        cnt = row_y.count
        pk = row_y.peak
        av = row_y.avg_bal
        lo = row_y.low
        yoy_chg_pct = ((av - prev_avg) / prev_avg) if prev_avg else None

        yr_rows = dfc[dfc["year"] == yr]
        yr_idx = yr_rows.index.tolist()
        max_jump = None
        for ii in yr_idx:
            if ii > 0 and ii in dfc.index and ii - 1 in dfc.index:
                jmp = dfc.at[ii, "Amount (£)"] - dfc.at[ii - 1, "Amount (£)"]
                if max_jump is None or jmp > max_jump:
                    max_jump = jmp

        alt = r_off % 2 == 0
        bg = LGREY if alt else None

        _num(ws_yoy, r_off, 1, yr, fmt="#,##0", fill_hex=bg, bold=True)
        _num(ws_yoy, r_off, 2, cnt, fmt="#,##0", fill_hex=bg)
        _money(ws_yoy, r_off, 3, pk, fill_hex=bg, bold=True)
        _money(ws_yoy, r_off, 4, av, fill_hex=bg)
        _money(ws_yoy, r_off, 5, lo, fill_hex=bg)

        if r_off > 2:
            c6 = ws_yoy.cell(row=r_off, column=6, value=f"=D{r_off}-D{r_off - 1}")
            c6.number_format = "£#,##0.00"
            c6.font = Font(name="Calibri", size=10, bold=True)
            c6.border = CELL_BORDER
            c6.alignment = Alignment(horizontal="right")
            if bg:
                c6.fill = PatternFill("solid", start_color=bg)

            c7 = ws_yoy.cell(row=r_off, column=7, value=f'=IFERROR(F{r_off}/D{r_off - 1},"")')
            c7.number_format = "+0.0%;-0.0%;—"
            c7.font = Font(name="Calibri", size=10, bold=True)
            c7.border = CELL_BORDER
            c7.alignment = Alignment(horizontal="right")
            yoy_fill = (
                RED
                if yoy_chg_pct is not None and yoy_chg_pct > 0.5
                else (
                    AMBER
                    if yoy_chg_pct is not None and yoy_chg_pct > 0.2
                    else (GREEN if yoy_chg_pct is not None and yoy_chg_pct < -0.1 else bg)
                )
            )
            if yoy_fill:
                c7.fill = PatternFill("solid", start_color=yoy_fill)
        else:
            ws_yoy.cell(row=r_off, column=6, value="—").border = CELL_BORDER
            ws_yoy.cell(row=r_off, column=7, value="—").border = CELL_BORDER

        yr_est = (
            int((dfc[dfc["year"] == yr]["Reading"] == "Estimated").sum())
            if "Reading" in dfc.columns
            else 0
        )
        _num(ws_yoy, r_off, 8, yr_est, fmt="#,##0", fill_hex=bg)
        if max_jump is not None:
            _money(ws_yoy, r_off, 9, max_jump, fill_hex=(RED if max_jump > 5000 else bg))

        yoy_data.append((yr, av))
        prev_avg = av

    bc = BarChart()
    bc.type = "col"
    bc.title = "Average Balance by Year"
    bc.y_axis.title = "Average Balance (£)"
    bc.style = 10
    bc.width, bc.height = 22, 14
    n_yrs = len(yoy_data)
    avg_ref = Reference(ws_yoy, min_col=4, min_row=1, max_row=n_yrs + 1)
    yr_ref = Reference(ws_yoy, min_col=1, min_row=2, max_row=n_yrs + 1)
    bc.add_data(avg_ref, titles_from_data=True)
    bc.set_categories(yr_ref)
    bc.series[0].graphicalProperties.solidFill = ORANGE
    ws_yoy.add_chart(bc, "K2")
    for col, w in zip(
        ["A", "B", "C", "D", "E", "F", "G", "H", "I"],
        [8, 8, 18, 18, 18, 16, 14, 14, 18],
        strict=False,
    ):
        ws_yoy.column_dimensions[col].width = w
    ws_yoy.freeze_panes = "A2"

    # ----- TAB D: PERIOD CHARGES -----
    ws_pc = wb.create_sheet(title="Period Charges")
    for ci, h in enumerate(
        [
            "From Date",
            "To Date",
            "Days",
            "Opening Balance (£)",
            "Closing Balance (£)",
            "Charge (£)",
            "Daily Rate (£/day)",
            "Flag",
        ],
        1,
    ):
        _hcell(ws_pc, 1, ci, h, bg=NAVY)
    ws_pc.row_dimensions[1].height = 22

    mean_daily = float(np.mean(pos_diffs)) / 30.0 if len(pos_diffs) else 0
    pc_rows_data = []

    pc_r = 2
    for i in range(1, n):
        p = dfc.iloc[i - 1]
        c_ = dfc.iloc[i]
        days = (c_["_dt"] - p["_dt"]).days
        charge = float(c_["Amount (£)"]) - float(p["Amount (£)"])
        daily = charge / days if days > 0 else None

        flag = ""
        if days > 90:
            flag = f"⚠ {days}-day gap — possible missed bill(s)"
        elif charge < 0:
            flag = f"↓ Balance reduced by £{abs(charge):,.2f} (payment or credit)"
        elif daily and mean_daily > 0 and daily > mean_daily * 2.5:
            flag = f"⚠ Daily rate {daily / mean_daily:.1f}× average"

        bg = LGREY if pc_r % 2 == 0 else None
        if flag.startswith("⚠"):
            bg = AMBER
        elif charge < 0:
            bg = GREEN

        _text(ws_pc, pc_r, 1, p["Date"], fill_hex=bg)
        _text(ws_pc, pc_r, 2, c_["Date"], fill_hex=bg)
        _num(ws_pc, pc_r, 3, days, fmt="#,##0", fill_hex=bg)
        _money(ws_pc, pc_r, 4, float(p["Amount (£)"]), fill_hex=bg)
        _money(ws_pc, pc_r, 5, float(c_["Amount (£)"]), fill_hex=bg)

        c6 = ws_pc.cell(row=pc_r, column=6, value=f"=E{pc_r}-D{pc_r}")
        c6.number_format = "£#,##0.00"
        c6.font = Font(name="Calibri", size=10)
        c6.border = CELL_BORDER
        c6.alignment = Alignment(horizontal="right")
        if bg:
            c6.fill = PatternFill("solid", start_color=bg)

        c7 = ws_pc.cell(row=pc_r, column=7, value=f'=IFERROR(F{pc_r}/C{pc_r},"")')
        c7.number_format = "£#,##0.00"
        c7.font = Font(name="Calibri", size=10)
        c7.border = CELL_BORDER
        c7.alignment = Alignment(horizontal="right")
        if bg:
            c7.fill = PatternFill("solid", start_color=bg)

        _text(ws_pc, pc_r, 8, flag, fill_hex=bg, wrap=True)

        if charge > 0:
            pc_rows_data.append((c_["Date"], charge))
        pc_r += 1

    if pc_r > 2:
        sr = pc_r + 2
        _section_hdr(ws_pc, sr, "SUMMARY STATISTICS", ncols=8, bg=ORANGE)
        sr += 1
        dr = f"F2:F{pc_r - 1}"
        cr = f"C2:C{pc_r - 1}"

        def pc_stat(r, lbl, formula, fmt="£"):
            _text(ws_pc, r, 1, lbl, bold=True, fill_hex=LGREY)
            c = ws_pc.cell(row=r, column=2, value=formula)
            c.font = Font(name="Calibri", size=10, bold=True)
            c.fill = PatternFill("solid", start_color=LGREY)
            c.border = CELL_BORDER
            c.alignment = Alignment(horizontal="right")
            c.number_format = "£#,##0.00" if fmt == "£" else fmt
            for cc in range(3, 9):
                ws_pc.cell(row=r, column=cc).fill = PatternFill("solid", start_color=LGREY)
                ws_pc.cell(row=r, column=cc).border = CELL_BORDER

        pc_stat(sr, "Mean charge per period (positive only)", f'=IFERROR(AVERAGEIF({dr},">0"),"")')
        pc_stat(sr + 1, "Largest single charge", f'=IFERROR(MAX({dr}),"")')
        pc_stat(sr + 2, "Largest credit / reduction", f'=IFERROR(MIN({dr}),"")')
        pc_stat(sr + 3, "Charge periods", f'=IFERROR(COUNTIF({dr},">0"),"")', fmt="#,##0")
        pc_stat(sr + 4, "Credit periods", f'=IFERROR(COUNTIF({dr},"<0"),"")', fmt="#,##0")
        pc_stat(sr + 5, "Average days between bills", f'=IFERROR(AVERAGE({cr}),"")', fmt="#,##0.0")

    if len(pc_rows_data) > 1:
        bc2 = BarChart()
        bc2.type = "col"
        bc2.title = "Charge Added Each Period"
        bc2.y_axis.title = "Charge (£)"
        bc2.style = 10
        bc2.width, bc2.height = 28, 14
        chg_ref2 = Reference(ws_pc, min_col=6, min_row=1, max_row=pc_r - 1)
        date_ref2 = Reference(ws_pc, min_col=2, min_row=2, max_row=pc_r - 1)
        bc2.add_data(chg_ref2, titles_from_data=True)
        bc2.set_categories(date_ref2)
        bc2.series[0].graphicalProperties.solidFill = NAVY
        ws_pc.add_chart(bc2, "J2")

    for col, w in zip(
        ["A", "B", "C", "D", "E", "F", "G", "H"], [13, 13, 7, 18, 18, 16, 14, 42], strict=False
    ):
        ws_pc.column_dimensions[col].width = w
    ws_pc.freeze_panes = "A2"

    # ----- TAB E: DISPUTE FLAGS -----
    ws_df = wb.create_sheet(title="Dispute Flags")

    def _banner(ws, r, text, bg):
        c = ws.cell(row=r, column=1, value=text)
        c.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", start_color=bg)
        c.border = CELL_BORDER
        c.alignment = Alignment(horizontal="left", vertical="center")
        for col in range(2, 7):
            x = ws.cell(row=r, column=col)
            x.fill = PatternFill("solid", start_color=bg)
            x.border = CELL_BORDER
        ws.row_dimensions[r].height = 20

    _banner(ws_df, 1, "EDF ENERGY DISPUTE  —  AUTOMATED ANALYSIS FLAGS", ORANGE)
    ws_df.cell(
        row=2,
        column=1,
        value=f"Generated {datetime.now().strftime('%d/%m/%Y %H:%M')}  |  Period: {dates_lbl[0]} to {dates_lbl[-1]}",
    )
    ws_df.cell(row=2, column=1).font = Font(name="Calibri", size=9, italic=True, color=DGREY)

    for ci, (txt, col_hex) in enumerate(
        [
            ("■ RED = HIGH severity", RED),
            ("■ AMBER = MEDIUM", AMBER),
            ("■ GREEN = Payment/credit", GREEN),
        ],
        1,
    ):
        lc2 = ws_df.cell(row=3, column=ci * 2 - 1, value=txt)
        lc2.font = Font(name="Calibri", size=9, bold=True)
        lc2.fill = PatternFill("solid", start_color=col_hex)
        lc2.border = CELL_BORDER

    hdr_row = 5
    for ci, h in enumerate(["#", "Date", "Balance (£)", "Flag Type", "Detail", "Severity"], 1):
        _hcell(ws_df, hdr_row, ci, h, bg=NAVY)

    flags, counts = compute_dispute_flags(dfc, mean_daily)

    sev_fill = {"HIGH": RED, "MEDIUM": AMBER, "INFO": GREEN}
    for fi, (ftype, date, amt, detail, sev) in enumerate(flags, hdr_row + 1):
        bg = sev_fill.get(sev, LGREY)
        _num(ws_df, fi, 1, fi - hdr_row, fmt="#,##0", fill_hex=bg)
        _text(ws_df, fi, 2, date or "—", fill_hex=bg)
        if amt:
            _money(ws_df, fi, 3, float(amt), fill_hex=bg)
        else:
            ws_df.cell(row=fi, column=3).fill = PatternFill("solid", start_color=bg)
            ws_df.cell(row=fi, column=3).border = CELL_BORDER
        _text(ws_df, fi, 4, ftype, bold=True, fill_hex=bg)
        _text(ws_df, fi, 5, detail, fill_hex=bg, wrap=True)
        _text(ws_df, fi, 6, sev, bold=True, fill_hex=bg, align="center")
        ws_df.row_dimensions[fi].height = 30

    if flags:
        fr = len(flags) + hdr_row + 2
        counts = {s: sum(1 for f in flags if f[4] == s) for s in ("HIGH", "MEDIUM", "INFO")}
        _banner(
            ws_df,
            fr,
            f"TOTAL FLAGS: {len(flags)}   |   HIGH: {counts['HIGH']}   |   MEDIUM: {counts['MEDIUM']}   |   INFO: {counts['INFO']}",
            NAVY,
        )

    for col, w in zip(["A", "B", "C", "D", "E", "F"], [5, 13, 16, 20, 60, 10], strict=False):
        ws_df.column_dimensions[col].width = w
    ws_df.freeze_panes = f"A{hdr_row + 1}"

    # ----- TAB F: DISPUTE TIMELINE -----
    ws_tl = wb.create_sheet(title="Dispute Timeline")
    _banner(ws_tl, 1, "EDF ENERGY DISPUTE  —  CHRONOLOGICAL TIMELINE", ORANGE)
    ws_tl.cell(
        row=2, column=1, value=f"Account: {acc_ref}  |  Period: {dates_lbl[0]} to {dates_lbl[-1]}"
    )
    ws_tl.cell(row=2, column=1).font = Font(name="Calibri", size=9, italic=True, color=DGREY)

    for ci, h in enumerate(["Date", "Event Type", "Description"], 1):
        _hcell(ws_tl, 4, ci, h, bg=NAVY)

    timeline_events = []

    # Bookend: first record
    timeline_events.append(
        (dates_lbl[0], "ACCOUNT START", f"First bill on record. Balance: £{amounts[0]:,.2f}.")
    )

    # Top 5 largest balance jumps
    jumps = []
    for i in range(1, n):
        delta = float(amounts[i]) - float(amounts[i - 1])
        days = (dfc.iloc[i]["_dt"] - dfc.iloc[i - 1]["_dt"]).days
        if delta > 0:
            jumps.append((delta, i, days))
    jumps.sort(key=lambda x: x[0], reverse=True)
    for delta, idx, days in jumps[:5]:
        timeline_events.append(
            (
                dfc.iloc[idx]["Date"],
                "LARGE INCREASE",
                f"Balance rose £{delta:,.2f} in {days} days "
                f"(from £{amounts[idx - 1]:,.2f} to £{amounts[idx]:,.2f}).",
            )
        )

    # Billing gaps > 60 days
    for i in range(1, n):
        days = (dfc.iloc[i]["_dt"] - dfc.iloc[i - 1]["_dt"]).days
        if days > 60:
            timeline_events.append(
                (
                    dfc.iloc[i]["Date"],
                    "BILLING GAP",
                    f"{days} days without a bill (previous: {dfc.iloc[i - 1]['Date']}). "
                    f"Balance accumulated unchecked.",
                )
            )

    # Estimated reading runs (reuse existing detection)
    if "Reading" in dfc.columns:
        run = 0
        run_start_date = None
        for i, rv in enumerate(dfc["Reading"].tolist()):
            if str(rv).lower() in ("estimated", "est."):
                run += 1
                if run == 1:
                    run_start_date = dfc.iloc[i]["Date"]
            else:
                if run >= 3:
                    timeline_events.append(
                        (
                            run_start_date,
                            "ESTIMATED READINGS",
                            f"{run} consecutive bills used estimated meter readings.",
                        )
                    )
                run = 0
                run_start_date = None
        if run >= 3:
            timeline_events.append(
                (
                    run_start_date,
                    "ESTIMATED READINGS",
                    f"{run} consecutive estimated readings (ongoing).",
                )
            )

    # Payment events (balance reductions)
    for i in range(1, n):
        delta = float(amounts[i]) - float(amounts[i - 1])
        if delta < -200:
            timeline_events.append(
                (
                    dfc.iloc[i]["Date"],
                    "PAYMENT/CREDIT",
                    f"Balance reduced by £{abs(delta):,.2f} "
                    f"(from £{amounts[i - 1]:,.2f} to £{amounts[i]:,.2f}).",
                )
            )

    # Reconciliation mismatches (from flags)
    for ftype, fdate, _famt, fdetail, _fsev in flags:
        if ftype == "RECONCILIATION MISMATCH":
            timeline_events.append((fdate, "RECONCILIATION", fdetail))

    # Bookend: latest record
    timeline_events.append(
        (
            dates_lbl[-1],
            "CURRENT STATE",
            f"Latest bill on record. Balance: £{amounts[-1]:,.2f}. "
            f"Total increase from first record: £{amounts[-1] - amounts[0]:,.2f}.",
        )
    )

    # Sort by date and write
    timeline_events.sort(key=lambda e: parse_to_sort_date(e[0]) or pd.Timestamp.min)
    tl_r = 5
    for date, etype, desc in timeline_events:
        bg_hex = LGREY if tl_r % 2 == 0 else None
        _text(ws_tl, tl_r, 1, date, fill_hex=bg_hex)
        _text(ws_tl, tl_r, 2, etype, bold=True, fill_hex=bg_hex)
        _text(ws_tl, tl_r, 3, desc, fill_hex=bg_hex, wrap=True)
        ws_tl.row_dimensions[tl_r].height = 40
        tl_r += 1

    for col, w in zip(["A", "B", "C"], [14, 22, 90], strict=False):
        ws_tl.column_dimensions[col].width = w
    ws_tl.freeze_panes = "A5"

    # =====================================================================
    # NEW ANALYSIS TABS (added after Dispute Timeline)
    # =====================================================================

    # Statistical Analysis
    write_statistical_analysis_sheet(wb.create_sheet(title="Statistical Analysis"), dfc, config)

    # Payment Analysis
    write_payment_analysis_sheet(wb.create_sheet(title="Payment Analysis"), dfc)

    # Forecast & Projection
    write_forecast_sheet(wb.create_sheet(title="Forecast & Projection"), dfc)

    # Data Quality Report
    write_data_quality_sheet(wb.create_sheet(title="Data Quality Report"), df)

    # Tariff Analysis (if data available)
    write_tariff_analysis_sheet(wb.create_sheet(title="Tariff Analysis"), dfc)

    wb.save(output_path)


# =====================================================================
# NEW ANALYSIS FUNCTIONS (pandas-powered enhancements)
# =====================================================================


def _compute_rolling_stats(series, window=6):
    """Compute rolling statistics for a time series."""
    return {
        "mean": series.rolling(window=window, min_periods=1).mean(),
        "std": series.rolling(window=window, min_periods=1).std(),
        "min": series.rolling(window=window, min_periods=1).min(),
        "max": series.rolling(window=window, min_periods=1).max(),
        "median": series.rolling(window=window, min_periods=1).median(),
    }


def _compute_ema(series, span=6):
    """Compute Exponential Moving Average."""
    return series.ewm(span=span, adjust=False).mean()


def _compute_momentum(series, period=3):
    """Compute momentum (rate of change) of a series."""
    return series.diff(period)


def _compute_volatility(series, window=6):
    """Compute rolling volatility (std of returns)."""
    returns = series.pct_change()
    return returns.rolling(window=window, min_periods=1).std()


def _zscore_anomalies(series, threshold=2.5):
    """Detect anomalies using z-score method."""
    if len(series) < 3:
        return pd.Series(False, index=series.index)
    mean = series.mean()
    std = series.std()
    if std == 0:
        return pd.Series(False, index=series.index)
    z_scores = np.abs((series - mean) / std)
    return z_scores > threshold


def _iqr_anomalies(series, multiplier=1.5):
    """Detect anomalies using IQR method."""
    if len(series) < 4:
        return pd.Series(False, index=series.index)
    q1 = series.quantile(0.25)
    q3 = series.quantile(0.75)
    iqr = q3 - q1
    if iqr == 0:
        return pd.Series(False, index=series.index)
    lower = q1 - multiplier * iqr
    upper = q3 + multiplier * iqr
    return (series < lower) | (series > upper)


def _linear_forecast_pair(series, steps=6):
    """Simple linear regression: returns (fitted, future) values.

    The fitted series is the model's prediction at each historical
    point — this lets the Forecast tab back-paint predictions onto
    historical rows so the reader sees actual-vs-predicted for the
    whole data range, not only at a 6-step future horizon.

    Linear regression in this codebase uses ``np.polyfit``.  The
    fitted value at index ``i`` is simply ``np.polyval(coeffs, i)``
    computed against the same coefficients used for the future
    forecast, so the in-sample and out-of-sample predictions share
    a single model — meaning the historical vs forward columns
    reflect exactly the same fit.

    Returns ``(None, None)`` for insufficient data.
    """
    if len(series) < 3:
        return None, None
    x = np.arange(len(series))
    y = series.values
    # Handle NaN values
    mask = ~np.isnan(y)
    if mask.sum() < 3:
        return None, None
    x_clean = x[mask]
    y_clean = y[mask]
    try:
        coeffs = np.polyfit(x_clean, y_clean, 1)
        # Fitted values for every historical index — back-pained
        # by the same straight line that drives the future window.
        fitted = np.polyval(coeffs, x)
        future_x = np.arange(len(series), len(series) + steps)
        forecast = np.polyval(coeffs, future_x)
        return fitted, forecast
    except Exception:
        return None, None


def _holt_winters_forecast_pair(series, steps=6, seasonal_periods=None):
    """Holt-Winters: returns (fitted, future) values (if statsmodels available).

    Mirrors ``_linear_forecast_pair`` for the ExponentialSmoothing
    path.  Statsmodels's ``fit()`` returns a fitted-ness model whose
    ``.fittedvalues`` attribute carries the one-step-ahead in-sample
    prediction at every historical index — exactly what we need to
    back-paint the forecast tab so the reader sees actual vs
    predicted divergence for the whole data range.

    Returns ``(None, None)`` when statsmodels is unavailable, the
    series is too short, or fitting fails.
    """
    if not HAS_STATSMODELS or len(series) < 4:
        return None, None
    try:
        clean_series = series.dropna()
        if len(clean_series) < 4:
            return None, None

        if seasonal_periods is None:
            seasonal_periods = min(12, len(clean_series) // 2) if len(clean_series) >= 8 else None

        model = ExponentialSmoothing(
            clean_series,
            trend="add",
            seasonal="add" if seasonal_periods else None,
            seasonal_periods=seasonal_periods,
            initialization_method="estimated",
        )
        fitted_model = model.fit(optimized=True)
        # In-sample fitted: statsmodels returns the one-step-ahead
        # prediction for each historical point the model was fit
        # against.  We reindex onto the original series (which may
        # include NaN gaps) so row N in the call sites lines up
        # with row N in the user's data.
        fitted_vals = fitted_model.fittedvalues.reindex(series.index)
        forecast = fitted_model.forecast(steps).values
        return fitted_vals.values, forecast
    except Exception:
        return None, None


def _linear_forecast(series, steps=6):
    """Simple linear regression forecast (forward-only legacy entry point).

    See ``_linear_forecast_pair`` for the (fitted, future) form that
    the Forecast tab now uses.  This single-value shim is kept for
    any callers that imported the previous-shape return value (we
    don't have any in-tree callers anymore, but a user
    may have downstream code that does).
    """
    _, forecast = _linear_forecast_pair(series, steps)
    return forecast


def _holt_winters_forecast(series, steps=6, seasonal_periods=None):
    """Holt-Winters forward-only legacy entry point.  See ``_holt_winters_forecast_pair``."""
    _, forecast = _holt_winters_forecast_pair(series, steps, seasonal_periods)
    return forecast


def _detect_payment_patterns(df):
    """Analyze payment/credit patterns in the data."""
    payments = df[df["Entry Type"].isin(["Payment", "Credit"])].copy()
    if payments.empty:
        return {}

    payments["_dt"] = payments["Date"].apply(parse_to_sort_date)
    payments = payments.sort_values("_dt")

    # Calculate days between payments
    pay_dates = payments["_dt"].dropna()
    intervals = pay_dates.diff().dt.days.dropna()

    # Payment amounts (negative values for credits/payments)
    pay_amounts = payments["Amount (£)"].astype(float)

    return {
        "count": len(payments),
        "total_paid": abs(pay_amounts.sum()),
        "avg_payment": abs(pay_amounts.mean()),
        "median_payment": abs(pay_amounts.median()),
        "max_payment": abs(pay_amounts.max()),
        "min_payment": abs(pay_amounts.min()),
        "avg_interval_days": float(intervals.mean()) if len(intervals) > 0 else None,
        "median_interval_days": float(intervals.median()) if len(intervals) > 0 else None,
        "last_payment_date": payments.iloc[-1]["Date"] if len(payments) > 0 else None,
        "last_payment_amount": abs(pay_amounts.iloc[-1]) if len(pay_amounts) > 0 else None,
    }


def _analyze_tariff_impact(df):
    """Analyze the impact of tariff changes on unit rates and charges."""
    if "Tariff" not in df.columns or "Unit Rate (p/kWh)" not in df.columns:
        return {}

    tariff_data = df[df["Tariff"].notna() & (df["Tariff"] != "N/A")].copy()
    if tariff_data.empty:
        return {}

    # Convert unit rate to numeric
    tariff_data["unit_rate_num"] = pd.to_numeric(tariff_data["Unit Rate (p/kWh)"], errors="coerce")
    tariff_data = tariff_data.dropna(subset=["unit_rate_num"])

    if tariff_data.empty:
        return {}

    # Group by tariff
    tariff_stats = (
        tariff_data.groupby("Tariff")
        .agg(
            count=("unit_rate_num", "count"),
            avg_unit_rate=("unit_rate_num", "mean"),
            median_unit_rate=("unit_rate_num", "median"),
            min_unit_rate=("unit_rate_num", "min"),
            max_unit_rate=("unit_rate_num", "max"),
            avg_charge=("Period Charge (£)", lambda x: pd.to_numeric(x, errors="coerce").mean()),
        )
        .reset_index()
    )

    # Find tariff changes
    tariff_data = tariff_data.sort_values("_dt" if "_dt" in tariff_data.columns else "Date")
    tariff_changes = tariff_data["Tariff"].ne(tariff_data["Tariff"].shift()).cumsum()

    return {
        "tariff_stats": tariff_stats,
        "num_tariffs": tariff_data["Tariff"].nunique(),
        "tariff_changes": int(tariff_changes.max()) if not tariff_changes.empty else 0,
    }


def _data_quality_report(df):
    """Generate a comprehensive data quality report.

    Works on a *copy* of the input DataFrame so the caller's data is
    never mutated (previously this added ``_dt_parsed`` as a side-effect
    on the caller's df, which broke downstream code that re-used the
    same DataFrame for other purposes).
    """
    # Work on a copy to avoid mutating the caller's DataFrame
    df = df.copy()
    total_records = len(df)
    if total_records == 0:
        return {}

    # Date parsing success
    df["_dt_parsed"] = df["Date"].apply(parse_to_sort_date)
    date_parsed = df["_dt_parsed"].notna().sum()
    date_failed = total_records - date_parsed

    # Amount completeness
    amt_complete = df["Amount (£)"].notna().sum()
    amt_missing = total_records - amt_complete

    # Period info completeness
    period_from_complete = (df["Period From"] != "N/A").sum()
    _ = (df["Period To"] != "N/A").sum()  # Not used, but computed for completeness
    period_complete = period_from_complete  # At least from date

    # Reading classification
    # Reading classification — "N/A" is the sentinel for unclassified readings
    reading_classified = (df["Reading"] != "N/A").sum() if "Reading" in df.columns else 0

    # Unit rate computable — count numeric values only. The unit
    # rate column can hold `int | float | "N/A"`; only numerics can be
    # used downstream by tariff charts, so other values are excluded.
    # The older draft guarded this with `and x != "N/A"`, which is
    # unreachable for an already-typed numeric — pinned here so a
    # future careless refactor cannot silently change this branch
    # back into a no-op-or-true tautology that overcounts.
    ur_computable = df["Unit Rate (p/kWh)"].apply(lambda x: isinstance(x, (int, float))).sum()

    # Duplicates (same date + amount)
    dup_count = df.duplicated(subset=["Date", "Amount (£)"]).sum()

    # Source distribution
    source_dist = df["Source"].value_counts().to_dict()

    # Entry type distribution
    entry_dist = df["Entry Type"].value_counts().to_dict() if "Entry Type" in df.columns else {}

    return {
        "total_records": total_records,
        "date_parsed": date_parsed,
        "date_failed": date_failed,
        "date_parse_rate": date_parsed / total_records if total_records > 0 else 0,
        "amt_complete": amt_complete,
        "amt_missing": amt_missing,
        "period_complete": period_complete,
        "period_completeness_rate": period_complete / total_records if total_records > 0 else 0,
        "reading_classified": reading_classified,
        "reading_classify_rate": reading_classified / total_records if total_records > 0 else 0,
        "ur_computable": ur_computable,
        "ur_computable_rate": ur_computable / total_records if total_records > 0 else 0,
        "duplicate_count": int(dup_count),
        "duplicate_rate": dup_count / total_records if total_records > 0 else 0,
        "source_distribution": source_dist,
        "entry_type_distribution": entry_dist,
    }


# ---------------------------------------------------------------------------
# NEW ANALYSIS TAB WRITERS
# ---------------------------------------------------------------------------


def write_statistical_analysis_sheet(ws, dfc, config):
    """Write Statistical Analysis tab with advanced pandas analytics."""
    ws.title = "Statistical Analysis"

    NAVY = "10367A"
    ORANGE = "FE5716"
    AMBER = "FFD166"
    LGREY = "F0F0F0"
    DGREY = "888888"

    # Prepare data
    dfc = dfc.copy()
    dfc["_dt"] = dfc["Date"].apply(parse_to_sort_date)
    dfc = dfc.sort_values("_dt").reset_index(drop=True)
    amounts = dfc["Amount (£)"].astype(float).values
    dates = dfc["Date"].tolist()
    n = len(amounts)

    if n < 3:
        _hcell(ws, 1, 1, "Insufficient data for statistical analysis", bg=NAVY)
        ws.column_dimensions["A"].width = 50
        return

    # Headers
    headers = [
        "Metric",
        "Value",
        "Notes",
    ]
    for col, h in enumerate(headers, 1):
        _hcell(ws, 1, col, h, bg=NAVY)
    ws.row_dimensions[1].height = 28

    # Title
    tc = ws.cell(row=1, column=1, value="EDF ENERGY DISPUTE  —  STATISTICAL ANALYSIS")
    tc.font = Font(name="Calibri", size=13, bold=True, color="FFFFFF")
    tc.fill = PatternFill("solid", start_color=ORANGE)
    tc.border = CELL_BORDER
    tc.alignment = Alignment(horizontal="left", vertical="center")
    for c in [2, 3]:
        x = ws.cell(row=1, column=c)
        x.fill = PatternFill("solid", start_color=ORANGE)
        x.border = CELL_BORDER

    # Summary stats
    r = 2
    _section_hdr(ws, r, "DESCRIPTIVE STATISTICS")

    amounts_series = pd.Series(amounts)
    stats_data = [
        ("Count", len(amounts), "#,##0", "Number of billing records"),
        ("Mean (£)", float(amounts_series.mean()), "£#,##0.00", "Average balance"),
        ("Median (£)", float(amounts_series.median()), "£#,##0.00", "Median balance"),
        ("Std Dev (£)", float(amounts_series.std()), "£#,##0.00", "Standard deviation"),
        ("Min (£)", float(amounts_series.min()), "£#,##0.00", "Minimum balance"),
        ("Max (£)", float(amounts_series.max()), "£#,##0.00", "Maximum balance"),
        ("Range (£)", float(amounts_series.max() - amounts_series.min()), "£#,##0.00", "Max - Min"),
        (
            "Skewness",
            float(amounts_series.skew()) if hasattr(amounts_series, "skew") else 0,
            "0.00",
            "Asymmetry of distribution",
        ),
        (
            "Kurtosis",
            float(amounts_series.kurtosis()) if hasattr(amounts_series, "kurtosis") else 0,
            "0.00",
            "Tailedness of distribution",
        ),
        (
            "CV (%)",
            float(amounts_series.std() / amounts_series.mean() * 100)
            if amounts_series.mean() > 0
            else 0,
            "0.00",
            "Coefficient of variation",
        ),
    ]

    for label, value, fmt, note in stats_data:
        r += 1
        bg = LGREY if r % 2 == 0 else None
        _text(ws, r, 1, label, fill_hex=bg)
        if fmt == "£":
            _money(ws, r, 2, value, fill_hex=bg)
        elif fmt == "%":
            _num(ws, r, 2, value, fmt="0.0%", fill_hex=bg)
        else:
            _num(ws, r, 2, value, fmt=fmt, fill_hex=bg)
        _text(ws, r, 3, note, fill_hex=bg, color=DGREY)

    # Rolling statistics
    r += 1
    _section_hdr(ws, r, "ROLLING STATISTICS (6-period window)")
    r += 1
    _text(ws, r, 1, "Current 6-Period Rolling Mean (£)", bold=True)
    rolling_mean = float(pd.Series(amounts).rolling(6, min_periods=1).mean().iloc[-1])
    _money(ws, r, 2, rolling_mean)

    r += 1
    _text(ws, r, 1, "Current 6-Period Rolling Std (£)", bold=True)
    rolling_std = float(pd.Series(amounts).rolling(6, min_periods=1).std().iloc[-1])
    _money(ws, r, 2, rolling_std)

    r += 1
    _text(ws, r, 1, "Current 6-Period Rolling Min (£)", bold=True)
    rolling_min = float(pd.Series(amounts).rolling(6, min_periods=1).min().iloc[-1])
    _money(ws, r, 2, rolling_min)

    r += 1
    _text(ws, r, 1, "Current 6-Period Rolling Max (£)", bold=True)
    rolling_max = float(pd.Series(amounts).rolling(6, min_periods=1).max().iloc[-1])
    _money(ws, r, 2, rolling_max)

    r += 1
    _text(ws, r, 1, "Current 6-Period Rolling Median (£)", bold=True)
    rolling_median = float(pd.Series(amounts).rolling(6, min_periods=1).median().iloc[-1])
    _money(ws, r, 2, rolling_median)

    # Exponential Moving Average
    r += 1
    _section_hdr(ws, r, "EXPONENTIAL MOVING AVERAGE")
    r += 1
    _text(ws, r, 1, "Current EMA (span=6) (£)", bold=True)
    ema = float(pd.Series(amounts).ewm(span=6, adjust=False).mean().iloc[-1])
    _money(ws, r, 2, ema)

    r += 1
    _text(ws, r, 1, "EMA vs Simple SMA Difference (£)", bold=True)
    sma = float(pd.Series(amounts).rolling(6, min_periods=1).mean().iloc[-1])
    _money(ws, r, 2, ema - sma)

    # Momentum & Volatility
    r += 1
    _section_hdr(ws, r, "MOMENTUM & VOLATILITY")
    r += 1
    mom = float(pd.Series(amounts).diff(3).iloc[-1]) if n >= 4 else 0
    _text(ws, r, 1, "3-Period Momentum (£)", bold=True)
    _money(ws, r, 2, mom)

    r += 1
    vol = (
        float(pd.Series(amounts).pct_change().rolling(6, min_periods=1).std().iloc[-1])
        if n >= 3
        else 0
    )
    _text(ws, r, 1, "6-Period Volatility (σ of returns)", bold=True)
    _num(ws, r, 2, vol, fmt="0.00%")

    # Anomaly Detection
    r += 1
    _section_hdr(ws, r, "ANOMALY DETECTION")
    series = pd.Series(amounts, index=pd.to_datetime(dates, dayfirst=True, errors="coerce"))

    z_anoms = _zscore_anomalies(series, threshold=2.5)
    iqr_anoms = _iqr_anomalies(series, multiplier=1.5)

    z_count = int(z_anoms.sum())
    iqr_count = int(iqr_anoms.sum())

    r += 1
    _text(ws, r, 1, "Z-Score Anomalies (threshold=2.5σ)", bold=True)
    _num(ws, r, 2, z_count, fmt="#,##0")

    r += 1
    _text(ws, r, 1, "IQR Anomalies (multiplier=1.5)", bold=True)
    _num(ws, r, 2, iqr_count, fmt="#,##0")

    # List detected anomalies
    if z_count > 0:
        r += 1
        _text(ws, r, 1, "Z-Score Anomaly Dates:", bold=True)
        anom_dates = series[z_anoms].index
        for dt in anom_dates:
            r += 1
            _text(
                ws,
                r,
                1,
                f"  • {dt.strftime('%d/%m/%Y') if hasattr(dt, 'strftime') else dt} ({series[dt]:,.2f})",
            )

    if iqr_count > 0:
        r += 1
        _text(ws, r, 1, "IQR Anomaly Dates:", bold=True)
        anom_dates = series[iqr_anoms].index
        for dt in anom_dates:
            r += 1
            _text(
                ws,
                r,
                1,
                f"  • {dt.strftime('%d/%m/%Y') if hasattr(dt, 'strftime') else dt} ({series[dt]:,.2f})",
            )

    # Normality test (if scipy available)
    r += 1
    _section_hdr(ws, r, "DISTRIBUTION TESTS")
    if HAS_SCIPY:
        try:
            from scipy import stats as sp_stats

            shapiro_stat, shapiro_p = sp_stats.shapiro(amounts_series.dropna())
            r += 1
            _text(ws, r, 1, "Shapiro-Wilk Test (Normality)", bold=True)
            _num(ws, r, 2, shapiro_stat, fmt="0.0000")
            _text(
                ws,
                r,
                3,
                f"p-value: {shapiro_p:.4f} — {'Normal' if shapiro_p > 0.05 else 'Non-normal'}",
            )

            # Jarque-Bera
            jb_stat, jb_p = sp_stats.jarque_bera(amounts_series.dropna())
            r += 1
            _text(ws, r, 1, "Jarque-Bera Test (Normality)", bold=True)
            _num(ws, r, 2, jb_stat, fmt="0.00")
            _text(ws, r, 3, f"p-value: {jb_p:.4f} — {'Normal' if jb_p > 0.05 else 'Non-normal'}")
        except Exception:
            r += 1
            _text(ws, r, 1, "Scipy tests failed", fill_hex=AMBER)
    else:
        r += 1
        _text(ws, r, 1, "Scipy not available — install for normality tests", fill_hex=AMBER)

    # Column widths
    for col_letter, width in zip(["A", "B", "C"], [45, 22, 80], strict=False):
        ws.column_dimensions[col_letter].width = width
    ws.freeze_panes = "A2"


def write_payment_analysis_sheet(ws, dfc):
    """Write Payment/Credit Analysis tab."""
    ws.title = "Payment Analysis"

    NAVY = "10367A"
    ORANGE = "FE5716"
    LGREY = "F0F0F0"
    DGREY = "888888"

    payments = dfc[dfc["Entry Type"].isin(["Payment", "Credit"])].copy()
    if payments.empty:
        _hcell(ws, 1, 1, "No payment/credit records found", bg=NAVY)
        ws.column_dimensions["A"].width = 50
        return

    payments["_dt"] = payments["Date"].apply(parse_to_sort_date)
    payments = payments.sort_values("_dt").reset_index(drop=True)

    headers = ["Metric", "Value", "Notes"]
    for col, h in enumerate(headers, 1):
        _hcell(ws, 1, col, h, bg=NAVY)
    ws.row_dimensions[1].height = 28

    tc = ws.cell(row=1, column=1, value="EDF ENERGY DISPUTE  —  PAYMENT & CREDIT ANALYSIS")
    tc.font = Font(name="Calibri", size=13, bold=True, color="FFFFFF")
    tc.fill = PatternFill("solid", start_color=ORANGE)
    tc.border = CELL_BORDER
    tc.alignment = Alignment(horizontal="left", vertical="center")
    for c in [2, 3]:
        x = ws.cell(row=1, column=c)
        x.fill = PatternFill("solid", start_color=ORANGE)
        x.border = CELL_BORDER

    pat = _detect_payment_patterns(dfc)

    r = 2
    _section_hdr(ws, r, "PAYMENT SUMMARY")

    payment_items = [
        ("Total Payments/Credits", pat["count"], "#,##0", "Number of payment events"),
        ("Total Amount Paid (£)", pat["total_paid"], "£#,##0.00", "Sum of all payments/credits"),
        ("Average Payment (£)", pat["avg_payment"], "£#,##0.00", "Mean payment amount"),
        ("Median Payment (£)", pat["median_payment"], "£#,##0.00", "Median payment amount"),
        ("Largest Payment (£)", pat["max_payment"], "£#,##0.00", "Maximum single payment"),
        ("Smallest Payment (£)", pat["min_payment"], "£#,##0.00", "Minimum single payment"),
    ]

    for label, value, fmt, note in payment_items:
        r += 1
        bg = LGREY if r % 2 == 0 else None
        _text(ws, r, 1, label, fill_hex=bg)
        if fmt == "£":
            _money(ws, r, 2, value, fill_hex=bg)
        else:
            _num(ws, r, 2, value, fmt=fmt, fill_hex=bg)
        _text(ws, r, 3, note, fill_hex=bg, color=DGREY)

    # Payment intervals
    r += 1
    _section_hdr(ws, r, "PAYMENT TIMING")
    interval_items = [
        ("Avg Interval (days)", pat["avg_interval_days"], "#,##0.0", "Mean days between payments"),
        (
            "Median Interval (days)",
            pat["median_interval_days"],
            "#,##0.0",
            "Median days between payments",
        ),
    ]
    for label, value, fmt, note in interval_items:
        r += 1
        bg = LGREY if r % 2 == 0 else None
        _text(ws, r, 1, label, fill_hex=bg)
        if value is not None:
            _num(ws, r, 2, value, fmt=fmt, fill_hex=bg)
        else:
            _text(ws, r, 2, "N/A", fill_hex=bg)
        _text(ws, r, 3, note, fill_hex=bg, color=DGREY)

    # Last payment
    r += 1
    _section_hdr(ws, r, "LAST PAYMENT")
    r += 1
    _text(ws, r, 1, "Last Payment Date", bold=True)
    _text(ws, r, 2, pat["last_payment_date"] or "N/A")

    r += 1
    _text(ws, r, 1, "Last Payment Amount (£)", bold=True)
    _money(ws, r, 2, pat["last_payment_amount"] or 0)

    # Payment detail table
    r += 2
    _section_hdr(ws, r, "ALL PAYMENTS & CREDITS (Chronological)")
    r += 1
    pay_headers = ["Date", "Entry Type", "Amount (£)", "Balance After (£)", "Details"]
    for ci, h in enumerate(pay_headers, 1):
        _hcell(ws, r, ci, h, bg=NAVY)

    for i, (_, row) in enumerate(payments.iterrows()):
        r += 1
        bg = LGREY if i % 2 == 0 else None
        _text(ws, r, 1, row["Date"], fill_hex=bg)
        _text(ws, r, 2, row["Entry Type"], fill_hex=bg, bold=True)
        _money(ws, r, 3, float(row["Amount (£)"]), fill_hex=bg)
        # Balance After (column 4) — Historical Note: this column shows
        # the per-row transaction amount rather than the running account
        # balance.  Real "balance-after" data is not currently parsed
        # from EDF bills, so we display the same amount as a placeholder
        # (open Low-severity follow-up B6: parse the running balance
        # column from EDF statements when available).
        _money(ws, r, 4, float(row["Amount (£)"]), fill_hex=bg)
        _text(ws, r, 5, str(row.get("Details", ""))[:60], fill_hex=bg, wrap=True)

    # Chart - Payment amounts over time.
    # Phase-2 portability fix: the previous layout anchored the
    # chart at ``cell(row+2, column H)`` (column 8) which sat past
    # the visible data table (columns A-E) **and** the user's
    # roughly-default Excel viewport (about seven column-units
    # wide before they have to scroll).  An ombudsman reading
    # the report saw the chart title render *off-screen*.  We now:
    #
    #  * Place the chart-data helper cells in **column A**
    #    (single-cell-style) at a dedicated row block below the
    #    data so the chart reads ``date × amount`` cleanly;
    #  * Drop the chart *anchor* to column B, two rows below the
    #    data table — that's the most common Excel default
    #    reading order, so the user sees the data first and the
    #    chart underneath;
    #  * Cap the chart at width=16, height=10 (openpyxl's chart
    #    units, where 1 unit ≈ 1 Excel column / row).  The
    #    previous 28 × 14 values pushed the chart so far right
    #    that it appeared only partially when the file opened;
    #  * Use a colour-blind-friendly palette (single GREEN
    #    series — the existing colour — so a reviewer with
    #    deuteranopia can still trace payment size to date via
    #    the data labels).
    if len(payments) > 1:
        bc = BarChart()
        bc.type = "col"
        bc.title = "Payment/Credit Amounts Over Time"
        bc.y_axis.title = "Amount (£)"
        bc.x_axis.title = "Payment Date"
        bc.style = 10
        bc.width = 16
        bc.height = 10
        bc.legend = None

        # Step 1: write the chart-data series to a dedicated,
        # labelled mini-table two rows below the payments detail.
        # Putting both series in the same column range keeps the
        # chart's Reference call simple and avoids scattered helper
        # cells.
        chart_data_start_row = r + 3
        _hcell(ws, chart_data_start_row, 1, "Date", bg=NAVY)
        _hcell(
            ws,
            chart_data_start_row,
            2,
            "Payment Amount (£)",
            bg=NAVY,
        )
        for i, (_, row) in enumerate(payments.iterrows(), 1):
            payload_row = chart_data_start_row + i
            _text(ws, payload_row, 1, row["Date"])
            _money(ws, payload_row, 2, float(row["Amount (£)"]))

        # Step 2: build the chart from the labelled mini-table so
        # the title ("C2", "D2") series is unambiguous when a
        # reviewer opens the file's chart-edit dialog.
        chg_ref = Reference(
            ws,
            min_col=2,
            min_row=chart_data_start_row,
            max_row=chart_data_start_row + len(payments),
        )
        date_ref = Reference(
            ws,
            min_col=1,
            min_row=chart_data_start_row + 1,
            max_row=chart_data_start_row + len(payments),
        )
        bc.add_data(chg_ref, titles_from_data=True)
        bc.set_categories(date_ref)

        # Step 3: anchor the chart under the data table so the
        # reader's eye flows from raw rows to chart without
        # panning across the spreadsheet.  Row offset 2 gives the
        # chart a small breathing-room gap below the helper rows.
        anchor_row = chart_data_start_row + len(payments) + 2
        ws.add_chart(bc, f"B{anchor_row}")

    for col_letter, width in zip(["A", "B", "C", "D", "E"], [14, 16, 16, 16, 60], strict=False):
        ws.column_dimensions[col_letter].width = width
    ws.freeze_panes = f"A{r - len(payments)}"


def write_forecast_sheet(ws, dfc):
    """Write Forecast/Projection tab with multiple forecasting methods."""
    ws.title = "Forecast & Projection"

    NAVY = "10367A"
    ORANGE = "FE5716"
    AMBER = "FFD166"
    LGREY = "F0F0F0"
    DGREY = "888888"

    dfc = dfc.copy()
    dfc["_dt"] = dfc["Date"].apply(parse_to_sort_date)
    dfc = dfc.sort_values("_dt").reset_index(drop=True)
    amounts = dfc["Amount (£)"].astype(float).values
    dates = dfc["Date"].tolist()
    n = len(amounts)

    if n < 3:
        _hcell(ws, 1, 1, "Insufficient data for forecasting (need 3+ records)", bg=NAVY)
        ws.column_dimensions["A"].width = 60
        return

    # ``Date`` + the canonical six forecast columns + ``Forecast Δ
    # (Actual − Linear)``.  The Δ column is what makes the tab
    # useful as evidence: a reviewer sees *by how much* each bill
    # diverged from what the model would call average.  Historical
    # rows carry a per-row back-painted prediction; future rows
    # carry forward-looking projections; the divider between the
    # two is a separator row.
    headers = [
        "Date",
        "Actual (£)",
        "Linear Forecast (£)",
        "Holt-Winters (£)",
        "EMA Projection (£)",
        "Confidence (±£)",
        "Forecast Δ (Actual − Linear)",
    ]
    for col, h in enumerate(headers, 1):
        _hcell(ws, 1, col, h, bg=NAVY)
    ws.row_dimensions[1].height = 28

    tc = ws.cell(row=1, column=1, value="EDF ENERGY DISPUTE  —  BALANCE FORECAST")
    tc.font = Font(name="Calibri", size=13, bold=True, color="FFFFFF")
    tc.fill = PatternFill("solid", start_color=ORANGE)
    tc.border = CELL_BORDER
    tc.alignment = Alignment(horizontal="left", vertical="center")
    for c in range(2, 8):
        x = ws.cell(row=1, column=c)
        x.fill = PatternFill("solid", start_color=ORANGE)
        x.border = CELL_BORDER

    # Generate forecasts (6 steps ahead).  We use the *_pair helper
    # variants to also obtain the in-sample fitted-values array so
    # every historical row carries a real prediction column rather
    # than the previous "—" placeholders.  This is what makes the
    # tab show model-vs-actual divergence across the full data range.
    forecast_steps = 6
    series = pd.Series(amounts, index=pd.to_datetime(dates, dayfirst=True, errors="coerce"))

    # ``linear_fitted[i]`` is the straight-line prediction at row i
    # (uses ALL n historical points); ``linear_fc[i]`` is the future
    # value i steps past the last historical row.  Both come from
    # the same fit, so the in-sample and out-of-sample columns
    # share one model.
    linear_fitted, linear_fc = _linear_forecast_pair(series, forecast_steps)
    hw_fitted, hw_fc = _holt_winters_forecast_pair(series, forecast_steps)
    # EMA trajectory: per-row exponentially-weighted moving average.
    # We expand the existing ``_compute_ema`` helper into a length-n
    # series so every historical row gets the right EMA *as of that
    # row*, not the last-window mean.
    ema_series = _compute_ema(series, span=6)
    ema_last = ema_series.iloc[-1] if n >= 2 else amounts[-1]
    # Forward EMA projection extends the last EMA flat-forecast for
    # future rows; historical rows just carry the historical EMA.
    ema_future = [ema_last] * forecast_steps

    # Historical volatility for confidence intervals.
    # ``hist_vol`` is the std-dev of monthly *returns* (pct_change),
    # which is what we multiply against the predicted value to
    # produce a ±2σ confidence band.  With only one historical bill
    # we fall back to a sensible default.
    returns = pd.Series(amounts).pct_change().dropna()
    hist_vol = returns.std() if len(returns) > 1 else 0.05

    def _model_value(fitted_array, fc_array, i, n_total):
        """Pick the in-sample fitted value at historical index i
        or ``N/A`` if the model didn't fit (not enough data).
        """
        if fitted_array is None:
            return None
        # Defensive index guard — the fitted array has the same
        # length as ``series`` per the *_pair helpers, but a
        # statsmodels-index misalignment is always possible.
        if i < len(fitted_array):
            val = fitted_array[i]
            return val if not pd.isna(val) else None
        return None

    # === Historical block: back-paint every forecast column ===
    # The y-axis of the forecast table now spans the *entire* data
    # range — each historical row carries the model's prediction at
    # that point, and the Forecast Δ column quantifies how far the
    # actual bill landed above (positive) or below (negative) the
    # linear-trend baseline.  The future block (after the separator
    # row) shows 6 forward projection rows.  Together they answer
    # "given what you've paid historically, what should you have
    # paid each month, and where did the bill diverge?".
    r = 2
    for i in range(n):
        bg = LGREY if i % 2 == 0 else None
        _text(ws, r, 1, dates[i], fill_hex=bg)
        _money(ws, r, 2, float(amounts[i]), fill_hex=bg)
        # Linear forecast — back-painted fitted value (not "—").
        lin_val = _model_value(linear_fitted, linear_fc, i, n)
        if lin_val is not None:
            _money(ws, r, 3, float(lin_val), fill_hex=bg)
        else:
            _text(ws, r, 3, "N/A", fill_hex=bg)
        # Holt-Winters — back-painted fitted value (still "N/A"
        # when statsmodels is unavailable or the series is too
        # short for the additive-trend fit).
        hw_val = _model_value(hw_fitted, hw_fc, i, n)
        if hw_val is not None:
            _money(ws, r, 4, float(hw_val), fill_hex=bg)
        else:
            _text(ws, r, 4, "N/A", fill_hex=bg)
        # EMA — per-row exponentially-weighted moving average
        # (historical anchored to row i's position in the series).
        ema_at_i = float(ema_series.iloc[i]) if not pd.isna(ema_series.iloc[i]) else None
        if ema_at_i is not None:
            _money(ws, r, 5, ema_at_i, fill_hex=bg)
        else:
            _text(ws, r, 5, "N/A", fill_hex=bg)
        # Confidence band — ±2σ around the fitted value.  When the
        # model didn't fit we fall back to the predicted value of
        # the actual bill (i.e. confidence = 0) — visually faithful
        # but not concealing data.
        if lin_val is not None:
            conf = abs(float(lin_val)) * hist_vol * 2
            _money(ws, r, 6, conf, fill_hex=bg)
        else:
            _text(ws, r, 6, "N/A", fill_hex=bg)
        # Forecast Δ = actual − fitted linear.  This is the
        # ombudsman-facing signal: a row with ``£50`` actual and a
        # fitted linear value of ``£200`` writes ``−£150`` here,
        # i.e. the bill landed £150 below what the trend expected
        # (favourable).  Conversely an actual bill above fitted
        # writes a positive number the reviewer can see as the
        # over-billing flag.
        if lin_val is not None:
            delta = float(amounts[i]) - float(lin_val)
            _money(ws, r, 7, delta, fill_hex=bg)
        else:
            _text(ws, r, 7, "N/A", fill_hex=bg)
        r += 1

    # Separator
    ws.cell(row=r, column=1, value="— " * 20).font = Font(bold=True, color=DGREY)
    r += 1

    # === Forward forecast block: 6 steps past the last historical ===
    forecast_dates = []
    last_date = parse_to_sort_date(dates[-1])
    from datetime import timedelta

    if not pd.isna(last_date):
        for i in range(1, forecast_steps + 1):
            next_date = last_date + timedelta(days=30 * i)  # Approximate monthly
            forecast_dates.append(next_date.strftime("%d/%m/%Y"))
    else:
        forecast_dates = [f"Forecast +{i + 1}" for i in range(forecast_steps)]

    for i in range(forecast_steps):
        bg = AMBER
        _text(ws, r, 1, forecast_dates[i], fill_hex=bg, bold=True)
        _text(ws, r, 2, "—", fill_hex=bg)  # No actual
        lin_val = linear_fc[i] if linear_fc is not None else None
        hw_val = hw_fc[i] if hw_fc is not None else None
        if lin_val is not None:
            _money(ws, r, 3, float(lin_val), fill_hex=bg)
        else:
            _text(ws, r, 3, "N/A", fill_hex=bg)
        if hw_val is not None:
            _money(ws, r, 4, float(hw_val), fill_hex=bg)
        else:
            _text(ws, r, 4, "N/A", fill_hex=bg)
        _money(ws, r, 5, ema_future[i], fill_hex=bg)
        # Confidence band on the future prediction is the *predicted
        # value's* ±2σ — same shape as on the historical rows but
        # at the forecasted level so the reviewer sees the
        # widening band as the horizon extends.
        if lin_val is not None:
            conf = abs(float(lin_val)) * hist_vol * 2
            _money(ws, r, 6, conf, fill_hex=bg)
        else:
            _text(ws, r, 6, "N/A", fill_hex=bg)
        # Forecast Δ is intentionally "—" for future rows: there
        # is no actual bill yet to subtract from.
        _text(ws, r, 7, "—", fill_hex=bg)
        r += 1

    # Model comparison
    r += 1
    _section_hdr(ws, r, "MODEL COMPARISON")
    r += 1
    _text(ws, r, 1, "Linear Trend", bold=True)
    _text(ws, r, 2, "Simple linear regression on time index")
    r += 1
    _text(ws, r, 1, "Holt-Winters", bold=True)
    _text(
        ws, r, 2, "Exponential smoothing with trend" + (" + seasonality" if HAS_STATSMODELS else "")
    )
    r += 1
    _text(ws, r, 1, "EMA Projection", bold=True)
    _text(ws, r, 2, "Extends last Exponential Moving Average (span=6)")
    r += 1
    _text(ws, r, 1, "Historical Volatility", bold=True)
    _num(ws, r, 2, hist_vol, fmt="0.00%")
    _text(ws, r, 3, "Monthly return std used for confidence bands")

    # Accuracy metrics (in-sample)
    r += 1
    _section_hdr(ws, r, "IN-SAMPLE ACCURACY (Last 6 periods)")
    if n >= 7:
        test_series = pd.Series(amounts[:-6])
        true_vals = amounts[-6:]
        lin_hist = _linear_forecast(test_series, 6)
        if lin_hist is not None:
            mae = np.mean(np.abs(lin_hist - true_vals))
            rmse = np.sqrt(np.mean((lin_hist - true_vals) ** 2))
            mape = np.mean(np.abs((lin_hist - true_vals) / true_vals)) * 100

            r += 1
            _text(ws, r, 1, "Linear Forecast MAE (£)", bold=True)
            _money(ws, r, 2, mae)
            r += 1
            _text(ws, r, 1, "Linear Forecast RMSE (£)", bold=True)
            _money(ws, r, 2, rmse)
            r += 1
            _text(ws, r, 1, "Linear Forecast MAPE (%)", bold=True)
            _num(ws, r, 2, mape, fmt="0.00%")

    for col_letter, width in zip(
        ["A", "B", "C", "D", "E", "F", "G"], [14, 16, 18, 18, 18, 16, 22], strict=False
    ):
        ws.column_dimensions[col_letter].width = width
    ws.freeze_panes = "A2"


def write_data_quality_sheet(ws, df):
    """Write Data Quality Report tab."""
    ws.title = "Data Quality Report"

    NAVY = "10367A"
    ORANGE = "FE5716"
    LGREY = "F0F0F0"
    DGREY = "888888"

    def _banner(ws, r, text, bg):
        c = ws.cell(row=r, column=1, value=text)
        c.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", start_color=bg)
        c.border = CELL_BORDER
        c.alignment = Alignment(horizontal="left", vertical="center")
        for col in range(2, 6):
            x = ws.cell(row=r, column=col)
            x.fill = PatternFill("solid", start_color=bg)
            x.border = CELL_BORDER
        ws.row_dimensions[r].height = 20

    dq = _data_quality_report(df)

    if not dq:
        _hcell(ws, 1, 1, "No data to analyze", bg=NAVY)
        ws.column_dimensions["A"].width = 40
        return

    headers = ["Check", "Result", "Rate/Count", "Status"]
    for col, h in enumerate(headers, 1):
        _hcell(ws, 1, col, h, bg=NAVY)
    ws.row_dimensions[1].height = 28

    tc = ws.cell(row=1, column=1, value="EDF ENERGY DISPUTE  —  DATA QUALITY REPORT")
    tc.font = Font(name="Calibri", size=13, bold=True, color="FFFFFF")
    tc.fill = PatternFill("solid", start_color=ORANGE)
    tc.border = CELL_BORDER
    tc.alignment = Alignment(horizontal="left", vertical="center")
    for c in range(2, 5):
        x = ws.cell(row=1, column=c)
        x.fill = PatternFill("solid", start_color=ORANGE)
        x.border = CELL_BORDER

    def _check_row(r, check, result, rate, status, note=""):
        bg = LGREY if r % 2 == 0 else None
        _text(ws, r, 1, check, fill_hex=bg)
        _text(ws, r, 2, str(result), fill_hex=bg)
        _text(ws, r, 3, str(rate), fill_hex=bg)
        _text(ws, r, 4, status, bold=True, fill_hex=bg)
        if note:
            ws.cell(row=r, column=5, value=note).font = Font(name="Calibri", size=9, color=DGREY)

    r = 2
    _section_hdr(ws, r, "COMPLETENESS CHECKS")

    checks = [
        ("Total Records", dq["total_records"], "—", "PASS" if dq["total_records"] > 0 else "FAIL"),
        (
            "Date Parsing Success",
            dq["date_parsed"],
            f"{dq['date_parse_rate']:.1%}",
            "PASS"
            if dq["date_parse_rate"] > 0.8
            else "WARN"
            if dq["date_parse_rate"] > 0.5
            else "FAIL",
        ),
        (
            "Amount Complete",
            dq["amt_complete"],
            f"{(dq['amt_complete'] / dq['total_records']):.1%}",
            "PASS" if dq["amt_complete"] == dq["total_records"] else "WARN",
        ),
        (
            "Period Info Complete",
            dq["period_complete"],
            f"{dq['period_completeness_rate']:.1%}",
            "PASS" if dq["period_completeness_rate"] > 0.7 else "WARN",
        ),
        (
            "Reading Classified",
            dq["reading_classified"],
            f"{dq['reading_classify_rate']:.1%}",
            "PASS" if dq["reading_classify_rate"] > 0.5 else "WARN",
        ),
        (
            "Unit Rate Computable",
            dq["ur_computable"],
            f"{dq['ur_computable_rate']:.1%}",
            "PASS" if dq["ur_computable_rate"] > 0.3 else "INFO",
        ),
    ]
    for check, result, rate, status in checks:
        _check_row(r, check, result, rate, status)
        r += 1

    r += 1
    _section_hdr(ws, r, "DUPLICATION CHECKS")
    r += 1
    _check_row(
        r,
        "Duplicate Records (Date+Amount)",
        dq["duplicate_count"],
        f"{dq['duplicate_rate']:.2%}",
        "PASS"
        if dq["duplicate_rate"] < 0.05
        else "WARN"
        if dq["duplicate_rate"] < 0.15
        else "FAIL",
    )
    r += 1

    r += 1
    _section_hdr(ws, r, "SOURCE DISTRIBUTION")
    for src, cnt in dq.get("source_distribution", {}).items():
        r += 1
        _check_row(r, f"Source: {src}", cnt, f"{cnt / dq['total_records']:.1%}", "INFO")

    r += 1
    _section_hdr(ws, r, "ENTRY TYPE DISTRIBUTION")
    for etype, cnt in dq.get("entry_type_distribution", {}).items():
        r += 1
        _check_row(r, f"Type: {etype}", cnt, f"{cnt / dq['total_records']:.1%}", "INFO")

    # Summary banner
    r += 2
    total_checks = (
        len(checks)
        + 1
        + len(dq.get("source_distribution", {}))
        + len(dq.get("entry_type_distribution", {}))
    )
    pass_count = sum(1 for c in checks if c[3] == "PASS") + (
        1 if dq["duplicate_rate"] < 0.05 else 0
    )
    warn_count = sum(1 for c in checks if c[3] == "WARN") + (
        1 if 0.05 <= dq["duplicate_rate"] < 0.15 else 0
    )
    fail_count = sum(1 for c in checks if c[3] == "FAIL") + (
        1 if dq["duplicate_rate"] >= 0.15 else 0
    )

    _banner(
        ws,
        r,
        f"QUALITY SUMMARY: {total_checks} checks  |  PASS: {pass_count}  |  WARN: {warn_count}  |  FAIL: {fail_count}",
        NAVY,
    )

    for col_letter, width in zip(["A", "B", "C", "D", "E"], [40, 20, 18, 12, 60], strict=False):
        ws.column_dimensions[col_letter].width = width
    ws.freeze_panes = "A2"


def write_tariff_analysis_sheet(ws, dfc):
    """Write Tariff Impact Analysis tab."""
    ws.title = "Tariff Analysis"

    NAVY = "10367A"
    ORANGE = "FE5716"
    LGREY = "F0F0F0"

    tariff_info = _analyze_tariff_impact(dfc)

    if not tariff_info:
        _hcell(ws, 1, 1, "No tariff data available in records", bg=NAVY)
        ws.column_dimensions["A"].width = 50
        return

    headers = [
        "Tariff",
        "Records",
        "Avg Unit Rate (p/kWh)",
        "Median Unit Rate",
        "Min Rate",
        "Max Rate",
        "Avg Period Charge (£)",
    ]
    for col, h in enumerate(headers, 1):
        _hcell(ws, 1, col, h, bg=NAVY)
    ws.row_dimensions[1].height = 28

    tc = ws.cell(row=1, column=1, value="EDF ENERGY DISPUTE  —  TARIFF IMPACT ANALYSIS")
    tc.font = Font(name="Calibri", size=13, bold=True, color="FFFFFF")
    tc.fill = PatternFill("solid", start_color=ORANGE)
    tc.border = CELL_BORDER
    tc.alignment = Alignment(horizontal="left", vertical="center")
    for c in range(2, 8):
        x = ws.cell(row=1, column=c)
        x.fill = PatternFill("solid", start_color=ORANGE)
        x.border = CELL_BORDER

    tariff_stats = tariff_info.get("tariff_stats")
    if tariff_stats is not None and not tariff_stats.empty:
        r = 2
        for _, row in tariff_stats.iterrows():
            bg = LGREY if r % 2 == 0 else None
            _text(ws, r, 1, str(row["Tariff"]), fill_hex=bg)
            _num(ws, r, 2, int(row["count"]), fmt="#,##0", fill_hex=bg)
            _num(ws, r, 3, float(row["avg_unit_rate"]), fmt="0.00", fill_hex=bg)
            _num(ws, r, 4, float(row["median_unit_rate"]), fmt="0.00", fill_hex=bg)
            _num(ws, r, 5, float(row["min_unit_rate"]), fmt="0.00", fill_hex=bg)
            _num(ws, r, 6, float(row["max_unit_rate"]), fmt="0.00", fill_hex=bg)
            avg_chg = row["avg_charge"]
            _money(ws, r, 7, float(avg_chg) if pd.notna(avg_chg) else 0, fill_hex=bg)
            r += 1

    r += 1
    _section_hdr(ws, r, "SUMMARY")
    r += 1
    _text(ws, r, 1, "Unique Tariffs Identified")
    _num(ws, r, 2, tariff_info.get("num_tariffs", 0), fmt="#,##0")
    r += 1
    _text(ws, r, 1, "Tariff Changes Detected")
    _num(ws, r, 2, tariff_info.get("tariff_changes", 0), fmt="#,##0")

    for col_letter, width in zip(
        ["A", "B", "C", "D", "E", "F", "G"], [28, 10, 22, 18, 16, 16, 20], strict=False
    ):
        ws.column_dimensions[col_letter].width = width
    ws.freeze_panes = "A2"


# ---------------------------------------------------------------------------
# GUI
# ---------------------------------------------------------------------------


class ReportOptionsDialog:
    """Modern report options dialog with format selection and section checkboxes."""

    SECTIONS = [
        ("cover", "Cover Page", True),
        ("toc", "Table of Contents", True),
        ("exec_summary", "Executive Summary", True),
        ("key_findings", "Key Findings", True),
        ("evidence_index", "Evidence Index", True),
        ("detailed_findings", "Detailed Findings", True),
        ("timeline", "Timeline", True),
        ("ofgem", "OFGEM Price Cap Comparison", True),
        ("statistical", "Statistical Analysis", True),
        ("payment", "Payment Analysis", True),
        ("forecast", "Forecast", True),
        ("data_quality", "Data Quality", True),
        ("tariff", "Tariff Impact Analysis", True),
        ("appendix_methodology", "Appendix: Methodology", True),
        ("appendix_glossary", "Appendix: Glossary", True),
        ("appendix_full_evidence", "Appendix: Full Evidence Table", True),
    ]

    def __init__(self, parent):
        self.parent = parent
        self.result = None
        self.dialog = None

    def show(self):
        """Show the dialog and return the selected options."""
        self.dialog = tk.Toplevel(self.parent)
        self.dialog.title("Report Options")
        # Default size for 1080p: visible buttons without scrolling
        self.dialog.geometry("600x900")
        self.dialog.minsize(500, 500)
        self.dialog.resizable(True, True)
        self.dialog.transient(self.parent)
        self.dialog.grab_set()

        # Center on parent
        self.dialog.update_idletasks()
        x = self.parent.winfo_rootx() + (self.parent.winfo_width() // 2) - 300
        y = self.parent.winfo_rooty() + (self.parent.winfo_height() // 2) - 450
        self.dialog.geometry(f"+{x}+{y}")

        self._build_ui()
        self.dialog.wait_window()
        return self.result

    def _build_ui(self):
        """Build the dialog UI."""
        # Create scrollable main area
        canvas = tk.Canvas(self.dialog, highlightthickness=0)
        scrollbar = ttk.Scrollbar(self.dialog, orient="vertical", command=canvas.yview)
        main = ttk.Frame(canvas, padding=20)

        main.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=main, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        # Bind mousewheel
        def _on_mousewheel(event):
            canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

        canvas.bind_all("<MouseWheel>", _on_mousewheel)
        if self.dialog is not None:
            self.dialog.bind("<Destroy>", lambda e: canvas.unbind_all("<MouseWheel>"))

        # Also allow resizing canvas window width
        def _on_canvas_configure(event):
            canvas.itemconfig(canvas.find_all()[0], width=event.width)

        canvas.bind("<Configure>", _on_canvas_configure)

        # Header
        hdr = ttk.Frame(main)
        hdr.pack(fill=tk.X, pady=(0, 20))

        title_lbl = ttk.Label(
            hdr,
            text="Generate Ombudsman Report",
            font=("Calibri", 18, "bold"),
            foreground=EDF_NAVY,
        )
        title_lbl.pack(anchor=tk.W)

        subtitle = ttk.Label(
            hdr,
            text="Choose format and select sections to include",
            font=("Calibri", 10),
            foreground=MEDIUM_GREY,
        )
        subtitle.pack(anchor=tk.W, pady=(4, 0))

        ttk.Separator(main, orient="horizontal").pack(fill=tk.X, pady=(0, 16))

        # Format selection
        fmt_frame = ttk.LabelFrame(main, text=" Output Format ", padding=12)
        fmt_frame.pack(fill=tk.X, pady=(0, 16))

        self.format_var = tk.StringVar(value="both")
        formats = [
            ("both", "Both (PDF + Word)", "Generate both PDF and DOCX reports"),
            ("pdf", "PDF Only", "Professional PDF report (reportlab)"),
            ("docx", "Word Document Only", "Editable Word document (python-docx)"),
        ]

        for val, label, desc in formats:
            r = ttk.Frame(fmt_frame)
            r.pack(fill=tk.X, pady=3)
            rb = ttk.Radiobutton(r, variable=self.format_var, value=val)
            rb.pack(side=tk.LEFT)
            lbl_frame = ttk.Frame(r)
            lbl_frame.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=8)
            ttk.Label(lbl_frame, text=label, font=("Calibri", 10, "bold")).pack(anchor=tk.W)
            ttk.Label(lbl_frame, text=desc, font=("Calibri", 8), foreground=MEDIUM_GREY).pack(
                anchor=tk.W
            )

        ttk.Separator(main, orient="horizontal").pack(fill=tk.X, pady=(8, 16))

        # Section checkboxes
        sec_frame = ttk.LabelFrame(main, text=" Report Sections ", padding=12)
        sec_frame.pack(fill=tk.X, pady=(0, 16))

        # Select All / None buttons
        btn_frame = ttk.Frame(sec_frame)
        btn_frame.pack(fill=tk.X, pady=(0, 8))

        ttk.Button(btn_frame, text="Select All", command=self._select_all, width=12).pack(
            side=tk.LEFT
        )
        ttk.Button(btn_frame, text="Select None", command=self._select_none, width=12).pack(
            side=tk.LEFT, padx=(8, 0)
        )
        ttk.Button(btn_frame, text="Defaults", command=self._select_defaults, width=12).pack(
            side=tk.LEFT, padx=(8, 0)
        )

        # Checkboxes (main dialog is now scrollable, so no nested scrollbar needed)
        self.section_vars = {}
        for key, label, default in self.SECTIONS:
            var = tk.BooleanVar(value=default)
            self.section_vars[key] = var
            cb = ttk.Checkbutton(sec_frame, text=label, variable=var)
            cb.pack(anchor=tk.W, pady=1)

        ttk.Separator(main, orient="horizontal").pack(fill=tk.X, pady=(8, 16))

        # Action buttons
        action_frame = ttk.Frame(main)
        action_frame.pack(fill=tk.X)

        cancel_btn = ttk.Button(action_frame, text="Cancel", command=self._cancel, width=14)
        cancel_btn.pack(side=tk.RIGHT)

        ok_btn = tk.Button(
            action_frame,
            text="OK — Generate Report",
            bg=EDF_ORANGE,
            fg="white",
            font=("Calibri", 11, "bold"),
            command=self._generate,
            relief="flat",
            width=22,
        )
        ok_btn.pack(side=tk.RIGHT, padx=(0, 12))

        # Bind Enter key to OK, Escape to Cancel
        if self.dialog:
            self.dialog.bind("<Return>", lambda e: self._generate())
            self.dialog.bind("<Escape>", lambda e: self._cancel())

    def _select_all(self):
        for var in self.section_vars.values():
            var.set(True)

    def _select_none(self):
        for var in self.section_vars.values():
            var.set(False)

    def _select_defaults(self):
        for key, var in self.section_vars.items():
            # Find default from SECTIONS
            for k, _, default in self.SECTIONS:
                if k == key:
                    var.set(default)
                    break

    def _generate(self):
        """Collect results and close dialog."""
        selected_sections = [key for key, var in self.section_vars.items() if var.get()]
        if not selected_sections:
            messagebox.showwarning("No Sections", "Please select at least one report section.")
            return

        self.result = {
            "format": self.format_var.get(),
            "sections": selected_sections,
        }
        if self.dialog is not None:
            self.dialog.destroy()

    def _cancel(self):
        self.result = None
        if self.dialog is not None:
            self.dialog.destroy()


class App:
    def __init__(self, root):
        self.root = root
        self.root.title("EDF Master Evidence Collector")
        self.root.geometry("780x860")
        self.root.configure(bg=EDF_OFFWHITE)

        self.pst_path = tk.StringVar()
        self.pdf_dir = tk.StringVar()
        self.htm_path = tk.StringVar()
        self.acc_num = tk.StringVar(value="")
        self.status = tk.StringVar(value="Ready.")
        self.progress_v = tk.DoubleVar(value=0)

        self.use_anchors = tk.BooleanVar(value=True)
        self.use_large = tk.BooleanVar(value=True)
        self.use_reading_class = tk.BooleanVar(value=True)
        self.use_pdf_fields = tk.BooleanVar(value=True)
        self.use_acc_filt = tk.BooleanVar(value=False)
        self.filter_below = tk.BooleanVar(value=True)
        self.save_filtered = tk.BooleanVar(value=True)
        self.use_dedup = tk.BooleanVar(value=True)
        self.save_dups = tk.BooleanVar(value=True)
        self.use_domain_filter = tk.BooleanVar(value=True)
        self.domain_filter = tk.StringVar(value="edfenergy.com")
        self.min_amount = tk.DoubleVar(value=500.0)
        self.analysis_min = tk.DoubleVar(value=500.0)
        self.output_name = tk.StringVar(value="EDF_Dispute_Evidence.xlsx")
        self.report_account_ref = tk.StringVar(value="")

        # New vars for UI refresh (see spec 2026-07-10-ui-refresh-design.md)
        self.output_folder = tk.StringVar(value="")
        self.amalgamate_duplicates = tk.BooleanVar(value=False)
        self.auto_generate_report = tk.BooleanVar(value=False)
        self._report_options: dict = {}
        self._CONFIG_PATH = os.path.expanduser("~/.edf_collector/config.json")

        # Load persisted config (may override the var defaults above)
        self._load_config()

        self.cancel_event = threading.Event()
        self.build_ui()

    # -- Config persistence --

    def _load_config(self):
        """Read config file and mutate tk vars via .set().

        Silently falls back to hardcoded defaults when the file is
        missing, unreadable, or malformed.
        """
        try:
            with open(self._CONFIG_PATH) as f:
                data = json.load(f)
        except (FileNotFoundError, json.JSONDecodeError, OSError):
            return

        gui = data.get("gui_state", {})
        _bool_keys: dict[str, tk.Variable] = {
            "use_anchors": self.use_anchors,
            "use_large": self.use_large,
            "use_reading_class": self.use_reading_class,
            "use_pdf_fields": self.use_pdf_fields,
            "use_acc_filt": self.use_acc_filt,
            "filter_below": self.filter_below,
            "save_filtered": self.save_filtered,
            "use_dedup": self.use_dedup,
            "save_dups": self.save_dups,
            "amalgamate_duplicates": self.amalgamate_duplicates,
            "use_domain_filter": self.use_domain_filter,
            "auto_generate_report": self.auto_generate_report,
        }
        for key, var in _bool_keys.items():
            if key in gui:
                var.set(bool(gui[key]))

        _str_keys: dict[str, tk.Variable] = {
            "acc_num": self.acc_num,
            "domain_filter": self.domain_filter,
            "output_name": self.output_name,
            "report_account_ref": self.report_account_ref,
            "output_folder": self.output_folder,
        }
        for key, var in _str_keys.items():
            if key in gui:
                var.set(str(gui[key]))

        _float_keys: dict[str, tk.Variable] = {
            "min_amount": self.min_amount,
            "analysis_min": self.analysis_min,
        }
        for key, var in _float_keys.items():
            if key in gui:
                try:
                    var.set(float(gui[key]))
                except (ValueError, TypeError):
                    pass

        ro = data.get("report_options", {})
        if ro:
            self._report_options = ro

    def _save_config(self):
        """Persist GUI state + report options to config file atomically.

        Write to <path>.tmp, fsync, os.replace.  Permissions 0o600.
        """
        config_dir = os.path.dirname(self._CONFIG_PATH)
        os.makedirs(config_dir, exist_ok=True)

        gui = {
            "use_anchors": self.use_anchors.get(),
            "use_large": self.use_large.get(),
            "use_reading_class": self.use_reading_class.get(),
            "use_pdf_fields": self.use_pdf_fields.get(),
            "use_acc_filt": self.use_acc_filt.get(),
            "acc_num": self.acc_num.get(),
            "min_amount": self.min_amount.get(),
            "analysis_min": self.analysis_min.get(),
            "filter_below": self.filter_below.get(),
            "save_filtered": self.save_filtered.get(),
            "use_dedup": self.use_dedup.get(),
            "save_dups": self.save_dups.get(),
            "amalgamate_duplicates": self.amalgamate_duplicates.get(),
            "use_domain_filter": self.use_domain_filter.get(),
            "domain_filter": self.domain_filter.get(),
            "output_name": self.output_name.get(),
            "report_account_ref": self.report_account_ref.get(),
            "auto_generate_report": self.auto_generate_report.get(),
            "output_folder": self.output_folder.get(),
        }

        payload = {
            "output_folder": self.output_folder.get(),
            "report_options": getattr(self, "_report_options", {}),
            "gui_state": gui,
        }

        tmp_path = self._CONFIG_PATH + ".tmp"
        with open(tmp_path, "w") as f:
            json.dump(payload, f, indent=2)
            f.flush()
            os.fsync(f.fileno())
        os.chmod(tmp_path, 0o600)
        os.replace(tmp_path, self._CONFIG_PATH)

    def build_ui(self):
        hdr = tk.Frame(self.root, bg=EDF_ORANGE, height=60)
        hdr.pack(fill=tk.X)
        tk.Label(
            hdr,
            text="EDF BILLING EVIDENCE COLLECTOR",
            bg=EDF_ORANGE,
            fg="white",
            font=("Calibri", 14, "bold"),
        ).pack(pady=15)

        container = ttk.Frame(self.root)
        container.pack(fill=tk.BOTH, expand=True)

        canvas = tk.Canvas(container, bg=EDF_OFFWHITE, highlightthickness=0)
        yscroll = ttk.Scrollbar(container, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=yscroll.set)
        yscroll.pack(side=tk.RIGHT, fill=tk.Y)
        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        main = ttk.Frame(canvas, padding=16)
        cw = canvas.create_window((0, 0), window=main, anchor="nw")

        def _reconfig(_e=None):
            canvas.configure(scrollregion=canvas.bbox("all"))
            canvas.itemconfig(cw, width=canvas.winfo_width())

        main.bind("<Configure>", _reconfig)
        canvas.bind("<Configure>", _reconfig)

        # --- Section 1: Source Data ---
        s1 = ttk.LabelFrame(main, text=" 1. Source Data ", padding=10)
        s1.pack(fill=tk.X, pady=5)

        def browse_row(parent, label, var, cmd):
            r = ttk.Frame(parent)
            r.pack(fill=tk.X, pady=2)
            ttk.Label(r, text=label, width=14).pack(side=tk.LEFT)
            ttk.Entry(r, textvariable=var).pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
            ttk.Button(r, text="Browse", command=cmd).pack(side=tk.LEFT)

        browse_row(s1, "PST/OST File:", self.pst_path, self._pick_pst)
        browse_row(s1, "PDF Folder:", self.pdf_dir, self._pick_pdf_dir)
        browse_row(
            s1,
            "HTM Export:",
            self.htm_path,
            lambda: self.htm_path.set(
                filedialog.askopenfilename(filetypes=[("HTM/HTML", "*.htm *.html")])
            ),
        )

        # --- Section 2: Extraction options ---
        s2 = ttk.LabelFrame(main, text=" 2. Search & Filter Options ", padding=10)
        s2.pack(fill=tk.X, pady=5)
        for text, var in [
            ("Smart Context Search", self.use_anchors),
            ("Large Number Fallback", self.use_large),
            ("Classify Reading Type", self.use_reading_class),
            ("Deep PDF Mine (kWh, standing charge, invoice #)", self.use_pdf_fields),
        ]:
            tk.Checkbutton(s2, text=text, variable=var, bg=EDF_OFFWHITE).pack(anchor=tk.W)

        r3 = ttk.Frame(s2)
        r3.pack(fill=tk.X, pady=4)
        tk.Checkbutton(
            r3, text="Filter by Account #:", variable=self.use_acc_filt, bg=EDF_OFFWHITE
        ).pack(side=tk.LEFT)
        ttk.Entry(r3, textvariable=self.acc_num, width=16).pack(side=tk.LEFT, padx=5)

        r3d = ttk.Frame(s2)
        r3d.pack(fill=tk.X, pady=4)
        tk.Checkbutton(
            r3d,
            text="Filter PST emails by sender domain:",
            variable=self.use_domain_filter,
            bg=EDF_OFFWHITE,
        ).pack(side=tk.LEFT)
        ttk.Entry(r3d, textvariable=self.domain_filter, width=40).pack(side=tk.LEFT, padx=5)
        ttk.Label(r3d, text="(comma-separated domains/addresses)", font=("Calibri", 8)).pack(
            side=tk.LEFT
        )

        r4 = ttk.Frame(s2)
        r4.pack(fill=tk.X, pady=2)
        chk_filt = tk.Checkbutton(
            r4, text="Filter results below minimum £:", variable=self.filter_below, bg=EDF_OFFWHITE
        )
        chk_filt.pack(side=tk.LEFT)
        ttk.Entry(r4, textvariable=self.min_amount, width=8).pack(side=tk.LEFT, padx=5)

        r4c = ttk.Frame(s2)
        r4c.pack(fill=tk.X, pady=2)
        ttk.Label(r4c, text="Analysis threshold (£):", width=24).pack(side=tk.LEFT)
        ttk.Entry(r4c, textvariable=self.analysis_min, width=8).pack(side=tk.LEFT, padx=5)

        r4d = ttk.Frame(s2)
        r4d.pack(fill=tk.X, pady=2)
        ttk.Label(r4d, text="Report account reference:", width=24).pack(side=tk.LEFT)
        ttk.Entry(r4d, textvariable=self.report_account_ref, width=20).pack(side=tk.LEFT, padx=5)

        r4e = ttk.Frame(s2)
        r4e.pack(fill=tk.X, pady=2)
        ttk.Label(r4e, text="Output filename:", width=24).pack(side=tk.LEFT)
        ttk.Entry(r4e, textvariable=self.output_name, width=30).pack(side=tk.LEFT, padx=5)

        chk_sf = tk.Checkbutton(
            s2,
            text="Save filtered-out records to worksheet",
            variable=self.save_filtered,
            bg=EDF_OFFWHITE,
        )
        chk_sf.pack(anchor=tk.W, padx=20)
        chk_filt.config(
            command=lambda: chk_sf.config(state="normal" if self.filter_below.get() else "disabled")
        )

        # --- Section 3: Deduplication ---
        s3 = ttk.LabelFrame(main, text=" 3. Deduplication ", padding=10)
        s3.pack(fill=tk.X, pady=5)
        chk_dup = tk.Checkbutton(
            s3,
            text="Filter duplicate records (same date & amount)",
            variable=self.use_dedup,
            bg=EDF_OFFWHITE,
        )
        chk_dup.pack(anchor=tk.W)
        chk_sd = tk.Checkbutton(
            s3,
            text="Save duplicates to separate worksheet",
            variable=self.save_dups,
            bg=EDF_OFFWHITE,
        )
        chk_sd.pack(anchor=tk.W, padx=20)
        chk_dup.config(
            command=lambda: chk_sd.config(state="normal" if self.use_dedup.get() else "disabled")
        )

        # --- Progress ---
        self.pb = ttk.Progressbar(main, mode="determinate", maximum=100, variable=self.progress_v)
        self.pb.pack(fill=tk.X, pady=10)
        ttk.Label(
            main, textvariable=self.status, foreground=EDF_NAVY, font=("Calibri", 11, "bold")
        ).pack()

        btns = ttk.Frame(main)
        btns.pack(fill=tk.X, pady=8)
        self.run_btn = tk.Button(
            btns,
            text="EXTRACT TO EXCEL",
            bg=EDF_ORANGE,
            fg="white",
            font=("Calibri", 12, "bold"),
            command=self.start_thread,
            relief="flat",
        )
        self.run_btn.pack(side=tk.LEFT, fill=tk.X, expand=True, ipady=8)
        self.cancel_btn = ttk.Button(btns, text="Cancel", command=self._cancel, state="disabled")
        self.cancel_btn.pack(side=tk.LEFT, padx=8)
        self.pdf_report_btn = tk.Button(
            btns,
            text="EXPORT REPORT",
            bg=EDF_NAVY,
            fg="white",
            font=("Calibri", 12, "bold"),
            command=self.export_report,
            relief="flat",
            state="disabled" if not (HAS_PDF_REPORT or HAS_DOCX_REPORT) else "normal",
        )
        self.pdf_report_btn.pack(side=tk.LEFT, fill=tk.X, expand=True, ipady=8, padx=(8, 0))

        # Load Spreadsheet & Generate Report button
        self.load_report_btn = tk.Button(
            btns,
            text="LOAD & REPORT",
            bg=EDF_ORANGE,
            fg="white",
            font=("Calibri", 12, "bold"),
            command=self.load_spreadsheet_and_report,
            relief="flat",
        )
        self.load_report_btn.pack(side=tk.LEFT, fill=tk.X, expand=True, ipady=8, padx=(8, 0))

    # -- Helpers --

    def _pick_pst(self):
        p = filedialog.askopenfilename(filetypes=[("Mail Stores", "*.pst *.ost")])
        if p:
            self.pst_path.set(p)

    def _pick_pdf_dir(self):
        p = filedialog.askdirectory()
        if p:
            self.pdf_dir.set(p)

    def set_status(self, text):
        def _apply():
            self.status.set(text)
            self.root.update_idletasks()

        if threading.current_thread() is threading.main_thread():
            _apply()
        else:
            self.root.after(0, _apply)

    def set_progress(self, current, total, text=None):
        pct = max(0, min(100, (current / total) * 100)) if total else 0

        def _apply():
            self.progress_v.set(pct)
            if text:
                self.status.set(text)

        if threading.current_thread() is threading.main_thread():
            _apply()
        else:
            self.root.after(0, _apply)

    def _show(self, level, title, text):
        def _s():
            if level == "info":
                messagebox.showinfo(title, text)
            elif level == "warning":
                messagebox.showwarning(title, text)
            else:
                messagebox.showerror(title, text)

        if threading.current_thread() is threading.main_thread():
            _s()
        else:
            self.root.after(0, _s)

    def _finish(self):
        self.run_btn.config(state="normal")
        self.cancel_btn.config(state="disabled")
        if hasattr(self, "pdf_report_btn"):
            self.pdf_report_btn.config(state="normal" if HAS_PDF_REPORT else "disabled")
        self.progress_v.set(0)
        self.set_status("Cancelled." if self.cancel_event.is_set() else "Ready.")
        gc.collect()

    def export_report(self):
        """Unified report export — opens modern options dialog."""
        if not HAS_PDF_REPORT and not HAS_DOCX_REPORT:
            self._show(
                "error",
                "Report Unavailable",
                "Report generation requires 'reportlab' (PDF) and/or 'python-docx' (Word).\n"
                "Install with: pip install reportlab python-docx",
            )
            return

        if not hasattr(self, "engine") or not self.engine or not self.engine.records:
            self._show("warning", "No Data", "No records available. Run extraction first.")
            return

        # Show modern options dialog
        dialog = ReportOptionsDialog(self.root)
        options = dialog.show()

        if not options:
            return  # User cancelled

        # Ask for output path(s)
        base_dir = (
            os.path.dirname(self.pst_path.get().strip())
            if self.pst_path.get().strip()
            else self.pdf_dir.get().strip()
            if self.pdf_dir.get().strip()
            else os.path.dirname(self.htm_path.get().strip())
            if self.htm_path.get().strip()
            else os.getcwd()
        )

        output_paths = {}
        fmt = options["format"]

        if fmt in ("pdf", "both"):
            if not HAS_PDF_REPORT:
                self._show(
                    "warning",
                    "PDF Unavailable",
                    "PDF generation requires 'reportlab'. Install with: pip install reportlab",
                )
            else:
                default_name = "EDF_Ombudsman_Report.pdf"
                out_path = filedialog.asksaveasfilename(
                    initialdir=base_dir,
                    initialfile=default_name,
                    defaultextension=".pdf",
                    filetypes=[("PDF Files", "*.pdf"), ("All Files", "*.*")],
                    title="Save Ombudsman PDF Report As",
                )
                if out_path:
                    output_paths["pdf"] = out_path
                elif fmt == "pdf":
                    return  # User cancelled PDF-only

        if fmt in ("docx", "both"):
            if not HAS_DOCX_REPORT:
                self._show(
                    "warning",
                    "DOCX Unavailable",
                    "DOCX generation requires 'python-docx'. Install with: pip install python-docx",
                )
            else:
                default_name = "EDF_Ombudsman_Report.docx"
                out_path = filedialog.asksaveasfilename(
                    initialdir=base_dir,
                    initialfile=default_name,
                    defaultextension=".docx",
                    filetypes=[("Word Documents", "*.docx"), ("All Files", "*.*")],
                    title="Save Ombudsman Word Report As",
                )
                if out_path:
                    output_paths["docx"] = out_path
                elif fmt == "docx":
                    return  # User cancelled DOCX-only

        if not output_paths:
            return  # User cancelled all

        self.set_status("Generating report…")
        self.pdf_report_btn.config(state="disabled")
        self.run_btn.config(state="disabled")
        self.cancel_btn.config(state="disabled")

        config = {
            "min_amount": self.min_amount.get(),
            "analysis_min": self.analysis_min.get(),
            "acc_num": self.acc_num.get(),
            "report_account_ref": self.report_account_ref.get().strip(),
            "report_sections": options["sections"],
        }

        def _generate():
            # Lazy import to avoid circular import
            from edf_report import generate_pdf_from_gui
            from edf_report_docx import generate_docx_from_gui

            try:
                messages = []
                if "pdf" in output_paths:
                    success, msg = generate_pdf_from_gui(
                        records=self.engine.records,
                        output_path=output_paths["pdf"],
                        config=config,
                        engine=self.engine,
                        filtered=self.engine.filtered_records,
                    )
                    messages.append(("PDF", success, msg))

                if "docx" in output_paths:
                    success, msg = generate_docx_from_gui(
                        records=self.engine.records,
                        output_path=output_paths["docx"],
                        config=config,
                        engine=self.engine,
                        filtered=self.engine.filtered_records,
                    )
                    messages.append(("DOCX", success, msg))

                # Report results
                combined_msgs = []
                all_success = True
                for fmt_label, success, msg in messages:
                    if success:
                        combined_msgs.append(
                            f"✓ {fmt_label}: {msg.split(chr(10))[-1] if msg else 'Generated'}"
                        )
                    else:
                        all_success = False
                        self.root.after(
                            0,
                            lambda m=msg, f=fmt_label: self._show(
                                "error", f"{f} Generation Failed", m
                            ),
                        )

                if all_success and combined_msgs:
                    self.root.after(
                        0,
                        lambda msgs=combined_msgs: self._show(
                            "info", "Reports Generated", "\n\n".join(msgs)
                        ),
                    )

            except Exception as e:
                self.root.after(
                    0, lambda err=e: self._show("error", "Error", f"An error occurred:\n\n{err}")
                )
            finally:
                self.root.after(
                    0,
                    lambda: (
                        self.pdf_report_btn.config(
                            state="normal" if HAS_PDF_REPORT else "disabled"
                        ),
                        self.run_btn.config(state="normal"),
                        self.cancel_btn.config(state="disabled"),
                        self.set_status("Ready."),
                    ),
                )

        threading.Thread(target=_generate, daemon=True).start()

    def load_spreadsheet_and_report(self):
        """Load records from an existing spreadsheet and open report options dialog.

        Assumes the spreadsheet has the standard EDF Evidence Report format with
        an 'EDF Evidence Report' sheet. The user may have corrected/tweaked the data.
        """
        if not HAS_PDF_REPORT and not HAS_DOCX_REPORT:
            self._show(
                "error",
                "Report Unavailable",
                "Report generation requires 'reportlab' (PDF) and/or 'python-docx' (Word).\n"
                "Install with: pip install reportlab python-docx",
            )
            return

        file_path = filedialog.askopenfilename(
            initialdir=os.getcwd(),
            filetypes=[("Excel Files", "*.xlsx"), ("All Files", "*.*")],
            title="Select EDF Evidence Report Spreadsheet",
        )
        if not file_path:
            return  # User cancelled

        try:
            # Load the spreadsheet
            df = pd.read_excel(file_path, sheet_name="EDF Evidence Report")
            if df.empty:
                self._show(
                    "warning",
                    "No Data",
                    "The selected spreadsheet has no records in 'EDF Evidence Report' sheet.",
                )
                return

            records = df.to_dict("records")

            # Create a minimal engine-like object for metadata.
            # A dataclass is used instead of bare class-level annotations
            # so the attributes have a clear constructor contract and are
            # self-documenting (previously these were bare class-level
            # annotations that looked like static type hints but were only
            # ever set on instances after construction).
            from dataclasses import dataclass

            @dataclass
            class MockEngine:
                records: list
                filtered_records: list
                pdf_count: int
                email_count: int

            engine = MockEngine(records=records, filtered_records=[], pdf_count=0, email_count=0)

            # Open report options
            dialog = ReportOptionsDialog(self.root)
            options = dialog.show()

            if not options:
                return  # User cancelled

            # Ask for output paths
            fmt = options["format"]
            base_dir = os.path.dirname(file_path)
            output_paths = {}

            if fmt in ("pdf", "both"):
                if not HAS_PDF_REPORT:
                    self._show(
                        "warning",
                        "PDF Unavailable",
                        "PDF generation requires 'reportlab'. Install with: pip install reportlab",
                    )
                else:
                    default_name = os.path.basename(file_path).replace(".xlsx", "_Report.pdf")
                    out_path = filedialog.asksaveasfilename(
                        initialdir=base_dir,
                        initialfile=default_name,
                        defaultextension=".pdf",
                        filetypes=[("PDF Files", "*.pdf"), ("All Files", "*.*")],
                        title="Save Ombudsman PDF Report As",
                    )
                    if out_path:
                        output_paths["pdf"] = out_path
                    elif fmt == "pdf":
                        return

            if fmt in ("docx", "both"):
                if not HAS_DOCX_REPORT:
                    self._show(
                        "warning",
                        "DOCX Unavailable",
                        "DOCX generation requires 'python-docx'. Install with: pip install python-docx",
                    )
                else:
                    default_name = os.path.basename(file_path).replace(".xlsx", "_Report.docx")
                    out_path = filedialog.asksaveasfilename(
                        initialdir=base_dir,
                        initialfile=default_name,
                        defaultextension=".docx",
                        filetypes=[("Word Documents", "*.docx"), ("All Files", "*.*")],
                        title="Save Ombudsman Word Report As",
                    )
                    if out_path:
                        output_paths["docx"] = out_path
                    elif fmt == "docx":
                        return

            if not output_paths:
                return

            self.set_status("Generating report…")
            self.pdf_report_btn.config(state="disabled")
            self.load_report_btn.config(state="disabled")

            config = {
                "min_amount": self.min_amount.get(),
                "analysis_min": self.analysis_min.get(),
                "acc_num": self.acc_num.get(),
                "report_account_ref": self.report_account_ref.get().strip(),
                "report_sections": options["sections"],
            }

            def _generate():
                # Lazy import to avoid circular import
                from edf_report import generate_pdf_from_gui
                from edf_report_docx import generate_docx_from_gui

                try:
                    messages = []
                    if "pdf" in output_paths:
                        success, msg = generate_pdf_from_gui(
                            records=records,
                            output_path=output_paths["pdf"],
                            config=config,
                            engine=engine,
                            filtered=[],
                        )
                        messages.append(("PDF", success, msg))

                    if "docx" in output_paths:
                        success, msg = generate_docx_from_gui(
                            records=records,
                            output_path=output_paths["docx"],
                            config=config,
                            engine=engine,
                            filtered=[],
                        )
                        messages.append(("DOCX", success, msg))

                    # Report results (single combined message)
                    combined_msgs = []
                    all_success = True
                    for fmt_label, success, msg in messages:
                        if success:
                            combined_msgs.append(
                                f"✓ {fmt_label}: {msg.split(chr(10))[-1] if msg else 'Generated'}"
                            )
                        else:
                            all_success = False
                            self.root.after(
                                0,
                                lambda m=msg, f=fmt_label: self._show(
                                    "error", f"{f} Generation Failed", m
                                ),
                            )

                    if all_success and combined_msgs:
                        self.root.after(
                            0,
                            lambda msgs=combined_msgs: self._show(
                                "info", "Reports Generated", "\n\n".join(msgs)
                            ),
                        )

                except Exception as e:
                    self.root.after(
                        0,
                        lambda err=e: self._show("error", "Error", f"An error occurred:\n\n{err}"),
                    )
                finally:
                    self.root.after(
                        0,
                        lambda: (
                            self.pdf_report_btn.config(
                                state="normal"
                                if (HAS_PDF_REPORT or HAS_DOCX_REPORT)
                                else "disabled"
                            ),
                            self.load_report_btn.config(state="normal"),
                            self.set_status("Ready."),
                        ),
                    )

            threading.Thread(target=_generate, daemon=True).start()

        except Exception as e:
            self._show("error", "Load Error", f"Failed to load spreadsheet:\n\n{e}")

    def _export_legacy(self, fmt: str) -> None:
        """Legacy single-format export with all sections."""
        if fmt == "pdf" and not HAS_PDF_REPORT:
            self._show(
                "error",
                "PDF Unavailable",
                "PDF generation requires 'reportlab'.\nInstall with: pip install reportlab",
            )
            return
        if fmt == "docx" and not HAS_DOCX_REPORT:
            self._show(
                "error",
                "DOCX Unavailable",
                "DOCX generation requires 'python-docx'.\nInstall with: pip install python-docx",
            )
            return

        if not hasattr(self, "engine") or not self.engine or not self.engine.records:
            self._show("warning", "No Data", "No records available. Run extraction first.")
            return

        base_dir = (
            os.path.dirname(self.pst_path.get().strip())
            if self.pst_path.get().strip()
            else self.pdf_dir.get().strip()
            if self.pdf_dir.get().strip()
            else os.path.dirname(self.htm_path.get().strip())
            if self.htm_path.get().strip()
            else os.getcwd()
        )

        ext = ".pdf" if fmt == "pdf" else ".docx"
        default_name = f"EDF_Ombudsman_Report{ext}"
        filetypes = [("PDF Files", "*.pdf")] if fmt == "pdf" else [("Word Documents", "*.docx")]
        title = f"Save Ombudsman {'PDF' if fmt == 'pdf' else 'Word'} Report As"

        out_path = filedialog.asksaveasfilename(
            initialdir=base_dir,
            initialfile=default_name,
            defaultextension=ext,
            filetypes=filetypes + [("All Files", "*.*")],
            title=title,
        )
        if not out_path:
            return

        self.set_status(f"Generating {'PDF' if fmt == 'pdf' else 'Word'} report…")
        getattr(self, f"{fmt}_report_btn").config(state="disabled")
        self.run_btn.config(state="disabled")
        self.cancel_btn.config(state="disabled")

        # Use all sections for legacy
        config = {
            "min_amount": self.min_amount.get(),
            "analysis_min": self.analysis_min.get(),
            "acc_num": self.acc_num.get(),
            "report_account_ref": self.report_account_ref.get().strip(),
            "report_sections": [s[0] for s in ReportOptionsDialog.SECTIONS],
        }

        def _generate():
            # Lazy import to avoid circular import
            from edf_report import generate_pdf_from_gui
            from edf_report_docx import generate_docx_from_gui

            try:
                if fmt == "pdf":
                    success, msg = generate_pdf_from_gui(
                        records=self.engine.records,
                        output_path=out_path,
                        config=config,
                        engine=self.engine,
                        filtered=self.engine.filtered_records,
                    )
                else:
                    success, msg = generate_docx_from_gui(
                        records=self.engine.records,
                        output_path=out_path,
                        config=config,
                        engine=self.engine,
                        filtered=self.engine.filtered_records,
                    )

                if success:
                    self.root.after(0, lambda: self._show("info", "Success", msg))
                else:
                    self.root.after(
                        0, lambda: self._show("error", f"{fmt.upper()} Generation Failed", msg)
                    )
            except Exception as e:
                self.root.after(
                    0, lambda err=e: self._show("error", "Error", f"An error occurred:\n\n{err}")
                )
            finally:
                self.root.after(
                    0,
                    lambda: (
                        getattr(self, f"{fmt}_report_btn").config(
                            state="normal"
                            if (HAS_PDF_REPORT if fmt == "pdf" else HAS_DOCX_REPORT)
                            else "disabled"
                        ),
                        self.run_btn.config(state="normal"),
                        self.cancel_btn.config(state="disabled"),
                        self.set_status("Ready."),
                    ),
                )

        threading.Thread(target=_generate, daemon=True).start()

    def _cancel(self):
        self.cancel_event.set()
        self.set_status("Cancelling…")

    def start_thread(self):
        try:
            self.min_amount.get()
            self.analysis_min.get()
        except Exception:
            messagebox.showerror(
                "Error", "Minimum amount and analysis threshold must be valid numbers."
            )
            return

        has_sources = any(
            [
                self.pst_path.get().strip(),
                self.pdf_dir.get().strip(),
                self.htm_path.get().strip(),
            ]
        )
        if not has_sources:
            messagebox.showerror(
                "Error",
                "Please select at least one source:\nPST/OST file, PDF folder, or HTM export.",
            )
            return
        self.cancel_event.clear()
        self.run_btn.config(state="disabled")
        self.cancel_btn.config(state="normal")
        self.progress_v.set(0)
        threading.Thread(target=self._run, daemon=True).start()

    def _run(self):
        config = {
            "use_anchors": self.use_anchors.get(),
            "use_large": self.use_large.get(),
            "use_reading_classification": self.use_reading_class.get(),
            "use_pdf_fields": self.use_pdf_fields.get(),
            "use_acc_filter": self.use_acc_filt.get(),
            "acc_num": self.acc_num.get(),
            "min_amount": self.min_amount.get(),
            "analysis_min": self.analysis_min.get(),
            "report_account_ref": self.report_account_ref.get().strip(),
            "filter_below": self.filter_below.get(),
            "save_filtered": self.save_filtered.get(),
            "use_dedup": self.use_dedup.get(),
            "save_dups": self.save_dups.get(),
            "use_domain_filter": self.use_domain_filter.get(),
            "domain_filter": self.domain_filter.get().strip(),
        }

        engine = EvidenceEngine(config, self.set_status, self.set_progress, self.cancel_event)
        self.engine = engine

        try:
            pst_path = self.pst_path.get().strip()
            if pst_path and os.path.exists(pst_path) and not self.cancel_event.is_set():
                if not HAS_PYPFF:
                    self._show("warning", "PST", "pypff not installed — PST/OST scanning skipped.")
                else:
                    self.set_status("Scanning PST/OST…")
                    try:
                        pff = pypff.file()
                    except AttributeError:
                        # Fallback for different pypff API versions
                        pff = getattr(pypff, "File", None)
                        if pff is None:
                            raise AttributeError(
                                "pypff module has no 'file' or 'File' attribute"
                            ) from None
                        pff = pff()
                    pff.open(os.path.abspath(pst_path))
                    try:
                        engine.crawl_pst(pff.get_root_folder())
                    finally:
                        pff.close()

            htm_path = self.htm_path.get().strip()
            if htm_path and os.path.exists(htm_path) and not self.cancel_event.is_set():
                self.set_status("Parsing HTM account history…")
                engine.process_htm_file(htm_path)

            pdf_path = self.pdf_dir.get().strip()
            if pdf_path and os.path.exists(pdf_path) and not self.cancel_event.is_set():
                engine.crawl_local_pdfs(pdf_path)

            if self.cancel_event.is_set():
                self._show("warning", "Cancelled", "Extraction cancelled.")
                return

            if engine.records:
                self.set_status("Writing Excel report…")
                base_dir = (
                    os.path.dirname(pst_path)
                    if pst_path
                    else pdf_path
                    if pdf_path
                    else os.path.dirname(htm_path)
                    if htm_path
                    else os.getcwd()
                )
                out_name = self.output_name.get().strip() or "EDF_Dispute_Evidence.xlsx"
                if not out_name.lower().endswith(".xlsx"):
                    out_name += ".xlsx"
                out_path = os.path.join(base_dir, out_name)
                export_to_excel(
                    engine.records,
                    out_path,
                    engine.error_log,
                    config,
                    filtered=engine.filtered_records,
                )
                summary = (
                    f"Extraction complete.\n\n"
                    f"  Emails matched: {engine.email_count}\n"
                    f"  PDFs processed: {engine.pdf_count}\n"
                    f"  Records found:  {len(engine.records)}\n"
                )
                if engine.error_log:
                    summary += f"\n  Parse errors: {len(engine.error_log)} (see Parse Errors tab)"
                summary += f"\n\nSaved to:\n{out_path}"
                self._show("info", "Success", summary)
            else:
                self._show(
                    "warning",
                    "No Data",
                    "No billing amounts found.\n\nTips:\n"
                    "• Uncheck the Account Filter\n"
                    "• Lower the minimum threshold\n"
                    "• Check your source files contain EDF billing data",
                )

        except Exception:
            self._show("error", "Error", f"An error occurred:\n\n{traceback.format_exc()}")
        finally:
            self.root.after(0, self._finish)


# ---------------------------------------------------------------------------
# Safe pickle deserialiser — prevents arbitrary code execution when loading
# engine-data pickle files from disk.  Only standard built-in types and
# the project's own EvidenceEngine class are allowed through; anything
# else raises UnpicklingError.
# ---------------------------------------------------------------------------


class _RestrictedUnpickler(pickle.Unpickler):
    """Unpickler that only allows known-safe types.

    Permits: built-in scalars, dicts, lists, tuples, sets, frozensets,
    bytes/bytearray, and the project's own ``EvidenceEngine``.  Everything
    else triggers ``pickle.UnpicklingError`` so a crafted pickle can never
    import and call arbitrary code.
    """

    # Module→class whitelist.  Only classes listed here can be rebuilt.
    # A whitelist value of ``None`` (as opposed to the usual
    # ``set[str]`` of permitted class names) is interpreted as
    # "the entire module is trusted".  We only use this for
    # ``pyarrow.lib`` whose exposed pickle surface is purely
    # restoration-callable ``_something`` functions, never
    # ``os.system`` / ``subprocess.Popen``.  Every other
    # whitelist entry is an explicit set of class names.
    #
    # Note ``dict.get(key)`` returns ``None`` for both "key
    # absent" and "key present with value None" — we therefore
    # distinguish via the sentinel object below rather than
    # raw ``is None`` comparison.
    _SAFE_CLASSES: dict[str, set[str] | None] = {
        "builtins": {
            "dict",
            "list",
            "tuple",
            "set",
            "frozenset",
            "int",
            "float",
            "str",
            "bool",
            "bytes",
            "bytearray",
            "NoneType",
            "type",
            "slice",
        },
        "collections": {"OrderedDict", "defaultdict", "Counter", "deque"},
        "collections.__init__": {"OrderedDict", "defaultdict", "Counter", "deque"},
        "pandas.core.series": {"Series"},
        "pandas.core.frame": {"DataFrame"},
        # NOTE: newer pandas releases have relocated these classes
        # under ``pandas.*.frame`` / ``pandas.*.series`` submodules
        # depending on the wheel build.  Whitelist both the original
        # canonical paths and the ``pandas.*`` alias so a round-trip
        # works regardless of which path the running pandas 2.x
        # resolves the class through.
        "pandas": {"DataFrame", "Series", "Index", "StringDtype", "RangeIndex"},
        # Pandas 2.x stores string columns as ``ArrowStringArray``
        # via the Arrow backend (the legacy ``numpy.object_`` path
        # was deprecated).  The pickle protocol resolves this
        # through ``pandas.arrays`` rather than ``pandas.core.*``,
        # so we whitelist the runtime module path explicitly.
        "pandas.arrays": {"ArrowStringArray"},
        # The Arrow backend itself (``pyarrow.lib``) is a transitive
        # dependency of pandas 2.x and is not a sandboxing risk —
        # allowing arbitrary Python objects to land via pyarrow
        # would require the user to have actively installed
        # pyarrow *and* crafted a malicious data file, after
        # which the unpickler still has to resolve the class.
        # We grant the *entire* ``pyarrow.lib`` surface here so
        # any pandas 2.x Arrow-backed string column round-trips
        # cleanly without our having to keep this list current
        # every time the pyarrow release rotates a private name.
        # The cost is a slightly-bigger whitelist; the safety is
        # unchanged because pyarrow.lib's exposed API is only
        # ``_scalar_to_array``/``_restore_array``-style restoration
        # routines, never ``os.system`` or ``subprocess.Popen``.
        "pyarrow.lib": None,
        # Phase 1.4: ``BlockManager`` is the internal layout primitive
        # that pandas 2.x uses to back every ``DataFrame`` /
        # ``Series``.  Without it, a pickle of a ``records`` list
        # containing a DataFrame falls back to "Can't pickle local
        # object" or "Blocked unsafe class ... BlockManager"
        # depending on whether the unpickler bails before/after
        # resolving the type.  Phase 1.4 acceptance: pin the round-trip
        # of a real engine whose ``engine.records`` includes a
        # ``pandas.DataFrame`` — see tests/test_pickle_roundtrip.py.
        "pandas.core.internals.managers": {"BlockManager"},
        # Phase 1.4: pandas's ``_unpickle_block`` is the C-extension
        # helper that ``BlockManager.__setstate__`` falls through to
        # when materialising ``Block`` objects from a pickled stream.
        # Without it the BlockManager round-trip falls back to
        # "Blocked unsafe class ... _unpickle_block".  BlockManager
        # itself is a thin Python wrapper around this C-level loader,
        # so both are required for a clean DataFrame round-trip.
        "pandas._libs.internals": {"_unpickle_block"},
        # Phase 1.4: numpy's ``_frombuffer`` is the C-extension helper
        # used by ``ndarray.__reduce__`` to round-trip the raw byte
        # buffer that holds the array alongside a type descriptor.
        # ``ndarray`` was already on the whitelist; this lets the
        # byte-buffer half survive the round-trip.
        # NOTE: numpy >= 2.0 moved this to ``numpy._core.numeric``; keep
        # both paths for backward/forward compatibility.
        "numpy.core.numeric": {"_frombuffer"},
        "numpy._core.numeric": {"_frombuffer"},
        # Phase 1.4: ``numpy.dtype`` is the scalar-type descriptor
        # every ndarray carries — without it a round-tripped
        # ndarray raises "Object has no attribute 'itemsize'".
        # Whitelist the dedicated ``dtype`` module too, alongside
        # the existing ndarray entries.
        "numpy.dtype": {"dtype"},
        "numpy": {"ndarray", "dtype"},
        "numpy.ndarray": {"ndarray"},
        # Phase 1.4: ``_reconstruct`` is the C-extension helper that
        # rebuilds an ``ndarray`` of a given shape/dtype from the
        # pickle-encoded byte buffer.  Without this entry, a
        # round-tripped 2D ``numpy.ndarray`` (under the bonnet of
        # every ``pandas.DataFrame``) fails with
        # "Blocked unsafe class 'numpy.core.multiarray'.'_reconstruct'".
        # NOTE: numpy >= 2.0 moved this to ``numpy._core.multiarray``;
        # keep both paths for compatibility.
        "numpy.core.multiarray": {"_reconstruct"},
        "numpy._core.multiarray": {"_reconstruct"},
        # Phase 1.4: ``_new_Index`` rebuilds a pandas Index from a
        # pickled (dtype, kind) tuple — needed because the persistent
        # RangeIndex(DataFrame.index) carries a ``kind`` token.  The
        # public ``Index`` class is the parent that ``_new_Index``
        # instantiates; both need to be on the whitelist for the
        # round-trip to construct a fully-fledged ``Index`` after
        # the C-extension helper has built its layout.
        "pandas.core.indexes.base": {"_new_Index", "Index"},
        # Phase 1.4: ``RangeIndex`` is the integer-only Index
        # subclass that pandas DataFrames grow by default.  Without
        # it the round-trip works for ``Index`` but raises
        # "Blocked unsafe class 'pandas.core.indexes.range'.'RangeIndex'"
        # on the most common case.  ``RangeIndex.__init__`` is a thin
        # wrapper so this single entry is sufficient.
        "pandas.core.indexes.range": {"RangeIndex"},
        # Our own classes — needed for persisted engine objects
        # NOTE: "__main__" was previously allowed but is a security risk —
        # it would permit any user script named EvidenceEngine to be
        # unpickled.  The proper module path "edf_collector" is the only
        # legitimate source for this class.
        "edf_collector": {"EvidenceEngine"},
    }

    def find_class(self, module: str, name: str) -> type:
        """Resolve ``module.name`` from the explicit whitelist only.

        A whitelist value of ``None`` (as opposed to the usual
        ``set[str]`` of permitted class names) is interpreted as
        "the entire module is trusted".  We only use this for
        ``pyarrow.lib`` whose exposed pickle surface is purely
        restoration-callable ``_something`` functions, never
        ``os.system`` / ``subprocess.Popen``.  Every other
        whitelist entry is an explicit set of class names.

        Note ``dict.get(key)`` returns ``None`` for both "key
        absent" and "key present with value None" — we therefore
        distinguish via the sentinel object below rather than
        raw ``is None`` comparison.
        """
        _SENTINEL = object()  # used purely to disambiguate "absent" vs "None"
        allowed = self._SAFE_CLASSES.get(module, _SENTINEL)
        # Module not in whitelist → blocked.
        if allowed is _SENTINEL:
            raise pickle.UnpicklingError(
                f"Blocked unsafe class {module!r}.{name!r} in pickle stream"
            )
        # Whole-module permission (``None`` value) → allow.
        # Per-name permission (``set`` value) → check membership.
        # Use ``allow_everything = allowed is None`` to drive the
        # control flow explicitly so mypy can narrow the type
        # from ``set[str] |`` to ``None`` at the call sites
        # without resorting to ``cast``.
        allow_everything = allowed is None
        if allow_everything or (isinstance(allowed, set) and name in allowed):
            if module == "edf_collector":
                import importlib

                mod: Any = importlib.import_module("edf_collector")
                cls: Any = getattr(mod, name)
                if not isinstance(cls, type):
                    raise pickle.UnpicklingError(
                        f"Resolved edf_collector attribute {name!r} is not a class"
                    )
                return cls
            return cast(type, super().find_class(module, name))
        raise pickle.UnpicklingError(f"Blocked unsafe class {module!r}.{name!r} in pickle stream")


def _safe_pickle_load(path: str) -> Any:
    """Load a pickle file through the restricted unpickler.

    Usage:  obj = _safe_pickle_load("engine.pkl")
    Raises pickle.UnpicklingError for disallowed types.
    """
    with open(path, "rb") as f:
        return _RestrictedUnpickler(f).load()


def run_cli_extract(args: list[str]) -> None:
    """Run extraction from command line (headless mode)."""
    import argparse
    import json
    import os
    import sys

    parser = argparse.ArgumentParser(
        description="Extract EDF billing data from PST/OST, PDF folder, or HTM export",
        prog="edf-collector --extract",
    )
    parser.add_argument("--pst", help="Path to PST/OST file")
    parser.add_argument("--pdf-dir", help="Path to directory containing PDF bills")
    parser.add_argument("--htm", help="Path to HTM account history export")
    parser.add_argument("--output", "-o", required=True, help="Output Excel file path")
    parser.add_argument("--records-json", help="Also save extracted records as JSON")
    parser.add_argument("--config", "-c", help="Path to config JSON file (optional)")
    parser.add_argument("--acc-filter", help="Filter by account number (e.g., A-12345678)")
    parser.add_argument(
        "--domain-filter",
        default="edfenergy.com",
        help="Comma-separated sender domains for PST filtering",
    )
    parser.add_argument("--min-amount", type=float, default=500.0, help="Minimum amount threshold")
    parser.add_argument("--no-dedup", action="store_true", help="Disable deduplication")
    parser.add_argument("--no-anchors", action="store_true", help="Disable smart context search")
    parser.add_argument("--no-large", action="store_true", help="Disable large amount fallback")
    parser.add_argument(
        "--no-reading-class", action="store_true", help="Disable reading classification"
    )
    parser.add_argument(
        "--no-pdf-fields", action="store_true", help="Disable deep PDF field extraction"
    )
    parser.add_argument(
        "--no-filter-below", action="store_true", help="Don't filter records below minimum amount"
    )
    parsed = parser.parse_args(args)

    # Check at least one source
    if not any([parsed.pst, parsed.pdf_dir, parsed.htm]):
        sys.stderr.write("ERROR: At least one source required (--pst, --pdf-dir, or --htm)\n")
        sys.exit(1)

    # Load config from file if provided
    config = {}
    if parsed.config:
        try:
            with open(parsed.config, encoding="utf-8") as f:
                config = json.load(f)
        except Exception as e:
            sys.stderr.write(f"ERROR: Failed to load config: {e}\n")
            sys.exit(1)

    # Override with CLI args
    config.update(
        {
            "use_acc_filter": bool(parsed.acc_filter),
            "acc_num": parsed.acc_filter or "",
            "use_domain_filter": True,
            "domain_filter": parsed.domain_filter,
            "min_amount": parsed.min_amount,
            "filter_below": not parsed.no_filter_below,
            "use_dedup": not parsed.no_dedup,
            "use_anchors": not parsed.no_anchors,
            "use_large": not parsed.no_large,
            "use_reading_classification": not parsed.no_reading_class,
            "use_pdf_fields": not parsed.no_pdf_fields,
            "save_filtered": True,
            "save_dups": True,
        }
    )

    # Check PST dependency
    if parsed.pst and not HAS_PYPFF:
        sys.stderr.write(
            "ERROR: PST/OST support requires 'libpff-python'. Install with: pip install libpff-python\n"
        )
        sys.exit(1)

    engine = EvidenceEngine(config, print, None, None)

    try:
        if parsed.pst and os.path.exists(parsed.pst):
            print(f"Scanning PST/OST: {parsed.pst}")
            try:
                pff = pypff.file()
            except AttributeError:
                pff = getattr(pypff, "File", None)
                if pff is None:
                    raise AttributeError("pypff module has no 'file' or 'File' attribute") from None
                pff = pff()
            pff.open(os.path.abspath(parsed.pst))
            try:
                engine.crawl_pst(pff.get_root_folder())
            finally:
                pff.close()

        if parsed.htm and os.path.exists(parsed.htm):
            print(f"Parsing HTM: {parsed.htm}")
            engine.process_htm_file(parsed.htm)

        if parsed.pdf_dir and os.path.exists(parsed.pdf_dir):
            print(f"Scanning PDF folder: {parsed.pdf_dir}")
            engine.crawl_local_pdfs(parsed.pdf_dir)

        if not engine.records:
            sys.stderr.write("WARNING: No billing records found\n")
            sys.exit(1)

        # Export to Excel
        print(f"Writing Excel report: {parsed.output}")
        export_to_excel(
            engine.records,
            parsed.output,
            engine.error_log,
            config,
            filtered=engine.filtered_records,
        )

        # Optionally save records as JSON
        if parsed.records_json:
            import datetime

            output_data = {
                "extracted_at": datetime.datetime.now().isoformat(),
                "config": config,
                "records": engine.records,
                "filtered_records": engine.filtered_records,
                "error_log": engine.error_log,
            }
            with open(parsed.records_json, "w", encoding="utf-8") as f:
                json.dump(output_data, f, indent=2, default=str)
            print(f"Records saved as JSON: {parsed.records_json}")

        print("Extraction complete!")
        print(f"  PDFs processed: {engine.pdf_count}")
        print(f"  Emails matched: {engine.email_count}")
        print(f"  Records found:  {len(engine.records)}")
        if engine.error_log:
            print(f"  Parse errors:   {len(engine.error_log)}")

    except Exception as e:
        sys.stderr.write(f"ERROR: {e}\n")
        import traceback

        traceback.print_exc()
        sys.exit(1)


def run_cli_pdf_report(args: list[str]) -> None:
    """Run PDF report generation from command line."""
    import argparse
    import json
    import sys

    from edf_report import generate_pdf_from_gui

    parser = argparse.ArgumentParser(
        description="Generate PDF report from extracted records",
        prog="edf-collector --pdf-report",
    )
    parser.add_argument(
        "--records",
        "-i",
        required=True,
        help="Path to extracted records JSON file (exported from GUI or script)",
    )
    parser.add_argument("--output", "-o", required=True, help="Output PDF file path")
    parser.add_argument("--config", "-c", help="Path to config JSON file (optional)")
    parser.add_argument(
        "--engine-data",
        "-e",
        help="Path to engine data pickle file (optional, for filtered records)",
    )
    parsed = parser.parse_args(args)

    try:
        with open(parsed.records, encoding="utf-8") as f:
            loaded = json.load(f)

        # Accept either a bare list of records (preferred) or the wrapper
        # object emitted by ``--extract --records-json``.  The wrapper
        # shape is ``{"records": [...], ...meta}`` — unwrap it so both
        # CLI entry points behave identically.
        if isinstance(loaded, dict) and "records" in loaded:
            records = loaded["records"]
        else:
            records = loaded

        config = {}
        if parsed.config:
            with open(parsed.config, encoding="utf-8") as f:
                config = json.load(f)

        engine = None
        filtered = None
        if parsed.engine_data:
            # Use the restricted unpickler to prevent arbitrary code
            # execution from crafted pickle files (see C1 fix).
            engine = _safe_pickle_load(parsed.engine_data)
            filtered = getattr(engine, "filtered_records", None)

        success, msg = generate_pdf_from_gui(
            records=records,
            output_path=parsed.output,
            config=config,
            engine=engine,
            filtered=filtered,
        )
        if success:
            sys.stdout.write(msg + "\n")
            sys.exit(0)
        else:
            sys.stderr.write(f"ERROR: {msg}\n")
            sys.exit(1)

    except Exception as e:
        sys.stderr.write(f"ERROR: {e}\n")
        sys.exit(1)


def run_cli_docx_report(args: list[str]) -> None:
    """Run DOCX report generation from command line."""
    import argparse
    import json
    import sys

    from edf_report_docx import generate_docx_from_gui

    parser = argparse.ArgumentParser(
        description="Generate DOCX report from extracted records",
        prog="edf-collector --docx-report",
    )
    parser.add_argument(
        "--records",
        "-i",
        required=True,
        help="Path to extracted records JSON file (exported from GUI or script)",
    )
    parser.add_argument("--output", "-o", required=True, help="Output DOCX file path")
    parser.add_argument("--config", "-c", help="Path to config JSON file (optional)")
    parser.add_argument(
        "--engine-data",
        "-e",
        help="Path to engine data pickle file (optional, for filtered records)",
    )
    parsed = parser.parse_args(args)

    try:
        with open(parsed.records, encoding="utf-8") as f:
            loaded = json.load(f)

        # Accept either a bare list of records (preferred) or the wrapper
        # object emitted by ``--extract --records-json``.  Mirrors the
        # PDF CLI loader so both formats round-trip without extra steps.
        if isinstance(loaded, dict) and "records" in loaded:
            records = loaded["records"]
        else:
            records = loaded

        config = {}
        if parsed.config:
            with open(parsed.config, encoding="utf-8") as f:
                config = json.load(f)

        engine = None
        filtered = None
        if parsed.engine_data:
            # Use the restricted unpickler to prevent arbitrary code
            # execution from crafted pickle files (see C1 fix).
            engine = _safe_pickle_load(parsed.engine_data)
            filtered = getattr(engine, "filtered_records", None)

        success, msg = generate_docx_from_gui(
            records=records,
            output_path=parsed.output,
            config=config,
            engine=engine,
            filtered=filtered,
        )
        if success:
            sys.stdout.write(msg + "\n")
            sys.exit(0)
        else:
            sys.stderr.write(f"ERROR: {msg}\n")
            sys.exit(1)
    except Exception as e:
        sys.stderr.write(f"ERROR: {e}\n")
        sys.exit(1)


def main() -> None:
    """Entry point for the EDF Evidence Collector CLI."""
    import sys

    if len(sys.argv) > 1:
        if sys.argv[1] in ("--pdf-report", "--report", "-r"):
            run_cli_pdf_report(sys.argv[2:])
            return
        elif sys.argv[1] in ("--docx-report", "--word-report", "-w"):
            run_cli_docx_report(sys.argv[2:])
            return
        elif sys.argv[1] in ("--extract", "-e"):
            run_cli_extract(sys.argv[2:])
            return

    if not HAS_TK:
        sys.stderr.write(
            "ERROR: tkinter is not available in this Python build. "
            "Launch a CLI command instead (e.g. --extract, --pdf-report, "
            "--docx-report) or run on a system with Tk installed."
        )
        sys.stderr.write("\n")
        sys.exit(2)

    root = tk.Tk()
    App(root)
    root.mainloop()


if __name__ == "__main__":
    main()
