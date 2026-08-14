"""Superseded Reconciliation sheet writer (2026-08-14 design)."""

from __future__ import annotations

from datetime import datetime

import openpyxl
import pandas as pd
from openpyxl.worksheet.worksheet import Worksheet

from edf_bill_fetcher.helpers.excel_utils import (
    Alignment,
    Font,
    PatternFill,
    set_column_widths_from_spec,
)
from edf_bill_fetcher.helpers.excel_utils import (
    money as _money,
)
from edf_bill_fetcher.helpers.excel_utils import (
    num as _num,
)
from edf_bill_fetcher.helpers.excel_utils import (
    pdf_hyperlink_cell as _pdf_hyperlink_cell,
)
from edf_bill_fetcher.helpers.excel_utils import (
    text as _text,
)
from edf_bill_fetcher.io.writers.sheet_layout import (
    freeze_at,
    write_banner,
    write_header_row,
    write_merged_text,
    write_section_label,
    write_trailing_total,
)
from edf_bill_fetcher.writers._helpers import _disclosed_label

_COLS = [
    "Invoice #",
    "Bill Date",
    "Period From",
    "Period To",
    "Days Billed",
    "Period Charge (£)",
    "Unlawful Charge (£)",
    "Excess Days",
    "Cancel/Rebill Disclosed",
    "Reason Assessment",
    "Killer on spreadsheet",
    "Original invoice on spreadsheet",
    "Original invoice PDF",
    "Killer invoice PDF",
    "Partial Overlap",
]


def write_superseded_reconciliation_sheet(
    ws: Worksheet,
    bb: pd.DataFrame,
    domination_map: dict[str, tuple[str, bool]],
    evidence_index: dict[str, int] | None = None,
    invoice_pdf_paths: dict[str, str] | None = None,
    live_row_map: dict[str, int] | None = None,
) -> dict[str, int]:
    """Render the Superseded Reconciliation worksheet.

    One row per superseded back-bill invoice (a key of ``domination_map``),
    grouped under a ``KILLER: <survivor>`` label row so a reader jumping from
    a live row's "View superseded" link lands on the right chain. Each row
    carries the superseded invoice's own data plus four navigation links: the
    killer row on Back-billing Analysis, the original invoice's row on the EDF
    Evidence Report, and ``file://`` links to both saved PDFs. The trailing
    total is the sum of the superseded rows' unlawful charges, labelled as
    absorbed/audit-only and never added to the live union total.

    Returns a ``{survivor_invoice: reconciliation_row}`` map recording the
    Excel row of each ``KILLER:`` header this writer actually emitted.  The
    caller (export) feeds that back to :func:`write_back_billing_sheet` as
    ``view_superseded_row`` so each live survivor's "View superseded" link
    lands on its own group header.  These header rows are NOT the survivor's
    Back-billing Analysis row: the reconciliation sheet intersperses a
    ``KILLER:`` header per group, so row numbers diverge once there is more
    than one group.
    """
    ws.title = "Superseded Reconciliation"
    NAVY = "10367A"
    ORANGE = "FE5716"
    write_banner(ws, "SUPERSEDED RECONCILIATION", 17, color=ORANGE, row=1, height=22)
    write_section_label(ws, 2, "LEGAL CONTEXT", 17)
    write_merged_text(
        ws,
        3,
        "Superseded back-billing invoices are earlier invoices that a later "
        "cancel-and-rebill invoice (the killer / survivor) has absorbed: the "
        "same consumption is re-covered by the surviving invoice, so these rows "
        "are excluded from the Back-billing Analysis total to avoid "
        "double-counting. This sheet records each superseded invoice for audit, "
        "with the survivor's row on the Back-billing Analysis sheet, the "
        "original invoice's row on the EDF Evidence Report, and links to the "
        "saved PDFs of both the original and the killer invoice.",
        17,
        height=60,
    )
    write_header_row(ws, 7, _COLS, bg=NAVY, height=28)
    r = 8
    superseded_total = 0.0
    alt_fill = PatternFill("solid", start_color="EEF2FF")
    survivor_row_map: dict[str, int] = {}
    # group by survivor preserving bb order
    survivors = sorted({s for s, _ in domination_map.values()})
    for survivor in survivors:
        label = ws.cell(row=r, column=1, value=f"KILLER: {survivor}")
        label.font = Font(bold=True, color=NAVY)
        label.alignment = Alignment(horizontal="left", vertical="center")
        survivor_row_map[survivor] = r
        r += 1
        group = bb[
            bb["Invoice #"]
            .astype(str)
            .isin([k for k, (s, _) in domination_map.items() if s == survivor])
        ]
        for _, row in group.iterrows():
            inv = str(row.get("Invoice #", ""))
            partial_overlap = bool(domination_map.get(inv, (survivor, False))[1])
            row_fill = alt_fill if r % 2 == 0 else PatternFill()
            bg = None if row_fill.fill_type is None else "EEF2FF"
            charge = float(row.get("Period Charge (£)", 0.0) or 0.0)
            unlawful = float(row.get("Unlawful Charge (£)", 0.0) or 0.0)
            superseded_total += unlawful
            bill_date_val = row.get("Bill Date", "")
            if isinstance(bill_date_val, pd.Timestamp | datetime):
                bill_date_val = bill_date_val.strftime("%d %b %Y")
            pf = row.get("Period From")
            if isinstance(pf, pd.Timestamp | datetime):
                pf = pf.strftime("%d %b %Y")
            pt = row.get("Period To")
            if isinstance(pt, pd.Timestamp | datetime):
                pt = pt.strftime("%d %b %Y")
            reason_assessment = str(row.get("Reason Assessment", "") or "")
            reason_assessment = (
                f"{reason_assessment} Superseded by {survivor}, which re-billed the same "
                "period; this invoice's unlawful charge is absorbed into the survivor's "
                "total on the Back-billing Analysis sheet."
            )

            _text(ws, r, 1, inv, fill_hex=bg)
            _text(ws, r, 2, bill_date_val, fill_hex=bg)
            _text(ws, r, 3, pf, fill_hex=bg)
            _text(ws, r, 4, pt, fill_hex=bg)
            _num(ws, r, 5, int(row.get("Days Billed", 0)), fmt="#,##0", fill_hex=bg)
            _money(ws, r, 6, charge, fill_hex=bg)
            _money(ws, r, 7, unlawful, fill_hex=bg)
            _num(ws, r, 8, int(row.get("Excess Days", 0)), fmt="#,##0", fill_hex=bg)
            if int(row.get("Excess Days", 0)) > 30:
                ws.cell(row=r, column=8).font = Font(
                    name="Calibri", size=10, bold=True, color="C00000"
                )
            disclosed = _disclosed_label(bool(row.get("Cancel/Rebill Admitted")), partial_overlap)
            _text(ws, r, 9, disclosed, fill_hex=bg)
            _text(ws, r, 10, reason_assessment, wrap=True, fill_hex=bg)

            # Killer on spreadsheet (col 11): jump to the survivor's row on
            # Back-billing Analysis.
            killer_row = (live_row_map or {}).get(survivor)
            if killer_row:
                cell = ws.cell(row=r, column=11, value="→")
                cell.hyperlink = openpyxl.worksheet.hyperlink.Hyperlink(
                    ref=cell.coordinate,
                    location=f"'Back-billing Analysis'!A{killer_row}",
                    display="→",
                    tooltip=f"Jump to Back-billing Analysis!A{killer_row}",
                )
                cell.font = Font(name="Calibri", size=10, color="0563C1", underline="single")
            else:
                cell = _text(ws, r, 11, "No match", fill_hex=bg)
                cell.font = Font(name="Calibri", size=10, italic=True, color="A6A6A6")

            # Original invoice on spreadsheet (col 12): jump to this invoice's
            # row on the EDF Evidence Report.
            ev_row = (evidence_index or {}).get(f"inv:{inv}")
            if ev_row:
                cell = ws.cell(row=r, column=12, value="→")
                cell.hyperlink = openpyxl.worksheet.hyperlink.Hyperlink(
                    ref=cell.coordinate,
                    location=f"'EDF Evidence Report'!A{ev_row}",
                    display="→",
                    tooltip=f"Jump to EDF Evidence Report!A{ev_row}",
                )
                cell.font = Font(name="Calibri", size=10, color="0563C1", underline="single")
            else:
                cell = _text(ws, r, 12, "No match", fill_hex=bg)
                cell.font = Font(name="Calibri", size=10, italic=True, color="A6A6A6")

            # Original invoice PDF (col 13).
            orig_pdf = (invoice_pdf_paths or {}).get(inv, "")
            if orig_pdf:
                _pdf_hyperlink_cell(ws, r, 13, orig_pdf)
            else:
                cell = _text(ws, r, 13, "No file", fill_hex=bg)
                cell.font = Font(name="Calibri", size=10, italic=True, color="A6A6A6")

            # Killer invoice PDF (col 14).
            killer_pdf = (invoice_pdf_paths or {}).get(survivor, "")
            if killer_pdf:
                _pdf_hyperlink_cell(ws, r, 14, killer_pdf)
            else:
                cell = _text(ws, r, 14, "No file", fill_hex=bg)
                cell.font = Font(name="Calibri", size=10, italic=True, color="A6A6A6")

            _text(ws, r, 15, "Yes" if partial_overlap else "", fill_hex=bg)
            r += 1

    write_trailing_total(
        ws,
        r,
        "TOTAL SUPERSEDED UNLAWFUL CHARGES (absorbed into survivors)",
        [(7, round(superseded_total, 2))],
        5,
        17,
    )
    widths: dict[str, float] = {
        "A": 18,
        "B": 14,
        "C": 14,
        "D": 14,
        "E": 12,
        "F": 16,
        "G": 16,
        "H": 12,
        "I": 22,
        "J": 60,
        "K": 22,
        "L": 26,
        "M": 22,
        "N": 20,
        "O": 16,
    }
    set_column_widths_from_spec(ws, widths)
    freeze_at(ws, "A8")
    return survivor_row_map
