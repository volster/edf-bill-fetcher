"""Meter readings and contract history sheet writers — extracted from writers/__init__.py.

Phase 5C of the modularization refactor. Contains the meter-rollover
detector, the contract-inference detector, and the sheet writers for
the meter readings and contract history tabs.

The public API exposed here matches the test contract at
``tests/test_io_writers_extraction.py``: ``write_meter_readings_sheet``
takes ``(ws, df)`` only. The underlying implementation accepts additional
optional arguments; sensible defaults (empty rollovers DataFrame) are
supplied by the adapter wrapper below.
"""

from __future__ import annotations

import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from edf_bill_fetcher.helpers.date_utils import _safe_to_datetime
from edf_bill_fetcher.helpers.excel_utils import (
    hcell as _hcell,
)
from edf_bill_fetcher.helpers.excel_utils import (
    num as _num,
)
from edf_bill_fetcher.helpers.excel_utils import (
    open_pdf_hyperlink_cell as _open_pdf_hyperlink_cell,
)
from edf_bill_fetcher.helpers.excel_utils import (
    text as _text,
)
from edf_bill_fetcher.helpers.theme import CELL_BORDER
from edf_bill_fetcher.processors.detection import detect_meter_rollover
from edf_bill_fetcher.processors.matching import infer_contracts
from edf_bill_fetcher.writers._helpers import _reading_type_to_aem

# ---- detect_meter_rollover, infer_contracts (re-exported from processors) ----


# ---- write_meter_readings_sheet (was L2607-2782) ----


def _write_meter_readings_sheet_impl(
    ws: Worksheet,
    df: pd.DataFrame,
    rollovers: pd.DataFrame,
    account: str = "",
    *,
    evidence_df: pd.DataFrame | None = None,
    evidence_index: dict[str, int] | None = None,
) -> None:
    """Render the Meter Readings tab (spec §4.3).

    Layout:
      row 1: title banner with account
      row 2: legend 'A = Actual, E = Estimated, M = Meter rollover'
      row 7: table header (8 cols)
      rows 8+: one row per evidence record, ordered by Date

    The 'Type (A/E/M)' column maps each evidence row's Reading
    column to A / E, with M overriding when this invoice appears in
    the ``rollovers`` table. The Estimated Source column carries
    Details verbatim (e.g. 'Automatic estimate' or 'SAP estimate')
    for Estimated rows, else blank.

    Open PDF column (col 7): hyperlink
    source-PDF-text + regex trace for that invoice, fetched from
    ``evidence_df`` via :func:`_open_pdf_hyperlink_cell` for hyperlink
    available.

    Spec §10.2 adds a "View on Evidence Report" column (col 8): a
    hyperlinked right-arrow that jumps to the matched row on the
    EDF Evidence Report sheet, looked up from ``evidence_index``.
    """
    ws.title = "Meter Readings"
    NAVY = "10367A"
    ORANGE = "FE5716"

    rollover_invoices: set[str] = set()
    if rollovers is not None and not rollovers.empty and "Invoice #" in rollovers.columns:
        rollover_invoices = {str(x) for x in rollovers["Invoice #"].tolist() if x}

    # Row 1: title banner with account
    title = "METER READING HISTORY \u2014 Actual vs Estimated"
    if account:
        title = f"{title}  |  Account {account}"
    t1 = ws.cell(row=1, column=1, value=title)
    t1.font = Font(name="Calibri", size=13, bold=True, color="FFFFFF")
    t1.fill = PatternFill("solid", start_color=ORANGE)
    t1.border = CELL_BORDER
    t1.alignment = Alignment(horizontal="left", vertical="center")
    for c in range(2, 7):
        x = ws.cell(row=1, column=c)
        x.fill = PatternFill("solid", start_color=ORANGE)
        x.border = CELL_BORDER
    ws.row_dimensions[1].height = 22

    # Row 2: legend subheader
    sub = (
        "A = Actual (supplier-confirmed reading)  |  E = Estimated  |  "
        "M = Meter rollover candidate (negative delta near rollover threshold)"
    )
    sub_cell = ws.cell(row=2, column=1, value=sub)
    sub_cell.font = Font(name="Calibri", size=10, italic=True)
    sub_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=6)
    ws.row_dimensions[2].height = 28

    # Row 7: table header (8 cols per spec \u00a74.3 + \u00a75.2 + \u00a710.2).
    headers = [
        "Date",
        "Reading (kWh)",
        "Type (A/E/M)",
        "Estimated Source",
        "Invoice #",
        "Notes",
        "Open PDF",
        "View on Evidence Report",
    ]
    for col, h in enumerate(headers, 1):
        _hcell(ws, 7, col, h, bg=NAVY)
    ws.row_dimensions[7].height = 28

    # Sort rows by Date before writing.
    work = df.copy() if df is not None and not df.empty else pd.DataFrame()
    if not work.empty:
        work["_dt"] = _safe_to_datetime(work.get("Date"))
        work = work.sort_values("_dt").drop(columns=["_dt"])

    r = 8
    for _, row in work.iterrows():
        bg = "EEF2FF" if r % 2 == 0 else None
        inv = str(row.get("Invoice #", ""))
        reading = row.get("Reading", "")
        units_raw = row.get("Units (kWh)", "N/A")
        try:
            units = float(units_raw)
        except (TypeError, ValueError):
            units = units_raw  # keep as-is in cell
        # Type code: M overrides if invoice flagged in rollovers.
        type_code = "M" if inv in rollover_invoices else _reading_type_to_aem(str(reading))
        est_src = ""
        if str(reading) == "Estimated":
            est_src = str(row.get("Details", "") or "")
        notes = (
            "Meter rollover candidate -- see rollover table." if inv in rollover_invoices else ""
        )
        _text(ws, r, 1, row.get("Date", ""), fill_hex=bg)
        if isinstance(units, int | float):
            _num(ws, r, 2, units, fmt="#,##0.0", fill_hex=bg)
        else:
            _text(ws, r, 2, str(units), fill_hex=bg)
        _text(ws, r, 3, type_code, fill_hex=bg)
        _text(ws, r, 4, est_src, fill_hex=bg)
        _text(ws, r, 5, inv, fill_hex=bg)
        _text(ws, r, 6, notes, wrap=True, fill_hex=bg)
        # Colour the type cell for clarity: amber for E, blue for M.
        type_cell = ws.cell(row=r, column=3)
        if type_code == "M":
            type_cell.font = Font(name="Calibri", size=10, bold=True, color="003F87")
        elif type_code == "E":
            type_cell.font = Font(name="Calibri", size=10, color="C08000")
        excerpt = ""
        if evidence_df is not None and not evidence_df.empty and "Invoice #" in evidence_df.columns:
            matches = evidence_df[evidence_df["Invoice #"].astype(str) == str(inv)]
            if not matches.empty:
                source_text = matches.iloc[0].get("Source PDF Text", "")
                if isinstance(source_text, str) and source_text:
                    excerpt = source_text[:400]
                    if len(source_text) > 400:
                        excerpt += " ..."
        if excerpt:
            _text(ws, r, 7, excerpt, wrap=True, fill_hex=bg)
        else:
            _open_pdf_hyperlink_cell(ws, r, 7, evidence_df, inv)
        # View on Evidence Report (col 8): hyperlinked right-arrow
        # that jumps to the matched invoice's row on the Evidence
        # Report sheet.
        target_row = None
        if evidence_index is not None:
            target_row = evidence_index.get(f"inv:{inv}")
            if target_row is None:
                # Fall back to date+units signature.
                try:
                    amt = float(units)
                    units_sig = int(round(amt))
                    key = f"date_units:{row.get('Date', '')}|{units_sig}"
                    target_row = evidence_index.get(key)
                except (TypeError, ValueError):
                    pass
        if target_row is not None:
            cell = ws.cell(row=r, column=8, value="\u2192")
            cell.hyperlink = openpyxl.worksheet.hyperlink.Hyperlink(
                ref=cell.coordinate,
                location=f"'EDF Evidence Report'!A{target_row}",
                display="\u2192",
                tooltip=f"Jump to EDF Evidence Report!A{target_row}",
            )
            cell.font = Font(name="Calibri", size=10, color="0563C1", underline="single")
        else:
            cell = ws.cell(row=r, column=8, value="No match")
            cell.font = Font(name="Calibri", size=10, italic=True, color="A6A6A6")
        r += 1

    # Column widths tailored for the table cells.
    widths = {
        "A": 14,
        "B": 16,
        "C": 16,
        "D": 26,
        "E": 20,
        "F": 50,
        "G": 60,  # Open PDF
        "H": 22,  # View on Evidence Report
    }
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width
    ws.freeze_panes = "A8"


# ---- write_contract_history_sheet (was L2785-2915) ----


def write_contract_history_sheet(
    ws: Worksheet,
    contracts: pd.DataFrame,
    account: str = "",
    *,
    evidence_df: pd.DataFrame | None = None,
    evidence_index: dict[str, int] | None = None,
) -> None:
    """Render the Contract History tab (spec \u00a74.4).

    Spec \u00a75.2 adds a "Source Excerpt" column populated from any
    invoice in ``evidence_df`` whose Period falls inside each
    inferred contract's [Contract From, Contract To] window.
    Spec \u00a710.2 adds a "View on Evidence Report" hotlink to
    that matching invoice's row.
    """
    ws.title = "Contract History"
    NAVY = "10367A"
    ORANGE = "FE5716"

    def _first_matching_invoice(cf: pd.Timestamp, ct: pd.Timestamp) -> str:
        if evidence_df is None or evidence_df.empty or "Invoice #" not in evidence_df.columns:
            return ""
        for _, er in evidence_df.iterrows():
            ipf = _safe_to_datetime(er.get("Period From"))
            ipt = _safe_to_datetime(er.get("Period To"))
            if pd.isna(ipf) or pd.isna(ipt):
                continue
            if (ipf <= ct) and (ipt >= cf):
                inv = str(er.get("Invoice #", ""))
                if inv:
                    return inv
        return ""

    # Row 1: title banner with account
    title = "INFERRED CONTRACT HISTORY"
    if account:
        title = f"{title}  |  Account {account}"
    t1 = ws.cell(row=1, column=1, value=title)
    t1.font = Font(name="Calibri", size=13, bold=True, color="FFFFFF")
    t1.fill = PatternFill("solid", start_color=ORANGE)
    t1.border = CELL_BORDER
    for c in range(2, 8):
        x = ws.cell(row=1, column=c)
        x.fill = PatternFill("solid", start_color=ORANGE)
        x.border = CELL_BORDER
    ws.row_dimensions[1].height = 22

    # Row 2: subheader
    sub = (
        "Contract periods inferred from tariff transitions in the parsed "
        "invoice stream. Boundaries are approximate (\u2264 30-day merges)."
    )
    sub_cell = ws.cell(row=2, column=1, value=sub)
    sub_cell.font = Font(name="Calibri", size=10, italic=True)
    sub_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=7)
    ws.row_dimensions[2].height = 30

    # Row 7: table headers (5 data + Open PDF + View on ER = 7
    # cols per spec \u00a74.4 + \u00a75.2 + \u00a710.2).
    headers = [
        "Contract From",
        "Contract To",
        "Tariff",
        "Days",
        "# Invoices",
        "Open PDF",
        "View on Evidence Report",
    ]
    for col, h in enumerate(headers, 1):
        _hcell(ws, 7, col, h, bg=NAVY)
    ws.row_dimensions[7].height = 28

    r = 8
    for _, row in contracts.iterrows() if contracts is not None and not contracts.empty else []:
        bg = "EEF2FF" if r % 2 == 0 else None
        cf = _safe_to_datetime(row.get("Contract From"))
        ct = _safe_to_datetime(row.get("Contract To"))
        cf_text = row.get("Contract From", "")
        if isinstance(cf, pd.Timestamp) and not pd.isna(cf):
            cf_text = cf.strftime("%d %b %Y")
        ct_text = row.get("Contract To", "")
        if isinstance(ct, pd.Timestamp) and not pd.isna(ct):
            ct_text = ct.strftime("%d %b %Y")
        _text(ws, r, 1, cf_text, fill_hex=bg)
        _text(ws, r, 2, ct_text, fill_hex=bg)
        _text(ws, r, 3, row.get("Tariff", ""), fill_hex=bg)
        _num(ws, r, 4, int(row.get("Days", 0)), fmt="#,##0", fill_hex=bg)
        _num(ws, r, 5, int(row.get("# Invoices", 0)), fmt="#,##0", fill_hex=bg)
        matched_inv = (
            _first_matching_invoice(cf, ct)
            if (
                isinstance(cf, pd.Timestamp)
                and not pd.isna(cf)
                and isinstance(ct, pd.Timestamp)
                and not pd.isna(ct)
            )
            else ""
        )
        _open_pdf_hyperlink_cell(ws, r, 6, evidence_df, matched_inv)
        target_row = None
        if evidence_index is not None and matched_inv:
            target_row = evidence_index.get(f"inv:{matched_inv}")
        if target_row is not None:
            cell = ws.cell(row=r, column=7, value="\u2192")
            cell.hyperlink = openpyxl.worksheet.hyperlink.Hyperlink(
                ref=cell.coordinate,
                location=f"'EDF Evidence Report'!A{target_row}",
                display="\u2192",
                tooltip=f"Jump to EDF Evidence Report!A{target_row}",
            )
            cell.font = Font(name="Calibri", size=10, color="0563C1", underline="single")
        else:
            cell = ws.cell(row=r, column=7, value="No match")
            cell.font = Font(name="Calibri", size=10, italic=True, color="A6A6A6")
        r += 1

    # Column widths.
    widths = {
        "A": 16,
        "B": 16,
        "C": 24,
        "D": 10,
        "E": 12,
        "F": 60,  # Open PDF
        "G": 22,  # View on Evidence Report
    }
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width
    ws.freeze_panes = "A8"


def write_meter_readings_sheet(
    ws: Worksheet,
    df: pd.DataFrame,
    rollovers: pd.DataFrame | None = None,
    account: str = "",
    *,
    evidence_df: pd.DataFrame | None = None,
    evidence_index: dict[str, int] | None = None,
) -> None:
    """Adapter: test contract uses ``(ws, df)``; supply defaults for the rest."""
    return _write_meter_readings_sheet_impl(
        ws,
        df,
        rollovers if rollovers is not None else pd.DataFrame(),
        account,
        evidence_df=evidence_df,
        evidence_index=evidence_index,
    )


__all__ = [
    "_write_meter_readings_sheet_impl",
    "detect_meter_rollover",
    "infer_contracts",
    "write_contract_history_sheet",
    "write_meter_readings_sheet",
]
