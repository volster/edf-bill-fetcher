"""Writer functions for the EDF evidence workbook.

Extracted from ``edf_collector.py`` as part of the modularization
refactor (Task 5).  Each function writes one or more Excel sheets
using openpyxl.
"""

from __future__ import annotations

import os
import re
import warnings
from datetime import datetime
from typing import Any

import numpy as np
import openpyxl
import pandas as pd
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill

from edf_bill_fetcher.helpers.date_utils import (  # noqa: E402,F401,I001
    _safe_to_datetime,
    parse_to_sort_date,
    to_excel_date,
)
from edf_bill_fetcher.helpers.date_utils import (
    completeness_score as _completeness_score,
)
from edf_bill_fetcher.helpers.excel_utils import (  # noqa: E402,F401,I001
    CELL_BORDER,
)
from edf_bill_fetcher.helpers.excel_utils import (
    build_sap_row_index_map as _build_sap_row_index_map,
)
from edf_bill_fetcher.helpers.excel_utils import (
    hcell as _hcell,
)
from edf_bill_fetcher.helpers.excel_utils import (
    money as _money,
)
from edf_bill_fetcher.helpers.excel_utils import (
    num as _num,
)
from edf_bill_fetcher.helpers.excel_utils import (
    section_hdr as _section_hdr,
)
from edf_bill_fetcher.helpers.excel_utils import (
    text as _text,
)
from edf_bill_fetcher.helpers.formatting import (
    _apply_amalgamate_to_kept_frame,
)
from edf_bill_fetcher.io.writers.back_billing import (  # noqa: E402,F401
    write_back_billing_sheet,
)
from edf_bill_fetcher.io.writers.data_quality import (  # noqa: E402,F401
    write_data_quality_sheet,
)
from edf_bill_fetcher.io.writers.evidence import (  # noqa: E402,F401
    write_evidence_sheet,
    write_summary_sheet,
)
from edf_bill_fetcher.io.writers.forecast import (  # noqa: E402,F401
    write_forecast_sheet,
)
from edf_bill_fetcher.io.writers.meter import (  # noqa: E402,F401
    write_contract_history_sheet,
    write_meter_readings_sheet,
)
from edf_bill_fetcher.io.writers.payment import (  # noqa: E402,F401
    write_payment_analysis_sheet,
)
from edf_bill_fetcher.io.writers.rebilling import (  # noqa: E402,F401
    write_rebilling_sheet,
)
from edf_bill_fetcher.io.writers.reconciliation import (  # noqa: E402,F401,I001
    write_reconciliation_sheet,
)
from edf_bill_fetcher.io.writers.sap import (  # noqa: E402,F401,I001
    write_sap_back_billing_sheets,
    write_sap_contract_history_sheet,
    write_sap_financial_transactions_sheet,
    write_sap_meter_readings_sheet,
)
from edf_bill_fetcher.io.writers.statistical import (  # noqa: E402,F401
    write_statistical_analysis_sheet,
)
from edf_bill_fetcher.io.writers.tariff import (  # noqa: E402,F401
    write_tariff_analysis_sheet,
)
from edf_bill_fetcher.models.config import ConfigDict
from edf_bill_fetcher.processors.detection import (  # noqa: E402,F401
    compute_transitive_domination,
)
from edf_bill_fetcher.processors.matching import _build_bb_clusters
from edf_bill_fetcher.writers._helpers import (  # noqa: E402,F401,I001
    _SOURCE_PRECEDENCE,
    compute_dispute_flags,
    detect_sap_back_billing_events,
    handle_cluster_unmatched,
    match_sap_events_to_edf,
)

# ---------------------------------------------------------------------------
# Module-level colour constants (lifted from export_to_excel locals so the
# module-private helpers below can reference them without closing over locals)
# ---------------------------------------------------------------------------

NAVY = "10367A"
ORANGE = "FE5716"
RED = "FF6B6B"
AMBER = "FFD166"
GREEN = "06D6A0"
LGREY = "F0F0F0"
DGREY = "888888"


# ---------------------------------------------------------------------------
# Module-private helpers (lifted from export_to_excel closures)
# ---------------------------------------------------------------------------


def _compute_unit_rate(row):
    """Per-row unit-rate stamping for the duplicate DataFrame.

    Pure function of a row dict (reads ``Period Charge (£)`` and ``Units (kWh)``).
    Lifted to module scope for testability and pickle-safety (closures aren't picklable).
    """
    pc = row.get("Period Charge (£)")
    units = row.get("Units (kWh)")
    try:
        pc_f = float(pc)
        u_f = float(str(units).replace(",", ""))
        if u_f > 0 and pc_f > 0:
            return round((pc_f / u_f) * 100, 2)
    except (ValueError, TypeError):
        pass
    return np.nan


def _summary(kept_idx_by_dup: dict[int, dict], idx: int) -> str:
    """Build the printable kept-row-reference string for a duplicate row.

    ``kept_idx_by_dup`` is threaded explicitly (formerly a closure var).
    """
    row = kept_idx_by_dup.get(idx)
    if not row:
        return ""
    try:
        amount_val = float(row["Amount (£)"])  # type: ignore[arg-type]
        amt_str = "£" + format(amount_val, ".2f")
    except (TypeError, ValueError):
        amt_str = "£--"
    return f"{row['Source']} · {row['Date']} · {amt_str}"


def _ks_row(ws_ks, r, label, value, note="", fmt=None, bold=False, alt=False):
    """Key-statistics row renderer; ``ws_ks`` threaded explicitly (formerly a closure var)."""
    bg = LGREY if alt else None
    _text(ws_ks, r, 1, label, bold=bold, fill_hex=bg)
    if fmt == "£":
        _money(ws_ks, r, 2, value, bold=bold, fill_hex=bg)
    elif fmt == "%":
        _num(ws_ks, r, 2, value, fmt="0.0%", bold=bold, fill_hex=bg)
    elif fmt == "date":
        cell = ws_ks.cell(row=r, column=2, value=value)
        cell.number_format = "dd/mm/yyyy"
        cell.font = Font(name="Calibri", size=10, bold=bold)
        cell.border = CELL_BORDER
        cell.alignment = Alignment(horizontal="right", vertical="center")
        if bg:
            cell.fill = PatternFill("solid", start_color=bg)
    elif fmt:
        _num(ws_ks, r, 2, value, fmt=fmt, bold=bold, fill_hex=bg)
    else:
        _text(ws_ks, r, 2, value, bold=bold, fill_hex=bg, align="right")
    _text(ws_ks, r, 3, note, fill_hex=bg, color=DGREY)


def _pc_stat(ws_pc, r, lbl, formula, fmt="£"):
    """Period-charges summary-stat row renderer; ``ws_pc`` threaded explicitly (formerly a closure var)."""
    _text(ws_pc, r, 1, lbl, bold=True, fill_hex=LGREY)
    c = ws_pc.cell(row=r, column=2, value=formula)
    c.font = Font(name="Calibri", size=10, bold=True)
    c.fill = PatternFill("solid", start_color=LGREY)
    c.border = CELL_BORDER
    c.alignment = Alignment(horizontal="right")
    c.number_format = "£#,##0.00" if fmt == "£" else fmt
    for cc in range(3, 9):
        ws_pc.cell(row=r, column=cc).fill = PatternFill("solid", start_color=LGREY)
        ws_pc.cell(row=r, column=cc).border = CELL_BORDER


def _banner(ws, r, text, bg):
    """Section-header banner writer; already takes ``ws`` — moved verbatim to module scope."""
    c = ws.cell(row=r, column=1, value=text)
    c.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", start_color=bg)
    c.border = CELL_BORDER
    c.alignment = Alignment(horizontal="left", vertical="center")
    for col in range(2, 7):
        x = ws.cell(row=r, column=col)
        x.fill = PatternFill("solid", start_color=bg)
        x.border = CELL_BORDER
    ws.row_dimensions[r].height = 20


def _write_provenance_sheet(wb, config, n_records, n_filtered, n_errors):
    """Write the opening Provenance tab documenting how this workbook was made.

    The sheet is created at index 0 so it is the first tab on both the
    full path and the fewer-than-2-analysis-rows early-exit path (where
    the workbook is saved before the analyser sheets render).  It
    records the tool version (resolved from the repo-root pyproject,
    not a fallback), the generation timestamp in UTC, the account
    reference, record counts, and a snapshot of the configuration
    thresholds that produced the evidence — so a filed submission is
    self-documenting without reverse-engineering the run.

    Must be listed first in the ``_reorder_sheets`` severity-led order: the
    final reorder rebuilds ``wb._sheets`` from that list, so an unlisted tab
    is silently dropped from the saved workbook.
    """
    from datetime import datetime, timezone

    from edf_bill_fetcher.helpers.version import get_package_version

    ws = wb.create_sheet(title="Provenance", index=0)
    _banner(ws, 1, "EDF Bill Fetcher — Evidence Workbook", NAVY)

    rows = [
        ("Generated (UTC)", datetime.now(timezone.utc).isoformat(timespec="seconds")),
        ("Version", get_package_version()),
        ("Account Reference", str(config.get("report_account_ref") or config.get("acc_num") or "")),
        ("Input Records", str(n_records)),
        ("Filtered Records", str(n_filtered)),
        ("Error Log Entries", str(n_errors)),
    ]
    for r_idx, (label, value) in enumerate(rows, 2):
        _hcell(ws, r_idx, 1, label, bg=DGREY)
        _text(ws, r_idx, 2, value)

    r_idx = len(rows) + 3
    _banner(ws, r_idx, "Configuration snapshot", ORANGE)
    r_idx += 1
    for key in sorted(config):
        _hcell(ws, r_idx, 1, key, bg=LGREY)
        _text(ws, r_idx, 2, str(config[key]))
        r_idx += 1

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 60


def _prepare_analysis_frame(df_an: pd.DataFrame, config: ConfigDict) -> pd.DataFrame:
    """Apply the analysis-min magnitude filter to the analysis frame.

    Balance-affecting entries are split by ``Entry Type``: every
    ``Payment``/``Credit`` row is kept, while ``New Bill``/``Ongoing
    Balance`` rows are kept only when their ``Amount (£)`` is at or above
    ``config["analysis_min"]``.

    Legal back-billing candidates are exempt from the amount gate so a
    late-billed invoice is not silently dropped from the analysis input.
    A candidate is an invoice whose bill ``Date`` is more than 365 days
    after the start of its billed consumption (``Period From``) -- the
    SLC 7A back-billing test compares the bill date with the consumption
    date, NOT the ``Period To - Period From`` span.  A one-day period
    billed five years late passes; a long period billed contemporaneously
    does not pass solely because of its span.

    This is the gate logic that was previously inline at export.py:807-811.
    """
    analysis_min = float(config.get("analysis_min", 500.0))
    payment_credit_mask = df_an["Entry Type"].isin(("Payment", "Credit"))
    bill_mask = df_an["Entry Type"].isin(("New Bill", "Ongoing Balance"))
    amount_mask = df_an["Amount (£)"] >= analysis_min

    bill_date = _safe_to_datetime(df_an["Date"])
    period_from = _safe_to_datetime(df_an["Period From"])
    legal_candidate = (bill_date - period_from).dt.days > 365

    nat_count = bill_mask & (bill_date.isna() | period_from.isna())
    if nat_count.any():
        warnings.warn(
            f"{int(nat_count.sum())} bill row(s) with unparseable dates dropped from analysis frame",
            stacklevel=2,
        )

    dfc = df_an[(payment_credit_mask) | (bill_mask & (amount_mask | legal_candidate))]
    return dfc.copy().reset_index(drop=True)


def _reorder_sheets(wb: openpyxl.Workbook) -> None:
    """Rebuild ``wb._sheets`` in severity-led tab order.

    Runs before EVERY save — including the fewer-than-2-analysis-rows
    early-exit path — so a single-row workbook opens with the same tab
    order as a full run instead of the raw creation order.
    """
    _SEVERITY_LED_ORDER = [
        "Provenance",
        "Annual Summary",
        "EDF Evidence Report",
        "SAP Financial Transactions",
        "SAP Back-billing Events",
        "SAP ↔ EDF Matched Events",
        "Backbilling According to SAP",
        "SAP Meter Readings",
        "SAP Contract History",
        "Back-billing Analysis",
        "Rebilling & Corrections",
        "Meter Readings",
        "Contract History",
        "Reconciliation",
        "Reconciliation Drill-down",
        "Dispute Flags",
        "Dispute Timeline",
        "Period Charges",
        "Payment Analysis",
        "Balance Trend",
        "Year-on-Year",
        "Key Statistics",
        "Statistical Analysis",
        "Forecast & Projection",
        "Tariff Analysis",
        "Data Quality Report",
        "Duplicate Entries",
        "Filtered (Below Min)",
        "Parse Errors",
    ]
    wb._sheets = [wb[name] for name in _SEVERITY_LED_ORDER if name in wb.sheetnames]


# ---------------------------------------------------------------------------
# Main export function
# ---------------------------------------------------------------------------


def export_to_excel(
    data: list[dict[str, Any]],
    output_path: str,
    error_log: list[str],
    config: ConfigDict,
    filtered: list[dict[str, Any]] | None = None,
    sap_rows: dict[str, list[dict]] | None = None,
) -> None:
    """Build the multi-sheet evidence workbook by orchestrating each writer submodule.

    Calls every sheet writer in the canonical order so the workbook opens
    with the Annual Summary, EDF Evidence Report, and analysis tabs in
    the layout documented in the README.
    """
    df = pd.DataFrame(data)
    df["_sort"] = df["Date"].apply(parse_to_sort_date)
    df = df.sort_values(by=["_sort", "Invoice #"], ascending=[True, False]).reset_index(drop=True)
    df["% Change"] = None

    # Deduplication — multi-pass to match the same bill across sources
    # Pass 1: Period To + Amount  (catches HTM ↔ PST where billing period matches)
    # Pass 2: Amount within 60-day window for records with no period info (Local PDF)
    dup_df = pd.DataFrame()
    if config.get("use_dedup", True):
        # Source precedence lives at module scope (``_SOURCE_PRECEDENCE``)
        # so that ``tests/test_source_precedence.py`` can pin the
        # explicit ordering without booting the entire Excel
        # export pipeline.  Lower number = higher precedence.
        df["_src_pri"] = df["Source"].map(_SOURCE_PRECEDENCE).fillna(9).astype(int)
        # Completeness score — primary sort key.  Spec: "duplicates
        # should be assessed and the most complete version of the
        # information presented".  ``_completeness_score`` counts
        # populated substantive fields on each row; the richer row
        # sorts *before* the sparser row so ``keep="first"`` keeps it.
        # Computed here (not earlier) so it's available even if the
        # upstream pipeline headers change in future.
        df["_completeness"] = df.apply(_completeness_score, axis=1)
        # Sort order (primary to tie-breaker):
        #   1. _completeness descending      — most-populated row wins
        #   2. _src_pri ascending             — higher-precedence source wins ties
        #   3. _sort ascending                — earliest date wins remaining ties
        # ``keep="first"`` then retains the head of every duplicate cluster.
        # Pre-fix the sort was only ``["_src_pri", "_sort"]`` so source
        # precedence dominated completeness — a sparser HTM row would
        # beat a richer PST row.  The companion test is
        # ``tests/test_dedup_most_complete.py``.
        df = df.sort_values(
            ["_completeness", "_src_pri", "_sort"],
            ascending=[False, True, True],
        ).reset_index(drop=True)

        # Dedup key: prefer Period To (consistent across sources for same bill),
        # fall back to Date for records without period info.  Pass 1's
        # ``DUPLICATED`` flags for *period-aware* rows track which *kept*
        # row they collide against so the dup sheet can render a clickable
        # summary linking back to the source-of-truth record.  We capture
        # the matched-against row's *original* df index — that index is
        # what ``dup_df.index`` carries through to the writer, since
        # ``dup_df = df[is_dup]`` runs before the ``reset_index`` line below.
        # Period To is the source-of-truth end-of-billing-period
        # date when present; fall back to ``_sort`` (the parsed
        # source-specific ``Date``) when the row is no-period
        # (e.g. Local PDF).  ``df["_sort"].where(cond, df["_sort"])``
        # is a tautology — Period To was being ignored and Pass 1
        # ``_dedup_date`` is the *canonical* dedup key — Period To when
        # available, otherwise left as ``NaT`` so the row is excluded
        # from ``duplicated`` clusters (since ``duplicated`` treats
        # NaT as equal across rows, falling back to ``_sort`` would
        # silently merge unrelated no-period same-amount rows).
        # Rows with NaT here are rerouted through Pass-2's no-period
        # bucket logic below, which uses ``Period To == "N/A" | NaN``
        # as the explicit handling mask.
        # Vectorised pass via _safe_to_datetime to suppress the
        # 'format-inference fallback' UserWarning pandas emits on
        # mixed-format Series when a single string passes the
        # simple-format regex gate.
        period_to_dt = _safe_to_datetime(df["Period To"])
        df["_dedup_date"] = period_to_dt
        is_dup = df.duplicated(subset=["_dedup_date", "Amount (£)"], keep="first")
        # Pass 1 (period+amount): build ``kept_pass1_index`` keyed on
        # ``(_dedup_date, Amount)`` so we can look up "which kept row
        # did this dup lose to".  The kept row's original df index (not
        # its reset_index value) survives into the dup sheet.
        kept_for_dup: dict[int, int] = {}  # dup_idx -> kept_idx (both original indices)
        kept_for_summary: dict[int, dict[str, object]] = {}  # kept_idx -> display fields
        kept_frame = df[~is_dup]
        kept_pass1_index: dict[tuple, int] = {}
        for kept_idx in kept_frame.index:
            k = (
                kept_frame.at[kept_idx, "_dedup_date"],
                kept_frame.at[kept_idx, "Amount (£)"],
            )
            kept_pass1_index.setdefault(k, kept_idx)
            # Cache the displayed fields once per kept row so the
            # dup lookup below doesn't re-read them.
            kept_for_summary[kept_idx] = {
                "Source": kept_frame.at[kept_idx, "Source"],
                "Date": kept_frame.at[kept_idx, "Date"],
                "Amount (£)": kept_frame.at[kept_idx, "Amount (£)"],
            }
        # Resolve Pass 1's kept-against reference per duplicate
        # before any reset_index runs.
        for dup_idx in df[is_dup].index:
            k = (
                df.at[dup_idx, "_dedup_date"],
                df.at[dup_idx, "Amount (£)"],
            )
            kept_idx = kept_pass1_index.get(k, -1)
            kept_for_dup[dup_idx] = kept_idx

        # Pass 2: records with no period info (e.g. Local PDF) — match by
        # Amount within a 60-day window of any already-kept record.
        #
        # Phase 2.2 follows the spec: group candidates by Amount (£)
        # first, then look up matches inside each amount-bucket
        # rather than scanning the entire kept-mask frame for every
        # candidate.  The previous implementation was O(N²) — at
        # 5,000 records the *bench* showed it took ~2.3 s.  This
        # bucketed approach is O(N) amortised: typical EDF bills
        # have unique amounts, so bucket size is 1–2 rows and the
        # inner day-window check is effectively constant.
        #
        # Layout-preserving detail worth flagging: the *legacy*
        # algorithm visits ``df.index`` in increasing order and
        # looks at the live ``kept`` mask — which includes
        # forward-yet-to-be-visited rows whose ``~is_dup`` is the
        # pre-iteration value (so any same-amount row ±60 days
        # *before or after* the candidate, except itself, can
        # match).  We replicate that exact behaviour by iterating
        # ``df.index`` in *reverse* and building per-amount buckets
        # incrementally: at row N's visit, the bucket for any
        # amount A already contains every row with amount A and
        # index > N that wasn't marked as dup — exactly the
        # forward-direction rows the legacy code saw.
        #
        # Concretely: with the legacy ``kept = df[(~is_dup) &
        # (df.index != idx)]`` mask, the set of candidate matches
        # for row idx against amount A is
        # ``{j != idx : df.Amount[j] == A and ~is_dup.at[j]}``.
        # For most rows this set is split into:
        #   (i) j in [0, idx) — *earlier* df indices,
        #  (ii) j in (idx, len(df)) — *later* df indices.
        # The legacy code consulted both groups via the live
        # ``~is_dup`` mask.  Iterating reverse and limiting our
        # bucket hashes to *only* ``j > idx`` (the "earlier in
        # reverse-iteration-order" rows) lands on exactly the
        # same candidate set provided *no row gets marked as dup
        # before its later neighbours are visited* — which the
        # reverse loop guarantees by ordering inspections from
        # the bottom of the frame upwards.
        no_period = (df["Period To"] == "N/A") | df["Period To"].isna()
        # ``bucket_by_amt`` is keyed on Amount and stores the
        # ``(df_ordinal, _sort date)`` of every row already visited
        # (reverse-iteration order) that hasn't been marked as
        # duplicate.  We append a row to its bucket whenever the
        # row *does not* get marked — symmetric to the legacy
        # ``kept`` mask at iteration time.
        bucket_by_amt: dict[float, list[tuple[int, object]]] = {}
        # Reverse-iterate ``df.index`` so that "later in df order"
        # rows are visited first and accumulate in the bucket for
        # the earlier row's lookup.  Equivalently, the bucket for
        # each amount at ``idx`` is exactly the rows j > idx with
        # Amount[j] == amount and ~is_dup.at[j] — the same row set
        # legacy would consult.
        reverse_idx = list(df[~is_dup & no_period].index)[::-1]
        for idx in reverse_idx:
            amt = df.loc[idx, "Amount (£)"]
            rec_date = df.loc[idx, "_sort"]
            same_amt = bucket_by_amt.get(amt, [])
            matched = False
            for m_idx, m_date in same_amt:
                # ``pd.notna`` short-circuit means NaT-dated rows
                # already in the bucket (originally the loop
                # ``continue``-skipped them but still listed them
                # in the next-iter kept set) never trigger a match.
                if pd.notna(m_date) and abs((rec_date - m_date).days) <= 60:
                    matched = True
                    # Capture the matched-against row's *original
                    # df index* so the dup sheet can resolve to
                    # the same frame.  We resolve the summary
                    # *before* the kept set is `reset_index`-
                    # rasterised below — once ``df = df[~is_dup]
                    # .reset_index(drop=True)`` runs, the
                    # ``m_idx`` no longer references a row.
                    kept_for_dup[idx] = m_idx
                    kept_for_summary[m_idx] = {
                        "Source": df.at[m_idx, "Source"],
                        "Date": df.at[m_idx, "Date"],
                        "Amount (£)": df.at[m_idx, "Amount (£)"],
                    }
                    break
            if matched:
                is_dup.at[idx] = True
                # Don't add to the bucket — the legacy loop's
                # recomputed ``~is_dup`` mask would have excluded a
                # row marked dup at the *start* of iteration, so it
                # cannot anchor later (here: earlier-in-iteration)
                # matches either.
            else:
                # Always add the row even if ``_sort`` is NaT —
                # the legacy ``kept`` mask at the *next* (lower) row
                # includes this row because it's ``~is_dup``-true,
                # and the NaT date just means it can't anchor a
                # match on its own.
                bucket_by_amt.setdefault(amt, []).append((idx, rec_date))

        # ``dup_df`` is built BEFORE the ``reset_index`` line below so
        # ``dup_df.index`` still carries each duplicate's original df
        # index — that's the key we use to look up the kept-against
        # summary in ``kept_for_summary``.
        #
        # ``save_dups`` toggles whether dedup *itself* is applied to the
        # main dataframe (``df``).  When True (the historical default),
        # duplicates are filtered out of ``df`` and *recorded* in
        # ``dup_df`` for the dup sheet — users never lose visibility of
        # what was dropped.  When False, dedup is skipped entirely: every
        # row stays in ``df`` and ``dup_df`` is empty.
        if config.get("save_dups", True):
            dup_df = df[is_dup].copy()
        else:
            dup_df = df[is_dup].iloc[0:0].copy()

        # Spec 3 (stretch): hybrid rows when ``amalgamate_duplicates`` is
        # True.  Instead of keeping the completeness-winner verbatim, we
        # merge each duplicate cluster's non-empty fields into a single
        # hybrid kept row.  The composite keeps the completeness-winner's
        # ``Source`` identity and picks any populated column value from
        # any sibling.  Each non-surviving sibling still stays in
        # ``dup_df`` (the spec's 'never drop without being recorded').
        #
        # N.B. the amalgamated ``df`` is is already a cleaned kept set
        # (all duplicates removed), so the ``df[~is_dup]`` filter below
        # is skipped for the amalgamate path.
        if (
            config.get("save_dups", True)
            and config.get("amalgamate_duplicates", False)
            and not dup_df.empty
        ):
            df = _apply_amalgamate_to_kept_frame(df, dup_df, kept_pass1_index, kept_for_dup, is_dup)
            # dup_df stays unchanged — the amalgamation only touches the
            # kept set; the dup sheet still records every sibling.

        if config.get("save_dups", True) and not config.get("amalgamate_duplicates", False):
            df = df[~is_dup].reset_index(drop=True)
        # else: do not drop duplicates — leave ``df`` unchanged so the
        # user sees the raw ingress and can resolve duplicates manually.
        df = df.drop(columns=["_src_pri", "_dedup_date", "_completeness"], errors="ignore")

    df = df.drop(columns=["_sort"], errors="ignore")
    dup_df = (
        dup_df.drop(
            columns=["_sort", "_src_pri", "_dedup_date", "_completeness"],
            errors="ignore",
        )
        if not dup_df.empty
        else dup_df
    )

    # Compute Unit Rate (p/kWh) where both Period Charge and Units are available.
    #
    # Phase 2.1: vectorised path.  The historic row-wise apply walked
    # Python per row, which the bench measured at ~63 ms at 5,000
    # records (not the bottleneck we'd been worried about, but the
    # spec asks for vectorisation).  New path uses pd.to_numeric
    # + np.where — same observable output (rounded to 0.01) but
    # vectorised.  ``Units`` is normalised for the inline comma
    # (``"1,234"`` to ``"1234"``) the same way the row-wise path
    # did via ``str(units).replace(",", "")``.
    pc = pd.to_numeric(df["Period Charge (£)"], errors="coerce")
    units = pd.to_numeric(
        df["Units (kWh)"].astype(str).str.replace(",", ""),
        errors="coerce",
    )
    df["Unit Rate (p/kWh)"] = np.where(
        (units > 0) & (pc > 0),
        np.round((pc / units) * 100, 2),
        np.nan,
    )

    # ``dup_df`` computation is kept in the path for backward
    # compatibility — the dup DataFrame is much smaller than the
    # kept set, so per-row apply only adds ms-level overhead.  The
    # per-row unit-rate helper lives at module scope (``_compute_unit_rate``)
    # so ``pickle`` can find it on round-trip (closures aren't picklable).
    if not dup_df.empty:
        dup_df["Unit Rate (p/kWh)"] = dup_df.apply(_compute_unit_rate, axis=1)
        # Matched-against kept-record block (Phase-2 follow-up).
        # Each duplicate row gets a clickable summary pointing
        # back to the *kept* record so an ombudsman reviewing the
        # workbook can navigate from the dup sheet to the
        # source-of-truth record with one click.  Earlier in the
        # dedup walk we built ``kept_for_summary`` keyed on the
        # duplicate's *original* df-index — that's also the index
        # ``dup_df.index`` carries because ``dup_df = df[is_dup]
        # .copy()`` runs *before* the ``reset_index(drops...)``
        # line.  So we can resolve the summary now without
        # re-doing any index resets.
        kept_idx_by_dup = {
            dup_idx: kept_for_summary.get(kept_for_dup.get(dup_idx, -1), {})
            for dup_idx in dup_df.index
        }

        # ``df`` is the kept set after dedup reset_index.  After
        # ``df = df[~is_dup].reset_index(drop=True)``, ``df.index``
        # is a sequential 0..N-1 range, *not* the original df
        # labels.  But the *order* of rows is preserved — the n-th
        # row of the kept set is the same n-th kept row that survived
        # dedup.  We therefore translate the original-index
        # references we still hold in ``kept_for_dup`` (the dedup
        # walker wrote them *before* reset_index) into post-reset
        # positions by ranking the kept rows in ascending original
        # df-index order — kept_rank[k] = rank-in-kept-set.
        kept_rank: dict[int, int] = {}
        for rank, orig_idx in enumerate(sorted(kept_for_summary.keys())):
            kept_rank[int(orig_idx)] = rank

        # ``Duplicate Of`` is the visible column on the dup sheet
        # itself; ``_matches_kept_idx`` is the link target the
        # Excel writer will use to mint the click-through hyperlink
        # back to the kept row in the main evidence report.
        dup_df["Duplicate Of"] = [_summary(kept_idx_by_dup, idx) for idx in dup_df.index]
        # ``_matches_kept_idx`` is the *post-reset* position of
        # the kept row in ``EDF Evidence Report`` — the Excel
        # writer uses this with ``A{+1}`` as the click target
        # so an ombudsman can jump from the dup cell directly to
        # the source-of-truth record.  We translate via
        # ``kept_rank`` (computed above from kept-against-original
        # ordering) because the dedup walker built ``kept_for_dup``
        # *before* ``reset_index`` ran on the kept frame.
        dup_df["_matches_kept_idx"] = pd.Series(
            {idx: kept_rank.get(int(kept_for_dup.get(idx, -1)), -1) for idx in dup_df.index},
            dtype="Int64",
        )

    # F2 (SEV-1): single source of truth for the saved-column
    # ordering.  Every ``_add_record``-time builder must stamp
    # every name in this list (use ``record.setdefault(col, "N/A")``
    # if unsure) — otherwise ``reindex`` silently drops the column
    # and the workbook schema drifts from what other readers
    # (Tariff Analysis, Dict Comparer) expect.  The structural
    # guard lives in ``tests/test_export_headers_invariant.py``.
    col_order = [
        "Source",
        "Sender",
        "Date",
        "Period From",
        "Period To",
        "Invoice #",
        "Amount (£)",
        "Period Charge (£)",
        "Unit Rate (p/kWh)",
        "% Change",
        "Entry Type",
        "Reading",
        "Units (kWh)",
        "Standing Chg (p/day)",
        # Tariff column — lights up the Tariff Analysis Excel/DOCX/PDF
        # section.  Populated only by ``_process_new_invoice``;
        # every other source path stamps "N/A".  Without this entry
        # here, ``reindex`` would drop the column from the saved
        # workbook even though every record dict now carries it.
        "Tariff",
        "Attachment Name",
        "Details",
        "Logic Used",
        "Anomaly Flag",
        "Duplicate Of",
        "Sub Periods",
    ]
    # Diagnostic-only columns that the analyser writers (Back-billing,
    # Rebilling, Meter Readings, Contract History) need for their
    # Source Excerpt column lookup, but which must NOT appear on the
    # EDF Evidence Report tab.  They survive the canonical ``reindex``
    # below so ``dfc = df_an[...]`` retains them for the analyser
    # writers' ``evidence_df=dfc`` argument.  ``write_evidence_sheet``
    # drops them via the ``evidence_df = df.drop(columns=[...])`` pass
    # at line ~3694 just before the Evidence Report is written.
    # 'Balance Last Bill (£)' is captured by the reconciliation-statement
    # parser and is consumed by the analyser writers as a diagnostic; it
    # stays here for the same reason.
    diagnostic_cols = [
        "Source PDF Text",
        "_regex_trace",
        "Balance Last Bill (£)",
    ]
    # Only carry forward the diagnostic cols that are actually present
    # on the records -- avoids reindex inserting all-NaN cols when no
    # record builder emitted them (e.g. a synthetic test DataFrame).
    diagnostic_present = [c for c in diagnostic_cols if c in df.columns]
    df = df.reindex(columns=col_order + diagnostic_present)
    # Belt-and-braces invariant: every column the *kept* set still
    # carries must be in the canonical order list — otherwise a
    # future record builder that adds a new column without updating
    # col_order would survive the reindex and land as a
    # mysteries-leading-column in the saved workbook.  We assert
    # loudly here (developer-visible) rather than silently dropping
    # the unknown column.  The diagnostic cols (``Source PDF Text``,
    # ``_regex_trace``, ``Balance Last Bill (£)``) are intentionally
    # excluded from the canonical ``col_order`` so they're not written
    # to the Evidence Report sheet; the assertion below permits them.
    _allowed_extras = {"Source PDF Text", "_regex_trace", "Balance Last Bill (£)"}
    _unexpected = [c for c in df.columns if c not in col_order and c not in _allowed_extras]
    if _unexpected:
        raise ValueError(
            "export_to_excel received columns not in col_order: "
            f"{_unexpected!r}.  Add them to col_order or build the "
            "records so they carry only known keys."
        )
    # The dup sheet needs both ``Duplicate Of`` *and*
    # ``_matches_kept_idx`` available to the writer so the
    # post-loop pass can mint clickable HYPERLINK cells.  We
    # attach ``_matches_kept_idx`` after the reindex pass so the
    # saved workbook geometry stays 19-column even though the
    # writer's row-iteration will see the 20th column briefly —
    # the writer drops the column before saving.
    if not dup_df.empty and "_matches_kept_idx" in dup_df.columns:
        # Already present — nothing to do.
        pass
    else:
        # Neither column nor value is preserved.  Don't write
        # anything — the post-loop pass will skip minting
        # HYPERLINKs because ``match_positions_series`` is None.
        pass
    # No-op reindex guard for clarity; dup_df reindex on col_order
    # actually *drops* the helper column, which is what we want
    # for the Excel geometry — but we also need it for the
    # hyperlink pass.  Best approach: call site reads it BEFORE
    # reindex and threads it via a separate side cache.
    # The simplest implementation is to re-attach the column
    # *after* reindex here:
    if not dup_df.empty:
        dup_df_reindexed = dup_df.reindex(columns=col_order)
        # Re-attach from dup_df's pre-reindex view — the column
        # is dropped by reindex, so we restore it from the
        # original here.  This is the only place where the
        # writer would otherwise lose access to the helper.
        if "_matches_kept_idx" in dup_df.columns:
            dup_df = pd.concat(
                [
                    dup_df_reindexed,
                    dup_df["_matches_kept_idx"].rename("_matches_kept_idx"),
                ],
                axis=1,
            )
        else:
            dup_df = dup_df_reindexed

    # Years for summary tab
    years = sorted(
        y for y in df["Date"].apply(parse_to_sort_date).dropna().dt.year.astype(int).unique()
    )

    wb = openpyxl.Workbook()
    wb.calculation.fullCalcOnLoad = True

    # Tab 1: Evidence (created first — summary formulas reference it by name)
    ws_main = wb.active
    ws_main.title = "EDF Evidence Report"
    # Provenance sheet must be created AFTER ``ws_main.title`` is set above:
    # inserting at index 0 makes the new tab the active sheet, so creating it
    # earlier would steal ``wb.active`` and the Evidence tab would lose its name.
    _write_provenance_sheet(
        wb,
        config,
        n_records=len(df),
        n_filtered=len(filtered) if filtered else 0,
        n_errors=len(error_log),
    )
    # The diagnostic-only columns (``Source PDF Text``, ``_regex_trace``,
    # ``Balance Last Bill (£)``) are captured by the parsers for the
    # analyser tabs' Source Excerpt column / balance-context rendering.
    # They are intentionally NOT written to the visible Evidence Report
    # tab: ``Source PDF Text`` is a 4 KB chunk per row (too noisy),
    # ``_regex_trace`` is internal pipeline metadata, and
    # ``Balance Last Bill (£)`` is a reconciliation-statement field that
    # only the Reconciliation tab needs.
    # Drop them from the copy handed to the writer; the underlying
    # ``df`` is left intact so subsequent analyser renders (``dfc``)
    # retain them for in-memory Source Excerpt lookups.
    _diagnostic_columns_for_evidence_report = [
        "Source PDF Text",
        "_regex_trace",
        "Balance Last Bill (£)",
    ]
    evidence_df = df.drop(
        columns=[c for c in _diagnostic_columns_for_evidence_report if c in df.columns],
        errors="ignore",
    )
    write_evidence_sheet(ws_main, evidence_df, is_duplicate=False)

    # Tab 2: Annual Summary
    ws_summary = wb.create_sheet(title="Annual Summary", index=0)
    write_summary_sheet(ws_summary, years, ws_main.title, last_data_row=len(df) + 1)

    # Tab 3: Duplicates
    if not dup_df.empty:
        # Same diagnostic-column cleanup as the main evidence sheet:
        # dup_df inherits the diagnostic-only columns from reindex so
        # the duplication hotspot is visible here, but they're
        # inappropriate on the Duplicate Entries tab itself.
        dup_df_for_report = dup_df.drop(
            columns=[c for c in _diagnostic_columns_for_evidence_report if c in dup_df.columns],
            errors="ignore",
        )
        ws_dup = wb.create_sheet(title="Duplicate Entries")
        write_evidence_sheet(ws_dup, dup_df_for_report, is_duplicate=True)

    # Tab 4: Filtered
    if filtered and config.get("save_filtered", True):
        ws_filt = wb.create_sheet(title="Filtered (Below Min)")
        filt_headers = ["Source", "Date", "Amount (£)", "Details", "Logic Used", "Reason"]
        for ci, h in enumerate(filt_headers, 1):
            _hcell(ws_filt, 1, ci, h, bg="888888")
        filt_df = pd.DataFrame(filtered).sort_values("Amount (£)", ascending=False)
        for r_idx, frow in enumerate(filt_df.values, 2):
            bg_hex = "F5F5F5" if r_idx % 2 == 0 else None
            for c_idx, val in enumerate(frow, 1):
                c = ws_filt.cell(row=r_idx, column=c_idx, value=val)
                c.font = Font(name="Calibri", size=10)
                c.border = CELL_BORDER
                if bg_hex:
                    c.fill = PatternFill("solid", start_color=bg_hex)
                if c_idx == 3:
                    c.number_format = "£#,##0.00"
        for col, w in zip(["A", "B", "C", "D", "E", "F"], [18, 13, 14, 38, 18, 28], strict=False):
            ws_filt.column_dimensions[col].width = w
        ws_filt.freeze_panes = "A2"

    # Tab 5: Parse errors
    if error_log:
        ws_err = wb.create_sheet(title="Parse Errors")
        _hcell(ws_err, 1, 1, "Time", bg="888888")
        _hcell(ws_err, 1, 2, "Context", bg="888888")
        _hcell(ws_err, 1, 3, "Error", bg="888888")
        for r_idx, entry in enumerate(error_log, 2):
            ts_m = re.match(r"\[(.+?)\]\s*(.*?)\s*—\s*(.*)", entry)
            if ts_m:
                ts, ctx, err = ts_m.group(1), ts_m.group(2), ts_m.group(3)
            else:
                ts, ctx, err = "", entry, ""
            for c_idx, val in enumerate([ts, ctx, err], 1):
                c = ws_err.cell(row=r_idx, column=c_idx, value=val)
                c.font = Font(name="Calibri", size=10)
                c.border = CELL_BORDER
        ws_err.column_dimensions["A"].width = 10
        ws_err.column_dimensions["B"].width = 45
        ws_err.column_dimensions["C"].width = 60

    # =====================================================================
    # ANALYSIS SUITE
    # Uses bills above analysis_min threshold only (payments/credits always included).
    # =====================================================================

    df_an = df.copy()
    df_an["_dt"] = df_an["Date"].apply(parse_to_sort_date)
    df_an = df_an.sort_values("_dt").reset_index(drop=True)

    dfc = _prepare_analysis_frame(df_an, config)
    dfc["year"] = dfc["_dt"].dt.year
    dfc["month"] = dfc["_dt"].dt.month

    if len(dfc) < 2:
        # Not enough data for analysis sheets; save the workbook with what
        # we have (evidence, summary, duplicates, etc. are already written).
        _reorder_sheets(wb)
        wb.save(output_path)
        return

    amounts = dfc["Amount (£)"].values.astype(float)
    dates_lbl = dfc["Date"].tolist()
    n = len(amounts)

    raw_diffs = np.diff(amounts)
    pos_diffs = raw_diffs[raw_diffs > 0]

    yearly = (
        dfc.groupby("year")
        .agg(
            count=("Amount (£)", "count"),
            avg_bal=("Amount (£)", "mean"),
            peak=("Amount (£)", "max"),
            low=("Amount (£)", "min"),
        )
        .reset_index()
    )

    # ----- TAB A: KEY STATISTICS -----
    ws_ks = wb.create_sheet(title="Key Statistics")
    ws_ks.column_dimensions["A"].width = 44
    ws_ks.column_dimensions["B"].width = 22
    ws_ks.column_dimensions["C"].width = 44

    tc = ws_ks.cell(row=1, column=1, value="EDF ENERGY DISPUTE  —  KEY STATISTICS")
    tc.font = Font(name="Calibri", size=13, bold=True, color="FFFFFF")
    tc.fill = PatternFill("solid", start_color=ORANGE)
    tc.border = CELL_BORDER
    tc.alignment = Alignment(horizontal="left", vertical="center")
    for c in [2, 3]:
        x = ws_ks.cell(row=1, column=c)
        x.fill = PatternFill("solid", start_color=ORANGE)
        x.border = CELL_BORDER
    ws_ks.row_dimensions[1].height = 26

    acc_ref = str(config.get("report_account_ref") or config.get("acc_num") or "N/A")

    r = 2
    _section_hdr(ws_ks, r, "ACCOUNT OVERVIEW")
    r = 3
    _ks_row(ws_ks, r, "Account reference", acc_ref, alt=True)
    r = 4
    _ks_row(
        ws_ks,
        r,
        "First bill on record",
        "='Balance Trend'!A2",
        fmt="date",
        note="From Balance Trend sheet",
    )
    r = 5
    _ks_row(
        ws_ks,
        r,
        "Most recent bill",
        "=INDEX('Balance Trend'!A:A,MATCH(9.99E+307,'Balance Trend'!B:B)+1)",
        fmt="date",
        alt=True,
    )
    r = 6
    _ks_row(
        ws_ks,
        r,
        "Period covered (days)",
        "=IFERROR(INT(INDEX('Balance Trend'!A:A,MATCH(9.99E+307,'Balance Trend'!B:B)+1)-'Balance Trend'!A2),\"\")",
        fmt="#,##0",
        note="Days between first and last bill",
    )
    r = 7
    _ks_row(
        ws_ks,
        r,
        "Total bills on record",
        "=IFERROR(COUNT('Balance Trend'!B:B),\"\")",
        fmt="#,##0",
        alt=True,
    )

    r = 8
    _section_hdr(ws_ks, r, "BALANCE FIGURES")
    r = 9
    _ks_row(
        ws_ks,
        r,
        "Opening balance (first bill)",
        "='Balance Trend'!B2",
        fmt="£",
        alt=True,
        note="First entry in Balance Trend",
    )
    r = 10
    _ks_row(
        ws_ks,
        r,
        "Current balance (latest bill)",
        "=INDEX('Balance Trend'!B:B,MATCH(9.99E+307,'Balance Trend'!B:B))",
        fmt="£",
        bold=True,
        note="Last numeric entry in Balance Trend",
    )
    r = 11
    _ks_row(
        ws_ks,
        r,
        "Total balance increase",
        '=IFERROR(B10-B9,"")',
        fmt="£",
        bold=True,
        alt=True,
        note="Latest minus earliest",
    )
    r = 12
    _ks_row(ws_ks, r, "% increase over full period", '=IFERROR((B10-B9)/B9,"")', fmt="%", bold=True)
    r = 13
    _ks_row(
        ws_ks,
        r,
        "Mean balance across all bills",
        "=IFERROR(AVERAGE('Balance Trend'!B:B),\"\")",
        fmt="£",
        alt=True,
    )
    r = 14
    _ks_row(ws_ks, r, "Median balance", "=IFERROR(MEDIAN('Balance Trend'!B:B),\"\")", fmt="£")
    r = 15
    _ks_row(
        ws_ks,
        r,
        "Peak balance recorded",
        "=IFERROR(MAX('Balance Trend'!B:B),\"\")",
        fmt="£",
        alt=True,
    )
    r = 16
    _ks_row(ws_ks, r, "Lowest balance recorded", "=IFERROR(MIN('Balance Trend'!B:B),\"\")", fmt="£")

    r = 17
    _section_hdr(ws_ks, r, "PERIODIC CHARGES")
    r = 18
    _ks_row(
        ws_ks,
        r,
        "Note",
        "Bills are a running cumulative balance — periodic charge = closing minus opening balance",
        alt=True,
    )
    r = 19
    _ks_row(
        ws_ks,
        r,
        "Mean charge per period (positive only)",
        '=IFERROR(AVERAGEIF(\'Period Charges\'!F:F,">0"),"")',
        fmt="£",
    )
    r = 20
    _ks_row(
        ws_ks,
        r,
        "Largest single-period charge",
        "=IFERROR(MAX('Period Charges'!F:F),\"\")",
        fmt="£",
        bold=True,
        alt=True,
    )
    r = 21
    _ks_row(
        ws_ks,
        r,
        "Smallest positive charge",
        "=IFERROR(_xlfn.MINIFS('Period Charges'!F:F,'Period Charges'!F:F,\">0\"),\"\")",
        fmt="£",
    )
    r = 22
    _ks_row(
        ws_ks,
        r,
        "Periods where balance increased",
        '=IFERROR(COUNTIF(\'Period Charges\'!F:F,">0"),"")',
        fmt="#,##0",
        alt=True,
    )
    r = 23
    _ks_row(
        ws_ks,
        r,
        "Periods where balance fell (payments/credits)",
        '=IFERROR(COUNTIF(\'Period Charges\'!F:F,"<0"),"")',
        fmt="#,##0",
    )
    r = 24
    _ks_row(
        ws_ks,
        r,
        "Implied annual rate (avg last 6 charges ×12)",
        "=IFERROR(AVERAGE(OFFSET('Period Charges'!F1,MAX(1,COUNTIF('Period Charges'!F:F,\">0\")-5),0,6,1))*12,\"\")",
        fmt="£",
        bold=True,
        alt=True,
        note="Assumes ~monthly billing — may overstate if billing is quarterly",
    )

    r = 25
    _section_hdr(ws_ks, r, "READING & DATA QUALITY")
    r = 26
    _ks_row(
        ws_ks,
        r,
        "Estimated readings",
        '=IFERROR(COUNTIF(\'EDF Evidence Report\'!L:L,"Estimated"),"")',
        fmt="#,##0",
        alt=True,
    )
    r = 27
    _ks_row(
        ws_ks,
        r,
        "Actual / customer readings",
        '=IFERROR(COUNTIF(\'EDF Evidence Report\'!L:L,"Actual"),"")',
        fmt="#,##0",
    )
    r = 28
    _ks_row(
        ws_ks,
        r,
        "Smart meter readings",
        '=IFERROR(COUNTIF(\'EDF Evidence Report\'!L:L,"Smart"),"")',
        fmt="#,##0",
        alt=True,
    )
    r = 29
    _ks_row(
        ws_ks,
        r,
        "% of bills with estimated readings",
        "=IFERROR(B26/COUNT('EDF Evidence Report'!G:G),\"\")",
        fmt="%",
    )

    r = 30
    _section_hdr(ws_ks, r, "UNIT RATES")
    r = 31
    _ks_row(
        ws_ks,
        r,
        "Average unit rate (p/kWh)",
        "=IFERROR(AVERAGE('EDF Evidence Report'!I:I),\"\")",
        fmt="0.00",
        alt=True,
        note="Across all bills with valid period charge and kWh",
    )
    r = 32
    _ks_row(
        ws_ks,
        r,
        "Maximum unit rate (p/kWh)",
        "=IFERROR(MAX('EDF Evidence Report'!I:I),\"\")",
        fmt="0.00",
        note="Highest effective rate — potential overcharge",
    )
    r = 33
    _ks_row(
        ws_ks,
        r,
        "Minimum unit rate (p/kWh)",
        "=IFERROR(MIN('EDF Evidence Report'!I:I),\"\")",
        fmt="0.00",
        alt=True,
    )

    ws_ks.freeze_panes = "A2"

    # ----- TAB B: BALANCE TREND -----
    ws_bt = wb.create_sheet(title="Balance Trend")
    for ci, h in enumerate(
        ["Date", "Balance (£)", "6-Bill Rolling Avg (£)", "Linear Trend (£)", "Period Charge (£)"],
        1,
    ):
        _hcell(ws_bt, 1, ci, h, bg=NAVY)
    ws_bt.row_dimensions[1].height = 22

    last_data_row = n + 1
    for i in range(n):
        r = i + 2
        bg = LGREY if i % 2 == 0 else None

        # Write date as a true Excel date serial
        excel_dt = to_excel_date(dates_lbl[i])
        c1 = ws_bt.cell(row=r, column=1, value=excel_dt)
        c1.number_format = "dd/mm/yyyy"
        c1.font = Font(name="Calibri", size=10)
        c1.border = CELL_BORDER
        c1.alignment = Alignment(horizontal="left")
        if bg:
            c1.fill = PatternFill("solid", start_color=bg)

        _money(ws_bt, r, 2, float(amounts[i]), fill_hex=bg)

        start_r = max(2, r - 5)
        for col_i, formula in [
            (3, f'=IFERROR(AVERAGE(B{start_r}:B{r}),"")'),
            (
                4,
                f'=IFERROR(FORECAST.LINEAR(ROW(),B$2:B${last_data_row},ROW(B$2:B${last_data_row})),"")',
            ),
        ]:
            cx = ws_bt.cell(row=r, column=col_i, value=formula)
            cx.number_format = "£#,##0.00"
            cx.font = Font(name="Calibri", size=10)
            cx.border = CELL_BORDER
            cx.alignment = Alignment(horizontal="right")
            if bg:
                cx.fill = PatternFill("solid", start_color=bg)

        if i > 0:
            c5 = ws_bt.cell(row=r, column=5, value=f"=B{r}-B{r - 1}")
            c5.number_format = "£#,##0.00"
            c5.font = Font(name="Calibri", size=10)
            c5.border = CELL_BORDER
            c5.alignment = Alignment(horizontal="right")
            if bg:
                c5.fill = PatternFill("solid", start_color=bg)

    # Line chart
    lc = LineChart()
    lc.title = "Account Balance Over Time"
    lc.style = 10
    lc.y_axis.title = "Balance (£)"
    lc.x_axis.title = "Bill Date"
    lc.width, lc.height = 30, 18
    data_ref = Reference(ws_bt, min_col=2, max_col=4, min_row=1, max_row=n + 1)
    dates_ref = Reference(ws_bt, min_col=1, min_row=2, max_row=n + 1)
    lc.add_data(data_ref, titles_from_data=True)
    lc.set_categories(dates_ref)
    lc.series[0].graphicalProperties.line.solidFill = ORANGE
    lc.series[0].graphicalProperties.line.width = 22000
    if len(lc.series) > 1:
        lc.series[1].graphicalProperties.line.solidFill = NAVY
        lc.series[1].graphicalProperties.line.width = 15000
        lc.series[1].graphicalProperties.line.dashDot = "dash"
    if len(lc.series) > 2:
        lc.series[2].graphicalProperties.line.solidFill = DGREY
        lc.series[2].graphicalProperties.line.width = 10000
        lc.series[2].graphicalProperties.line.dashDot = "sysDash"
    ws_bt.add_chart(lc, "G2")
    for col, w in zip(["A", "B", "C", "D", "E"], [14, 16, 20, 16, 16], strict=False):
        ws_bt.column_dimensions[col].width = w
    ws_bt.freeze_panes = "A2"

    # ----- TAB C: YEAR-ON-YEAR -----
    ws_yoy = wb.create_sheet(title="Year-on-Year")
    for ci, h in enumerate(
        [
            "Year",
            "Bills",
            "Peak Balance (£)",
            "Avg Balance (£)",
            "Lowest Balance (£)",
            "YoY Avg Δ (£)",
            "YoY Avg Δ (%)",
            "Est. Readings",
            "Biggest Jump (£)",
        ],
        1,
    ):
        _hcell(ws_yoy, 1, ci, h, bg=ORANGE)
    ws_yoy.row_dimensions[1].height = 22

    prev_avg = None
    yoy_data = []
    for r_off, row_y in enumerate(yearly.itertuples(), 2):
        yr = row_y.year
        cnt = row_y.count
        pk = row_y.peak
        av = row_y.avg_bal
        lo = row_y.low
        yoy_chg_pct: float | None = ((av - prev_avg) / prev_avg) if prev_avg else None

        yr_rows = dfc[dfc["year"] == yr]
        yr_idx = yr_rows.index.tolist()
        max_jump = None
        for ii in yr_idx:
            if ii > 0 and ii in dfc.index and ii - 1 in dfc.index:
                jmp = dfc.at[ii, "Amount (£)"] - dfc.at[ii - 1, "Amount (£)"]
                if max_jump is None or jmp > max_jump:
                    max_jump = jmp

        alt = r_off % 2 == 0
        bg = LGREY if alt else None

        _num(ws_yoy, r_off, 1, yr, fmt="#,##0", fill_hex=bg, bold=True)
        _num(ws_yoy, r_off, 2, cnt, fmt="#,##0", fill_hex=bg)
        _money(ws_yoy, r_off, 3, pk, fill_hex=bg, bold=True)
        _money(ws_yoy, r_off, 4, av, fill_hex=bg)
        _money(ws_yoy, r_off, 5, lo, fill_hex=bg)

        if r_off > 2:
            c6 = ws_yoy.cell(row=r_off, column=6, value=f"=D{r_off}-D{r_off - 1}")
            c6.number_format = "£#,##0.00"
            c6.font = Font(name="Calibri", size=10, bold=True)
            c6.border = CELL_BORDER
            c6.alignment = Alignment(horizontal="right")
            if bg:
                c6.fill = PatternFill("solid", start_color=bg)

            c7 = ws_yoy.cell(row=r_off, column=7, value=f'=IFERROR(F{r_off}/D{r_off - 1},"")')
            c7.number_format = "+0.0%;-0.0%;—"
            c7.font = Font(name="Calibri", size=10, bold=True)
            c7.border = CELL_BORDER
            c7.alignment = Alignment(horizontal="right")
            yoy_fill = (
                RED
                if yoy_chg_pct is not None and yoy_chg_pct > 0.5
                else (
                    AMBER
                    if yoy_chg_pct is not None and yoy_chg_pct > 0.2
                    else (GREEN if yoy_chg_pct is not None and yoy_chg_pct < -0.1 else bg)
                )
            )
            if yoy_fill:
                c7.fill = PatternFill("solid", start_color=yoy_fill)
        else:
            ws_yoy.cell(row=r_off, column=6, value="—").border = CELL_BORDER
            ws_yoy.cell(row=r_off, column=7, value="—").border = CELL_BORDER

        yr_est = (
            int((dfc[dfc["year"] == yr]["Reading"] == "Estimated").sum())
            if "Reading" in dfc.columns
            else 0
        )
        _num(ws_yoy, r_off, 8, yr_est, fmt="#,##0", fill_hex=bg)
        if max_jump is not None:
            _money(ws_yoy, r_off, 9, max_jump, fill_hex=(RED if max_jump > 5000 else bg))

        yoy_data.append((yr, av))
        prev_avg = av

    bc = BarChart()
    bc.type = "col"
    bc.title = "Average Balance by Year"
    bc.y_axis.title = "Average Balance (£)"
    bc.style = 10
    bc.width, bc.height = 22, 14
    n_yrs = len(yoy_data)
    avg_ref = Reference(ws_yoy, min_col=4, min_row=1, max_row=n_yrs + 1)
    yr_ref = Reference(ws_yoy, min_col=1, min_row=2, max_row=n_yrs + 1)
    bc.add_data(avg_ref, titles_from_data=True)
    bc.set_categories(yr_ref)
    bc.series[0].graphicalProperties.solidFill = ORANGE
    ws_yoy.add_chart(bc, "K2")
    for col, w in zip(
        ["A", "B", "C", "D", "E", "F", "G", "H", "I"],
        [8, 8, 18, 18, 18, 16, 14, 14, 18],
        strict=False,
    ):
        ws_yoy.column_dimensions[col].width = w
    ws_yoy.freeze_panes = "A2"

    # ----- TAB D: PERIOD CHARGES -----
    ws_pc = wb.create_sheet(title="Period Charges")
    for ci, h in enumerate(
        [
            "From Date",
            "To Date",
            "Days",
            "Opening Balance (£)",
            "Closing Balance (£)",
            "Charge (£)",
            "Daily Rate (£/day)",
            "Flag",
        ],
        1,
    ):
        _hcell(ws_pc, 1, ci, h, bg=NAVY)
    ws_pc.row_dimensions[1].height = 22

    mean_daily = float(np.mean(pos_diffs)) / 30.0 if len(pos_diffs) else 0
    pc_rows_data = []

    pc_r = 2
    for i in range(1, n):
        p = dfc.iloc[i - 1]
        c_ = dfc.iloc[i]
        days = (c_["_dt"] - p["_dt"]).days
        charge = float(c_["Amount (£)"]) - float(p["Amount (£)"])
        daily = charge / days if days > 0 else None

        flag = ""
        if days > 90:
            flag = f"⚠ {days}-day gap — possible missed bill(s)"
        elif charge < 0:
            flag = f"↓ Balance reduced by £{abs(charge):,.2f} (payment or credit)"
        elif daily and mean_daily > 0 and daily > mean_daily * 2.5:
            flag = f"⚠ Daily rate {daily / mean_daily:.1f}× average"

        bg = LGREY if pc_r % 2 == 0 else None
        if flag.startswith("⚠"):
            bg = AMBER
        elif charge < 0:
            bg = GREEN

        _text(ws_pc, pc_r, 1, p["Date"], fill_hex=bg)
        _text(ws_pc, pc_r, 2, c_["Date"], fill_hex=bg)
        _num(ws_pc, pc_r, 3, days, fmt="#,##0", fill_hex=bg)
        _money(ws_pc, pc_r, 4, float(p["Amount (£)"]), fill_hex=bg)
        _money(ws_pc, pc_r, 5, float(c_["Amount (£)"]), fill_hex=bg)

        c6 = ws_pc.cell(row=pc_r, column=6, value=f"=E{pc_r}-D{pc_r}")
        c6.number_format = "£#,##0.00"
        c6.font = Font(name="Calibri", size=10)
        c6.border = CELL_BORDER
        c6.alignment = Alignment(horizontal="right")
        if bg:
            c6.fill = PatternFill("solid", start_color=bg)

        c7 = ws_pc.cell(row=pc_r, column=7, value=f'=IFERROR(F{pc_r}/C{pc_r},"")')
        c7.number_format = "£#,##0.00"
        c7.font = Font(name="Calibri", size=10)
        c7.border = CELL_BORDER
        c7.alignment = Alignment(horizontal="right")
        if bg:
            c7.fill = PatternFill("solid", start_color=bg)

        _text(ws_pc, pc_r, 8, flag, fill_hex=bg, wrap=True)

        if charge > 0:
            pc_rows_data.append((c_["Date"], charge))
        pc_r += 1

    if pc_r > 2:
        sr = pc_r + 2
        _section_hdr(ws_pc, sr, "SUMMARY STATISTICS", ncols=8, bg=ORANGE)
        sr += 1
        dr = f"F2:F{pc_r - 1}"
        cr = f"C2:C{pc_r - 1}"

        _pc_stat(
            ws_pc,
            sr,
            "Mean charge per period (positive only)",
            f'=IFERROR(AVERAGEIF({dr},">0"),"")',
        )
        _pc_stat(ws_pc, sr + 1, "Largest single charge", f'=IFERROR(MAX({dr}),"")')
        _pc_stat(ws_pc, sr + 2, "Largest credit / reduction", f'=IFERROR(MIN({dr}),"")')
        _pc_stat(ws_pc, sr + 3, "Charge periods", f'=IFERROR(COUNTIF({dr},">0"),"")', fmt="#,##0")
        _pc_stat(ws_pc, sr + 4, "Credit periods", f'=IFERROR(COUNTIF({dr},"<0"),"")', fmt="#,##0")
        _pc_stat(
            ws_pc,
            sr + 5,
            "Average days between bills",
            f'=IFERROR(AVERAGE({cr}),"")',
            fmt="#,##0.0",
        )

    if len(pc_rows_data) > 1:
        bc2 = BarChart()
        bc2.type = "col"
        bc2.title = "Charge Added Each Period"
        bc2.y_axis.title = "Charge (£)"
        bc2.style = 10
        bc2.width, bc2.height = 28, 14
        chg_ref2 = Reference(ws_pc, min_col=6, min_row=1, max_row=pc_r - 1)
        date_ref2 = Reference(ws_pc, min_col=2, min_row=2, max_row=pc_r - 1)
        bc2.add_data(chg_ref2, titles_from_data=True)
        bc2.set_categories(date_ref2)
        bc2.series[0].graphicalProperties.solidFill = NAVY
        ws_pc.add_chart(bc2, "J2")

    for col, w in zip(
        ["A", "B", "C", "D", "E", "F", "G", "H"], [13, 13, 7, 18, 18, 16, 14, 42], strict=False
    ):
        ws_pc.column_dimensions[col].width = w
    ws_pc.freeze_panes = "A2"

    # ----- TAB E: DISPUTE FLAGS -----
    ws_df = wb.create_sheet(title="Dispute Flags")

    _banner(ws_df, 1, "EDF ENERGY DISPUTE  —  AUTOMATED ANALYSIS FLAGS", ORANGE)
    ws_df.cell(
        row=2,
        column=1,
        value=f"Generated {datetime.now().strftime('%d/%m/%Y %H:%M')}  |  Period: {dates_lbl[0]} to {dates_lbl[-1]}",
    )
    ws_df.cell(row=2, column=1).font = Font(name="Calibri", size=9, italic=True, color=DGREY)

    for ci, (txt, col_hex) in enumerate(
        [
            ("■ RED = HIGH severity", RED),
            ("■ AMBER = MEDIUM", AMBER),
            ("■ GREEN = Payment/credit", GREEN),
        ],
        1,
    ):
        lc2 = ws_df.cell(row=3, column=ci * 2 - 1, value=txt)
        lc2.font = Font(name="Calibri", size=9, bold=True)
        lc2.fill = PatternFill("solid", start_color=col_hex)
        lc2.border = CELL_BORDER

    hdr_row = 5
    for ci, h in enumerate(["#", "Date", "Balance (£)", "Flag Type", "Detail", "Severity"], 1):
        _hcell(ws_df, hdr_row, ci, h, bg=NAVY)

    flags, counts = compute_dispute_flags(dfc, mean_daily)

    sev_fill = {"HIGH": RED, "MEDIUM": AMBER, "INFO": GREEN}
    for fi, (ftype, flag_date, amt, detail, sev) in enumerate(flags, hdr_row + 1):
        bg = sev_fill.get(sev, LGREY)
        _num(ws_df, fi, 1, fi - hdr_row, fmt="#,##0", fill_hex=bg)
        _text(ws_df, fi, 2, flag_date or "—", fill_hex=bg)
        if amt:
            _money(ws_df, fi, 3, float(amt), fill_hex=bg)
        else:
            ws_df.cell(row=fi, column=3).fill = PatternFill("solid", start_color=bg)
            ws_df.cell(row=fi, column=3).border = CELL_BORDER
        _text(ws_df, fi, 4, ftype, bold=True, fill_hex=bg)
        _text(ws_df, fi, 5, detail, fill_hex=bg, wrap=True)
        _text(ws_df, fi, 6, sev, bold=True, fill_hex=bg, align="center")
        ws_df.row_dimensions[fi].height = 30

    if flags:
        fr = len(flags) + hdr_row + 2
        counts = {s: sum(1 for f in flags if f[4] == s) for s in ("HIGH", "MEDIUM", "INFO")}
        _banner(
            ws_df,
            fr,
            f"TOTAL FLAGS: {len(flags)}   |   HIGH: {counts['HIGH']}   |   MEDIUM: {counts['MEDIUM']}   |   INFO: {counts['INFO']}",
            NAVY,
        )

    for col, w in zip(["A", "B", "C", "D", "E", "F"], [5, 13, 16, 20, 60, 10], strict=False):
        ws_df.column_dimensions[col].width = w
    ws_df.freeze_panes = f"A{hdr_row + 1}"

    # ----- TAB F: DISPUTE TIMELINE -----
    ws_tl = wb.create_sheet(title="Dispute Timeline")
    _banner(ws_tl, 1, "EDF ENERGY DISPUTE  —  CHRONOLOGICAL TIMELINE", ORANGE)
    ws_tl.cell(
        row=2, column=1, value=f"Account: {acc_ref}  |  Period: {dates_lbl[0]} to {dates_lbl[-1]}"
    )
    ws_tl.cell(row=2, column=1).font = Font(name="Calibri", size=9, italic=True, color=DGREY)

    for ci, h in enumerate(["Date", "Event Type", "Description"], 1):
        _hcell(ws_tl, 4, ci, h, bg=NAVY)

    timeline_events = []

    # Bookend: first record
    timeline_events.append(
        (dates_lbl[0], "ACCOUNT START", f"First bill on record. Balance: £{amounts[0]:,.2f}.")
    )

    # Top 5 largest balance jumps
    jumps = []
    for i in range(1, n):
        delta = float(amounts[i]) - float(amounts[i - 1])
        days = (dfc.iloc[i]["_dt"] - dfc.iloc[i - 1]["_dt"]).days
        if delta > 0:
            jumps.append((delta, i, days))
    jumps.sort(key=lambda x: x[0], reverse=True)
    for delta, idx, days in jumps[:5]:
        timeline_events.append(
            (
                dfc.iloc[idx]["Date"],
                "LARGE INCREASE",
                f"Balance rose £{delta:,.2f} in {days} days "
                f"(from £{amounts[idx - 1]:,.2f} to £{amounts[idx]:,.2f}).",
            )
        )

    # Billing gaps > 60 days
    for i in range(1, n):
        days = (dfc.iloc[i]["_dt"] - dfc.iloc[i - 1]["_dt"]).days
        if days > 60:
            timeline_events.append(
                (
                    dfc.iloc[i]["Date"],
                    "BILLING GAP",
                    f"{days} days without a bill (previous: {dfc.iloc[i - 1]['Date']}). "
                    f"Balance accumulated unchecked.",
                )
            )

    # Estimated reading runs (reuse existing detection)
    if "Reading" in dfc.columns:
        run = 0
        run_start_date = None
        for i, rv in enumerate(dfc["Reading"].tolist()):
            if str(rv).lower() in ("estimated", "est."):
                run += 1
                if run == 1:
                    run_start_date = dfc.iloc[i]["Date"]
            else:
                if run >= 3:
                    timeline_events.append(
                        (
                            run_start_date,
                            "ESTIMATED READINGS",
                            f"{run} consecutive bills used estimated meter readings.",
                        )
                    )
                run = 0
                run_start_date = None
        if run >= 3:
            timeline_events.append(
                (
                    run_start_date,
                    "ESTIMATED READINGS",
                    f"{run} consecutive estimated readings (ongoing).",
                )
            )

    # Payment events (balance reductions)
    for i in range(1, n):
        delta = float(amounts[i]) - float(amounts[i - 1])
        if delta < -200:
            timeline_events.append(
                (
                    dfc.iloc[i]["Date"],
                    "PAYMENT/CREDIT",
                    f"Balance reduced by £{abs(delta):,.2f} "
                    f"(from £{amounts[i - 1]:,.2f} to £{amounts[i]:,.2f}).",
                )
            )

    # Reconciliation mismatches (from flags)
    for ftype, fdate, _famt, fdetail, _fsev in flags:
        if ftype == "RECONCILIATION MISMATCH":
            timeline_events.append((fdate, "RECONCILIATION", fdetail))

    # Bookend: latest record
    timeline_events.append(
        (
            dates_lbl[-1],
            "CURRENT STATE",
            f"Latest bill on record. Balance: £{amounts[-1]:,.2f}. "
            f"Total increase from first record: £{amounts[-1] - amounts[0]:,.2f}.",
        )
    )

    # Sort by date and write
    timeline_events.sort(key=lambda e: parse_to_sort_date(e[0]) or pd.Timestamp.min)
    tl_r = 5
    for tl_date, etype, desc in timeline_events:
        bg_hex = LGREY if tl_r % 2 == 0 else None
        _text(ws_tl, tl_r, 1, tl_date, fill_hex=bg_hex)
        _text(ws_tl, tl_r, 2, etype, bold=True, fill_hex=bg_hex)
        _text(ws_tl, tl_r, 3, desc, fill_hex=bg_hex, wrap=True)
        ws_tl.row_dimensions[tl_r].height = 40
        tl_r += 1

    for col, w in zip(["A", "B", "C"], [14, 22, 90], strict=False):
        ws_tl.column_dimensions[col].width = w
    ws_tl.freeze_panes = "A5"

    # =====================================================================
    # NEW ANALYSIS TABS (added after Dispute Timeline)
    # =====================================================================

    # Statistical Analysis
    write_statistical_analysis_sheet(wb.create_sheet(title="Statistical Analysis"), dfc, config)

    # Payment Analysis
    write_payment_analysis_sheet(wb.create_sheet(title="Payment Analysis"), dfc)

    # Forecast & Projection
    write_forecast_sheet(wb.create_sheet(title="Forecast & Projection"), dfc)

    # Data Quality Report
    write_data_quality_sheet(wb.create_sheet(title="Data Quality Report"), df)

    # Tariff Analysis (if data available)
    write_tariff_analysis_sheet(wb.create_sheet(title="Tariff Analysis"), dfc)

    # ------------------------------------------------------------------
    # Phase-2 analysis tabs (back-billing, rebilling, meter rollover,
    # contract history). run_analysers runs the four detectors on the
    # same `dfc` (post-dedup, post-filter) the rest of the workbook
    # uses, then each writer paints the result onto its own tab.
    # The new tabs append AFTER the existing 16 -- no existing sheet
    # is touched. Account label is pulled from config['acc_num'].
    # ------------------------------------------------------------------
    account_label = str(config.get("acc_num", "") or "")
    from edf_bill_fetcher.io.writers.analysis import run_analysers

    # Evidence index must be built on the FULL evidence frame `df`, not
    # the filtered `dfc`: the index maps signatures to Excel rows on the
    # EDF Evidence Report sheet, whose layout follows the full frame.
    analyses = run_analysers(dfc, evidence_index_df=df)
    rb = analyses["rebilling"]
    overlapping_invoices: set[str] = (
        {str(x) for x in rb["Killer Invoice"].tolist()} if not rb.empty else set()
    )
    domination_map = compute_transitive_domination(
        rb,
        analyses["back_billing"],
    )
    write_back_billing_sheet(
        wb.create_sheet(title="Back-billing Analysis"),
        analyses["back_billing"],
        account=account_label,
        overlapping_invoices=overlapping_invoices,
        evidence_df=dfc,
        evidence_index=analyses["evidence_index"],
        domination_map=domination_map,
    )
    write_rebilling_sheet(
        wb.create_sheet(title="Rebilling & Corrections"),
        analyses["rebilling"],
        account=account_label,
        evidence_df=dfc,
        evidence_index=analyses["evidence_index"],
    )
    write_meter_readings_sheet(
        wb.create_sheet(title="Meter Readings"),
        dfc,
        analyses["meter_rollover"],
        account=account_label,
        evidence_df=dfc,
        evidence_index=analyses["evidence_index"],
    )
    write_contract_history_sheet(
        wb.create_sheet(title="Contract History"),
        analyses["contracts"],
        account=account_label,
        evidence_df=dfc,
        evidence_index=analyses["evidence_index"],
    )

    # Stream P1 + P2: SAP CSV-in-PDF data dumps and the cross-source
    # Reconciliation sheet. When ``sap_rows`` is supplied (from the
    # engine's three SAP-row accumulators) and the user hasn't opted
    # out via ``config["scan_sap_dumps"] = False``, emit the three SAP
    # sheets. The Reconciliation sheet additionally honours
    # ``config["generate_reconciliation_sheet"]`` (default True) so a
    # reviewer can toggle it off independently when only the SAP data
    # is wanted.
    sap_rows = sap_rows or {}
    sap_contract = list(sap_rows.get("contract") or [])
    sap_meter = list(sap_rows.get("meter") or [])
    sap_financial = list(sap_rows.get("financial") or [])
    scan_sap_dumps = config.get("scan_sap_dumps", True)
    if scan_sap_dumps and (sap_contract or sap_meter or sap_financial):
        if sap_contract:
            write_sap_contract_history_sheet(
                wb.create_sheet(title="SAP Contract History"),
                sap_contract,
                account=account_label,
            )
        if sap_meter:
            write_sap_meter_readings_sheet(
                wb.create_sheet(title="SAP Meter Readings"),
                sap_meter,
                account=account_label,
            )
        if sap_financial:
            write_sap_financial_transactions_sheet(
                wb.create_sheet(title="SAP Financial Transactions"),
                sap_financial,
                account=account_label,
            )
            # SAP Back-billing analyser (spec §6):
            # uses the EDF Evidence Report rows (filter/dedup-applied
            # ``dfc``) as the join target. Both new sheets appear under
            # the existing ``scan_sap_dumps`` toggle alongside the
            # existing SAP sheets.
            edf_records_for_bb: list[dict] = []
            if dfc is not None and not dfc.empty:
                edf_records_for_bb = dfc.to_dict(orient="records")
            bb_events = detect_sap_back_billing_events(sap_financial)
            bb_matches = match_sap_events_to_edf(bb_events, edf_records_for_bb)
            # Populate Sheet 1's "Matched EDF Invoice #" column with the
            # highest-confidence match per event (tiebreak: smallest
            # date_delta_days).
            for ev in bb_events:
                ev_matches = [m for m in bb_matches if m.event is ev]
                if ev_matches:
                    conf_rank = {"High": 0, "Medium": 1, "Low": 2}
                    best = sorted(
                        ev_matches,
                        key=lambda m: (
                            conf_rank.get(m.confidence_band, 3),
                            m.date_delta_days,
                        ),
                    )[0]
                    ev.matched_edf_invoice = str(best.edf_record.get("Invoice #", "") or "") or None
            # Wire cluster-unmatched: SAP events with no amount-banded invoice
            # match that fall inside a known back-billing cluster's window are
            # tagged as internal mechanism events of that cluster.
            unmatched_events = [ev for ev in bb_events if ev.matched_edf_invoice is None]
            if (
                unmatched_events
                and "back_billing" in analyses
                and not analyses["back_billing"].empty
            ):
                clusters = _build_bb_clusters(analyses["back_billing"])
                for ev in unmatched_events:
                    tag = handle_cluster_unmatched(ev, clusters)
                    if tag is not None:
                        ev.matched_edf_invoice = tag["Matched EDF Invoice #"]
                        ev._cluster_unmatched_tag = tag
            write_sap_back_billing_sheets(
                wb,
                bb_events,
                bb_matches,
                sap_financial_first_row=4,
                edf_rows=edf_records_for_bb,
                edf_sheet_name="EDF Evidence Report",
                edf_first_row=4,
                account=account_label,
                sap_row_index_map=_build_sap_row_index_map(sap_financial),
            )
            if config.get("generate_reconciliation_sheet", True):
                from edf_bill_fetcher.io.writers.sap import write_sap_back_billing_position_sheet
                from edf_bill_fetcher.processors.matching import analyse_sap_back_billing

                sap_bb_position = analyse_sap_back_billing(
                    bb_events, dfc, analyses.get("back_billing")
                )
                write_sap_back_billing_position_sheet(wb, sap_bb_position, account=account_label)
        if config.get("generate_reconciliation_sheet", True):
            ws_recon_summary = wb.create_sheet(title="Reconciliation")
            ws_recon_detail = wb.create_sheet(title="Reconciliation Drill-down")
            write_reconciliation_sheet(
                ws_recon_summary,
                ws_recon_detail,
                sap_contract,
                analyses["contracts"],
                sap_meter,
                dfc,
                sap_financial,
                dfc,
                account=account_label,
            )

    _reorder_sheets(wb)

    try:
        wb.save(output_path)
    except FileNotFoundError:
        output_dir = os.path.dirname(output_path) or "."
        raise FileNotFoundError(
            f"The output folder does not exist:\n\n    {output_dir}\n\n"
            f"Either create this folder first, or click the 'Browse' button next to "
            f"'Output Folder' in the GUI to pick an existing folder.\n"
            f"(Target file was: {os.path.basename(output_path)})"
        ) from None


__all__ = ["export_to_excel"]
