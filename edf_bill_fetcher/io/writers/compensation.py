"""Indicative Compensation evidence sheet writer (Wave 6d, Task 7).

Renders the claims produced by
:func:`edf_bill_fetcher.processors.compensation.estimate_compensation`
onto a ``Compensation`` tab: rows grouped by ``category`` under label
rows, one row per claim, with a trailing total of ``indicative_amount``
and a DISCLAIMER row so the workbook is self-documenting that the
figures are indicative, not legal advice.

Follows the canonical sheet-writer pattern in ``io/writers/superseded.py``:
banner / section-label / merged-text / header-row / trailing-total /
freeze from ``io/writers/sheet_layout.py``; cell primitives (``_text``,
``_money``, ``_num``) from ``helpers/excel_utils.py``; widths via
``set_column_widths_from_spec``.
"""

from __future__ import annotations

from typing import Any

from openpyxl.styles import Font
from openpyxl.worksheet.worksheet import Worksheet

from edf_bill_fetcher.helpers.excel_utils import (
    money as _money,
)
from edf_bill_fetcher.helpers.excel_utils import (
    num as _num,
)
from edf_bill_fetcher.helpers.excel_utils import (
    set_column_widths_from_spec,
)
from edf_bill_fetcher.helpers.excel_utils import (
    text as _text,
)
from edf_bill_fetcher.io.writers.sheet_layout import (
    freeze_at,
    write_banner,
    write_header_row,
    write_merged_text,
    write_section_label,
    write_trailing_total,
)
from edf_bill_fetcher.processors.compensation import DISCLAIMER

# Column order for the data rows.  ``rate`` is rendered as a percentage
# when present (credit-interest rows) and blank for back-billing excess.
_COLS = [
    "Invoice #",
    "Date",
    "Base Amount (£)",
    "Days",
    "Rate",
    "Indicative Amount (£)",
    "Legal Basis",
]

_CATEGORY_LABELS: dict[str, str] = {
    "back_billing_excess": "BACK-BILLING EXCESS",
    "credit_hold_interest": "CREDIT-HOLD INTEREST",
    "late_credit_interest": "LATE-CREDIT INTEREST",
}

# Stable category display order (matches the estimator's emission order).
_CATEGORY_ORDER = [
    "back_billing_excess",
    "credit_hold_interest",
    "late_credit_interest",
]

_NAVY = "10367A"
_ORANGE = "FE5716"
_ALT_FILL = "EEF2FF"


def write_compensation_sheet(
    ws: Worksheet,
    comp_rows: list[dict[str, Any]] | None,
) -> None:
    """Render the indicative-compensation worksheet.

    ``comp_rows`` is the list of row dicts produced by
    :func:`estimate_compensation` — each ``{category, invoice_ref, date,
    base_amount, days, rate, indicative_amount, legal_basis, disclaimer}``.
    Rows are grouped by ``category`` under a label row, one row per claim,
    with a trailing total of ``indicative_amount`` and a DISCLAIMER row.

    An empty or ``None`` ``comp_rows`` writes nothing (the caller gates
    sheet creation on non-empty rows, so this is a defensive no-op).
    """
    rows = comp_rows or []
    if not rows:
        return

    ws.title = "Compensation"
    write_banner(ws, "INDICATIVE COMPENSATION", len(_COLS), color=_ORANGE, row=1, height=22)
    write_section_label(ws, 2, "LEGAL CONTEXT", len(_COLS))
    write_merged_text(
        ws,
        3,
        "Indicative compensation claims estimated deterministically from the "
        "extracted evidence records. Figures are indicative only and are not "
        "legal advice; verify against original documents before use in any "
        "formal dispute.",
        len(_COLS),
        height=40,
    )
    write_header_row(ws, 5, _COLS, bg=_NAVY, height=28)

    r = 6
    total = 0.0
    for category in _CATEGORY_ORDER:
        group = [row for row in rows if row.get("category") == category]
        if not group:
            continue
        label = _CATEGORY_LABELS.get(category, category.upper())
        label_cell = ws.cell(row=r, column=1, value=label)
        label_cell.font = Font(bold=True, color=_NAVY)
        r += 1
        for row in group:
            bg = _ALT_FILL if r % 2 == 0 else None
            invoice_ref = str(row.get("invoice_ref", "") or "")
            date = str(row.get("date", "") or "")
            base_amount = float(row.get("base_amount", 0.0) or 0.0)
            days = int(row.get("days", 0) or 0)
            rate = row.get("rate")
            indicative = float(row.get("indicative_amount", 0.0) or 0.0)
            legal_basis = str(row.get("legal_basis", "") or "")
            total += indicative

            _text(ws, r, 1, invoice_ref, fill_hex=bg)
            _text(ws, r, 2, date, fill_hex=bg)
            _money(ws, r, 3, base_amount, fill_hex=bg)
            _num(ws, r, 4, days, fmt="#,##0", fill_hex=bg)
            if rate is not None:
                _num(ws, r, 5, float(rate), fmt="0.00%", fill_hex=bg)
            else:
                _text(ws, r, 5, "", fill_hex=bg)
            _money(ws, r, 6, indicative, fill_hex=bg)
            _text(ws, r, 7, legal_basis, wrap=True, fill_hex=bg)
            r += 1

    write_trailing_total(
        ws,
        r,
        "TOTAL INDICATIVE COMPENSATION",
        [(6, round(total, 2))],
        5,
        len(_COLS),
    )
    r += 1
    # DISCLAIMER row spanning the full width.
    disclaimer_cell = ws.cell(row=r, column=1, value=DISCLAIMER)
    disclaimer_cell.font = Font(name="Calibri", size=9, italic=True, color="595959")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(_COLS))

    widths: dict[str, float] = {
        "A": 18,
        "B": 14,
        "C": 16,
        "D": 10,
        "E": 12,
        "F": 18,
        "G": 60,
    }
    set_column_widths_from_spec(ws, widths)
    freeze_at(ws, "A6")


__all__ = ["write_compensation_sheet"]
