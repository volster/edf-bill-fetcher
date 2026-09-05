"""Statistical analysis sheet writer — extracted from writers/__init__.py.

Contains: write_statistical_analysis_sheet — renders the "Statistical Analysis"
worksheet with anomaly detection, distribution moments, and JB/Shapiro tests.
"""

from __future__ import annotations

import importlib.util

from openpyxl.styles import Alignment, Font, PatternFill

from edf_bill_fetcher.helpers.date_utils import parse_to_sort_date
from edf_bill_fetcher.helpers.excel_utils import (
    hcell as _hcell,
)
from edf_bill_fetcher.helpers.excel_utils import (
    money as _money,
)
from edf_bill_fetcher.helpers.excel_utils import (
    num as _num,
)
from edf_bill_fetcher.helpers.excel_utils import (
    section_hdr as _section_hdr,
)
from edf_bill_fetcher.helpers.excel_utils import (
    text as _text,
)
from edf_bill_fetcher.helpers.theme import CELL_BORDER

# Local scipy-availability probe — mirrors the module-level
# ``HAS_SCIPY`` constant that lived in ``writers/__init__.py`` before
# this function was extracted. Kept private to this module because
# no other writer needs it.
_HAS_SCIPY = importlib.util.find_spec("scipy") is not None


# --- write_statistical_analysis_sheet (was writers/__init__.py L1721-1956) ---


def write_statistical_analysis_sheet(ws, dfc, config):
    """Write Statistical Analysis tab with advanced pandas analytics."""
    from edf_bill_fetcher.models.report_models import compute_statistical_analysis

    ws.title = "Statistical Analysis"

    NAVY = "10367A"
    ORANGE = "FE5716"
    AMBER = "FFD166"
    LGREY = "F0F0F0"
    DGREY = "888888"

    # Prepare data
    dfc = dfc.copy()
    dfc["_dt"] = dfc["Date"].apply(parse_to_sort_date)
    dfc = dfc.sort_values("_dt").reset_index(drop=True)
    n = len(dfc)

    if n < 3:
        _hcell(ws, 1, 1, "Insufficient data for statistical analysis", bg=NAVY)
        ws.column_dimensions["A"].width = 50
        return

    sa = compute_statistical_analysis(dfc)

    # Headers
    headers = [
        "Metric",
        "Value",
        "Notes",
    ]
    for col, h in enumerate(headers, 1):
        _hcell(ws, 2, col, h, bg=NAVY)
    ws.row_dimensions[2].height = 28

    # Title
    tc = ws.cell(row=1, column=1, value="EDF ENERGY DISPUTE  —  STATISTICAL ANALYSIS")
    tc.font = Font(name="Calibri", size=13, bold=True, color="FFFFFF")
    tc.fill = PatternFill("solid", start_color=ORANGE)
    tc.border = CELL_BORDER
    tc.alignment = Alignment(horizontal="left", vertical="center")
    for c in [2, 3]:
        x = ws.cell(row=1, column=c)
        x.fill = PatternFill("solid", start_color=ORANGE)
        x.border = CELL_BORDER

    # Summary stats
    r = 3
    _section_hdr(ws, r, "DESCRIPTIVE STATISTICS")

    stats_data = [
        ("Count", sa.count, "#,##0", "Number of billing records"),
        ("Mean (£)", sa.mean, "£#,##0.00", "Average balance"),
        ("Median (£)", sa.median, "£#,##0.00", "Median balance"),
        ("Std Dev (£)", sa.std, "£#,##0.00", "Standard deviation"),
        ("Min (£)", sa.minimum, "£#,##0.00", "Minimum balance"),
        ("Max (£)", sa.maximum, "£#,##0.00", "Maximum balance"),
        ("Range (£)", sa.range, "£#,##0.00", "Max - Min"),
        ("Skewness", sa.skewness, "0.00", "Asymmetry of distribution"),
        ("Kurtosis", sa.kurtosis, "0.00", "Tailedness of distribution"),
        (
            "CV (%)",
            sa.cv if sa.cv is not None else 0.0,
            "0.00",
            "Coefficient of variation",
        ),
    ]

    for label, value, fmt, note in stats_data:
        r += 1
        bg = LGREY if r % 2 == 0 else None
        _text(ws, r, 1, label, fill_hex=bg)
        _num(ws, r, 2, value, fmt=fmt, fill_hex=bg)
        _text(ws, r, 3, note, fill_hex=bg, color=DGREY)

    # Rolling statistics
    r += 1
    _section_hdr(ws, r, "ROLLING STATISTICS (6-period window)")
    r += 1
    _text(ws, r, 1, "Current 6-Period Rolling Mean (£)", bold=True)
    _money(ws, r, 2, sa.rolling["mean"])

    r += 1
    _text(ws, r, 1, "Current 6-Period Rolling Std (£)", bold=True)
    _money(ws, r, 2, sa.rolling["std"])

    r += 1
    _text(ws, r, 1, "Current 6-Period Rolling Min (£)", bold=True)
    _money(ws, r, 2, sa.rolling["min"])

    r += 1
    _text(ws, r, 1, "Current 6-Period Rolling Max (£)", bold=True)
    _money(ws, r, 2, sa.rolling["max"])

    r += 1
    _text(ws, r, 1, "Current 6-Period Rolling Median (£)", bold=True)
    _money(ws, r, 2, sa.rolling_median)

    # Exponential Moving Average
    r += 1
    _section_hdr(ws, r, "EXPONENTIAL MOVING AVERAGE")
    r += 1
    _text(ws, r, 1, "Current EMA (span=6) (£)", bold=True)
    _money(ws, r, 2, sa.ema)

    r += 1
    _text(ws, r, 1, "EMA vs Simple SMA Difference (£)", bold=True)
    _money(ws, r, 2, sa.ema - sa.rolling["mean"])

    # Momentum & Volatility
    r += 1
    _section_hdr(ws, r, "MOMENTUM & VOLATILITY")
    r += 1
    _text(ws, r, 1, "3-Period Momentum (£)", bold=True)
    _money(ws, r, 2, sa.momentum)

    r += 1
    _text(ws, r, 1, "6-Period Volatility (σ of returns)", bold=True)
    _num(ws, r, 2, sa.volatility, fmt="0.00%")

    # Anomaly Detection
    r += 1
    _section_hdr(ws, r, "ANOMALY DETECTION")

    r += 1
    _text(ws, r, 1, "Z-Score Anomalies (threshold=2.5σ)", bold=True)
    _num(ws, r, 2, sa.z_count, fmt="#,##0")

    r += 1
    _text(ws, r, 1, "IQR Anomalies (multiplier=1.5)", bold=True)
    _num(ws, r, 2, sa.iqr_count, fmt="#,##0")

    # List detected anomalies
    if sa.z_count > 0:
        r += 1
        _text(ws, r, 1, "Z-Score Anomaly Dates:", bold=True)
        for entry in sa.z_dates:
            r += 1
            _text(ws, r, 1, f"  • {entry}")

    if sa.iqr_count > 0:
        r += 1
        _text(ws, r, 1, "IQR Anomaly Dates:", bold=True)
        for entry in sa.iqr_dates:
            r += 1
            _text(ws, r, 1, f"  • {entry}")

    # Normality test (if scipy available)
    r += 1
    _section_hdr(ws, r, "DISTRIBUTION TESTS")
    if _HAS_SCIPY:
        shapiro_stat = sa.shapiro_stat
        shapiro_p = sa.shapiro_p
        jb_stat = sa.jb_stat
        jb_p = sa.jb_p
        if (
            shapiro_stat is not None
            and shapiro_p is not None
            and jb_stat is not None
            and jb_p is not None
        ):
            r += 1
            _text(ws, r, 1, "Shapiro-Wilk Test (Normality)", bold=True)
            _num(ws, r, 2, shapiro_stat, fmt="0.0000")
            _text(
                ws,
                r,
                3,
                f"p-value: {shapiro_p:.4f} — {'Normal' if shapiro_p > 0.05 else 'Non-normal'}",
            )

            r += 1
            _text(ws, r, 1, "Jarque-Bera Test (Normality)", bold=True)
            _num(ws, r, 2, jb_stat, fmt="0.00")
            _text(ws, r, 3, f"p-value: {jb_p:.4f} — {'Normal' if jb_p > 0.05 else 'Non-normal'}")
        else:
            r += 1
            _text(ws, r, 1, "Scipy tests failed", fill_hex=AMBER)
    else:
        r += 1
        _text(ws, r, 1, "Scipy not available — install for normality tests", fill_hex=AMBER)

    # Column widths
    for col_letter, width in zip(["A", "B", "C"], [45, 22, 80], strict=False):
        ws.column_dimensions[col_letter].width = width
    ws.freeze_panes = "A3"


__all__ = ["write_statistical_analysis_sheet"]
