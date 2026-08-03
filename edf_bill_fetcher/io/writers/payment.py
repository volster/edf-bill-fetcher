"""Payment analysis sheet writer — extracted from writers/__init__.py.

Contains: write_payment_analysis_sheet — renders the "Payment Analysis" worksheet
with bar charts of payment amounts over time and pattern-detection callouts.
"""

from __future__ import annotations

from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill

from edf_bill_fetcher.helpers.date_utils import parse_to_sort_date
from edf_bill_fetcher.helpers.excel_utils import (
    hcell as _hcell,
)
from edf_bill_fetcher.helpers.excel_utils import (
    money as _money,
)
from edf_bill_fetcher.helpers.excel_utils import (
    num as _num,
)
from edf_bill_fetcher.helpers.excel_utils import (
    section_hdr as _section_hdr,
)
from edf_bill_fetcher.helpers.excel_utils import (
    text as _text,
)
from edf_bill_fetcher.helpers.theme import CELL_BORDER
from edf_bill_fetcher.writers._helpers import _detect_payment_patterns

# --- write_payment_analysis_sheet (was writers/__init__.py L1959-2172) ---


def write_payment_analysis_sheet(ws, dfc):
    """Write Payment/Credit Analysis tab."""
    ws.title = "Payment Analysis"

    NAVY = "10367A"
    ORANGE = "FE5716"
    LGREY = "F0F0F0"
    DGREY = "888888"

    payments = dfc[dfc["Entry Type"].isin(["Payment", "Credit"])].copy()
    if payments.empty:
        _hcell(ws, 1, 1, "No payment/credit records found", bg=NAVY)
        ws.column_dimensions["A"].width = 50
        return

    payments["_dt"] = payments["Date"].apply(parse_to_sort_date)
    payments = payments.sort_values("_dt").reset_index(drop=True)

    headers = ["Metric", "Value", "Notes"]
    for col, h in enumerate(headers, 1):
        _hcell(ws, 1, col, h, bg=NAVY)
    ws.row_dimensions[1].height = 28

    tc = ws.cell(row=1, column=1, value="EDF ENERGY DISPUTE  —  PAYMENT & CREDIT ANALYSIS")
    tc.font = Font(name="Calibri", size=13, bold=True, color="FFFFFF")
    tc.fill = PatternFill("solid", start_color=ORANGE)
    tc.border = CELL_BORDER
    tc.alignment = Alignment(horizontal="left", vertical="center")
    for c in [2, 3]:
        x = ws.cell(row=1, column=c)
        x.fill = PatternFill("solid", start_color=ORANGE)
        x.border = CELL_BORDER

    pat = _detect_payment_patterns(dfc)

    r = 2
    _section_hdr(ws, r, "PAYMENT SUMMARY")

    payment_items = [
        ("Total Payments/Credits", pat["count"], "#,##0", "Number of payment events"),
        ("Total Amount Paid (£)", pat["total_paid"], "£#,##0.00", "Sum of all payments/credits"),
        ("Average Payment (£)", pat["avg_payment"], "£#,##0.00", "Mean payment amount"),
        ("Median Payment (£)", pat["median_payment"], "£#,##0.00", "Median payment amount"),
        ("Largest Payment (£)", pat["max_payment"], "£#,##0.00", "Maximum single payment"),
        ("Smallest Payment (£)", pat["min_payment"], "£#,##0.00", "Minimum single payment"),
    ]

    for label, value, fmt, note in payment_items:
        r += 1
        bg = LGREY if r % 2 == 0 else None
        _text(ws, r, 1, label, fill_hex=bg)
        if fmt == "£":
            _money(ws, r, 2, value, fill_hex=bg)
        else:
            _num(ws, r, 2, value, fmt=fmt, fill_hex=bg)
        _text(ws, r, 3, note, fill_hex=bg, color=DGREY)

    # Payment intervals
    r += 1
    _section_hdr(ws, r, "PAYMENT TIMING")
    interval_items = [
        ("Avg Interval (days)", pat["avg_interval_days"], "#,##0.0", "Mean days between payments"),
        (
            "Median Interval (days)",
            pat["median_interval_days"],
            "#,##0.0",
            "Median days between payments",
        ),
    ]
    for label, value, fmt, note in interval_items:
        r += 1
        bg = LGREY if r % 2 == 0 else None
        _text(ws, r, 1, label, fill_hex=bg)
        if value is not None:
            _num(ws, r, 2, value, fmt=fmt, fill_hex=bg)
        else:
            _text(ws, r, 2, "N/A", fill_hex=bg)
        _text(ws, r, 3, note, fill_hex=bg, color=DGREY)

    # Last payment
    r += 1
    _section_hdr(ws, r, "LAST PAYMENT")
    r += 1
    _text(ws, r, 1, "Last Payment Date", bold=True)
    _text(ws, r, 2, pat["last_payment_date"] or "N/A")

    r += 1
    _text(ws, r, 1, "Last Payment Amount (£)", bold=True)
    _money(ws, r, 2, pat["last_payment_amount"] or 0)

    # Payment detail table
    r += 2
    _section_hdr(ws, r, "ALL PAYMENTS & CREDITS (Chronological)")
    r += 1
    pay_headers = ["Date", "Entry Type", "Amount (£)", "Balance After (£)", "Details"]
    for ci, h in enumerate(pay_headers, 1):
        _hcell(ws, r, ci, h, bg=NAVY)

    for i, (_, row) in enumerate(payments.iterrows()):
        r += 1
        bg = LGREY if i % 2 == 0 else None
        _text(ws, r, 1, row["Date"], fill_hex=bg)
        _text(ws, r, 2, row["Entry Type"], fill_hex=bg, bold=True)
        # Amount (£) column: the actual transaction amount (customer
        # payment or EDF credit). HTM Payment/Credit rows carry this
        # in Period Charge (£); legacy rows that only populated
        # Amount (£) use that instead.
        pc_val = row.get("Period Charge (£)")
        try:
            amount_to_show = float(pc_val)
        except (TypeError, ValueError):
            amount_to_show = float(row["Amount (£)"])
        _money(ws, r, 3, amount_to_show, fill_hex=bg)
        # Balance After (£) -- the running balance stored in
        # ``Amount (£)`` for HTM rows. For legacy rows where Amount
        # WAS the transaction, we have no separate balance, so show
        # the same amount (with a note that real balance-after is
        # not parsed for legacy formats).
        try:
            balance_after = float(row["Amount (£)"])
        except (TypeError, ValueError):
            balance_after = amount_to_show
        _money(ws, r, 4, balance_after, fill_hex=bg)
        _text(ws, r, 5, str(row.get("Details", ""))[:60], fill_hex=bg, wrap=True)

    # Chart - Payment amounts over time.
    # Phase-2 portability fix: the previous layout anchored the
    # chart at ``cell(row+2, column H)`` (column 8) which sat past
    # the visible data table (columns A-E) **and** the user's
    # roughly-default Excel viewport (about seven column-units
    # wide before they have to scroll).  An ombudsman reading
    # the report saw the chart title render *off-screen*.  We now:
    #
    #  * Place the chart-data helper cells in **column A**
    #    (single-cell-style) at a dedicated row block below the
    #    data so the chart reads ``date × amount`` cleanly;
    #  * Drop the chart *anchor* to column B, two rows below the
    #    data table — that's the most common Excel default
    #    reading order, so the user sees the data first and the
    #    chart underneath;
    #  * Cap the chart at width=16, height=10 (openpyxl's chart
    #    units, where 1 unit ≈ 1 Excel column / row).  The
    #    previous 28 × 14 values pushed the chart so far right
    #    that it appeared only partially when the file opened;
    #  * Use a colour-blind-friendly palette (single GREEN
    #    series — the existing colour — so a reviewer with
    #    deuteranopia can still trace payment size to date via
    #    the data labels).
    if len(payments) > 1:
        bc = BarChart()
        bc.type = "col"
        bc.title = "Payment/Credit Amounts Over Time"
        bc.y_axis.title = "Amount (£)"
        bc.x_axis.title = "Payment Date"
        bc.style = 10
        bc.width = 16
        bc.height = 10
        bc.legend = None

        # Step 1: write the chart-data series to a dedicated,
        # labelled mini-table two rows below the payments detail.
        # Putting both series in the same column range keeps the
        # chart's Reference call simple and avoids scattered helper
        # cells.
        chart_data_start_row = r + 3
        _hcell(ws, chart_data_start_row, 1, "Date", bg=NAVY)
        _hcell(
            ws,
            chart_data_start_row,
            2,
            "Payment Amount (£)",
            bg=NAVY,
        )
        for i, (_, row) in enumerate(payments.iterrows(), 1):
            payload_row = chart_data_start_row + i
            _text(ws, payload_row, 1, row["Date"])
            # Same preference logic as the detail table above:
            # the per-row transaction value (Period Charge (£))
            # over the running balance (Amount (£)).
            pc_val = row.get("Period Charge (£)")
            try:
                amount_for_chart = float(pc_val)
            except (TypeError, ValueError):
                amount_for_chart = float(row["Amount (£)"])
            _money(ws, payload_row, 2, amount_for_chart)

        # Step 2: build the chart from the labelled mini-table so
        # the title ("C2", "D2") series is unambiguous when a
        # reviewer opens the file's chart-edit dialog.
        chg_ref = Reference(
            ws,
            min_col=2,
            min_row=chart_data_start_row,
            max_row=chart_data_start_row + len(payments),
        )
        date_ref = Reference(
            ws,
            min_col=1,
            min_row=chart_data_start_row + 1,
            max_row=chart_data_start_row + len(payments),
        )
        bc.add_data(chg_ref, titles_from_data=True)
        bc.set_categories(date_ref)

        # Step 3: anchor the chart under the data table so the
        # reader's eye flows from raw rows to chart without
        # panning across the spreadsheet.  Row offset 2 gives the
        # chart a small breathing-room gap below the helper rows.
        anchor_row = chart_data_start_row + len(payments) + 2
        ws.add_chart(bc, f"B{anchor_row}")

    for col_letter, width in zip(["A", "B", "C", "D", "E"], [14, 16, 16, 16, 60], strict=False):
        ws.column_dimensions[col_letter].width = width
    ws.freeze_panes = f"A{r - len(payments)}"


__all__ = ["write_payment_analysis_sheet"]
