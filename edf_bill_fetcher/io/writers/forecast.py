"""Forecast sheet writer — extracted from writers/__init__.py.

Contains: write_forecast_sheet — renders the "Forecast" worksheet with
linear and Holt-Winters forecasts + forecast chart.
"""

from __future__ import annotations

from datetime import timedelta

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill

from edf_bill_fetcher.helpers.date_utils import parse_to_sort_date
from edf_bill_fetcher.helpers.excel_utils import (
    hcell as _hcell,
)
from edf_bill_fetcher.helpers.excel_utils import (
    money as _money,
)
from edf_bill_fetcher.helpers.excel_utils import (
    num as _num,
)
from edf_bill_fetcher.helpers.excel_utils import (
    section_hdr as _section_hdr,
)
from edf_bill_fetcher.helpers.excel_utils import (
    text as _text,
)
from edf_bill_fetcher.helpers.theme import CELL_BORDER
from edf_bill_fetcher.processors.forecasting import HAS_STATSMODELS

# --- write_forecast_sheet (was writers/__init__.py L2175-2422) ---


def write_forecast_sheet(ws, dfc):
    """Write Forecast/Projection tab with multiple forecasting methods."""
    from edf_bill_fetcher.models.report_models import compute_forecast

    ws.title = "Forecast & Projection"

    NAVY = "10367A"
    ORANGE = "FE5716"
    AMBER = "FFD166"
    LGREY = "F0F0F0"
    DGREY = "888888"

    dfc = dfc.copy()
    dfc["_dt"] = dfc["Date"].apply(parse_to_sort_date)
    dfc = dfc.sort_values("_dt").reset_index(drop=True)
    amounts = dfc["Amount (£)"].astype(float).values
    dates = dfc["Date"].tolist()
    n = len(amounts)

    if n < 3:
        _hcell(ws, 1, 1, "Insufficient data for forecasting (need 3+ records)", bg=NAVY)
        ws.column_dimensions["A"].width = 60
        return

    fc = compute_forecast(dfc)
    linear_fitted = fc.linear_fitted
    linear_fc = fc.linear_forecast
    hw_fitted = fc.hw_fitted
    hw_fc = fc.hw_forecast
    ema_series = fc.ema_series
    ema_future = fc.ema_forecast
    hist_vol = fc.hist_vol
    forecast_steps = 6

    # ``Date`` + the canonical six forecast columns + ``Forecast Δ
    # (Actual − Linear)``.  The Δ column is what makes the tab
    # useful as evidence: a reviewer sees *by how much* each bill
    # diverged from what the model would call average.  Historical
    # rows carry a per-row back-painted prediction; future rows
    # carry forward-looking projections; the divider between the
    # two is a separator row.
    headers = [
        "Date",
        "Actual (£)",
        "Linear Forecast (£)",
        "Holt-Winters (£)",
        "EMA Projection (£)",
        "Confidence (±£)",
        "Forecast Δ (Actual − Linear)",
    ]
    for col, h in enumerate(headers, 1):
        _hcell(ws, 2, col, h, bg=NAVY)
    ws.row_dimensions[2].height = 28

    tc = ws.cell(row=1, column=1, value="EDF ENERGY DISPUTE  —  BALANCE FORECAST")
    tc.font = Font(name="Calibri", size=13, bold=True, color="FFFFFF")
    tc.fill = PatternFill("solid", start_color=ORANGE)
    tc.border = CELL_BORDER
    tc.alignment = Alignment(horizontal="left", vertical="center")
    for c in range(2, 8):
        x = ws.cell(row=1, column=c)
        x.fill = PatternFill("solid", start_color=ORANGE)
        x.border = CELL_BORDER

    def _model_value(fitted_list, i):
        """Return the fitted value at index ``i`` or ``None`` when the model didn't fit."""
        if fitted_list is None:
            return None
        if i < len(fitted_list):
            val = fitted_list[i]
            return val if not pd.isna(val) else None
        return None

    # === Historical block: back-paint every forecast column ===
    r = 3
    for i in range(n):
        bg = LGREY if i % 2 == 0 else None
        _text(ws, r, 1, dates[i], fill_hex=bg)
        _money(ws, r, 2, float(amounts[i]), fill_hex=bg)
        lin_val = _model_value(linear_fitted, i)
        if lin_val is not None:
            _money(ws, r, 3, float(lin_val), fill_hex=bg)
        else:
            _text(ws, r, 3, "N/A", fill_hex=bg)
        hw_val = _model_value(hw_fitted, i)
        if hw_val is not None:
            _money(ws, r, 4, float(hw_val), fill_hex=bg)
        else:
            _text(ws, r, 4, "N/A", fill_hex=bg)
        ema_at_i = ema_series[i] if not pd.isna(ema_series[i]) else None
        if ema_at_i is not None:
            _money(ws, r, 5, ema_at_i, fill_hex=bg)
        else:
            _text(ws, r, 5, "N/A", fill_hex=bg)
        if lin_val is not None:
            conf = abs(float(lin_val)) * hist_vol * 2
            _money(ws, r, 6, conf, fill_hex=bg)
        else:
            _text(ws, r, 6, "N/A", fill_hex=bg)
        if lin_val is not None:
            delta = float(amounts[i]) - float(lin_val)
            _money(ws, r, 7, delta, fill_hex=bg)
        else:
            _text(ws, r, 7, "N/A", fill_hex=bg)
        r += 1

    # Separator
    ws.cell(row=r, column=1, value="— " * 20).font = Font(bold=True, color=DGREY)
    r += 1

    # === Forward forecast block: 6 steps past the last historical ===
    forecast_dates = []
    last_date = parse_to_sort_date(dates[-1])

    if not pd.isna(last_date):
        for i in range(1, forecast_steps + 1):
            next_date = last_date + timedelta(days=30 * i)  # Approximate monthly
            forecast_dates.append(next_date.strftime("%d/%m/%Y"))
    else:
        forecast_dates = [f"Forecast +{i + 1}" for i in range(forecast_steps)]

    for i in range(forecast_steps):
        bg = AMBER
        _text(ws, r, 1, forecast_dates[i], fill_hex=bg, bold=True)
        _text(ws, r, 2, "—", fill_hex=bg)  # No actual
        lin_val = linear_fc[i] if linear_fc and i < len(linear_fc) else None
        hw_val = hw_fc[i] if hw_fc and i < len(hw_fc) else None
        if lin_val is not None:
            _money(ws, r, 3, float(lin_val), fill_hex=bg)
        else:
            _text(ws, r, 3, "N/A", fill_hex=bg)
        if hw_val is not None:
            _money(ws, r, 4, float(hw_val), fill_hex=bg)
        else:
            _text(ws, r, 4, "N/A", fill_hex=bg)
        _money(ws, r, 5, ema_future[i], fill_hex=bg)
        if lin_val is not None:
            conf = abs(float(lin_val)) * hist_vol * 2
            _money(ws, r, 6, conf, fill_hex=bg)
        else:
            _text(ws, r, 6, "N/A", fill_hex=bg)
        _text(ws, r, 7, "—", fill_hex=bg)
        r += 1

    # Model comparison
    r += 1
    _section_hdr(ws, r, "MODEL COMPARISON")
    r += 1
    _text(ws, r, 1, "Linear Trend", bold=True)
    _text(ws, r, 2, "Simple linear regression on time index")
    r += 1
    _text(ws, r, 1, "Holt-Winters", bold=True)
    _text(
        ws, r, 2, "Exponential smoothing with trend" + (" + seasonality" if HAS_STATSMODELS else "")
    )
    r += 1
    _text(ws, r, 1, "EMA Projection", bold=True)
    _text(ws, r, 2, "Extends last Exponential Moving Average (span=6)")
    r += 1
    _text(ws, r, 1, "Historical Volatility", bold=True)
    _num(ws, r, 2, hist_vol, fmt="0.00%")
    _text(ws, r, 3, "Monthly return std used for confidence bands")

    # Accuracy metrics (in-sample)
    r += 1
    _section_hdr(ws, r, "IN-SAMPLE ACCURACY (Last 6 periods)")
    if n >= 7:
        mae = fc.mae
        rmse = fc.rmse
        mape = fc.mape
        if mae is not None and rmse is not None and mape is not None:
            r += 1
            _text(ws, r, 1, "Linear Forecast MAE (£)", bold=True)
            _money(ws, r, 2, mae)
            r += 1
            _text(ws, r, 1, "Linear Forecast RMSE (£)", bold=True)
            _money(ws, r, 2, rmse)
            r += 1
            _text(ws, r, 1, "Linear Forecast MAPE (%)", bold=True)
            _num(ws, r, 2, mape, fmt="0.00%")

    for col_letter, width in zip(
        ["A", "B", "C", "D", "E", "F", "G"], [14, 16, 18, 18, 18, 16, 22], strict=False
    ):
        ws.column_dimensions[col_letter].width = width
    ws.freeze_panes = "A3"


__all__ = ["write_forecast_sheet"]
