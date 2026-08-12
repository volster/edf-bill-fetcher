"""Rebilling analysis writer — extracted from writers/__init__.py.

Contains: detect_rebilling (pure-pandas detector for cancel-and-repost
invoice pairs), write_rebilling_sheet (renders the "Rebilling &
Corrections" worksheet), and the private _reversal_match helper that
checks whether a reversal-credit row in the evidence DataFrame matches
a killed invoice well enough to count as rebilling evidence.
"""

from __future__ import annotations

import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from edf_bill_fetcher.helpers.date_utils import _safe_to_datetime
from edf_bill_fetcher.helpers.excel_utils import (
    hcell as _hcell,
)
from edf_bill_fetcher.helpers.excel_utils import (
    num as _num,
)
from edf_bill_fetcher.helpers.excel_utils import (
    open_pdf_hyperlink_cell as _open_pdf_hyperlink_cell,
)
from edf_bill_fetcher.helpers.excel_utils import (
    text as _text,
)
from edf_bill_fetcher.helpers.theme import CELL_BORDER
from edf_bill_fetcher.processors.detection import detect_rebilling  # noqa: F401

# --- _reversal_match (was writers/__init__.py L3021-3059) ---


def _reversal_match(
    evidence_df: pd.DataFrame | None,
    killed_inv: str,
    killed_amount: float | None,
    killed_pf: pd.Timestamp,
    killed_pt: pd.Timestamp,
) -> bool:
    """Return whether a reversal-credit row in *evidence_df* matches the killed invoice well enough to count as rebilling evidence.

    Spec ref: 2026-07-16 §11. A reversal credit accepts the killed
    invoice when its amount is within ±£0.50 AND either its period
    overlaps the killed period by ≥ 30 days OR its period is
    unparseable (so we accept on amount alone, Entry Type == Credit).
    """
    if evidence_df is None or evidence_df.empty:
        return False
    if "Entry Type" not in evidence_df.columns:
        return False
    try:
        amount = abs(float(killed_amount or 0.0))
    except (TypeError, ValueError):
        return False
    matching = evidence_df[evidence_df["Entry Type"].isin(["Credit", "Payment"])]
    for _, row in matching.iterrows():
        try:
            row_amt = abs(float(row.get("Amount (£)", 0) or 0))
        except (TypeError, ValueError):
            continue
        if abs(row_amt - amount) > 0.50:
            continue
        rpf = _safe_to_datetime(row.get("Period From"))
        rpt = _safe_to_datetime(row.get("Period To"))
        if pd.isna(rpf) or pd.isna(rpt):
            return True
        overlap = (min(killed_pt, rpt) - max(killed_pf, rpf)).days
        if overlap >= 30:
            return True
    return False


# --- detect_rebilling (re-exported from processors.detection) ---


# --- write_rebilling_sheet (was writers/__init__.py L3375-3500) ---


def write_rebilling_sheet(
    ws: Worksheet,
    rb: pd.DataFrame,
    account: str = "",
    evidence_df: pd.DataFrame | None = None,
    evidence_index: dict[str, int] | None = None,
) -> None:
    """Render the Rebilling / Corrections tab (spec §4.2)."""
    ws.title = "Rebilling & Corrections"
    NAVY = "10367A"
    ORANGE = "FE5716"

    # Row 1: banner with account
    title = "REBILLING / CORRECTION EVENTS — Cancel-and-Repost Patterns"
    if account:
        title = f"{title}  |  Account {account}"
    t1 = ws.cell(row=1, column=1, value=title)
    t1.font = Font(name="Calibri", size=13, bold=True, color="FFFFFF")
    t1.fill = PatternFill("solid", start_color=ORANGE)
    t1.border = CELL_BORDER
    t1.alignment = Alignment(horizontal="left", vertical="center")
    for c in range(2, 10):
        x = ws.cell(row=1, column=c)
        x.fill = PatternFill("solid", start_color=ORANGE)
        x.border = CELL_BORDER
    ws.row_dimensions[1].height = 22

    # Row 2: subheader (merged)
    sub = (
        "Each row identifies a pair of invoices where the later invoice "
        "effectively cancelled and re-billed an earlier invoice's period. "
        "Heuristic: period overlap > 30 days OR billing starts >30 days "
        "earlier than the new invoice. Trigger Reason lists every "
        "matching heuristic."
    )
    sub_cell = ws.cell(row=2, column=1, value=sub)
    sub_cell.font = Font(name="Calibri", size=10, italic=True)
    sub_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=9)
    ws.row_dimensions[2].height = 45

    # Row 7: table headers (9 cols incl. Open PDF + View on Evidence Report).
    headers = [
        "Killer Invoice",
        "Killed Invoice",
        "Killer Date",
        "Killed Date",
        "Period Overlap (days)",
        "Jump-back (days)",
        "Trigger Reason",
        "Open PDF",
        "View on Evidence Report",
    ]
    for col, h in enumerate(headers, 1):
        _hcell(ws, 7, col, h, bg=NAVY)
    ws.row_dimensions[7].height = 28

    r = 8
    for _, row in rb.iterrows():
        bg = "EEF2FF" if r % 2 == 0 else None
        killer = str(row.get("Killer Invoice", ""))
        killed = str(row.get("Killed Invoice", ""))
        _text(ws, r, 1, killer, fill_hex=bg)
        _text(ws, r, 2, killed, fill_hex=bg)
        _text(ws, r, 3, row.get("Killer Date", ""), fill_hex=bg)
        _text(ws, r, 4, row.get("Killed Date", ""), fill_hex=bg)
        _num(
            ws,
            r,
            5,
            int(row.get("Period Overlap (days)", 0)),
            fmt="#,##0",
            fill_hex=bg,
        )
        _num(
            ws,
            r,
            6,
            int(row.get("Jump-back (days)", 0)),
            fmt="#,##0",
            fill_hex=bg,
        )
        _text(
            ws,
            r,
            7,
            str(row.get("Trigger Reason", "")),
            wrap=True,
            fill_hex=bg,
        )
        if bool(row.get("Cancel/Rebill Admitted (Killer)", False)):
            ws.cell(row=r, column=7).font = Font(name="Calibri", size=10, bold=True, color="C00000")
        _open_pdf_hyperlink_cell(ws, r, 8, evidence_df, killer)
        # View on Evidence Report (col 9): hotlink on the killer invoice row.
        target_row = None
        if evidence_index is not None:
            target_row = evidence_index.get(f"inv:{killer}") or evidence_index.get(f"inv:{killed}")
        if target_row is not None:
            cell = ws.cell(row=r, column=9, value="\u2192")
            cell.hyperlink = openpyxl.worksheet.hyperlink.Hyperlink(
                ref=cell.coordinate,
                location=f"'EDF Evidence Report'!A{target_row}",
                display="→",
                tooltip=f"Jump to EDF Evidence Report!A{target_row}",
            )
            cell.font = Font(name="Calibri", size=10, color="0563C1", underline="single")
        else:
            cell = ws.cell(row=r, column=9, value="No match")
            cell.font = Font(name="Calibri", size=10, italic=True, color="A6A6A6")
        r += 1

    # Column widths tailored for the table cells.
    widths = {
        "A": 18,
        "B": 18,
        "C": 14,
        "D": 14,
        "E": 18,
        "F": 16,
        "G": 50,
        "H": 60,  # Open PDF
        "I": 22,  # View on Evidence Report
    }
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width
    ws.freeze_panes = "A8"
