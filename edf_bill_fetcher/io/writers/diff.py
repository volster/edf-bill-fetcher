"""Diff workbook writer — renders a run-diff result as a 3-sheet Excel file.

Companion to the CLI ``--diff`` subcommand: ``write_diff_workbook`` takes
the ``{"added": [...], "removed": [...], "changed": [...]}`` dict returned
by ``edf_bill_fetcher.processors.run_diff.diff_records`` and writes one
sheet per category:

- **Added Records** / **Removed Records** — one row per record with a
  plain header.
- **Changed Records** — paired ``<field> (old)`` / ``<field> (new)``
  columns per record field plus a trailing ``Changed Fields`` column
  summarising each change as ``field: old → new``.

Openpyxl-only (sister of the other ``io.writers`` sheet modules), so it
stays testable without a workbook fixture.
"""

from __future__ import annotations

import math
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ADDED_SHEET = "Added Records"
REMOVED_SHEET = "Removed Records"
CHANGED_SHEET = "Changed Records"
CHANGED_FIELDS_HEADER = "Changed Fields"

_NAVY = "10367A"
_PREFERRED_FIELDS = ("Date", "Amount (£)", "Source")


def _iter_fields(records: list[dict[str, Any]]) -> list[str]:
    """Union of record fields, Date/Amount/Source first, rest alphabetical."""
    all_fields: set[str] = set()
    for record in records:
        all_fields.update(str(field) for field in record)
    return list(_PREFERRED_FIELDS) + sorted(all_fields - set(_PREFERRED_FIELDS))


def _cell_value(value: Any) -> Any:
    """Convert a record cell to an Excel-safe value (None for missing)."""
    if value is None:
        return None
    if isinstance(value, float) and math.isnan(value):
        return None
    if isinstance(value, (str, int, float, bool)):
        return value
    return str(value)


def _append_row(ws: Any, row: int, values: list[Any]) -> None:
    for col, value in enumerate(values, 1):
        ws.cell(row=row, column=col, value=value)


def _style_header_row(ws: Any, ncols: int, row: int = 1) -> None:
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", start_color=_NAVY)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.row_dimensions[row].height = 24


def _autosize(ws: Any, ncols: int) -> None:
    for col in range(1, ncols + 1):
        width = 0
        for row in range(1, ws.max_row + 1):
            value = ws.cell(row=row, column=col).value
            if value is not None:
                width = max(width, len(str(value)))
        ws.column_dimensions[get_column_letter(col)].width = min(max(width + 2, 10), 40)


def _write_records_sheet(ws: Any, records: list[dict[str, Any]]) -> None:
    """Plain sheet: one header row, one row per record."""
    fields = _iter_fields(records)
    _append_row(ws, 1, fields)
    for row_index, record in enumerate(records, 2):
        _append_row(ws, row_index, [_cell_value(record.get(field)) for field in fields])
    _style_header_row(ws, len(fields))
    _autosize(ws, len(fields))


def _summary_value(field: str, value: Any) -> str:
    """Render one delta value for the Changed Fields summary cell."""
    if value is None:
        return "-"
    if field == "Amount (£)":
        try:
            return f"£{float(value):.2f}"
        except (TypeError, ValueError):
            return str(value)
    return str(value)


def _changed_cell_summary(entry: dict[str, Any]) -> str:
    """Summarise one changed entry's deltas as ``field: old → new`` items."""
    parts: list[str] = []
    for field, old_value, new_value in zip(
        entry["changed_fields"], entry["old_values"], entry["new_values"], strict=True
    ):
        parts.append(
            f"{field}: {_summary_value(field, old_value)} → {_summary_value(field, new_value)}"
        )
    return ", ".join(parts)


def _write_changed_sheet(ws: Any, changed: list[dict[str, Any]]) -> None:
    """Paired old/new columns per field plus a Changed Fields summary column."""
    old_rows = [entry["old_row"] for entry in changed]
    new_rows = [entry["new_row"] for entry in changed]
    fields = _iter_fields(old_rows + new_rows)
    headers = [header for field in fields for header in (f"{field} (old)", f"{field} (new)")]
    headers.append(CHANGED_FIELDS_HEADER)
    _append_row(ws, 1, headers)

    for row_index, entry in enumerate(changed, 2):
        values: list[Any] = []
        for field in fields:
            values.append(_cell_value(entry["old_row"].get(field)))
            values.append(_cell_value(entry["new_row"].get(field)))
        values.append(_changed_cell_summary(entry))
        _append_row(ws, row_index, values)
    _style_header_row(ws, len(headers))
    _autosize(ws, len(headers))


def write_diff_workbook(diff: dict[str, list[dict[str, Any]]], output_path: str) -> None:
    """Write the added/removed/changed sheets of a run-diff result to ``output_path``."""
    wb = Workbook()
    added_ws = wb.active
    added_ws.title = ADDED_SHEET
    _write_records_sheet(added_ws, diff.get("added", []))
    removed_ws = wb.create_sheet(REMOVED_SHEET)
    _write_records_sheet(removed_ws, diff.get("removed", []))
    changed_ws = wb.create_sheet(CHANGED_SHEET)
    _write_changed_sheet(changed_ws, diff.get("changed", []))
    wb.save(output_path)
