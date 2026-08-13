"""SAP sheet writers — extracted from writers/__init__.py (Phase 5D).

Contains the writers for SAP contract history / meter readings / financial
transactions / back-billing sheets, plus the helpers (_write_sap_header_row,
_bb_invoice_value) and SAP color constants they depend on.
"""

from __future__ import annotations

from typing import Any

import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.worksheet import Worksheet

from edf_bill_fetcher.helpers.theme import (
    CELL_BORDER,
    EDF_OFFWHITE,
    MEDIUM_GREY,
)
from edf_bill_fetcher.io.writers.sheet_layout import freeze_at, write_header_row
from edf_bill_fetcher.models.events import SapBackBillingEvent, SapEdfMatch

# ---- SAP color constants (was writers/__init__.py L2003-2009) ----
ORANGE = "FE5716"
NAVY_BLUE = "10367A"
SAP_BB_SUMMARY_FILL_PAIR = ("EFF4FB", "ffffff")
SAP_BB_DETAIL_FILL_PAIR = ("F8FAFC", "ffffff")
SAP_BB_MEDIUM_BORDER = Side(style="medium", color="10367A")


def _bb_invoice_value(rec: dict, key: str) -> object:
    """Look up an invoice field, tolerating 'N/A' / None."""
    v = rec.get(key)
    if v in (None, "", "N/A", "None"):
        return ""
    return v


# ---- _write_sap_header_row (was L1976-2002) ----
def _write_sap_header_row(ws: Worksheet, row: int, columns: list) -> None:
    """Header row — left-aligned variant of the shared helper."""
    write_header_row(ws, row, columns, align="left")


# ---------------------------------------------------------------------------
# SAP Back-billing sheet writers (spec: 2026-07-21-sap-back-billing-analysis-design.md)
# ---------------------------------------------------------------------------
#
# Two adjacent sheets:
# 1. "SAP Back-billing Events" (Sheet 1) — one summary row per SAP
#    Clearing Document cluster, with the underlying SAP rows rendered
#    as collapsed outline groups beneath each summary.
# 2. "SAP ↔ EDF Matched Events" (Sheet 2) — one row per (SAP event ×
#    matched EDF invoice candidate), only High/Medium/Low confidence
#    pairs included.
#
# Hyperlinks cross-reference each other, the source SAP Financial
# Transactions sheet (using the first underlying row's row index),
# and the matched EDF Evidence Report row (using the pre-dedup index
# into ``evidence_rows`` passed by the caller).


# ---- write_sap_contract_history_sheet (IMPL — was L1811-1866) ----
def _write_sap_contract_history_sheet_impl(
    ws: Worksheet,
    rows: list[dict],
    account: str = "",
) -> None:
    """Render the SAP Contract History tab.

    Layout:
      row 1: title banner noting SAP/Kraken origin
      row 2: empty
      row 3: column headers (9 cols)
      rows 4+: one row per contract
    """
    ws.title = "SAP Contract History"
    ORANGE = "FE5716"
    columns = [
        "Contract From",
        "Contract To",
        "Product Code",
        "Product Description",
        "Contract Reason",
        "Set Up By",
        "Notes",
        "Cancelled Flag",
        "Source File",
    ]
    ncol = len(columns)

    title = "EDF SAP CONTRACT AND PRODUCT CHANGE HISTORY"
    if account:
        title = f"{title}  |  Account {account}"
    t1 = ws.cell(row=1, column=1, value=title)
    t1.font = Font(name="Calibri", size=13, bold=True, color="FFFFFF")
    t1.fill = PatternFill("solid", start_color=ORANGE)
    t1.border = CELL_BORDER
    for c in range(2, ncol + 1):
        x = ws.cell(row=1, column=c)
        x.fill = PatternFill("solid", start_color=ORANGE)
        x.border = CELL_BORDER
    ws.row_dimensions[1].height = 22
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncol)

    ws.row_dimensions[2].height = 4

    _write_sap_header_row(ws, row=3, columns=columns)

    for i, row in enumerate(rows):
        r = 4 + i
        for j, col in enumerate(columns):
            cell = ws.cell(row=r, column=j + 1, value=row.get(col, ""))
            cell.font = Font(name="Calibri", size=10)
            cell.border = CELL_BORDER
            if i % 2 == 0:
                cell.fill = PatternFill("solid", start_color=EDF_OFFWHITE.lstrip("#"))


# ---- write_sap_meter_readings_sheet (was L1867-1936) ----
def write_sap_meter_readings_sheet(
    ws: Worksheet,
    rows: list[dict],
    account: str = "",
) -> None:
    """Render the SAP Meter Readings tab.

    Layout:
      row 1: banner
      row 2: legend
      row 3: column headers (9 cols)
      rows 4+: data
    """
    ws.title = "SAP Meter Readings"
    ORANGE = "FE5716"
    columns = [
        "Scheduled Read Date",
        "Meter Read Date",
        "Meter Read Reason",
        "Reading (kWh)",
        "Read Type",
        "Read Source",
        "Read Status",
        "Register",
        "Source File",
    ]
    ncol = len(columns)

    title = "EDF SAP METER-READ HISTORY"
    if account:
        title = f"{title}  |  Account {account}"
    t1 = ws.cell(row=1, column=1, value=title)
    t1.font = Font(name="Calibri", size=13, bold=True, color="FFFFFF")
    t1.fill = PatternFill("solid", start_color=ORANGE)
    t1.border = CELL_BORDER
    for c in range(2, ncol + 1):
        x = ws.cell(row=1, column=c)
        x.fill = PatternFill("solid", start_color=ORANGE)
        x.border = CELL_BORDER
    ws.row_dimensions[1].height = 22
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncol)

    l2 = ws.cell(row=2, column=1, value="A = Actual (supplier-confirmed)  |  E = Estimated")
    l2.font = Font(name="Calibri", size=9, italic=True, color=MEDIUM_GREY.lstrip("#"))
    for c in range(2, ncol + 1):
        ws.cell(row=2, column=c).border = CELL_BORDER

    _write_sap_header_row(ws, row=3, columns=columns)

    for i, row in enumerate(rows):
        r = 4 + i
        for j, col in enumerate(columns):
            c = j + 1
            val = row.get(col, "")
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = Font(name="Calibri", size=10)
            cell.border = CELL_BORDER
            if i % 2 == 0:
                cell.fill = PatternFill("solid", start_color=EDF_OFFWHITE.lstrip("#"))
            if col == "Read Type":
                if val == "A":
                    cell.font = Font(name="Calibri", size=10, bold=True, color="006100")
                    cell.fill = PatternFill("solid", start_color="C6EFCE")
                if val == "E":
                    cell.font = Font(
                        name="Calibri", size=10, italic=True, color=MEDIUM_GREY.lstrip("#")
                    )
                    cell.fill = PatternFill("solid", start_color=EDF_OFFWHITE.lstrip("#"))


# ---- write_sap_financial_transactions_sheet (was L1937-1975) ----
def write_sap_financial_transactions_sheet(
    ws: Worksheet,
    rows: list[dict],
    account: str = "",
) -> None:
    """Render the SAP Financial Transactions tab."""
    ws.title = "SAP Financial Transactions"
    ORANGE = "FE5716"
    columns = list(rows[0].keys()) if rows else ["Source File"]
    ncol = len(columns)

    title = "EDF SAP FINANCIAL TRANSACTIONS"
    if account:
        title = f"{title}  |  Account {account}"
    t1 = ws.cell(row=1, column=1, value=title)
    t1.font = Font(name="Calibri", size=13, bold=True, color="FFFFFF")
    t1.fill = PatternFill("solid", start_color=ORANGE)
    t1.border = CELL_BORDER
    for c in range(2, ncol + 1):
        x = ws.cell(row=1, column=c)
        x.fill = PatternFill("solid", start_color=ORANGE)
        x.border = CELL_BORDER
    ws.row_dimensions[1].height = 22
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncol)

    ws.row_dimensions[2].height = 4

    _write_sap_header_row(ws, row=3, columns=columns)

    for i, row in enumerate(rows):
        r = 4 + i
        for j, col in enumerate(columns):
            cell = ws.cell(row=r, column=j + 1, value=row.get(col, ""))
            cell.font = Font(name="Calibri", size=10)
            cell.border = CELL_BORDER
            if i % 2 == 0:
                cell.fill = PatternFill("solid", start_color=EDF_OFFWHITE.lstrip("#"))


# ---- write_sap_back_billing_sheets (was L2018-2055) ----
def write_sap_back_billing_sheets(
    wb: openpyxl.Workbook,
    events: list[SapBackBillingEvent],
    matches: list[SapEdfMatch],
    sap_financial_first_row: int,
    edf_rows: list[dict],
    edf_sheet_name: str = "EDF Evidence Report",
    edf_first_row: int = 4,
    account: str = "",
    sap_row_index_map: dict[int, int] | None = None,
) -> tuple[Worksheet, Worksheet]:
    """Write both new sheets to the workbook (spec §4).

    Returns ``(sheet1_ws, sheet2_ws)`` so callers can pass them around
    if needed (e.g. the GUI's status line).
    """
    ws1 = wb.create_sheet(title="SAP Back-billing Events")
    ws2 = wb.create_sheet(title="SAP ↔ EDF Matched Events")
    _write_sap_bb_events_sheet(
        ws1,
        events,
        sap_financial_first_row,
        sap_row_index_map=sap_row_index_map,
        account=account,
    )
    _write_sap_bb_matches_sheet(
        ws2,
        events,
        matches,
        ws1,
        edf_rows,
        edf_sheet_name=edf_sheet_name,
        edf_first_row=edf_first_row,
        account=account,
    )
    return ws1, ws2


# ---- _write_sap_bb_events_sheet (was L2056-2250) ----
def _write_sap_bb_events_sheet(
    ws: Worksheet,
    events: list[SapBackBillingEvent],
    sap_financial_first_row: int,
    sap_row_index_map: dict[int, int] | None = None,
    account: str = "",
) -> None:
    """Sheet 1 of the SAP Back-billing pair."""
    ws.title = "SAP Back-billing Events"
    # Outline groups sit ABOVE (summary at top of each cluster).
    ws.sheet_properties.outlinePr.summaryBelow = False

    summary_cols = [
        "Clearing Doc #",
        "Clearing Date",
        "Clearing Reason",
        "# SAP Rows",
        "Net Amount (£)",
        "Has Cr-Credit for Consum Billing?",
        "Has Account Maintenance?",
        "Largest Single Posting (£)",
        "Posting Date Range",
        "Evidence Trail",
        "Matched EDF Invoice #",
        "Link to SAP Financial Transactions",
    ]
    ncol = len(summary_cols)

    # Spec §3.3 — title summary counts (replaces the deleted legal block).
    net_zero_count = sum(1 for ev in events if abs(ev.net_amount) < 1.0)
    with_credit_count = sum(1 for ev in events if ev.has_credit_for_consum_billing)
    event_count = len(events)

    # ---------- rows 1-3: banner + spacer + header ----------
    # Spec §3.3 — the legal-context block (was row 3) and the italic
    # intro paragraph (was row 5) are entirely REMOVED. The title row
    # now carries the event-count summary; header moves from row 7 to
    # row 3; data from row 8 to row 4; freeze panes from A8 to A4.
    title = (
        "EDF SAP BACK-BILLING EVENTS  |  Account {acc}  |  "
        "{n} events ({nz} net-zero, {wc} with credit)"
    ).format(
        acc=account or "(no account)",
        n=event_count,
        nz=net_zero_count,
        wc=with_credit_count,
    )
    t1 = ws.cell(row=1, column=1, value=title)
    t1.font = Font(name="Calibri", size=13, bold=True, color="FFFFFF")
    t1.fill = PatternFill("solid", start_color=ORANGE)
    t1.border = CELL_BORDER
    for c in range(2, ncol + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = PatternFill("solid", start_color=ORANGE)
        cell.border = CELL_BORDER
    ws.row_dimensions[1].height = 22
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncol)

    ws.row_dimensions[2].height = 4

    # Row 3: header (was row 7 per spec §3.3)
    _write_sap_header_row(ws, row=3, columns=summary_cols)

    r = 4  # was 8 — header moved from row 7 to row 3 (spec §3.3)
    # Need to look up the index of each underlying row inside the
    # global sap_financial_rows list to point at the SAP Financial
    # Transactions sheet. The caller passes sap_financial_first_row
    # (typically 4, matching the existing writer's first_row=4).
    # We rely on the events' rows retaining insertion order so we
    # can show the first underlying row's hyperlink.

    for ev_i, ev in enumerate(events):
        # Alternating summary tint per event (option C)
        summary_fill = SAP_BB_SUMMARY_FILL_PAIR[ev_i % 2]
        # ----- summary row -----
        cd_iso = (
            pd.Timestamp(ev.clearing_date).strftime("%Y-%m-%d")
            if not pd.isna(ev.clearing_date)
            else ""
        )
        posting_range = (
            f"{ev.posting_date_range[0]} … {ev.posting_date_range[1]}"
            if ev.posting_date_range[0] and ev.posting_date_range[1]
            else ""
        )
        summary_vals = [
            ev.clearing_doc,
            cd_iso,
            ev.clearing_reason,
            len(ev.rows),
            ev.net_amount,
            "Yes" if ev.has_credit_for_consum_billing else "No",
            "Yes" if ev.has_account_maintenance else "No",
            ev.largest_single_posting,
            posting_range,
            ev.evidence_trail,
            ev.matched_edf_invoice or "",
        ]
        # Write values
        for j, val in enumerate(summary_vals, start=1):
            cell = ws.cell(row=r, column=j, value=val)
            cell.font = Font(name="Calibri", size=10, bold=True)
            cell.border = CELL_BORDER
            cell.fill = PatternFill("solid", start_color=summary_fill)
            cell.alignment = Alignment(
                horizontal="left" if j == 10 else "center", vertical="top", wrap_text=True
            )
        # Cluster-unmatched tags (set by handle_cluster_unmatched in the
        # export pipeline) are rendered italic + grey in the Matched EDF
        # Invoice # column (col 11) to distinguish them from real matches.
        if getattr(ev, "_cluster_unmatched_tag", None) is not None:
            tag_cell = ws.cell(row=r, column=11)
            tag_cell.font = Font(name="Calibri", size=10, bold=False, italic=True, color="666666")
        # Last column: hyperlink (added below once we know target row)
        # Add medium top border on the first event visible in a band
        # — actually we want a top border on every summary row to
        # visually separate events.
        for j in range(1, ncol + 1):
            cell = ws.cell(row=r, column=j)
            existing = cell.border
            cell.border = Border(
                left=existing.left,
                right=existing.right,
                bottom=existing.bottom,
                top=SAP_BB_MEDIUM_BORDER,
            )
        summary_row_idx = r
        r += 1

        # ----- underlying rows (outline level 1, hidden by default) -----
        detail_fill_pair_idx = ev_i % 2
        for k, row_dict in enumerate(ev.rows):
            # Mirror the visible columns of the SAP Financial Transactions sheet.
            # We pick a reduced column set that mirrors what readers will recognise.
            detail_cols = [
                "Document No.",
                "Item",
                "Document Date",
                "Posting Date",
                "Net Due Date",
                "Main Transaction",
                "Sub Transaction",
                "Transaction Text",
                "Amount",
                "Clearing Status",
                "Clearing Document",
                "Clearing Date",
                "Clearing Reason",
                "Document Type",
                "Document Type Description",
            ]
            # Pad with blanks where this row's columns run out vs ncol header.
            for j, col_name in enumerate(detail_cols, start=1):
                v = row_dict.get(col_name, "")
                cell = ws.cell(row=r, column=j, value=v)
                cell.font = Font(name="Calibri", size=9, color="333333")
                cell.fill = PatternFill(
                    "solid",
                    start_color=SAP_BB_DETAIL_FILL_PAIR[(detail_fill_pair_idx + k) % 2],
                )
            # Blank-fill remaining cells so the row visually ends at the
            # right margin (otherwise alternating-band widths match the
            # summary band, which has more columns than detail).
            for j in range(len(detail_cols) + 1, ncol + 1):
                cell = ws.cell(row=r, column=j)
                cell.fill = PatternFill(
                    "solid",
                    start_color=SAP_BB_DETAIL_FILL_PAIR[(detail_fill_pair_idx + k) % 2],
                )
            ws.row_dimensions[r].outline_level = 1
            ws.row_dimensions[r].hidden = True
            r += 1
        # Link to the first underlying row on the SAP Financial Transactions sheet,
        # using the actual Excel row of that underlying row (via ``id(row)`` lookup
        # into the caller-provided ``sap_row_index_map``).
        first_doc = (ev.rows[0].get("Document No.") or "") if ev.rows else ""
        cell = ws.cell(row=summary_row_idx, column=ncol, value="→")
        cell.font = Font(name="Calibri", size=10, color="0563C1", underline="single")
        cell.alignment = Alignment(horizontal="center", vertical="top")
        if sap_row_index_map is not None and ev.rows:
            target_excel_row = sap_row_index_map.get(id(ev.rows[0]), sap_financial_first_row)
        else:
            target_excel_row = sap_financial_first_row
        summary_cell_location = f"'SAP Financial Transactions'!A{target_excel_row}"
        cell.hyperlink = openpyxl.worksheet.hyperlink.Hyperlink(
            ref=cell.coordinate,
            location=summary_cell_location,
            display="→",
            tooltip=f"Jump to SAP Financial Transactions row {target_excel_row} (DOC {first_doc})",
        )

    # Widths
    widths = [18, 12, 28, 9, 13, 14, 14, 14, 22, 60, 18, 20]
    for i, w in enumerate(widths, start=1):
        col_letter = openpyxl.utils.get_column_letter(i)
        ws.column_dimensions[col_letter].width = w

    # Freeze top header
    freeze_at(ws, "A4")  # was A8 — header moved to row 3 (spec §3.3)


# ---- _write_sap_bb_matches_sheet (was L2251-2448) ----
def _write_sap_bb_matches_sheet(
    ws: Worksheet,
    events: list[SapBackBillingEvent],
    matches: list[SapEdfMatch],
    sheet1: Worksheet,
    edf_rows: list[dict],
    edf_sheet_name: str = "EDF Evidence Report",
    edf_first_row: int = 4,
    account: str = "",
) -> None:
    """Sheet 2 of the SAP Back-billing pair."""
    ws.title = "SAP ↔ EDF Matched Events"

    header_cols = [
        "SAP Clearing Doc #",
        "SAP Clearing Date",
        "SAP Event Net Amount (£)",
        "EDF Invoice #",
        "EDF Bill Date",
        "EDF Period From → To",
        "EDF Invoice Amount (£)",
        "Amount Δ (£)",
        "Date Δ (days)",
        "Match Confidence",
        "Has Cr-Credit for Consum Billing?",
        "Notes",
    ]
    ncol = len(header_cols)

    # Rows 1-3: banner + intro
    title = "SAP ↔ EDF MATCHED EVENTS"
    if account:
        title = f"{title}  |  Account {account}"
    t1 = ws.cell(row=1, column=1, value=title)
    t1.font = Font(name="Calibri", size=13, bold=True, color="FFFFFF")
    t1.fill = PatternFill("solid", start_color=ORANGE)
    t1.border = CELL_BORDER
    for c in range(2, ncol + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = PatternFill("solid", start_color=ORANGE)
        cell.border = CELL_BORDER
    ws.row_dimensions[1].height = 22
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncol)

    ws.row_dimensions[2].height = 4

    intro = (
        "Each row pairs a SAP back-billing event (sheet 'SAP Back-billing\n"
        "Events') with the EDF invoice(s) whose period overlaps the\n"
        "event's clear date. Confidence is computed from date proximity\n"
        "and amount agreement; Low-confidence rows may be coincidental.\n"
        "Click the SAP Clearing Doc link to view the event's underlying\n"
        "SAP rows on sheet 'SAP Back-billing Events'. Click the EDF\n"
        "Invoice # link to view the matched row on the EDF Evidence\n"
        "Report."
    )
    intro_cell = ws.cell(row=3, column=1, value=intro)
    intro_cell.font = Font(name="Calibri", size=10, italic=True)
    intro_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=ncol)
    ws.row_dimensions[3].height = 90

    # Row 5: header
    _write_sap_header_row(ws, row=5, columns=header_cols)

    # Sort matches: by Clearing Date asc, then by score desc within date.
    def sort_key(m: SapEdfMatch) -> tuple:
        cd = m.event.clearing_date
        cd_iso = pd.Timestamp(cd).strftime("%Y-%m-%d") if not pd.isna(cd) else "9999"
        # Score desc → negate.
        conf_order = {"High": 0, "Medium": 1, "Low": 2}.get(m.confidence_band, 3)
        return (cd_iso, conf_order, -m.confidence_score)

    sorted_matches = sorted(matches, key=sort_key)

    # Map Clearing Doc -> summary row on Sheet 1 so we can link back.
    # Summary row index = 4 + sum_{j<i}(1 + len(events[j].rows))
    # (summary row + N underlying rows per event). Header row is row 3
    # and the first summary row is row 4 per spec §3.3.
    summary_row_for_doc: dict[str, int] = {}
    sheet1_summary_row = 4
    for ev in events:
        summary_row_for_doc[ev.clearing_doc] = sheet1_summary_row
        sheet1_summary_row += 1 + len(ev.rows)

    confidence_band_fill = {
        "High": "E5F5E5",  # pale green
        "Medium": "FFF4D6",  # pale amber
        "Low": "F2EAEA",  # pale grey-pink
    }

    r = 6
    for m in sorted_matches:
        ev = m.event
        rec = m.edf_record
        # Build EDF period From→To display
        pf = rec.get("Period From")
        pt = rec.get("Period To")
        period_span = ""
        if pf and pt:
            period_span = f"{pf} → {pt}"
        elif pt:
            period_span = f"? → {pt}"
        elif pf:
            period_span = f"{pf} → ?"
        cd_iso = (
            pd.Timestamp(ev.clearing_date).strftime("%Y-%m-%d")
            if not pd.isna(ev.clearing_date)
            else ""
        )
        edf_amt_val = rec.get("Amount (£)")
        vals = [
            ev.clearing_doc,
            cd_iso,
            ev.net_amount,
            rec.get("Invoice #", ""),
            rec.get("Date", ""),
            period_span,
            edf_amt_val if edf_amt_val not in (None, "") else "",
            m.amount_delta,
            m.date_delta_days,
            m.confidence_band,
            "Yes" if ev.has_credit_for_consum_billing else "No",
            m.notes,
        ]
        row_fill = confidence_band_fill.get(m.confidence_band, "")
        for j, v in enumerate(vals, start=1):
            cell = ws.cell(row=r, column=j, value=v)
            cell.font = Font(name="Calibri", size=10)
            cell.border = CELL_BORDER
            if row_fill:
                cell.fill = PatternFill("solid", start_color=row_fill)
            cell.alignment = Alignment(
                horizontal="left" if j in (6, 12) else "center",
                vertical="center",
                wrap_text=(j == 12),
            )
        # Hyperlinks:
        # col 1 (SAP Clearing Doc #) -> Sheet 1 summary row for this event
        sap_target_row = summary_row_for_doc.get(ev.clearing_doc, 8)
        c1 = ws.cell(row=r, column=1)
        summary_loc = f"'SAP Back-billing Events'!A{sap_target_row}"
        c1.hyperlink = openpyxl.worksheet.hyperlink.Hyperlink(
            ref=c1.coordinate,
            location=summary_loc,
            display=ev.clearing_doc,
            tooltip=f"Jump to SAP event summary row {sap_target_row}",
        )
        c1.font = Font(name="Calibri", size=10, color="0563C1", underline="single")
        # col 4 (EDF Invoice #) -> EDF Evidence Report row for this invoice
        # The row index on EDF Evidence Report = edf_first_row + position in edf_rows.
        # Need to find the index. We use the edf_records list (which match_sap_events_to_edf
        # was given) — the caller has passed the same list as edf_rows.
        edf_idx = None
        for i, er in enumerate(edf_rows):
            if er is rec:
                edf_idx = i
                break
        if edf_idx is None:
            # Try matching by Invoice # as a fallback
            target_inv = str(rec.get("Invoice #", "")).strip()
            if target_inv:
                for i, er in enumerate(edf_rows):
                    if str(er.get("Invoice #", "")).strip() == target_inv:
                        edf_idx = i
                        break
        if edf_idx is not None:
            edf_target_row = edf_first_row + edf_idx
            c4 = ws.cell(row=r, column=4)
            c4_loc = f"'{edf_sheet_name}'!A{edf_target_row}"
            c4.hyperlink = openpyxl.worksheet.hyperlink.Hyperlink(
                ref=c4.coordinate,
                location=c4_loc,
                display=str(rec.get("Invoice #", "")),
                tooltip=f"Jump to {edf_sheet_name} row {edf_target_row}",
            )
            c4.font = Font(name="Calibri", size=10, color="0563C1", underline="single")
        r += 1

    # Column widths
    widths = [18, 14, 18, 20, 14, 30, 18, 14, 12, 16, 16, 50]
    for i, w in enumerate(widths, start=1):
        col_letter = openpyxl.utils.get_column_letter(i)
        ws.column_dimensions[col_letter].width = w
    freeze_at(ws, "A6")


# ---------------------------------------------------------------------------
# Cross-source reconciliation sheet writer
# ---------------------------------------------------------------------------
# Compares rows from the three SAP dump writers against the inferred analyser
# tables (Contract History / Meter Readings) and the EDF Evidence Report, line
# by line, with one Matched/Discrepancy/Missing row per comparison. Each
# matched row carries an openpyxl Hyperlink whose ``location`` points at the
# row on the source sheet that owns the matched side, so a reviewer can jump
# straight from a Discrepancy on the Reconciliation tab to the underlying row.


# ---- Public adapter for write_sap_contract_history_sheet ----
def write_sap_contract_history_sheet(
    ws: Worksheet,
    df_or_rows: pd.DataFrame | list[dict[str, Any]],
    account: str = "",
) -> None:
    """Adapter: test contract uses ``(ws, df)``; convert DataFrame to rows."""
    rows = (
        df_or_rows.to_dict(orient="records") if isinstance(df_or_rows, pd.DataFrame) else df_or_rows
    )
    return _write_sap_contract_history_sheet_impl(ws, rows, account)


# ---- write_sap_back_billing_position_sheet (Task 8) ----
def write_sap_back_billing_position_sheet(
    wb: openpyxl.Workbook,
    result: dict,
    account: str = "",
) -> Worksheet:
    """Render the 'Backbilling According to SAP' cross-referenced position.

    Three sections: title banner (with event-count summary), the SAP
    back-billing events table (reversal-containing clusters), and the
    reconciliation table against our PDF-derived Back-billing Analysis.
    """
    ws = wb.create_sheet(title="Backbilling According to SAP")
    ORANGE = "FE5716"
    NAVY = "10367A"

    summary = result.get("summary", {})
    title = (
        "BACKBILLING ACCORDING TO SAP  |  Account {acc}  |  "
        "{n} event(s)  |  SAP net total £{net:,.2f}  |  "
        "{rec} reconciled"
    ).format(
        acc=account or "(no account)",
        n=summary.get("sap_events", 0),
        net=summary.get("sap_net_total", 0.0),
        rec=summary.get("reconciled", 0),
    )
    _write_sap_header_row(ws, 1, [title])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
    c1 = ws.cell(row=1, column=1)
    c1.font = Font(name="Calibri", size=13, bold=True, color="FFFFFF")
    c1.fill = PatternFill("solid", start_color=ORANGE)

    events = result.get("events", [])
    ev_cols = [
        "Clearing Doc #",
        "Clearing Date",
        "Clearing Reason",
        "# Rows",
        "Net Amount (£)",
        "Has Credit for Consum Billing",
        "Period(s)",
        "Matched EDF Invoice #",
    ]
    _write_sap_header_row(ws, 3, ev_cols)
    r = 4
    for i, ev in enumerate(events):
        for j, col in enumerate(ev_cols, start=1):
            cell = ws.cell(row=r, column=j, value=ev.get(col, ""))
            cell.font = Font(name="Calibri", size=10)
            cell.border = CELL_BORDER
            if i % 2 == 0:
                cell.fill = PatternFill("solid", start_color=SAP_BB_SUMMARY_FILL_PAIR[0])
        r += 1

    r += 1
    rec_cols = [
        "SAP Event",
        "EDF Invoice #",
        "EDF Unlawful Charge (£)",
        "SAP Net (£)",
        "Verdict",
    ]
    _write_sap_header_row(ws, r, rec_cols)
    r += 1
    for i, rec in enumerate(result.get("reconciliation", [])):
        for j, col in enumerate(rec_cols, start=1):
            cell = ws.cell(row=r, column=j, value=rec.get(col, ""))
            cell.font = Font(name="Calibri", size=10)
            cell.border = CELL_BORDER
            if i % 2 == 0:
                cell.fill = PatternFill("solid", start_color=SAP_BB_DETAIL_FILL_PAIR[0])
        r += 1

    for idx, width in enumerate((18, 14, 22, 10, 16, 26, 40, 20), start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = width
    return ws


__all__ = [
    "_bb_invoice_value",
    "_write_sap_bb_events_sheet",
    "_write_sap_bb_matches_sheet",
    "_write_sap_header_row",
    "_write_sap_contract_history_sheet_impl",
    "write_sap_back_billing_position_sheet",
    "write_sap_back_billing_sheets",
    "write_sap_contract_history_sheet",
    "write_sap_financial_transactions_sheet",
    "write_sap_meter_readings_sheet",
]
