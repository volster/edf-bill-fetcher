"""Shared row-layout helpers for the Excel sheet writers.

Encapsulates the banner / section-label / merged-text / header-row /
trailing-total / freeze patterns that every ``write_*_sheet`` in this
package previously hand-rolled, so the rendered workbooks stay
identical while the call sites shrink.

Cell-level primitives (``hcell``, ``text``, ``num``, ``money``,
``section_hdr``, ``set_column_widths_from_spec``) live in
``helpers/excel_utils.py``; this module composes them into the
recurring row layouts.
"""

from __future__ import annotations

from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from edf_bill_fetcher.helpers.theme import CELL_BORDER


def write_banner(
    ws: Worksheet,
    title: str,
    ncols: int,
    color: str = "FE5716",
    row: int = 1,
    height: int | None = None,
    size: int = 13,
) -> None:
    """Write a title banner spanning ``ncols`` columns at ``row``.

    Row 1 of ``write_back_billing_sheet`` (ORANGE title banner) and the
    payment writer's row 1 are the canonical examples.  The title cell
    is bold white on the fill; the remaining columns carry the same
    fill + border so the bar reads as one unit.
    """
    t = ws.cell(row=row, column=1, value=title)
    t.font = Font(name="Calibri", size=size, bold=True, color="FFFFFF")
    t.fill = PatternFill("solid", start_color=color)
    t.border = CELL_BORDER
    t.alignment = Alignment(horizontal="left", vertical="center")
    for c in range(2, ncols + 1):
        x = ws.cell(row=row, column=c)
        x.fill = PatternFill("solid", start_color=color)
        x.border = CELL_BORDER
    if height is not None:
        ws.row_dimensions[row].height = height


def write_section_label(
    ws: Worksheet,
    row: int,
    label: str,
    ncols: int,
    bg: str = "10367A",
    size: int = 11,
) -> None:
    """Write a section label spanning ``ncols`` columns at ``row``.

    Row 2 'LEGAL CONTEXT' of ``write_back_billing_sheet`` (NAVY,
    size 11) is the canonical example.  Note this is the size-11
    treatment, distinct from ``excel_utils.section_hdr`` (size 10,
    used by the payment writer) — the two sheet families genuinely
    differ, so both helpers stay.
    """
    cell = ws.cell(row=row, column=1, value=label)
    cell.font = Font(name="Calibri", size=size, bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", start_color=bg)
    cell.border = CELL_BORDER
    cell.alignment = Alignment(horizontal="left", vertical="center")
    for c in range(2, ncols + 1):
        x = ws.cell(row=row, column=c)
        x.fill = PatternFill("solid", start_color=bg)
        x.border = CELL_BORDER


def write_merged_text(
    ws: Worksheet,
    row: int,
    text_value: str,
    ncols: int,
    height: int | None = None,
    italic: bool = False,
    border: bool = True,
) -> None:
    """Write a merged, wrapped text cell spanning ``ncols`` at ``row``.

    Rows 3 (legal context) and 5 (instruction) of
    ``write_back_billing_sheet`` are the canonical examples.
    """
    cell = ws.cell(row=row, column=1, value=text_value)
    cell.font = Font(name="Calibri", size=10, italic=italic)
    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    if border:
        cell.border = CELL_BORDER
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    if height is not None:
        ws.row_dimensions[row].height = height


def write_header_row(
    ws: Worksheet,
    row: int,
    headers: list[str],
    bg: str = "10367A",
    height: int | None = None,
    align: str = "center",
) -> None:
    """Write a header row with bold white text on ``bg``.

    Absorbs ``_write_sap_header_row`` (io/writers/sap.py:41, which is
    **left**-aligned — pass ``align="left"``) and the ``_hcell``-based
    header loops in the back-billing / evidence / payment writers
    (**center**-aligned — the default).  Row 7 of
    ``write_back_billing_sheet`` is the canonical center example.
    """
    for j, col in enumerate(headers):
        cell = ws.cell(row=row, column=j + 1, value=col)
        cell.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", start_color=bg)
        cell.border = CELL_BORDER
        cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    if height is not None:
        ws.row_dimensions[row].height = height


def write_trailing_total(
    ws: Worksheet,
    row: int,
    label: str,
    values: list[tuple[int, float]],
    label_span: int,
    ncols: int,
    bg: str = "10367A",
    fmt: str = "#,##0.00",
) -> None:
    """Write a NAVY trailing-total row.

    The 'TOTAL RETROSPECTIVE CHARGES…' row of
    ``write_back_billing_sheet`` is the canonical example: the label
    spans ``label_span`` columns starting at column 1, and ``values``
    lists ``(column, amount)`` pairs for the value cell(s) — the
    back-billing total has two (Period Charge at col 6, Unlawful
    Charge at col 10).  Every non-label, non-value column carries the
    fill so the row reads as one unit.
    """
    value_cols = {col for col, _ in values}
    label_cell = ws.cell(row=row, column=1, value=label)
    label_cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    label_cell.fill = PatternFill("solid", start_color=bg)
    label_cell.border = CELL_BORDER
    label_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=label_span)
    # Fill EVERY non-value column (col 1 is the label cell — re-setting
    # its fill/border is harmless since the style is identical — and
    # cols 2..label_span sit inside the merged label range, matching
    # the original back-billing total which filled cols 2-5 and 7-17).
    for c in range(1, ncols + 1):
        if c in value_cols:
            continue
        x = ws.cell(row=row, column=c)
        x.fill = PatternFill("solid", start_color=bg)
        x.border = CELL_BORDER
    for col, amount in values:
        vc = ws.cell(row=row, column=col, value=amount)
        vc.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        vc.fill = PatternFill("solid", start_color=bg)
        vc.border = CELL_BORDER
        vc.number_format = fmt
    ws.row_dimensions[row].height = 22


def freeze_at(ws: Worksheet, cell: str) -> None:
    """Freeze panes at ``cell`` (e.g. ``"A8"``, ``"A2"``)."""
    ws.freeze_panes = cell


__all__ = [
    "write_banner",
    "write_section_label",
    "write_merged_text",
    "write_header_row",
    "write_trailing_total",
    "freeze_at",
]
