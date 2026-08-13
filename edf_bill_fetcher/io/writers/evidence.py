"""Evidence sheet + summary sheet writers for the EDF workbook.

Extracted from ``edf_bill_fetcher/writers/__init__.py`` (Phase 5A) during the
modularization refactor.  These two functions are the heart of the Evidence tab
and Summary tab.
"""

from __future__ import annotations

import openpyxl
import pandas as pd
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# Re-export shared helpers / theme constants used by these two functions.
from edf_bill_fetcher.helpers.date_utils import to_excel_date
from edf_bill_fetcher.helpers.excel_utils import (
    hcell as _hcell,
)
from edf_bill_fetcher.helpers.excel_utils import (
    set_column_widths_from_spec,
)
from edf_bill_fetcher.helpers.theme import CELL_BORDER, DUP_GREY
from edf_bill_fetcher.io.writers.sheet_layout import (
    freeze_at,
    write_header_row,
)
from edf_bill_fetcher.writers._helpers import EST_YELLOW, JUMP_RED

# Column letter map for the evidence sheet (matches ``EVIDENCE_HEADERS`` below):
# A=Source B=Sender C=Date D=PeriodFrom E=PeriodTo F=Invoice#
# G=Amount H=PeriodCharge I=UnitRate J=%Change K=EntryType
# L=Reading M=Units N=StandingChg O=Tariff P=AttachmentName
# Q=Details R=LogicUsed S=AnomalyFlag
# (Duplicate-of-link cells are rendered in a post-loop pass for
# the ``is_duplicate=True`` branch and don't appear in this
# header list.)
#
# F1 (SEV-1):  every COL_* is derived from the headers list, not
# hard-coded.  Inserting a new column at any position requires
# updating exactly one place (the headers list) — the conditional
# formatting range, formula references, column widths and the
# dedup hyperlink pass all read the same index.  Verified by
# ``tests/test_evidence_sheet_columns.py``.
EVIDENCE_HEADERS: list[str] = [
    "Source",
    "Sender",
    "Date",
    "Period From",
    "Period To",
    "Invoice #",
    "Amount (£)",
    "Period Charge (£)",
    "Unit Rate (p/kWh)",
    "% Change",
    "Entry Type",
    "Reading",
    "Units (kWh)",
    "Standing Chg (p/day)",
    # Tariff price-plan name (e.g. "Freedom", "Standard");
    # extracted by ``extract_new_invoice_fields`` on KI-style
    # bills. See ``_process_new_invoice``.
    "Tariff",
    "Attachment Name",
    "Details",
    "Logic Used",
    "Anomaly Flag",
    "Sub Periods",
]
COL_AMOUNT = EVIDENCE_HEADERS.index("Amount (£)") + 1

# --- function body 1: write_evidence_sheet (was writers/__init__.py L182-422) ---


def write_evidence_sheet(ws, df, is_duplicate=False):
    """Render the EDF Evidence Report (or Duplicate Entries) worksheet."""
    headers = EVIDENCE_HEADERS
    COL_PERIOD_CHG = headers.index("Period Charge (£)") + 1
    COL_UNIT_RATE = headers.index("Unit Rate (p/kWh)") + 1
    COL_PCT_CHANGE = headers.index("% Change") + 1
    COL_READING_IDX = headers.index("Reading") + 1
    COL_ANOMALY = headers.index("Anomaly Flag") + 1
    # Phase 2 follow-on: dup sheets carry a "Duplicate Of"
    # printable summary cell per row plus a clickable hyperlink
    # back to the matched kept record in ``EDF Evidence Report``.
    # The matched-against position lands in a parallel
    # ``_matches_kept_idx`` Series the caller passes alongside the
    # dup_df — we render the column in a *post-loop* pass below so
    # we don't have to count on the row-iteration matching a
    # fixed column index (which previously conflicted with the
    # constant ``COL_ANOMALY = 18`` in the writer).
    has_match_col = "Duplicate Of" in df.columns
    # Capture the writer-helper ``_matches_kept_idx`` Series *before*
    # the row iteration so the post-loop pass can mint HYPERLINK
    # cells.  We then strip the column from the in-scope ``df``
    # so row iteration only sees the reader-facing schema (no
    # 20th column leaks into the saved workbook).
    if "_matches_kept_idx" in df.columns:
        match_positions_series: pd.Series = df["_matches_kept_idx"].copy()
    else:
        match_positions_series = None
    df = df.drop(columns=["_matches_kept_idx"], errors="ignore")
    bg = "888888" if is_duplicate else "FE5716"
    write_header_row(ws, 1, headers, bg=bg, height=28)

    alt_fill = PatternFill("solid", start_color="FFF3EE")

    last_data_row = len(df) + 1
    for r_idx, row in enumerate(df.values, 2):
        row_fill = alt_fill if r_idx % 2 == 0 else PatternFill()

        for c_idx, val in enumerate(row, 1):
            if c_idx == COL_PCT_CHANGE and not is_duplicate:
                # % Change as live formula — Amount is col G (derived
                # from COL_AMOUNT), not col E (Period To serials).
                amt_col_letter = get_column_letter(COL_AMOUNT)
                c = ws.cell(
                    row=r_idx,
                    column=COL_PCT_CHANGE,
                    value=(
                        f"=IFERROR(({amt_col_letter}{r_idx}-"
                        f"{amt_col_letter}{r_idx - 1})/"
                        f'{amt_col_letter}{r_idx - 1},"")'
                    ),
                )
                c.number_format = "0.0%"
                c.alignment = Alignment(horizontal="right", vertical="top")
                c.font = Font(name="Calibri", size=10)
                c.border = CELL_BORDER
                c.fill = row_fill
            else:
                # Convert date columns to real Excel date serials (C=3, D=4, E=5)
                excel_val = val
                if c_idx in (3, 4, 5):
                    dt = to_excel_date(val)
                    if dt is not None:
                        excel_val = dt
                c = ws.cell(row=r_idx, column=c_idx, value=excel_val)
                # Phase 2.x — formula-injection guard on the
                # generic evidence-sheet row path.  openpyxl
                # auto-promotes any cell whose text value starts
                # with ``=``/``+``/``-``/``@`` to ``data_type='f'``
                # (formula).  Without this fix, a bill whose
                # Invoice # or Details field begins with ``=cmd
                # |'/c calc'!A1`` would render as a real formula
                # when an ombudsman opens the workbook.  Same
                # belt-and-braces policy as ``_text``: coerce
                # textual leads to ``str`` first, then pin
                # ``data_type='s'`` and prefix apostrophe on
                # leading special chars.
                if isinstance(excel_val, str) and excel_val:
                    safe_val = excel_val
                    if safe_val[0] in "+-=@":
                        safe_val = "'" + safe_val
                    c.value = safe_val
                    c.data_type = "s"
                if c_idx == COL_AMOUNT and isinstance(val, int | float):
                    c.number_format = "£#,##0.00"
                if c_idx == COL_PERIOD_CHG and isinstance(val, int | float):
                    c.number_format = "£#,##0.00"
                if c_idx == COL_UNIT_RATE and isinstance(val, int | float):
                    c.number_format = "0.00"
                if c_idx in (3, 4, 5) and hasattr(excel_val, "year"):
                    c.number_format = "dd/mm/yyyy"
                c.font = Font(name="Calibri", size=10)
                c.fill = (
                    row_fill if not is_duplicate else PatternFill("solid", start_color=DUP_GREY)
                )
                c.border = CELL_BORDER
                c.alignment = Alignment(vertical="top")

            # Highlight estimated readings (Reading is col L = 0-based index 11)
            if (
                not is_duplicate
                and len(row) > COL_READING_IDX
                and row[COL_READING_IDX] == "Estimated"
            ):
                c.fill = PatternFill("solid", start_color=EST_YELLOW)

        # Anomaly flag col S (19) — Amount is col G (derived from
        # COL_AMOUNT), not col E (Period To serials).  Anomaly Flag
        # shifted right by one when the Tariff column was inserted at
        # column O; see the column-letter map at the top of this module.
        if not is_duplicate and r_idx > 2:
            amt_col_letter = get_column_letter(COL_AMOUNT)
            ca = ws.cell(
                row=r_idx,
                column=COL_ANOMALY,
                value=(
                    f"=IF(AND({amt_col_letter}{r_idx - 1}>0,"
                    f"{amt_col_letter}{r_idx}>{amt_col_letter}{r_idx - 1}*2),"
                    '"⚠ >100% INCREASE","")'
                ),
            )
            ca.font = Font(name="Calibri", size=10, bold=True)
            ca.border = CELL_BORDER
            ca.fill = row_fill

    # Conditional formatting: only colour anomaly column red when non-empty
    if not is_duplicate and last_data_row > 2:
        anomaly_col_letter = get_column_letter(COL_ANOMALY)
        ws.conditional_formatting.add(
            f"{anomaly_col_letter}2:{anomaly_col_letter}{last_data_row}",
            FormulaRule(
                formula=[f'${anomaly_col_letter}2<>""'],
                fill=PatternFill("solid", start_color=JUMP_RED),
                font=Font(name="Calibri", size=10, bold=True),
            ),
        )

    # Phase 2 follow-on: post-loop pass to render the "Duplicate
    # Of" column.  The matched-against keystrokes live in
    # ``match_positions_series`` (a pd.Series keyed on the dup
    # sheet's df-index by df-positional index) so the click-through
    # target always aligns with the writer's row indexing scheme.
    # We render this ``Duplicate Of`` column only when
    # ``is_duplicate`` is True — main evidence reports never get
    # one.
    if is_duplicate and has_match_col and match_positions_series is not None:
        last_data_row = len(df) + 1
        col_idx_duplicate_of = len(headers) + 1
        # Header cell
        bg = "888888"
        _hcell(ws, 1, col_idx_duplicate_of, "Duplicate Of", bg=bg)
        # Materialise columns once
        dup_text = df["Duplicate Of"].tolist()
        for r_idx, (match_val, summary) in enumerate(
            zip(match_positions_series.tolist(), dup_text, strict=True), 2
        ):
            target_row_excel: int | None = None
            try:
                # ``-1`` sentinel from the dedup walker = no
                # match (Pass 1 dedup found a duplicate tuple
                # but Pass 2's kept set dropped it before the
                # post-reset lookup fired).
                mi = int(match_val)
                target_row_excel = mi + 2 if mi >= 0 else None
            except (TypeError, ValueError):
                target_row_excel = None
            if not summary:
                continue
            c = ws.cell(row=r_idx, column=col_idx_duplicate_of, value=summary)
            if target_row_excel:
                c.hyperlink = openpyxl.worksheet.hyperlink.Hyperlink(
                    ref=f"{c.coordinate}",
                    location=f"'EDF Evidence Report'!A{target_row_excel}",
                    display=summary,
                    tooltip=(f"Jump to the kept record at EDF Evidence Report!A{target_row_excel}"),
                )
                c.font = Font(name="Calibri", size=10, color="0000FF", underline="single")
            else:
                c.font = Font(name="Calibri", size=10)
            c.alignment = Alignment(vertical="top", wrap_text=True)
            c.border = CELL_BORDER
            # Dup cells read like the rest of the dup sheet
            # (greyed out so they stand out from the kept set).
            c.fill = PatternFill("solid", start_color=DUP_GREY)
        # Widen the column to fit the longest summary. After the
        # Tariff insertion, "Duplicate Of" lives at column T (was S).
        ws.column_dimensions["T"].width = 50

    widths: dict[str, float] = {
        "A": 18,
        "B": 26,
        "C": 13,
        "D": 13,
        "E": 13,
        "F": 16,
        "G": 13,
        "H": 15,
        "I": 15,
        "J": 10,
        "K": 14,
        "L": 11,
        "M": 12,
        "N": 18,
        # Tariff price-plan column — short enough to fit
        # "Standard Variable", "Freedom Tariff", etc.
        "O": 22,
        "P": 28,
        "Q": 38,
        "R": 18,
        "S": 20,
    }
    set_column_widths_from_spec(ws, widths)
    freeze_at(ws, "A2")


# ---------------------------------------------------------------------------
# Write summary sheet — uses _xlfn.MAXIFS/_xlfn.MINIFS so Excel evaluates
# on load without the dynamic-array compatibility dialog
# ---------------------------------------------------------------------------

# --- function body 2: write_summary_sheet (was writers/__init__.py L431-519) ---


def write_summary_sheet(ws, years, evidence_sheet_name, last_data_row=5000):
    """Render the Annual Summary worksheet with per-year balance aggregates."""
    ws.title = "Annual Summary"

    headers = [
        "Year",
        "Balance Range (£)",
        "Records",
        "Avg Balance (£)",
        "Peak Balance (£)",
        "Lowest Balance (£)",
        "Drill down",
    ]
    for col, h in enumerate(headers, 1):
        _hcell(ws, 1, col, h, bg="10367A")
    ws.row_dimensions[1].height = 28

    alt_fill = PatternFill("solid", start_color="EEF2FF")
    esn = evidence_sheet_name

    date_col = f"'{esn}'!$C$2:$C${last_data_row}"
    # Amount lives in col G of the evidence sheet (not col E, which is
    # Period To); derive it from the shared COL_AMOUNT constant.
    amt_col_letter = get_column_letter(COL_AMOUNT)
    amt_col = f"'{esn}'!${amt_col_letter}$2:${amt_col_letter}${last_data_row}"

    for r_idx, year_val in enumerate(years, 2):
        row_fill = alt_fill if r_idx % 2 == 0 else PatternFill()
        yr_cell = f"A{r_idx}"

        # _xlfn. prefix tells Excel to evaluate MAXIFS/MINIFS on load without
        # the dynamic-array compatibility dialog.
        peak_f = f'=IFERROR(_xlfn.MAXIFS({amt_col},{date_col},">="&DATE({yr_cell},1,1),{date_col},"<"&DATE({yr_cell}+1,1,1)),"")'
        low_f = f'=IFERROR(_xlfn.MINIFS({amt_col},{date_col},">="&DATE({yr_cell},1,1),{date_col},"<"&DATE({yr_cell}+1,1,1)),"")'
        range_f = f'=IFERROR(_xlfn.MAXIFS({amt_col},{date_col},">="&DATE({yr_cell},1,1),{date_col},"<"&DATE({yr_cell}+1,1,1))-_xlfn.MINIFS({amt_col},{date_col},">="&DATE({yr_cell},1,1),{date_col},"<"&DATE({yr_cell}+1,1,1)),"")'

        row_values = [
            int(year_val),
            range_f,
            f'=COUNTIFS({date_col},">="&DATE({yr_cell},1,1),{date_col},"<"&DATE({yr_cell}+1,1,1))',
            f'=IFERROR(AVERAGEIFS({amt_col},{date_col},">="&DATE({yr_cell},1,1),{date_col},"<"&DATE({yr_cell}+1,1,1)),"")',
            peak_f,
            low_f,
            "→ Drill down",
        ]
        for c_idx, val in enumerate(row_values, 1):
            c = ws.cell(row=r_idx, column=c_idx, value=val)
            c.font = Font(name="Calibri", size=10)
            c.fill = row_fill
            c.border = CELL_BORDER
            c.alignment = Alignment(
                horizontal="center" if c_idx == 1 else "right",
                vertical="top",
            )
            if c_idx == 2:
                c.number_format = "£#,##0.00"
            elif c_idx == 3:
                c.number_format = "#,##0"
            elif c_idx > 3:
                c.number_format = "£#,##0.00"
            if c_idx == 7:
                c.hyperlink = openpyxl.worksheet.hyperlink.Hyperlink(
                    ref=c.coordinate,
                    location="'Reconciliation Drill-down'!A2",
                    display="→ Drill down",
                    tooltip="Jump to Reconciliation Drill-down sheet",
                )

    # Grand total row — SUM/MAX/MIN over the year rows only, no dynamic-array functions
    n = len(years) + 2
    first_r = 2
    last_r = n - 1
    tot_fill = PatternFill("solid", start_color="10367A")
    tot_specs = [
        ("OVERALL", None, "center"),
        (f'=IFERROR(MAX(E{first_r}:E{last_r})-MIN(F{first_r}:F{last_r}),"")', "£#,##0.00", "right"),
        (f"=SUM(C{first_r}:C{last_r})", "#,##0", "right"),
        (f'=IFERROR(AVERAGE(D{first_r}:D{last_r}),"")', "£#,##0.00", "right"),
        (f'=IFERROR(MAX(E{first_r}:E{last_r}),"")', "£#,##0.00", "right"),
        (f'=IFERROR(MIN(F{first_r}:F{last_r}),"")', "£#,##0.00", "right"),
    ]
    for c_idx, (val, num_fmt, align) in enumerate(tot_specs, 1):
        c = ws.cell(row=n, column=c_idx, value=val)
        c.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        c.fill = tot_fill
        c.border = CELL_BORDER
        c.alignment = Alignment(horizontal=align)
        if num_fmt:
            c.number_format = num_fmt

    for col_letter in ["A", "B", "C", "D", "E", "F"]:
        ws.column_dimensions[col_letter].width = 22
    ws.freeze_panes = "A2"


# ---------------------------------------------------------------------------
# Main export function
# ---------------------------------------------------------------------------
__all__ = ["write_evidence_sheet", "write_summary_sheet"]
