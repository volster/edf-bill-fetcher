"""Data quality sheet writer — extracted from writers/__init__.py.

Contains: write_data_quality_sheet — renders the "Data Quality" worksheet
with row checks (missing dates, missing amounts, sign mismatch) and summary banner.
"""

from __future__ import annotations

from typing import Any as _Any

from openpyxl.styles import Alignment, Font, PatternFill

from edf_bill_fetcher.helpers.excel_utils import (
    hcell as _hcell,
)
from edf_bill_fetcher.helpers.excel_utils import (
    section_hdr as _section_hdr,
)
from edf_bill_fetcher.helpers.excel_utils import (
    text as _text,
)
from edf_bill_fetcher.helpers.theme import CELL_BORDER
from edf_bill_fetcher.writers._helpers import _data_quality_report

# --- write_data_quality_sheet (was writers/__init__.py L2425-2578) ---


def write_data_quality_sheet(ws, df):
    """Write Data Quality Report tab."""
    ws.title = "Data Quality Report"

    NAVY = "10367A"
    ORANGE = "FE5716"
    LGREY = "F0F0F0"
    DGREY = "888888"

    def _banner(ws, r, text, bg):
        c = ws.cell(row=r, column=1, value=text)
        c.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", start_color=bg)
        c.border = CELL_BORDER
        c.alignment = Alignment(horizontal="left", vertical="center")
        for col in range(2, 6):
            x = ws.cell(row=r, column=col)
            x.fill = PatternFill("solid", start_color=bg)
            x.border = CELL_BORDER
        ws.row_dimensions[r].height = 20

    dq: dict[str, _Any] = _data_quality_report(df)

    if not dq:
        _hcell(ws, 1, 1, "No data to analyze", bg=NAVY)
        ws.column_dimensions["A"].width = 40
        return

    headers = ["Check", "Result", "Rate/Count", "Status"]
    for col, h in enumerate(headers, 1):
        _hcell(ws, 1, col, h, bg=NAVY)
    ws.row_dimensions[1].height = 28

    tc = ws.cell(row=1, column=1, value="EDF ENERGY DISPUTE  —  DATA QUALITY REPORT")
    tc.font = Font(name="Calibri", size=13, bold=True, color="FFFFFF")
    tc.fill = PatternFill("solid", start_color=ORANGE)
    tc.border = CELL_BORDER
    tc.alignment = Alignment(horizontal="left", vertical="center")
    for c in range(2, 5):
        x = ws.cell(row=1, column=c)
        x.fill = PatternFill("solid", start_color=ORANGE)
        x.border = CELL_BORDER

    def _check_row(r, check, result, rate, status, note=""):
        bg = LGREY if r % 2 == 0 else None
        _text(ws, r, 1, check, fill_hex=bg)
        _text(ws, r, 2, str(result), fill_hex=bg)
        _text(ws, r, 3, str(rate), fill_hex=bg)
        _text(ws, r, 4, status, bold=True, fill_hex=bg)
        if note:
            ws.cell(row=r, column=5, value=note).font = Font(name="Calibri", size=9, color=DGREY)

    r = 2
    _section_hdr(ws, r, "COMPLETENESS CHECKS")

    checks = [
        ("Total Records", dq["total_records"], "—", "PASS" if dq["total_records"] > 0 else "FAIL"),
        (
            "Date Parsing Success",
            dq["date_parsed"],
            f"{dq['date_parse_rate']:.1%}",
            "PASS"
            if dq["date_parse_rate"] > 0.8
            else "WARN"
            if dq["date_parse_rate"] > 0.5
            else "FAIL",
        ),
        (
            "Amount Complete",
            dq["amt_complete"],
            f"{(dq['amt_complete'] / dq['total_records']):.1%}",
            "PASS" if dq["amt_complete"] == dq["total_records"] else "WARN",
        ),
        (
            "Period Info Complete",
            dq["period_complete"],
            f"{dq['period_completeness_rate']:.1%}",
            "PASS" if dq["period_completeness_rate"] > 0.7 else "WARN",
        ),
        (
            "Reading Classified",
            dq["reading_classified"],
            f"{dq['reading_classify_rate']:.1%}",
            "PASS" if dq["reading_classify_rate"] > 0.5 else "WARN",
        ),
        (
            "Unit Rate Computable",
            dq["ur_computable"],
            f"{dq['ur_computable_rate']:.1%}",
            "PASS" if dq["ur_computable_rate"] > 0.3 else "INFO",
        ),
    ]
    for check, result, rate, status in checks:
        _check_row(r, check, result, rate, status)
        r += 1

    r += 1
    _section_hdr(ws, r, "DUPLICATION CHECKS")
    r += 1
    _check_row(
        r,
        "Duplicate Records (Date+Amount)",
        dq["duplicate_count"],
        f"{dq['duplicate_rate']:.2%}",
        "PASS"
        if dq["duplicate_rate"] < 0.05
        else "WARN"
        if dq["duplicate_rate"] < 0.15
        else "FAIL",
    )
    r += 1

    r += 1
    _section_hdr(ws, r, "SOURCE DISTRIBUTION")
    for src, cnt in dq.get("source_distribution", {}).items():
        r += 1
        _check_row(r, f"Source: {src}", cnt, f"{cnt / dq['total_records']:.1%}", "INFO")

    r += 1
    _section_hdr(ws, r, "ENTRY TYPE DISTRIBUTION")
    for etype, cnt in dq.get("entry_type_distribution", {}).items():
        r += 1
        _check_row(r, f"Type: {etype}", cnt, f"{cnt / dq['total_records']:.1%}", "INFO")

    # Summary banner
    r += 2
    total_checks = (
        len(checks)
        + 1
        + len(dq.get("source_distribution", {}))
        + len(dq.get("entry_type_distribution", {}))
    )
    pass_count = sum(1 for c in checks if c[3] == "PASS") + (
        1 if dq["duplicate_rate"] < 0.05 else 0
    )
    warn_count = sum(1 for c in checks if c[3] == "WARN") + (
        1 if 0.05 <= dq["duplicate_rate"] < 0.15 else 0
    )
    fail_count = sum(1 for c in checks if c[3] == "FAIL") + (
        1 if dq["duplicate_rate"] >= 0.15 else 0
    )

    _banner(
        ws,
        r,
        f"QUALITY SUMMARY: {total_checks} checks  |  PASS: {pass_count}  |  WARN: {warn_count}  |  FAIL: {fail_count}",
        NAVY,
    )

    for col_letter, width in zip(["A", "B", "C", "D", "E"], [40, 20, 18, 12, 60], strict=False):
        ws.column_dimensions[col_letter].width = width
    ws.freeze_panes = "A2"


__all__ = ["write_data_quality_sheet"]
