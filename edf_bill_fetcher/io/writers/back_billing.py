"""Back-billing analysis writer — extracted from writers/__init__.py.

Contains: write_back_billing_sheet (renders the "Back-billing Analysis"
worksheet).  The pure-pandas detector ``detect_back_billing`` and its
private helpers ``_assess_reason`` / ``_pull_period_charge`` are
re-exported from :mod:`edf_bill_fetcher.processors.detection` (the
canonical home) so there is exactly one definition in the codebase.
"""

from __future__ import annotations

from datetime import datetime

import openpyxl
import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from edf_bill_fetcher.helpers.excel_utils import (
    evidence_report_hyperlink_cell as _evidence_report_hyperlink_cell,
)
from edf_bill_fetcher.helpers.excel_utils import (
    money as _money,
)
from edf_bill_fetcher.helpers.excel_utils import (
    num as _num,
)
from edf_bill_fetcher.helpers.excel_utils import (
    set_column_widths_from_spec,
)
from edf_bill_fetcher.helpers.excel_utils import (
    text as _text,
)
from edf_bill_fetcher.io.adapters.pdf import legal_context
from edf_bill_fetcher.io.writers.sheet_layout import (
    freeze_at,
    write_banner,
    write_header_row,
    write_merged_text,
    write_section_label,
    write_trailing_total,
)

# Re-exported from processors.detection (canonical home) for backwards
# compatibility — existing imports of these names from this module keep
# working, but the pipeline only ever uses the processors.detection copy.
from edf_bill_fetcher.processors.detection import (
    _assess_reason,  # noqa: F401 — re-exported for backwards compatibility
    _pull_period_charge,  # noqa: F401
    detect_back_billing,  # noqa: F401
)
from edf_bill_fetcher.writers._helpers import _disclosed_label

# --- write_back_billing_sheet (was writers/__init__.py L2809-3018) ---


def write_back_billing_sheet(
    ws: Worksheet,
    bb: pd.DataFrame,
    account: str = "",
    overlapping_invoices: set[str] | None = None,
    evidence_df: pd.DataFrame | None = None,
    evidence_index: dict[str, int] | None = None,
    domination_map: dict[str, tuple[str, bool]] | None = None,
    view_superseded_row: dict[str, int] | None = None,
) -> None:
    """Render the Back-billing Analysis tab.

    Layout follows the design spec (§4.1):
      row 1: title banner with SAP account
      row 2: 'LEGAL CONTEXT' section label
      row 3: legal_context() body (one merged paragraph)
      row 4: empty
      row 5: short instruction
      row 6: empty
      row 7: column headers (18 cols incl. Unlawful Charge, Open PDF,
              View on Evidence Report, Status, Superseded By,
              Partial Overlap, View Superseded)
      rows 8+: data rows (sorted by Bill Date as produced by
              :func:`detect_back_billing`)
      trailing: 'TOTAL UNLAWFUL CHARGES — UNION (no double count)'

    The ``Cancel/Rebill Disclosed`` cell (col 11) is the
    :func:`_disclosed_label` value taking the row's
    ``Cancel/Rebill Admitted`` bool AND whether this invoice also
    appears in ``overlapping_invoices`` (a set populated by the
    rebilling detector; defaults to empty).

    Open PDF column (col 13) carries hyperlink
    the first ~400 chars of the source PDF text so a reviewer can
    see why N/A entries were N/A and which regex produced which value.

    ``domination_map`` (from :func:`compute_transitive_domination`)
    maps ``superseded_invoice_id -> (survivor_invoice_id, partial_overlap)``.
    This worksheet renders ONLY the live rows — invoices that appear as a
    KEY in the map are superseded and are excluded here (they move to the
    reconciliation view); every rendered row therefore carries
    ``Status="Live"``.  A live row whose invoice number appears as a
    SURVIVOR value in the map gets a blue-underline ``View superseded``
    cell (col 18) pointing at its chain. ``view_superseded_row`` maps each
    survivor invoice to the Excel row of its ``KILLER:`` group on the
    Superseded Reconciliation sheet; when present, the col-18 cell
    hyperlinks there.  Rows with no chain leave col 18 blank.

    The single trailing total covers ONLY the live rows. ``Period Charge
    (£)`` is summed into col 6; col 10 carries the union of the live rows'
    unlawful consumption days computed by
    :func:`compute_unlawful_union_total` over the live subset, so the same
    consumption is never counted twice.  When every row is superseded (or
    ``bb`` is otherwise empty of live rows) the total row is still written
    with ``£0.00`` values.
    """
    ws.title = "Back-billing Analysis"
    NAVY = "10367A"
    ORANGE = "FE5716"
    overlaps = overlapping_invoices or set()

    # Row 1: banner with account
    title = "BACK-BILLING EVENTS ANALYSIS"
    if account:
        title = f"{title}  |  Account {account}"
    write_banner(ws, title, 17, color=ORANGE, row=1, height=22)

    # Row 2: 'LEGAL CONTEXT' label
    write_section_label(ws, 2, "LEGAL CONTEXT", 17)

    # Row 3: legal_context body (merged across the whole width so the
    # paragraph is readable in one cell).
    lc_text = legal_context()
    write_merged_text(ws, 3, lc_text, 17, height=90)

    # Row 5: instruction
    inst = (
        "Each row identifies an invoice where EDF billed more than 12 "
        "months retrospectively. The Excess Days column shows by how "
        "many days beyond the Standard Licence Condition 7A (SLC 7A) "
        "12-month limit the invoice went. Where a later invoice "
        "cancelled and re-billed an earlier one, the earlier invoice "
        "is excluded from this analysis (see the reconciliation "
        "worksheet) because the same consumption is re-covered by the "
        "surviving invoice; live rows that supersede another carry a "
        "'View superseded' link. The trailing total is the union over "
        "the surviving invoices so the same consumption is not counted "
        "twice."
    )
    write_merged_text(ws, 5, inst, 17, height=70, italic=True, border=False)

    # Row 7: headers
    headers = [
        "Invoice #",
        "Bill Date",
        "Period From",
        "Period To",
        "Days Billed",
        "Period Charge (£)",
        "Value Source",
        "12-Month Limit (days)",
        "Excess Days",
        "Unlawful Charge (£)",
        "Cancel/Rebill Disclosed",
        "Reason Assessment",
        "Open PDF",
        "View on Evidence Report",
        "Status",
        "Superseded By",
        "Partial Overlap",
        "View Superseded",
        "Sub-Period Basis",
    ]
    write_header_row(ws, 7, headers, bg=NAVY, height=28)

    # Data rows + running total
    r = 8
    total = 0.0
    alt_fill = PatternFill("solid", start_color="EEF2FF")
    # Live rows that appear as a survivor VALUE in domination_map carry a
    # 'View superseded' placeholder cell (hyperlink wired in a later task).
    survivors = {survivor for survivor, _ in (domination_map or {}).values()}
    for _, row in bb.iterrows():
        inv = str(row.get("Invoice #", ""))
        # Superseded invoices (keys of domination_map) are NOT rendered on
        # this worksheet — they move to the reconciliation view. Only live
        # rows appear, so the loop below never touches a superseded row.
        if domination_map is not None and inv in domination_map:
            continue
        row_fill = alt_fill if r % 2 == 0 else PatternFill()
        bg = None if row_fill.fill_type is None else "EEF2FF"
        overlap_flag = inv in overlaps
        disclosed = _disclosed_label(bool(row.get("Cancel/Rebill Admitted")), overlap_flag)
        charge = float(row.get("Period Charge (£)", 0.0) or 0.0)
        total += charge
        value_src = str(row.get("Value Source", ""))
        bill_date_val = row.get("Bill Date", "")
        if isinstance(bill_date_val, pd.Timestamp | datetime):
            bill_date_val = bill_date_val.strftime("%d %b %Y")
        pf = row.get("Period From")
        if isinstance(pf, pd.Timestamp | datetime):
            pf = pf.strftime("%d %b %Y")
        pt = row.get("Period To")
        if isinstance(pt, pd.Timestamp | datetime):
            pt = pt.strftime("%d %b %Y")
        reason_assessment = str(row.get("Reason Assessment", ""))
        status = "Live"

        _text(ws, r, 1, inv, fill_hex=bg)
        _text(ws, r, 2, bill_date_val, fill_hex=bg)
        _text(ws, r, 3, pf, fill_hex=bg)
        _text(ws, r, 4, pt, fill_hex=bg)
        _num(ws, r, 5, int(row.get("Days Billed", 0)), fmt="#,##0", fill_hex=bg)
        _money(ws, r, 6, charge, fill_hex=bg)
        _text(ws, r, 7, value_src, fill_hex=bg)
        _num(ws, r, 8, int(row.get("12-Month Limit (days)", 365)), fmt="#,##0", fill_hex=bg)
        _num(ws, r, 9, int(row.get("Excess Days", 0)), fmt="#,##0", fill_hex=bg)
        # Highlight excess-days when >30 (i.e. back-billing is materially over)
        if int(row.get("Excess Days", 0)) > 30:
            ws.cell(row=r, column=9).font = Font(name="Calibri", size=10, bold=True, color="C00000")
        # Unlawful Charge (£): prorated share of Period Charge for the
        # Excess Days. Rendered as money; NOT summed into the trailing total.
        unlawful = float(row.get("Unlawful Charge (£)", 0.0) or 0.0)
        _money(ws, r, 10, unlawful, fill_hex=bg)
        _text(ws, r, 11, disclosed, fill_hex=bg)
        _text(ws, r, 12, reason_assessment, wrap=True, fill_hex=bg)
        _evidence_report_hyperlink_cell(ws, r, 13, evidence_df, inv)
        # View on Evidence Report (col 14): bidirectional hotlink back to the
        # row on the EDF Evidence Report sheet. Match by Invoice # first,
        # falling back to the amt|days signature.
        target_row = None
        if evidence_index is not None:
            target_row = evidence_index.get(f"inv:{inv}")
            if target_row is None:
                try:
                    amt = float(row.get("Amount (£)", 0.0) or 0.0)
                    days = int(row.get("Days Billed", 0) or 0)
                    key = f"amt_days:{amt:.2f}|{days}"
                    target_row = evidence_index.get(key)
                except (TypeError, ValueError):
                    pass
        if target_row is not None:
            cell = ws.cell(row=r, column=14, value="→")
            cell.hyperlink = openpyxl.worksheet.hyperlink.Hyperlink(
                ref=cell.coordinate,
                location=f"'EDF Evidence Report'!A{target_row}",
                display="→",
                tooltip=f"Jump to EDF Evidence Report!A{target_row}",
            )
            cell.font = Font(name="Calibri", size=10, color="0563C1", underline="single")
        else:
            cell = ws.cell(row=r, column=14, value="No match")
            cell.font = Font(name="Calibri", size=10, italic=True, color="A6A6A6")
        # Status / Superseded By / Partial Overlap (cols 15-17). Every row on
        # this sheet is Live (superseded rows are excluded above), so
        # Superseded By / Partial Overlap are always blank here — their
        # content lives on the reconciliation view.
        _text(ws, r, 15, status, fill_hex=bg)
        _text(ws, r, 16, "", fill_hex=bg)
        _text(ws, r, 17, "", fill_hex=bg)
        # View Superseded (col 18): a blue-underline placeholder on live rows
        # that are a survivor in domination_map. When ``view_superseded_row``
        # maps this invoice to a row on the Superseded Reconciliation sheet,
        # the cell hyperlinks there so a reader can jump to the group.
        if inv in survivors:
            cell = _text(ws, r, 18, "View superseded", fill_hex=bg)
            cell.font = Font(name="Calibri", size=10, color="0563C1", underline="single")
            target_row = (view_superseded_row or {}).get(inv)
            if target_row:
                cell.hyperlink = openpyxl.worksheet.hyperlink.Hyperlink(
                    ref=cell.coordinate,
                    location=f"'Superseded Reconciliation'!A{target_row}",
                    display="View superseded",
                    tooltip=f"Jump to Superseded Reconciliation!A{target_row}",
                )
        else:
            _text(ws, r, 18, "", fill_hex=bg)
        # Sub-Period Basis (col 19): how Unlawful Charge (£) was derived —
        # "Sub-period × rate" when the invoice carried a per-sub-period table,
        # "Day-ratio fallback" when the old proration was used. Provenance
        # only; never summed into the trailing total.
        _text(ws, r, 19, str(row.get("Sub-Period Basis", "") or ""), fill_hex=bg)
        r += 1

    # Trailing total row: a single union total over LIVE rows only.
    if not bb.empty:
        from edf_bill_fetcher.processors.detection import compute_unlawful_union_total

        bb_live = bb[~bb["Invoice #"].astype(str).isin(domination_map or {})]
        union_total = compute_unlawful_union_total(bb_live)
        total_label = "TOTAL UNLAWFUL CHARGES — UNION (no double count)"
        # Period Charge total (col 6) sums the live rows; the Unlawful column
        # (col 10) carries the union of the live rows' unlawful consumption
        # days so the same consumption is never counted twice. When there are
        # no live rows both values render as £0.00.
        write_trailing_total(
            ws,
            r,
            total_label,
            [(6, round(total, 2)), (10, round(union_total, 2))],
            5,
            19,
        )

    # Column widths
    widths: dict[str, float] = {
        "A": 18,
        "B": 14,
        "C": 14,
        "D": 14,
        "E": 12,
        "F": 16,
        "G": 18,
        "H": 14,
        "I": 12,
        "J": 16,  # Unlawful Charge (£)
        "K": 22,  # Cancel/Rebill Disclosed
        "L": 60,  # Reason Assessment
        "M": 60,  # Open PDF
        "N": 22,  # View on Evidence Report
        "O": 14,  # Status
        "P": 16,  # Superseded By
        "Q": 16,  # Partial Overlap
        "R": 18,  # View Superseded
        "S": 20,  # Sub-Period Basis
    }
    set_column_widths_from_spec(ws, widths)
    freeze_at(ws, "A8")
