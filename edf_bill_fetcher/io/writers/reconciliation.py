"""Reconciliation sheet writer — extracted from writers/__init__.py (Phase 5E).

Contains the writer plus two small _recon_* helpers for date/amount parsing.
The shared `_recon_hyperlink` helper stays in writers/_helpers.py and is
imported (not re-extracted) here.
"""

from __future__ import annotations

import re

import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from edf_bill_fetcher.writers._helpers import _recon_hyperlink

# ---- _recon_parse_iso_date (was writers/__init__.py L1822-1833) ----

def _recon_parse_iso_date(s: str) -> pd.Timestamp | pd._libs.tslibs.nattype.NaTType:
    if not s:
        return pd.NaT
    s = str(s).strip()
    if not s:
        return pd.NaT
    # ISO first (YYYY-MM-DD), then day-first for DD/MM/YYYY.
    if re.match(r"^\d{4}-\d{2}-\d{2}$", s):
        return pd.to_datetime(s, errors="coerce")
    return pd.to_datetime(s, dayfirst=True, errors="coerce")




# ---- _recon_amount_to_float (was L1834-1844) ----

def _recon_amount_to_float(v: object) -> float:
    if v is None:
        return 0.0
    if isinstance(v, int | float):
        return float(v)
    try:
        return float(str(v).replace(",", "").strip().lstrip("£"))
    except ValueError:
        return 0.0




# ---- write_reconciliation_sheet (was L1845-2217) ----

def write_reconciliation_sheet(
    ws_summary: Worksheet,
    ws_detail: Worksheet,
    sap_contract: list[dict],
    inferred_contract: pd.DataFrame,
    sap_meter: list[dict],
    inferred_meter: pd.DataFrame,
    sap_financial: list[dict],
    evidence_df: pd.DataFrame,
    account: str = "",
) -> None:
    """Render the two-sheet cross-source Reconciliation pair.

    Spec §3.2.  Sheet 1 is a compact 3-entity summary with verdict
    and drill-down hyperlink; Sheet 2 carries unmatched-only rows
    across three sections with AutoFilter.
    """
    ORANGE = "FE5716"
    NAVY = "10367A"

    def _banner(ws: Worksheet, row: int, text: str) -> None:
        ws.cell(row=row, column=1, value=text)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        for c in range(1, 9):
            cell = ws.cell(row=row, column=c)
            cell.fill = PatternFill("solid", start_color=NAVY)
            cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="left", vertical="center")

    def _section_banner(ws: Worksheet, row: int, text: str) -> None:
        ws.cell(row=row, column=1, value=f"\u25a0 {text} \u25a0")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        for c in range(1, 9):
            cell = ws.cell(row=row, column=c)
            cell.fill = PatternFill("solid", start_color=ORANGE)
            cell.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="left", vertical="center")

    def _header(ws: Worksheet, row: int, cols: list[str]) -> None:
        for i, name in enumerate(cols, start=1):
            cell = ws.cell(row=row, column=i, value=name)
            cell.fill = PatternFill("solid", start_color="DDDDDD")
            cell.font = Font(name="Calibri", size=10, bold=True, color="000000")

    # ---- Sheet 1 — summary ----
    ws_summary.title = "Reconciliation"
    _banner(ws_summary, 1, "EDF CROSS-SOURCE RECONCILIATION")
    second_line = "SAP dumps vs inferred evidence + reconciled evidence_df rows"
    if account:
        second_line += f"  •  Account: {account}"
    ws_summary.cell(row=2, column=1, value=second_line)
    ws_summary.merge_cells(start_row=2, start_column=1, end_row=2, end_column=8)

    sum_headers = [
        "Entity",
        "SAP-side count",
        "EDF-inferred count",
        "Matched",
        "Unmatched (SAP)",
        "Unmatched (EDF)",
        "Verdict",
        "Drill down",
    ]
    _header(ws_summary, 3, sum_headers)

    # ---- Sheet 2 — unmatched detail ----
    ws_detail.title = "Reconciliation Drill-down"
    _banner(ws_detail, 1, "EDF CROSS-SOURCE RECONCILIATION")
    ws_detail.cell(row=2, column=1, value=second_line)
    ws_detail.merge_cells(start_row=2, start_column=1, end_row=2, end_column=8)

    # Pre-compute detail row counters; each closure below mutates
    # these so the summary can read the tallies after all three
    # sections are built.
    detail_row = 3  # header row will be written here first
    contract_counts: dict[str, int] = {}
    meter_counts: dict[str, int] = {}
    financial_counts: dict[str, int] = {}

    def _build_section(
        ws: Worksheet,
        title: str,
        sap_items: list[dict],
        inferred_df: pd.DataFrame,
        inferred_cols: list[str],
        sap_date_key: str,
        sap_ref_key: str,
        sap_read_key: str | None,
        inferred_date_key: str,
        inferred_read_key: str,
        source_label: str,
        inferred_label: str,
        date_tolerance_days: int = 7,
        amount_tolerance: float | None = None,
    ) -> dict[str, int]:
        nonlocal detail_row
        counts = {"matched": 0, "unmatched_sap": 0, "unmatched_edf": 0}

        _section_banner(ws, detail_row, title)
        detail_row += 1
        header_cols = [
            "Status",
            f"SAP {sap_date_key}",
            f"SAP {sap_read_key or sap_ref_key}",
            f"Inferred {inferred_date_key}",
            f"Inferred {inferred_read_key}",
            "Notes",
            "Source",
            "Hyperlink",
        ]
        _header(ws, detail_row, header_cols)
        detail_row += 1
        section_start = detail_row

        inferred_rows: list[dict] = []
        if inferred_df is not None and not inferred_df.empty:
            inferred_rows = inferred_df.to_dict(orient="records")

        # Build matched stacks.
        unmatched_inferred = list(range(len(inferred_rows)))
        for sap in sap_items:
            sap_date = _recon_parse_iso_date(sap.get(sap_date_key, ""))
            sap_read = _recon_amount_to_float(sap.get(sap_read_key, "")) if sap_read_key else None
            matched = False
            for i in unmatched_inferred[:]:
                inf = inferred_rows[i]
                inf_date = _recon_parse_iso_date(inf.get(inferred_date_key, ""))
                date_close = (
                    abs((sap_date - inf_date).days) <= date_tolerance_days
                    if not pd.isna(sap_date) and not pd.isna(inf_date)
                    else False
                )
                if not date_close:
                    continue
                # Amount-close check (if applicable).
                if amount_tolerance is not None and sap_read is not None:
                    inf_read = _recon_amount_to_float(inf.get(inferred_read_key, ""))
                    if abs(sap_read - inf_read) > amount_tolerance:
                        continue
                # Successful match — suppress on detail sheet per spec 3.2;
                # only the summary counts carry this information.
                unmatched_inferred.remove(i)
                counts["matched"] += 1
                break
            if not matched:
                sap_idx = sap_items.index(sap) + 4
                ws.cell(row=detail_row, column=1, value="Missing in Inferred")
                ws.cell(row=detail_row, column=2, value=sap.get(sap_date_key, ""))
                ws.cell(
                    row=detail_row,
                    column=3,
                    value=sap.get(sap_read_key or sap_ref_key, ""),
                )
                for col in (4, 5):
                    ws.cell(row=detail_row, column=col, value="—")
                ws.cell(
                    row=detail_row,
                    column=6,
                    value=f"SAP {title.lower()} row not present in {inferred_label}",
                )
                ws.cell(row=detail_row, column=7, value=source_label)
                _recon_hyperlink(ws, detail_row, 8, source_label, sap_idx)
                detail_row += 1
                counts["unmatched_sap"] += 1
        for i in unmatched_inferred:
            inf = inferred_rows[i]
            inf_target = i + 4
            ws.cell(row=detail_row, column=1, value="Missing in SAP")
            for col in (2, 3):
                ws.cell(row=detail_row, column=col, value="—")
            ws.cell(
                row=detail_row,
                column=4,
                value=inf.get(inferred_date_key, ""),
            )
            ws.cell(
                row=detail_row,
                column=5,
                value=inf.get(inferred_read_key, ""),
            )
            ws.cell(
                row=detail_row,
                column=6,
                value=f"{inferred_label} row not present in SAP dump",
            )
            ws.cell(row=detail_row, column=7, value=inferred_label)
            _recon_hyperlink(ws, detail_row, 8, inferred_label, inf_target)
            detail_row += 1
            counts["unmatched_edf"] += 1

        # Compact spacer after each section.
        detail_row += 1
        return {**counts, "section_start": section_start}

    # Contract section
    contract_section = _build_section(
        ws_detail,
        "Contract Reconciliation",
        sap_contract,
        inferred_contract,
        ["Contract From", "Contract To", "Product Code"],
        "Contract From",
        "Product Code",
        None,
        "Contract From",
        "Product Code",
        "SAP Contract History",
        "Contract History",
    )
    contract_counts = contract_section

    # Meter Read section
    meter_section = _build_section(
        ws_detail,
        "Meter Read Reconciliation",
        sap_meter,
        inferred_meter,
        ["Meter Read Date", "Reading (kWh)"],
        "Meter Read Date",
        "Reading (kWh)",
        "Reading (kWh)",
        "Meter Read Date",
        "Reading (kWh)",
        "SAP Meter Readings",
        "Meter Readings",
    )
    meter_counts = meter_section

    # Financial reconciliation section
    _section_banner(ws_detail, detail_row, "Financial Reconciliation")
    detail_row += 1
    _header(
        ws_detail,
        detail_row,
        [
            "Status",
            "SAP Document No.",
            "SAP Posting Date",
            "SAP Amount",
            "Evidence Date",
            "Evidence Amount",
            "Notes",
            "Hyperlink",
        ],
    )
    detail_row += 1
    financial_section_start = detail_row

    evidence_rows_list: list[dict] = []
    if evidence_df is not None and not evidence_df.empty:
        evidence_rows_list = evidence_df.to_dict(orient="records")

    unmatched_ev = list(range(len(evidence_rows_list)))
    financial_counts = {"matched": 0, "unmatched_sap": 0, "unmatched_edf": 0}
    for sap in sap_financial:
        sap_date = _recon_parse_iso_date(sap.get("Posting Date", ""))
        sap_amt = _recon_amount_to_float(sap.get("Amount", ""))
        matched = False
        for i in unmatched_ev[:]:
            ev = evidence_rows_list[i]
            ev_date = _recon_parse_iso_date(ev.get("Date", ""))
            ev_amt = _recon_amount_to_float(ev.get("Amount (£)", 0))
            date_close = (
                abs((sap_date - ev_date).days) <= 7
                if not pd.isna(sap_date) and not pd.isna(ev_date)
                else False
            )
            amount_close = abs(sap_amt - ev_amt) <= 0.50
            if date_close and amount_close:
                matched = True
                abs_amt_diff = abs(sap_amt - ev_amt)
                if abs_amt_diff >= 0.01:
                    # Discrepancy rows stay on the detail sheet.
                    status = "Discrepancy"
                    note = f"Amount Δ = £{abs_amt_diff:,.2f}"
                    ws_detail.cell(row=detail_row, column=1, value=status)
                    ws_detail.cell(row=detail_row, column=2, value=sap.get("Document No.", ""))
                    ws_detail.cell(row=detail_row, column=3, value=sap.get("Posting Date", ""))
                    ws_detail.cell(row=detail_row, column=4, value=sap_amt)
                    ws_detail.cell(row=detail_row, column=5, value=ev.get("Date", ""))
                    ws_detail.cell(row=detail_row, column=6, value=ev_amt)
                    ws_detail.cell(row=detail_row, column=7, value=note)
                    _recon_hyperlink(
                        ws_detail, detail_row, 8, "EDF Evidence Report", financial_section_start + i
                    )
                    detail_row += 1
                unmatched_ev.remove(i)
                financial_counts["matched"] += 1
                break
        if not matched:
            sap_idx = sap_financial.index(sap) + 4
            ws_detail.cell(row=detail_row, column=1, value="Missing in Evidence")
            ws_detail.cell(row=detail_row, column=2, value=sap.get("Document No.", ""))
            ws_detail.cell(row=detail_row, column=3, value=sap.get("Posting Date", ""))
            ws_detail.cell(row=detail_row, column=4, value=sap_amt)
            for col in (5, 6):
                ws_detail.cell(row=detail_row, column=col, value="—")
            ws_detail.cell(
                row=detail_row,
                column=7,
                value="SAP financial row not on Evidence Report",
            )
            _recon_hyperlink(ws_detail, detail_row, 8, "SAP Financial Transactions", sap_idx)
            detail_row += 1
            financial_counts["unmatched_sap"] += 1
    for i in unmatched_ev:
        ev = evidence_rows_list[i]
        ev_target = financial_section_start + i
        ws_detail.cell(row=detail_row, column=1, value="Missing in SAP")
        for col in (2, 3, 4):
            ws_detail.cell(row=detail_row, column=col, value="—")
        ws_detail.cell(row=detail_row, column=5, value=ev.get("Date", ""))
        ws_detail.cell(row=detail_row, column=6, value=ev.get("Amount (£)", ""))
        ws_detail.cell(
            row=detail_row,
            column=7,
            value="Evidence row not present in SAP Financial dump",
        )
        _recon_hyperlink(ws_detail, detail_row, 8, "EDF Evidence Report", ev_target)
        detail_row += 1
        financial_counts["unmatched_edf"] += 1

    # ---- AutoFilter + freeze on detail sheet ----
    if detail_row > 3:
        ws_detail.auto_filter.ref = f"A3:H{detail_row}"
    ws_detail.freeze_panes = "A4"

    # ---- Build summary rows (3 entities) ----
    section_starts = {
        "Contract": contract_section["section_start"],
        "Meter Read": meter_section["section_start"],
        "Financial": financial_section_start,
    }

    def _verdict(entity: str, counts: dict[str, int]) -> str:
        if counts["matched"] == 0 and counts["unmatched_sap"] == 0:
            return f"No SAP-side {entity.lower()} rows to reconcile"
        if counts["unmatched_edf"] == 0 and counts["unmatched_sap"] == 0:
            return f"All {counts['matched']} SAP {entity.lower()} rows matched"
        parts = []
        if counts["unmatched_sap"]:
            parts.append(f"{counts['unmatched_sap']} SAP {entity.lower()} row(s) missing in EDF")
        if counts["unmatched_edf"]:
            parts.append(f"{counts['unmatched_edf']} EDF {entity.lower()} row(s) missing in SAP")
        return "; ".join(parts)

    summary_rows = [
        ("Contract", len(sap_contract), len(inferred_contract), contract_counts),
        ("Meter Read", len(sap_meter), len(inferred_meter), meter_counts),
        ("Financial", len(sap_financial), len(evidence_rows_list), financial_counts),
    ]
    for i, (entity, sap_cnt, edf_cnt, counts) in enumerate(summary_rows, start=4):
        ws_summary.cell(row=i, column=1, value=entity)
        ws_summary.cell(row=i, column=2, value=sap_cnt)
        ws_summary.cell(row=i, column=3, value=edf_cnt)
        ws_summary.cell(row=i, column=4, value=counts["matched"])
        ws_summary.cell(row=i, column=5, value=counts["unmatched_sap"])
        ws_summary.cell(row=i, column=6, value=counts["unmatched_edf"])
        ws_summary.cell(row=i, column=7, value=_verdict(entity, counts))
        target_row = section_starts[entity]
        cell = ws_summary.cell(row=i, column=8)
        cell.value = "→ Drill down"
        cell.hyperlink = openpyxl.worksheet.hyperlink.Hyperlink(
            ref=cell.coordinate,
            location=f"'Reconciliation Drill-down'!A{target_row}",
            display="→ Drill down",
        )
        cell.font = Font(color="0563C1", underline="single")

    ws_summary.freeze_panes = "A4"


# ---------------------------------------------------------------------------


__all__ = [
    "_recon_amount_to_float",
    "_recon_parse_iso_date",
    "write_reconciliation_sheet",
]
