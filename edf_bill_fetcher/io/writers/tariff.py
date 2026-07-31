"""Tariff analysis sheet writer — extracted from writers/__init__.py.

Contains: write_tariff_analysis_sheet — renders the "Tariff Analysis" worksheet
with average tariff-change stats and tariff impact breakdown.
"""
from __future__ import annotations

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill

from edf_bill_fetcher.helpers.excel_utils import (
    hcell as _hcell,
)
from edf_bill_fetcher.helpers.excel_utils import (
    money as _money,
)
from edf_bill_fetcher.helpers.excel_utils import (
    num as _num,
)
from edf_bill_fetcher.helpers.excel_utils import (
    section_hdr as _section_hdr,
)
from edf_bill_fetcher.helpers.excel_utils import (
    text as _text,
)
from edf_bill_fetcher.helpers.theme import CELL_BORDER
from edf_bill_fetcher.writers._helpers import _analyze_tariff_impact

# --- write_tariff_analysis_sheet (was writers/__init__.py L2581-2651) ---


def write_tariff_analysis_sheet(ws, dfc):
    """Write Tariff Impact Analysis tab."""
    ws.title = "Tariff Analysis"

    NAVY = "10367A"
    ORANGE = "FE5716"
    LGREY = "F0F0F0"

    tariff_info = _analyze_tariff_impact(dfc)

    if not tariff_info:
        _hcell(ws, 1, 1, "No tariff data available in records", bg=NAVY)
        ws.column_dimensions["A"].width = 50
        return

    headers = [
        "Tariff",
        "Records",
        "Avg Unit Rate (p/kWh)",
        "Median Unit Rate",
        "Min Rate",
        "Max Rate",
        "Avg Period Charge (£)",
    ]
    for col, h in enumerate(headers, 1):
        _hcell(ws, 1, col, h, bg=NAVY)
    ws.row_dimensions[1].height = 28

    tc = ws.cell(row=1, column=1, value="EDF ENERGY DISPUTE  —  TARIFF IMPACT ANALYSIS")
    tc.font = Font(name="Calibri", size=13, bold=True, color="FFFFFF")
    tc.fill = PatternFill("solid", start_color=ORANGE)
    tc.border = CELL_BORDER
    tc.alignment = Alignment(horizontal="left", vertical="center")
    for c in range(2, 8):
        x = ws.cell(row=1, column=c)
        x.fill = PatternFill("solid", start_color=ORANGE)
        x.border = CELL_BORDER

    tariff_stats = tariff_info.get("tariff_stats")
    if tariff_stats is not None:
        assert isinstance(tariff_stats, pd.DataFrame)
        if not tariff_stats.empty:
            r = 2
            for _, row in tariff_stats.iterrows():
                bg = LGREY if r % 2 == 0 else None
                _text(ws, r, 1, str(row["Tariff"]), fill_hex=bg)
                _num(ws, r, 2, int(row["count"]), fmt="#,##0", fill_hex=bg)
                _num(ws, r, 3, float(row["avg_unit_rate"]), fmt="0.00", fill_hex=bg)
                _num(ws, r, 4, float(row["median_unit_rate"]), fmt="0.00", fill_hex=bg)
                _num(ws, r, 5, float(row["min_unit_rate"]), fmt="0.00", fill_hex=bg)
                _num(ws, r, 6, float(row["max_unit_rate"]), fmt="0.00", fill_hex=bg)
                avg_chg = row["avg_charge"]
                _money(ws, r, 7, float(avg_chg) if pd.notna(avg_chg) else 0, fill_hex=bg)
                r += 1

    r += 1
    _section_hdr(ws, r, "SUMMARY")
    r += 1
    _text(ws, r, 1, "Unique Tariffs Identified")
    _num(ws, r, 2, tariff_info.get("num_tariffs", 0), fmt="#,##0")
    r += 1
    _text(ws, r, 1, "Tariff Changes Detected")
    _num(ws, r, 2, tariff_info.get("tariff_changes", 0), fmt="#,##0")

    for col_letter, width in zip(
        ["A", "B", "C", "D", "E", "F", "G"], [28, 10, 22, 18, 16, 16, 20], strict=False
    ):
        ws.column_dimensions[col_letter].width = width
    ws.freeze_panes = "A2"


__all__ = ["write_tariff_analysis_sheet"]
