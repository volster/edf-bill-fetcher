"""Writer functions for the EDF evidence workbook.

Extracted from ``edf_collector.py`` as part of the modularization
refactor (Task 5).  Each function writes one or more Excel sheets
using openpyxl.
"""

from __future__ import annotations

import gc
import glob
import json
import os
import pickle
import re
import threading
import traceback
from datetime import date, datetime
from typing import Any, cast

import numpy as np
import openpyxl
import pandas as pd

try:
    import tkinter as tk
    from tkinter import filedialog, messagebox, ttk

    HAS_TK = True
except ImportError:
    HAS_TK = False

try:
    import pypff

    HAS_PYPFF = True
except ImportError:
    HAS_PYPFF = False

try:
    import importlib.util

    HAS_SCIPY = importlib.util.find_spec("scipy") is not None
except ImportError:
    HAS_SCIPY = False

try:
    importlib.util.find_spec("statsmodels.tsa.holtwinters")

    HAS_STATSMODELS = True
except ImportError:
    HAS_STATSMODELS = False

try:
    HAS_PDF_REPORT = importlib.util.find_spec("edf_report") is not None
    HAS_DOCX_REPORT = importlib.util.find_spec("edf_report_docx") is not None
except ImportError:
    HAS_PDF_REPORT = False
    HAS_DOCX_REPORT = False

from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.worksheet import Worksheet

from edf_bill_fetcher.helpers.date_utils import (  # noqa: E402,F401,I001
    _ISO_DATE_RE,
    _safe_to_datetime,
    parse_to_display_date,
    parse_to_sort_date,
    to_excel_date,
)
from edf_bill_fetcher.helpers.date_utils import (
    completeness_score as _completeness_score,
)
from edf_bill_fetcher.helpers.excel_utils import (  # noqa: E402,F401,I001
    _TEXT_SUPPRESSION_QUEUE,
    CELL_BORDER,
)
from edf_bill_fetcher.helpers.excel_utils import (
    build_sap_row_index_map as _build_sap_row_index_map,
)
from edf_bill_fetcher.helpers.excel_utils import (
    hcell as _hcell,
)
from edf_bill_fetcher.helpers.excel_utils import (
    money as _money,
)
from edf_bill_fetcher.helpers.excel_utils import (
    num as _num,
)
from edf_bill_fetcher.helpers.excel_utils import (  # noqa: F401
    open_pdf_hyperlink_cell as _open_pdf_hyperlink_cell,
)
from edf_bill_fetcher.helpers.excel_utils import (
    section_hdr as _section_hdr,
)
from edf_bill_fetcher.helpers.excel_utils import (
    text as _text,
)
from edf_bill_fetcher.helpers.formatting import (
    _apply_amalgamate_to_kept_frame,
)
from edf_bill_fetcher.helpers.formatting import (  # noqa: E402,F401,I001
    account_number_matches as _account_number_matches,
)
from edf_bill_fetcher.io.adapters.pdf import legal_context  # noqa: E402,F401,I001
from edf_bill_fetcher.writers._helpers import (  # noqa: E402,F401,I001
    _SOURCE_PRECEDENCE,
    DUP_GREY,
    EDF_NAVY,
    EDF_OFFWHITE,
    EDF_ORANGE,
    EST_YELLOW,
    JUMP_RED,
    MEDIUM_GREY,
    _analyze_tariff_impact,
    _compute_volatility,
    _data_quality_report,
    _detect_payment_patterns,
    _disclosed_label,
    _holt_winters_forecast,
    _holt_winters_forecast_pair,
    _iqr_anomalies,
    _linear_forecast,
    _linear_forecast_pair,
    _parse_amount_for_event,
    _reading_type_to_aem,
    _recon_hyperlink,
    _zscore_anomalies,
    build_evidence_index,
    compute_dispute_flags,
    detect_sap_back_billing_events,
    match_sap_events_to_edf,
)
from edf_bill_fetcher.io.writers.evidence import (  # noqa: E402,F401
    write_evidence_sheet,
    write_summary_sheet,
)
from edf_bill_fetcher.models.events import SapBackBillingEvent, SapEdfMatch  # noqa: I001
from evidence_bundle import build_bundle_index, save_evidence_files  # noqa: E402,F401

__all__ = [
    "_analyze_tariff_impact",
    "_compute_volatility",
    "_data_quality_report",
    "_detect_payment_patterns",
    "_disclosed_label",
    "_holt_winters_forecast",
    "_holt_winters_forecast_pair",
    "_iqr_anomalies",
    "_linear_forecast",
    "_linear_forecast_pair",
    "_recon_hyperlink",
    "_write_sap_bb_events_sheet",
    "_write_sap_bb_matches_sheet",
    "_write_sap_header_row",
    "build_evidence_index",
    "compute_dispute_flags",
    "detect_back_billing",
    "detect_sap_back_billing_events",
    "export_to_excel",
    "match_sap_events_to_edf",
    "write_back_billing_sheet",
    "write_contract_history_sheet",
    "write_data_quality_sheet",
    "write_evidence_sheet",
    "write_forecast_sheet",
    "write_meter_readings_sheet",
    "write_payment_analysis_sheet",
    "write_rebilling_sheet",
    "write_reconciliation_sheet",
    "write_sap_back_billing_sheets",
    "write_sap_contract_history_sheet",
    "write_sap_financial_transactions_sheet",
    "write_sap_meter_readings_sheet",
    "write_statistical_analysis_sheet",
    "write_summary_sheet",
    "write_tariff_analysis_sheet",
]


# ---------------------------------------------------------------------------
# Main export function
# ---------------------------------------------------------------------------


def export_to_excel(data, output_path, error_log, config, filtered=None, sap_rows=None):
    NAVY = "10367A"
    ORANGE = "FE5716"
    RED = "FF6B6B"
    AMBER = "FFD166"
    GREEN = "06D6A0"
    LGREY = "F0F0F0"
    DGREY = "888888"

    df = pd.DataFrame(data)
    df["_sort"] = df["Date"].apply(parse_to_sort_date)
    df = df.sort_values(by=["_sort", "Invoice #"], ascending=[True, False]).reset_index(drop=True)
    df["% Change"] = None

    # Deduplication — multi-pass to match the same bill across sources
    # Pass 1: Period To + Amount  (catches HTM ↔ PST where billing period matches)
    # Pass 2: Amount within 60-day window for records with no period info (Local PDF)
    dup_df = pd.DataFrame()
    if config.get("use_dedup", True):
        # Source precedence lives at module scope (``_SOURCE_PRECEDENCE``)
        # so that ``tests/test_source_precedence.py`` can pin the
        # explicit ordering without booting the entire Excel
        # export pipeline.  Lower number = higher precedence.
        df["_src_pri"] = df["Source"].map(_SOURCE_PRECEDENCE).fillna(9).astype(int)
        # Completeness score — primary sort key.  Spec: "duplicates
        # should be assessed and the most complete version of the
        # information presented".  ``_completeness_score`` counts
        # populated substantive fields on each row; the richer row
        # sorts *before* the sparser row so ``keep="first"`` keeps it.
        # Computed here (not earlier) so it's available even if the
        # upstream pipeline headers change in future.
        df["_completeness"] = df.apply(_completeness_score, axis=1)
        # Sort order (primary to tie-breaker):
        #   1. _completeness descending      — most-populated row wins
        #   2. _src_pri ascending             — higher-precedence source wins ties
        #   3. _sort ascending                — earliest date wins remaining ties
        # ``keep="first"`` then retains the head of every duplicate cluster.
        # Pre-fix the sort was only ``["_src_pri", "_sort"]`` so source
        # precedence dominated completeness — a sparser HTM row would
        # beat a richer PST row.  The companion test is
        # ``tests/test_dedup_most_complete.py``.
        df = df.sort_values(
            ["_completeness", "_src_pri", "_sort"],
            ascending=[False, True, True],
        ).reset_index(drop=True)

        # Dedup key: prefer Period To (consistent across sources for same bill),
        # fall back to Date for records without period info.  Pass 1's
        # ``DUPLICATED`` flags for *period-aware* rows track which *kept*
        # row they collide against so the dup sheet can render a clickable
        # summary linking back to the source-of-truth record.  We capture
        # the matched-against row's *original* df index — that index is
        # what ``dup_df.index`` carries through to the writer, since
        # ``dup_df = df[is_dup]`` runs before the ``reset_index`` line below.
        # Period To is the source-of-truth end-of-billing-period
        # date when present; fall back to ``_sort`` (the parsed
        # source-specific ``Date``) when the row is no-period
        # (e.g. Local PDF).  ``df["_sort"].where(cond, df["_sort"])``
        # is a tautology — Period To was being ignored and Pass 1
        # ``_dedup_date`` is the *canonical* dedup key — Period To when
        # available, otherwise left as ``NaT`` so the row is excluded
        # from ``duplicated`` clusters (since ``duplicated`` treats
        # NaT as equal across rows, falling back to ``_sort`` would
        # silently merge unrelated no-period same-amount rows).
        # Rows with NaT here are rerouted through Pass-2's no-period
        # bucket logic below, which uses ``Period To == "N/A" | NaN``
        # as the explicit handling mask.
        # Vectorised pass via _safe_to_datetime to suppress the
        # 'format-inference fallback' UserWarning pandas emits on
        # mixed-format Series when a single string passes the
        # simple-format regex gate.
        period_to_dt = _safe_to_datetime(df["Period To"])
        df["_dedup_date"] = period_to_dt
        is_dup = df.duplicated(subset=["_dedup_date", "Amount (£)"], keep="first")
        # Pass 1 (period+amount): build ``kept_pass1_index`` keyed on
        # ``(_dedup_date, Amount)`` so we can look up "which kept row
        # did this dup lose to".  The kept row's original df index (not
        # its reset_index value) survives into the dup sheet.
        kept_for_dup: dict[int, int] = {}  # dup_idx -> kept_idx (both original indices)
        kept_for_summary: dict[int, dict[str, object]] = {}  # kept_idx -> display fields
        kept_frame = df[~is_dup]
        kept_pass1_index: dict[tuple, int] = {}
        for kept_idx in kept_frame.index:
            k = (
                kept_frame.at[kept_idx, "_dedup_date"],
                kept_frame.at[kept_idx, "Amount (£)"],
            )
            kept_pass1_index.setdefault(k, kept_idx)
            # Cache the displayed fields once per kept row so the
            # dup lookup below doesn't re-read them.
            kept_for_summary[kept_idx] = {
                "Source": kept_frame.at[kept_idx, "Source"],
                "Date": kept_frame.at[kept_idx, "Date"],
                "Amount (£)": kept_frame.at[kept_idx, "Amount (£)"],
            }
        # Resolve Pass 1's kept-against reference per duplicate
        # before any reset_index runs.
        for dup_idx in df[is_dup].index:
            k = (
                df.at[dup_idx, "_dedup_date"],
                df.at[dup_idx, "Amount (£)"],
            )
            kept_idx = kept_pass1_index.get(k, -1)
            kept_for_dup[dup_idx] = kept_idx

        # Pass 2: records with no period info (e.g. Local PDF) — match by
        # Amount within a 60-day window of any already-kept record.
        #
        # Phase 2.2 follows the spec: group candidates by Amount (£)
        # first, then look up matches inside each amount-bucket
        # rather than scanning the entire kept-mask frame for every
        # candidate.  The previous implementation was O(N²) — at
        # 5,000 records the *bench* showed it took ~2.3 s.  This
        # bucketed approach is O(N) amortised: typical EDF bills
        # have unique amounts, so bucket size is 1–2 rows and the
        # inner day-window check is effectively constant.
        #
        # Layout-preserving detail worth flagging: the *legacy*
        # algorithm visits ``df.index`` in increasing order and
        # looks at the live ``kept`` mask — which includes
        # forward-yet-to-be-visited rows whose ``~is_dup`` is the
        # pre-iteration value (so any same-amount row ±60 days
        # *before or after* the candidate, except itself, can
        # match).  We replicate that exact behaviour by iterating
        # ``df.index`` in *reverse* and building per-amount buckets
        # incrementally: at row N's visit, the bucket for any
        # amount A already contains every row with amount A and
        # index > N that wasn't marked as dup — exactly the
        # forward-direction rows the legacy code saw.
        #
        # Concretely: with the legacy ``kept = df[(~is_dup) &
        # (df.index != idx)]`` mask, the set of candidate matches
        # for row idx against amount A is
        # ``{j != idx : df.Amount[j] == A and ~is_dup.at[j]}``.
        # For most rows this set is split into:
        #   (i) j in [0, idx) — *earlier* df indices,
        #  (ii) j in (idx, len(df)) — *later* df indices.
        # The legacy code consulted both groups via the live
        # ``~is_dup`` mask.  Iterating reverse and limiting our
        # bucket hashes to *only* ``j > idx`` (the "earlier in
        # reverse-iteration-order" rows) lands on exactly the
        # same candidate set provided *no row gets marked as dup
        # before its later neighbours are visited* — which the
        # reverse loop guarantees by ordering inspections from
        # the bottom of the frame upwards.
        no_period = (df["Period To"] == "N/A") | df["Period To"].isna()
        # ``bucket_by_amt`` is keyed on Amount and stores the
        # ``(df_ordinal, _sort date)`` of every row already visited
        # (reverse-iteration order) that hasn't been marked as
        # duplicate.  We append a row to its bucket whenever the
        # row *does not* get marked — symmetric to the legacy
        # ``kept`` mask at iteration time.
        bucket_by_amt: dict[float, list[tuple[int, object]]] = {}
        # Reverse-iterate ``df.index`` so that "later in df order"
        # rows are visited first and accumulate in the bucket for
        # the earlier row's lookup.  Equivalently, the bucket for
        # each amount at ``idx`` is exactly the rows j > idx with
        # Amount[j] == amount and ~is_dup.at[j] — the same row set
        # legacy would consult.
        reverse_idx = list(df[~is_dup & no_period].index)[::-1]
        for idx in reverse_idx:
            amt = df.loc[idx, "Amount (£)"]
            rec_date = df.loc[idx, "_sort"]
            same_amt = bucket_by_amt.get(amt, [])
            matched = False
            for m_idx, m_date in same_amt:
                # ``pd.notna`` short-circuit means NaT-dated rows
                # already in the bucket (originally the loop
                # ``continue``-skipped them but still listed them
                # in the next-iter kept set) never trigger a match.
                if pd.notna(m_date) and abs((rec_date - m_date).days) <= 60:
                    matched = True
                    # Capture the matched-against row's *original
                    # df index* so the dup sheet can resolve to
                    # the same frame.  We resolve the summary
                    # *before* the kept set is `reset_index`-
                    # rasterised below — once ``df = df[~is_dup]
                    # .reset_index(drop=True)`` runs, the
                    # ``m_idx`` no longer references a row.
                    kept_for_dup[idx] = m_idx
                    kept_for_summary[m_idx] = {
                        "Source": df.at[m_idx, "Source"],
                        "Date": df.at[m_idx, "Date"],
                        "Amount (£)": df.at[m_idx, "Amount (£)"],
                    }
                    break
            if matched:
                is_dup.at[idx] = True
                # Don't add to the bucket — the legacy loop's
                # recomputed ``~is_dup`` mask would have excluded a
                # row marked dup at the *start* of iteration, so it
                # cannot anchor later (here: earlier-in-iteration)
                # matches either.
            else:
                # Always add the row even if ``_sort`` is NaT —
                # the legacy ``kept`` mask at the *next* (lower) row
                # includes this row because it's ``~is_dup``-true,
                # and the NaT date just means it can't anchor a
                # match on its own.
                bucket_by_amt.setdefault(amt, []).append((idx, rec_date))

        # ``dup_df`` is built BEFORE the ``reset_index`` line below so
        # ``dup_df.index`` still carries each duplicate's original df
        # index — that's the key we use to look up the kept-against
        # summary in ``kept_for_summary``.
        #
        # ``save_dups`` toggles whether dedup *itself* is applied to the
        # main dataframe (``df``).  When True (the historical default),
        # duplicates are filtered out of ``df`` and *recorded* in
        # ``dup_df`` for the dup sheet — users never lose visibility of
        # what was dropped.  When False, dedup is skipped entirely: every
        # row stays in ``df`` and ``dup_df`` is empty.
        if config.get("save_dups", True):
            dup_df = df[is_dup].copy()
        else:
            dup_df = df[is_dup].iloc[0:0].copy()

        # Spec 3 (stretch): hybrid rows when ``amalgamate_duplicates`` is
        # True.  Instead of keeping the completeness-winner verbatim, we
        # merge each duplicate cluster's non-empty fields into a single
        # hybrid kept row.  The composite keeps the completeness-winner's
        # ``Source`` identity and picks any populated column value from
        # any sibling.  Each non-surviving sibling still stays in
        # ``dup_df`` (the spec's 'never drop without being recorded').
        #
        # N.B. the amalgamated ``df`` is is already a cleaned kept set
        # (all duplicates removed), so the ``df[~is_dup]`` filter below
        # is skipped for the amalgamate path.
        if (
            config.get("save_dups", True)
            and config.get("amalgamate_duplicates", False)
            and not dup_df.empty
        ):
            df = _apply_amalgamate_to_kept_frame(df, dup_df, kept_pass1_index, kept_for_dup, is_dup)
            # dup_df stays unchanged — the amalgamation only touches the
            # kept set; the dup sheet still records every sibling.

        if config.get("save_dups", True) and not config.get("amalgamate_duplicates", False):
            df = df[~is_dup].reset_index(drop=True)
        # else: do not drop duplicates — leave ``df`` unchanged so the
        # user sees the raw ingress and can resolve duplicates manually.
        df = df.drop(columns=["_src_pri", "_dedup_date", "_completeness"], errors="ignore")

    df = df.drop(columns=["_sort"], errors="ignore")
    dup_df = (
        dup_df.drop(
            columns=["_sort", "_src_pri", "_dedup_date", "_completeness"],
            errors="ignore",
        )
        if not dup_df.empty
        else dup_df
    )

    # Compute Unit Rate (p/kWh) where both Period Charge and Units are available.
    #
    # Phase 2.1: vectorised path.  The historic row-wise apply walked
    # Python per row, which the bench measured at ~63 ms at 5,000
    # records (not the bottleneck we'd been worried about, but the
    # spec asks for vectorisation).  New path uses pd.to_numeric
    # + np.where — same observable output (rounded to 0.01) but
    # vectorised.  ``Units`` is normalised for the inline comma
    # (``"1,234"`` to ``"1234"``) the same way the row-wise path
    # did via ``str(units).replace(",", "")``.
    pc = pd.to_numeric(df["Period Charge (£)"], errors="coerce")
    units = pd.to_numeric(
        df["Units (kWh)"].astype(str).str.replace(",", ""),
        errors="coerce",
    )
    df["Unit Rate (p/kWh)"] = np.where(
        (units > 0) & (pc > 0),
        np.round((pc / units) * 100, 2),
        np.nan,
    )

    # ``dup_df`` computation is kept in the path for backward
    # compatibility — the dup DataFrame is much smaller than the
    # kept set, so per-row apply only adds ms-level overhead.  We
    # use a tiny module-scope helper rather than a closure so
    # ``pickle`` can find it on round-trip (the spec used to break
    # here because closures aren't picklable).
    def _compute_unit_rate(row):
        pc = row.get("Period Charge (£)")
        units = row.get("Units (kWh)")
        try:
            pc_f = float(pc)
            u_f = float(str(units).replace(",", ""))
            if u_f > 0 and pc_f > 0:
                return round((pc_f / u_f) * 100, 2)
        except (ValueError, TypeError):
            pass
        return np.nan

    if not dup_df.empty:
        dup_df["Unit Rate (p/kWh)"] = dup_df.apply(_compute_unit_rate, axis=1)
        # Matched-against kept-record block (Phase-2 follow-up).
        # Each duplicate row gets a clickable summary pointing
        # back to the *kept* record so an ombudsman reviewing the
        # workbook can navigate from the dup sheet to the
        # source-of-truth record with one click.  Earlier in the
        # dedup walk we built ``kept_for_summary`` keyed on the
        # duplicate's *original* df-index — that's also the index
        # ``dup_df.index`` carries because ``dup_df = df[is_dup]
        # .copy()`` runs *before* the ``reset_index(drops...)``
        # line.  So we can resolve the summary now without
        # re-doing any index resets.
        kept_idx_by_dup = {
            dup_idx: kept_for_summary.get(kept_for_dup.get(dup_idx, -1), {})
            for dup_idx in dup_df.index
        }

        # ``df`` is the kept set after dedup reset_index.  After
        # ``df = df[~is_dup].reset_index(drop=True)``, ``df.index``
        # is a sequential 0..N-1 range, *not* the original df
        # labels.  But the *order* of rows is preserved — the n-th
        # row of the kept set is the same n-th kept row that survived
        # dedup.  We therefore translate the original-index
        # references we still hold in ``kept_for_dup`` (the dedup
        # walker wrote them *before* reset_index) into post-reset
        # positions by ranking the kept rows in ascending original
        # df-index order — kept_rank[k] = rank-in-kept-set.
        kept_rank: dict[int, int] = {}
        for rank, orig_idx in enumerate(sorted(kept_for_summary.keys())):
            kept_rank[int(orig_idx)] = rank

        def _summary(idx: int) -> str:
            # Build the printable kept-row-reference string.  Falls
            # back to an empty string if the matched-against kept
            # row was rolled up by Pass 1 *after* the lookup
            # captured -1 (a corner case where the pattern matched
            # but no kept frame picked it up).
            row = kept_idx_by_dup.get(idx)
            if not row:
                return ""
            try:
                amount_val = float(row["Amount (£)"])  # type: ignore[arg-type]
                amt_str = "£" + format(amount_val, ".2f")
            except (TypeError, ValueError):
                amt_str = "£--"
            return f"{row['Source']} · {row['Date']} · {amt_str}"

        # ``Duplicate Of`` is the visible column on the dup sheet
        # itself; ``_matches_kept_idx`` is the link target the
        # Excel writer will use to mint the click-through hyperlink
        # back to the kept row in the main evidence report.
        dup_df["Duplicate Of"] = [_summary(idx) for idx in dup_df.index]
        # ``_matches_kept_idx`` is the *post-reset* position of
        # the kept row in ``EDF Evidence Report`` — the Excel
        # writer uses this with ``A{+1}`` as the click target
        # so an ombudsman can jump from the dup cell directly to
        # the source-of-truth record.  We translate via
        # ``kept_rank`` (computed above from kept-against-original
        # ordering) because the dedup walker built ``kept_for_dup``
        # *before* ``reset_index`` ran on the kept frame.
        dup_df["_matches_kept_idx"] = pd.Series(
            {idx: kept_rank.get(int(kept_for_dup.get(idx, -1)), -1) for idx in dup_df.index},
            dtype="Int64",
        )

    # F2 (SEV-1): single source of truth for the saved-column
    # ordering.  Every ``_add_record``-time builder must stamp
    # every name in this list (use ``record.setdefault(col, "N/A")``
    # if unsure) — otherwise ``reindex`` silently drops the column
    # and the workbook schema drifts from what other readers
    # (Tariff Analysis, Dict Comparer) expect.  The structural
    # guard lives in ``tests/test_export_headers_invariant.py``.
    col_order = [
        "Source",
        "Sender",
        "Date",
        "Period From",
        "Period To",
        "Invoice #",
        "Amount (£)",
        "Period Charge (£)",
        "Unit Rate (p/kWh)",
        "% Change",
        "Entry Type",
        "Reading",
        "Units (kWh)",
        "Standing Chg (p/day)",
        # Tariff column — lights up the Tariff Analysis Excel/DOCX/PDF
        # section.  Populated only by ``_process_new_invoice``;
        # every other source path stamps "N/A".  Without this entry
        # here, ``reindex`` would drop the column from the saved
        # workbook even though every record dict now carries it.
        "Tariff",
        "Attachment Name",
        "Details",
        "Logic Used",
        "Anomaly Flag",
        "Duplicate Of",
    ]
    # Diagnostic-only columns that the analyser writers (Back-billing,
    # Rebilling, Meter Readings, Contract History) need for their
    # Source Excerpt column lookup, but which must NOT appear on the
    # EDF Evidence Report tab.  They survive the canonical ``reindex``
    # below so ``dfc = df_an[...]`` retains them for the analyser
    # writers' ``evidence_df=dfc`` argument.  ``write_evidence_sheet``
    # drops them via the ``evidence_df = df.drop(columns=[...])`` pass
    # at line ~3694 just before the Evidence Report is written.
    # 'Balance Last Bill (£)' is captured by the reconciliation-statement
    # parser and is consumed by the analyser writers as a diagnostic; it
    # stays here for the same reason.
    diagnostic_cols = [
        "Source PDF Text",
        "_regex_trace",
        "Balance Last Bill (£)",
    ]
    # Only carry forward the diagnostic cols that are actually present
    # on the records -- avoids reindex inserting all-NaN cols when no
    # record builder emitted them (e.g. a synthetic test DataFrame).
    diagnostic_present = [c for c in diagnostic_cols if c in df.columns]
    df = df.reindex(columns=col_order + diagnostic_present)
    # Belt-and-braces invariant: every column the *kept* set still
    # carries must be in the canonical order list — otherwise a
    # future record builder that adds a new column without updating
    # col_order would survive the reindex and land as a
    # mysteries-leading-column in the saved workbook.  We assert
    # loudly here (developer-visible) rather than silently dropping
    # the unknown column.  The diagnostic cols (``Source PDF Text``,
    # ``_regex_trace``, ``Balance Last Bill (£)``) are intentionally
    # excluded from the canonical ``col_order`` so they're not written
    # to the Evidence Report sheet; the assertion below permits them.
    _allowed_extras = {"Source PDF Text", "_regex_trace", "Balance Last Bill (£)"}
    _unexpected = [c for c in df.columns if c not in col_order and c not in _allowed_extras]
    if _unexpected:
        raise ValueError(
            "export_to_excel received columns not in col_order: "
            f"{_unexpected!r}.  Add them to col_order or build the "
            "records so they carry only known keys."
        )
    # The dup sheet needs both ``Duplicate Of`` *and*
    # ``_matches_kept_idx`` available to the writer so the
    # post-loop pass can mint clickable HYPERLINK cells.  We
    # attach ``_matches_kept_idx`` after the reindex pass so the
    # saved workbook geometry stays 19-column even though the
    # writer's row-iteration will see the 20th column briefly —
    # the writer drops the column before saving.
    if not dup_df.empty and "_matches_kept_idx" in dup_df.columns:
        # Already present — nothing to do.
        pass
    else:
        # Neither column nor value is preserved.  Don't write
        # anything — the post-loop pass will skip minting
        # HYPERLINKs because ``match_positions_series`` is None.
        pass
    # No-op reindex guard for clarity; dup_df reindex on col_order
    # actually *drops* the helper column, which is what we want
    # for the Excel geometry — but we also need it for the
    # hyperlink pass.  Best approach: call site reads it BEFORE
    # reindex and threads it via a separate side cache.
    # The simplest implementation is to re-attach the column
    # *after* reindex here:
    if not dup_df.empty:
        dup_df_reindexed = dup_df.reindex(columns=col_order)
        # Re-attach from dup_df's pre-reindex view — the column
        # is dropped by reindex, so we restore it from the
        # original here.  This is the only place where the
        # writer would otherwise lose access to the helper.
        if "_matches_kept_idx" in dup_df.columns:
            dup_df = pd.concat(
                [
                    dup_df_reindexed,
                    dup_df["_matches_kept_idx"].rename("_matches_kept_idx"),
                ],
                axis=1,
            )
        else:
            dup_df = dup_df_reindexed

    # Years for summary tab
    years = sorted(
        y for y in df["Date"].apply(parse_to_sort_date).dropna().dt.year.astype(int).unique()
    )

    wb = openpyxl.Workbook()
    wb.calculation.fullCalcOnLoad = True

    # Tab 1: Evidence (created first — summary formulas reference it by name)
    ws_main = wb.active
    ws_main.title = "EDF Evidence Report"
    # The diagnostic-only columns (``Source PDF Text``, ``_regex_trace``,
    # ``Balance Last Bill (£)``) are captured by the parsers for the
    # analyser tabs' Source Excerpt column / balance-context rendering.
    # They are intentionally NOT written to the visible Evidence Report
    # tab: ``Source PDF Text`` is a 4 KB chunk per row (too noisy),
    # ``_regex_trace`` is internal pipeline metadata, and
    # ``Balance Last Bill (£)`` is a reconciliation-statement field that
    # only the Reconciliation tab needs.
    # Drop them from the copy handed to the writer; the underlying
    # ``df`` is left intact so subsequent analyser renders (``dfc``)
    # retain them for in-memory Source Excerpt lookups.
    _diagnostic_columns_for_evidence_report = [
        "Source PDF Text",
        "_regex_trace",
        "Balance Last Bill (£)",
    ]
    evidence_df = df.drop(
        columns=[c for c in _diagnostic_columns_for_evidence_report if c in df.columns],
        errors="ignore",
    )
    write_evidence_sheet(ws_main, evidence_df, is_duplicate=False)

    # Tab 2: Annual Summary
    ws_summary = wb.create_sheet(title="Annual Summary", index=0)
    write_summary_sheet(ws_summary, years, ws_main.title, last_data_row=len(df) + 1)

    # Tab 3: Duplicates
    if not dup_df.empty:
        # Same diagnostic-column cleanup as the main evidence sheet:
        # dup_df inherits the diagnostic-only columns from reindex so
        # the duplication hotspot is visible here, but they're
        # inappropriate on the Duplicate Entries tab itself.
        dup_df_for_report = dup_df.drop(
            columns=[c for c in _diagnostic_columns_for_evidence_report if c in dup_df.columns],
            errors="ignore",
        )
        ws_dup = wb.create_sheet(title="Duplicate Entries")
        write_evidence_sheet(ws_dup, dup_df_for_report, is_duplicate=True)

    # Tab 4: Filtered
    if filtered and config.get("save_filtered", True):
        ws_filt = wb.create_sheet(title="Filtered (Below Min)")
        filt_headers = ["Source", "Date", "Amount (£)", "Details", "Logic Used", "Reason"]
        for ci, h in enumerate(filt_headers, 1):
            _hcell(ws_filt, 1, ci, h, bg="888888")
        filt_df = pd.DataFrame(filtered).sort_values("Amount (£)", ascending=False)
        for r_idx, frow in enumerate(filt_df.values, 2):
            bg_hex = "F5F5F5" if r_idx % 2 == 0 else None
            for c_idx, val in enumerate(frow, 1):
                c = ws_filt.cell(row=r_idx, column=c_idx, value=val)
                c.font = Font(name="Calibri", size=10)
                c.border = CELL_BORDER
                if bg_hex:
                    c.fill = PatternFill("solid", start_color=bg_hex)
                if c_idx == 3:
                    c.number_format = "£#,##0.00"
        for col, w in zip(["A", "B", "C", "D", "E", "F"], [18, 13, 14, 38, 18, 28], strict=False):
            ws_filt.column_dimensions[col].width = w
        ws_filt.freeze_panes = "A2"

    # Tab 5: Parse errors
    if error_log:
        ws_err = wb.create_sheet(title="Parse Errors")
        _hcell(ws_err, 1, 1, "Time", bg="888888")
        _hcell(ws_err, 1, 2, "Context", bg="888888")
        _hcell(ws_err, 1, 3, "Error", bg="888888")
        for r_idx, entry in enumerate(error_log, 2):
            ts_m = re.match(r"\[(.+?)\]\s*(.*?)\s*—\s*(.*)", entry)
            if ts_m:
                ts, ctx, err = ts_m.group(1), ts_m.group(2), ts_m.group(3)
            else:
                ts, ctx, err = "", entry, ""
            for c_idx, val in enumerate([ts, ctx, err], 1):
                c = ws_err.cell(row=r_idx, column=c_idx, value=val)
                c.font = Font(name="Calibri", size=10)
                c.border = CELL_BORDER
        ws_err.column_dimensions["A"].width = 10
        ws_err.column_dimensions["B"].width = 45
        ws_err.column_dimensions["C"].width = 60

    # =====================================================================
    # ANALYSIS SUITE
    # Uses bills above analysis_min threshold only (payments/credits always included).
    # =====================================================================

    df_an = df.copy()
    df_an["_dt"] = df_an["Date"].apply(parse_to_sort_date)
    df_an = df_an.sort_values("_dt").reset_index(drop=True)
    analysis_min = float(config.get("analysis_min", 500.0))

    # For balance-affecting entries: include all Payments/Credits, but filter
    # New Bill/Ongoing Balance by analysis_min threshold
    payment_credit_mask = df_an["Entry Type"].isin(("Payment", "Credit"))
    bill_mask = df_an["Entry Type"].isin(("New Bill", "Ongoing Balance"))
    amount_mask = df_an["Amount (£)"] >= analysis_min

    dfc = df_an[(payment_credit_mask) | (bill_mask & amount_mask)].copy().reset_index(drop=True)
    dfc["year"] = dfc["_dt"].dt.year
    dfc["month"] = dfc["_dt"].dt.month

    if len(dfc) < 2:
        # Not enough data for analysis sheets; save the workbook with what
        # we have (evidence, summary, duplicates, etc. are already written).
        wb.save(output_path)
        return

    amounts = dfc["Amount (£)"].values.astype(float)
    dates_lbl = dfc["Date"].tolist()
    n = len(amounts)

    raw_diffs = np.diff(amounts)
    pos_diffs = raw_diffs[raw_diffs > 0]

    yearly = (
        dfc.groupby("year")
        .agg(
            count=("Amount (£)", "count"),
            avg_bal=("Amount (£)", "mean"),
            peak=("Amount (£)", "max"),
            low=("Amount (£)", "min"),
        )
        .reset_index()
    )

    # ----- TAB A: KEY STATISTICS -----
    ws_ks = wb.create_sheet(title="Key Statistics")
    ws_ks.column_dimensions["A"].width = 44
    ws_ks.column_dimensions["B"].width = 22
    ws_ks.column_dimensions["C"].width = 44

    tc = ws_ks.cell(row=1, column=1, value="EDF ENERGY DISPUTE  —  KEY STATISTICS")
    tc.font = Font(name="Calibri", size=13, bold=True, color="FFFFFF")
    tc.fill = PatternFill("solid", start_color=ORANGE)
    tc.border = CELL_BORDER
    tc.alignment = Alignment(horizontal="left", vertical="center")
    for c in [2, 3]:
        x = ws_ks.cell(row=1, column=c)
        x.fill = PatternFill("solid", start_color=ORANGE)
        x.border = CELL_BORDER
    ws_ks.row_dimensions[1].height = 26

    def ks_row(r, label, value, note="", fmt=None, bold=False, alt=False):
        bg = LGREY if alt else None
        _text(ws_ks, r, 1, label, bold=bold, fill_hex=bg)
        if fmt == "£":
            _money(ws_ks, r, 2, value, bold=bold, fill_hex=bg)
        elif fmt == "%":
            _num(ws_ks, r, 2, value, fmt="0.0%", bold=bold, fill_hex=bg)
        elif fmt == "date":
            cell = ws_ks.cell(row=r, column=2, value=value)
            cell.number_format = "dd/mm/yyyy"
            cell.font = Font(name="Calibri", size=10, bold=bold)
            cell.border = CELL_BORDER
            cell.alignment = Alignment(horizontal="right", vertical="center")
            if bg:
                cell.fill = PatternFill("solid", start_color=bg)
        elif fmt:
            _num(ws_ks, r, 2, value, fmt=fmt, bold=bold, fill_hex=bg)
        else:
            _text(ws_ks, r, 2, value, bold=bold, fill_hex=bg, align="right")
        _text(ws_ks, r, 3, note, fill_hex=bg, color=DGREY)

    acc_ref = str(config.get("report_account_ref") or config.get("acc_num") or "N/A")

    r = 2
    _section_hdr(ws_ks, r, "ACCOUNT OVERVIEW")
    r = 3
    ks_row(r, "Account reference", acc_ref, alt=True)
    r = 4
    ks_row(
        r,
        "First bill on record",
        "='Balance Trend'!A2",
        fmt="date",
        note="From Balance Trend sheet",
    )
    r = 5
    ks_row(
        r,
        "Most recent bill",
        "=INDEX('Balance Trend'!A:A,MATCH(9.99E+307,'Balance Trend'!B:B)+1)",
        fmt="date",
        alt=True,
    )
    r = 6
    ks_row(
        r,
        "Period covered (days)",
        "=IFERROR(INT(INDEX('Balance Trend'!A:A,MATCH(9.99E+307,'Balance Trend'!B:B)+1)-'Balance Trend'!A2),\"\")",
        fmt="#,##0",
        note="Days between first and last bill",
    )
    r = 7
    ks_row(
        r,
        "Total bills on record",
        "=IFERROR(COUNT('Balance Trend'!B:B),\"\")",
        fmt="#,##0",
        alt=True,
    )

    r = 8
    _section_hdr(ws_ks, r, "BALANCE FIGURES")
    r = 9
    ks_row(
        r,
        "Opening balance (first bill)",
        "='Balance Trend'!B2",
        fmt="£",
        alt=True,
        note="First entry in Balance Trend",
    )
    r = 10
    ks_row(
        r,
        "Current balance (latest bill)",
        "=INDEX('Balance Trend'!B:B,MATCH(9.99E+307,'Balance Trend'!B:B))",
        fmt="£",
        bold=True,
        note="Last numeric entry in Balance Trend",
    )
    r = 11
    ks_row(
        r,
        "Total balance increase",
        '=IFERROR(B10-B9,"")',
        fmt="£",
        bold=True,
        alt=True,
        note="Latest minus earliest",
    )
    r = 12
    ks_row(r, "% increase over full period", '=IFERROR((B10-B9)/B9,"")', fmt="%", bold=True)
    r = 13
    ks_row(
        r,
        "Mean balance across all bills",
        "=IFERROR(AVERAGE('Balance Trend'!B:B),\"\")",
        fmt="£",
        alt=True,
    )
    r = 14
    ks_row(r, "Median balance", "=IFERROR(MEDIAN('Balance Trend'!B:B),\"\")", fmt="£")
    r = 15
    ks_row(r, "Peak balance recorded", "=IFERROR(MAX('Balance Trend'!B:B),\"\")", fmt="£", alt=True)
    r = 16
    ks_row(r, "Lowest balance recorded", "=IFERROR(MIN('Balance Trend'!B:B),\"\")", fmt="£")

    r = 17
    _section_hdr(ws_ks, r, "PERIODIC CHARGES")
    r = 18
    ks_row(
        r,
        "Note",
        "Bills are a running cumulative balance — periodic charge = closing minus opening balance",
        alt=True,
    )
    r = 19
    ks_row(
        r,
        "Mean charge per period (positive only)",
        '=IFERROR(AVERAGEIF(\'Period Charges\'!F:F,">0"),"")',
        fmt="£",
    )
    r = 20
    ks_row(
        r,
        "Largest single-period charge",
        "=IFERROR(MAX('Period Charges'!F:F),\"\")",
        fmt="£",
        bold=True,
        alt=True,
    )
    r = 21
    ks_row(
        r,
        "Smallest positive charge",
        "=IFERROR(_xlfn.MINIFS('Period Charges'!F:F,'Period Charges'!F:F,\">0\"),\"\")",
        fmt="£",
    )
    r = 22
    ks_row(
        r,
        "Periods where balance increased",
        '=IFERROR(COUNTIF(\'Period Charges\'!F:F,">0"),"")',
        fmt="#,##0",
        alt=True,
    )
    r = 23
    ks_row(
        r,
        "Periods where balance fell (payments/credits)",
        '=IFERROR(COUNTIF(\'Period Charges\'!F:F,"<0"),"")',
        fmt="#,##0",
    )
    r = 24
    ks_row(
        r,
        "Implied annual rate (avg last 6 charges ×12)",
        "=IFERROR(AVERAGE(OFFSET('Period Charges'!F1,MAX(1,COUNTIF('Period Charges'!F:F,\">0\")-5),0,6,1))*12,\"\")",
        fmt="£",
        bold=True,
        alt=True,
        note="Assumes ~monthly billing — may overstate if billing is quarterly",
    )

    r = 25
    _section_hdr(ws_ks, r, "READING & DATA QUALITY")
    r = 26
    ks_row(
        r,
        "Estimated readings",
        '=IFERROR(COUNTIF(\'EDF Evidence Report\'!L:L,"Estimated"),"")',
        fmt="#,##0",
        alt=True,
    )
    r = 27
    ks_row(
        r,
        "Actual / customer readings",
        '=IFERROR(COUNTIF(\'EDF Evidence Report\'!L:L,"Actual"),"")',
        fmt="#,##0",
    )
    r = 28
    ks_row(
        r,
        "Smart meter readings",
        '=IFERROR(COUNTIF(\'EDF Evidence Report\'!L:L,"Smart"),"")',
        fmt="#,##0",
        alt=True,
    )
    r = 29
    ks_row(
        r,
        "% of bills with estimated readings",
        "=IFERROR(B26/COUNT('EDF Evidence Report'!G:G),\"\")",
        fmt="%",
    )

    r = 30
    _section_hdr(ws_ks, r, "UNIT RATES")
    r = 31
    ks_row(
        r,
        "Average unit rate (p/kWh)",
        "=IFERROR(AVERAGE('EDF Evidence Report'!I:I),\"\")",
        fmt="0.00",
        alt=True,
        note="Across all bills with valid period charge and kWh",
    )
    r = 32
    ks_row(
        r,
        "Maximum unit rate (p/kWh)",
        "=IFERROR(MAX('EDF Evidence Report'!I:I),\"\")",
        fmt="0.00",
        note="Highest effective rate — potential overcharge",
    )
    r = 33
    ks_row(
        r,
        "Minimum unit rate (p/kWh)",
        "=IFERROR(MIN('EDF Evidence Report'!I:I),\"\")",
        fmt="0.00",
        alt=True,
    )

    ws_ks.freeze_panes = "A2"

    # ----- TAB B: BALANCE TREND -----
    ws_bt = wb.create_sheet(title="Balance Trend")
    for ci, h in enumerate(
        ["Date", "Balance (£)", "6-Bill Rolling Avg (£)", "Linear Trend (£)", "Period Charge (£)"],
        1,
    ):
        _hcell(ws_bt, 1, ci, h, bg=NAVY)
    ws_bt.row_dimensions[1].height = 22

    last_data_row = n + 1
    for i in range(n):
        r = i + 2
        bg = LGREY if i % 2 == 0 else None

        # Write date as a true Excel date serial
        excel_dt = to_excel_date(dates_lbl[i])
        c1 = ws_bt.cell(row=r, column=1, value=excel_dt)
        c1.number_format = "dd/mm/yyyy"
        c1.font = Font(name="Calibri", size=10)
        c1.border = CELL_BORDER
        c1.alignment = Alignment(horizontal="left")
        if bg:
            c1.fill = PatternFill("solid", start_color=bg)

        _money(ws_bt, r, 2, float(amounts[i]), fill_hex=bg)

        start_r = max(2, r - 5)
        for col_i, formula in [
            (3, f'=IFERROR(AVERAGE(B{start_r}:B{r}),"")'),
            (
                4,
                f'=IFERROR(FORECAST.LINEAR(ROW(),B$2:B${last_data_row},ROW(B$2:B${last_data_row})),"")',
            ),
        ]:
            cx = ws_bt.cell(row=r, column=col_i, value=formula)
            cx.number_format = "£#,##0.00"
            cx.font = Font(name="Calibri", size=10)
            cx.border = CELL_BORDER
            cx.alignment = Alignment(horizontal="right")
            if bg:
                cx.fill = PatternFill("solid", start_color=bg)

        if i > 0:
            c5 = ws_bt.cell(row=r, column=5, value=f"=B{r}-B{r - 1}")
            c5.number_format = "£#,##0.00"
            c5.font = Font(name="Calibri", size=10)
            c5.border = CELL_BORDER
            c5.alignment = Alignment(horizontal="right")
            if bg:
                c5.fill = PatternFill("solid", start_color=bg)

    # Line chart
    lc = LineChart()
    lc.title = "Account Balance Over Time"
    lc.style = 10
    lc.y_axis.title = "Balance (£)"
    lc.x_axis.title = "Bill Date"
    lc.width, lc.height = 30, 18
    data_ref = Reference(ws_bt, min_col=2, max_col=4, min_row=1, max_row=n + 1)
    dates_ref = Reference(ws_bt, min_col=1, min_row=2, max_row=n + 1)
    lc.add_data(data_ref, titles_from_data=True)
    lc.set_categories(dates_ref)
    lc.series[0].graphicalProperties.line.solidFill = ORANGE
    lc.series[0].graphicalProperties.line.width = 22000
    if len(lc.series) > 1:
        lc.series[1].graphicalProperties.line.solidFill = NAVY
        lc.series[1].graphicalProperties.line.width = 15000
        lc.series[1].graphicalProperties.line.dashDot = "dash"
    if len(lc.series) > 2:
        lc.series[2].graphicalProperties.line.solidFill = DGREY
        lc.series[2].graphicalProperties.line.width = 10000
        lc.series[2].graphicalProperties.line.dashDot = "sysDash"
    ws_bt.add_chart(lc, "G2")
    for col, w in zip(["A", "B", "C", "D", "E"], [14, 16, 20, 16, 16], strict=False):
        ws_bt.column_dimensions[col].width = w
    ws_bt.freeze_panes = "A2"

    # ----- TAB C: YEAR-ON-YEAR -----
    ws_yoy = wb.create_sheet(title="Year-on-Year")
    for ci, h in enumerate(
        [
            "Year",
            "Bills",
            "Peak Balance (£)",
            "Avg Balance (£)",
            "Lowest Balance (£)",
            "YoY Avg Δ (£)",
            "YoY Avg Δ (%)",
            "Est. Readings",
            "Biggest Jump (£)",
        ],
        1,
    ):
        _hcell(ws_yoy, 1, ci, h, bg=ORANGE)
    ws_yoy.row_dimensions[1].height = 22

    prev_avg = None
    yoy_data = []
    for r_off, row_y in enumerate(yearly.itertuples(), 2):
        yr = row_y.year
        cnt = row_y.count
        pk = row_y.peak
        av = row_y.avg_bal
        lo = row_y.low
        yoy_chg_pct = ((av - prev_avg) / prev_avg) if prev_avg else None

        yr_rows = dfc[dfc["year"] == yr]
        yr_idx = yr_rows.index.tolist()
        max_jump = None
        for ii in yr_idx:
            if ii > 0 and ii in dfc.index and ii - 1 in dfc.index:
                jmp = dfc.at[ii, "Amount (£)"] - dfc.at[ii - 1, "Amount (£)"]
                if max_jump is None or jmp > max_jump:
                    max_jump = jmp

        alt = r_off % 2 == 0
        bg = LGREY if alt else None

        _num(ws_yoy, r_off, 1, yr, fmt="#,##0", fill_hex=bg, bold=True)
        _num(ws_yoy, r_off, 2, cnt, fmt="#,##0", fill_hex=bg)
        _money(ws_yoy, r_off, 3, pk, fill_hex=bg, bold=True)
        _money(ws_yoy, r_off, 4, av, fill_hex=bg)
        _money(ws_yoy, r_off, 5, lo, fill_hex=bg)

        if r_off > 2:
            c6 = ws_yoy.cell(row=r_off, column=6, value=f"=D{r_off}-D{r_off - 1}")
            c6.number_format = "£#,##0.00"
            c6.font = Font(name="Calibri", size=10, bold=True)
            c6.border = CELL_BORDER
            c6.alignment = Alignment(horizontal="right")
            if bg:
                c6.fill = PatternFill("solid", start_color=bg)

            c7 = ws_yoy.cell(row=r_off, column=7, value=f'=IFERROR(F{r_off}/D{r_off - 1},"")')
            c7.number_format = "+0.0%;-0.0%;—"
            c7.font = Font(name="Calibri", size=10, bold=True)
            c7.border = CELL_BORDER
            c7.alignment = Alignment(horizontal="right")
            yoy_fill = (
                RED
                if yoy_chg_pct is not None and yoy_chg_pct > 0.5
                else (
                    AMBER
                    if yoy_chg_pct is not None and yoy_chg_pct > 0.2
                    else (GREEN if yoy_chg_pct is not None and yoy_chg_pct < -0.1 else bg)
                )
            )
            if yoy_fill:
                c7.fill = PatternFill("solid", start_color=yoy_fill)
        else:
            ws_yoy.cell(row=r_off, column=6, value="—").border = CELL_BORDER
            ws_yoy.cell(row=r_off, column=7, value="—").border = CELL_BORDER

        yr_est = (
            int((dfc[dfc["year"] == yr]["Reading"] == "Estimated").sum())
            if "Reading" in dfc.columns
            else 0
        )
        _num(ws_yoy, r_off, 8, yr_est, fmt="#,##0", fill_hex=bg)
        if max_jump is not None:
            _money(ws_yoy, r_off, 9, max_jump, fill_hex=(RED if max_jump > 5000 else bg))

        yoy_data.append((yr, av))
        prev_avg = av

    bc = BarChart()
    bc.type = "col"
    bc.title = "Average Balance by Year"
    bc.y_axis.title = "Average Balance (£)"
    bc.style = 10
    bc.width, bc.height = 22, 14
    n_yrs = len(yoy_data)
    avg_ref = Reference(ws_yoy, min_col=4, min_row=1, max_row=n_yrs + 1)
    yr_ref = Reference(ws_yoy, min_col=1, min_row=2, max_row=n_yrs + 1)
    bc.add_data(avg_ref, titles_from_data=True)
    bc.set_categories(yr_ref)
    bc.series[0].graphicalProperties.solidFill = ORANGE
    ws_yoy.add_chart(bc, "K2")
    for col, w in zip(
        ["A", "B", "C", "D", "E", "F", "G", "H", "I"],
        [8, 8, 18, 18, 18, 16, 14, 14, 18],
        strict=False,
    ):
        ws_yoy.column_dimensions[col].width = w
    ws_yoy.freeze_panes = "A2"

    # ----- TAB D: PERIOD CHARGES -----
    ws_pc = wb.create_sheet(title="Period Charges")
    for ci, h in enumerate(
        [
            "From Date",
            "To Date",
            "Days",
            "Opening Balance (£)",
            "Closing Balance (£)",
            "Charge (£)",
            "Daily Rate (£/day)",
            "Flag",
        ],
        1,
    ):
        _hcell(ws_pc, 1, ci, h, bg=NAVY)
    ws_pc.row_dimensions[1].height = 22

    mean_daily = float(np.mean(pos_diffs)) / 30.0 if len(pos_diffs) else 0
    pc_rows_data = []

    pc_r = 2
    for i in range(1, n):
        p = dfc.iloc[i - 1]
        c_ = dfc.iloc[i]
        days = (c_["_dt"] - p["_dt"]).days
        charge = float(c_["Amount (£)"]) - float(p["Amount (£)"])
        daily = charge / days if days > 0 else None

        flag = ""
        if days > 90:
            flag = f"⚠ {days}-day gap — possible missed bill(s)"
        elif charge < 0:
            flag = f"↓ Balance reduced by £{abs(charge):,.2f} (payment or credit)"
        elif daily and mean_daily > 0 and daily > mean_daily * 2.5:
            flag = f"⚠ Daily rate {daily / mean_daily:.1f}× average"

        bg = LGREY if pc_r % 2 == 0 else None
        if flag.startswith("⚠"):
            bg = AMBER
        elif charge < 0:
            bg = GREEN

        _text(ws_pc, pc_r, 1, p["Date"], fill_hex=bg)
        _text(ws_pc, pc_r, 2, c_["Date"], fill_hex=bg)
        _num(ws_pc, pc_r, 3, days, fmt="#,##0", fill_hex=bg)
        _money(ws_pc, pc_r, 4, float(p["Amount (£)"]), fill_hex=bg)
        _money(ws_pc, pc_r, 5, float(c_["Amount (£)"]), fill_hex=bg)

        c6 = ws_pc.cell(row=pc_r, column=6, value=f"=E{pc_r}-D{pc_r}")
        c6.number_format = "£#,##0.00"
        c6.font = Font(name="Calibri", size=10)
        c6.border = CELL_BORDER
        c6.alignment = Alignment(horizontal="right")
        if bg:
            c6.fill = PatternFill("solid", start_color=bg)

        c7 = ws_pc.cell(row=pc_r, column=7, value=f'=IFERROR(F{pc_r}/C{pc_r},"")')
        c7.number_format = "£#,##0.00"
        c7.font = Font(name="Calibri", size=10)
        c7.border = CELL_BORDER
        c7.alignment = Alignment(horizontal="right")
        if bg:
            c7.fill = PatternFill("solid", start_color=bg)

        _text(ws_pc, pc_r, 8, flag, fill_hex=bg, wrap=True)

        if charge > 0:
            pc_rows_data.append((c_["Date"], charge))
        pc_r += 1

    if pc_r > 2:
        sr = pc_r + 2
        _section_hdr(ws_pc, sr, "SUMMARY STATISTICS", ncols=8, bg=ORANGE)
        sr += 1
        dr = f"F2:F{pc_r - 1}"
        cr = f"C2:C{pc_r - 1}"

        def pc_stat(r, lbl, formula, fmt="£"):
            _text(ws_pc, r, 1, lbl, bold=True, fill_hex=LGREY)
            c = ws_pc.cell(row=r, column=2, value=formula)
            c.font = Font(name="Calibri", size=10, bold=True)
            c.fill = PatternFill("solid", start_color=LGREY)
            c.border = CELL_BORDER
            c.alignment = Alignment(horizontal="right")
            c.number_format = "£#,##0.00" if fmt == "£" else fmt
            for cc in range(3, 9):
                ws_pc.cell(row=r, column=cc).fill = PatternFill("solid", start_color=LGREY)
                ws_pc.cell(row=r, column=cc).border = CELL_BORDER

        pc_stat(sr, "Mean charge per period (positive only)", f'=IFERROR(AVERAGEIF({dr},">0"),"")')
        pc_stat(sr + 1, "Largest single charge", f'=IFERROR(MAX({dr}),"")')
        pc_stat(sr + 2, "Largest credit / reduction", f'=IFERROR(MIN({dr}),"")')
        pc_stat(sr + 3, "Charge periods", f'=IFERROR(COUNTIF({dr},">0"),"")', fmt="#,##0")
        pc_stat(sr + 4, "Credit periods", f'=IFERROR(COUNTIF({dr},"<0"),"")', fmt="#,##0")
        pc_stat(sr + 5, "Average days between bills", f'=IFERROR(AVERAGE({cr}),"")', fmt="#,##0.0")

    if len(pc_rows_data) > 1:
        bc2 = BarChart()
        bc2.type = "col"
        bc2.title = "Charge Added Each Period"
        bc2.y_axis.title = "Charge (£)"
        bc2.style = 10
        bc2.width, bc2.height = 28, 14
        chg_ref2 = Reference(ws_pc, min_col=6, min_row=1, max_row=pc_r - 1)
        date_ref2 = Reference(ws_pc, min_col=2, min_row=2, max_row=pc_r - 1)
        bc2.add_data(chg_ref2, titles_from_data=True)
        bc2.set_categories(date_ref2)
        bc2.series[0].graphicalProperties.solidFill = NAVY
        ws_pc.add_chart(bc2, "J2")

    for col, w in zip(
        ["A", "B", "C", "D", "E", "F", "G", "H"], [13, 13, 7, 18, 18, 16, 14, 42], strict=False
    ):
        ws_pc.column_dimensions[col].width = w
    ws_pc.freeze_panes = "A2"

    # ----- TAB E: DISPUTE FLAGS -----
    ws_df = wb.create_sheet(title="Dispute Flags")

    def _banner(ws, r, text, bg):
        c = ws.cell(row=r, column=1, value=text)
        c.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", start_color=bg)
        c.border = CELL_BORDER
        c.alignment = Alignment(horizontal="left", vertical="center")
        for col in range(2, 7):
            x = ws.cell(row=r, column=col)
            x.fill = PatternFill("solid", start_color=bg)
            x.border = CELL_BORDER
        ws.row_dimensions[r].height = 20

    _banner(ws_df, 1, "EDF ENERGY DISPUTE  —  AUTOMATED ANALYSIS FLAGS", ORANGE)
    ws_df.cell(
        row=2,
        column=1,
        value=f"Generated {datetime.now().strftime('%d/%m/%Y %H:%M')}  |  Period: {dates_lbl[0]} to {dates_lbl[-1]}",
    )
    ws_df.cell(row=2, column=1).font = Font(name="Calibri", size=9, italic=True, color=DGREY)

    for ci, (txt, col_hex) in enumerate(
        [
            ("■ RED = HIGH severity", RED),
            ("■ AMBER = MEDIUM", AMBER),
            ("■ GREEN = Payment/credit", GREEN),
        ],
        1,
    ):
        lc2 = ws_df.cell(row=3, column=ci * 2 - 1, value=txt)
        lc2.font = Font(name="Calibri", size=9, bold=True)
        lc2.fill = PatternFill("solid", start_color=col_hex)
        lc2.border = CELL_BORDER

    hdr_row = 5
    for ci, h in enumerate(["#", "Date", "Balance (£)", "Flag Type", "Detail", "Severity"], 1):
        _hcell(ws_df, hdr_row, ci, h, bg=NAVY)

    flags, counts = compute_dispute_flags(dfc, mean_daily)

    sev_fill = {"HIGH": RED, "MEDIUM": AMBER, "INFO": GREEN}
    for fi, (ftype, flag_date, amt, detail, sev) in enumerate(flags, hdr_row + 1):
        bg = sev_fill.get(sev, LGREY)
        _num(ws_df, fi, 1, fi - hdr_row, fmt="#,##0", fill_hex=bg)
        _text(ws_df, fi, 2, flag_date or "—", fill_hex=bg)
        if amt:
            _money(ws_df, fi, 3, float(amt), fill_hex=bg)
        else:
            ws_df.cell(row=fi, column=3).fill = PatternFill("solid", start_color=bg)
            ws_df.cell(row=fi, column=3).border = CELL_BORDER
        _text(ws_df, fi, 4, ftype, bold=True, fill_hex=bg)
        _text(ws_df, fi, 5, detail, fill_hex=bg, wrap=True)
        _text(ws_df, fi, 6, sev, bold=True, fill_hex=bg, align="center")
        ws_df.row_dimensions[fi].height = 30

    if flags:
        fr = len(flags) + hdr_row + 2
        counts = {s: sum(1 for f in flags if f[4] == s) for s in ("HIGH", "MEDIUM", "INFO")}
        _banner(
            ws_df,
            fr,
            f"TOTAL FLAGS: {len(flags)}   |   HIGH: {counts['HIGH']}   |   MEDIUM: {counts['MEDIUM']}   |   INFO: {counts['INFO']}",
            NAVY,
        )

    for col, w in zip(["A", "B", "C", "D", "E", "F"], [5, 13, 16, 20, 60, 10], strict=False):
        ws_df.column_dimensions[col].width = w
    ws_df.freeze_panes = f"A{hdr_row + 1}"

    # ----- TAB F: DISPUTE TIMELINE -----
    ws_tl = wb.create_sheet(title="Dispute Timeline")
    _banner(ws_tl, 1, "EDF ENERGY DISPUTE  —  CHRONOLOGICAL TIMELINE", ORANGE)
    ws_tl.cell(
        row=2, column=1, value=f"Account: {acc_ref}  |  Period: {dates_lbl[0]} to {dates_lbl[-1]}"
    )
    ws_tl.cell(row=2, column=1).font = Font(name="Calibri", size=9, italic=True, color=DGREY)

    for ci, h in enumerate(["Date", "Event Type", "Description"], 1):
        _hcell(ws_tl, 4, ci, h, bg=NAVY)

    timeline_events = []

    # Bookend: first record
    timeline_events.append(
        (dates_lbl[0], "ACCOUNT START", f"First bill on record. Balance: £{amounts[0]:,.2f}.")
    )

    # Top 5 largest balance jumps
    jumps = []
    for i in range(1, n):
        delta = float(amounts[i]) - float(amounts[i - 1])
        days = (dfc.iloc[i]["_dt"] - dfc.iloc[i - 1]["_dt"]).days
        if delta > 0:
            jumps.append((delta, i, days))
    jumps.sort(key=lambda x: x[0], reverse=True)
    for delta, idx, days in jumps[:5]:
        timeline_events.append(
            (
                dfc.iloc[idx]["Date"],
                "LARGE INCREASE",
                f"Balance rose £{delta:,.2f} in {days} days "
                f"(from £{amounts[idx - 1]:,.2f} to £{amounts[idx]:,.2f}).",
            )
        )

    # Billing gaps > 60 days
    for i in range(1, n):
        days = (dfc.iloc[i]["_dt"] - dfc.iloc[i - 1]["_dt"]).days
        if days > 60:
            timeline_events.append(
                (
                    dfc.iloc[i]["Date"],
                    "BILLING GAP",
                    f"{days} days without a bill (previous: {dfc.iloc[i - 1]['Date']}). "
                    f"Balance accumulated unchecked.",
                )
            )

    # Estimated reading runs (reuse existing detection)
    if "Reading" in dfc.columns:
        run = 0
        run_start_date = None
        for i, rv in enumerate(dfc["Reading"].tolist()):
            if str(rv).lower() in ("estimated", "est."):
                run += 1
                if run == 1:
                    run_start_date = dfc.iloc[i]["Date"]
            else:
                if run >= 3:
                    timeline_events.append(
                        (
                            run_start_date,
                            "ESTIMATED READINGS",
                            f"{run} consecutive bills used estimated meter readings.",
                        )
                    )
                run = 0
                run_start_date = None
        if run >= 3:
            timeline_events.append(
                (
                    run_start_date,
                    "ESTIMATED READINGS",
                    f"{run} consecutive estimated readings (ongoing).",
                )
            )

    # Payment events (balance reductions)
    for i in range(1, n):
        delta = float(amounts[i]) - float(amounts[i - 1])
        if delta < -200:
            timeline_events.append(
                (
                    dfc.iloc[i]["Date"],
                    "PAYMENT/CREDIT",
                    f"Balance reduced by £{abs(delta):,.2f} "
                    f"(from £{amounts[i - 1]:,.2f} to £{amounts[i]:,.2f}).",
                )
            )

    # Reconciliation mismatches (from flags)
    for ftype, fdate, _famt, fdetail, _fsev in flags:
        if ftype == "RECONCILIATION MISMATCH":
            timeline_events.append((fdate, "RECONCILIATION", fdetail))

    # Bookend: latest record
    timeline_events.append(
        (
            dates_lbl[-1],
            "CURRENT STATE",
            f"Latest bill on record. Balance: £{amounts[-1]:,.2f}. "
            f"Total increase from first record: £{amounts[-1] - amounts[0]:,.2f}.",
        )
    )

    # Sort by date and write
    timeline_events.sort(key=lambda e: parse_to_sort_date(e[0]) or pd.Timestamp.min)
    tl_r = 5
    for tl_date, etype, desc in timeline_events:
        bg_hex = LGREY if tl_r % 2 == 0 else None
        _text(ws_tl, tl_r, 1, tl_date, fill_hex=bg_hex)
        _text(ws_tl, tl_r, 2, etype, bold=True, fill_hex=bg_hex)
        _text(ws_tl, tl_r, 3, desc, fill_hex=bg_hex, wrap=True)
        ws_tl.row_dimensions[tl_r].height = 40
        tl_r += 1

    for col, w in zip(["A", "B", "C"], [14, 22, 90], strict=False):
        ws_tl.column_dimensions[col].width = w
    ws_tl.freeze_panes = "A5"

    # =====================================================================
    # NEW ANALYSIS TABS (added after Dispute Timeline)
    # =====================================================================

    # Statistical Analysis
    write_statistical_analysis_sheet(wb.create_sheet(title="Statistical Analysis"), dfc, config)

    # Payment Analysis
    write_payment_analysis_sheet(wb.create_sheet(title="Payment Analysis"), dfc)

    # Forecast & Projection
    write_forecast_sheet(wb.create_sheet(title="Forecast & Projection"), dfc)

    # Data Quality Report
    write_data_quality_sheet(wb.create_sheet(title="Data Quality Report"), df)

    # Tariff Analysis (if data available)
    write_tariff_analysis_sheet(wb.create_sheet(title="Tariff Analysis"), dfc)

    # ------------------------------------------------------------------
    # Phase-2 analysis tabs (back-billing, rebilling, meter rollover,
    # contract history). run_analysers runs the four detectors on the
    # same `dfc` (post-dedup, post-filter) the rest of the workbook
    # uses, then each writer paints the result onto its own tab.
    # The new tabs append AFTER the existing 16 -- no existing sheet
    # is touched. Account label is pulled from config['acc_num'].
    # ------------------------------------------------------------------
    account_label = str(config.get("acc_num", "") or "")
    analyses = run_analysers(dfc)
    rb = analyses["rebilling"]
    overlapping_invoices: set[str] = (
        {str(x) for x in rb["Killer Invoice"].tolist()} if not rb.empty else set()
    )
    write_back_billing_sheet(
        wb.create_sheet(title="Back-billing Analysis"),
        analyses["back_billing"],
        account=account_label,
        overlapping_invoices=overlapping_invoices,
        evidence_df=dfc,
        evidence_index=analyses["evidence_index"],
    )
    write_rebilling_sheet(
        wb.create_sheet(title="Rebilling & Corrections"),
        analyses["rebilling"],
        account=account_label,
        evidence_df=dfc,
        evidence_index=analyses["evidence_index"],
    )
    write_meter_readings_sheet(
        wb.create_sheet(title="Meter Readings"),
        dfc,
        analyses["meter_rollover"],
        account=account_label,
        evidence_df=dfc,
        evidence_index=analyses["evidence_index"],
    )
    write_contract_history_sheet(
        wb.create_sheet(title="Contract History"),
        analyses["contracts"],
        account=account_label,
        evidence_df=dfc,
        evidence_index=analyses["evidence_index"],
    )

    # Stream P1 + P2: SAP CSV-in-PDF data dumps and the cross-source
    # Reconciliation sheet. When ``sap_rows`` is supplied (from the
    # engine's three SAP-row accumulators) and the user hasn't opted
    # out via ``config["scan_sap_dumps"] = False``, emit the three SAP
    # sheets. The Reconciliation sheet additionally honours
    # ``config["generate_reconciliation_sheet"]`` (default True) so a
    # reviewer can toggle it off independently when only the SAP data
    # is wanted.
    sap_rows = sap_rows or {}
    sap_contract = list(sap_rows.get("contract") or [])
    sap_meter = list(sap_rows.get("meter") or [])
    sap_financial = list(sap_rows.get("financial") or [])
    scan_sap_dumps = config.get("scan_sap_dumps", True)
    if scan_sap_dumps and (sap_contract or sap_meter or sap_financial):
        if sap_contract:
            write_sap_contract_history_sheet(
                wb.create_sheet(title="SAP Contract History"),
                sap_contract,
                account=account_label,
            )
        if sap_meter:
            write_sap_meter_readings_sheet(
                wb.create_sheet(title="SAP Meter Readings"),
                sap_meter,
                account=account_label,
            )
        if sap_financial:
            write_sap_financial_transactions_sheet(
                wb.create_sheet(title="SAP Financial Transactions"),
                sap_financial,
                account=account_label,
            )
            # SAP Back-billing analyser (spec §6):
            # uses the EDF Evidence Report rows (filter/dedup-applied
            # ``dfc``) as the join target. Both new sheets appear under
            # the existing ``scan_sap_dumps`` toggle alongside the
            # existing SAP sheets.
            edf_records_for_bb: list[dict] = []
            if dfc is not None and not dfc.empty:
                edf_records_for_bb = dfc.to_dict(orient="records")
            bb_events = detect_sap_back_billing_events(sap_financial)
            bb_matches = match_sap_events_to_edf(bb_events, edf_records_for_bb)
            # Populate Sheet 1's "Matched EDF Invoice #" column with the
            # highest-confidence match per event (tiebreak: smallest
            # date_delta_days).
            for ev in bb_events:
                ev_matches = [m for m in bb_matches if m.event is ev]
                if ev_matches:
                    conf_rank = {"High": 0, "Medium": 1, "Low": 2}
                    best = sorted(
                        ev_matches,
                        key=lambda m: (
                            conf_rank.get(m.confidence_band, 3),
                            m.date_delta_days,
                        ),
                    )[0]
                    ev.matched_edf_invoice = str(best.edf_record.get("Invoice #", "") or "") or None
            write_sap_back_billing_sheets(
                wb,
                bb_events,
                bb_matches,
                sap_financial_first_row=4,
                edf_rows=edf_records_for_bb,
                edf_sheet_name="EDF Evidence Report",
                edf_first_row=4,
                account=account_label,
                sap_row_index_map=_build_sap_row_index_map(sap_financial),
            )
        if config.get("generate_reconciliation_sheet", True):
            ws_recon_summary = wb.create_sheet(title="Reconciliation")
            ws_recon_detail = wb.create_sheet(title="Reconciliation Drill-down")
            write_reconciliation_sheet(
                ws_recon_summary,
                ws_recon_detail,
                sap_contract,
                analyses["contracts"],
                sap_meter,
                dfc,
                sap_financial,
                dfc,
                account=account_label,
            )

    _SEVERITY_LED_ORDER = [
        "Annual Summary",
        "EDF Evidence Report",
        "SAP Financial Transactions",
        "SAP Back-billing Events",
        "SAP ↔ EDF Matched Events",
        "SAP Meter Readings",
        "SAP Contract History",
        "Back-billing Analysis",
        "Rebilling & Corrections",
        "Meter Readings",
        "Contract History",
        "Reconciliation",
        "Reconciliation Drill-down",
        "Dispute Flags",
        "Dispute Timeline",
        "Period Charges",
        "Payment Analysis",
        "Balance Trend",
        "Year-on-Year",
        "Key Statistics",
        "Statistical Analysis",
        "Forecast & Projection",
        "Tariff Analysis",
        "Data Quality Report",
        "Duplicate Entries",
        "Filtered (Below Min)",
        "Parse Errors",
    ]
    wb._sheets = [wb[name] for name in _SEVERITY_LED_ORDER if name in wb.sheetnames]

    wb.save(output_path)


# =====================================================================
# NEW ANALYSIS FUNCTIONS (pandas-powered enhancements)
from edf_bill_fetcher.io.writers.back_billing import (  # noqa: E402,F401
    _assess_reason,
    detect_back_billing,
    write_back_billing_sheet,
)
from edf_bill_fetcher.io.writers.data_quality import (  # noqa: E402,F401
    write_data_quality_sheet,
)
from edf_bill_fetcher.io.writers.forecast import (  # noqa: E402,F401
    write_forecast_sheet,
)
from edf_bill_fetcher.io.writers.meter import (  # noqa: E402,F401
    detect_meter_rollover,
    infer_contracts,
    write_contract_history_sheet,
    write_meter_readings_sheet,
)
from edf_bill_fetcher.io.writers.payment import (  # noqa: E402,F401
    write_payment_analysis_sheet,
)
from edf_bill_fetcher.io.writers.rebilling import (  # noqa: E402,F401
    _reversal_match,
    detect_rebilling,
    write_rebilling_sheet,
)
from edf_bill_fetcher.io.writers.statistical import (  # noqa: E402,F401
    write_statistical_analysis_sheet,
)
from edf_bill_fetcher.io.writers.tariff import (  # noqa: E402,F401
    write_tariff_analysis_sheet,
)

# ---------------------------------------------------------------------------
# Back-billing, rebilling, meter-rollover, and contract-inference analysis.
# Pure-pandas detectors (``detect_*``) feed thin sheet writers
# (``write_*_sheet``).  Both halves keep the existing module style: a
# detector returns a tidy DataFrame, a writer paints it onto an
# openpyxl worksheet using the shared cell helpers above.
# ---------------------------------------------------------------------------


# ---------------------------------------------------------------------------
# A ``dict[str, int]`` mapping per-row signatures to the 1-indexed Excel row
# on the ``EDF Evidence Report`` sheet so the 4 analyser writers can emit a
# ``View on Evidence Report`` hotlink on each row. Two signatures per row are
# emitted:
#   - ``inv:<Invoice #>`` -- exact Invoice # match, used when the analyser
#     DataFrame carries the Invoice # in its key column.
#   - ``amt_days:<Amount £ with 2dp>|<days>`` -- fallback signature keyed on
#     the diagnostic pair `` (£, days-billed)``, used when Invoice # is N/A
#     or otherwise unparseable. ``setdefault`` keeps the first hit.














def run_analysers(df: pd.DataFrame) -> dict[str, Any]:
    """Run all Phase-2 detection analyses on the deduplicated
    DataFrame and return their outputs in a dict.

    The orchestrator is a thin wrapper so :func:`export_to_excel` can
    call four detectors with one line and downstream tests can
    inspect the full set without re-running each individually.

    Returns:
        dict with keys ``back_billing``, ``rebilling``,
        ``meter_rollover``, ``contracts``, ``evidence_index``. The
        first four are tidy DataFrames; ``evidence_index`` is a
        ``dict[str, int]`` mapping per-row signatures to the Excel row
        on the ``EDF Evidence Report`` sheet so the analyser tabs can
        emit a ``View on Evidence Report`` hotlink.
    """
    import edf_collector as _edc

    return {
        "back_billing": _edc.detect_back_billing(df),
        "rebilling": _edc.detect_rebilling(df, evidence_df=df),
        "meter_rollover": _edc.detect_meter_rollover(df),
        "contracts": _edc.infer_contracts(df),
        "evidence_index": build_evidence_index(df, header_row_offset=1),
    }






def write_sap_contract_history_sheet(
    ws: Worksheet,
    rows: list[dict],
    account: str = "",
) -> None:
    """Render the SAP Contract History tab.

    Layout:
      row 1: title banner noting SAP/Kraken origin
      row 2: empty
      row 3: column headers (9 cols)
      rows 4+: one row per contract
    """
    ws.title = "SAP Contract History"
    ORANGE = "FE5716"
    columns = [
        "Contract From",
        "Contract To",
        "Product Code",
        "Product Description",
        "Contract Reason",
        "Set Up By",
        "Notes",
        "Cancelled Flag",
        "Source File",
    ]
    ncol = len(columns)

    title = "EDF SAP CONTRACT AND PRODUCT CHANGE HISTORY"
    if account:
        title = f"{title}  |  Account {account}"
    t1 = ws.cell(row=1, column=1, value=title)
    t1.font = Font(name="Calibri", size=13, bold=True, color="FFFFFF")
    t1.fill = PatternFill("solid", start_color=ORANGE)
    t1.border = CELL_BORDER
    for c in range(2, ncol + 1):
        x = ws.cell(row=1, column=c)
        x.fill = PatternFill("solid", start_color=ORANGE)
        x.border = CELL_BORDER
    ws.row_dimensions[1].height = 22
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncol)

    ws.row_dimensions[2].height = 4

    _write_sap_header_row(ws, row=3, columns=columns)

    for i, row in enumerate(rows):
        r = 4 + i
        for j, col in enumerate(columns):
            cell = ws.cell(row=r, column=j + 1, value=row.get(col, ""))
            cell.font = Font(name="Calibri", size=10)
            cell.border = CELL_BORDER
            if i % 2 == 0:
                cell.fill = PatternFill("solid", start_color=EDF_OFFWHITE.lstrip("#"))


def write_sap_meter_readings_sheet(
    ws: Worksheet,
    rows: list[dict],
    account: str = "",
) -> None:
    """Render the SAP Meter Readings tab.

    Layout:
      row 1: banner
      row 2: legend
      row 3: column headers (9 cols)
      rows 4+: data
    """
    ws.title = "SAP Meter Readings"
    ORANGE = "FE5716"
    columns = [
        "Scheduled Read Date",
        "Meter Read Date",
        "Meter Read Reason",
        "Reading (kWh)",
        "Read Type",
        "Read Source",
        "Read Status",
        "Register",
        "Source File",
    ]
    ncol = len(columns)

    title = "EDF SAP METER-READ HISTORY"
    if account:
        title = f"{title}  |  Account {account}"
    t1 = ws.cell(row=1, column=1, value=title)
    t1.font = Font(name="Calibri", size=13, bold=True, color="FFFFFF")
    t1.fill = PatternFill("solid", start_color=ORANGE)
    t1.border = CELL_BORDER
    for c in range(2, ncol + 1):
        x = ws.cell(row=1, column=c)
        x.fill = PatternFill("solid", start_color=ORANGE)
        x.border = CELL_BORDER
    ws.row_dimensions[1].height = 22
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncol)

    l2 = ws.cell(row=2, column=1, value="A = Actual (supplier-confirmed)  |  E = Estimated")
    l2.font = Font(name="Calibri", size=9, italic=True, color=MEDIUM_GREY.lstrip("#"))
    for c in range(2, ncol + 1):
        ws.cell(row=2, column=c).border = CELL_BORDER

    _write_sap_header_row(ws, row=3, columns=columns)

    for i, row in enumerate(rows):
        r = 4 + i
        for j, col in enumerate(columns):
            c = j + 1
            val = row.get(col, "")
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = Font(name="Calibri", size=10)
            cell.border = CELL_BORDER
            if i % 2 == 0:
                cell.fill = PatternFill("solid", start_color=EDF_OFFWHITE.lstrip("#"))
            if col == "Read Type":
                if val == "A":
                    cell.font = Font(name="Calibri", size=10, bold=True, color="006100")
                    cell.fill = PatternFill("solid", start_color="C6EFCE")
                if val == "E":
                    cell.font = Font(
                        name="Calibri", size=10, italic=True, color=MEDIUM_GREY.lstrip("#")
                    )
                    cell.fill = PatternFill("solid", start_color=EDF_OFFWHITE.lstrip("#"))


def write_sap_financial_transactions_sheet(
    ws: Worksheet,
    rows: list[dict],
    account: str = "",
) -> None:
    """Render the SAP Financial Transactions tab."""
    ws.title = "SAP Financial Transactions"
    ORANGE = "FE5716"
    columns = list(rows[0].keys()) if rows else ["Source File"]
    ncol = len(columns)

    title = "EDF SAP FINANCIAL TRANSACTIONS"
    if account:
        title = f"{title}  |  Account {account}"
    t1 = ws.cell(row=1, column=1, value=title)
    t1.font = Font(name="Calibri", size=13, bold=True, color="FFFFFF")
    t1.fill = PatternFill("solid", start_color=ORANGE)
    t1.border = CELL_BORDER
    for c in range(2, ncol + 1):
        x = ws.cell(row=1, column=c)
        x.fill = PatternFill("solid", start_color=ORANGE)
        x.border = CELL_BORDER
    ws.row_dimensions[1].height = 22
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncol)

    ws.row_dimensions[2].height = 4

    _write_sap_header_row(ws, row=3, columns=columns)

    for i, row in enumerate(rows):
        r = 4 + i
        for j, col in enumerate(columns):
            cell = ws.cell(row=r, column=j + 1, value=row.get(col, ""))
            cell.font = Font(name="Calibri", size=10)
            cell.border = CELL_BORDER
            if i % 2 == 0:
                cell.fill = PatternFill("solid", start_color=EDF_OFFWHITE.lstrip("#"))


def _write_sap_header_row(ws: Worksheet, row: int, columns: list) -> None:
    NAVY = "10367A"
    for j, col in enumerate(columns):
        cell = ws.cell(row=row, column=j + 1, value=col)
        cell.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", start_color=NAVY)
        cell.border = CELL_BORDER
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)


# ---------------------------------------------------------------------------
# SAP Back-billing sheet writers (spec: 2026-07-21-sap-back-billing-analysis-design.md)
# ---------------------------------------------------------------------------
#
# Two adjacent sheets:
# 1. "SAP Back-billing Events" (Sheet 1) — one summary row per SAP
#    Clearing Document cluster, with the underlying SAP rows rendered
#    as collapsed outline groups beneath each summary.
# 2. "SAP ↔ EDF Matched Events" (Sheet 2) — one row per (SAP event ×
#    matched EDF invoice candidate), only High/Medium/Low confidence
#    pairs included.
#
# Hyperlinks cross-reference each other, the source SAP Financial
# Transactions sheet (using the first underlying row's row index),
# and the matched EDF Evidence Report row (using the pre-dedup index
# into ``evidence_rows`` passed by the caller).

ORANGE = "FE5716"
NAVY_BLUE = "10367A"
SAP_BB_SUMMARY_FILL_PAIR = ("EFF4FB", "ffffff")
SAP_BB_DETAIL_FILL_PAIR = ("F8FAFC", "ffffff")
SAP_BB_MEDIUM_BORDER = Side(style="medium", color="10367A")


def _bb_invoice_value(rec: dict, key: str) -> object:
    """Look up an invoice field, tolerating 'N/A' / None."""
    v = rec.get(key)
    if v in (None, "", "N/A", "None"):
        return ""
    return v


def write_sap_back_billing_sheets(
    wb: openpyxl.Workbook,
    events: list[SapBackBillingEvent],
    matches: list[SapEdfMatch],
    sap_financial_first_row: int,
    edf_rows: list[dict],
    edf_sheet_name: str = "EDF Evidence Report",
    edf_first_row: int = 4,
    account: str = "",
    sap_row_index_map: dict[int, int] | None = None,
) -> tuple[Worksheet, Worksheet]:
    """Write both new sheets to the workbook (spec §4).

    Returns ``(sheet1_ws, sheet2_ws)`` so callers can pass them around
    if needed (e.g. the GUI's status line).
    """
    ws1 = wb.create_sheet(title="SAP Back-billing Events")
    ws2 = wb.create_sheet(title="SAP ↔ EDF Matched Events")
    _write_sap_bb_events_sheet(
        ws1,
        events,
        sap_financial_first_row,
        sap_row_index_map=sap_row_index_map,
        account=account,
    )
    _write_sap_bb_matches_sheet(
        ws2,
        events,
        matches,
        ws1,
        edf_rows,
        edf_sheet_name=edf_sheet_name,
        edf_first_row=edf_first_row,
        account=account,
    )
    return ws1, ws2


def _write_sap_bb_events_sheet(
    ws: Worksheet,
    events: list[SapBackBillingEvent],
    sap_financial_first_row: int,
    sap_row_index_map: dict[int, int] | None = None,
    account: str = "",
) -> None:
    """Sheet 1 of the SAP Back-billing pair."""
    ws.title = "SAP Back-billing Events"
    # Outline groups sit ABOVE (summary at top of each cluster).
    ws.sheet_properties.outlinePr.summaryBelow = False

    summary_cols = [
        "Clearing Doc #",
        "Clearing Date",
        "Clearing Reason",
        "# SAP Rows",
        "Net Amount (£)",
        "Has Cr-Credit for Consum Billing?",
        "Has Account Maintenance?",
        "Largest Single Posting (£)",
        "Posting Date Range",
        "Evidence Trail",
        "Matched EDF Invoice #",
        "Link to SAP Financial Transactions",
    ]
    ncol = len(summary_cols)

    # Spec §3.3 — title summary counts (replaces the deleted legal block).
    net_zero_count = sum(1 for ev in events if abs(ev.net_amount) < 1.0)
    with_credit_count = sum(1 for ev in events if ev.has_credit_for_consum_billing)
    event_count = len(events)

    # ---------- rows 1-3: banner + spacer + header ----------
    # Spec §3.3 — the legal-context block (was row 3) and the italic
    # intro paragraph (was row 5) are entirely REMOVED. The title row
    # now carries the event-count summary; header moves from row 7 to
    # row 3; data from row 8 to row 4; freeze panes from A8 to A4.
    title = (
        "EDF SAP BACK-BILLING EVENTS  |  Account {acc}  |  "
        "{n} events ({nz} net-zero, {wc} with credit)"
    ).format(
        acc=account or "(no account)",
        n=event_count,
        nz=net_zero_count,
        wc=with_credit_count,
    )
    t1 = ws.cell(row=1, column=1, value=title)
    t1.font = Font(name="Calibri", size=13, bold=True, color="FFFFFF")
    t1.fill = PatternFill("solid", start_color=ORANGE)
    t1.border = CELL_BORDER
    for c in range(2, ncol + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = PatternFill("solid", start_color=ORANGE)
        cell.border = CELL_BORDER
    ws.row_dimensions[1].height = 22
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncol)

    ws.row_dimensions[2].height = 4

    # Row 3: header (was row 7 per spec §3.3)
    _write_sap_header_row(ws, row=3, columns=summary_cols)

    r = 4  # was 8 — header moved from row 7 to row 3 (spec §3.3)
    # Need to look up the index of each underlying row inside the
    # global sap_financial_rows list to point at the SAP Financial
    # Transactions sheet. The caller passes sap_financial_first_row
    # (typically 4, matching the existing writer's first_row=4).
    # We rely on the events' rows retaining insertion order so we
    # can show the first underlying row's hyperlink.

    for ev_i, ev in enumerate(events):
        # Alternating summary tint per event (option C)
        summary_fill = SAP_BB_SUMMARY_FILL_PAIR[ev_i % 2]
        # ----- summary row -----
        cd_iso = (
            pd.Timestamp(ev.clearing_date).strftime("%Y-%m-%d")
            if not pd.isna(ev.clearing_date)
            else ""
        )
        posting_range = (
            f"{ev.posting_date_range[0]} … {ev.posting_date_range[1]}"
            if ev.posting_date_range[0] and ev.posting_date_range[1]
            else ""
        )
        summary_vals = [
            ev.clearing_doc,
            cd_iso,
            ev.clearing_reason,
            len(ev.rows),
            ev.net_amount,
            "Yes" if ev.has_credit_for_consum_billing else "No",
            "Yes" if ev.has_account_maintenance else "No",
            ev.largest_single_posting,
            posting_range,
            ev.evidence_trail,
            ev.matched_edf_invoice or "",
        ]
        # Write values
        for j, val in enumerate(summary_vals, start=1):
            cell = ws.cell(row=r, column=j, value=val)
            cell.font = Font(name="Calibri", size=10, bold=True)
            cell.border = CELL_BORDER
            cell.fill = PatternFill("solid", start_color=summary_fill)
            cell.alignment = Alignment(
                horizontal="left" if j == 10 else "center", vertical="top", wrap_text=True
            )
        # Last column: hyperlink (added below once we know target row)
        # Add medium top border on the first event visible in a band
        # — actually we want a top border on every summary row to
        # visually separate events.
        for j in range(1, ncol + 1):
            cell = ws.cell(row=r, column=j)
            existing = cell.border
            cell.border = Border(
                left=existing.left,
                right=existing.right,
                bottom=existing.bottom,
                top=SAP_BB_MEDIUM_BORDER,
            )
        summary_row_idx = r
        r += 1

        # ----- underlying rows (outline level 1, hidden by default) -----
        detail_fill_pair_idx = ev_i % 2
        for k, row_dict in enumerate(ev.rows):
            # Mirror the visible columns of the SAP Financial Transactions sheet.
            # We pick a reduced column set that mirrors what readers will recognise.
            detail_cols = [
                "Document No.",
                "Item",
                "Document Date",
                "Posting Date",
                "Net Due Date",
                "Main Transaction",
                "Sub Transaction",
                "Transaction Text",
                "Amount",
                "Clearing Status",
                "Clearing Document",
                "Clearing Date",
                "Clearing Reason",
                "Document Type",
                "Document Type Description",
            ]
            # Pad with blanks where this row's columns run out vs ncol header.
            for j, col_name in enumerate(detail_cols, start=1):
                v = row_dict.get(col_name, "")
                cell = ws.cell(row=r, column=j, value=v)
                cell.font = Font(name="Calibri", size=9, color="333333")
                cell.fill = PatternFill(
                    "solid",
                    start_color=SAP_BB_DETAIL_FILL_PAIR[(detail_fill_pair_idx + k) % 2],
                )
            # Blank-fill remaining cells so the row visually ends at the
            # right margin (otherwise alternating-band widths match the
            # summary band, which has more columns than detail).
            for j in range(len(detail_cols) + 1, ncol + 1):
                cell = ws.cell(row=r, column=j)
                cell.fill = PatternFill(
                    "solid",
                    start_color=SAP_BB_DETAIL_FILL_PAIR[(detail_fill_pair_idx + k) % 2],
                )
            ws.row_dimensions[r].outline_level = 1
            ws.row_dimensions[r].hidden = True
            r += 1
        # Link to the first underlying row on the SAP Financial Transactions sheet,
        # using the actual Excel row of that underlying row (via ``id(row)`` lookup
        # into the caller-provided ``sap_row_index_map``).
        first_doc = (ev.rows[0].get("Document No.") or "") if ev.rows else ""
        cell = ws.cell(row=summary_row_idx, column=ncol, value="→")
        cell.font = Font(name="Calibri", size=10, color="0563C1", underline="single")
        cell.alignment = Alignment(horizontal="center", vertical="top")
        if sap_row_index_map is not None and ev.rows:
            target_excel_row = sap_row_index_map.get(id(ev.rows[0]), sap_financial_first_row)
        else:
            target_excel_row = sap_financial_first_row
        summary_cell_location = f"'SAP Financial Transactions'!A{target_excel_row}"
        cell.hyperlink = openpyxl.worksheet.hyperlink.Hyperlink(
            ref=cell.coordinate,
            location=summary_cell_location,
            display="→",
            tooltip=f"Jump to SAP Financial Transactions row {target_excel_row} (DOC {first_doc})",
        )

    # Widths
    widths = [18, 12, 28, 9, 13, 14, 14, 14, 22, 60, 18, 20]
    for i, w in enumerate(widths, start=1):
        col_letter = openpyxl.utils.get_column_letter(i)
        ws.column_dimensions[col_letter].width = w

    # Freeze top header
    ws.freeze_panes = "A4"  # was A8 — header moved to row 3 (spec §3.3)


def _write_sap_bb_matches_sheet(
    ws: Worksheet,
    events: list[SapBackBillingEvent],
    matches: list[SapEdfMatch],
    sheet1: Worksheet,
    edf_rows: list[dict],
    edf_sheet_name: str = "EDF Evidence Report",
    edf_first_row: int = 4,
    account: str = "",
) -> None:
    """Sheet 2 of the SAP Back-billing pair."""
    ws.title = "SAP ↔ EDF Matched Events"

    header_cols = [
        "SAP Clearing Doc #",
        "SAP Clearing Date",
        "SAP Event Net Amount (£)",
        "EDF Invoice #",
        "EDF Bill Date",
        "EDF Period From → To",
        "EDF Invoice Amount (£)",
        "Amount Δ (£)",
        "Date Δ (days)",
        "Match Confidence",
        "Has Cr-Credit for Consum Billing?",
        "Notes",
    ]
    ncol = len(header_cols)

    # Rows 1-3: banner + intro
    title = "SAP ↔ EDF MATCHED EVENTS"
    if account:
        title = f"{title}  |  Account {account}"
    t1 = ws.cell(row=1, column=1, value=title)
    t1.font = Font(name="Calibri", size=13, bold=True, color="FFFFFF")
    t1.fill = PatternFill("solid", start_color=ORANGE)
    t1.border = CELL_BORDER
    for c in range(2, ncol + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = PatternFill("solid", start_color=ORANGE)
        cell.border = CELL_BORDER
    ws.row_dimensions[1].height = 22
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncol)

    ws.row_dimensions[2].height = 4

    intro = (
        "Each row pairs a SAP back-billing event (sheet 'SAP Back-billing\n"
        "Events') with the EDF invoice(s) whose period overlaps the\n"
        "event's clear date. Confidence is computed from date proximity\n"
        "and amount agreement; Low-confidence rows may be coincidental.\n"
        "Click the SAP Clearing Doc link to view the event's underlying\n"
        "SAP rows on sheet 'SAP Back-billing Events'. Click the EDF\n"
        "Invoice # link to view the matched row on the EDF Evidence\n"
        "Report."
    )
    intro_cell = ws.cell(row=3, column=1, value=intro)
    intro_cell.font = Font(name="Calibri", size=10, italic=True)
    intro_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=ncol)
    ws.row_dimensions[3].height = 90

    # Row 5: header
    _write_sap_header_row(ws, row=5, columns=header_cols)

    # Sort matches: by Clearing Date asc, then by score desc within date.
    def sort_key(m: SapEdfMatch) -> tuple:
        cd = m.event.clearing_date
        cd_iso = pd.Timestamp(cd).strftime("%Y-%m-%d") if not pd.isna(cd) else "9999"
        # Score desc → negate.
        conf_order = {"High": 0, "Medium": 1, "Low": 2}.get(m.confidence_band, 3)
        return (cd_iso, conf_order, -m.confidence_score)

    sorted_matches = sorted(matches, key=sort_key)

    # Map Clearing Doc -> summary row on Sheet 1 so we can link back.
    # Summary row index = 4 + sum_{j<i}(1 + len(events[j].rows))
    # (summary row + N underlying rows per event). Header row is row 3
    # and the first summary row is row 4 per spec §3.3.
    summary_row_for_doc: dict[str, int] = {}
    sheet1_summary_row = 4
    for ev in events:
        summary_row_for_doc[ev.clearing_doc] = sheet1_summary_row
        sheet1_summary_row += 1 + len(ev.rows)

    confidence_band_fill = {
        "High": "E5F5E5",  # pale green
        "Medium": "FFF4D6",  # pale amber
        "Low": "F2EAEA",  # pale grey-pink
    }

    r = 6
    for m in sorted_matches:
        ev = m.event
        rec = m.edf_record
        # Build EDF period From→To display
        pf = rec.get("Period From")
        pt = rec.get("Period To")
        period_span = ""
        if pf and pt:
            period_span = f"{pf} → {pt}"
        elif pt:
            period_span = f"? → {pt}"
        elif pf:
            period_span = f"{pf} → ?"
        cd_iso = (
            pd.Timestamp(ev.clearing_date).strftime("%Y-%m-%d")
            if not pd.isna(ev.clearing_date)
            else ""
        )
        edf_amt_val = rec.get("Amount (£)")
        vals = [
            ev.clearing_doc,
            cd_iso,
            ev.net_amount,
            rec.get("Invoice #", ""),
            rec.get("Date", ""),
            period_span,
            edf_amt_val if edf_amt_val not in (None, "") else "",
            m.amount_delta,
            m.date_delta_days,
            m.confidence_band,
            "Yes" if ev.has_credit_for_consum_billing else "No",
            m.notes,
        ]
        row_fill = confidence_band_fill.get(m.confidence_band, "")
        for j, v in enumerate(vals, start=1):
            cell = ws.cell(row=r, column=j, value=v)
            cell.font = Font(name="Calibri", size=10)
            cell.border = CELL_BORDER
            if row_fill:
                cell.fill = PatternFill("solid", start_color=row_fill)
            cell.alignment = Alignment(
                horizontal="left" if j in (6, 12) else "center",
                vertical="center",
                wrap_text=(j == 12),
            )
        # Hyperlinks:
        # col 1 (SAP Clearing Doc #) -> Sheet 1 summary row for this event
        sap_target_row = summary_row_for_doc.get(ev.clearing_doc, 8)
        c1 = ws.cell(row=r, column=1)
        summary_loc = f"'SAP Back-billing Events'!A{sap_target_row}"
        c1.hyperlink = openpyxl.worksheet.hyperlink.Hyperlink(
            ref=c1.coordinate,
            location=summary_loc,
            display=ev.clearing_doc,
            tooltip=f"Jump to SAP event summary row {sap_target_row}",
        )
        c1.font = Font(name="Calibri", size=10, color="0563C1", underline="single")
        # col 4 (EDF Invoice #) -> EDF Evidence Report row for this invoice
        # The row index on EDF Evidence Report = edf_first_row + position in edf_rows.
        # Need to find the index. We use the edf_records list (which match_sap_events_to_edf
        # was given) — the caller has passed the same list as edf_rows.
        edf_idx = None
        for i, er in enumerate(edf_rows):
            if er is rec:
                edf_idx = i
                break
        if edf_idx is None:
            # Try matching by Invoice # as a fallback
            target_inv = str(rec.get("Invoice #", "")).strip()
            if target_inv:
                for i, er in enumerate(edf_rows):
                    if str(er.get("Invoice #", "")).strip() == target_inv:
                        edf_idx = i
                        break
        if edf_idx is not None:
            edf_target_row = edf_first_row + edf_idx
            c4 = ws.cell(row=r, column=4)
            c4_loc = f"'{edf_sheet_name}'!A{edf_target_row}"
            c4.hyperlink = openpyxl.worksheet.hyperlink.Hyperlink(
                ref=c4.coordinate,
                location=c4_loc,
                display=str(rec.get("Invoice #", "")),
                tooltip=f"Jump to {edf_sheet_name} row {edf_target_row}",
            )
            c4.font = Font(name="Calibri", size=10, color="0563C1", underline="single")
        r += 1

    # Column widths
    widths = [18, 14, 18, 20, 14, 30, 18, 14, 12, 16, 16, 50]
    for i, w in enumerate(widths, start=1):
        col_letter = openpyxl.utils.get_column_letter(i)
        ws.column_dimensions[col_letter].width = w
    ws.freeze_panes = "A6"


# ---------------------------------------------------------------------------
# Cross-source reconciliation sheet writer
# ---------------------------------------------------------------------------
# Compares rows from the three SAP dump writers against the inferred analyser
# tables (Contract History / Meter Readings) and the EDF Evidence Report, line
# by line, with one Matched/Discrepancy/Missing row per comparison. Each
# matched row carries an openpyxl Hyperlink whose ``location`` points at the
# row on the source sheet that owns the matched side, so a reviewer can jump
# straight from a Discrepancy on the Reconciliation tab to the underlying row.


def _recon_parse_iso_date(s: str) -> pd.Timestamp | pd._libs.tslibs.nattype.NaTType:
    if not s:
        return pd.NaT
    s = str(s).strip()
    if not s:
        return pd.NaT
    # ISO first (YYYY-MM-DD), then day-first for DD/MM/YYYY.
    if re.match(r"^\d{4}-\d{2}-\d{2}$", s):
        return pd.to_datetime(s, errors="coerce")
    return pd.to_datetime(s, dayfirst=True, errors="coerce")


def _recon_amount_to_float(v: object) -> float:
    if v is None:
        return 0.0
    if isinstance(v, int | float):
        return float(v)
    try:
        return float(str(v).replace(",", "").strip().lstrip("£"))
    except ValueError:
        return 0.0


def write_reconciliation_sheet(
    ws_summary: Worksheet,
    ws_detail: Worksheet,
    sap_contract: list[dict],
    inferred_contract: pd.DataFrame,
    sap_meter: list[dict],
    inferred_meter: pd.DataFrame,
    sap_financial: list[dict],
    evidence_df: pd.DataFrame,
    account: str = "",
) -> None:
    """Render the two-sheet cross-source Reconciliation pair.

    Spec §3.2.  Sheet 1 is a compact 3-entity summary with verdict
    and drill-down hyperlink; Sheet 2 carries unmatched-only rows
    across three sections with AutoFilter.
    """
    ORANGE = "FE5716"
    NAVY = "10367A"

    def _banner(ws: Worksheet, row: int, text: str) -> None:
        ws.cell(row=row, column=1, value=text)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        for c in range(1, 9):
            cell = ws.cell(row=row, column=c)
            cell.fill = PatternFill("solid", start_color=NAVY)
            cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="left", vertical="center")

    def _section_banner(ws: Worksheet, row: int, text: str) -> None:
        ws.cell(row=row, column=1, value=f"\u25a0 {text} \u25a0")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        for c in range(1, 9):
            cell = ws.cell(row=row, column=c)
            cell.fill = PatternFill("solid", start_color=ORANGE)
            cell.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="left", vertical="center")

    def _header(ws: Worksheet, row: int, cols: list[str]) -> None:
        for i, name in enumerate(cols, start=1):
            cell = ws.cell(row=row, column=i, value=name)
            cell.fill = PatternFill("solid", start_color="DDDDDD")
            cell.font = Font(name="Calibri", size=10, bold=True, color="000000")

    # ---- Sheet 1 — summary ----
    ws_summary.title = "Reconciliation"
    _banner(ws_summary, 1, "EDF CROSS-SOURCE RECONCILIATION")
    second_line = "SAP dumps vs inferred evidence + reconciled evidence_df rows"
    if account:
        second_line += f"  •  Account: {account}"
    ws_summary.cell(row=2, column=1, value=second_line)
    ws_summary.merge_cells(start_row=2, start_column=1, end_row=2, end_column=8)

    sum_headers = [
        "Entity",
        "SAP-side count",
        "EDF-inferred count",
        "Matched",
        "Unmatched (SAP)",
        "Unmatched (EDF)",
        "Verdict",
        "Drill down",
    ]
    _header(ws_summary, 3, sum_headers)

    # ---- Sheet 2 — unmatched detail ----
    ws_detail.title = "Reconciliation Drill-down"
    _banner(ws_detail, 1, "EDF CROSS-SOURCE RECONCILIATION")
    ws_detail.cell(row=2, column=1, value=second_line)
    ws_detail.merge_cells(start_row=2, start_column=1, end_row=2, end_column=8)

    # Pre-compute detail row counters; each closure below mutates
    # these so the summary can read the tallies after all three
    # sections are built.
    detail_row = 3  # header row will be written here first
    contract_counts: dict[str, int] = {}
    meter_counts: dict[str, int] = {}
    financial_counts: dict[str, int] = {}

    def _build_section(
        ws: Worksheet,
        title: str,
        sap_items: list[dict],
        inferred_df: pd.DataFrame,
        inferred_cols: list[str],
        sap_date_key: str,
        sap_ref_key: str,
        sap_read_key: str | None,
        inferred_date_key: str,
        inferred_read_key: str,
        source_label: str,
        inferred_label: str,
        date_tolerance_days: int = 7,
        amount_tolerance: float | None = None,
    ) -> dict[str, int]:
        nonlocal detail_row
        counts = {"matched": 0, "unmatched_sap": 0, "unmatched_edf": 0}

        _section_banner(ws, detail_row, title)
        detail_row += 1
        header_cols = [
            "Status",
            f"SAP {sap_date_key}",
            f"SAP {sap_read_key or sap_ref_key}",
            f"Inferred {inferred_date_key}",
            f"Inferred {inferred_read_key}",
            "Notes",
            "Source",
            "Hyperlink",
        ]
        _header(ws, detail_row, header_cols)
        detail_row += 1
        section_start = detail_row

        inferred_rows: list[dict] = []
        if inferred_df is not None and not inferred_df.empty:
            inferred_rows = inferred_df.to_dict(orient="records")

        # Build matched stacks.
        unmatched_inferred = list(range(len(inferred_rows)))
        for sap in sap_items:
            sap_date = _recon_parse_iso_date(sap.get(sap_date_key, ""))
            sap_read = _recon_amount_to_float(sap.get(sap_read_key, "")) if sap_read_key else None
            matched = False
            for i in unmatched_inferred[:]:
                inf = inferred_rows[i]
                inf_date = _recon_parse_iso_date(inf.get(inferred_date_key, ""))
                date_close = (
                    abs((sap_date - inf_date).days) <= date_tolerance_days
                    if not pd.isna(sap_date) and not pd.isna(inf_date)
                    else False
                )
                if not date_close:
                    continue
                # Amount-close check (if applicable).
                if amount_tolerance is not None and sap_read is not None:
                    inf_read = _recon_amount_to_float(inf.get(inferred_read_key, ""))
                    if abs(sap_read - inf_read) > amount_tolerance:
                        continue
                # Successful match — suppress on detail sheet per spec 3.2;
                # only the summary counts carry this information.
                unmatched_inferred.remove(i)
                counts["matched"] += 1
                break
            if not matched:
                sap_idx = sap_items.index(sap) + 4
                ws.cell(row=detail_row, column=1, value="Missing in Inferred")
                ws.cell(row=detail_row, column=2, value=sap.get(sap_date_key, ""))
                ws.cell(
                    row=detail_row,
                    column=3,
                    value=sap.get(sap_read_key or sap_ref_key, ""),
                )
                for col in (4, 5):
                    ws.cell(row=detail_row, column=col, value="—")
                ws.cell(
                    row=detail_row,
                    column=6,
                    value=f"SAP {title.lower()} row not present in {inferred_label}",
                )
                ws.cell(row=detail_row, column=7, value=source_label)
                _recon_hyperlink(ws, detail_row, 8, source_label, sap_idx)
                detail_row += 1
                counts["unmatched_sap"] += 1
        for i in unmatched_inferred:
            inf = inferred_rows[i]
            inf_target = i + 4
            ws.cell(row=detail_row, column=1, value="Missing in SAP")
            for col in (2, 3):
                ws.cell(row=detail_row, column=col, value="—")
            ws.cell(
                row=detail_row,
                column=4,
                value=inf.get(inferred_date_key, ""),
            )
            ws.cell(
                row=detail_row,
                column=5,
                value=inf.get(inferred_read_key, ""),
            )
            ws.cell(
                row=detail_row,
                column=6,
                value=f"{inferred_label} row not present in SAP dump",
            )
            ws.cell(row=detail_row, column=7, value=inferred_label)
            _recon_hyperlink(ws, detail_row, 8, inferred_label, inf_target)
            detail_row += 1
            counts["unmatched_edf"] += 1

        # Compact spacer after each section.
        detail_row += 1
        return {**counts, "section_start": section_start}

    # Contract section
    contract_section = _build_section(
        ws_detail,
        "Contract Reconciliation",
        sap_contract,
        inferred_contract,
        ["Contract From", "Contract To", "Product Code"],
        "Contract From",
        "Product Code",
        None,
        "Contract From",
        "Product Code",
        "SAP Contract History",
        "Contract History",
    )
    contract_counts = contract_section

    # Meter Read section
    meter_section = _build_section(
        ws_detail,
        "Meter Read Reconciliation",
        sap_meter,
        inferred_meter,
        ["Meter Read Date", "Reading (kWh)"],
        "Meter Read Date",
        "Reading (kWh)",
        "Reading (kWh)",
        "Meter Read Date",
        "Reading (kWh)",
        "SAP Meter Readings",
        "Meter Readings",
    )
    meter_counts = meter_section

    # Financial reconciliation section
    _section_banner(ws_detail, detail_row, "Financial Reconciliation")
    detail_row += 1
    _header(
        ws_detail,
        detail_row,
        [
            "Status",
            "SAP Document No.",
            "SAP Posting Date",
            "SAP Amount",
            "Evidence Date",
            "Evidence Amount",
            "Notes",
            "Hyperlink",
        ],
    )
    detail_row += 1
    financial_section_start = detail_row

    evidence_rows_list: list[dict] = []
    if evidence_df is not None and not evidence_df.empty:
        evidence_rows_list = evidence_df.to_dict(orient="records")

    unmatched_ev = list(range(len(evidence_rows_list)))
    financial_counts = {"matched": 0, "unmatched_sap": 0, "unmatched_edf": 0}
    for sap in sap_financial:
        sap_date = _recon_parse_iso_date(sap.get("Posting Date", ""))
        sap_amt = _recon_amount_to_float(sap.get("Amount", ""))
        matched = False
        for i in unmatched_ev[:]:
            ev = evidence_rows_list[i]
            ev_date = _recon_parse_iso_date(ev.get("Date", ""))
            ev_amt = _recon_amount_to_float(ev.get("Amount (£)", 0))
            date_close = (
                abs((sap_date - ev_date).days) <= 7
                if not pd.isna(sap_date) and not pd.isna(ev_date)
                else False
            )
            amount_close = abs(sap_amt - ev_amt) <= 0.50
            if date_close and amount_close:
                matched = True
                abs_amt_diff = abs(sap_amt - ev_amt)
                if abs_amt_diff >= 0.01:
                    # Discrepancy rows stay on the detail sheet.
                    status = "Discrepancy"
                    note = f"Amount Δ = £{abs_amt_diff:,.2f}"
                    ws_detail.cell(row=detail_row, column=1, value=status)
                    ws_detail.cell(row=detail_row, column=2, value=sap.get("Document No.", ""))
                    ws_detail.cell(row=detail_row, column=3, value=sap.get("Posting Date", ""))
                    ws_detail.cell(row=detail_row, column=4, value=sap_amt)
                    ws_detail.cell(row=detail_row, column=5, value=ev.get("Date", ""))
                    ws_detail.cell(row=detail_row, column=6, value=ev_amt)
                    ws_detail.cell(row=detail_row, column=7, value=note)
                    _recon_hyperlink(
                        ws_detail, detail_row, 8, "EDF Evidence Report", financial_section_start + i
                    )
                    detail_row += 1
                unmatched_ev.remove(i)
                financial_counts["matched"] += 1
                break
        if not matched:
            sap_idx = sap_financial.index(sap) + 4
            ws_detail.cell(row=detail_row, column=1, value="Missing in Evidence")
            ws_detail.cell(row=detail_row, column=2, value=sap.get("Document No.", ""))
            ws_detail.cell(row=detail_row, column=3, value=sap.get("Posting Date", ""))
            ws_detail.cell(row=detail_row, column=4, value=sap_amt)
            for col in (5, 6):
                ws_detail.cell(row=detail_row, column=col, value="—")
            ws_detail.cell(
                row=detail_row,
                column=7,
                value="SAP financial row not on Evidence Report",
            )
            _recon_hyperlink(ws_detail, detail_row, 8, "SAP Financial Transactions", sap_idx)
            detail_row += 1
            financial_counts["unmatched_sap"] += 1
    for i in unmatched_ev:
        ev = evidence_rows_list[i]
        ev_target = financial_section_start + i
        ws_detail.cell(row=detail_row, column=1, value="Missing in SAP")
        for col in (2, 3, 4):
            ws_detail.cell(row=detail_row, column=col, value="—")
        ws_detail.cell(row=detail_row, column=5, value=ev.get("Date", ""))
        ws_detail.cell(row=detail_row, column=6, value=ev.get("Amount (£)", ""))
        ws_detail.cell(
            row=detail_row,
            column=7,
            value="Evidence row not present in SAP Financial dump",
        )
        _recon_hyperlink(ws_detail, detail_row, 8, "EDF Evidence Report", ev_target)
        detail_row += 1
        financial_counts["unmatched_edf"] += 1

    # ---- AutoFilter + freeze on detail sheet ----
    if detail_row > 3:
        ws_detail.auto_filter.ref = f"A3:H{detail_row}"
    ws_detail.freeze_panes = "A4"

    # ---- Build summary rows (3 entities) ----
    section_starts = {
        "Contract": contract_section["section_start"],
        "Meter Read": meter_section["section_start"],
        "Financial": financial_section_start,
    }

    def _verdict(entity: str, counts: dict[str, int]) -> str:
        if counts["matched"] == 0 and counts["unmatched_sap"] == 0:
            return f"No SAP-side {entity.lower()} rows to reconcile"
        if counts["unmatched_edf"] == 0 and counts["unmatched_sap"] == 0:
            return f"All {counts['matched']} SAP {entity.lower()} rows matched"
        parts = []
        if counts["unmatched_sap"]:
            parts.append(f"{counts['unmatched_sap']} SAP {entity.lower()} row(s) missing in EDF")
        if counts["unmatched_edf"]:
            parts.append(f"{counts['unmatched_edf']} EDF {entity.lower()} row(s) missing in SAP")
        return "; ".join(parts)

    summary_rows = [
        ("Contract", len(sap_contract), len(inferred_contract), contract_counts),
        ("Meter Read", len(sap_meter), len(inferred_meter), meter_counts),
        ("Financial", len(sap_financial), len(evidence_rows_list), financial_counts),
    ]
    for i, (entity, sap_cnt, edf_cnt, counts) in enumerate(summary_rows, start=4):
        ws_summary.cell(row=i, column=1, value=entity)
        ws_summary.cell(row=i, column=2, value=sap_cnt)
        ws_summary.cell(row=i, column=3, value=edf_cnt)
        ws_summary.cell(row=i, column=4, value=counts["matched"])
        ws_summary.cell(row=i, column=5, value=counts["unmatched_sap"])
        ws_summary.cell(row=i, column=6, value=counts["unmatched_edf"])
        ws_summary.cell(row=i, column=7, value=_verdict(entity, counts))
        target_row = section_starts[entity]
        cell = ws_summary.cell(row=i, column=8)
        cell.value = "→ Drill down"
        cell.hyperlink = openpyxl.worksheet.hyperlink.Hyperlink(
            ref=cell.coordinate,
            location=f"'Reconciliation Drill-down'!A{target_row}",
            display="→ Drill down",
        )
        cell.font = Font(color="0563C1", underline="single")

    ws_summary.freeze_panes = "A4"


# ---------------------------------------------------------------------------
# GUI
# ---------------------------------------------------------------------------


class ReportOptionsDialog:
    """Modern report options dialog with format selection and section checkboxes."""

    SECTIONS = [
        ("cover", "Cover Page", True),
        ("toc", "Table of Contents", True),
        ("exec_summary", "Executive Summary", True),
        ("key_findings", "Key Findings", True),
        ("evidence_index", "Evidence Index", True),
        ("detailed_findings", "Detailed Findings", True),
        ("timeline", "Timeline", True),
        ("ofgem", "OFGEM Price Cap Comparison", True),
        ("statistical", "Statistical Analysis", True),
        ("payment", "Payment Analysis", True),
        ("forecast", "Forecast", True),
        ("data_quality", "Data Quality", True),
        ("tariff", "Tariff Impact Analysis", True),
        ("appendix_methodology", "Appendix: Methodology", True),
        ("appendix_glossary", "Appendix: Glossary", True),
        ("appendix_full_evidence", "Appendix: Full Evidence Table", True),
    ]

    def __init__(self, parent):
        self.parent = parent
        self.result = None
        self.dialog = None

    def show(self):
        """Show the dialog and return the selected options."""
        self.dialog = tk.Toplevel(self.parent)
        self.dialog.title("Report Options")
        # Default size for 1080p: visible buttons without scrolling
        self.dialog.geometry("600x900")
        self.dialog.minsize(500, 500)
        self.dialog.resizable(True, True)
        self.dialog.transient(self.parent)
        self.dialog.grab_set()

        # Center on parent
        self.dialog.update_idletasks()
        x = self.parent.winfo_rootx() + (self.parent.winfo_width() // 2) - 300
        y = self.parent.winfo_rooty() + (self.parent.winfo_height() // 2) - 450
        self.dialog.geometry(f"+{x}+{y}")

        self._build_ui()
        self.dialog.wait_window()
        return self.result

    def _build_ui(self):
        """Build the dialog UI."""
        # Create scrollable main area
        canvas = tk.Canvas(self.dialog, highlightthickness=0)
        scrollbar = ttk.Scrollbar(self.dialog, orient="vertical", command=canvas.yview)
        main = ttk.Frame(canvas, padding=20)

        main.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=main, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        # Bind mousewheel
        def _on_mousewheel(event):
            canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

        canvas.bind_all("<MouseWheel>", _on_mousewheel)
        if self.dialog is not None:
            self.dialog.bind("<Destroy>", lambda e: canvas.unbind_all("<MouseWheel>"))

        # Also allow resizing canvas window width
        def _on_canvas_configure(event):
            canvas.itemconfig(canvas.find_all()[0], width=event.width)

        canvas.bind("<Configure>", _on_canvas_configure)

        # Header
        hdr = ttk.Frame(main)
        hdr.pack(fill=tk.X, pady=(0, 20))

        title_lbl = ttk.Label(
            hdr,
            text="Generate Ombudsman Report",
            font=("Calibri", 18, "bold"),
            foreground=EDF_NAVY,
        )
        title_lbl.pack(anchor=tk.W)

        subtitle = ttk.Label(
            hdr,
            text="Choose format and select sections to include",
            font=("Calibri", 10),
            foreground=MEDIUM_GREY,
        )
        subtitle.pack(anchor=tk.W, pady=(4, 0))

        ttk.Separator(main, orient="horizontal").pack(fill=tk.X, pady=(0, 16))

        # Format selection
        fmt_frame = ttk.LabelFrame(main, text=" Output Format ", padding=12)
        fmt_frame.pack(fill=tk.X, pady=(0, 16))

        self.format_var = tk.StringVar(value="both")
        formats = [
            ("both", "Both (PDF + Word)", "Generate both PDF and DOCX reports"),
            ("pdf", "PDF Only", "Professional PDF report (reportlab)"),
            ("docx", "Word Document Only", "Editable Word document (python-docx)"),
        ]

        for val, label, desc in formats:
            r = ttk.Frame(fmt_frame)
            r.pack(fill=tk.X, pady=3)
            rb = ttk.Radiobutton(r, variable=self.format_var, value=val)
            rb.pack(side=tk.LEFT)
            lbl_frame = ttk.Frame(r)
            lbl_frame.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=8)
            ttk.Label(lbl_frame, text=label, font=("Calibri", 10, "bold")).pack(anchor=tk.W)
            ttk.Label(lbl_frame, text=desc, font=("Calibri", 8), foreground=MEDIUM_GREY).pack(
                anchor=tk.W
            )

        ttk.Separator(main, orient="horizontal").pack(fill=tk.X, pady=(8, 16))

        # Section checkboxes
        sec_frame = ttk.LabelFrame(main, text=" Report Sections ", padding=12)
        sec_frame.pack(fill=tk.X, pady=(0, 16))

        # Select All / None buttons
        btn_frame = ttk.Frame(sec_frame)
        btn_frame.pack(fill=tk.X, pady=(0, 8))

        ttk.Button(btn_frame, text="Select All", command=self._select_all, width=12).pack(
            side=tk.LEFT
        )
        ttk.Button(btn_frame, text="Select None", command=self._select_none, width=12).pack(
            side=tk.LEFT, padx=(8, 0)
        )
        ttk.Button(btn_frame, text="Defaults", command=self._select_defaults, width=12).pack(
            side=tk.LEFT, padx=(8, 0)
        )

        # Checkboxes (main dialog is now scrollable, so no nested scrollbar needed)
        self.section_vars = {}
        for key, label, default in self.SECTIONS:
            var = tk.BooleanVar(value=default)
            self.section_vars[key] = var
            cb = ttk.Checkbutton(sec_frame, text=label, variable=var)
            cb.pack(anchor=tk.W, pady=1)

        ttk.Separator(main, orient="horizontal").pack(fill=tk.X, pady=(8, 16))

        # Action buttons
        action_frame = ttk.Frame(main)
        action_frame.pack(fill=tk.X)

        cancel_btn = ttk.Button(action_frame, text="Cancel", command=self._cancel, width=14)
        cancel_btn.pack(side=tk.RIGHT)

        ok_btn = tk.Button(
            action_frame,
            text="OK — Generate Report",
            bg=EDF_ORANGE,
            fg="white",
            font=("Calibri", 11, "bold"),
            command=self._generate,
            relief="flat",
            width=22,
        )
        ok_btn.pack(side=tk.RIGHT, padx=(0, 12))

        # Bind Enter key to OK, Escape to Cancel
        if self.dialog:
            self.dialog.bind("<Return>", lambda e: self._generate())
            self.dialog.bind("<Escape>", lambda e: self._cancel())

    def _select_all(self):
        for var in self.section_vars.values():
            var.set(True)

    def _select_none(self):
        for var in self.section_vars.values():
            var.set(False)

    def _select_defaults(self):
        for key, var in self.section_vars.items():
            # Find default from SECTIONS
            for k, _, default in self.SECTIONS:
                if k == key:
                    var.set(default)
                    break

    def _generate(self):
        """Collect results and close dialog."""
        selected_sections = [key for key, var in self.section_vars.items() if var.get()]
        if not selected_sections:
            messagebox.showwarning("No Sections", "Please select at least one report section.")
            return

        self.result = {
            "format": self.format_var.get(),
            "sections": selected_sections,
        }
        if self.dialog is not None:
            self.dialog.destroy()

    def _cancel(self):
        self.result = None
        if self.dialog is not None:
            self.dialog.destroy()


class App:
    def __init__(self, root):
        self.root = root
        self.root.title("EDF Master Evidence Collector")
        self.root.geometry("780x860")
        self.root.configure(bg=EDF_OFFWHITE)

        self.pst_path = tk.StringVar()
        self.pdf_dir = tk.StringVar()
        self.htm_path = tk.StringVar()
        self.acc_num = tk.StringVar(value="")
        self.status = tk.StringVar(value="Ready.")
        self.progress_v = tk.DoubleVar(value=0)

        self.use_anchors = tk.BooleanVar(value=True)
        self.use_large = tk.BooleanVar(value=True)
        self.use_reading_class = tk.BooleanVar(value=True)
        self.use_pdf_fields = tk.BooleanVar(value=True)
        self.use_acc_filt = tk.BooleanVar(value=False)
        self.filter_below = tk.BooleanVar(value=True)
        self.save_filtered = tk.BooleanVar(value=True)
        self.use_dedup = tk.BooleanVar(value=True)
        self.save_dups = tk.BooleanVar(value=True)
        self.use_domain_filter = tk.BooleanVar(value=True)
        self.domain_filter = tk.StringVar(value="edfenergy.com")
        self.min_amount = tk.DoubleVar(value=500.0)
        self.analysis_min = tk.DoubleVar(value=500.0)
        self.output_name = tk.StringVar(value="EDF_Dispute_Evidence.xlsx")
        self.report_account_ref = tk.StringVar(value="")

        # New vars for UI refresh (see spec 2026-07-10-ui-refresh-design.md)
        self.output_folder = tk.StringVar(value="")
        self.amalgamate_duplicates = tk.BooleanVar(value=False)
        self.auto_generate_report = tk.BooleanVar(value=False)
        # Stream P5: save evidence files referenced by the workbook into a
        # flat ``output/evidence_files/`` folder and a themed DOCX index.
        # Defaults to True so the bundle is produced alongside the workbook
        # by default; reviewer can uncheck if they only want the XLSX.
        self.save_evidence_files_var = tk.BooleanVar(value=True)
        # Stream P1/P2 GUI toggles. SAP CSV-in-PDF data dumps render
        # their own dedicated sheets when "scan_sap_dumps" is set; the
        # cross-source Reconciliation sheet is independently controllable
        # via "generate_reconciliation_sheet" so a reviewer can keep the
        # SAP data without the cross-sheet matching view if desired.
        # Both default to True so the new sheets appear in the standard
        # extraction output; toggle off if the reviewer doesn't want
        # the legacy SAP dump analysis at all (e.g. on a clean run with
        # only invoice PDFs).
        self.scan_sap_dumps_var = tk.BooleanVar(value=True)
        self.generate_reconciliation_sheet_var = tk.BooleanVar(value=True)
        self._report_options: dict = {}
        self._CONFIG_PATH = os.path.expanduser("~/.edf_collector/config.json")

        # Load persisted config (may override the var defaults above)
        self._load_config()

        self.cancel_event = threading.Event()
        self.build_ui()

    # -- Config persistence --

    def _load_config(self):
        """Read config file and mutate tk vars via .set().

        Silently falls back to hardcoded defaults when the file is
        missing, unreadable, or malformed.
        """
        try:
            with open(self._CONFIG_PATH) as f:
                data = json.load(f)
        except (FileNotFoundError, json.JSONDecodeError, OSError):
            return

        gui = data.get("gui_state", {})
        _bool_keys: dict[str, tk.Variable] = {
            "use_anchors": self.use_anchors,
            "use_large": self.use_large,
            "use_reading_class": self.use_reading_class,
            "use_pdf_fields": self.use_pdf_fields,
            "use_acc_filt": self.use_acc_filt,
            "filter_below": self.filter_below,
            "save_filtered": self.save_filtered,
            "use_dedup": self.use_dedup,
            "save_dups": self.save_dups,
            "amalgamate_duplicates": self.amalgamate_duplicates,
            "use_domain_filter": self.use_domain_filter,
            "auto_generate_report": self.auto_generate_report,
            "save_evidence_files": self.save_evidence_files_var,
            "scan_sap_dumps": self.scan_sap_dumps_var,
            "generate_reconciliation_sheet": self.generate_reconciliation_sheet_var,
        }
        for key, var in _bool_keys.items():
            if key in gui:
                var.set(bool(gui[key]))

        _str_keys: dict[str, tk.Variable] = {
            "acc_num": self.acc_num,
            "domain_filter": self.domain_filter,
            "output_name": self.output_name,
            "report_account_ref": self.report_account_ref,
            "output_folder": self.output_folder,
        }
        for key, var in _str_keys.items():
            if key in gui:
                var.set(str(gui[key]))

        _float_keys: dict[str, tk.Variable] = {
            "min_amount": self.min_amount,
            "analysis_min": self.analysis_min,
        }
        for key, var in _float_keys.items():
            if key in gui:
                try:
                    var.set(float(gui[key]))
                except (ValueError, TypeError):
                    pass

        ro = data.get("report_options", {})
        if ro:
            self._report_options = ro

    def _save_config(self):
        """Persist GUI state + report options to config file atomically.

        Write to <path>.tmp, fsync, os.replace.  Permissions 0o600.
        """
        config_dir = os.path.dirname(self._CONFIG_PATH)
        os.makedirs(config_dir, exist_ok=True)

        gui = {
            "use_anchors": self.use_anchors.get(),
            "use_large": self.use_large.get(),
            "use_reading_class": self.use_reading_class.get(),
            "use_pdf_fields": self.use_pdf_fields.get(),
            "use_acc_filt": self.use_acc_filt.get(),
            "acc_num": self.acc_num.get(),
            "min_amount": self.min_amount.get(),
            "analysis_min": self.analysis_min.get(),
            "filter_below": self.filter_below.get(),
            "save_filtered": self.save_filtered.get(),
            "use_dedup": self.use_dedup.get(),
            "save_dups": self.save_dups.get(),
            "amalgamate_duplicates": self.amalgamate_duplicates.get(),
            "use_domain_filter": self.use_domain_filter.get(),
            "domain_filter": self.domain_filter.get(),
            "output_name": self.output_name.get(),
            "report_account_ref": self.report_account_ref.get(),
            "auto_generate_report": self.auto_generate_report.get(),
            "output_folder": self.output_folder.get(),
            "save_evidence_files": self.save_evidence_files_var.get(),
            "scan_sap_dumps": self.scan_sap_dumps_var.get(),
            "generate_reconciliation_sheet": self.generate_reconciliation_sheet_var.get(),
        }

        payload = {
            "output_folder": self.output_folder.get(),
            "report_options": getattr(self, "_report_options", {}),
            "gui_state": gui,
        }

        tmp_path = self._CONFIG_PATH + ".tmp"
        with open(tmp_path, "w") as f:
            json.dump(payload, f, indent=2)
            f.flush()
            os.fsync(f.fileno())
        os.chmod(tmp_path, 0o600)
        os.replace(tmp_path, self._CONFIG_PATH)

    def _resolve_output_path(
        self,
        stem: str,
        ext: str,
        batch_n: int | None = None,
        is_report: bool = False,
    ) -> str:
        """Build a sequential non-overwriting output path.

        Naming: {folder}/{stem}_{date}_{N}[{_Report}].{ext}
        If batch_n is passed, use it (shared counter for a batch).
        If None, scan folder for max existing N and use N+1.
        If output_folder is empty, falls back to os.getcwd().
        """
        folder = self.output_folder.get().strip() or os.getcwd()
        date_stamp = date.today().isoformat()
        suffix = "_Report" if is_report else ""

        if batch_n is not None:
            n = batch_n
        else:
            pattern = os.path.join(folder, f"{stem}_{date_stamp}_*{suffix}.{ext}")
            existing = glob.glob(pattern)
            max_n = 0
            for f in existing:
                basename = os.path.basename(f)
                prefix = f"{stem}_{date_stamp}_"
                rest = basename[len(prefix) :]
                if suffix:
                    rest = rest[: rest.index(suffix)]
                rest = rest.rsplit(".", 1)[0]
                if rest.isdigit():
                    max_n = max(max_n, int(rest))
            n = max_n + 1

        filename = f"{stem}_{date_stamp}_{n}{suffix}.{ext}"
        return os.path.join(folder, filename)

    def build_ui(self):
        hdr = tk.Frame(self.root, bg=EDF_ORANGE, height=60)
        hdr.pack(fill=tk.X)
        tk.Label(
            hdr,
            text="EDF BILLING EVIDENCE COLLECTOR",
            bg=EDF_ORANGE,
            fg="white",
            font=("Calibri", 14, "bold"),
        ).pack(pady=15)

        container = ttk.Frame(self.root)
        container.pack(fill=tk.BOTH, expand=True)

        canvas = tk.Canvas(container, bg=EDF_OFFWHITE, highlightthickness=0)
        yscroll = ttk.Scrollbar(container, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=yscroll.set)
        yscroll.pack(side=tk.RIGHT, fill=tk.Y)
        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        main = ttk.Frame(canvas, padding=16)
        cw = canvas.create_window((0, 0), window=main, anchor="nw")

        def _reconfig(_e=None):
            canvas.configure(scrollregion=canvas.bbox("all"))
            canvas.itemconfig(cw, width=canvas.winfo_width())

        main.bind("<Configure>", _reconfig)
        canvas.bind("<Configure>", _reconfig)

        # --- Section 1: Source Data ---
        s1 = ttk.LabelFrame(main, text=" 1. Source Data ", padding=10)
        s1.pack(fill=tk.X, pady=5)

        def browse_row(parent, label, var, cmd):
            r = ttk.Frame(parent)
            r.pack(fill=tk.X, pady=2)
            ttk.Label(r, text=label, width=14).pack(side=tk.LEFT)
            ttk.Entry(r, textvariable=var).pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
            ttk.Button(r, text="Browse", command=cmd).pack(side=tk.LEFT)

        browse_row(s1, "PST/OST File:", self.pst_path, self._pick_pst)
        browse_row(s1, "PDF Folder:", self.pdf_dir, self._pick_pdf_dir)
        browse_row(
            s1,
            "HTM Export:",
            self.htm_path,
            lambda: self.htm_path.set(
                filedialog.askopenfilename(filetypes=[("HTM/HTML", "*.htm *.html")])
            ),
        )

        # Output Folder picker (new - spec Design Section 1)
        browse_row(s1, "Output Folder:", self.output_folder, self._pick_output_folder)

        # Output filename row relocated from Section 2 to Section 1
        r_out = ttk.Frame(s1)
        r_out.pack(fill=tk.X, pady=2)
        ttk.Label(r_out, text="Output filename:", width=14).pack(side=tk.LEFT)
        ttk.Entry(r_out, textvariable=self.output_name, width=30).pack(side=tk.LEFT, padx=5)

        # --- Section 2: Extraction options ---
        s2 = ttk.LabelFrame(main, text=" 2. Search & Filter Options ", padding=10)
        s2.pack(fill=tk.X, pady=5)
        for text, var in [
            ("Smart Context Search", self.use_anchors),
            ("Large Number Fallback", self.use_large),
            ("Classify Reading Type", self.use_reading_class),
            ("Deep PDF Mine (kWh, standing charge, invoice #)", self.use_pdf_fields),
        ]:
            tk.Checkbutton(s2, text=text, variable=var, bg=EDF_OFFWHITE).pack(anchor=tk.W)

        r3 = ttk.Frame(s2)
        r3.pack(fill=tk.X, pady=4)
        tk.Checkbutton(
            r3, text="Filter by Account #:", variable=self.use_acc_filt, bg=EDF_OFFWHITE
        ).pack(side=tk.LEFT)
        ttk.Entry(r3, textvariable=self.acc_num, width=16).pack(side=tk.LEFT, padx=5)

        r3d = ttk.Frame(s2)
        r3d.pack(fill=tk.X, pady=4)
        tk.Checkbutton(
            r3d,
            text="Filter PST emails by sender domain:",
            variable=self.use_domain_filter,
            bg=EDF_OFFWHITE,
        ).pack(side=tk.LEFT)
        ttk.Entry(r3d, textvariable=self.domain_filter, width=40).pack(side=tk.LEFT, padx=5)
        ttk.Label(r3d, text="(comma-separated domains/addresses)", font=("Calibri", 8)).pack(
            side=tk.LEFT
        )

        r4 = ttk.Frame(s2)
        r4.pack(fill=tk.X, pady=2)
        chk_filt = tk.Checkbutton(
            r4, text="Filter results below minimum £:", variable=self.filter_below, bg=EDF_OFFWHITE
        )
        chk_filt.pack(side=tk.LEFT)
        ttk.Entry(r4, textvariable=self.min_amount, width=8).pack(side=tk.LEFT, padx=5)

        r4c = ttk.Frame(s2)
        r4c.pack(fill=tk.X, pady=2)
        ttk.Label(r4c, text="Analysis threshold (£):", width=24).pack(side=tk.LEFT)
        ttk.Entry(r4c, textvariable=self.analysis_min, width=8).pack(side=tk.LEFT, padx=5)

        r4d = ttk.Frame(s2)
        r4d.pack(fill=tk.X, pady=2)
        ttk.Label(r4d, text="Report account reference:", width=24).pack(side=tk.LEFT)
        ttk.Entry(r4d, textvariable=self.report_account_ref, width=20).pack(side=tk.LEFT, padx=5)

        chk_sf = tk.Checkbutton(
            s2,
            text="Keep filtered-out records on side sheet (Filtered (Below Min))",
            variable=self.save_filtered,
            bg=EDF_OFFWHITE,
        )
        chk_sf.pack(anchor=tk.W, padx=20)

        def _update_sf_state() -> None:
            chk_sf.config(state="normal" if self.filter_below.get() else "disabled")

        chk_filt.config(command=_update_sf_state)
        _update_sf_state()

        # Auto-generate report after extraction (spec Design Section 2)
        tk.Checkbutton(
            s2,
            text="Auto-generate report after extraction",
            variable=self.auto_generate_report,
            bg=EDF_OFFWHITE,
            command=self._save_config,
        ).pack(anchor=tk.W)

        # Stream P5: save evidence files + themed DOCX bundle index alongside
        # the workbook (spec Design Section 2 + §7). Defaults True.
        tk.Checkbutton(
            s2,
            text="Save evidence files + bundle index (output/evidence_files + evidence_index.docx)",
            variable=self.save_evidence_files_var,
            bg=EDF_OFFWHITE,
            command=self._save_config,
        ).pack(anchor=tk.W)

        # Stream P1: detect + render the three SAP CSV-in-PDF data
        # dumps (Contract / Meter-Read / Financial-Transactions) on
        # their own dedicated sheets.
        tk.Checkbutton(
            s2,
            text="Scan SAP CSV-in-PDF data dumps (adds SAP Contract History / Meter Readings / Financial Transactions sheets)",
            variable=self.scan_sap_dumps_var,
            bg=EDF_OFFWHITE,
            command=self._save_config,
        ).pack(anchor=tk.W)

        # Stream P2: cross-source Reconciliation sheet (SAP rows vs
        # inferred analyser data). Independent of the SAP-scan
        # toggle so a reviewer can keep the SAP data without the
        # cross-source match view if they want only the raw SAP
        # signals.
        tk.Checkbutton(
            s2,
            text="Generate cross-source Reconciliation sheet (SAP vs inferred analyser rows)",
            variable=self.generate_reconciliation_sheet_var,
            bg=EDF_OFFWHITE,
            command=self._save_config,
        ).pack(anchor=tk.W)

        self.report_options_section2_btn = tk.Button(
            s2,
            text="Report Options...",
            bg=EDF_NAVY,
            fg="white",
            font=("Calibri", 10),
            command=self._open_report_options,
            relief="flat",
        )
        self.report_options_section2_btn.pack(anchor=tk.W, padx=20, pady=4)

        # --- Section 3: Deduplication (relabelled + amalgamate child) ---
        s3 = ttk.LabelFrame(main, text=" 3. Deduplication ", padding=10)
        s3.pack(fill=tk.X, pady=5)
        chk_dup = tk.Checkbutton(
            s3,
            text="Drop duplicates found across sources",
            variable=self.use_dedup,
            bg=EDF_OFFWHITE,
        )
        chk_dup.pack(anchor=tk.W)
        chk_sd = tk.Checkbutton(
            s3,
            text="Record dropped duplicates on side sheet (Duplicate Entries)",
            variable=self.save_dups,
            bg=EDF_OFFWHITE,
        )
        chk_sd.pack(anchor=tk.W, padx=20)
        chk_am = tk.Checkbutton(
            s3,
            text="Build hybrid row per duplicate cluster (merge columns from every sibling)",
            variable=self.amalgamate_duplicates,
            bg=EDF_OFFWHITE,
            command=self._save_config,
        )
        chk_am.pack(anchor=tk.W, padx=40)

        def _update_dedup_state() -> None:
            dedup_on = self.use_dedup.get()
            chk_sd.config(state="normal" if dedup_on else "disabled")
            chk_am.config(state="normal" if (dedup_on and self.save_dups.get()) else "disabled")

        def _update_amalgamate_state() -> None:
            chk_am.config(
                state="normal" if (self.use_dedup.get() and self.save_dups.get()) else "disabled"
            )

        chk_dup.config(command=_update_dedup_state)
        chk_sd.config(command=_update_amalgamate_state)
        _update_dedup_state()

        # --- Progress ---
        self.pb = ttk.Progressbar(main, mode="determinate", maximum=100, variable=self.progress_v)
        self.pb.pack(fill=tk.X, pady=10)
        ttk.Label(
            main, textvariable=self.status, foreground=EDF_NAVY, font=("Calibri", 11, "bold")
        ).pack()

        btns = ttk.Frame(main)
        btns.pack(fill=tk.X, pady=8)
        self.run_btn = tk.Button(
            btns,
            text="EXTRACT TO EXCEL",
            bg=EDF_ORANGE,
            fg="white",
            font=("Calibri", 12, "bold"),
            command=self.start_thread,
            relief="flat",
        )
        self.run_btn.pack(side=tk.LEFT, fill=tk.X, expand=True, ipady=8)

        self.report_options_btn = tk.Button(
            btns,
            text="Report Options",
            bg=EDF_NAVY,
            fg="white",
            font=("Calibri", 12, "bold"),
            command=self._open_report_options,
            relief="flat",
            state="normal" if (HAS_PDF_REPORT or HAS_DOCX_REPORT) else "disabled",
        )
        self.report_options_btn.pack(side=tk.LEFT, fill=tk.X, expand=True, ipady=8, padx=(8, 0))

        # Load Spreadsheet & Generate Report button
        self.load_report_btn = tk.Button(
            btns,
            text="LOAD & REPORT",
            bg=EDF_ORANGE,
            fg="white",
            font=("Calibri", 12, "bold"),
            command=self.load_spreadsheet_and_report,
            relief="flat",
        )
        self.load_report_btn.pack(side=tk.LEFT, fill=tk.X, expand=True, ipady=8, padx=(8, 0))

    # -- Helpers --

    def _pick_pst(self):
        p = filedialog.askopenfilename(filetypes=[("Mail Stores", "*.pst *.ost")])
        if p:
            self.pst_path.set(p)

    def _pick_pdf_dir(self):
        p = filedialog.askdirectory()
        if p:
            self.pdf_dir.set(p)

    def _pick_output_folder(self):
        p = filedialog.askdirectory()
        if p:
            self.output_folder.set(p)
            self._save_config()

    def _open_report_options(self):
        """Open ReportOptionsDialog and persist selection on OK."""
        dialog = ReportOptionsDialog(self.root)
        options = dialog.show()
        if options:
            self._report_options = options
            self._save_config()

    def set_status(self, text):
        def _apply():
            self.status.set(text)
            self.root.update_idletasks()

        if threading.current_thread() is threading.main_thread():
            _apply()
        else:
            self.root.after(0, _apply)

    def set_progress(self, current, total, text=None):
        pct = max(0, min(100, (current / total) * 100)) if total else 0

        def _apply():
            self.progress_v.set(pct)
            if text:
                self.status.set(text)

        if threading.current_thread() is threading.main_thread():
            _apply()
        else:
            self.root.after(0, _apply)

    def _show(self, level, title, text):
        def _s():
            if level == "info":
                messagebox.showinfo(title, text)
            elif level == "warning":
                messagebox.showwarning(title, text)
            else:
                messagebox.showerror(title, text)

        if threading.current_thread() is threading.main_thread():
            _s()
        else:
            self.root.after(0, _s)

    def _finish(self):
        self._set_extract_idle()
        self.progress_v.set(0)
        self.set_status("Cancelled." if self.cancel_event.is_set() else "Ready.")
        gc.collect()

    def _set_extract_idle(self):
        """Flip run_btn to Idle: orange, EXTRACT TO EXCEL."""
        self.run_btn.config(
            text="EXTRACT TO EXCEL",
            bg=EDF_ORANGE,
            fg="white",
            command=self.start_thread,
            state="normal",
        )

    def _set_extract_running(self):
        """Flip run_btn to Running: navy, Cancel."""
        self.run_btn.config(
            text="Cancel",
            bg=EDF_NAVY,
            fg="white",
            command=self._cancel,
            state="normal",
        )

    def _set_extract_cancelling(self):
        """Flip run_btn to Cancelling: grey, Cancelling..."""
        self.run_btn.config(
            text="Cancelling...",
            bg=MEDIUM_GREY,
            fg="white",
            state="disabled",
        )

    def load_spreadsheet_and_report(self):
        """Load records from an existing spreadsheet and auto-generate reports.

        Assumes the spreadsheet has standard EDF Evidence Report format with
        an 'EDF Evidence Report' sheet.  Sequential-named reports written
        into output_folder (or the picked file's directory if unset).
        """
        if not HAS_PDF_REPORT and not HAS_DOCX_REPORT:
            self._show(
                "error",
                "Report Unavailable",
                "Report generation requires 'reportlab' (PDF) and/or 'python-docx' (Word).\n"
                "Install with: pip install reportlab python-docx",
            )
            return

        file_path = filedialog.askopenfilename(
            initialdir=self.output_folder.get() or os.getcwd(),
            filetypes=[("Excel Files", "*.xlsx"), ("All Files", "*.*")],
            title="Select EDF Evidence Report Spreadsheet",
        )
        if not file_path:
            return

        try:
            df = pd.read_excel(file_path, sheet_name="EDF Evidence Report")
            if df.empty:
                self._show(
                    "warning",
                    "No Data",
                    "The selected spreadsheet has no records in 'EDF Evidence Report' sheet.",
                )
                return

            records = df.to_dict("records")
            ro = getattr(self, "_report_options", {})
            fmt = ro.get("format", "both")
            sections = ro.get("sections", [s[0] for s in ReportOptionsDialog.SECTIONS])

            base_dir = self.output_folder.get().strip() or os.path.dirname(file_path)
            self.output_folder.set(base_dir)
            stem = os.path.basename(file_path).replace(".xlsx", "")

            output_paths: dict[str, str] = {}
            if fmt in ("pdf", "both") and HAS_PDF_REPORT:
                output_paths["pdf"] = self._resolve_output_path(stem, "pdf", is_report=True)
            if fmt in ("docx", "both") and HAS_DOCX_REPORT:
                output_paths["docx"] = self._resolve_output_path(
                    stem, "docx", batch_n=1, is_report=True
                )

            if not output_paths:
                self._show(
                    "warning",
                    "No Reports",
                    "No report paths resolved (check pdf/docx availability).",
                )
                return

            self.set_status("Generating report…")
            self.load_report_btn.config(state="disabled")

            config = {
                "min_amount": self.min_amount.get(),
                "analysis_min": self.analysis_min.get(),
                "acc_num": self.acc_num.get(),
                "report_account_ref": self.report_account_ref.get().strip(),
                "report_sections": sections,
            }

            from dataclasses import dataclass

            @dataclass
            class MockEngine:
                records: list
                filtered_records: list
                pdf_count: int
                email_count: int
                error_log: list

            engine = MockEngine(
                records=records, filtered_records=[], pdf_count=0, email_count=0, error_log=[]
            )

            def _generate():
                from edf_report import generate_pdf_from_gui
                from edf_report_docx import generate_docx_from_gui

                try:
                    msgs = []
                    if "pdf" in output_paths:
                        s, m = generate_pdf_from_gui(
                            records=records,
                            output_path=output_paths["pdf"],
                            config=config,
                            engine=engine,
                            filtered=[],
                        )
                        msgs.append(("PDF", s, m))
                    if "docx" in output_paths:
                        s, m = generate_docx_from_gui(
                            records=records,
                            output_path=output_paths["docx"],
                            config=config,
                            engine=engine,
                            filtered=[],
                        )
                        msgs.append(("DOCX", s, m))

                    combined = []
                    all_ok = True
                    for label, ok, m in msgs:
                        if ok:
                            combined.append(
                                f"✓ {label}: {m.split(chr(10))[-1] if m else 'Generated'}\n{output_paths[label.lower()]}"
                            )
                        else:
                            all_ok = False
                            self.root.after(
                                0,
                                lambda mn=m, lb=label: self._show(
                                    "error", f"{lb} Generation Failed", mn
                                ),
                            )
                    if all_ok and combined:
                        self.root.after(
                            0,
                            lambda c=combined: self._show(
                                "info", "Reports Generated", "\n\n".join(c)
                            ),
                        )
                except Exception as e:
                    self.root.after(
                        0,
                        lambda err=e: self._show("error", "Error", f"An error occurred:\n\n{err}"),
                    )
                finally:
                    self.root.after(
                        0,
                        lambda: (
                            self.load_report_btn.config(state="normal"),
                            self.set_status("Ready."),
                        ),
                    )

            threading.Thread(target=_generate, daemon=True).start()

        except Exception as e:
            self._show("error", "Load Error", f"Failed to load spreadsheet:\n\n{e}")

    def _cancel(self):
        self.cancel_event.set()
        self._set_extract_cancelling()
        self.set_status("Cancelling…")

    def start_thread(self):
        try:
            self.min_amount.get()
            self.analysis_min.get()
        except Exception:
            messagebox.showerror(
                "Error", "Minimum amount and analysis threshold must be valid numbers."
            )
            return

        has_sources = any(
            [
                self.pst_path.get().strip(),
                self.pdf_dir.get().strip(),
                self.htm_path.get().strip(),
            ]
        )
        if not has_sources:
            messagebox.showerror(
                "Error",
                "Please select at least one source:\nPST/OST file, PDF folder, or HTM export.",
            )
            return
        self.cancel_event.clear()
        self._set_extract_running()
        self.progress_v.set(0)
        threading.Thread(target=self._run, daemon=True).start()

    def _run(self):
        config = {
            "use_anchors": self.use_anchors.get(),
            "use_large": self.use_large.get(),
            "use_reading_classification": self.use_reading_class.get(),
            "use_pdf_fields": self.use_pdf_fields.get(),
            "use_acc_filter": self.use_acc_filt.get(),
            "acc_num": self.acc_num.get(),
            "min_amount": self.min_amount.get(),
            "analysis_min": self.analysis_min.get(),
            "report_account_ref": self.report_account_ref.get().strip(),
            "filter_below": self.filter_below.get(),
            "save_filtered": self.save_filtered.get(),
            "use_dedup": self.use_dedup.get(),
            "save_dups": self.save_dups.get(),
            "amalgamate_duplicates": self.amalgamate_duplicates.get(),
            "use_domain_filter": self.use_domain_filter.get(),
            "domain_filter": self.domain_filter.get().strip(),
            # Stream P1/P2 toggles -- threaded through to
            # export_to_excel which gates SAP sheet writes + the
            # Reconciliation sheet.
            "save_evidence_files": self.save_evidence_files_var.get(),
            "scan_sap_dumps": self.scan_sap_dumps_var.get(),
            "generate_reconciliation_sheet": self.generate_reconciliation_sheet_var.get(),
        }

        from edf_collector import EvidenceEngine  # noqa: F401,E402

        engine = EvidenceEngine(config, self.set_status, self.set_progress, self.cancel_event)
        self.engine = engine

        try:
            pst_path = self.pst_path.get().strip()
            if pst_path and os.path.exists(pst_path) and not self.cancel_event.is_set():
                if not HAS_PYPFF:
                    self._show("warning", "PST", "pypff not installed — PST/OST scanning skipped.")
                else:
                    self.set_status("Scanning PST/OST…")
                    try:
                        pff = pypff.file()
                    except AttributeError:
                        pff = getattr(pypff, "File", None)
                        if pff is None:
                            raise AttributeError(
                                "pypff module has no 'file' or 'File' attribute"
                            ) from None
                        pff = pff()
                    pff.open(os.path.abspath(pst_path))
                    try:
                        engine.crawl_pst(pff.get_root_folder())
                    finally:
                        pff.close()

            htm_path = self.htm_path.get().strip()
            if htm_path and os.path.exists(htm_path) and not self.cancel_event.is_set():
                self.set_status("Parsing HTM account history…")
                engine.process_htm_file(htm_path)

            pdf_path = self.pdf_dir.get().strip()
            if pdf_path and os.path.exists(pdf_path) and not self.cancel_event.is_set():
                engine.crawl_local_pdfs(pdf_path)

            if self.cancel_event.is_set():
                self._show("warning", "Cancelled", "Extraction cancelled.")
                return

            if engine.records:
                self.set_status("Writing Excel report…")
                # Fall back to source dir when output_folder unset
                if not self.output_folder.get().strip():
                    base_dir = (
                        os.path.dirname(pst_path)
                        if pst_path
                        else pdf_path
                        if pdf_path
                        else os.path.dirname(htm_path)
                        if htm_path
                        else os.getcwd()
                    )
                    self.output_folder.set(base_dir)
                stem = self.output_name.get().strip() or "EDF_Dispute_Evidence"
                if stem.lower().endswith(".xlsx"):
                    stem = stem[:-5]
                xlsx_path = self._resolve_output_path(stem, "xlsx")
                export_to_excel(
                    engine.records,
                    xlsx_path,
                    engine.error_log,
                    config,
                    filtered=engine.filtered_records,
                    sap_rows={
                        "contract": engine.sap_contract_rows,
                        "meter": engine.sap_meter_rows,
                        "financial": engine.sap_financial_rows,
                    },
                )
                self._save_config()
                summary = (
                    f"Extraction complete.\n\n"
                    f"  Emails matched: {engine.email_count}\n"
                    f"  PDFs processed: {engine.pdf_count}\n"
                    f"  Records found:  {len(engine.records)}\n"
                )
                if engine.error_log:
                    summary += f"\n  Parse errors: {len(engine.error_log)} (see Parse Errors tab)"
                summary += f"\n\nSaved to:\n{xlsx_path}"

                # Stream P5: save evidence files + themed DOCX bundle index
                # into a sibling ``evidence_files/`` folder when the toggle is
                # set on (default True).
                if self.save_evidence_files_var.get():
                    try:
                        import pandas as pd

                        out_dir = os.path.dirname(xlsx_path) or os.getcwd()
                        ev_dir = os.path.join(out_dir, "evidence_files")
                        dfc = pd.DataFrame(engine.records)
                        # Build the source-paths reverse-lookup from the
                        # crawl attribute the engine carries internally.
                        source_paths = getattr(engine, "source_paths", {}) or {}
                        saved = save_evidence_files(dfc, source_paths, ev_dir)
                        index_docx = os.path.join(out_dir, "evidence_index.docx")
                        build_bundle_index(
                            dfc, saved, index_docx, account=str(config.get("acc_num", ""))
                        )
                        summary += f"\n\nSaved {len(saved)} evidence files to:\n{ev_dir}"
                        summary += f"\nBundle index:\n{index_docx}"
                    except Exception as bundle_err:
                        # Don't lose the run if the bundle step blows up --
                        # log it loudly but still keep the XLSX.
                        self._show(
                            "warning",
                            "Bundle step failed",
                            (
                                f"Evidence file save failed:\n{bundle_err}"
                                f"\n\nThe XLSX workbook is still saved at:\n{xlsx_path}"
                            ),
                        )

                if self.auto_generate_report.get():
                    report_paths = self._run_auto_report(engine, stem, 1)
                    if report_paths:
                        summary += "\n\nReports:\n" + "\n".join(report_paths)

                self._show("info", "Success", summary)
            else:
                self._show(
                    "warning",
                    "No Data",
                    "No billing amounts found.\n\nTips:\n"
                    "• Uncheck the Account Filter\n"
                    "• Lower the minimum threshold\n"
                    "• Check your source files contain EDF billing data",
                )

        except Exception:
            self._show("error", "Error", f"An error occurred:\n\n{traceback.format_exc()}")
        finally:
            self.root.after(0, self._finish)

    def _run_auto_report(self, engine, stem, batch_n):
        """Run report generation for the auto-generate flow.

        Uses persisted _report_options; writes to output_folder;
        returns list of written paths.
        """
        from edf_report import generate_pdf_from_gui
        from edf_report_docx import generate_docx_from_gui

        ro = getattr(self, "_report_options", {})
        fmt = ro.get("format", "both")
        sections = ro.get("sections", [s[0] for s in ReportOptionsDialog.SECTIONS])

        config = {
            "min_amount": self.min_amount.get(),
            "analysis_min": self.analysis_min.get(),
            "acc_num": self.acc_num.get(),
            "report_account_ref": self.report_account_ref.get().strip(),
            "report_sections": sections,
        }

        written: list[str] = []
        if fmt in ("pdf", "both") and HAS_PDF_REPORT:
            pdf_path = self._resolve_output_path(stem, "pdf", batch_n=batch_n, is_report=True)
            success, _ = generate_pdf_from_gui(
                records=engine.records,
                output_path=pdf_path,
                config=config,
                engine=engine,
                filtered=engine.filtered_records,
            )
            if success:
                written.append(pdf_path)

        if fmt in ("docx", "both") and HAS_DOCX_REPORT:
            docx_path = self._resolve_output_path(stem, "docx", batch_n=batch_n, is_report=True)
            success, _ = generate_docx_from_gui(
                records=engine.records,
                output_path=docx_path,
                config=config,
                engine=engine,
                filtered=engine.filtered_records,
            )
            if success:
                written.append(docx_path)

        return written


# ---------------------------------------------------------------------------
# Safe pickle deserialiser — prevents arbitrary code execution when loading
# engine-data pickle files from disk.  Only standard built-in types and
# the project's own EvidenceEngine class are allowed through; anything
# else raises UnpicklingError.
# ---------------------------------------------------------------------------


class _RestrictedUnpickler(pickle.Unpickler):
    """Unpickler that only allows known-safe types.

    Permits: built-in scalars, dicts, lists, tuples, sets, frozensets,
    bytes/bytearray, and the project's own ``EvidenceEngine``.  Everything
    else triggers ``pickle.UnpicklingError`` so a crafted pickle can never
    import and call arbitrary code.
    """

    # Module→class whitelist.  Only classes listed here can be rebuilt.
    # A whitelist value of ``None`` (as opposed to the usual
    # ``set[str]`` of permitted class names) is interpreted as
    # "the entire module is trusted".  We only use this for
    # ``pyarrow.lib`` whose exposed pickle surface is purely
    # restoration-callable ``_something`` functions, never
    # ``os.system`` / ``subprocess.Popen``.  Every other
    # whitelist entry is an explicit set of class names.
    #
    # Note ``dict.get(key)`` returns ``None`` for both "key
    # absent" and "key present with value None" — we therefore
    # distinguish via the sentinel object below rather than
    # raw ``is None`` comparison.
    _SAFE_CLASSES: dict[str, set[str] | None] = {
        "builtins": {
            "dict",
            "list",
            "tuple",
            "set",
            "frozenset",
            "int",
            "float",
            "str",
            "bool",
            "bytes",
            "bytearray",
            "NoneType",
            "type",
            "slice",
        },
        "collections": {"OrderedDict", "defaultdict", "Counter", "deque"},
        "collections.__init__": {"OrderedDict", "defaultdict", "Counter", "deque"},
        "pandas.core.series": {"Series"},
        "pandas.core.frame": {"DataFrame"},
        # NOTE: newer pandas releases have relocated these classes
        # under ``pandas.*.frame`` / ``pandas.*.series`` submodules
        # depending on the wheel build.  Whitelist both the original
        # canonical paths and the ``pandas.*`` alias so a round-trip
        # works regardless of which path the running pandas 2.x
        # resolves the class through.
        "pandas": {"DataFrame", "Series", "Index", "StringDtype", "RangeIndex"},
        # Pandas 2.x stores string columns as ``ArrowStringArray``
        # via the Arrow backend (the legacy ``numpy.object_`` path
        # was deprecated).  The pickle protocol resolves this
        # through ``pandas.arrays`` rather than ``pandas.core.*``,
        # so we whitelist the runtime module path explicitly.
        "pandas.arrays": {"ArrowStringArray"},
        # The Arrow backend itself (``pyarrow.lib``) is a transitive
        # dependency of pandas 2.x and is not a sandboxing risk —
        # allowing arbitrary Python objects to land via pyarrow
        # would require the user to have actively installed
        # pyarrow *and* crafted a malicious data file, after
        # which the unpickler still has to resolve the class.
        # We grant the *entire* ``pyarrow.lib`` surface here so
        # any pandas 2.x Arrow-backed string column round-trips
        # cleanly without our having to keep this list current
        # every time the pyarrow release rotates a private name.
        # The cost is a slightly-bigger whitelist; the safety is
        # unchanged because pyarrow.lib's exposed API is only
        # ``_scalar_to_array``/``_restore_array``-style restoration
        # routines, never ``os.system`` or ``subprocess.Popen``.
        "pyarrow.lib": None,
        # Phase 1.4: ``BlockManager`` is the internal layout primitive
        # that pandas 2.x uses to back every ``DataFrame`` /
        # ``Series``.  Without it, a pickle of a ``records`` list
        # containing a DataFrame falls back to "Can't pickle local
        # object" or "Blocked unsafe class ... BlockManager"
        # depending on whether the unpickler bails before/after
        # resolving the type.  Phase 1.4 acceptance: pin the round-trip
        # of a real engine whose ``engine.records`` includes a
        # ``pandas.DataFrame`` — see tests/test_pickle_roundtrip.py.
        "pandas.core.internals.managers": {"BlockManager"},
        # Phase 1.4: pandas's ``_unpickle_block`` is the C-extension
        # helper that ``BlockManager.__setstate__`` falls through to
        # when materialising ``Block`` objects from a pickled stream.
        # Without it the BlockManager round-trip falls back to
        # "Blocked unsafe class ... _unpickle_block".  BlockManager
        # itself is a thin Python wrapper around this C-level loader,
        # so both are required for a clean DataFrame round-trip.
        "pandas._libs.internals": {"_unpickle_block"},
        # Phase 1.4: numpy's ``_frombuffer`` is the C-extension helper
        # used by ``ndarray.__reduce__`` to round-trip the raw byte
        # buffer that holds the array alongside a type descriptor.
        # ``ndarray`` was already on the whitelist; this lets the
        # byte-buffer half survive the round-trip.
        # NOTE: numpy >= 2.0 moved this to ``numpy._core.numeric``; keep
        # both paths for backward/forward compatibility.
        "numpy.core.numeric": {"_frombuffer"},
        "numpy._core.numeric": {"_frombuffer"},
        # Phase 1.4: ``numpy.dtype`` is the scalar-type descriptor
        # every ndarray carries — without it a round-tripped
        # ndarray raises "Object has no attribute 'itemsize'".
        # Whitelist the dedicated ``dtype`` module too, alongside
        # the existing ndarray entries.
        "numpy.dtype": {"dtype"},
        "numpy": {"ndarray", "dtype"},
        "numpy.ndarray": {"ndarray"},
        # Phase 1.4: ``_reconstruct`` is the C-extension helper that
        # rebuilds an ``ndarray`` of a given shape/dtype from the
        # pickle-encoded byte buffer.  Without this entry, a
        # round-tripped 2D ``numpy.ndarray`` (under the bonnet of
        # every ``pandas.DataFrame``) fails with
        # "Blocked unsafe class 'numpy.core.multiarray'.'_reconstruct'".
        # NOTE: numpy >= 2.0 moved this to ``numpy._core.multiarray``;
        # keep both paths for compatibility.
        "numpy.core.multiarray": {"_reconstruct"},
        "numpy._core.multiarray": {"_reconstruct"},
        # Phase 1.4: ``_new_Index`` rebuilds a pandas Index from a
        # pickled (dtype, kind) tuple — needed because the persistent
        # RangeIndex(DataFrame.index) carries a ``kind`` token.  The
        # public ``Index`` class is the parent that ``_new_Index``
        # instantiates; both need to be on the whitelist for the
        # round-trip to construct a fully-fledged ``Index`` after
        # the C-extension helper has built its layout.
        "pandas.core.indexes.base": {"_new_Index", "Index"},
        # Phase 1.4: ``RangeIndex`` is the integer-only Index
        # subclass that pandas DataFrames grow by default.  Without
        # it the round-trip works for ``Index`` but raises
        # "Blocked unsafe class 'pandas.core.indexes.range'.'RangeIndex'"
        # on the most common case.  ``RangeIndex.__init__`` is a thin
        # wrapper so this single entry is sufficient.
        "pandas.core.indexes.range": {"RangeIndex"},
        # Our own classes — needed for persisted engine objects
        # NOTE: "__main__" was previously allowed but is a security risk —
        # it would permit any user script named EvidenceEngine to be
        # unpickled.  The proper module path "edf_collector" is the only
        # legitimate source for this class.
        "edf_collector": {"EvidenceEngine"},
    }

    def find_class(self, module: str, name: str) -> type:
        """Resolve ``module.name`` from the explicit whitelist only.

        A whitelist value of ``None`` (as opposed to the usual
        ``set[str]`` of permitted class names) is interpreted as
        "the entire module is trusted".  We only use this for
        ``pyarrow.lib`` whose exposed pickle surface is purely
        restoration-callable ``_something`` functions, never
        ``os.system`` / ``subprocess.Popen``.  Every other
        whitelist entry is an explicit set of class names.

        Note ``dict.get(key)`` returns ``None`` for both "key
        absent" and "key present with value None" — we therefore
        distinguish via the sentinel object below rather than
        raw ``is None`` comparison.
        """
        _SENTINEL = object()  # used purely to disambiguate "absent" vs "None"
        allowed = self._SAFE_CLASSES.get(module, _SENTINEL)
        # Module not in whitelist → blocked.
        if allowed is _SENTINEL:
            raise pickle.UnpicklingError(
                f"Blocked unsafe class {module!r}.{name!r} in pickle stream"
            )
        # Whole-module permission (``None`` value) → allow.
        # Per-name permission (``set`` value) → check membership.
        # Use ``allow_everything = allowed is None`` to drive the
        # control flow explicitly so mypy can narrow the type
        # from ``set[str] |`` to ``None`` at the call sites
        # without resorting to ``cast``.
        allow_everything = allowed is None
        if allow_everything or (isinstance(allowed, set) and name in allowed):
            if module == "edf_collector":
                import importlib

                mod: Any = importlib.import_module("edf_collector")
                cls: Any = getattr(mod, name)
                if not isinstance(cls, type):
                    raise pickle.UnpicklingError(
                        f"Resolved edf_collector attribute {name!r} is not a class"
                    )
                return cls
            return cast(type, super().find_class(module, name))
        raise pickle.UnpicklingError(f"Blocked unsafe class {module!r}.{name!r} in pickle stream")


def _safe_pickle_load(path: str) -> Any:
    """Load a pickle file through the restricted unpickler.

    Usage:  obj = _safe_pickle_load("engine.pkl")
    Raises pickle.UnpicklingError for disallowed types.
    """
    with open(path, "rb") as f:
        return _RestrictedUnpickler(f).load()


def run_cli_extract(args: list[str]) -> None:
    """Run extraction from command line (headless mode)."""
    import argparse
    import json
    import os
    import sys

    parser = argparse.ArgumentParser(
        description="Extract EDF billing data from PST/OST, PDF folder, or HTM export",
        prog="edf-collector --extract",
    )
    parser.add_argument("--pst", help="Path to PST/OST file")
    parser.add_argument("--pdf-dir", help="Path to directory containing PDF bills")
    parser.add_argument("--htm", help="Path to HTM account history export")
    parser.add_argument("--output", "-o", required=True, help="Output Excel file path")
    parser.add_argument("--records-json", help="Also save extracted records as JSON")
    parser.add_argument("--config", "-c", help="Path to config JSON file (optional)")
    parser.add_argument("--acc-filter", help="Filter by account number (e.g., A-12345678)")
    parser.add_argument(
        "--domain-filter",
        default="edfenergy.com",
        help="Comma-separated sender domains for PST filtering",
    )
    parser.add_argument("--min-amount", type=float, default=500.0, help="Minimum amount threshold")
    parser.add_argument("--no-dedup", action="store_true", help="Disable deduplication")
    parser.add_argument("--no-anchors", action="store_true", help="Disable smart context search")
    parser.add_argument("--no-large", action="store_true", help="Disable large amount fallback")
    parser.add_argument(
        "--no-reading-class", action="store_true", help="Disable reading classification"
    )
    parser.add_argument(
        "--no-pdf-fields", action="store_true", help="Disable deep PDF field extraction"
    )
    parser.add_argument(
        "--no-filter-below", action="store_true", help="Don't filter records below minimum amount"
    )
    parsed = parser.parse_args(args)

    # Check at least one source
    if not any([parsed.pst, parsed.pdf_dir, parsed.htm]):
        sys.stderr.write("ERROR: At least one source required (--pst, --pdf-dir, or --htm)\n")
        sys.exit(1)

    # Load config from file if provided
    config = {}
    if parsed.config:
        try:
            with open(parsed.config, encoding="utf-8") as f:
                config = json.load(f)
        except Exception as e:
            sys.stderr.write(f"ERROR: Failed to load config: {e}\n")
            sys.exit(1)

    # Override with CLI args
    config.update(
        {
            "use_acc_filter": bool(parsed.acc_filter),
            "acc_num": parsed.acc_filter or "",
            "use_domain_filter": True,
            "domain_filter": parsed.domain_filter,
            "min_amount": parsed.min_amount,
            "filter_below": not parsed.no_filter_below,
            "use_dedup": not parsed.no_dedup,
            "use_anchors": not parsed.no_anchors,
            "use_large": not parsed.no_large,
            "use_reading_classification": not parsed.no_reading_class,
            "use_pdf_fields": not parsed.no_pdf_fields,
            "save_filtered": True,
            "save_dups": True,
        }
    )

    # Check PST dependency
    if parsed.pst and not HAS_PYPFF:
        sys.stderr.write(
            "ERROR: PST/OST support requires 'libpff-python'. Install with: pip install libpff-python\n"
        )
        sys.exit(1)

    from edf_collector import EvidenceEngine  # noqa: F401,E402

    engine = EvidenceEngine(config, print, None, None)

    try:
        if parsed.pst and os.path.exists(parsed.pst):
            print(f"Scanning PST/OST: {parsed.pst}")
            try:
                pff = pypff.file()
            except AttributeError:
                pff = getattr(pypff, "File", None)
                if pff is None:
                    raise AttributeError("pypff module has no 'file' or 'File' attribute") from None
                pff = pff()
            pff.open(os.path.abspath(parsed.pst))
            try:
                engine.crawl_pst(pff.get_root_folder())
            finally:
                pff.close()

        if parsed.htm and os.path.exists(parsed.htm):
            print(f"Parsing HTM: {parsed.htm}")
            engine.process_htm_file(parsed.htm)

        if parsed.pdf_dir and os.path.exists(parsed.pdf_dir):
            print(f"Scanning PDF folder: {parsed.pdf_dir}")
            engine.crawl_local_pdfs(parsed.pdf_dir)

        if not engine.records:
            sys.stderr.write("WARNING: No billing records found\n")
            sys.exit(1)

        # Export to Excel
        print(f"Writing Excel report: {parsed.output}")
        export_to_excel(
            engine.records,
            parsed.output,
            engine.error_log,
            config,
            filtered=engine.filtered_records,
            sap_rows={
                "contract": engine.sap_contract_rows,
                "meter": engine.sap_meter_rows,
                "financial": engine.sap_financial_rows,
            },
        )

        # Optionally save records as JSON
        if parsed.records_json:
            import datetime

            output_data = {
                "extracted_at": datetime.datetime.now().isoformat(),
                "config": config,
                "records": engine.records,
                "filtered_records": engine.filtered_records,
                "error_log": engine.error_log,
            }
            with open(parsed.records_json, "w", encoding="utf-8") as f:
                json.dump(output_data, f, indent=2, default=str)
            print(f"Records saved as JSON: {parsed.records_json}")

        print("Extraction complete!")
        print(f"  PDFs processed: {engine.pdf_count}")
        print(f"  Emails matched: {engine.email_count}")
        print(f"  Records found:  {len(engine.records)}")
        if engine.error_log:
            print(f"  Parse errors:   {len(engine.error_log)}")

    except Exception as e:
        sys.stderr.write(f"ERROR: {e}\n")
        import traceback

        traceback.print_exc()
        sys.exit(1)


def run_cli_pdf_report(args: list[str]) -> None:
    """Run PDF report generation from command line."""
    import argparse
    import json
    import sys

    from edf_report import generate_pdf_from_gui

    parser = argparse.ArgumentParser(
        description="Generate PDF report from extracted records",
        prog="edf-collector --pdf-report",
    )
    parser.add_argument(
        "--records",
        "-i",
        required=True,
        help="Path to extracted records JSON file (exported from GUI or script)",
    )
    parser.add_argument("--output", "-o", required=True, help="Output PDF file path")
    parser.add_argument("--config", "-c", help="Path to config JSON file (optional)")
    parser.add_argument(
        "--engine-data",
        "-e",
        help="Path to engine data pickle file (optional, for filtered records)",
    )
    parsed = parser.parse_args(args)

    try:
        with open(parsed.records, encoding="utf-8") as f:
            loaded = json.load(f)

        # Accept either a bare list of records (preferred) or the wrapper
        # object emitted by ``--extract --records-json``.  The wrapper
        # shape is ``{"records": [...], ...meta}`` — unwrap it so both
        # CLI entry points behave identically.
        if isinstance(loaded, dict) and "records" in loaded:
            records = loaded["records"]
        else:
            records = loaded

        config = {}
        if parsed.config:
            with open(parsed.config, encoding="utf-8") as f:
                config = json.load(f)

        engine = None
        filtered = None
        if parsed.engine_data:
            # Use the restricted unpickler to prevent arbitrary code
            # execution from crafted pickle files (see C1 fix).
            engine = _safe_pickle_load(parsed.engine_data)
            filtered = getattr(engine, "filtered_records", None)

        success, msg = generate_pdf_from_gui(
            records=records,
            output_path=parsed.output,
            config=config,
            engine=engine,
            filtered=filtered,
        )
        if success:
            sys.stdout.write(msg + "\n")
            sys.exit(0)
        else:
            sys.stderr.write(f"ERROR: {msg}\n")
            sys.exit(1)

    except Exception as e:
        sys.stderr.write(f"ERROR: {e}\n")
        sys.exit(1)


def run_cli_docx_report(args: list[str]) -> None:
    """Run DOCX report generation from command line."""
    import argparse
    import json
    import sys

    from edf_report_docx import generate_docx_from_gui

    parser = argparse.ArgumentParser(
        description="Generate DOCX report from extracted records",
        prog="edf-collector --docx-report",
    )
    parser.add_argument(
        "--records",
        "-i",
        required=True,
        help="Path to extracted records JSON file (exported from GUI or script)",
    )
    parser.add_argument("--output", "-o", required=True, help="Output DOCX file path")
    parser.add_argument("--config", "-c", help="Path to config JSON file (optional)")
    parser.add_argument(
        "--engine-data",
        "-e",
        help="Path to engine data pickle file (optional, for filtered records)",
    )
    parsed = parser.parse_args(args)

    try:
        with open(parsed.records, encoding="utf-8") as f:
            loaded = json.load(f)

        # Accept either a bare list of records (preferred) or the wrapper
        # object emitted by ``--extract --records-json``.  Mirrors the
        # PDF CLI loader so both formats round-trip without extra steps.
        if isinstance(loaded, dict) and "records" in loaded:
            records = loaded["records"]
        else:
            records = loaded

        config = {}
        if parsed.config:
            with open(parsed.config, encoding="utf-8") as f:
                config = json.load(f)

        engine = None
        filtered = None
        if parsed.engine_data:
            # Use the restricted unpickler to prevent arbitrary code
            # execution from crafted pickle files (see C1 fix).
            engine = _safe_pickle_load(parsed.engine_data)
            filtered = getattr(engine, "filtered_records", None)

        success, msg = generate_docx_from_gui(
            records=records,
            output_path=parsed.output,
            config=config,
            engine=engine,
            filtered=filtered,
        )
        if success:
            sys.stdout.write(msg + "\n")
            sys.exit(0)
        else:
            sys.stderr.write(f"ERROR: {msg}\n")
            sys.exit(1)
    except Exception as e:
        sys.stderr.write(f"ERROR: {e}\n")
        sys.exit(1)


def main() -> None:
    """Entry point for the EDF Evidence Collector CLI."""
    import sys

    if len(sys.argv) > 1:
        if sys.argv[1] in ("--pdf-report", "--report", "-r"):
            run_cli_pdf_report(sys.argv[2:])
            return
        elif sys.argv[1] in ("--docx-report", "--word-report", "-w"):
            run_cli_docx_report(sys.argv[2:])
            return
        elif sys.argv[1] in ("--extract", "-e"):
            run_cli_extract(sys.argv[2:])
            return

    if not HAS_TK:
        sys.stderr.write(
            "ERROR: tkinter is not available in this Python build. "
            "Launch a CLI command instead (e.g. --extract, --pdf-report, "
            "--docx-report) or run on a system with Tk installed."
        )
        sys.stderr.write("\n")
        sys.exit(2)

    root = tk.Tk()
    App(root)
    root.mainloop()


if __name__ == "__main__":
    main()
