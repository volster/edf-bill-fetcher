"""Excel cell primitives and helpers used across the evidence workbook.

Extracted from ``edf_collector.py`` as part of the modularization refactor
(Task 3).  These cover:

- Cell primitives: ``hcell``, ``text``, ``num``, ``money``, ``section_hdr``
- Column-width and formula-injection guards
- Text-warning suppression queue + post-save ``<ignoredErrors>`` injection
- SAP row index map builder
- PDF hyperlink cell helper

All functions are openpyxl-based; they have no internal cross-dependencies
with ``date_utils.py``, ``formatting.py``, or ``pdf_utils.py``.
"""

from __future__ import annotations

import os
import tempfile
import xml.etree.ElementTree as ET
import zipfile
from typing import Any

import openpyxl
import openpyxl.cell
import openpyxl.worksheet.hyperlink
import openpyxl.worksheet.worksheet
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill

# CELL_BORDER moved to edf_bill_fetcher.helpers.theme (Task 2).  Re-export
# here so existing call sites that import it from excel_utils keep working
# until Phase 7 deletes the compat shim.
from edf_bill_fetcher.helpers.theme import CELL_BORDER  # noqa: E402,F401

_TEXT_SUPPRESSION_QUEUE: dict[str, list[tuple[str, int, int]]] = {}


def hcell(ws: Any, row: int, col: int, value: Any, bg: str = "FE5716") -> Any:
    """Header cell helper — bold white text on coloured fill."""
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
    c.fill = PatternFill("solid", start_color=bg)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = CELL_BORDER
    return c


def text(
    ws: Any,
    r: int,
    c: int,
    val: Any,
    bold: bool = False,
    fill_hex: str | None = None,
    wrap: bool = False,
    align: str = "left",
    color: str = "000000",
) -> Any:
    """Text cell with formula-injection guard.

    Phase 2.x — formula-injection guard. External text (PDF/PST/email)
    can start with ``=``, ``+``, ``-`` or ``@`` and Excel will silently
    evaluate the cell as a formula. We coerce non-strings via ``str()``
    and prefix a leading ``=``, ``+``, ``-`` or ``@`` with an apostrophe.
    """
    safe_val: str
    if val is None:
        safe_val = ""
    else:
        safe_val = str(val)
        if safe_val and safe_val[0] in "+-=@":
            safe_val = "'" + safe_val
    cell = ws.cell(row=r, column=c, value=safe_val)
    cell.data_type = "s"
    cell.font = Font(name="Calibri", size=10, bold=bold, color=color)
    cell.border = CELL_BORDER
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    if fill_hex:
        cell.fill = PatternFill("solid", start_color=fill_hex)
    return cell


def num(
    ws: Any,
    r: int,
    c: int,
    val: Any,
    fmt: str = "#,##0",
    bold: bool = False,
    fill_hex: str | None = None,
) -> Any:
    """Numeric cell helper."""
    cell = ws.cell(row=r, column=c, value=val)
    cell.font = Font(name="Calibri", size=10, bold=bold)
    cell.border = CELL_BORDER
    cell.number_format = fmt
    cell.alignment = Alignment(horizontal="right", vertical="center")
    if fill_hex:
        cell.fill = PatternFill("solid", start_color=fill_hex)
    return cell


def money(
    ws: Any, r: int, c: int, val: Any, bold: bool = False, fill_hex: str | None = None
) -> Any:
    """Money-formatted cell (£#,##0.00)."""
    cell = ws.cell(row=r, column=c, value=val)
    cell.font = Font(name="Calibri", size=10, bold=bold)
    cell.border = CELL_BORDER
    cell.number_format = "£#,##0.00"
    cell.alignment = Alignment(horizontal="right", vertical="center")
    if fill_hex:
        cell.fill = PatternFill("solid", start_color=fill_hex)
    return cell


def section_hdr(ws: Any, r: int, label: str, ncols: int = 3, bg: str = "10367A") -> None:
    """Write a section header spanning ncols columns at row r."""
    for c in range(1, ncols + 1):
        cell = ws.cell(row=r, column=c, value=label if c == 1 else "")
        cell.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", start_color=bg)
        cell.border = CELL_BORDER
        cell.alignment = Alignment(horizontal="left", vertical="center")


def set_column_widths_from_spec(
    ws: openpyxl.worksheet.worksheet.Worksheet, widths: dict[str, float]
) -> None:
    """Apply column-width pins from a ``{col_letter: width}`` dict."""
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width


def suppress_text_warning(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    col_letter: str,
    start_row: int,
    end_row: int,
) -> None:
    """Queue a text-ID suppression for a column/row range to be injected after save."""
    key = ws.title
    _TEXT_SUPPRESSION_QUEUE.setdefault(key, []).append((col_letter, start_row, end_row))


def _sheet_title_by_xml_path(zin: zipfile.ZipFile) -> dict[str, str]:
    """Map ``xl/worksheets/sheetN.xml`` paths to their sheet titles.

    Worksheet XML parts carry no title of their own; the mapping lives in
    ``xl/workbook.xml`` (``<sheet name="..." r:id="..."/>``) joined with
    ``xl/_rels/workbook.xml.rels`` (``rId`` -> ``Target``).  Returns an
    empty dict when either part is missing or unparseable.
    """
    try:
        workbook = ET.fromstring(zin.read("xl/workbook.xml"))
        rels = ET.fromstring(zin.read("xl/_rels/workbook.xml.rels"))
    except (KeyError, ET.ParseError):
        return {}

    rid_to_target: dict[str, str] = {}
    for rel in rels:
        if not rel.tag.endswith("}Relationship") and rel.tag != "Relationship":
            continue
        rid = rel.attrib.get("Id")
        target = rel.attrib.get("Target")
        if rid and target:
            target = target.lstrip("/")
            if not target.startswith("xl/"):
                target = "xl/" + target
            rid_to_target[rid] = target

    title_by_path: dict[str, str] = {}
    for sheet in workbook.iter():
        if not sheet.tag.endswith("}sheet") and sheet.tag != "sheet":
            continue
        name = sheet.attrib.get("name")
        rid = sheet.attrib.get(
            "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id"
        )
        if name and rid and rid in rid_to_target:
            title_by_path[rid_to_target[rid]] = name
    return title_by_path


def suppress_text_warnings_post_save(output_path: str) -> None:
    """Post-save zip injection for ``<ignoredErrors>`` blocks.

    openpyxl 3.1.5 silently drops ``ws.ignored_errors`` on save, but
    the round-trip through a reopen-and-re-inject approach persists.
    """
    tmp = tempfile.NamedTemporaryFile(
        suffix=".xlsx", delete=False, dir=os.path.dirname(output_path) or "."
    )
    tmp.close()
    os.replace(output_path, tmp.name)

    with zipfile.ZipFile(tmp.name, "r") as zin:
        title_by_path = _sheet_title_by_xml_path(zin)
        with zipfile.ZipFile(output_path, "w") as zout:
            for item in zin.namelist():
                data = zin.read(item)
                if item.startswith("xl/worksheets/sheet") and item.endswith(".xml"):
                    xml = data.decode("utf-8", errors="replace")
                    sheet_title = title_by_path.get(item)
                    suppressions = _TEXT_SUPPRESSION_QUEUE.get(sheet_title) if sheet_title else None
                    if suppressions:
                        for col_letter, start_row, end_row in suppressions:
                            sqref = f"{col_letter}{start_row}:{col_letter}{end_row}"
                            block = (
                                f'<ignoredErrors><ignoredError sqref="{sqref}" '
                                f'numberStoredAsText="1"/></ignoredErrors>'
                            )
                            xml = xml.replace("</worksheet>", f"{block}</worksheet>")
                    data = xml.encode("utf-8")
                zout.writestr(item, data)

    _TEXT_SUPPRESSION_QUEUE.clear()
    os.unlink(tmp.name)


def build_sap_row_index_map(sap_financial: list[dict]) -> dict[int, int]:
    """Return a map from ``id(sap_row)`` -> Excel row on SAP Financial Transactions.

    The SAP Financial Transactions sheet writes the header at row 3
    and the first data row at row 4.
    """
    out: dict[int, int] = {}
    for i, r in enumerate(sap_financial):
        out[id(r)] = 4 + i
    return out


def open_pdf_hyperlink_cell(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    row: int,
    col: int,
    evidence_df: pd.DataFrame | None,
    invoice_number: str,
) -> None:
    """Emit a ``→`` hyperlink cell at (``row``, ``col``) jumping to the matching row on EDF Evidence Report.

    ``None``/empty invoice numbers are skipped. Display is a right-arrow
    glyph; tooltip carries the invoice number for accessibility.
    """
    if not invoice_number:
        return
    target_row = None
    if evidence_df is not None and not evidence_df.empty and "Invoice #" in evidence_df.columns:
        matches = evidence_df[evidence_df["Invoice #"].astype(str) == str(invoice_number)]
        if not matches.empty:
            target_row = matches.iloc[0].name + 2  # +2 for header rows
    if target_row is not None:
        cell = ws.cell(row=row, column=col, value="\u2192")
        cell.hyperlink = openpyxl.worksheet.hyperlink.Hyperlink(
            ref=cell.coordinate,
            location=f"'EDF Evidence Report'!A{target_row}",
            display="\u2192",
            tooltip=f"Jump to EDF Evidence Report!A{target_row}",
        )
        cell.font = Font(name="Calibri", size=10, color="0000FF", underline="single")


__all__ = [
    "hcell",
    "text",
    "num",
    "money",
    "section_hdr",
    "set_column_widths_from_spec",
    "suppress_text_warning",
    "suppress_text_warnings_post_save",
    "build_sap_row_index_map",
    "open_pdf_hyperlink_cell",
    "CELL_BORDER",
]
