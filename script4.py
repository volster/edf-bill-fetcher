#!/usr/bin/env python3
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import pypff
import pandas as pd
import re
import os
import gc
import threading
import pdfplumber
import tempfile
import traceback
from bs4 import BeautifulSoup
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.series import SeriesLabel

# --- Branding & Colors ---
EDF_ORANGE, EDF_NAVY, EDF_OFFWHITE = "#FE5716", "#10367A", "#F5F5F5"
EST_YELLOW, JUMP_RED, DUP_GREY = "FFFF99", "FF9999", "E0E0E0"

# --- Extraction Patterns ---
AMOUNT_PATTERNS = [
    r"balance[\s\S]{0,30}?£\s?([\d,]+(?:\.\d{2})?)",
    r"total charges[\s\S]{0,30}?£\s?([\d,]+(?:\.\d{2})?)",
    r"total amount due[\s\S]{0,30}?£\s?([\d,]+(?:\.\d{2})?)",
    r"amount to pay[\s\S]{0,30}?£\s?([\d,]+(?:\.\d{2})?)",
    r"£\s?([\d,]+(?:\.\d{2})?)\s*(?:in\s*)?debit",
    r"current balance[\s\S]{0,30}?£\s?([\d,]+(?:\.\d{2})?)"
]

READING_PATTERNS = {
    "Estimated": re.compile(r"estimated|est\.|estimate", re.IGNORECASE),
    "Actual":    re.compile(r"actual|customer reading|your reading", re.IGNORECASE),
    "Smart":     re.compile(r"smart meter|automated reading|smart reading", re.IGNORECASE)
}

# Billing period — catches "1 June 2023 to 30 June 2023" and "01/06/2023 to 30/06/2023"
PERIOD_RE = re.compile(
    r'(\d{1,2}(?:\s+\w+\s+\d{4}|\s*/\s*\d{2}\s*/\s*\d{4}|\s*-\s*\d{2}\s*-\s*\d{4}))'
    r'\s*(?:to|to:|–|-)\s*'
    r'(\d{1,2}(?:\s+\w+\s+\d{4}|\s*/\s*\d{2}\s*/\s*\d{4}|\s*-\s*\d{2}\s*-\s*\d{4}))',
    re.IGNORECASE
)


_ISO_DATE_RE = re.compile(r'^\d{4}-\d{2}-\d{2}$')

def parse_to_display_date(date_input):
    """Converts any date string or datetime to DD/MM/YYYY. Returns original on failure.
    ISO format (YYYY-MM-DD) is parsed without dayfirst to avoid a pandas UserWarning;
    all other formats use dayfirst=True so DD/MM/YYYY strings are interpreted correctly.
    """
    if not date_input or str(date_input).strip() in ("Unknown", ""):
        return date_input
    s = str(date_input).strip()
    try:
        if _ISO_DATE_RE.match(s):
            return pd.to_datetime(s, format='%Y-%m-%d').strftime('%d/%m/%Y')
        return pd.to_datetime(s, dayfirst=True, format='mixed').strftime('%d/%m/%Y')
    except Exception:
        return s


def parse_to_sort_date(date_input):
    """Returns a sortable datetime for internal use only."""
    s = str(date_input).strip() if date_input else ''
    try:
        if _ISO_DATE_RE.match(s):
            return pd.to_datetime(s, format='%Y-%m-%d')
        return pd.to_datetime(s, dayfirst=True, format='mixed')
    except Exception:
        return pd.NaT


class EvidenceEngine:
    def __init__(self, config, update_ui_cb):
        self.config      = config
        self.records     = []
        self.update_ui   = update_ui_cb
        self.pdf_count   = 0
        self.email_count = 0
        self.error_log   = []

    def log_error(self, context, err):
        self.error_log.append(f"[{datetime.now().strftime('%H:%M:%S')}] {context} — {err}")

    def find_billing_period(self, text):
        """Extract billing period start/end from text if present."""
        m = PERIOD_RE.search(text)
        if m:
            return (
                parse_to_display_date(m.group(1).strip()),
                parse_to_display_date(m.group(2).strip())
            )
        return "N/A", "N/A"

    def process_text(self, text, source_type, detail, fallback_date):
        if not text:
            return

        clean_text = re.sub(r'\s+', ' ', text)

        # Account filter (strips spaces/dashes for PDF robustness)
        if self.config["use_acc_filter"]:
            acc = re.sub(r'[\s\-]', '', self.config["acc_num"])
            if acc and acc not in re.sub(r'[\s\-]', '', clean_text):
                return

        found_amt, strategy = None, ""

        # Pass 1: Smart context anchors
        if self.config["use_anchors"]:
            for p in AMOUNT_PATTERNS:
                m = re.search(p, clean_text, re.IGNORECASE)
                if m:
                    try:
                        found_amt = float(m.group(1).replace(',', ''))
                        strategy  = "Smart Context"
                        break
                    except Exception:
                        continue

        # Pass 2: Largest amount above threshold
        if not found_amt and self.config["use_large"]:
            matches = re.findall(r'£\s?(\d{1,3}(?:,\d{3})*(?:\.\d{2})?)', clean_text)
            if matches:
                floats = [float(x.replace(',', '')) for x in matches]
                highs  = [x for x in floats if x >= self.config["min_amount"]]
                if highs:
                    found_amt = max(highs)
                    strategy  = "Large Amount Fallback"

        if not found_amt:
            return

        # Post-extraction filter: discard records below minimum threshold.
        # filter_strategy controls which extraction methods this applies to:
        #   "all"         — strip anything below threshold regardless of how it was found
        #   "anchor_only" — only strip low values found via Smart Context
        #   "large_only"  — only strip low values found via Large Amount Fallback
        if self.config.get("filter_below", False) and found_amt < self.config["min_amount"]:
            fs = self.config.get("filter_strategy", "all")
            if fs == "all":
                return
            elif fs == "anchor_only" and strategy == "Smart Context":
                return
            elif fs == "large_only" and strategy == "Large Amount Fallback":
                return

        # Reading type
        r_type = "Unknown"
        if self.config["use_readings"]:
            for label, pat in READING_PATTERNS.items():
                if pat.search(clean_text):
                    r_type = label
                    break

        # Deep PDF fields
        units_used = standing_charge = inv_num = "N/A"
        date_to_use = fallback_date

        if "PDF" in source_type:
            date_m = re.search(
                r"(?:Bill date|Date issued):\s*[\",]*\s*(\d{1,2}\s\w+\s\d{4})",
                clean_text, re.IGNORECASE
            )
            if date_m:
                date_to_use = parse_to_display_date(date_m.group(1))

        u_m  = re.search(r'([\d,]+)\s*kWh',                           clean_text, re.IGNORECASE)
        sc_m = re.search(r'(\d+\.\d{2})p\s*per day',                  clean_text, re.IGNORECASE)
        in_m = re.search(r'Invoice number[\s:,\"\'\n]*([A-Z0-9\-]+)', clean_text, re.IGNORECASE)

        if u_m:  units_used      = u_m.group(1)
        if sc_m: standing_charge = sc_m.group(1)
        if in_m: inv_num         = in_m.group(1)

        period_from, period_to = self.find_billing_period(clean_text)

        self.records.append({
            "Source":               source_type,
            "Date":                 date_to_use,
            "Period From":          period_from,
            "Period To":            period_to,
            "Invoice #":            inv_num,
            "Amount (£)":           found_amt,
            "Reading":              r_type,
            "Units (kWh)":          units_used,
            "Standing Chg (p/day)": standing_charge,
            "Details":              detail[:60],
            "Logic Used":           strategy
        })

    def process_pdf_file(self, path, source_label, detail_label, fallback_date):
        try:
            with pdfplumber.open(path) as pdf:
                pdf_text = " ".join([p.extract_text() or "" for p in pdf.pages])
                self.process_text(pdf_text, source_label, detail_label, fallback_date)
        except Exception as e:
            self.log_error(f"PDF: {detail_label}", str(e))

    def crawl_pst(self, folder):
        for i in range(folder.get_number_of_sub_messages()):
            try:
                msg  = folder.get_sub_message(i)
                subj = str(msg.get_subject() or "")

                d_time   = msg.get_delivery_time()
                date_str = parse_to_display_date(d_time.strftime('%Y-%m-%d')) if d_time else "Unknown"

                # Email body — filter to EDF-related subjects
                if any(k in subj.upper() for k in ["EDF", "BILL", "STATEMENT", "ACCOUNT", "INVOICE"]):
                    self.email_count += 1
                    html  = msg.get_html_body()
                    plain = msg.get_plain_text_body()

                    if html:
                        body_text = BeautifulSoup(html, 'html.parser').get_text(separator=' ')
                        self.process_text(body_text, "Email Body", subj, date_str)
                    elif plain:
                        self.process_text(plain.decode('utf-8', errors='ignore'), "Email Body", subj, date_str)
                    else:
                        # Older emails (pre-2021) are often RTF only — strip control words to extract text
                        rtf_body = None
                        try:
                            rtf_body = msg.get_rtf_body()
                        except Exception:
                            pass
                        if rtf_body:
                            try:
                                rtf_str  = rtf_body.decode('utf-8', errors='replace')
                                rtf_text = re.sub(r'\\[a-z]+[-\d]*\s?', ' ', rtf_str)
                                rtf_text = re.sub(r'[{}\\]', ' ', rtf_text)
                                self.process_text(rtf_text, "Email Body (RTF)", subj, date_str)
                            except Exception as e:
                                self.log_error(f"Email: {subj} ({date_str})", f"RTF decode failed: {e}")
                        else:
                            self.log_error(f"Email: {subj} ({date_str})", "No readable body (tried HTML, plain, RTF)")

                # PDF attachments — detected by magic bytes, not file extension
                for a_idx in range(msg.get_number_of_attachments()):
                    try:
                        att  = msg.get_attachment(a_idx)
                        size = att.get_size()
                        if size > 4:
                            buf = att.read_buffer(size)
                            if buf and buf.startswith(b'%PDF'):
                                self.pdf_count += 1
                                # pypff uses get_long_filename / get_short_filename, not get_name
                                try:
                                    att_name = att.get_long_filename() or att.get_short_filename() or f"Attachment_{self.pdf_count}.pdf"
                                except Exception:
                                    att_name = f"Attachment_{self.pdf_count}.pdf"

                                with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp:
                                    tmp.write(buf)
                                    tmp_path = tmp.name
                                try:
                                    self.process_pdf_file(tmp_path, "PST PDF Attachment", att_name, date_str)
                                finally:
                                    if os.path.exists(tmp_path):
                                        os.remove(tmp_path)
                    except Exception as e:
                        self.log_error(f"Attachment in '{subj}' ({date_str})", str(e))

            except Exception as e:
                self.log_error(f"Message index {i} in folder", str(e))

        self.update_ui(f"Scanned {self.email_count} emails, {self.pdf_count} attached PDFs…")

        for j in range(folder.get_number_of_sub_folders()):
            self.crawl_pst(folder.get_sub_folder(j))

    def crawl_local_pdfs(self, path):
        if not path or not os.path.exists(path):
            return
        for file in os.listdir(path):
            if file.lower().endswith(".pdf"):
                self.pdf_count += 1
                file_path     = os.path.join(path, file)
                fallback_date = parse_to_display_date(
                    datetime.fromtimestamp(os.path.getmtime(file_path)).strftime('%Y-%m-%d')
                )
                self.process_pdf_file(file_path, "Local PDF Folder", file, fallback_date)
        self.update_ui(f"Scanned {self.email_count} emails, {self.pdf_count} total PDFs…")


# ---------------------------------------------------------------------------
# Excel export helpers
# ---------------------------------------------------------------------------

THIN        = Side(style='thin', color='DDDDDD')
CELL_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _hcell(ws, row, col, value, bg="FE5716"):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
    c.fill      = PatternFill("solid", start_color=bg)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border    = CELL_BORDER
    return c


def write_evidence_sheet(ws, df, is_duplicate=False):
    headers = [
        "Source", "Date", "Period From", "Period To", "Invoice #",
        "Amount (£)", "% Change", "Reading", "Units (kWh)",
        "Standing Chg (p/day)", "Details", "Logic Used", "Anomaly Flag"
    ]
    bg = "888888" if is_duplicate else "FE5716"
    for col, h in enumerate(headers, 1):
        _hcell(ws, 1, col, h, bg=bg)
    ws.row_dimensions[1].height = 28

    alt_fill = PatternFill("solid", start_color="FFF3EE")

    for r_idx, row in enumerate(df.values, 2):
        row_fill = alt_fill if r_idx % 2 == 0 else PatternFill()

        for c_idx, val in enumerate(row, 1):
            c = ws.cell(row=r_idx, column=c_idx, value=val)
            c.font      = Font(name="Calibri", size=10)
            c.fill      = row_fill if not is_duplicate else PatternFill("solid", start_color=DUP_GREY)
            c.border    = CELL_BORDER
            c.alignment = Alignment(vertical="top")

            if c_idx == 6 and isinstance(val, (int, float)):   # Amount
                c.number_format = '£#,##0.00'
            if c_idx == 7 and isinstance(val, (int, float)):   # % Change
                c.number_format = '0.0%'
                c.alignment = Alignment(horizontal="right", vertical="top")

            # Estimated reading rows — yellow tint
            if not is_duplicate and len(row) > 7 and row[7] == "Estimated":
                c.fill = PatternFill("solid", start_color=EST_YELLOW)

        # Anomaly flag — >100% jump (col 13)
        if not is_duplicate and r_idx > 2:
            prev = ws.cell(row=r_idx - 1, column=6).value
            curr = row[5]  # Amount (£) in df column order
            if isinstance(prev, (int, float)) and isinstance(curr, (int, float)) and prev > 0:
                if curr > prev * 2:
                    c = ws.cell(row=r_idx, column=13, value="⚠ >100% INCREASE")
                    c.fill   = PatternFill("solid", start_color=JUMP_RED)
                    c.font   = Font(name="Calibri", size=10, bold=True)
                    c.border = CELL_BORDER

    # Column widths
    widths = {
        'A': 18, 'B': 13, 'C': 13, 'D': 13, 'E': 16,
        'F': 13, 'G': 10, 'H': 11, 'I': 12, 'J': 18,
        'K': 38, 'L': 18, 'M': 20
    }
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width

    ws.freeze_panes = "A2"


def write_summary_sheet(ws, df):
    ws.title = "Annual Summary"

    df = df.copy()
    df['_sort'] = pd.to_datetime(df['Date'], dayfirst=True, format='mixed', errors='coerce')
    df['Year']  = df['_sort'].dt.year.astype('Int64')

    yearly = (
        df.dropna(subset=['Year'])
          .groupby('Year', as_index=False)
          .agg(
              Total_Billed=('Amount (£)', 'sum'),
              Bill_Count  =('Amount (£)', 'count'),
              Average_Bill=('Amount (£)', 'mean'),
              Highest_Bill=('Amount (£)', 'max'),
              Lowest_Bill =('Amount (£)', 'min'),
          )
          .sort_values('Year')
    )

    headers = ["Year", "Total Billed (£)", "Number of Bills",
               "Average Bill (£)", "Highest Bill (£)", "Lowest Bill (£)"]
    for col, h in enumerate(headers, 1):
        _hcell(ws, 1, col, h, bg="10367A")
    ws.row_dimensions[1].height = 28

    alt_fill = PatternFill("solid", start_color="EEF2FF")
    for r_idx, row in enumerate(yearly.values, 2):
        row_fill = alt_fill if r_idx % 2 == 0 else PatternFill()
        for c_idx, val in enumerate(row, 1):
            c = ws.cell(row=r_idx, column=c_idx,
                        value=int(val) if c_idx == 1 else val)
            c.font      = Font(name="Calibri", size=10)
            c.fill      = row_fill
            c.border    = CELL_BORDER
            c.alignment = Alignment(
                horizontal="center" if c_idx == 1 else "right",
                vertical="top"
            )
            if c_idx == 2:
                c.number_format = '£#,##0.00'
            elif c_idx == 3:
                c.number_format = '#,##0'
            elif c_idx > 3:
                c.number_format = '£#,##0.00'

    # Grand total row
    n = len(yearly) + 2
    totals = [
        "TOTAL",
        yearly['Total_Billed'].sum(),
        int(yearly['Bill_Count'].sum()),
        yearly['Average_Bill'].mean(),
        yearly['Highest_Bill'].max(),
        yearly['Lowest_Bill'].min(),
    ]
    total_fill = PatternFill("solid", start_color="10367A")
    for c_idx, val in enumerate(totals, 1):
        c = ws.cell(row=n, column=c_idx, value=val)
        c.font      = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        c.fill      = total_fill
        c.border    = CELL_BORDER
        c.alignment = Alignment(
            horizontal="center" if c_idx == 1 else "right"
        )
        if c_idx == 2 and isinstance(val, float):
            c.number_format = '£#,##0.00'
        elif c_idx == 3:
            c.number_format = '#,##0'
        elif c_idx > 3 and isinstance(val, float):
            c.number_format = '£#,##0.00'

    for col_letter in ['A', 'B', 'C', 'D', 'E', 'F']:
        ws.column_dimensions[col_letter].width = 20
    ws.freeze_panes = "A2"


def export_to_excel(data, output_path, error_log, config):
    df = pd.DataFrame(data)

    # Sort chronologically
    df['_sort'] = pd.to_datetime(df['Date'], dayfirst=True, format='mixed', errors='coerce')
    df = df.sort_values(by=['_sort', 'Invoice #'], ascending=[True, False]).reset_index(drop=True)

    # % change column — calculated here so it survives manual row deletion gracefully
    # (stored as a decimal so Excel formats it as percentage)
    pct_changes = [None]
    for i in range(1, len(df)):
        prev = df.at[i - 1, 'Amount (£)']
        curr = df.at[i,     'Amount (£)']
        if isinstance(prev, (int, float)) and prev > 0:
            pct_changes.append((curr - prev) / prev)
        else:
            pct_changes.append(None)
    df['% Change'] = pct_changes

    # Deduplication
    dup_df = pd.DataFrame()
    if config["use_dedup"]:
        is_dup = df.duplicated(subset=['Date', 'Amount (£)'], keep='first')
        if config["save_dups"]:
            dup_df = df[is_dup].copy()
        df = df[~is_dup].reset_index(drop=True)

    # Drop internal sort column
    df     = df.drop(columns=['_sort'], errors='ignore')
    dup_df = dup_df.drop(columns=['_sort'], errors='ignore') if not dup_df.empty else dup_df

    # Column order — % Change sits right after Amount for easy reading
    col_order = [
        "Source", "Date", "Period From", "Period To", "Invoice #",
        "Amount (£)", "% Change", "Reading", "Units (kWh)",
        "Standing Chg (p/day)", "Details", "Logic Used"
    ]
    df     = df.reindex(columns=col_order)
    dup_df = dup_df.reindex(columns=col_order) if not dup_df.empty else dup_df

    wb = openpyxl.Workbook()

    # Tab 1: Annual Summary (first thing anyone sees)
    ws_summary = wb.active
    write_summary_sheet(ws_summary, df)

    # Tab 2: Full evidence
    ws_main = wb.create_sheet(title="EDF Evidence Report")
    write_evidence_sheet(ws_main, df, is_duplicate=False)

    # Tab 3: Duplicates
    if not dup_df.empty:
        ws_dup = wb.create_sheet(title="Duplicate Entries")
        write_evidence_sheet(ws_dup, dup_df, is_duplicate=True)

    # Tab 4: Parse errors (only created if there were errors)
    if error_log:
        ws_err = wb.create_sheet(title="Parse Errors")
        _hcell(ws_err, 1, 1, "Time",    bg="888888")
        _hcell(ws_err, 1, 2, "Context", bg="888888")
        _hcell(ws_err, 1, 3, "Error",   bg="888888")
        for r_idx, entry in enumerate(error_log, 2):
            # Format: "[HH:MM:SS] context — error"
            ts_m = re.match(r'\[(.+?)\]\s*(.*?)\s*—\s*(.*)', entry)
            if ts_m:
                ts, ctx, err = ts_m.group(1), ts_m.group(2), ts_m.group(3)
            else:
                ts, ctx, err = "", entry, ""
            for c_idx, val in enumerate([ts, ctx, err], 1):
                c = ws_err.cell(row=r_idx, column=c_idx, value=val)
                c.font   = Font(name="Calibri", size=10)
                c.border = CELL_BORDER
        ws_err.column_dimensions['A'].width = 10
        ws_err.column_dimensions['B'].width = 45
        ws_err.column_dimensions['C'].width = 60

    # =========================================================================
    # ANALYSIS SUITE — 4 additional sheets written after the evidence tabs
    # All analysis uses only bills ≥ £5,000 (noise filter).
    # The bills represent a CUMULATIVE ACCOUNT BALANCE, so the true periodic
    # charge for any period = closing_balance − opening_balance.
    # =========================================================================
    import numpy as np
    import statistics as _stats
    from openpyxl.chart import BarChart, LineChart, Reference

    NAVY   = "10367A"
    ORANGE = "FE5716"
    RED    = "FF6B6B"
    AMBER  = "FFD166"
    GREEN  = "06D6A0"
    LGREY  = "F0F0F0"
    DGREY  = "888888"

    def _money(ws, r, c, val, bold=False, fill_hex=None):
        cell = ws.cell(row=r, column=c, value=val)
        cell.font   = Font(name="Calibri", size=10, bold=bold)
        cell.border = CELL_BORDER
        cell.number_format = '£#,##0.00'
        cell.alignment = Alignment(horizontal="right", vertical="center")
        if fill_hex:
            cell.fill = PatternFill("solid", start_color=fill_hex)
        return cell

    def _text(ws, r, c, val, bold=False, fill_hex=None, wrap=False, align="left", color="000000"):
        cell = ws.cell(row=r, column=c, value=val)
        cell.font   = Font(name="Calibri", size=10, bold=bold, color=color)
        cell.border = CELL_BORDER
        cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
        if fill_hex:
            cell.fill = PatternFill("solid", start_color=fill_hex)
        return cell

    def _num(ws, r, c, val, fmt="#,##0", bold=False, fill_hex=None):
        cell = ws.cell(row=r, column=c, value=val)
        cell.font   = Font(name="Calibri", size=10, bold=bold)
        cell.border = CELL_BORDER
        cell.number_format = fmt
        cell.alignment = Alignment(horizontal="right", vertical="center")
        if fill_hex:
            cell.fill = PatternFill("solid", start_color=fill_hex)
        return cell

    def _section_hdr(ws, r, label, ncols=3, bg=NAVY):
        for c in range(1, ncols+1):
            cell = ws.cell(row=r, column=c, value=label if c==1 else "")
            cell.font   = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
            cell.fill   = PatternFill("solid", start_color=bg)
            cell.border = CELL_BORDER
            cell.alignment = Alignment(horizontal="left", vertical="center")

    # ------------------------------------------------------------------
    # Build clean analysis frame
    # ------------------------------------------------------------------
    df_an = df.copy()
    df_an['_dt'] = pd.to_datetime(df_an['Date'], dayfirst=True, format='mixed', errors='coerce')
    df_an = df_an.sort_values('_dt').reset_index(drop=True)
    dfc   = df_an[df_an['Amount (£)'] >= 5000].copy().reset_index(drop=True)
    dfc['year']  = dfc['_dt'].dt.year
    dfc['month'] = dfc['_dt'].dt.month

    if len(dfc) < 2:
        return   # not enough data to analyse; wb.save called unconditionally below

    amounts  = dfc['Amount (£)'].values.astype(float)
    dates_dt = dfc['_dt'].tolist()
    dates_lbl= dfc['Date'].tolist()
    n        = len(amounts)

    # Periodic charges (balance diffs between consecutive bills)
    raw_diffs  = np.diff(amounts)                        # can be negative (payments)
    pos_diffs  = raw_diffs[raw_diffs > 0]               # genuine charges only

    # Rolling 6-bill average
    rolling6 = pd.Series(amounts).rolling(6, min_periods=3).mean().tolist()

    # Linear trend
    x_idx   = np.arange(n)
    slope, intercept = np.polyfit(x_idx, amounts, 1)
    trend_line = (slope * x_idx + intercept).tolist()

    # Year groups
    yearly = dfc.groupby('year').agg(
        count   = ('Amount (£)', 'count'),
        avg_bal = ('Amount (£)', 'mean'),
        peak    = ('Amount (£)', 'max'),
        low     = ('Amount (£)', 'min'),
        total   = ('Amount (£)', 'sum'),
    ).reset_index()

    # ==========================================================================
    # TAB A: KEY STATISTICS
    # ==========================================================================
    ws_ks = wb.create_sheet(title="Key Statistics")
    ws_ks.column_dimensions['A'].width = 42
    ws_ks.column_dimensions['B'].width = 22
    ws_ks.column_dimensions['C'].width = 44

    # Title banner
    tc = ws_ks.cell(row=1, column=1, value="EDF ENERGY DISPUTE  —  KEY STATISTICS")
    tc.font  = Font(name="Calibri", size=13, bold=True, color="FFFFFF")
    tc.fill  = PatternFill("solid", start_color=ORANGE)
    tc.border= CELL_BORDER
    tc.alignment = Alignment(horizontal="left", vertical="center")
    for c in [2,3]:
        x = ws_ks.cell(row=1, column=c)
        x.fill   = PatternFill("solid", start_color=ORANGE)
        x.border = CELL_BORDER
    ws_ks.row_dimensions[1].height = 26

    def ks_row(r, label, value, note="", fmt=None, bold=False, alt=False):
        bg = LGREY if alt else None
        _text(ws_ks, r, 1, label, bold=bold, fill_hex=bg)
        if fmt == '£':
            _money(ws_ks, r, 2, value, bold=bold, fill_hex=bg)
        elif fmt == '%':
            _num(ws_ks, r, 2, value, fmt="0.0%", bold=bold, fill_hex=bg)
        elif fmt:
            _num(ws_ks, r, 2, value, fmt=fmt, bold=bold, fill_hex=bg)
        else:
            _text(ws_ks, r, 2, value, bold=bold, fill_hex=bg, align="right")
        _text(ws_ks, r, 3, note, fill_hex=bg, color=DGREY)

    first_bill = dfc.iloc[0]
    last_bill  = dfc.iloc[-1]
    span_days  = (last_bill['_dt'] - first_bill['_dt']).days
    span_months= span_days / 30.44

    r = 2;  _section_hdr(ws_ks, r, "ACCOUNT OVERVIEW")
    r = 3;  ks_row(r, "Account reference", "671078701920", alt=True)
    r = 4;  ks_row(r, "First bill on record", first_bill['Date'])
    r = 5;  ks_row(r, "Most recent bill", last_bill['Date'], alt=True)
    r = 6;  ks_row(r, "Period covered", f"{int(span_months)} months  ({span_days:,} days)")
    r = 7;  ks_row(r, "Total bills analysed (≥£5,000)", n, fmt="#,##0", alt=True)

    r = 8;  _section_hdr(ws_ks, r, "BALANCE FIGURES")
    r = 9;  ks_row(r, "Opening balance", first_bill['Amount (£)'], f"Bill dated {first_bill['Date']}", fmt='£', alt=True)
    r = 10; ks_row(r, "Current balance (latest bill)", last_bill['Amount (£)'], f"Bill dated {last_bill['Date']}", fmt='£', bold=True)
    r = 11; ks_row(r, "Total balance increase", last_bill['Amount (£)'] - first_bill['Amount (£)'], "Latest minus earliest", fmt='£', bold=True, alt=True)
    r = 12; ks_row(r, "% increase over full period", (last_bill['Amount (£)']-first_bill['Amount (£)'])/first_bill['Amount (£)'], "", fmt='%', bold=True)
    r = 13; ks_row(r, "Mean balance across all bills", float(np.mean(amounts)), "", fmt='£', alt=True)
    r = 14; ks_row(r, "Median balance", float(np.median(amounts)), "", fmt='£')
    r = 15; ks_row(r, "Peak balance recorded", float(np.max(amounts)), f"Bill dated {dfc.loc[dfc['Amount (£)']==np.max(amounts),'Date'].iloc[0]}", fmt='£', alt=True)
    r = 16; ks_row(r, "Lowest balance recorded (clean)", float(np.min(amounts)), f"Bill dated {dfc.loc[dfc['Amount (£)']==np.min(amounts),'Date'].iloc[0]}", fmt='£')

    r = 17; _section_hdr(ws_ks, r, "PERIODIC CHARGES  (balance difference between consecutive bills)")
    note_charges = "Bills represent a running cumulative balance; charge = closing − opening"
    r = 18; ks_row(r, "Note", note_charges, alt=True)
    if len(pos_diffs):
        r = 19; ks_row(r, "Mean charge per period", float(np.mean(pos_diffs)), "Average across all positive-diff periods", fmt='£')
        r = 20; ks_row(r, "Median charge per period", float(np.median(pos_diffs)), "", fmt='£', alt=True)
        r = 21; ks_row(r, "Largest single-period charge", float(np.max(pos_diffs)), "", fmt='£', bold=True)
        r = 22; ks_row(r, "Smallest single-period charge", float(np.min(pos_diffs)), "", fmt='£', alt=True)
        # Annualised estimate from most recent 6 positive diffs
        recent6 = raw_diffs[-6:]
        recent6_pos = recent6[recent6 > 0]
        if len(recent6_pos):
            ann = float(np.mean(recent6_pos)) * 12
            r = 23; ks_row(r, "Implied annual rate (latest 6 periods)", ann, "Mean recent positive charge × 12", fmt='£', bold=True)

    r = 24; _section_hdr(ws_ks, r, "READING & DATA QUALITY")
    if 'Reading' in dfc.columns:
        est_n  = int((dfc['Reading']=='Estimated').sum())
        act_n  = int((dfc['Reading']=='Actual').sum())
        smt_n  = int((dfc['Reading']=='Smart').sum())
        unk_n  = n - est_n - act_n - smt_n
        est_pct= est_n / n
        r = 25; ks_row(r, "Estimated readings", est_n, f"{est_pct*100:.0f}% of bills — EDF may have charged without actual read",
                       fmt="#,##0", bold=(est_pct>0.4), alt=True)
        r = 26; ks_row(r, "Actual / customer readings", act_n, fmt="#,##0")
        r = 27; ks_row(r, "Smart meter readings", smt_n, fmt="#,##0", alt=True)
        r = 28; ks_row(r, "Unknown / not extracted", unk_n, fmt="#,##0")

    ws_ks.freeze_panes = "A2"

    # ==========================================================================
    # TAB B: BALANCE TREND  (data + line chart)
    # ==========================================================================
    ws_bt = wb.create_sheet(title="Balance Trend")

    bt_headers = ["Date", "Balance (£)", "6-Bill Rolling Avg (£)", "Linear Trend (£)", "Period Charge (£)"]
    for ci, h in enumerate(bt_headers, 1):
        _hcell(ws_bt, 1, ci, h, bg=NAVY)
    ws_bt.row_dimensions[1].height = 22

    period_charges = [None] + list(raw_diffs)   # first row has no prior

    for i in range(n):
        r  = i + 2
        bg = LGREY if i % 2 == 0 else None
        _text(ws_bt,  r, 1, dates_lbl[i], fill_hex=bg)
        _money(ws_bt, r, 2, float(amounts[i]), fill_hex=bg)
        rv = rolling6[i]
        if rv is not None and not (isinstance(rv, float) and np.isnan(rv)):
            _money(ws_bt, r, 3, round(float(rv), 2), fill_hex=bg)
        else:
            ws_bt.cell(row=r, column=3).fill = PatternFill("solid", start_color=bg) if bg else PatternFill()
        _money(ws_bt, r, 4, round(float(trend_line[i]), 2), fill_hex=bg)
        pc = period_charges[i]
        if pc is not None:
            c = _money(ws_bt, r, 5, float(pc), fill_hex=bg)
            if float(pc) > float(np.mean(pos_diffs)) * 2 if len(pos_diffs) else False:
                c.fill = PatternFill("solid", start_color=AMBER)

    # Line chart
    lc = LineChart()
    lc.title         = "Account Balance Over Time"
    lc.style         = 10
    lc.y_axis.title  = "Balance (£)"
    lc.x_axis.title  = "Bill Date"
    lc.width, lc.height = 30, 18

    data_ref  = Reference(ws_bt, min_col=2, max_col=4, min_row=1, max_row=n+1)
    dates_ref = Reference(ws_bt, min_col=1, min_row=2, max_row=n+1)
    lc.add_data(data_ref, titles_from_data=True)
    lc.set_categories(dates_ref)

    # Series styling: balance=orange, rolling=navy dashed, trend=grey dotted
    lc.series[0].graphicalProperties.line.solidFill = ORANGE
    lc.series[0].graphicalProperties.line.width     = 22000
    if len(lc.series) > 1:
        lc.series[1].graphicalProperties.line.solidFill = NAVY
        lc.series[1].graphicalProperties.line.width     = 15000
        lc.series[1].graphicalProperties.line.dashDot   = "dash"
    if len(lc.series) > 2:
        lc.series[2].graphicalProperties.line.solidFill = DGREY
        lc.series[2].graphicalProperties.line.width     = 10000
        lc.series[2].graphicalProperties.line.dashDot   = "sysDash"

    ws_bt.add_chart(lc, "G2")
    for col, w in zip(['A','B','C','D','E'], [14, 16, 20, 16, 16]):
        ws_bt.column_dimensions[col].width = w
    ws_bt.freeze_panes = "A2"

    # ==========================================================================
    # TAB C: YEAR-ON-YEAR ANALYSIS  (table + bar chart)
    # ==========================================================================
    ws_yoy = wb.create_sheet(title="Year-on-Year")

    yoy_hdrs = ["Year", "Bills", "Peak Balance (£)", "Avg Balance (£)",
                "Lowest Balance (£)", "YoY Avg Δ (£)", "YoY Avg Δ (%)",
                "Est. Readings", "Biggest Jump (£)"]
    for ci, h in enumerate(yoy_hdrs, 1):
        _hcell(ws_yoy, 1, ci, h, bg=ORANGE)
    ws_yoy.row_dimensions[1].height = 22

    prev_avg = None
    yoy_data = []   # for chart

    for r_off, row_y in enumerate(yearly.itertuples(), 2):
        yr   = row_y.year
        cnt  = row_y.count
        pk   = row_y.peak
        av   = row_y.avg_bal
        lo   = row_y.low
        yoy_chg_abs = av - prev_avg if prev_avg else None
        yoy_chg_pct = (yoy_chg_abs / prev_avg) if (prev_avg and prev_avg > 0) else None

        # Biggest single-period jump within year
        yr_rows = dfc[dfc['year']==yr]
        yr_idx  = yr_rows.index.tolist()
        max_jump = None
        for ii in yr_idx:
            if ii > 0 and ii in dfc.index and ii-1 in dfc.index:
                jmp = dfc.at[ii,'Amount (£)'] - dfc.at[ii-1,'Amount (£)']
                if max_jump is None or jmp > max_jump:
                    max_jump = jmp

        alt = (r_off % 2 == 0)
        bg  = LGREY if alt else None

        _num(ws_yoy,   r_off, 1, yr,   fmt="#,##0", fill_hex=bg, bold=True)
        _num(ws_yoy,   r_off, 2, cnt,  fmt="#,##0", fill_hex=bg)
        _money(ws_yoy, r_off, 3, pk,   fill_hex=bg, bold=True)
        _money(ws_yoy, r_off, 4, av,   fill_hex=bg)
        _money(ws_yoy, r_off, 5, lo,   fill_hex=bg)

        if yoy_chg_abs is not None:
            _money(ws_yoy, r_off, 6, yoy_chg_abs, fill_hex=bg, bold=(abs(yoy_chg_abs)>5000))
        else:
            ws_yoy.cell(row=r_off, column=6, value="—")

        if yoy_chg_pct is not None:
            c7 = _num(ws_yoy, r_off, 7, yoy_chg_pct, fmt="+0.0%;-0.0%;—", bold=True,
                      fill_hex=(RED if yoy_chg_pct > 0.5 else (AMBER if yoy_chg_pct > 0.2 else (GREEN if yoy_chg_pct < -0.1 else bg))))
        else:
            ws_yoy.cell(row=r_off, column=7, value="—")

        yr_est = int((dfc[dfc['year']==yr]['Reading']=='Estimated').sum()) if 'Reading' in dfc.columns else 0
        _num(ws_yoy, r_off, 8, yr_est, fmt="#,##0", fill_hex=bg)

        if max_jump is not None:
            _money(ws_yoy, r_off, 9, max_jump, fill_hex=(RED if max_jump > 5000 else bg))

        yoy_data.append((yr, av))
        prev_avg = av

    # Bar chart — avg balance per year
    bc = BarChart()
    bc.type   = "col"
    bc.title  = "Average Balance by Year"
    bc.y_axis.title = "Average Balance (£)"
    bc.x_axis.title = "Year"
    bc.style  = 10
    bc.width, bc.height = 22, 14
    n_yrs = len(yoy_data)
    avg_ref  = Reference(ws_yoy, min_col=4, min_row=1, max_row=n_yrs+1)
    yr_ref   = Reference(ws_yoy, min_col=1, min_row=2, max_row=n_yrs+1)
    bc.add_data(avg_ref, titles_from_data=True)
    bc.set_categories(yr_ref)
    bc.series[0].graphicalProperties.solidFill = ORANGE
    ws_yoy.add_chart(bc, "K2")

    for col, w in zip(['A','B','C','D','E','F','G','H','I'], [8,8,18,18,18,16,14,14,18]):
        ws_yoy.column_dimensions[col].width = w
    ws_yoy.freeze_panes = "A2"

    # ==========================================================================
    # TAB D: PERIOD-BY-PERIOD CHARGES  (the real money: what EDF charged each period)
    # ==========================================================================
    ws_pc = wb.create_sheet(title="Period Charges")

    pc_hdrs = ["From Date", "To Date", "Days", "Opening Balance (£)",
               "Closing Balance (£)", "Charge (£)", "Daily Rate (£/day)", "Flag"]
    for ci, h in enumerate(pc_hdrs, 1):
        _hcell(ws_pc, 1, ci, h, bg=NAVY)
    ws_pc.row_dimensions[1].height = 22

    mean_daily = float(np.mean(pos_diffs)) / 30.0 if len(pos_diffs) else 0
    pc_rows_data = []    # (date_label, charge) for chart

    pc_r = 2
    for i in range(1, n):
        p  = dfc.iloc[i-1]
        c_ = dfc.iloc[i]
        days   = (c_['_dt'] - p['_dt']).days
        charge = float(c_['Amount (£)']) - float(p['Amount (£)'])
        daily  = charge / days if days > 0 else None

        flag = ""
        if days > 90:
            flag = f"⚠ {days}-day gap — possible missed bill(s)"
        elif charge < 0:
            flag = f"↓ Balance reduced by £{abs(charge):,.2f} (payment or credit)"
        elif daily and mean_daily > 0 and daily > mean_daily * 2.5:
            flag = f"⚠ Daily rate {daily/mean_daily:.1f}× average"

        bg = LGREY if pc_r % 2 == 0 else None
        if flag.startswith("⚠"):
            bg = AMBER
        elif charge < 0:
            bg = GREEN

        _text(ws_pc,  pc_r, 1, p['Date'],               fill_hex=bg)
        _text(ws_pc,  pc_r, 2, c_['Date'],               fill_hex=bg)
        _num(ws_pc,   pc_r, 3, days,       fmt="#,##0",  fill_hex=bg)
        _money(ws_pc, pc_r, 4, float(p['Amount (£)']),   fill_hex=bg)
        _money(ws_pc, pc_r, 5, float(c_['Amount (£)']),  fill_hex=bg)
        chg_cell = _money(ws_pc, pc_r, 6, charge,        fill_hex=bg, bold=(charge > float(np.mean(pos_diffs))*1.5 if len(pos_diffs) else False))
        if daily is not None:
            _money(ws_pc, pc_r, 7, daily, fill_hex=bg)
        _text(ws_pc,  pc_r, 8, flag, fill_hex=bg, wrap=True, color=(DGREY if not flag else "000000"))

        if charge > 0:
            pc_rows_data.append((c_['Date'], charge))
        pc_r += 1

    # Summary stats below the table
    if len(pos_diffs):
        gap = 2
        sr  = pc_r + gap
        _section_hdr(ws_pc, sr, "SUMMARY STATISTICS", ncols=8, bg=ORANGE)
        sr += 1
        def pc_stat(r, lbl, val, fmt='£'):
            _text(ws_pc, r, 1, lbl, bold=True, fill_hex=LGREY)
            if fmt == '£':
                _money(ws_pc, r, 2, val, fill_hex=LGREY, bold=True)
            else:
                _num(ws_pc, r, 2, val, fmt=fmt, fill_hex=LGREY, bold=True)
            for cc in range(3,9):
                ws_pc.cell(row=r, column=cc).fill = PatternFill("solid", start_color=LGREY)
                ws_pc.cell(row=r, column=cc).border = CELL_BORDER

        pc_stat(sr,   "Mean charge per period",    float(np.mean(pos_diffs)))
        pc_stat(sr+1, "Median charge per period",  float(np.median(pos_diffs)))
        pc_stat(sr+2, "Largest single charge",     float(np.max(pos_diffs)))
        pc_stat(sr+3, "Smallest single charge",    float(np.min(pos_diffs)))
        avg_days = float(np.mean([(dfc.iloc[i]['_dt']-dfc.iloc[i-1]['_dt']).days for i in range(1,n)]))
        pc_stat(sr+4, "Average days between bills", avg_days, fmt="#,##0.0")

    # Bar chart — charge per period
    if len(pc_rows_data) > 1:
        bc2 = BarChart()
        bc2.type  = "col"
        bc2.title = "Charge Added Each Period"
        bc2.y_axis.title = "Charge (£)"
        bc2.style = 10
        bc2.width, bc2.height = 28, 14
        chg_ref2  = Reference(ws_pc, min_col=6, min_row=1, max_row=pc_r-1)
        date_ref2 = Reference(ws_pc, min_col=2, min_row=2, max_row=pc_r-1)
        bc2.add_data(chg_ref2, titles_from_data=True)
        bc2.set_categories(date_ref2)
        bc2.series[0].graphicalProperties.solidFill = NAVY
        ws_pc.add_chart(bc2, "J2")

    for col, w in zip(['A','B','C','D','E','F','G','H'], [13,13,7,18,18,16,14,42]):
        ws_pc.column_dimensions[col].width = w
    ws_pc.freeze_panes = "A2"

    # ==========================================================================
    # TAB E: DISPUTE FLAGS
    # ==========================================================================
    ws_df = wb.create_sheet(title="Dispute Flags")

    # Title block
    def _banner(ws, r, text, bg):
        c = ws.cell(row=r, column=1, value=text)
        c.font  = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        c.fill  = PatternFill("solid", start_color=bg)
        c.border= CELL_BORDER
        c.alignment = Alignment(horizontal="left", vertical="center")
        for col in range(2, 7):
            x = ws.cell(row=r, column=col)
            x.fill   = PatternFill("solid", start_color=bg)
            x.border = CELL_BORDER
        ws.row_dimensions[r].height = 20

    _banner(ws_df, 1, "EDF ENERGY DISPUTE  —  AUTOMATED ANALYSIS FLAGS", ORANGE)
    ws_df.cell(row=2, column=1, value=f"Generated {datetime.now().strftime('%d/%m/%Y %H:%M')}  |  {n} bills analysed  |  Period: {dates_lbl[0]} → {dates_lbl[-1]}")
    ws_df.cell(row=2, column=1).font = Font(name="Calibri", size=9, italic=True, color=DGREY)

    legend_row = 3
    legend_items = [("■ RED = HIGH severity", RED), ("■ AMBER = MEDIUM severity", AMBER), ("■ GREEN = Payment / credit", GREEN)]
    for ci, (txt, col_hex) in enumerate(legend_items, 1):
        lc2 = ws_df.cell(row=legend_row, column=ci*2-1, value=txt)
        lc2.font  = Font(name="Calibri", size=9, bold=True)
        lc2.fill  = PatternFill("solid", start_color=col_hex)
        lc2.border= CELL_BORDER

    hdr_row = 5
    for ci, h in enumerate(["#", "Date", "Balance (£)", "Flag Type", "Detail", "Severity"], 1):
        _hcell(ws_df, hdr_row, ci, h, bg=NAVY)

    flags = []

    # F1: Large balance jumps (>25% increase, gap ≤90 days)
    for i in range(1, n):
        p  = dfc.iloc[i-1]
        c_ = dfc.iloc[i]
        chg  = float(c_['Amount (£)']) - float(p['Amount (£)'])
        pct  = chg / float(p['Amount (£)']) if float(p['Amount (£)']) > 0 else 0
        days = (c_['_dt'] - p['_dt']).days
        if pct > 0.25 and 0 < days <= 90:
            flags.append(("LARGE JUMP", c_['Date'], c_['Amount (£)'],
                f"+£{chg:,.2f}  (+{pct*100:.1f}%)  in {days} days  (from {p['Date']}: £{p['Amount (£)']:,.2f})",
                "HIGH" if pct > 0.5 else "MEDIUM"))

    # F2: Billing gaps >60 days
    for i in range(1, n):
        p  = dfc.iloc[i-1]
        c_ = dfc.iloc[i]
        days = (c_['_dt'] - p['_dt']).days
        if days > 60:
            flags.append(("BILLING GAP", c_['Date'], c_['Amount (£)'],
                f"{days} days without a bill  (previous: {p['Date']}).  Balance charge accumulated unchecked.",
                "HIGH" if days > 120 else "MEDIUM"))

    # F3: Consecutive estimated readings (run of ≥3)
    if 'Reading' in dfc.columns:
        run = 0; run_start = None
        for i, rv in enumerate(dfc['Reading'].tolist()):
            if str(rv).lower() in ('estimated', 'est.'):
                run += 1
                if run == 1:
                    run_start = dfc.iloc[i]['Date']
            else:
                if run >= 3:
                    flags.append(("ESTIMATED RUN", run_start, None,
                        f"{run} consecutive estimated readings from {run_start}.  No actual meter read taken for entire period.",
                        "HIGH"))
                run = 0; run_start = None
        if run >= 3:
            flags.append(("ESTIMATED RUN", run_start, None,
                f"{run} consecutive estimated readings from {run_start}  (still ongoing at latest bill).",
                "HIGH"))

    # F4: Unusually high daily charge rate (>2.5× mean)
    if mean_daily > 0:
        for i in range(1, n):
            p  = dfc.iloc[i-1]
            c_ = dfc.iloc[i]
            days   = (c_['_dt'] - p['_dt']).days
            charge = float(c_['Amount (£)']) - float(p['Amount (£)'])
            if days > 0 and charge > 0:
                daily = charge / days
                ratio = daily / mean_daily
                if ratio > 2.5:
                    flags.append(("HIGH DAILY RATE", c_['Date'], c_['Amount (£)'],
                        f"£{daily:,.2f}/day  ({ratio:.1f}× average of £{mean_daily:,.2f}/day)  over {days} days",
                        "HIGH" if ratio > 4 else "MEDIUM"))

    # F5: Balance reduction (payment or credit — useful to note)
    for i in range(1, n):
        p  = dfc.iloc[i-1]
        c_ = dfc.iloc[i]
        chg = float(c_['Amount (£)']) - float(p['Amount (£)'])
        if chg < -500:
            flags.append(("BALANCE REDUCTION", c_['Date'], c_['Amount (£)'],
                f"Balance fell by £{abs(chg):,.2f}  (from £{p['Amount (£)']:,.2f} to £{c_['Amount (£)']:,.2f}).  Payment or credit applied.",
                "INFO"))

    sev_fill = {"HIGH": RED, "MEDIUM": AMBER, "INFO": GREEN, "LOW": "CCFFCC"}

    for fi, (ftype, date, amt, detail, sev) in enumerate(flags, hdr_row + 1):
        bg = sev_fill.get(sev, LGREY)
        _num(ws_df,  fi, 1, fi - hdr_row, fmt="#,##0", fill_hex=bg)
        _text(ws_df, fi, 2, date or "—",  fill_hex=bg)
        if amt:
            _money(ws_df, fi, 3, float(amt), fill_hex=bg)
        else:
            ws_df.cell(row=fi, column=3).fill  = PatternFill("solid", start_color=bg)
            ws_df.cell(row=fi, column=3).border= CELL_BORDER
        _text(ws_df, fi, 4, ftype,  bold=True, fill_hex=bg)
        _text(ws_df, fi, 5, detail, fill_hex=bg, wrap=True)
        _text(ws_df, fi, 6, sev,    bold=True,  fill_hex=bg, align="center")
        ws_df.row_dimensions[fi].height = 30

    # Summary footer
    if flags:
        fr = len(flags) + hdr_row + 2
        counts = {s: sum(1 for f in flags if f[4]==s) for s in ("HIGH","MEDIUM","INFO")}
        _banner(ws_df, fr, f"TOTAL FLAGS: {len(flags)}   |   HIGH: {counts['HIGH']}   |   MEDIUM: {counts['MEDIUM']}   |   INFO: {counts['INFO']}", NAVY)

    for col, w in zip(['A','B','C','D','E','F'], [5, 13, 16, 20, 60, 10]):
        ws_df.column_dimensions[col].width = w
    ws_df.freeze_panes = f"A{hdr_row+1}"

    wb.save(output_path)



# ---------------------------------------------------------------------------
# GUI
# ---------------------------------------------------------------------------

class App:
    def __init__(self, root):
        self.root = root
        self.root.title("EDF Master Evidence Collector")
        self.root.geometry("680x760")
        self.root.configure(bg=EDF_OFFWHITE)

        self.pst_path = tk.StringVar()
        self.pdf_dir  = tk.StringVar()
        self.acc_num  = tk.StringVar(value="671078701920")
        self.status   = tk.StringVar(value="Ready.")

        self.use_anchors  = tk.BooleanVar(value=True)
        self.use_large    = tk.BooleanVar(value=True)
        self.use_readings = tk.BooleanVar(value=True)
        self.use_acc_filt = tk.BooleanVar(value=False)
        self.use_dedup    = tk.BooleanVar(value=True)
        self.save_dups    = tk.BooleanVar(value=True)
        self.min_amount   = tk.DoubleVar(value=100.0)

        self.build_ui()

    def build_ui(self):
        hdr = tk.Frame(self.root, bg=EDF_ORANGE, height=60)
        hdr.pack(fill=tk.X)
        tk.Label(hdr, text="EDF BILLING EVIDENCE COLLECTOR",
                 bg=EDF_ORANGE, fg="white",
                 font=("Calibri", 14, "bold")).pack(pady=15)

        main = ttk.Frame(self.root, padding=20)
        main.pack(fill=tk.BOTH, expand=True)

        # Section 1 — Source files
        s1 = ttk.LabelFrame(main, text=" 1. Source Data ", padding=10)
        s1.pack(fill=tk.X, pady=5)

        r1 = ttk.Frame(s1); r1.pack(fill=tk.X, pady=2)
        ttk.Label(r1, text="PST Export:", width=12).pack(side=tk.LEFT)
        ttk.Entry(r1, textvariable=self.pst_path).pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        ttk.Button(r1, text="Browse",
                   command=lambda: self.pst_path.set(
                       filedialog.askopenfilename(filetypes=[("PST", "*.pst")])
                   )).pack(side=tk.LEFT)

        r2 = ttk.Frame(s1); r2.pack(fill=tk.X, pady=2)
        ttk.Label(r2, text="PDF Folder:", width=12).pack(side=tk.LEFT)
        ttk.Entry(r2, textvariable=self.pdf_dir).pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        ttk.Button(r2, text="Browse",
                   command=lambda: self.pdf_dir.set(filedialog.askdirectory())
                   ).pack(side=tk.LEFT)

        # Section 2 — Search options
        s2 = ttk.LabelFrame(main, text=" 2. Search & Filter Options ", padding=10)
        s2.pack(fill=tk.X, pady=5)

        tk.Checkbutton(s2,
            text="Smart Context Search  (looks for 'Balance', 'Debit', etc.)",
            variable=self.use_anchors, bg=EDF_OFFWHITE).pack(anchor=tk.W)
        tk.Checkbutton(s2,
            text="Large Number Fallback  (catch any large £ amount)",
            variable=self.use_large, bg=EDF_OFFWHITE).pack(anchor=tk.W)
        tk.Checkbutton(s2,
            text="Deep PDF Mine  (extract kWh, standing charges & invoice #)",
            variable=self.use_readings, bg=EDF_OFFWHITE).pack(anchor=tk.W)

        r3 = ttk.Frame(s2); r3.pack(fill=tk.X, pady=5)
        tk.Checkbutton(r3, text="Filter by Account #:",
                       variable=self.use_acc_filt, bg=EDF_OFFWHITE).pack(side=tk.LEFT)
        ttk.Entry(r3, textvariable=self.acc_num, width=15).pack(side=tk.LEFT, padx=5)

        r4 = ttk.Frame(s2); r4.pack(fill=tk.X)
        ttk.Label(r4, text="Minimum £ for 'Large Number' rule:").pack(side=tk.LEFT)
        ttk.Entry(r4, textvariable=self.min_amount, width=8).pack(side=tk.LEFT, padx=5)

        # Section 3 — Deduplication
        s3 = ttk.LabelFrame(main, text=" 3. Deduplication ", padding=10)
        s3.pack(fill=tk.X, pady=5)

        chk_dup = tk.Checkbutton(s3,
            text="Filter duplicate records  (same date & amount)",
            variable=self.use_dedup, bg=EDF_OFFWHITE)
        chk_dup.pack(anchor=tk.W)

        chk_save_dup = tk.Checkbutton(s3,
            text="Save duplicates to a separate worksheet for review",
            variable=self.save_dups, bg=EDF_OFFWHITE)
        chk_save_dup.pack(anchor=tk.W, padx=20)

        def toggle_dup_save():
            chk_save_dup.config(state="normal" if self.use_dedup.get() else "disabled")
        chk_dup.config(command=toggle_dup_save)

        # Progress & status
        self.pb = ttk.Progressbar(main, mode='indeterminate')
        self.pb.pack(fill=tk.X, pady=15)
        ttk.Label(main, textvariable=self.status,
                  foreground=EDF_NAVY,
                  font=("Calibri", 11, "bold")).pack()

        self.run_btn = tk.Button(
            main, text="EXTRACT TO EXCEL",
            bg=EDF_ORANGE, fg="white",
            font=("Calibri", 12, "bold"),
            command=self.start_thread,
            relief="flat"
        )
        self.run_btn.pack(fill=tk.X, pady=10, ipady=8)

    def set_status(self, text):
        self.status.set(text)
        self.root.update_idletasks()

    def start_thread(self):
        if not self.pst_path.get() and not self.pdf_dir.get():
            messagebox.showerror("Error", "Please select a PST file or PDF folder.")
            return
        self.run_btn.config(state="disabled")
        self.pb.start()
        threading.Thread(target=self.run_process, daemon=True).start()

    def run_process(self):
        config = {
            "use_anchors":    self.use_anchors.get(),
            "use_large":      self.use_large.get(),
            "use_readings":   self.use_readings.get(),
            "use_acc_filter": self.use_acc_filt.get(),
            "acc_num":        self.acc_num.get(),
            "min_amount":     self.min_amount.get(),
            "use_dedup":      self.use_dedup.get(),
            "save_dups":      self.save_dups.get()
        }

        engine = EvidenceEngine(config, self.set_status)

        try:
            pst_path = self.pst_path.get().strip()
            if pst_path and os.path.exists(pst_path):
                clean_path = os.path.abspath(os.path.normpath(pst_path))
                pst = pypff.file()
                pst.open(clean_path)
                engine.crawl_pst(pst.get_root_folder())
                pst.close()

            pdf_path = self.pdf_dir.get().strip()
            if pdf_path and os.path.exists(pdf_path):
                engine.crawl_local_pdfs(pdf_path)

            if engine.records:
                self.set_status("Writing Excel report…")
                save_dir = os.path.dirname(pst_path) if pst_path else pdf_path
                out_path = os.path.join(save_dir, "EDF_Dispute_Evidence.xlsx")
                export_to_excel(engine.records, out_path, engine.error_log, config)

                summary = (
                    f"Extraction complete.\n\n"
                    f"  Emails matched:   {engine.email_count}\n"
                    f"  PDFs processed:   {engine.pdf_count}\n"
                    f"  Records found:    {len(engine.records)}\n"
                )
                if engine.error_log:
                    summary += f"\n  Parse errors:     {len(engine.error_log)}  (see 'Parse Errors' tab)"
                summary += f"\n\nSaved to:\n{out_path}"
                messagebox.showinfo("Success", summary)
            else:
                messagebox.showwarning("No Data",
                    "No billing amounts found.\n\nTry unchecking the Account Filter.")

        except Exception:
            messagebox.showerror("System Error",
                f"An error occurred:\n\n{traceback.format_exc()}\n\n"
                "Ensure Outlook is closed and paths are correct.")
        finally:
            self.pb.stop()
            self.run_btn.config(state="normal")
            self.set_status("Ready.")
            gc.collect()


if __name__ == "__main__":
    root = tk.Tk()
    App(root)
    root.mainloop()