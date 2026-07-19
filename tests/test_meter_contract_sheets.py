from __future__ import annotations

import pandas as pd
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from edf_collector import (
    detect_meter_rollover,
    write_contract_history_sheet,
    write_meter_readings_sheet,
)


def _evidence_df() -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "Date": "01 Jan 2023",
                "Invoice #": "INV-001",
                "Reading": "Actual",
                "Units (kWh)": 300.0,
                "Details": "Customer reading",
                "Tariff": "Standard",
                "Cancel/Rebill Admitted": False,
                "Attachment Name": "INV-001.pdf",
                "Source PDF Text": "Invoice number: INV-001 Period 01 Jan 2023 - 31 Jan 2023",
                "Period From": "01 Jan 2023",
                "Period To": "31 Jan 2023",
            },
            {
                "Date": "01 Feb 2023",
                "Invoice #": "INV-002",
                "Reading": "Estimated",
                "Units (kWh)": 350.0,
                "Details": "Automatic estimate",
                "Tariff": "Standard",
                "Cancel/Rebill Admitted": False,
                "Attachment Name": "INV-002.pdf",
                "Source PDF Text": "Invoice number: INV-002 Period 01 Feb 2023 - 28 Feb 2023",
                "Period From": "01 Feb 2023",
                "Period To": "28 Feb 2023",
            },
            {
                "Date": "01 Mar 2023",
                "Invoice #": "INV-003",
                "Reading": "Actual",
                "Units (kWh)": -200000.0,
                "Details": "Reading was actual",
                "Tariff": "Standard",
                "Cancel/Rebill Admitted": True,
                "Attachment Name": "INV-003.pdf",
                "Source PDF Text": "Invoice number: INV-003 Period 01 Mar 2023 - 31 Mar 2023",
                "Period From": "01 Mar 2023",
                "Period To": "31 Mar 2023",
            },
        ]
    )


# Evidence-index fixture for the View-on-Evidence-Report hotlink.
def _evidence_index() -> dict[str, int]:
    return {
        "inv:INV-001": 2,
        "inv:INV-002": 3,
        "inv:INV-003": 4,
    }


def _open_ws(title: str = "Meter Readings") -> Worksheet:
    wb = Workbook()
    ws = wb.active
    ws.title = title
    return ws


# ---------------------------------------------------------------------------
# Meter Readings sheet (8 cols: 6 originals + Source Excerpt + View on ER)
# ---------------------------------------------------------------------------


def test_write_meter_readings_sheet_renders_title_with_account() -> None:
    ws = _open_ws()
    rollovers = detect_meter_rollover(_evidence_df())
    write_meter_readings_sheet(ws, _evidence_df(), rollovers, account="ACC1")
    a1 = ws.cell(row=1, column=1).value
    assert isinstance(a1, str)
    assert "METER READING" in a1.upper()
    assert "ACC1" in a1


def test_write_meter_readings_sheet_renders_legend_subheader() -> None:
    ws = _open_ws()
    rollovers = detect_meter_rollover(_evidence_df())
    write_meter_readings_sheet(ws, _evidence_df(), rollovers, account="ACC1")
    # Row 2 should mention A/E/M legend.
    a2 = ws.cell(row=2, column=1).value
    assert isinstance(a2, str)
    low = a2.lower()
    assert "actual" in low
    assert "estimated" in low
    assert "rollover" in low


def test_write_meter_readings_sheet_writes_eight_table_headers_including_excerpt_and_hotlink() -> (
    None
):
    """Spec \u00a75.2 wants Source Excerpt on the four analyser tabs.
    Spec \u00a710.2 wants hotlinks from each analyser-tab row back to
    the Evidence Report."""
    ws = _open_ws()
    rollovers = detect_meter_rollover(_evidence_df())
    write_meter_readings_sheet(
        ws, _evidence_df(), rollovers, account="ACC1", evidence_index=_evidence_index()
    )
    headers = [ws.cell(row=7, column=c).value for c in range(1, 9)]
    expected = [
        "Date",
        "Reading (kWh)",
        "Type (A/E/M)",
        "Estimated Source",
        "Invoice #",
        "Notes",
        "Source Excerpt",
        "View on Evidence Report",
    ]
    assert headers == expected


def test_write_meter_readings_sheet_one_row_per_evidence() -> None:
    ws = _open_ws()
    rollovers = detect_meter_rollover(_evidence_df())
    write_meter_readings_sheet(
        ws, _evidence_df(), rollovers, account="ACC1", evidence_index=_evidence_index()
    )
    # Sample has 3 evidence rows. Row 8 has row 1, etc.
    a8 = ws.cell(row=8, column=5).value
    assert a8 == "INV-001"
    a9 = ws.cell(row=9, column=5).value
    assert a9 == "INV-002"
    a10 = ws.cell(row=10, column=5).value
    assert a10 == "INV-003"
    # No row 11.
    assert ws.cell(row=11, column=1).value in (None, "")


def test_write_meter_readings_sheet_actual_row_type_is_a() -> None:
    ws = _open_ws()
    rollovers = detect_meter_rollover(_evidence_df())
    write_meter_readings_sheet(
        ws, _evidence_df(), rollovers, account="ACC1", evidence_index=_evidence_index()
    )
    assert ws.cell(row=8, column=3).value == "A"
    assert ws.cell(row=9, column=3).value == "E"


def test_write_meter_readings_sheet_rollover_row_marked_m() -> None:
    ws = _open_ws()
    rollovers = detect_meter_rollover(_evidence_df())
    write_meter_readings_sheet(
        ws, _evidence_df(), rollovers, account="ACC1", evidence_index=_evidence_index()
    )
    # INV-003 is in the rollover set.
    type_v = ws.cell(row=10, column=3).value
    assert type_v == "M"


def test_write_meter_readings_sheet_estimated_source_from_details() -> None:
    ws = _open_ws()
    rollovers = detect_meter_rollover(_evidence_df())
    write_meter_readings_sheet(
        ws, _evidence_df(), rollovers, account="ACC1", evidence_index=_evidence_index()
    )
    # INV-002 is Estimated with Details='Automatic estimate'
    est_src = ws.cell(row=9, column=4).value
    assert isinstance(est_src, str)
    assert "automatic" in est_src.lower()


def test_write_meter_readings_sheet_source_excerpt_column_populated() -> None:
    """When ``evidence_df`` is provided, the Source Excerpt column
    (col 7) carries the truncated source PDF text + regex trace for
    the row's Invoice # so reviewers can diagnose N/A entries."""
    ws = _open_ws()
    rollovers = detect_meter_rollover(_evidence_df())
    write_meter_readings_sheet(
        ws,
        _evidence_df(),
        rollovers,
        account="ACC1",
        evidence_df=_evidence_df(),
        evidence_index=_evidence_index(),
    )
    excerpt = ws.cell(row=8, column=7).value
    assert isinstance(excerpt, str)
    assert "INV-001" in excerpt  # source text contains the invoice number


def test_write_meter_readings_sheet_view_on_evidence_report_hyperlink_emitted() -> None:
    """Col 8 carries a '->' hyperlink that jumps to the matched row on
    EDF Evidence Report."""
    ws = _open_ws()
    rollovers = detect_meter_rollover(_evidence_df())
    write_meter_readings_sheet(
        ws,
        _evidence_df(),
        rollovers,
        account="ACC1",
        evidence_index=_evidence_index(),
    )
    cell = ws.cell(row=8, column=8)
    assert cell.value == "\u2192"  # right-arrow
    assert cell.hyperlink is not None
    assert "EDF Evidence Report" in str(cell.hyperlink.location or "")
    assert "A2" in str(cell.hyperlink.location or "")  # INV-001 -> row 2


def test_write_meter_readings_sheet_empty_evidence_renders_headers() -> None:
    ws = _open_ws()
    empty = pd.DataFrame()
    write_meter_readings_sheet(ws, empty, pd.DataFrame(), account="ACC1")
    a1 = ws.cell(row=1, column=1).value
    assert isinstance(a1, str)
    assert "METER READING" in a1.upper()
    headers = [ws.cell(row=7, column=c).value for c in range(1, 9)]
    assert headers[0] == "Date"
    assert ws.cell(row=8, column=1).value in (None, "")


# ---------------------------------------------------------------------------
# Contract History sheet (7 cols: 5 originals + Source Excerpt + View on ER)
# ---------------------------------------------------------------------------


def test_write_contract_history_sheet_renders_title_and_headers() -> None:
    ws = _open_ws()
    contracts = pd.DataFrame(
        [
            {
                "Contract From": "01 Jan 2022",
                "Contract To": "31 Dec 2022",
                "Tariff": "Standard",
                "Days": 365,
                "# Invoices": 12,
            }
        ]
    )
    write_contract_history_sheet(ws, contracts, account="ACC1")
    a1 = ws.cell(row=1, column=1).value
    assert isinstance(a1, str)
    assert "CONTRACT" in a1.upper()
    assert "ACC1" in a1
    # Headers at row 7 for layout consistency. Spec \u00a75.2 + \u00a710.2
    # require Source Excerpt + View on Evidence Report columns.
    headers = [ws.cell(row=7, column=c).value for c in range(1, 8)]
    expected = [
        "Contract From",
        "Contract To",
        "Tariff",
        "Days",
        "# Invoices",
        "Source Excerpt",
        "View on Evidence Report",
    ]
    assert headers == expected


def test_write_contract_history_sheet_one_row_per_contract() -> None:
    ws = _open_ws()
    contracts = pd.DataFrame(
        [
            {
                "Contract From": "01 Jan 2022",
                "Contract To": "30 Jun 2022",
                "Tariff": "Standard",
                "Days": 181,
                "# Invoices": 6,
            },
            {
                "Contract From": "01 Jul 2022",
                "Contract To": "31 Dec 2022",
                "Tariff": "Fixed",
                "Days": 184,
                "# Invoices": 6,
            },
        ]
    )
    write_contract_history_sheet(ws, contracts, account="ACC1")
    # Row 8 = first contract, row 9 = second
    a8 = ws.cell(row=8, column=3).value
    assert a8 == "Standard"
    a9 = ws.cell(row=9, column=3).value
    assert a9 == "Fixed"
    assert ws.cell(row=10, column=1).value in (None, "")


def test_write_contract_history_sheet_empty_renders_headers() -> None:
    ws = _open_ws()
    empty = pd.DataFrame(columns=["Contract From", "Contract To", "Tariff", "Days", "# Invoices"])
    write_contract_history_sheet(ws, empty, account="ACC1")
    a1 = ws.cell(row=1, column=1).value
    assert isinstance(a1, str)
    assert "CONTRACT" in a1.upper()
    # 7 cols
    headers = [ws.cell(row=7, column=c).value for c in range(1, 8)]
    assert headers[0] == "Contract From"
    # Source Excerpt + hotlink headers present even when no rows
    assert "Source Excerpt" in headers
    assert "View on Evidence Report" in headers
    assert ws.cell(row=8, column=1).value in (None, "")


def test_write_contract_history_sheet_view_on_evidence_report_hyperlink() -> None:
    """The contract-inferred analyser matches each inferred contract
    against any invoice in the evidence dataframe whose Period falls
    inside that contract. The hotlink should pick up the first
    matching invoice's row on the Evidence Report when available."""
    ws = _open_ws()
    contracts = pd.DataFrame(
        [
            {
                "Contract From": "01 Jan 2023",
                "Contract To": "31 Jan 2023",
                "Tariff": "Standard",
                "Days": 31,
                "# Invoices": 1,
            }
        ]
    )
    # evidence_index maps an invoice from this contract's window.
    evidence_index = {"inv:INV-001": 2}
    write_contract_history_sheet(
        ws,
        contracts,
        account="ACC1",
        evidence_index=evidence_index,
        evidence_df=_evidence_df(),
    )
    cell = ws.cell(row=8, column=7)
    assert cell.value == "\u2192"
    assert cell.hyperlink is not None
    assert "A2" in str(cell.hyperlink.location or "")
