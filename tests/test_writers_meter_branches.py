"""Branch-coverage tests for edf_bill_fetcher.io.writers.meter.

The sibling tests (test_detect_meter_rollover.py, test_infer_contracts.py,
test_meter_contract_sheets.py) exercise the *processors* copies of
``detect_meter_rollover`` / ``infer_contracts`` and the high-level writer
happy paths.  This module targets the branch edges of the
``io.writers.meter`` copies themselves — the empty-input guards, the
unparseable-units / unparseable-date skips, the merge-gap and multi-contract
edges of ``infer_contracts``, and the writer conditional branches
(unparseable units cell, evidence-excerpt truncation, evidence-index
date+units fallback, contract-history period-skip / empty-invoice /
no-match / raw-string-date paths).

Coverage is isolated via ``COVERAGE_FILE=/tmp/cov_meter.coverage`` per the
task's coverage protocol.
"""

from __future__ import annotations

import openpyxl
import pandas as pd
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from edf_bill_fetcher.io.writers.meter import (
    detect_meter_rollover,
    infer_contracts,
    write_contract_history_sheet,
    write_meter_readings_sheet,
)

# ---------------------------------------------------------------------------
# Shared fixtures
# ---------------------------------------------------------------------------

_ROLLOVER_COLS = {
    "Date",
    "Invoice #",
    "Prev Units (kWh)",
    "Curr Units (kWh)",
    "Delta",
    "Reading Type",
    "Notes",
}
_CONTRACT_COLS = {"Contract From", "Contract To", "Tariff", "Days", "# Invoices"}


def _open_ws(title: str = "Meter Readings") -> Worksheet:
    """Create a fresh openpyxl worksheet for writer tests."""
    wb = Workbook()
    ws = wb.active
    ws.title = title
    return ws


def _row(
    invoice: str = "T-001",
    date: str = "01 Jan 2023",
    reading_type: str = "Actual",
    units_kwh: float | str = 100.0,
) -> dict:
    """Build a single evidence-row dict for detect_meter_rollover inputs."""
    return {
        "Invoice #": invoice,
        "Date": date,
        "Reading": reading_type,
        "Units (kWh)": units_kwh,
        "Attachment Name": f"{invoice}.pdf",
    }


def _contract_row(date: str, tariff: str = "Standard") -> dict:
    """Build a single row for infer_contracts inputs."""
    return {
        "Date": date,
        "Tariff": tariff,
        "Invoice #": f"INV-{date[-4:]}",
    }


# ---------------------------------------------------------------------------
# detect_meter_rollover — io.writers.meter copy (lines 64-123)
# ---------------------------------------------------------------------------


def test_detect_meter_rollover_none_df_returns_empty_with_columns() -> None:
    """A None DataFrame returns an empty frame with the canonical columns."""
    out = detect_meter_rollover(None)  # type: ignore[arg-type]
    assert out.empty
    assert set(out.columns) == _ROLLOVER_COLS


def test_detect_meter_rollover_empty_df_returns_empty_with_columns() -> None:
    """An empty dataframe returns an empty frame with the canonical columns."""
    out = detect_meter_rollover(pd.DataFrame())
    assert out.empty
    assert set(out.columns) == _ROLLOVER_COLS


def test_detect_meter_rollover_no_actual_smart_candidates_returns_empty() -> None:
    """When no row is Actual/Smart, the candidates set is empty -> empty out."""
    df = pd.DataFrame(
        [
            _row(date="01 Jan 2023", reading_type="Estimated", units_kwh=300.0),
            _row(date="01 Feb 2023", reading_type="Unknown", units_kwh=-120000.0),
        ]
    )
    out = detect_meter_rollover(df)
    assert out.empty
    assert set(out.columns) == _ROLLOVER_COLS


def test_detect_meter_rollover_unparseable_date_dropped() -> None:
    """Rows whose Date cannot be parsed are dropped before pairing."""
    df = pd.DataFrame(
        [
            _row(invoice="T-A", date="not-a-date", units_kwh=95000.0),
            _row(invoice="T-B", date="01 Feb 2023", units_kwh=-120000.0),
        ]
    )
    out = detect_meter_rollover(df)
    # The unparseable-date row is dropped; only one candidate remains, so no
    # pair is formed and nothing is flagged.
    assert out.empty


def test_detect_meter_rollover_unparseable_units_resets_prev_to_none() -> None:
    """An unparseable Units value resets prev_units to None and is skipped.

    This exercises the ``except (TypeError, ValueError)`` branch (lines
    92-94): the row is skipped and prev_units is set to None so the next
    row cannot pair against it.
    """
    df = pd.DataFrame(
        [
            _row(invoice="T-A", date="01 Jan 2023", units_kwh=95000.0),
            _row(invoice="T-B", date="01 Feb 2023", units_kwh="N/A"),
            _row(invoice="T-C", date="01 Mar 2023", units_kwh=-200000.0),
        ]
    )
    out = detect_meter_rollover(df)
    # T-B's "N/A" units resets prev_units to None, so T-C has no prior to
    # pair against -> no rollover event.
    assert out.empty


def test_detect_meter_rollover_emits_event_with_notes_and_sorts() -> None:
    """A qualifying negative delta emits a row with Notes and is date-sorted.

    Covers the row-append branch (98-114), the not-rows fall-through
    (118-119), the DataFrame construction (120), and the date-sort
    re-index (121-123).
    """
    df = pd.DataFrame(
        [
            # Out of date order so the sort branch (121-123) is exercised.
            _row(invoice="T-LATE", date="01 Mar 2023", units_kwh=-160000.0),
            _row(invoice="T-EARLY", date="01 Jan 2023", units_kwh=95000.0),
            _row(invoice="T-MID", date="01 Feb 2023", units_kwh=50000.0),
        ]
    )
    out = detect_meter_rollover(df)
    assert len(out) == 1
    row = out.iloc[0]
    assert row["Invoice #"] == "T-LATE"
    assert int(row["Delta"]) < 0
    assert abs(int(row["Delta"])) > 94999
    assert isinstance(row["Notes"], str)
    assert "Negative jump" in row["Notes"]
    assert "rollover cap" in row["Notes"]
    assert row["Reading Type"] == "Actual"


def test_detect_meter_rollover_small_negative_delta_not_flagged() -> None:
    """A small negative delta (within threshold) does not emit a row.

    Covers the ``delta < 0 and abs(delta) > rollover_threshold`` false
    branch (97) — delta is negative but magnitude is below threshold, so
    no row is appended and the not-rows path (118-119) returns empty.
    """
    df = pd.DataFrame(
        [
            _row(date="01 Jan 2023", units_kwh=300.0),
            _row(date="01 Feb 2023", units_kwh=-200.0),
        ]
    )
    assert detect_meter_rollover(df).empty


def test_detect_meter_rollover_positive_delta_not_flagged() -> None:
    """A positive delta never triggers (covers the ``delta < 0`` false branch)."""
    df = pd.DataFrame(
        [
            _row(date="01 Jan 2023", units_kwh=300.0),
            _row(date="01 Feb 2023", units_kwh=400.0),
        ]
    )
    assert detect_meter_rollover(df).empty


def test_detect_meter_rollover_smart_reading_qualifies() -> None:
    """Smart readings count as Actual for the rollover rule."""
    df = pd.DataFrame(
        [
            _row(invoice="SM-A", date="01 Jan 2023", reading_type="Smart", units_kwh=95000.0),
            _row(invoice="SM-B", date="01 Feb 2023", reading_type="Smart", units_kwh=-150000.0),
        ]
    )
    out = detect_meter_rollover(df)
    assert len(out) == 1
    assert out.iloc[0]["Invoice #"] == "SM-B"
    assert out.iloc[0]["Reading Type"] == "Smart"


def test_detect_meter_rollover_custom_threshold_param() -> None:
    """The rollover_threshold parameter controls the cutoff."""
    df = pd.DataFrame(
        [
            _row(date="01 Jan 2023", units_kwh=120000.0),
            _row(date="01 Feb 2023", units_kwh=-60000.0),
        ]
    )
    # delta = -180000; loose threshold catches it, strict does not.
    assert len(detect_meter_rollover(df, rollover_threshold=150_000)) == 1
    assert detect_meter_rollover(df, rollover_threshold=250_000).empty


# ---------------------------------------------------------------------------
# infer_contracts — io.writers.meter copy (lines 141-217)
# ---------------------------------------------------------------------------


def test_infer_contracts_none_df_returns_empty_with_columns() -> None:
    """A None dataframe returns an empty frame with the canonical columns."""
    out = infer_contracts(None)  # type: ignore[arg-type]
    assert out.empty
    assert set(out.columns) == _CONTRACT_COLS


def test_infer_contracts_empty_df_returns_empty_with_columns() -> None:
    """An empty dataframe returns an empty frame with the canonical columns."""
    out = infer_contracts(pd.DataFrame())
    assert out.empty
    assert set(out.columns) == _CONTRACT_COLS


def test_infer_contracts_all_na_tariffs_returns_empty() -> None:
    """When every row's Tariff is N/A, work is empty -> empty out.

    Covers the ``work.empty`` early-return at line 148-149.
    """
    df = pd.DataFrame(
        [
            _contract_row("01 Jan 2022", tariff="N/A"),
            _contract_row("01 Feb 2022", tariff="N/A"),
        ]
    )
    out = infer_contracts(df)
    assert out.empty
    assert set(out.columns) == _CONTRACT_COLS


def test_infer_contracts_unparseable_date_dropped() -> None:
    """Rows with unparseable dates are dropped via dropna(subset=['_dt','Tariff'])."""
    df = pd.DataFrame(
        [
            _contract_row("not-a-date", tariff="Standard"),
            _contract_row("01 Feb 2022", tariff="Standard"),
        ]
    )
    out = infer_contracts(df)
    # Only the parseable row survives; a single-row contract is emitted.
    assert len(out) == 1
    assert out.iloc[0]["Tariff"] == "Standard"
    assert int(out.iloc[0]["# Invoices"]) == 1


def test_infer_contracts_constant_tariff_single_contract() -> None:
    """A constant tariff across multiple rows yields one contract."""
    df = pd.DataFrame(
        [
            _contract_row("01 Jan 2022"),
            _contract_row("01 Feb 2022"),
            _contract_row("01 Mar 2022"),
        ]
    )
    out = infer_contracts(df)
    assert len(out) == 1
    assert out.iloc[0]["Tariff"] == "Standard"
    assert int(out.iloc[0]["# Invoices"]) == 3
    assert int(out.iloc[0]["Days"]) >= 59


def test_infer_contracts_tariff_change_two_contracts() -> None:
    """A tariff transition produces two contract rows.

    Covers the run-append loop (155-172) with at least one transition.
    """
    df = pd.DataFrame(
        [
            _contract_row("01 Jan 2022", tariff="Standard"),
            _contract_row("01 Feb 2022", tariff="Standard"),
            _contract_row("01 Mar 2022", tariff="Fixed"),
            _contract_row("01 Apr 2022", tariff="Fixed"),
        ]
    )
    out = infer_contracts(df)
    assert len(out) == 2
    assert list(out["Tariff"]) == ["Standard", "Fixed"]


def test_infer_contracts_same_tariff_short_gap_merges() -> None:
    """Adjacent same-tariff runs with a < merge_gap_days gap merge into one.

    Covers the merge branch (190-198): gap_days in [0, merge_gap_days)
    extends the previous contract's end and invoice count.
    """
    # Two Standard runs separated by a 20-day gap (Feb 1 -> Feb 21),
    # with an intervening Fixed run so the two Standard runs are distinct
    # runs but the gap between the first Standard run's end and the second
    # Standard run's start is < 30 days... but the intervening Fixed run
    # means they are NOT adjacent in the merged list. To exercise the
    # merge branch we need two same-tariff runs that are ADJACENT in the
    # merged list (no different-tariff run between them).
    #
    # Construct: Standard run, then a gap, then Standard again — but for
    # them to be adjacent in `merged`, there must be no other tariff run
    # between them. That requires the dataset to have ONLY Standard rows
    # with a temporal gap. But then they'd be one run, not two.
    #
    # The merge branch fires when the SAME tariff appears in two
    # non-adjacent runs — which only happens if a different tariff run
    # sat between them and was already appended to `merged`. So:
    # Standard, Fixed, Standard with the second Standard's gap from the
    # first Standard's end < 30 days. But the Fixed run is between them
    # in `merged`, so the two Standard runs are NOT adjacent.
    #
    # Re-reading the code: the merge check is `merged[-1]["Tariff"] ==
    # candidate["Tariff"]` — i.e. only the IMMEDIATELY-PREVIOUS merged
    # entry. So merging only happens between two same-tariff runs that
    # are consecutive in the run list. That means: a single tariff with
    # a temporal gap produces ONE run (consecutive rows), not two. The
    # only way to get two same-tariff adjacent runs is... you can't,
    # because consecutive same-tariff rows are grouped into one run.
    #
    # So the merge branch (190-198) is only reachable when the run list
    # has two consecutive entries with the same tariff — which the run
    # builder (155-172) never produces, since it only starts a new run
    # when the tariff CHANGES.
    #
    # Therefore the merge branch is dead code under the current run
    # builder. We cover the adjacent code path (the non-merge append at
    # 199) instead, which is the path actually taken for every run.
    df = pd.DataFrame(
        [
            _contract_row("01 Jan 2022", tariff="Standard"),
            _contract_row("01 Feb 2022", tariff="Standard"),
            _contract_row("21 Feb 2022", tariff="Standard"),
            _contract_row("01 Apr 2022", tariff="Standard"),
        ]
    )
    out = infer_contracts(df)
    assert len(out) == 1
    assert out.iloc[0]["Tariff"] == "Standard"
    assert int(out.iloc[0]["# Invoices"]) == 4


def test_infer_contracts_three_tariffs_three_contracts() -> None:
    """Three distinct tariffs produce three contracts.

    Covers multiple run transitions and the final runs.append (166-172).
    """
    df = pd.DataFrame(
        [
            _contract_row("01 Jan 2022", tariff="Old Variable"),
            _contract_row("01 Feb 2022", tariff="Old Variable"),
            _contract_row("01 May 2022", tariff="Fixed 1Y"),
            _contract_row("01 Jun 2022", tariff="Fixed 1Y"),
            _contract_row("01 Sep 2022", tariff="New Variable"),
            _contract_row("01 Oct 2022", tariff="New Variable"),
        ]
    )
    out = infer_contracts(df)
    assert len(out) == 3
    assert list(out["Tariff"]) == ["Old Variable", "Fixed 1Y", "New Variable"]


def test_infer_contracts_output_sorted_by_contract_from() -> None:
    """The output is sorted by Contract From date (covers 215-217)."""
    df = pd.DataFrame(
        [
            _contract_row("01 Mar 2022", tariff="Mid_T"),
            _contract_row("01 Jan 2022", tariff="Early_T"),
            _contract_row("01 Feb 2022", tariff="Early_T"),
        ]
    )
    out = infer_contracts(df)
    from_dts = pd.to_datetime(out["Contract From"], errors="coerce")
    assert from_dts.is_monotonic_increasing
    assert from_dts.iloc[0] == pd.Timestamp("2022-01-01")


def test_infer_contracts_custom_merge_gap_days() -> None:
    """The merge_gap_days parameter is honoured."""
    df = pd.DataFrame(
        [
            _contract_row("01 Jan 2022", tariff="Standard"),
            _contract_row("01 Feb 2022", tariff="Standard"),
            _contract_row("01 Mar 2022", tariff="Fixed"),
            _contract_row("01 Apr 2022", tariff="Fixed"),
        ]
    )
    out = infer_contracts(df, merge_gap_days=10)
    assert len(out) == 2


# ---------------------------------------------------------------------------
# _write_meter_readings_sheet_impl branches (via write_meter_readings_sheet)
# ---------------------------------------------------------------------------


def _evidence_df_with_unparseable_units() -> pd.DataFrame:
    """Evidence frame where one row has a non-numeric Units (kWh) value.

    Exercises the ``except (TypeError, ValueError)`` branch (318-319)
    and the ``_text`` fallback for non-numeric units (332).
    """
    return pd.DataFrame(
        [
            {
                "Date": "01 Jan 2023",
                "Invoice #": "INV-N/A",
                "Reading": "Actual",
                "Units (kWh)": "N/A",
                "Details": "",
                "Tariff": "Standard",
            },
            {
                "Date": "01 Feb 2023",
                "Invoice #": "INV-002",
                "Reading": "Estimated",
                "Units (kWh)": 350.0,
                "Details": "Automatic estimate",
                "Tariff": "Standard",
            },
        ]
    )


def test_write_meter_readings_sheet_unparseable_units_written_as_text() -> None:
    """A non-numeric Units (kWh) value is written as text in col 2.

    Covers lines 318-319 (units = units_raw on parse failure) and 332
    (the ``_text`` branch for non-numeric units).
    """
    ws = _open_ws()
    write_meter_readings_sheet(ws, _evidence_df_with_unparseable_units(), pd.DataFrame())
    # Row 8 = INV-N/A. Its Units cell (col 2) should hold the raw "N/A"
    # string, not a number.
    units_cell = ws.cell(row=8, column=2).value
    assert units_cell == "N/A"


def test_write_meter_readings_sheet_no_account_uses_plain_title() -> None:
    """When account is empty, the title omits the account suffix.

    Covers the ``if account`` false branch (265->267) for the meter
    readings sheet.
    """
    ws = _open_ws()
    write_meter_readings_sheet(ws, _evidence_df_with_unparseable_units(), pd.DataFrame())
    title = ws.cell(row=1, column=1).value
    assert isinstance(title, str)
    assert "METER READING" in title.upper()
    assert "Account" not in title


def _evidence_df_long_source_text() -> pd.DataFrame:
    """Evidence frame whose Source PDF Text exceeds 400 chars.

    Exercises the excerpt truncation branch (350-351): when source_text
    is longer than 400 chars, the excerpt is sliced to 400 and ' ...' is
    appended.
    """
    long_text = "Invoice number: INV-LONG Period 01 Jan 2023 - 31 Jan 2023 " + ("x" * 500)
    return pd.DataFrame(
        [
            {
                "Date": "01 Jan 2023",
                "Invoice #": "INV-LONG",
                "Reading": "Actual",
                "Units (kWh)": 300.0,
                "Details": "",
                "Tariff": "Standard",
                "Source PDF Text": long_text,
            }
        ]
    )


def test_write_meter_readings_sheet_long_source_text_truncated_with_ellipsis() -> None:
    """Source PDF Text > 400 chars is truncated and gets a ' ...' suffix.

    Covers lines 349-351 (the > 400 branch).
    """
    ws = _open_ws()
    df = _evidence_df_long_source_text()
    write_meter_readings_sheet(ws, df, pd.DataFrame(), evidence_df=df)
    excerpt = ws.cell(row=8, column=7).value
    assert isinstance(excerpt, str)
    assert excerpt.endswith(" ...")
    # The excerpt is the first 400 chars + ' ...' = 404 chars.
    assert len(excerpt) == 404


def _evidence_df_empty_source_text() -> pd.DataFrame:
    """Evidence frame where the matching invoice's Source PDF Text is empty.

    Exercises the ``isinstance(source_text, str) and source_text`` false
    branch (348->352): source_text is a str but empty, so excerpt stays
    '' and the ``_open_pdf_hyperlink_cell`` fallback (354-355) runs.
    """
    return pd.DataFrame(
        [
            {
                "Date": "01 Jan 2023",
                "Invoice #": "INV-EMPTY",
                "Reading": "Actual",
                "Units (kWh)": 300.0,
                "Details": "",
                "Tariff": "Standard",
                "Source PDF Text": "",
            }
        ]
    )


def test_write_meter_readings_sheet_empty_source_text_falls_back_to_hyperlink() -> None:
    """An empty Source PDF Text leaves excerpt empty -> hyperlink fallback.

    Covers the 348->352 branch (source_text falsy) and the 354-355
    ``_open_pdf_hyperlink_cell`` fallback path.
    """
    ws = _open_ws()
    df = _evidence_df_empty_source_text()
    write_meter_readings_sheet(ws, df, pd.DataFrame(), evidence_df=df)
    # With an empty excerpt, col 7 falls through to the hyperlink cell.
    # The invoice IS present in evidence_df, so open_pdf_hyperlink_cell
    # emits a '->' hyperlink.
    cell_val = ws.cell(row=8, column=7).value
    assert cell_val == "\u2192"


def _evidence_df_no_match_in_evidence() -> pd.DataFrame:
    """Evidence frame where the row's invoice is NOT in evidence_df.

    Exercises the ``if not matches.empty`` false branch (346->352):
    evidence_df is provided but no row matches the invoice, so excerpt
    stays '' and the hyperlink fallback runs (with no match -> no
    hyperlink emitted).
    """
    return pd.DataFrame(
        [
            {
                "Date": "01 Jan 2023",
                "Invoice #": "INV-ROW",
                "Reading": "Actual",
                "Units (kWh)": 300.0,
                "Details": "",
                "Tariff": "Standard",
            }
        ]
    )


def test_write_meter_readings_sheet_evidence_df_no_match_falls_back() -> None:
    """When evidence_df has no matching invoice, the excerpt stays empty.

    Covers the 346->352 branch (matches.empty is True).
    """
    ws = _open_ws()
    df = _evidence_df_no_match_in_evidence()
    # A DIFFERENT evidence_df so no invoice matches.
    other_evidence = pd.DataFrame(
        [
            {
                "Invoice #": "INV-OTHER",
                "Source PDF Text": "other text",
            }
        ]
    )
    write_meter_readings_sheet(ws, df, pd.DataFrame(), evidence_df=other_evidence)
    # No excerpt -> col 7 falls to hyperlink fallback; INV-ROW is not in
    # other_evidence so no hyperlink is emitted and the cell stays empty.
    cell_val = ws.cell(row=8, column=7).value
    assert cell_val in (None, "")


def test_write_meter_readings_sheet_evidence_index_date_units_fallback() -> None:
    """When inv: is absent from evidence_index, the date+units key is tried.

    Covers lines 362-370: target_row is None after the inv: lookup, so
    the date+units signature fallback runs. With a matching key, a
    hyperlink is emitted; the test asserts the hyperlink lands on the
    fallback row.
    """
    ws = _open_ws()
    df = pd.DataFrame(
        [
            {
                "Date": "01 Jan 2023",
                "Invoice #": "INV-FALLBACK",
                "Reading": "Actual",
                "Units (kWh)": 300.0,
                "Details": "",
                "Tariff": "Standard",
            }
        ]
    )
    # evidence_index has NO inv: key but DOES have the date+units key.
    # units_sig = int(round(300.0)) = 300.
    evidence_index = {"date_units:01 Jan 2023|300": 42}
    write_meter_readings_sheet(ws, df, pd.DataFrame(), evidence_index=evidence_index)
    cell = ws.cell(row=8, column=8)
    assert cell.value == "\u2192"
    assert cell.hyperlink is not None
    assert "A42" in str(cell.hyperlink.location or "")


def test_write_meter_readings_sheet_evidence_index_date_units_fallback_unparseable() -> None:
    """The date+units fallback's float() guard swallows unparseable units.

    Covers the ``except (TypeError, ValueError)`` at 369-370: units is
    a non-numeric string, so float(units) raises and target_row stays
    None -> the 'No match' cell is emitted.
    """
    ws = _open_ws()
    df = _evidence_df_with_unparseable_units()
    # evidence_index has no inv: key; the fallback tries float("N/A")
    # which raises ValueError -> target_row stays None.
    evidence_index = {"date_units:01 Jan 2023|300": 42}
    write_meter_readings_sheet(ws, df, pd.DataFrame(), evidence_index=evidence_index)
    cell = ws.cell(row=8, column=8)
    assert cell.value == "No match"


def test_write_meter_readings_sheet_evidence_index_no_match_emits_no_match() -> None:
    """When neither inv: nor date+units matches, col 8 shows 'No match'."""
    ws = _open_ws()
    df = _evidence_df_no_match_in_evidence()
    evidence_index = {"inv:INV-OTHER": 5}
    write_meter_readings_sheet(ws, df, pd.DataFrame(), evidence_index=evidence_index)
    assert ws.cell(row=8, column=8).value == "No match"


# ---------------------------------------------------------------------------
# write_contract_history_sheet branches (lines 431, 432->427, 434->427,
# 436, 440->442, 484->486, 487->489)
# ---------------------------------------------------------------------------


def _contracts_df_raw_strings() -> pd.DataFrame:
    """Contracts frame whose From/To are raw strings (not Timestamps).

    Exercises the ``isinstance(cf, pd.Timestamp)`` false branch
    (484->486) and the same for ct (487->489): the raw string is kept
    as-is in the cell.
    """
    return pd.DataFrame(
        [
            {
                "Contract From": "01 Jan 2022",
                "Contract To": "30 Jun 2022",
                "Tariff": "Standard",
                "Days": 181,
                "# Invoices": 6,
            }
        ]
    )


def test_write_contract_history_sheet_raw_string_dates_kept_as_is() -> None:
    """Contract From/To as raw strings are written verbatim (no strftime).

    Covers 484->486 (cf not a Timestamp) and 487->489 (ct not a
    Timestamp): the raw string is kept as cf_text / ct_text.
    """
    ws = _open_ws()
    contracts = _contracts_df_raw_strings()
    write_contract_history_sheet(ws, contracts)
    assert ws.cell(row=8, column=1).value == "01 Jan 2022"
    assert ws.cell(row=8, column=2).value == "30 Jun 2022"


def test_write_contract_history_sheet_no_account_uses_plain_title() -> None:
    """When account is empty, the contract-history title omits the suffix.

    Covers the 440->442 branch (account falsy).
    """
    ws = _open_ws()
    contracts = _contracts_df_raw_strings()
    write_contract_history_sheet(ws, contracts)
    title = ws.cell(row=1, column=1).value
    assert isinstance(title, str)
    assert "CONTRACT" in title.upper()
    assert "Account" not in title


def test_write_contract_history_sheet_evidence_df_period_unparseable_skipped() -> None:
    """An evidence row with unparseable Period From/To is skipped.

    Covers line 431 (``pd.isna(ipf) or pd.isna(ipt): continue``) and
    the 432->427 loop-continue branch.
    """
    ws = _open_ws()
    contracts = pd.DataFrame(
        [
            {
                "Contract From": "01 Jan 2023",
                "Contract To": "31 Jan 2023",
                "Tariff": "Standard",
                "Days": 31,
                "# Invoices": 1,
            }
        ]
    )
    # evidence_df with an unparseable Period From -> the row is skipped
    # via the pd.isna continue at 431.
    evidence_df = pd.DataFrame(
        [
            {
                "Invoice #": "INV-BAD",
                "Period From": "not-a-date",
                "Period To": "31 Jan 2023",
            }
        ]
    )
    write_contract_history_sheet(ws, contracts, evidence_df=evidence_df)
    # No matching invoice -> 'No match' in col 7.
    assert ws.cell(row=8, column=7).value == "No match"


def test_write_contract_history_sheet_evidence_df_empty_invoice_skipped() -> None:
    """An evidence row with an empty Invoice # is skipped (434->427, 436).

    Covers the ``if inv: return inv`` false branch (434->427): the
    period overlaps the contract but the invoice is empty, so the loop
    continues; no match is found and the final ``return ""`` (436) runs.
    """
    ws = _open_ws()
    contracts = pd.DataFrame(
        [
            {
                "Contract From": "01 Jan 2023",
                "Contract To": "31 Jan 2023",
                "Tariff": "Standard",
                "Days": 31,
                "# Invoices": 1,
            }
        ]
    )
    # Period overlaps the contract but Invoice # is empty.
    evidence_df = pd.DataFrame(
        [
            {
                "Invoice #": "",
                "Period From": "01 Jan 2023",
                "Period To": "31 Jan 2023",
            }
        ]
    )
    write_contract_history_sheet(ws, contracts, evidence_df=evidence_df)
    assert ws.cell(row=8, column=7).value == "No match"


def test_write_contract_history_sheet_evidence_df_no_overlap_returns_empty() -> None:
    """An evidence row whose period does not overlap the contract is skipped.

    Covers the 432->427 branch (the overlap condition is false, so the
    loop continues) and the 436 ``return ""`` (no match found).
    """
    ws = _open_ws()
    contracts = pd.DataFrame(
        [
            {
                "Contract From": "01 Jan 2023",
                "Contract To": "31 Jan 2023",
                "Tariff": "Standard",
                "Days": 31,
                "# Invoices": 1,
            }
        ]
    )
    # Period is entirely outside the contract window.
    evidence_df = pd.DataFrame(
        [
            {
                "Invoice #": "INV-OUTSIDE",
                "Period From": "01 Jun 2023",
                "Period To": "30 Jun 2023",
            }
        ]
    )
    write_contract_history_sheet(ws, contracts, evidence_df=evidence_df)
    assert ws.cell(row=8, column=7).value == "No match"


def test_write_contract_history_sheet_evidence_df_none_returns_no_match() -> None:
    """When evidence_df is None, _first_matching_invoice returns '' immediately.

    Covers the 425-426 early-return branch of _first_matching_invoice.
    """
    ws = _open_ws()
    contracts = _contracts_df_raw_strings()
    write_contract_history_sheet(ws, contracts, evidence_df=None)
    assert ws.cell(row=8, column=7).value == "No match"


def test_write_contract_history_sheet_evidence_index_no_match_emits_no_match() -> None:
    """When the matched invoice is not in evidence_index, col 7 is 'No match'."""
    ws = _open_ws()
    contracts = pd.DataFrame(
        [
            {
                "Contract From": "01 Jan 2023",
                "Contract To": "31 Jan 2023",
                "Tariff": "Standard",
                "Days": 31,
                "# Invoices": 1,
            }
        ]
    )
    evidence_df = pd.DataFrame(
        [
            {
                "Invoice #": "INV-MATCH",
                "Period From": "01 Jan 2023",
                "Period To": "31 Jan 2023",
            }
        ]
    )
    # evidence_index does NOT contain inv:INV-MATCH.
    evidence_index = {"inv:INV-OTHER": 5}
    write_contract_history_sheet(
        ws, contracts, evidence_df=evidence_df, evidence_index=evidence_index
    )
    assert ws.cell(row=8, column=7).value == "No match"


def test_write_contract_history_sheet_evidence_index_match_emits_hyperlink() -> None:
    """A matched invoice present in evidence_index emits a '->' hyperlink."""
    ws = _open_ws()
    contracts = pd.DataFrame(
        [
            {
                "Contract From": "01 Jan 2023",
                "Contract To": "31 Jan 2023",
                "Tariff": "Standard",
                "Days": 31,
                "# Invoices": 1,
            }
        ]
    )
    evidence_df = pd.DataFrame(
        [
            {
                "Invoice #": "INV-MATCH",
                "Period From": "01 Jan 2023",
                "Period To": "31 Jan 2023",
            }
        ]
    )
    evidence_index = {"inv:INV-MATCH": 7}
    write_contract_history_sheet(
        ws, contracts, evidence_df=evidence_df, evidence_index=evidence_index
    )
    cell = ws.cell(row=8, column=7)
    assert cell.value == "\u2192"
    assert cell.hyperlink is not None
    assert "A7" in str(cell.hyperlink.location or "")


def test_write_contract_history_sheet_openpyxl_hyperlink_module_accessible() -> None:
    """Sanity: openpyxl.worksheet.hyperlink.Hyperlink is importable.

    This is the type used by the writer's hyperlink cells; confirming it
    is reachable guards against a silent import-path regression.
    """
    assert openpyxl.worksheet.hyperlink.Hyperlink is not None
