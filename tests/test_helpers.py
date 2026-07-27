"""Tests for edf_bill_fetcher.helpers submodule.

Verifies that helpers import correctly and behave correctly.
"""

from __future__ import annotations

import openpyxl

from edf_bill_fetcher.helpers.date_utils import (
    build_evidence_trail,
    completeness_score,
    compute_ema,
    compute_momentum,
    compute_rolling_stats,
)
from edf_bill_fetcher.helpers.excel_utils import (
    _TEXT_SUPPRESSION_QUEUE,
    build_sap_row_index_map,
    hcell,
    money,
    num,
    open_pdf_hyperlink_cell,
    section_hdr,
    set_column_widths_from_spec,
    suppress_text_warning,
)
from edf_bill_fetcher.helpers.excel_utils import (
    text as excel_text,
)
from edf_bill_fetcher.helpers.formatting import (
    account_number_matches,
    apply_currency_format,
    apply_int_format,
)


def test_formatting_module_imports():
    assert apply_currency_format is not None
    assert apply_int_format is not None
    assert account_number_matches is not None


def test_apply_currency_format_on_cell():
    wb = openpyxl.Workbook()
    ws = wb.active
    cell = ws.cell(row=1, column=1, value="123.45")
    apply_currency_format(cell)
    assert cell.number_format == "\u00a3#,##0.00"


def test_apply_int_format_on_cell():
    wb = openpyxl.Workbook()
    ws = wb.active
    cell = ws.cell(row=1, column=1, value="42")
    apply_int_format(cell)
    assert cell.number_format == "#,##0"


def test_account_number_matches_basic():
    assert account_number_matches("31", "Account number: 31 555 4444") is True
    assert account_number_matches("31", "Account: 31") is True
    assert account_number_matches("999", "Account: 31") is False


def test_date_utils_imports():
    assert completeness_score is not None
    assert compute_ema is not None
    assert compute_momentum is not None
    assert compute_rolling_stats is not None
    assert build_evidence_trail is not None


def test_compute_ema_basic():
    import pandas as pd

    s = pd.Series([1.0, 2.0, 3.0, 4.0, 5.0])
    ema = compute_ema(s, span=2)
    assert len(ema) == len(s)
    assert ema.iloc[-1] > ema.iloc[0]  # trending up


def test_excel_utils_imports():
    assert hcell is not None
    assert excel_text is not None
    assert num is not None
    assert money is not None
    assert section_hdr is not None
    assert set_column_widths_from_spec is not None
    assert suppress_text_warning is not None
    assert open_pdf_hyperlink_cell is not None
    assert build_sap_row_index_map is not None


def test_hcell_writes_header():
    wb = openpyxl.Workbook()
    ws = wb.active
    cell = hcell(ws, 1, 1, "Header", bg="FF0000")
    assert cell.value == "Header"
    assert cell.font.bold is True


def test_excel_text_formula_guard():
    wb = openpyxl.Workbook()
    ws = wb.active
    cell = excel_text(ws, 1, 1, "=SUM(A1:A10)")
    assert cell.value.startswith("'")


def test_open_pdf_hyperlink_cell_empty_invoice():
    wb = openpyxl.Workbook()
    ws = wb.active
    open_pdf_hyperlink_cell(ws, 1, 1, None, "")
    # Should not raise and no hyperlink should be set
    assert ws.cell(row=1, column=1).value is None or ws.cell(row=1, column=1).hyperlink is None


def test_suppress_text_warning_queues():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "TestSheet"
    _TEXT_SUPPRESSION_QUEUE.pop("TestSheet", None)
    suppress_text_warning(ws, "F", 2, 225)
    assert "TestSheet" in _TEXT_SUPPRESSION_QUEUE
    assert ("F", 2, 225) in _TEXT_SUPPRESSION_QUEUE["TestSheet"]
    _TEXT_SUPPRESSION_QUEUE.pop("TestSheet", None)


def test_pdf_utils_module_exists():
    from edf_bill_fetcher.helpers import pdf_utils

    assert pdf_utils is not None


def test_helpers_submodule_top_level_imports():
    from edf_bill_fetcher.helpers import date_utils, excel_utils, formatting, pdf_utils

    assert date_utils is not None
    assert excel_utils is not None
    assert formatting is not None
    assert pdf_utils is not None
