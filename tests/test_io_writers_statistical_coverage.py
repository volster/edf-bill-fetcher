"""Coverage tests for the statistical analysis writer — closes the
25-missed-line gap in ``edf_bill_fetcher/io/writers/statistical.py``
(79% -> ~100%).

Targets the 4 missed regions directly:
  * L126, L128 — pandas ``.skew()`` / ``.kurtosis()`` calls (need n>=3
    valid amounts beyond the early-return at ``n < 3``).
  * L212-220 — Z-Score Anomaly Dates block (need at least one outlier
    that crosses the z-score threshold).
  * L228-236 — IQR Anomaly Dates block (need at least one value in the
    IQR whisker's outside).
  * L267-272 — Scipy Jarque-Bera / Shapiro-Wilk normality tests (need
    scipy installed AND enough non-NaN data points).

All tests use synthetic ``pandas.DataFrame`` fixtures with the column
shape the writer expects; an in-process ``openpyxl.Workbook`` active
sheet is the ``ws`` argument. No mocking required.
"""

from __future__ import annotations

import openpyxl
import pandas as pd
import pytest

from edf_bill_fetcher.io.writers.statistical import (
    _HAS_SCIPY,
    write_statistical_analysis_sheet,
)


def _make_workbook() -> openpyxl.Workbook:
    """Fresh workbook with a single unnamed sheet for the writer to rename."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "placeholder"
    return wb


def _make_dataframe(rows: list[dict]) -> pd.DataFrame:
    """Build a DataFrame from a list of dicts with the writer's required columns."""
    return pd.DataFrame(rows)


# ---------- early-return branch (n < 3) ----------


def test_writer_returns_early_for_insufficient_data() -> None:
    """When dfc has fewer than 3 records, the writer only writes the insufficiency banner."""
    wb = _make_workbook()
    ws = wb.active
    dfc = _make_dataframe(
        [
            {"Date": "01 Jan 2024", "Amount (£)": "10.00"},
            {"Date": "15 Jan 2024", "Amount (£)": "20.00"},
        ]
    )
    write_statistical_analysis_sheet(ws, dfc, config={})
    assert ws["A1"].value == "Insufficient data for statistical analysis"
    assert ws.column_dimensions["A"].width == 50
    assert ws["A2"].value is None


# ---------- full statistical pass (covers L126, L128) ----------


def test_writer_writes_descriptive_statistics_with_enough_rows() -> None:
    """5 rows are enough to exercise the skewness/kurtosis branches (L126, L128)."""
    wb = _make_workbook()
    ws = wb.active
    dfc = _make_dataframe(
        [
            {"Date": "01 Jan 2024", "Amount (£)": "100.00"},
            {"Date": "01 Feb 2024", "Amount (£)": "105.00"},
            {"Date": "01 Mar 2024", "Amount (£)": "98.00"},
            {"Date": "01 Apr 2024", "Amount (£)": "102.50"},
            {"Date": "01 May 2024", "Amount (£)": "110.00"},
        ]
    )
    write_statistical_analysis_sheet(ws, dfc, config={})
    assert ws["A1"].value is not None
    assert "STATISTICAL ANALYSIS" in str(ws["A1"].value)
    descriptive_cells = [
        str(ws.cell(row=r, column=1).value or "")
        for r in range(2, 30)
        if ws.cell(row=r, column=1).value
    ]
    assert any("DESCRIPTIVE STATISTICS" in s for s in descriptive_cells)
    all_row_values = [ws.cell(row=r, column=1).value for r in range(2, 50)]
    assert any(v == "Skewness" for v in all_row_values)
    assert any(v == "Kurtosis" for v in all_row_values)


# ---------- Z-Score anomaly block (L212-220) ----------


def test_writer_lists_z_score_anomalies_when_outliers_present() -> None:
    """Z-Score Anomaly Dates block (L212-220) fires when 15 normal values + 1 large outlier
    push the z-score past 2.5σ (a single 8-row series isn't enough: the outlier pulls std
    high enough that z stays below 2.5)."""
    wb = _make_workbook()
    ws = wb.active
    normal_rows = [
        {"Date": f"01 {mon} 2024", "Amount (£)": f"{100.00 + i * 0.50:.2f}"}
        for i, mon in enumerate(
            [
                "Jan",
                "Feb",
                "Mar",
                "Apr",
                "May",
                "Jun",
                "Jul",
                "Aug",
                "Sep",
                "Oct",
                "Nov",
                "Dec",
            ]
        )
    ]
    outlier_row = [{"Date": "01 Jan 2025", "Amount (£)": "5000.00"}]
    # 13 normal values from Jan–Dec plus a single large outlier at the end
    rows = normal_rows[:13] + outlier_row
    dfc = _make_dataframe(rows)
    write_statistical_analysis_sheet(ws, dfc, config={})
    all_row_values = [ws.cell(row=r, column=1).value for r in range(1, 80)]
    assert any(v and "Z-Score Anomaly" in str(v) for v in all_row_values), (
        "Z-Score Anomaly Dates header should be written when z_count > 0"
    )


# ---------- IQR anomaly block (L228-236) ----------


def test_writer_lists_iqr_anomalies_when_outliers_present() -> None:
    """A series with a clear outlier triggers the IQR Anomaly Dates block (L228-236)."""
    wb = _make_workbook()
    ws = wb.active
    dfc = _make_dataframe(
        [
            {"Date": "01 Jan 2024", "Amount (£)": "100.00"},
            {"Date": "01 Feb 2024", "Amount (£)": "101.00"},
            {"Date": "01 Mar 2024", "Amount (£)": "99.00"},
            {"Date": "01 Apr 2024", "Amount (£)": "100.50"},
            {"Date": "01 May 2024", "Amount (£)": "98.50"},
            {"Date": "01 Jun 2024", "Amount (£)": "99999.00"},
        ]
    )
    write_statistical_analysis_sheet(ws, dfc, config={})
    all_row_values = [ws.cell(row=r, column=1).value for r in range(1, 80)]
    assert any(v and "IQR Anomaly" in str(v) for v in all_row_values), (
        "IQR Anomaly Dates header should be written when iqr_count > 0"
    )


# ---------- Scipy normality tests block (L267-272) ----------


@pytest.mark.skipif(not _HAS_SCIPY, reason="scipy not installed — L267-272 branch unreachable")
def test_writer_runs_scipy_normality_tests_when_available() -> None:
    """When scipy is installed, the Jarque-Bera / Shapiro-Wilk block executes (L267-272)."""
    wb = _make_workbook()
    ws = wb.active
    dfc = _make_dataframe(
        [
            {"Date": "01 Jan 2024", "Amount (£)": "100.00"},
            {"Date": "01 Feb 2024", "Amount (£)": "105.00"},
            {"Date": "01 Mar 2024", "Amount (£)": "98.00"},
            {"Date": "01 Apr 2024", "Amount (£)": "102.50"},
            {"Date": "01 May 2024", "Amount (£)": "110.00"},
            {"Date": "01 Jun 2024", "Amount (£)": "97.50"},
            {"Date": "01 Jul 2024", "Amount (£)": "103.00"},
            {"Date": "01 Aug 2024", "Amount (£)": "100.50"},
        ]
    )
    write_statistical_analysis_sheet(ws, dfc, config={})
    all_row_values = [ws.cell(row=r, column=1).value for r in range(1, 100)]
    assert any(v and "Jarque-Bera" in str(v) for v in all_row_values), (
        "Jarque-Bera normality test header should be written when scipy is installed"
    )


@pytest.mark.skipif(_HAS_SCIPY, reason="scipy IS installed — branch writes Jarque-Bera result")
def test_writer_prints_scipy_unavailable_message_when_not_installed() -> None:
    """When scipy is NOT installed, the writer writes the 'Scipy not available' banner (L272)."""
    wb = _make_workbook()
    ws = wb.active
    dfc = _make_dataframe(
        [
            {"Date": "01 Jan 2024", "Amount (£)": "100.00"},
            {"Date": "01 Feb 2024", "Amount (£)": "105.00"},
            {"Date": "01 Mar 2024", "Amount (£)": "98.00"},
            {"Date": "01 Apr 2024", "Amount (£)": "102.50"},
        ]
    )
    write_statistical_analysis_sheet(ws, dfc, config={})
    all_row_values = [ws.cell(row=r, column=1).value for r in range(1, 100)]
    assert any(v and "Scipy not available" in str(v) for v in all_row_values), (
        "Scipy-not-available banner should be written when _HAS_SCIPY is False"
    )
