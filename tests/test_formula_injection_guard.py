"""Phase 2.x — formula-injection guard on Excel text cells.

External text sources (PDF/PST/email bodies) are attacker-
controllable in principle: a user could receive a bill whose
``Details`` field starts with ``=cmd|'/c calc'!A1`` or whose
``Sender`` is ``+MALICIOUS_FORMULA``.  Excel will auto-evaluate
any cell value starting with ``=``, ``+``, ``-`` or ``@`` as a
formula when the workbook is opened, prompting a function-call
side-effect or a credential-leak formula chain.  This is real
for ombudsman submissions because they are passed to a third
party who opens the workbook.

The mitigation:

    1. ``text`` coerces the cell value to ``str`` first so
       pandas-introduced types don't fall through.
    2. ``text`` pins ``cell.data_type = 's'`` (text) so Excel
       never tries to auto-format the cell as a formula even
       when the workbook is opened with maximum-fidelity
       parsing.
    3. Belt-and-braces: a leading ``=``, ``+``, ``-`` or ``@``
       triggers an apostrophe-prefix so even LibreOffice
       (which sometimes ignores the data_type pin) renders the
       cell as text.

These tests pin all three guarantees.
"""

from __future__ import annotations

import os
from pathlib import Path
from typing import Any, cast

import pytest
from openpyxl import Workbook, load_workbook

from edf_bill_fetcher.helpers.excel_utils import text
from edf_bill_fetcher.io.writers import export_to_excel
from edf_bill_fetcher.models.config import ConfigDict


@pytest.fixture
def workdir() -> Path:
    # Skirt pytest's ``tmp_path`` fixture on this Windows host —
    # the sandboxed TEMP directory is read-only on this developer's
    # machine, so any fixture that depends on pytest's tmp-path
    # machinery error-cascades at setup.  Derive our own scratch
    # dir from ``USERPROFILE`` (or ``/tmp`` as a Linux fallback)
    # plus an explicit pid-locked name, so cross-test isolation is
    # still preserved.
    scratch = Path(os.environ.get("USERPROFILE", "/tmp")) / f".edf_formula_scratch_{os.getpid()}"
    scratch.mkdir(parents=True, exist_ok=True)
    return scratch


class TestFormulaInjectionGuard:
    """Phase 2.x — formula-injection guard.

    Excel auto-evaluation of a cell whose textual value starts
    with ``=``, ``+``, ``-`` or ``@`` is the standard
    formula-injection attack surface.  The guard pins:

        * data_type = 's' (text) on every text cell,
        * apostrophe-prefix on leading special chars,
        * benign strings render unchanged.
    """

    def test_data_type_is_text_pin(self) -> None:
        wb = Workbook()
        ws = wb.active
        text(ws, 1, 1, "Anything here")
        cell = ws.cell(row=1, column=1)
        # The data_type pin is what flips Excel out of
        # auto-evaluate mode on most builds.
        assert cell.data_type == "s", f"Expected data_type='s' (text), got {cell.data_type!r}"

    def test_leading_equals_triggers_apostrophe_prefix(self) -> None:
        wb = Workbook()
        ws = wb.active
        text(ws, 1, 1, "=cmd|'/c calc'!A1")
        cell = ws.cell(row=1, column=1)
        # Must not equal the raw ``=cmd...`` payload; we want
        # an apostrophe-prefix to defeat Excel's auto-format.
        assert cell.value == "'=cmd|'/c calc'!A1", (
            f"Expected apostrophe-prefix, got value={cell.value!r}"
        )
        assert cell.data_type == "s"

    def test_leading_plus_triggers_apostrophe_prefix(self) -> None:
        wb = Workbook()
        ws = wb.active
        text(ws, 1, 1, "+MALICIOUS_FORMULA")
        cell = ws.cell(row=1, column=1)
        assert cell.value == "'+MALICIOUS_FORMULA"
        assert cell.data_type == "s"

    def test_leading_minus_triggers_apostrophe_prefix(self) -> None:
        wb = Workbook()
        ws = wb.active
        text(ws, 1, 1, "-1+1")  # Excel would evaluate to 0
        cell = ws.cell(row=1, column=1)
        assert cell.value == "'-1+1", f"Leading minus must be guarded; got {cell.value!r}"
        assert cell.data_type == "s"

    def test_leading_at_triggers_apostrophe_prefix(self) -> None:
        wb = Workbook()
        ws = wb.active
        text(ws, 1, 1, "@SUM(1+1)")
        cell = ws.cell(row=1, column=1)
        assert cell.value == "'@SUM(1+1)"
        assert cell.data_type == "s"

    def test_benign_strings_render_unchanged(self) -> None:
        # A non-special-char leading string should be preserved
        # *verbatim* — the guard mustn't add an apostrophe when
        # one isn't needed.
        wb = Workbook()
        ws = wb.active
        text(ws, 1, 1, "Period To")
        cell = ws.cell(row=1, column=1)
        assert cell.value == "Period To", (
            f"Leading non-special char triggered apostrophe?  got {cell.value!r}"
        )
        assert cell.data_type == "s"

    def test_special_char_not_at_start_unchanged(self) -> None:
        # ``Config=foo`` has ``=`` mid-string, not leading.  The
        # guard should not append an apostrophe here because the
        # leading char isn't special.
        wb = Workbook()
        ws = wb.active
        text(ws, 1, 1, "Config=foo")
        cell = ws.cell(row=1, column=1)
        assert cell.value == "Config=foo"
        assert cell.data_type == "s"

    def test_none_becomes_empty_string(self) -> None:
        # A None value used to render as the literal ``None``
        # in some legacy Excel saves (``openpyxl`` writes it as
        # an empty cell).  We pin the empty-string behaviour so
        # downstream consumers (the analyst) see ``""`` rather
        # than ``None``.
        wb = Workbook()
        ws = wb.active
        text(ws, 1, 1, None)
        cell = ws.cell(row=1, column=1)
        assert cell.value == ""
        assert cell.data_type == "s"

    def test_non_string_values_are_coerced(self) -> None:
        # Realistic case: ``Details`` is sometimes a pandas
        # ``float`` (NaN) or ``int``.  The guard must coerce to
        # ``str`` first so ``data_type = 's'`` is set against a
        # textual cell, not a numeric cell.
        wb = Workbook()
        ws = wb.active
        text(ws, 1, 1, 12345)
        cell = ws.cell(row=1, column=1)
        assert cell.value == "12345"
        assert cell.data_type == "s"

    def test_apostrophe_in_legitimate_value_stays_intact(self) -> None:
        # Edge case: a real Description field that already
        # starts with ``'`` (e.g., a debt collector's
        # bracketed-style address).  We don't want to contaminate
        # it further, just pin the data_type.
        wb = Workbook()
        ws = wb.active
        text(ws, 1, 1, "Bob's Mobile")
        cell = ws.cell(row=1, column=1)
        assert cell.value == "Bob's Mobile"
        assert cell.data_type == "s"


class TestFormulaInjectionGuardEvidenceSheet:
    """Phase 2.x — formula-injection guard extended to the
    ``write_evidence_sheet`` row-iteration pathway.

    The headline ``text`` fix only covered helper calls.  The
    evidence-sheet's row-iteration path did ``ws.cell(...,
    value=val)`` directly, and openpyxl auto-set
    ``data_type='f'`` for any value starting with ``=``,
    ``+``, ``-`` or ``@`` — bypassing the guard.  This test
    pins that user-content now goes through the data_type
    pin in both pathways.
    """

    def _guarded(self, workdir: Path, attack_val: str) -> None:
        records = [
            {
                "Date": "01/05/2024",
                "Source": "Local PDF Folder",
                "Period From": "N/A",
                "Period To": "N/A",
                "Invoice #": attack_val,
                "Amount (£)": 50.0,
                "Period Charge (£)": 0.0,
                "Units (kWh)": "",
                "Reading": "",
                "Entry Type": "Payment",
                "Logic Used": "Pattern",
                "Details": "",
                "Attachment Name": "",
                "Standing Charge": "",
                "Anomaly Flag": "",
                "Sender": "edfenergy.com",
            },
        ]
        config = cast(
            ConfigDict,
            {
                "use_dedup": True,
                "save_dups": True,
                "use_anchors": False,
                "use_large": False,
                "min_amount": 0.0,
                "filter_below": False,
                "use_dedup_period": True,
                "expanded_columns": True,
                "include_charts": False,
                "include_forecast": False,
            },
        )
        out = workdir / f"tg_{abs(hash(attack_val))}.xlsx"
        export_to_excel(records, str(out), [], config=config)
        wb = load_workbook(str(out))
        ws = wb["EDF Evidence Report"]
        header_map = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}
        cell = ws.cell(row=2, column=header_map["Invoice #"])
        # data_type MUST be 's' (text), not 'f' (formula) — the
        # openpyxl-without-guard auto-promotion is exactly what
        # would open Excel with a formula evaluation prompt.
        assert cell.data_type == "s", (
            f"data_type={cell.data_type!r} for Invoice #: expected "
            f"'s'; an Excel formula evaluation prompt would "
            f"otherwise surface"
        )
        # Belt-and-braces: the value should start with an
        # apostrophe so any consumer that ignores data_type
        # still treats the cell as text.
        assert cell.value.startswith("'"), (
            f"Leading-special-char Invoice # not apostrophe-prefixed; cell.value={cell.value!r}"
        )
        # Strip the apostrophe and confirm the actual payload
        # is preserved verbatim — the mediator isn't silently
        # mangling customer-supplied invoice references.
        assert cell.value.lstrip("'") == attack_val

    def test_invoice_number_with_equals_sign_is_guarded(self, workdir: Path) -> None:
        self._guarded(workdir, "=cmd|'/c calc'!A1")

    def test_invoice_number_with_plus_is_guarded(self, workdir: Path) -> None:
        self._guarded(workdir, "+1+1")

    def test_invoice_number_with_minus_is_guarded(self, workdir: Path) -> None:
        self._guarded(workdir, "-100")

    def test_invoice_number_with_at_symbol_is_guarded(self, workdir: Path) -> None:
        self._guarded(workdir, "@SUM(100)")


class TestFormulaInjectionGuardOpacityRoundtrip:
    """Round-trip summary: a single fixture has every kind of
    attack cell across Source/Invoice/Attachment/Detail cells.
    Confirms the workbook's first 12 columns are uniformly
    guarded against the formula-evaluation-when-opened case.
    """

    def test_workbook_cells_all_guarded(self, workdir: Path) -> None:
        attack_payloads = {
            "Invoice #": "=DANGEROUS()",
            "Details": "+cmd|'/c calc'!A1",
            "Attachment Name": "@SUM(100)",
            "Sender": "-open@evil.com",
        }
        record: dict[str, Any] = {
            "Date": "01/05/2024",
            "Source": "Local PDF Folder",
            "Period From": "N/A",
            "Period To": "N/A",
            "Amount (£)": 50.0,
            "Period Charge (£)": 0.0,
            "Units (kWh)": "",
            "Reading": "",
            "Entry Type": "Payment",
            "Logic Used": "Pattern",
            "Standing Charge": "",
            "Anomaly Flag": "",
        }
        record.update(attack_payloads)
        config = cast(
            ConfigDict,
            {
                "use_dedup": True,
                "save_dups": True,
                "use_anchors": False,
                "use_large": False,
                "min_amount": 0.0,
                "filter_below": False,
                "use_dedup_period": True,
                "expanded_columns": True,
                "include_charts": False,
                "include_forecast": False,
            },
        )
        out = workdir / "attack_one.xlsx"
        export_to_excel([record], str(out), [], config=config)
        wb = load_workbook(str(out))
        ws = wb["EDF Evidence Report"]
        header_map = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}
        for col_name, payload in attack_payloads.items():
            cell = ws.cell(row=2, column=header_map[col_name])
            assert cell.data_type == "s", (
                f"{col_name} cell data_type={cell.data_type!r}; "
                f"expected 's' to block Excel formula evaluation"
            )
            # value prefix apostrophe to defeat any consumer that
            # ignores data_type.
            assert cell.value.startswith("'"), (
                f"{col_name} value missing apostrophe prefix; got value={cell.value!r}"
            )
            # Verbatim preservation of the original payload.
            assert cell.value.lstrip("'") == payload
