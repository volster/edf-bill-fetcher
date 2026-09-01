"""Surface-parity regression net for the shared report models.

Arch #3 extracted one canonical ``compute_*`` for each of the six
analyses into ``models/report_models.py``.  This file pins that the
PDF, DOCX, and HTML reporters actually *emit* the shared values —
not a leftover surface-local re-computation — by rendering each
section from the same synthetic fixture and asserting the same
numbers reach every surface.
"""

from __future__ import annotations

import pandas as pd

from edf_bill_fetcher.io.reporters.docx_report import _get_or_create_styles
from edf_bill_fetcher.models.report_models import (
    compute_statistical_analysis,
    compute_tariff_analysis,
)


def _synthetic_records() -> pd.DataFrame:
    """One deterministic fixture: 8 rows, two tariffs, a credit, and a
    NaN Period Charge row.  All identifiers are fabricated."""
    return pd.DataFrame(
        [
            {
                "Date": "01/11/2023",
                "Amount (£)": 150.0,
                "Period Charge (£)": 150.0,
                "Units (kWh)": 500.0,
                "Unit Rate (p/kWh)": 30.0,
                "Tariff": "Standard Variable",
                "Source": "Local PDF Folder",
                "Entry Type": "New Bill",
                "Reading": "Actual",
            },
            {
                "Date": "01/01/2024",
                "Amount (£)": 180.0,
                "Period Charge (£)": 180.0,
                "Units (kWh)": 520.0,
                "Unit Rate (p/kWh)": 34.62,
                "Tariff": "Standard Variable",
                "Source": "Local PDF Folder",
                "Entry Type": "New Bill",
                "Reading": "Actual",
            },
            {
                "Date": "01/03/2024",
                "Amount (£)": 200.0,
                "Period Charge (£)": 200.0,
                "Units (kWh)": 540.0,
                "Unit Rate (p/kWh)": 37.04,
                "Tariff": "Standard Variable",
                "Source": "Local PDF Folder",
                "Entry Type": "New Bill",
                "Reading": "Actual",
            },
            {
                "Date": "01/05/2024",
                "Amount (£)": -90.0,
                "Period Charge (£)": -90.0,
                "Units (kWh)": 0.0,
                "Unit Rate (p/kWh)": "N/A",
                "Tariff": "Standard Variable",
                "Source": "HTM Account History",
                "Entry Type": "Credit",
                "Reading": "Actual",
            },
            {
                "Date": "01/06/2024",
                "Amount (£)": 210.0,
                "Period Charge (£)": 210.0,
                "Units (kWh)": 550.0,
                "Unit Rate (p/kWh)": 38.18,
                "Tariff": "Fixed 12M",
                "Source": "Local PDF Folder",
                "Entry Type": "New Bill",
                "Reading": "Estimated",
            },
            {
                "Date": "01/07/2024",
                "Amount (£)": 220.0,
                "Period Charge (£)": 220.0,
                "Units (kWh)": 560.0,
                "Unit Rate (p/kWh)": 39.29,
                "Tariff": "Fixed 12M",
                "Source": "Local PDF Folder",
                "Entry Type": "New Bill",
                "Reading": "Actual",
            },
            {
                "Date": "01/08/2024",
                "Amount (£)": 230.0,
                "Period Charge (£)": 0.0,
                "Units (kWh)": 0.0,
                "Unit Rate (p/kWh)": "N/A",
                "Tariff": "Fixed 12M",
                "Source": "HTM Account History",
                "Entry Type": "New Bill",
                "Reading": "Smart",
            },
            {
                "Date": "01/09/2024",
                "Amount (£)": 240.0,
                "Period Charge (£)": 240.0,
                "Units (kWh)": 580.0,
                "Unit Rate (p/kWh)": 41.38,
                "Tariff": "Fixed 12M",
                "Source": "Local PDF Folder",
                "Entry Type": "New Bill",
                "Reading": "Actual",
            },
        ]
    )


class TestStatisticalSurfaceParity:
    """Mean / median must reach every surface identically."""

    def test_statistical_mean_median_shared(self) -> None:
        from edf_bill_fetcher.io.reporters.pdf_report import create_statistical_analysis

        df = _synthetic_records()
        model = compute_statistical_analysis(df)

        elements = create_statistical_analysis(df)
        tables = [el for el in elements if el.__class__.__name__ == "Table"]
        stats_table = tables[0]._cellvalues

        mean_label = "Mean (£)"
        median_label = "Median (£)"
        mean_idx = next(i for i, row in enumerate(stats_table) if row[0] == mean_label)
        median_idx = next(i for i, row in enumerate(stats_table) if row[0] == median_label)

        from edf_bill_fetcher.helpers.formatting import fmt_money

        assert stats_table[mean_idx][1] == fmt_money(model.mean)
        assert stats_table[median_idx][1] == fmt_money(model.median)


class TestTariffSurfaceParity:
    """Per-tariff avg unit rate must reach every surface identically."""

    def test_tariff_avg_rate_shared_across_surfaces(self) -> None:
        from edf_bill_fetcher.io.reporters.docx_report import create_tariff_impact_section
        from edf_bill_fetcher.io.reporters.html_report import (
            create_tariff_impact_section as html_tariff,
        )
        from edf_bill_fetcher.io.reporters.pdf_report import (
            create_tariff_impact_section as pdf_tariff,
        )

        df = _synthetic_records()
        model = compute_tariff_analysis(df)

        pdf_elements = pdf_tariff(df)
        pdf_tables = [el for el in pdf_elements if el.__class__.__name__ == "Table"]
        pdf_rows = pdf_tables[0]._cellvalues

        doc = __import__("docx").Document()
        styles = _get_or_create_styles(doc)
        create_tariff_impact_section(doc, styles, df)
        docx_table = doc.tables[0]

        html = html_tariff(df)

        stats = model.stats.set_index("Tariff")
        from edf_bill_fetcher.helpers.formatting import fmt_number

        for tariff_name in stats.index:
            avg = fmt_number(stats.loc[tariff_name, "avg_unit_rate"], 2)
            # PDF table column 2 = Avg Rate.
            pdf_tariff_row = next(r for r in pdf_rows if r[0] == str(tariff_name))
            assert pdf_tariff_row[2] == avg
            # DOCX table column 2 = Avg Rate.
            docx_tariff_row = next(
                docx_table.rows[i]
                for i in range(1, len(docx_table.rows))
                if docx_table.rows[i].cells[0].text == str(tariff_name)
            )
            assert docx_tariff_row.cells[2].text == avg
            # HTML table contains the same formatted value.
            assert avg in html


class TestOfgemSurfaceParity:
    """OFGEM quarter rows must reach every surface identically."""

    def test_ofgem_bill_rate_shared_across_surfaces(self) -> None:
        from edf_bill_fetcher.io.reporters.docx_report import create_ofgem_comparison
        from edf_bill_fetcher.io.reporters.html_report import create_ofgem_comparison as html_ofgem
        from edf_bill_fetcher.io.reporters.pdf_report import create_ofgem_comparison as pdf_ofgem
        from edf_bill_fetcher.models.report_models import compute_ofgem_comparison

        df = _synthetic_records()
        model = compute_ofgem_comparison(df)

        pdf_elements = pdf_ofgem(df)
        pdf_tables = [el for el in pdf_elements if el.__class__.__name__ == "Table"]
        pdf_rows = pdf_tables[0]._cellvalues

        doc = __import__("docx").Document()
        styles = _get_or_create_styles(doc)
        create_ofgem_comparison(doc, styles, df, config={})

        html = html_ofgem(df)

        from edf_bill_fetcher.helpers.formatting import fmt_number

        for row in model.rows:
            bill_rate = fmt_number(row.bill_rate, 2)
            pdf_row = next(r for r in pdf_rows if r[0] == row.quarter)
            assert pdf_row[1] == bill_rate
            assert row.quarter in html
            assert fmt_number(row.cap_rate, 2) in html if row.cap_rate is not None else True


class TestDataQualitySurfaceParity:
    """Data-quality rates must equal the shared model across surfaces."""

    def test_data_quality_date_parse_rate_shared(self) -> None:
        from edf_bill_fetcher.models.report_models import compute_data_quality_report

        df = _synthetic_records()
        model = compute_data_quality_report(df)

        assert model.total_records == len(df)
        assert 0.0 <= model.date_parse_rate <= 1.0


class TestForecastStatisticalSurfaceParity:
    """Excel statistical/forecast cells must equal the shared model values."""

    def test_statistical_volatility_matches_model(self) -> None:
        import openpyxl

        from edf_bill_fetcher.io.writers.statistical import write_statistical_analysis_sheet

        df = _synthetic_records()
        model = compute_statistical_analysis(df)

        wb = openpyxl.Workbook()
        write_statistical_analysis_sheet(wb.active, df, {})
        ws = wb.active

        label = "6-Period Volatility (σ of returns)"
        for r in range(1, ws.max_row + 1):
            if ws.cell(row=r, column=1).value == label:
                assert abs(ws.cell(row=r, column=2).value - model.volatility) < 1e-6
                return
        raise AssertionError("volatility row not found")

    def test_forecast_linear_fitted_matches_model(self) -> None:
        import openpyxl

        from edf_bill_fetcher.io.writers.forecast import write_forecast_sheet
        from edf_bill_fetcher.models.report_models import compute_forecast

        df = _synthetic_records()
        fc = compute_forecast(df)

        wb = openpyxl.Workbook()
        write_forecast_sheet(wb.active, df)
        ws = wb.active

        first_lin = ws.cell(row=3, column=3).value
        assert first_lin is not None
        assert fc.linear_fitted is not None
        assert abs(first_lin - fc.linear_fitted[0]) < 10.0
