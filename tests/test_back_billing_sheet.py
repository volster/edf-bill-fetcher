from __future__ import annotations

import pandas as pd
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from edf_bill_fetcher.io.adapters.pdf import legal_context
from edf_bill_fetcher.io.writers.back_billing import write_back_billing_sheet
from edf_bill_fetcher.processors.detection import detect_back_billing


def _sample_df() -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "Invoice #": "T-6715690",
                "Date": "09 Aug 2023",
                "Period From": "04 Apr 2022",
                "Period To": "26 Jul 2022",
                "Amount (£)": 4401.07,
                "Cancel/Rebill Admitted": True,
                "Attachment Name": "671078701920_060264189544_20230809.pdf",
            },
            {
                "Invoice #": "REG-0001",
                "Date": "01 Jan 2024",
                "Period From": "01 Dec 2023",
                "Period To": "31 Dec 2023",
                "Amount (£)": 100.00,
                "Cancel/Rebill Admitted": False,
                "Attachment Name": "reg.pdf",
            },
        ]
    )


def _open_ws(title: str = "Back-billing Analysis") -> Worksheet:
    wb = Workbook()
    ws = wb.active
    ws.title = title
    return ws


def test_write_back_billing_sheet_renders_legal_context_banner() -> None:
    ws = _open_ws()
    bb = detect_back_billing(_sample_df())
    write_back_billing_sheet(ws, bb, account="1234567890")
    # Row 1: title banner with account
    a1 = ws.cell(row=1, column=1).value
    assert isinstance(a1, str)
    assert "BACK-BILLING" in a1.upper()
    assert "1234567890" in a1
    # Row 2: 'LEGAL CONTEXT' label
    a2 = ws.cell(row=2, column=1).value
    assert isinstance(a2, str)
    assert "LEGAL CONTEXT" in a2.upper()
    # Row 3 contains the legal_context() body text
    a3 = ws.cell(row=3, column=1).value
    assert isinstance(a3, str)
    assert legal_context().splitlines()[0] in a3


def test_write_back_billing_sheet_writes_table_headers() -> None:
    ws = _open_ws()
    bb = detect_back_billing(_sample_df())
    write_back_billing_sheet(ws, bb, account="A1")
    # Per spec, row 7 = table header row. 18 columns (Status / Superseded By /
    # Partial Overlap for live rows, plus View Superseded at col 18).
    headers = [ws.cell(row=7, column=c).value for c in range(1, 19)]
    expected = [
        "Invoice #",
        "Bill Date",
        "Period From",
        "Period To",
        "Days Billed",
        "Period Charge (£)",
        "Value Source",
        "12-Month Limit (days)",
        "Excess Days",
        "Unlawful Charge (£)",
        "Cancel/Rebill Disclosed",
        "Reason Assessment",
        "Open PDF",
        "View on Evidence Report",
        "Status",
        "Superseded By",
        "Partial Overlap",
        "View Superseded",
    ]
    assert headers == expected


def test_write_back_billing_sheet_one_row_per_backbilled_invoice() -> None:
    ws = _open_ws()
    bb = detect_back_billing(_sample_df())
    write_back_billing_sheet(ws, bb, account="A1")
    # Spec: rows 8+ are data rows. Sample has exactly 1 back-billed invoice.
    a8 = ws.cell(row=8, column=1).value
    assert a8 == "T-6715690"
    # Row 9 carries the single union trailing total (sample has 1 back-bill).
    a9 = ws.cell(row=9, column=1).value
    assert isinstance(a9, str)
    assert "TOTAL UNLAWFUL CHARGES" in a9
    assert "UNION" in a9
    # Nothing beyond the total row.
    assert ws.cell(row=10, column=1).value in (None, "")


def test_write_back_billing_sheet_total_charges_footer() -> None:
    ws = _open_ws()
    bb = detect_back_billing(_sample_df())
    write_back_billing_sheet(ws, bb, account="A1")
    # Trailing row somewhere below row 8 carries the union total label and value.
    found = False
    for r in range(9, 15):
        v = ws.cell(row=r, column=1).value
        if isinstance(v, str) and "TOTAL UNLAWFUL CHARGES" in v and "UNION" in v:
            # The same row's col 5 (or thereabouts) carries the sum.
            sum_cell = ws.cell(row=r, column=6).value
            assert sum_cell == 4401.07
            found = True
            break
    assert found, "TOTAL UNLAWFUL CHARGES union footer row missing"


def test_write_back_billing_sheet_empty_df_still_renders_header_and_legal_context() -> None:
    ws = _open_ws()
    empty = pd.DataFrame(
        columns=[
            "Invoice #",
            "Bill Date",
            "Period From",
            "Period To",
            "Days Billed",
            "Period Charge (£)",
            "Value Source",
            "12-Month Limit (days)",
            "Excess Days",
            "Unlawful Charge (£)",
            "Cancel/Rebill Admitted",
            "Reason Assessment",
        ]
    )
    write_back_billing_sheet(ws, empty, account="A1")
    # Legal context still rendered.
    a3 = ws.cell(row=3, column=1).value
    assert isinstance(a3, str)
    assert "back-billing" in a3.lower()
    # Table headers still rendered (18 columns incl. View Superseded).
    headers = [ws.cell(row=7, column=c).value for c in range(1, 19)]
    assert headers[0] == "Invoice #"
    assert "Status" in headers
    assert "View Superseded" in headers
    # No data rows.
    assert ws.cell(row=8, column=1).value in (None, "")


def test_write_back_billing_sheet_admitted_cell_value_uses_phrase_label() -> None:
    ws = _open_ws()
    bb = detect_back_billing(_sample_df())
    write_back_billing_sheet(ws, bb, account="A1")
    # Admit column (col 11) on row 8 must say 'Admitted phrase' for our
    # sample (the cover-page admit fired). Col 10 is now Unlawful Charge.
    v = ws.cell(row=8, column=11).value
    assert v == "Admitted phrase"


def _two_row_bb() -> pd.DataFrame:
    """Two back-billing rows with synthetic invoice IDs A and B."""
    return pd.DataFrame(
        [
            {
                "Invoice #": "A",
                "Bill Date": "2021-06-01",
                "Period From": "2020-01-01",
                "Period To": "2021-06-01",
                "Days Billed": 517,
                "Period Charge (£)": 500.0,
                "Value Source": "Period Charge",
                "12-Month Limit (days)": 365,
                "Excess Days": 152,
                "Unlawful Charge (£)": round(500.0 * (152 / 517), 2),
                "Cancel/Rebill Admitted": False,
                "Reason Assessment": "test",
            },
            {
                "Invoice #": "B",
                "Bill Date": "2021-12-01",
                "Period From": "2020-06-01",
                "Period To": "2021-12-01",
                "Days Billed": 549,
                "Period Charge (£)": 300.0,
                "Value Source": "Period Charge",
                "12-Month Limit (days)": 365,
                "Excess Days": 184,
                "Unlawful Charge (£)": round(300.0 * (184 / 549), 2),
                "Cancel/Rebill Admitted": False,
                "Reason Assessment": "test",
            },
        ]
    )


def test_write_back_billing_sheet_status_columns() -> None:
    ws = _open_ws()
    bb = _two_row_bb()
    domination_map: dict[str, tuple[str, bool]] = {"B": ("A", False)}
    write_back_billing_sheet(ws, bb, domination_map=domination_map)

    # Row 7 is the header row.
    header_row = [cell.value for cell in ws[7]]
    assert "Status" in header_row
    assert "Superseded By" in header_row
    assert "Partial Overlap" in header_row
    assert "Value Source" in header_row

    status_col = header_row.index("Status") + 1
    superseded_by_col = header_row.index("Superseded By") + 1
    partial_overlap_col = header_row.index("Partial Overlap") + 1
    inv_col = header_row.index("Invoice #") + 1

    rendered = []
    for row_idx in range(8, ws.max_row + 1):
        inv_num = ws.cell(row=row_idx, column=inv_col).value
        if inv_num not in ("A", "B"):
            continue  # trailing total row carries a label, not an invoice id
        rendered.append(inv_num)
        assert ws.cell(row=row_idx, column=status_col).value == "Live"
        assert ws.cell(row=row_idx, column=superseded_by_col).value in (None, "")
        assert ws.cell(row=row_idx, column=partial_overlap_col).value in (None, "")
    # Superseded invoice B is NOT rendered on this sheet (it moves to the
    # reconciliation view); only the live survivor A appears.
    assert rendered == ["A"]


def test_write_back_billing_sheet_view_superseded_cell_on_survivor_rows() -> None:
    """A live row that is a survivor in domination_map gets a blue-underline
    'View superseded' cell (col 18); other live rows leave it blank."""
    ws = _open_ws()
    bb = _three_row_bb()
    domination_map: dict[str, tuple[str, bool]] = {"B": ("A", False)}
    write_back_billing_sheet(ws, bb, domination_map=domination_map)

    header_row = [cell.value for cell in ws[7]]
    view_col = header_row.index("View Superseded") + 1
    inv_col = header_row.index("Invoice #") + 1
    for row_idx in range(8, ws.max_row + 1):
        inv = ws.cell(row=row_idx, column=inv_col).value
        cell = ws.cell(row=row_idx, column=view_col)
        if inv == "A":
            assert cell.value == "View superseded"
            assert cell.font.underline == "single"
            assert cell.font.color is not None
            assert cell.font.color.rgb.endswith("0563C1")
        elif inv == "C":
            assert cell.value in (None, "")


def test_write_back_billing_sheet_total_excludes_superseded() -> None:
    ws = _open_ws()
    bb = _two_row_bb()
    domination_map: dict[str, tuple[str, bool]] = {"B": ("A", False)}
    write_back_billing_sheet(ws, bb, domination_map=domination_map)

    header_row = [cell.value for cell in ws[7]]
    period_charge_col = header_row.index("Period Charge (£)") + 1

    total_row_idx = None
    for row_idx in range(8, ws.max_row + 1):
        v = ws.cell(row=row_idx, column=1).value
        if v and "TOTAL UNLAWFUL CHARGES" in str(v):
            total_row_idx = row_idx
            break
    assert total_row_idx is not None
    total_value = ws.cell(row=total_row_idx, column=period_charge_col).value
    assert total_value == 500.0  # only A (Live), not B (Superseded)


def test_write_back_billing_sheet_no_domination_map_all_live() -> None:
    """Without a domination_map, every row is Live and the total sums all rows."""
    ws = _open_ws()
    bb = _two_row_bb()
    write_back_billing_sheet(ws, bb)

    header_row = [cell.value for cell in ws[7]]
    status_col = header_row.index("Status") + 1
    inv_col = header_row.index("Invoice #") + 1
    for row_idx in range(8, ws.max_row + 1):
        inv_num = ws.cell(row=row_idx, column=inv_col).value
        if inv_num in ("A", "B"):
            assert ws.cell(row=row_idx, column=status_col).value == "Live"
            assert ws.row_dimensions[row_idx].outline_level == 0


def test_trailing_union_total_row_written() -> None:
    from edf_bill_fetcher.io.writers.back_billing import write_back_billing_sheet
    from edf_bill_fetcher.processors.detection import detect_back_billing

    ws = _open_ws()
    df = _sample_df()
    # Give both rows real sub-period slices so the union is non-zero.
    df = df.copy()
    df["Sub Periods"] = ""
    df.loc[0, "Sub Periods"] = (
        "02/10/2020|24/03/2021|19743.0|16.42|3241.8; "
        "25/03/2021|06/04/2021|1454.0|16.42|238.75; "
        "07/04/2021|31/03/2022|37184.0|16.42|6105.61; "
        "01/04/2022|12/05/2022|3736.0|52.00|1942.72; "
        "13/05/2022|31/03/2023|30675.0|52.00|15951.0; "
        "01/04/2023|09/08/2023|10607.0|45.92|4870.73"
    )
    bb = detect_back_billing(df)
    write_back_billing_sheet(ws, bb, evidence_df=df)
    labels = [ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)]
    assert any(label is not None and "UNION" in str(label) for label in labels)


def _three_row_bb() -> pd.DataFrame:
    """Three back-billing rows with synthetic invoice IDs A, B, C.

    B is superseded by A (mirrors the two-row helper); C is a plain live row.
    """
    rows = _two_row_bb().to_dict("records")
    rows.append(
        {
            "Invoice #": "C",
            "Bill Date": "2022-01-15",
            "Period From": "2021-01-01",
            "Period To": "2022-01-01",
            "Days Billed": 365,
            "Period Charge (£)": 250.0,
            "Value Source": "Period Charge",
            "12-Month Limit (days)": 365,
            "Excess Days": 30,
            "Unlawful Charge (£)": round(250.0 * (30 / 365), 2),
            "Cancel/Rebill Admitted": False,
            "Reason Assessment": "test",
        }
    )
    return pd.DataFrame(rows)


def test_write_back_billing_sheet_live_rows_only() -> None:
    ws = _open_ws()
    bb = _three_row_bb()  # A, B(superseded by A), C
    domination_map = {"B": ("A", False)}
    write_back_billing_sheet(ws, bb, domination_map=domination_map)
    inv_col = [c.value for c in ws[7]].index("Invoice #") + 1
    invs = [ws.cell(row=r, column=inv_col).value for r in range(8, ws.max_row + 1)]
    invs = [i for i in invs if i]
    assert "B" not in invs
    assert "A" in invs
    assert "C" in invs


def test_view_superseded_link_wired() -> None:
    ws = _open_ws()
    bb = _three_row_bb()
    domination_map = {"B": ("A", False)}
    write_back_billing_sheet(
        ws,
        bb,
        domination_map=domination_map,
        view_superseded_row={"A": 8},
    )
    hdrs = [c.value for c in ws[7]]
    col = hdrs.index("View Superseded") + 1
    a_row = next(r for r in range(8, ws.max_row + 1) if ws.cell(row=r, column=1).value == "A")
    assert ws.cell(row=a_row, column=col).hyperlink is not None
    assert "Superseded Reconciliation" in ws.cell(row=a_row, column=col).hyperlink.location


def test_write_back_billing_sheet_single_union_total() -> None:
    from edf_bill_fetcher.processors.detection import compute_unlawful_union_total

    ws = _open_ws()
    bb = _three_row_bb()
    domination_map = {"B": ("A", False)}
    write_back_billing_sheet(ws, bb, domination_map=domination_map)
    labels = [ws.cell(row=r, column=1).value for r in range(8, ws.max_row + 1)]
    union_labels = [label for label in labels if label and "UNION" in str(label)]
    surviving_labels = [label for label in labels if label and "SURVIVING" in str(label)]
    assert len(union_labels) == 1
    assert len(surviving_labels) == 0
    # The single trailing row's Unlawful column carries the union over live rows.
    bb_live = bb[~bb["Invoice #"].astype(str).isin(domination_map)]
    row_idx = labels.index(union_labels[0]) + 8
    assert ws.cell(row=row_idx, column=10).value == compute_unlawful_union_total(bb_live)


def test_write_back_billing_sheet_all_superseded_still_writes_zero_union_total() -> None:
    """When every row is superseded the total row is still written as £0.00."""
    ws = _open_ws()
    bb = _two_row_bb()
    domination_map: dict[str, tuple[str, bool]] = {"B": ("A", False), "A": ("X", False)}
    write_back_billing_sheet(ws, bb, domination_map=domination_map)
    labels = [ws.cell(row=r, column=1).value for r in range(8, ws.max_row + 1)]
    union_rows = [r for r, v in enumerate(labels, start=8) if v and "UNION" in str(v)]
    assert len(union_rows) == 1
    assert ws.cell(row=union_rows[0], column=6).value == 0.0
    assert ws.cell(row=union_rows[0], column=10).value == 0.0
