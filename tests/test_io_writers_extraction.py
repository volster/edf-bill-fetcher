"""Tests for the planned io/writers submodule extraction.

Each test targets a planned writer submodule under
``edf_bill_fetcher.io.writers``.  All tests are RED at Phase 0
because those submodules do not yet exist.
"""

from __future__ import annotations

import pandas as pd
from openpyxl import Workbook


def _fixture_df() -> pd.DataFrame:
    """Minimal fixture DataFrame with all columns the writers expect."""
    return pd.DataFrame(
        {
            "Date": ["01 Jan 2024", "15 Feb 2024"],
            "Invoice #": ["T-001", "T-002"],
            "Amount (£)": [100.0, 200.0],
            "Period From": ["01 Jan 2024", "01 Feb 2024"],
            "Period To": ["31 Jan 2024", "29 Feb 2024"],
            "Period Charge (£)": [100.0, 200.0],
            "Units (kWh)": [100, 200],
            "Entry Type": ["New Bill", "New Bill"],
            "Source": ["PST", "PST"],
            "Sender": ["edf@example.com", "edf@example.com"],
            "Tariff": ["Standard", "Standard"],
            "Standing Chg (p/day)": [50.0, 50.0],
            "Unit Rate (p/kWh)": [24.0, 24.0],
            "Reading": [1000, 1200],
            "Details": ["Bill 1", "Bill 2"],
            "Attachment Name": ["bill1.pdf", "bill2.pdf"],
        }
    )


def test_evidence_writer_importable() -> None:
    from edf_bill_fetcher.io.writers.evidence import write_evidence_sheet

    wb = Workbook()
    ws = wb.active
    write_evidence_sheet(ws, _fixture_df())
    assert ws["A1"].value is not None, "Expected header in A1"


def test_statistical_writer_importable() -> None:
    from edf_bill_fetcher.io.writers.statistical import write_statistical_analysis_sheet

    wb = Workbook()
    ws = wb.active
    write_statistical_analysis_sheet(ws, _fixture_df(), {})
    assert ws["A1"].value is not None, "Expected header in A1"


def test_payment_writer_importable() -> None:
    from edf_bill_fetcher.io.writers.payment import write_payment_analysis_sheet

    wb = Workbook()
    ws = wb.active
    write_payment_analysis_sheet(ws, _fixture_df())
    assert ws["A1"].value is not None, "Expected header in A1"


def test_forecast_writer_importable() -> None:
    from edf_bill_fetcher.io.writers.forecast import write_forecast_sheet

    wb = Workbook()
    ws = wb.active
    write_forecast_sheet(ws, _fixture_df())
    assert ws["A1"].value is not None, "Expected header in A1"


def test_data_quality_writer_importable() -> None:
    from edf_bill_fetcher.io.writers.data_quality import write_data_quality_sheet

    wb = Workbook()
    ws = wb.active
    write_data_quality_sheet(ws, _fixture_df())
    assert ws["A1"].value is not None, "Expected header in A1"


def test_tariff_writer_importable() -> None:
    from edf_bill_fetcher.io.writers.tariff import write_tariff_analysis_sheet

    wb = Workbook()
    ws = wb.active
    write_tariff_analysis_sheet(ws, _fixture_df())
    assert ws["A1"].value is not None, "Expected header in A1"


def test_back_billing_writer_importable() -> None:
    from edf_bill_fetcher.io.writers.back_billing import write_back_billing_sheet

    wb = Workbook()
    ws = wb.active
    write_back_billing_sheet(ws, _fixture_df())
    assert ws["A1"].value is not None, "Expected header in A1"


def test_rebilling_writer_importable() -> None:
    from edf_bill_fetcher.io.writers.rebilling import write_rebilling_sheet

    wb = Workbook()
    ws = wb.active
    write_rebilling_sheet(ws, _fixture_df())
    assert ws["A1"].value is not None, "Expected header in A1"


def test_meter_writer_importable() -> None:
    from edf_bill_fetcher.io.writers.meter import write_meter_readings_sheet

    wb = Workbook()
    ws = wb.active
    write_meter_readings_sheet(ws, _fixture_df())
    assert ws["A1"].value is not None, "Expected header in A1"


def test_sap_writer_importable() -> None:
    from edf_bill_fetcher.io.writers.sap import write_sap_contract_history_sheet

    wb = Workbook()
    ws = wb.active
    write_sap_contract_history_sheet(ws, _fixture_df())
    assert ws["A1"].value is not None, "Expected header in A1"


def test_export_writer_importable(tmp_path: object) -> None:
    from edf_bill_fetcher.io.writers.export import export_to_excel

    df = _fixture_df()
    out = tmp_path / "export_writer.xlsx"  # type: ignore[operator]
    export_to_excel(df.to_dict(orient="records"), str(out), [], {})
    assert out.exists()  # export_to_excel writes to disk; success = file present


def test_evidence_sheet_has_sub_periods_column() -> None:
    from edf_bill_fetcher.io.writers.evidence import write_evidence_sheet

    df = _fixture_df()
    df["Sub Periods"] = ""
    df.loc[0, "Sub Periods"] = "02/10/2020|24/03/2021|19743.0|16.42|3241.8"
    ws = Workbook()
    write_evidence_sheet(ws.active, df)
    headers = [ws.active.cell(row=1, column=c).value for c in range(1, ws.active.max_column + 1)]
    assert "Sub Periods" in headers
