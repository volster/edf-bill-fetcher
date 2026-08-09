"""Tests for the bidirectional hotlink feature (Stream P4 / Task 7)."""

from __future__ import annotations

import pandas as pd
from openpyxl import Workbook

from edf_bill_fetcher.io.writers import write_rebilling_sheet
from edf_bill_fetcher.io.writers.back_billing import write_back_billing_sheet
from edf_bill_fetcher.processors.matching import build_evidence_index


def _ev_df() -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "Invoice #": "KI-31105244-0001-3",
                "Date": "14/05/2024",
                "Period From": "14/05/2024",
                "Period To": "30/06/2024",
                "Amount (£)": 1347.96,
                "Source": "Local PDF Folder",
            },
            {
                "Invoice #": "KI-31105244-0002-3",
                "Date": "01/07/2024",
                "Period From": "01/07/2024",
                "Period To": "31/07/2024",
                "Amount (£)": 841.36,
                "Source": "Local PDF Folder",
            },
            {
                "Invoice #": "N/A",
                "Date": "01/08/2024",
                "Period From": "01/08/2024",
                "Period To": "31/08/2024",
                "Amount (£)": 900.00,
                "Source": "Local PDF Folder",
            },
        ]
    )


def test_build_evidence_index_returns_inv_key() -> None:
    df = _ev_df()
    idx = build_evidence_index(df, header_row_offset=1)
    assert "inv:KI-31105244-0001-3" in idx
    # Evidence sheet header row = 1, first body row = 2.
    assert idx["inv:KI-31105244-0001-3"] == 2


def test_build_evidence_index_skips_na_invoice() -> None:
    df = _ev_df()
    idx = build_evidence_index(df, header_row_offset=1)
    assert "inv:N/A" not in idx
    # Fallback amt_days signature also produced for that row.
    assert "amt_days:900.00|30" in idx


def test_build_evidence_index_includes_amt_days_signature() -> None:
    df = _ev_df()
    idx = build_evidence_index(df, header_row_offset=1)
    assert idx["amt_days:1347.96|47"] == 2
    assert idx["amt_days:841.36|30"] == 3


def test_build_evidence_index_handles_empty_df() -> None:
    df = pd.DataFrame()
    assert build_evidence_index(df, header_row_offset=1) == {}  # type: ignore[arg-type]


def test_back_billing_view_on_evidence_report_column_present() -> None:
    bb = pd.DataFrame(
        [
            {
                "Invoice #": "KI-31105244-0001-3",
                "Bill Date": "01/01/2024",
                "Period From": pd.Timestamp("2022-01-01"),
                "Period To": pd.Timestamp("2024-01-01"),
                "Days Billed": 730,
                "Net Charge (£)": 1347.96,
                "12-Month Limit (days)": 365,
                "Excess Days": 365,
                "Cancel/Rebill Admitted": False,
                "Reason Assessment": "back-billing",
            }
        ]
    )
    ev = _ev_df()
    ev_idx = build_evidence_index(ev, header_row_offset=1)
    wb = Workbook()
    ws = wb.active
    write_back_billing_sheet(ws, bb, account="A-31105244", evidence_df=ev, evidence_index=ev_idx)
    # Header row=7. The hotlink column is "View on Evidence Report" at col 13
    # (post Task 3 + Task 4 the sheet has 16 columns: Open PDF is col 12,
    # View on Evidence Report is col 13, Status/Superseded By/Partial Overlap
    # are cols 14-16).
    hdr = ws.cell(row=7, column=13).value
    assert hdr == "View on Evidence Report"
    body = ws.cell(row=8, column=13)
    assert body.value == "→"
    assert body.hyperlink is not None
    # Excel report sheet name wrapped in single quotes with leading #.
    assert "EDF Evidence Report" in body.hyperlink.location
    assert body.hyperlink.location.endswith("!A2")  # row 2 in the synthetic ev


def test_back_billing_unmatched_emits_no_match() -> None:
    bb = pd.DataFrame(
        [
            {
                "Invoice #": "UNKNOWN-INVOICE-12345",
                "Bill Date": "01/01/2024",
                "Period From": pd.Timestamp("2022-01-01"),
                "Period To": pd.Timestamp("2024-01-01"),
                "Days Billed": 730,
                "Net Charge (£)": 1347.96,
                "12-Month Limit (days)": 365,
                "Excess Days": 365,
                "Cancel/Rebill Admitted": False,
                "Reason Assessment": "back-billing",
            }
        ]
    )
    ev = _ev_df()
    ev_idx = build_evidence_index(ev, header_row_offset=1)
    wb = Workbook()
    ws = wb.active
    write_back_billing_sheet(ws, bb, account="A-31105244", evidence_df=ev, evidence_index=ev_idx)
    body = ws.cell(row=8, column=13).value
    assert body == "No match"


def test_rebilling_view_on_evidence_report_column_present() -> None:
    reb = pd.DataFrame(
        [
            {
                "Killer Invoice": "KI-31105244-0002-3",
                "Killed Invoice": "KI-31105244-0001-3",
                "Killer Date": "01/07/2024",
                "Killed Date": "14/05/2024",
                "Period Overlap (days)": 10,
                "Jump-back (days)": 20,
                "Trigger Reason": "overlap",
            }
        ]
    )
    ev = _ev_df()
    ev_idx = build_evidence_index(ev, header_row_offset=1)
    wb = Workbook()
    ws = wb.active
    write_rebilling_sheet(ws, reb, account="A-31105244", evidence_df=ev, evidence_index=ev_idx)
    # Col 9 is the new View-on-Evidence-Report column.
    hdr = ws.cell(row=7, column=9).value
    assert hdr == "View on Evidence Report"
    body = ws.cell(row=8, column=9)
    assert body.value == "→"
    assert body.hyperlink is not None
    assert "EDF Evidence Report" in body.hyperlink.location
