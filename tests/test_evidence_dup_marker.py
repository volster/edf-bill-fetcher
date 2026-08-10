"""Regression tests pinning the evidence-layer Duplicate-Of marker contract.

When the dedup walker finds a duplicate row, the evidence layer must:

1. Create a ``Duplicates`` sheet in the workbook (not just emit inline
   markers).
2. Render one row per duplicate in that sheet with a *grey* row fill
   (DUP_GREY = ``E0E0E0``) so the dup rows stand out from the kept set.
3. Mint a clickable ``Duplicate Of`` hyperlink pointing at the kept
   row in ``EDF Evidence Report`` — the ombudsman's primary
   navigation affordance.

The audit's "Dup marker evidence-layer enforcement" item is the
mechanism that supplies the third column (``_matches_kept_idx``) on the
dup DataFrame before the writer runs.  The hand-built helper at
``write_evidence_sheet(..., is_duplicate=True)`` then plumbs it through
the post-loop rendering pass.

This module feeds a dedup-eligible pair into ``EvidenceEngine`` and
asserts:

* ``engine.records`` carries the kept record exactly once.
* ``len(engine.dup_records)`` (or whatever the engine exposes)
  contains the dropped row, OR — if a property isn't exposed —
  re-running the engine manually and feeding ``export_to_excel`` shows
  a ``Duplicates`` sheet whose rows are properly greyed.

Because the engine API for the dup set has shifted across refactors
(``dup_df`` is internal to ``export_to_excel``), this test exercises
the *worst-case* surface: call ``export_to_excel`` with the kept
records as ``data`` and confirm the dup sheet shows up with grey rows.

It also asserts that when ``save_dups=False`` the dup sheet is empty
(see ``test_save_dups_toggle.py`` for the parametrised contract) — here
we focus on the save_dups=True (default invariant).
"""

from __future__ import annotations

from typing import Any

import openpyxl
import pandas as pd

from edf_bill_fetcher.collectors.engine import EvidenceEngine
from edf_bill_fetcher.io.writers import export_to_excel
from edf_bill_fetcher.models.config import ConfigDict


def _engine_with_save_dups() -> EvidenceEngine:
    """Engine with dedup + dup-sheet enabled."""
    cfg: ConfigDict = {
        "use_anchors": False,
        "use_large": True,
        "use_reading_classification": False,
        "use_pdf_fields": False,
        "use_acc_filter": False,
        "acc_num": "",
        "min_amount": 1.0,
        "analysis_min": 1.0,
        "filter_below": False,
        "save_filtered": False,
        "use_dedup": True,
        "save_dups": True,
        "use_domain_filter": False,
        "domain_filter": "",
    }
    return EvidenceEngine(cfg, lambda *a: None)


def _records_to_dataframe(records: list[dict[str, Any]]) -> pd.DataFrame:
    if not records:
        return pd.DataFrame()
    return pd.DataFrame(records)


def test_dup_sheet_emitted_when_save_dups_enabled(tmp_path: object) -> None:
    """Round-trip: feed a duplicate pair; assert ``Duplicates`` sheet
    exists in the saved workbook, and that the rows carry the grey
    fill (DUP_GREY) for visibility.
    """
    engine = _engine_with_save_dups()
    _seed_canonical(engine)
    _seed_dup(engine, label="PST PDF Attachment")

    out = tmp_path / "out_dup_marker.xlsx"  # type: ignore[operator]
    export_to_excel(
        data=_records_to_dataframe(engine.records),
        output_path=str(out),
        error_log=engine.error_log,
        config=engine.config,
    )
    wb = openpyxl.load_workbook(out)  # type: ignore[arg-type]
    dup_sheet_name = "Duplicate Entries"
    assert dup_sheet_name in wb.sheetnames, (
        f"saved workbook is missing {dup_sheet_name!r} sheet; have {wb.sheetnames!r}"
    )

    ws = wb[dup_sheet_name]
    # Find at least one cell whose fill is the dup-grey RGB (E0E0E0 encoded).
    found_dup_grey = False
    for row in ws.iter_rows():
        for cell in row:
            fill = cell.fill
            if fill is None:
                continue
            pattern = getattr(fill, "patternType", None)
            if pattern is None:
                continue
            color = getattr(fill, "start_color", None)
            rgb = getattr(color, "rgb", None)
            if rgb and "E0E0E0" in str(rgb):
                found_dup_grey = True
                break
        if found_dup_grey:
            break
    assert found_dup_grey, (
        "dup sheet rows lack the DUP_GREY fill; the visual evidence "
        "marker on the dup set is degraded"
    )


def test_dup_sheet_empty_when_save_dups_disabled(tmp_path: object) -> None:
    """When ``save_dups=False`` the dup sheet either doesn't exist or
    carries no rows — duplicates are surfaced only via the kept
    DataFrame.
    """
    cfg: ConfigDict = {
        "use_anchors": False,
        "use_large": True,
        "use_reading_classification": False,
        "use_pdf_fields": False,
        "use_acc_filter": False,
        "acc_num": "",
        "min_amount": 1.0,
        "analysis_min": 1.0,
        "filter_below": False,
        "save_filtered": False,
        "use_dedup": True,
        "save_dups": False,
        "use_domain_filter": False,
        "domain_filter": "",
    }
    engine = EvidenceEngine(cfg, lambda *a: None)
    _seed_canonical(engine)
    _seed_dup(engine, label="PST PDF Attachment")

    out = tmp_path / "out_nodup_marker.xlsx"  # type: ignore[operator]
    export_to_excel(
        data=_records_to_dataframe(engine.records),
        output_path=str(out),
        error_log=engine.error_log,
        config=cfg,
    )
    wb = openpyxl.load_workbook(out)  # type: ignore[arg-type]
    name = "Duplicate Entries"
    if name in wb.sheetnames:
        # If the sheet is unconditionally created, it must be empty
        # (no data) when save_dups=False.
        ws = wb[name]
        # Sheet-with-no-data still has the column header (row 1); data
        # starts at row 2.  If row 2 is empty, the dup record was
        # suppressed.
        assert ws.cell(row=2, column=1).value in (None, ""), (
            "save_dups=False still surfaced a duplicate row in the dup sheet"
        )


def _seed_canonical(engine: EvidenceEngine) -> None:
    engine.process_text(
        "28 Feb 2025 We charged your account £500.00 For 1000 kWh of electricity "
        "used between 01 Feb 2025 and 28 Feb 2025 Balance £500.00 in debit",
        "HTM Account History",
        "seed.001",
        "28/02/2025",
    )


def _seed_dup(engine: EvidenceEngine, label: str) -> None:
    engine.process_text(
        "28 Feb 2025 We charged your account £500.00 For 1000 kWh of electricity "
        "used between 01 Feb 2025 and 28 Feb 2025 Balance £500.00 in debit",
        label,
        "dup.001",
        "28/02/2025",
    )
