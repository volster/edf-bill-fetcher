"""Tests for the reconciliation cross-source sheet writer (PR #4:
two-sheet summary + drill-down shape per spec §3.2)."""

from __future__ import annotations

import pandas as pd
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from edf_collector import write_reconciliation_sheet

_SAP_CONTRACT = [
    {
        "Contract From": "2024-05-14",
        "Contract To": "2024-06-30",
        "Product Code": "PRD_FXD24",
        "Product Description": "Fixed Online 2 Year",
        "Contract Reason": "New Sales",
        "Set Up By": "agent01",
        "Notes": "",
        "Cancelled Flag": "",
        "Source File": "Contract-and-Product-Change-History.pdf",
    },
    {
        "Contract From": "2024-07-01",
        "Contract To": "2024-07-31",
        "Product Code": "PRD_FREE",
        "Product Description": "Freedom",
        "Contract Reason": "Tariff Switch",
        "Set Up By": "agent02",
        "Notes": "",
        "Cancelled Flag": "",
        "Source File": "Contract-and-Product-Change-History.pdf",
    },
]

_INFERRED_CONTRACT = pd.DataFrame(
    [
        {
            "Contract From": "2024-05-14",
            "Contract To": "2024-06-30",
            "Product Code": "PRD_FXD24",
            "Product Description": "Fixed Online 2 Year",
            "Contract Reason": "Inferred from invoice body",
            "Set Up By": "N/A",
            "Notes": "",
            "Cancelled Flag": "",
            "Source File": "i-T12345.pdf",
        }
    ]
)

_SAP_METER = [
    {
        "Scheduled Read Date": "2024-05-14",
        "Meter Read Date": "2024-05-14",
        "Reading (kWh)": "1234.5000",
        "Read Type": "Periodic scheduled",
        "Read Source": "Metering System",
        "Read Status": "Posted",
        "Meter Read Reason": "Move-In",
        "Register": "01",
        "Source File": "Meter-Read-History.pdf",
    }
]

_INFERRED_METER = pd.DataFrame(
    [
        {
            "Scheduled Read Date": "2024-05-14",
            "Meter Read Date": "2024-05-14",
            "Reading (kWh)": "1234.5000",
            "Read Type": "A",
            "Read Source": "Customer",
            "Read Status": "Posted",
            "Meter Read Reason": "Move-In",
            "Register": "01",
        }
    ]
)

_SAP_FINANCIAL = [
    {
        "Document No.": "9000012345",
        "Item": "001",
        "Document Date": "2024-05-14",
        "Posting Date": "2024-05-14",
        "Net Due Date": "2024-05-21",
        "Main Transaction": "Credit Memo",
        "Sub Transaction": "Reversal",
        "Transaction Text": "Reversal Inv T12345",
        "Amount": "1347.96",
        "Clearing Status": "Not Cleared",
        "Clearing Document": "",
        "Clearing Date": "",
        "Clearing Reason": "",
        "Document Type": "CM",
        "Document Type Description": "Credit Memo",
        "Source File": "Financial-Transactions.pdf",
    }
]

_EVIDENCE_DF = pd.DataFrame(
    [
        {
            "Date": "14/05/2024",
            "Invoice #": "T12345",
            "Period From": "14/05/2024",
            "Period To": "30/06/2024",
            "Amount (£)": 1347.96,
            "Entry Type": "Charge",
            "Logic Used": "New Invoice Format",
        },
    ]
)


def _build_two_sheets(account: str = "") -> tuple[Workbook, Worksheet, Worksheet]:
    """Construct a fresh wb with both Reconciliation sheets."""
    wb = Workbook()
    ws_summary = wb.create_sheet(title="Reconciliation")
    ws_detail = wb.create_sheet(title="Reconciliation Drill-down")
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    write_reconciliation_sheet(
        ws_summary,
        ws_detail,
        _SAP_CONTRACT,
        _INFERRED_CONTRACT,
        _SAP_METER,
        _INFERRED_METER,
        _SAP_FINANCIAL,
        _EVIDENCE_DF,
        account=account,
    )
    return wb, ws_summary, ws_detail


def test_summary_sheet_has_three_entity_rows() -> None:
    """Spec §3.2: 3 entity rows (Contract / Meter Read / Financial)
    at rows 4-6, each with a distinct entity name."""
    _, summ, _ = _build_two_sheets()
    names = {summ.cell(row=r, column=1).value for r in (4, 5, 6)}
    assert names == {"Contract", "Meter Read", "Financial"}, names


def test_summary_sheet_each_row_links_to_detail_section() -> None:
    """Spec §3.2: col 8 (Drill down) on each summary row is a
    hyperlink pointing at the Reconciliation Drill-down sheet."""
    _, summ, _ = _build_two_sheets()
    for r in (4, 5, 6):
        cell = summ.cell(row=r, column=8)
        assert cell.hyperlink is not None, f"row {r} missing hyperlink"
        loc = cell.hyperlink.location or ""
        assert "Reconciliation Drill-down" in loc, f"row {r}: {loc}"


def test_detail_sheet_only_contains_unmatched_rows() -> None:
    """Spec §3.2: the detail sheet emits only 'Missing in Inferred',
    'Missing in SAP', and 'Discrepancy' rows — no 'Matched' rows."""
    _, _, detail = _build_two_sheets()
    for r in range(1, detail.max_row + 1):
        v = detail.cell(row=r, column=1).value
        if v is None:
            continue
        assert v != "Matched", f"row {r} on detail sheet has Matched status"


def test_summary_sheet_verdict_text_present() -> None:
    """Each summary entity row carries a plain-English verdict in col 7."""
    _, summ, _ = _build_two_sheets()
    for r in (4, 5, 6):
        verdict = summ.cell(row=r, column=7).value
        assert isinstance(verdict, str), f"row {r} verdict is not a string: {verdict!r}"
        assert len(verdict) > 10, f"row {r} verdict too short ({verdict!r})"


def test_eater_typo_fixed_in_subtitle() -> None:
    """The subtitle must read 'evidence' not 'eater' (spec §3.2 typo fix)."""
    wb = Workbook()
    ws_summ = wb.create_sheet(title="Reconciliation")
    ws_det = wb.create_sheet(title="Reconciliation Drill-down")
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    write_reconciliation_sheet(
        ws_summ,
        ws_det,
        sap_contract=[],
        inferred_contract=pd.DataFrame(),
        sap_meter=[],
        inferred_meter=pd.DataFrame(),
        sap_financial=[],
        evidence_df=pd.DataFrame(),
    )
    subtitle = str(ws_summ.cell(row=2, column=1).value)
    assert "evidence" in subtitle.lower(), subtitle
    assert "eater" not in subtitle.lower(), subtitle
