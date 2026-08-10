"""Characterization tests pinning the observable workbook output of
nested closures inside ``edf_bill_fetcher/io/writers/export.py``.

The export writer defines three closures *inside* ``export_to_excel``
that are not exposed at module scope:

* ``_compute_unit_rate`` (L388) — per-row unit-rate stamping applied to
  the duplicate DataFrame only; the result lands in the
  ``Duplicate Entries`` sheet's ``Unit Rate (p/kWh)`` column.
* ``ks_row`` (L730) — key-statistics row renderer for the
  ``Key Statistics`` sheet; dispatches on ``fmt`` (``"£"``, ``"%"``,
  ``"date"``, ``"#,##0"``, plain) and honours ``alt`` (LGREY fill) and
  ``bold`` (bold font).
* ``_banner`` (L1264) — section-header banner writer used by the
  ``Dispute Flags`` sheet (and others); writes text into row ``r``,
  fills columns 1-6 with the bg colour, sets a white bold font, and
  fixes the row height at 20.

These tests exercise the closures *through* the public
``export_to_excel`` surface and assert on the saved workbook via
``openpyxl.load_workbook`` — they never import or invoke the closures
directly (the closures do not exist at module scope, which is the
point of the characterization). They pin the CURRENT observable
behaviour so a future refactor that moves the closures out or alters
their output is caught.
"""

from __future__ import annotations

from typing import Any

import openpyxl
import pandas as pd

from edf_bill_fetcher.io.writers import export_to_excel
from edf_bill_fetcher.models.config import ConfigDict

# Colour constants mirrored from the LOCAL block at the top of
# ``export_to_excel`` (L109-115). They are not module-scope in export
# either, so the tests carry their own copies to assert against.
ORANGE = "FE5716"
LGREY = "F0F0F0"


def _base_config() -> ConfigDict:
    """Return a config dict that turns on dedup + dup-sheet saving and
    supplies an account reference so the Key Statistics ``Account
    reference`` row renders with a known value.
    """
    return {
        "use_anchors": False,
        "use_large": True,
        "use_reading_classification": False,
        "use_pdf_fields": False,
        "use_acc_filter": False,
        "acc_num": "",
        "min_amount": 1.0,
        "analysis_min": 1.0,
        "filter_below": False,
        "save_filtered": False,
        "use_dedup": True,
        "save_dups": True,
        "use_domain_filter": False,
        "domain_filter": "",
        "report_account_ref": "A-TEST-REF-123",
    }


def _record(
    *,
    invoice: str,
    amount: float,
    period_to: str,
    period_charge: float,
    units: float | None,
    date: str = "01/06/2024",
    source: str = "HTM Account History",
) -> dict[str, Any]:
    """Build a single evidence record dict with the canonical column
    set the writer expects (see ``evidence.py`` headers). ``units=None``
    yields a row with no usage — used to exercise the NaN branch of
    ``_compute_unit_rate``.
    """
    return {
        "Source": source,
        "Sender": "",
        "Date": date,
        "Period From": "01/05/2024",
        "Period To": period_to,
        "Invoice #": invoice,
        "Amount (£)": amount,
        "Period Charge (£)": period_charge,
        "Unit Rate (p/kWh)": None,
        "% Change": None,
        "Entry Type": "New Bill",
        "Reading": "Actual",
        "Units (kWh)": units,
        "Standing Chg (p/day)": 0.0,
        "Tariff": "Standard",
        "Attachment Name": "",
        "Details": "",
        "Logic Used": "",
        "Anomaly Flag": "",
    }


def _export(
    records: list[dict[str, Any]],
    tmp_path: object,
    config: ConfigDict | None = None,
) -> Any:
    """Run ``export_to_excel`` into a tmp workbook and return the
    loaded workbook for assertion. Centralises the path-ops ``type:
    ignore`` so individual tests stay readable.
    """
    out = tmp_path / "characterization.xlsx"  # type: ignore[operator]
    export_to_excel(
        data=pd.DataFrame(records),
        output_path=str(out),
        error_log=[],
        config=config or _base_config(),
    )
    return openpyxl.load_workbook(out)  # type: ignore[arg-type]


def _dup_unit_rate_cell(ws: Any) -> Any:
    """Locate the ``Unit Rate (p/kWh)`` data cell on the Duplicate
    Entries sheet. The header sits on row 1; the first dup row is row
    2. Returns the row-2 cell of the unit-rate column (or raises if the
    sheet/column is missing — the test then fails loudly).
    """
    header_row = next(ws.iter_rows(min_row=1, max_row=1))
    ur_col = next(
        (c.column for c in header_row if c.value == "Unit Rate (p/kWh)"),
        None,
    )
    assert ur_col is not None, "Duplicate Entries sheet lacks 'Unit Rate (p/kWh)' header"
    return ws.cell(row=2, column=ur_col)


# ---------------------------------------------------------------------------
# (a) _compute_unit_rate — unit-rate stamping on the dup DataFrame
# ---------------------------------------------------------------------------


def test_unit_rate_with_usage(tmp_path: object) -> None:
    """Pin ``_compute_unit_rate`` for a duplicate row that carries
    both ``Period Charge (£)`` and ``Units (kWh)``.

    Given: two records with the same Amount (£) and Period To — the
    second is flagged DUPLICATED and lands in the Duplicate Entries
    sheet. Period Charge = 100, Units = 500 → unit rate =
    round((100 / 500) * 100, 2) = 20.0 p/kWh.
    When: export_to_excel runs with dedup + save_dups on.
    Then: the Duplicate Entries sheet's row-2 ``Unit Rate (p/kWh)``
    cell holds 20.0.
    """
    records = [
        _record(
            invoice="KEPT-001",
            amount=500.0,
            period_to="31/05/2024",
            period_charge=100.0,
            units=500.0,
        ),
        _record(
            invoice="DUP-001",
            amount=500.0,
            period_to="31/05/2024",
            period_charge=100.0,
            units=500.0,
        ),
        _record(
            invoice="UNIQ-001",
            amount=300.0,
            period_to="30/04/2024",
            period_charge=60.0,
            units=300.0,
        ),
    ]
    wb = _export(records, tmp_path)
    assert "Duplicate Entries" in wb.sheetnames, wb.sheetnames
    cell = _dup_unit_rate_cell(wb["Duplicate Entries"])
    assert cell.value == 20.0, f"unit rate with usage: expected 20.0, got {cell.value!r}"


def test_unit_rate_without_usage(tmp_path: object) -> None:
    """Pin ``_compute_unit_rate`` for a duplicate row with NO usable
    usage (``Units (kWh)`` is None → NaN-coerced).

    Given: a duplicate pair where the dup row has ``Units (kWh)=None``
    and ``Period Charge (£)=100``. The closure's ``float(str(units))``
    path raises, the ``except`` swallows it, and ``np.nan`` is
    returned — which openpyxl serialises as an empty cell.
    When: export_to_excel runs with dedup + save_dups on.
    Then: the Duplicate Entries sheet's row-2 ``Unit Rate (p/kWh)``
    cell is empty (None), NOT a numeric zero.
    """
    records = [
        _record(
            invoice="KEPT-002",
            amount=500.0,
            period_to="31/05/2024",
            period_charge=100.0,
            units=500.0,
        ),
        _record(
            invoice="DUP-002", amount=500.0, period_to="31/05/2024", period_charge=100.0, units=None
        ),
        _record(
            invoice="UNIQ-002",
            amount=300.0,
            period_to="30/04/2024",
            period_charge=60.0,
            units=300.0,
        ),
    ]
    wb = _export(records, tmp_path)
    assert "Duplicate Entries" in wb.sheetnames, wb.sheetnames
    cell = _dup_unit_rate_cell(wb["Duplicate Entries"])
    assert cell.value is None, f"unit rate without usage: expected None (NaN), got {cell.value!r}"


# ---------------------------------------------------------------------------
# (b) ks_row — Key Statistics row rendering
# ---------------------------------------------------------------------------


def test_key_statistics_row_variants(tmp_path: object) -> None:
    """Pin the observable output of ``ks_row`` across its ``alt``,
    ``fmt``, and ``bold`` variants on the Key Statistics sheet.

    Given: a config with ``report_account_ref`` set and ≥2 records so
    the Key Statistics sheet renders.
    When: export_to_excel runs.
    Then:
      * ``alt`` row (Account reference, r=3) — column 1 label text is
        ``"Account reference"`` and column 1 carries the LGREY
        (``F0F0F0``) fill.
      * ``fmt="£"`` + ``bold`` row (Current balance, r=10) — column 2
        number_format is ``"£#,##0.00"`` and the font is bold.
      * ``fmt="%"`` + ``bold`` row (% increase, r=12) — column 2
        number_format is ``"0.0%"`` and the font is bold.
      * ``fmt="#,##0"`` row (Period covered, r=6) — column 2
        number_format is ``"#,##0"``.
      * The account-reference row's column-2 value is the configured
        reference string ``"A-TEST-REF-123"``.
    """
    records = [
        _record(
            invoice="KS-001",
            amount=500.0,
            period_to="31/05/2024",
            period_charge=100.0,
            units=500.0,
            date="01/05/2024",
        ),
        _record(
            invoice="KS-002",
            amount=300.0,
            period_to="30/04/2024",
            period_charge=60.0,
            units=300.0,
            date="01/04/2024",
        ),
    ]
    wb = _export(records, tmp_path)
    assert "Key Statistics" in wb.sheetnames, wb.sheetnames
    ws = wb["Key Statistics"]

    # alt row — Account reference (r=3, alt=True, fmt=None).
    label_cell = ws.cell(row=3, column=1)
    assert label_cell.value == "Account reference", (
        f"ks_row alt label: expected 'Account reference', got {label_cell.value!r}"
    )
    fill = label_cell.fill
    assert fill.patternType == "solid", fill.patternType
    assert str(fill.start_color.rgb).endswith(LGREY), (
        f"ks_row alt fill: expected LGREY {LGREY}, got {fill.start_color.rgb!r}"
    )
    # The account-reference value lands in column 2 as plain text.
    ref_cell = ws.cell(row=3, column=2)
    assert ref_cell.value == "A-TEST-REF-123", (
        f"ks_row account ref value: expected 'A-TEST-REF-123', got {ref_cell.value!r}"
    )

    # fmt="£" + bold — Current balance (r=10).
    money_cell = ws.cell(row=10, column=2)
    assert money_cell.number_format == "£#,##0.00", (
        f"ks_row £ number_format: expected '£#,##0.00', got {money_cell.number_format!r}"
    )
    assert money_cell.font.bold is True, (
        f"ks_row £ bold: expected True, got {money_cell.font.bold!r}"
    )

    # fmt="%" + bold — % increase (r=12).
    pct_cell = ws.cell(row=12, column=2)
    assert pct_cell.number_format == "0.0%", (
        f"ks_row % number_format: expected '0.0%', got {pct_cell.number_format!r}"
    )
    assert pct_cell.font.bold is True, f"ks_row % bold: expected True, got {pct_cell.font.bold!r}"

    # fmt="#,##0" — Period covered (r=6).
    int_cell = ws.cell(row=6, column=2)
    assert int_cell.number_format == "#,##0", (
        f"ks_row #,##0 number_format: expected '#,##0', got {int_cell.number_format!r}"
    )


# ---------------------------------------------------------------------------
# (c) _banner — section-header banner on the Dispute Flags sheet
# ---------------------------------------------------------------------------


def test_dispute_flags_banner_layout(tmp_path: object) -> None:
    """Pin the ``_banner`` closure's observable layout on the Dispute
    Flags sheet's row-1 banner.

    Given: ≥2 records so the Dispute Flags sheet renders.
    When: export_to_excel runs.
    Then: row 1 column 1 holds the verbatim banner text
      ``"EDF ENERGY DISPUTE  —  AUTOMATED ANALYSIS FLAGS"``, the cell
      font is white + bold, columns 1-6 all carry the ORANGE
      (``FE5716``) solid fill, and the row height is 20.
    """
    records = [
        _record(
            invoice="BNR-001",
            amount=500.0,
            period_to="31/05/2024",
            period_charge=100.0,
            units=500.0,
            date="01/05/2024",
        ),
        _record(
            invoice="BNR-002",
            amount=300.0,
            period_to="30/04/2024",
            period_charge=60.0,
            units=300.0,
            date="01/04/2024",
        ),
    ]
    wb = _export(records, tmp_path)
    assert "Dispute Flags" in wb.sheetnames, wb.sheetnames
    ws = wb["Dispute Flags"]

    text_cell = ws.cell(row=1, column=1)
    assert text_cell.value == "EDF ENERGY DISPUTE  —  AUTOMATED ANALYSIS FLAGS", (
        f"banner text: got {text_cell.value!r}"
    )
    assert text_cell.font.bold is True, (
        f"banner font bold: expected True, got {text_cell.font.bold!r}"
    )
    assert str(text_cell.font.color.rgb).endswith("FFFFFF"), (
        f"banner font color: expected white FFFFFF, got {text_cell.font.color.rgb!r}"
    )

    # Columns 1-6 must all carry the ORANGE solid fill.
    for col in range(1, 7):
        cell = ws.cell(row=1, column=col)
        fill = cell.fill
        assert fill.patternType == "solid", f"col {col}: {fill.patternType!r}"
        assert str(fill.start_color.rgb).endswith(ORANGE), (
            f"banner col {col} fill: expected ORANGE {ORANGE}, got {fill.start_color.rgb!r}"
        )

    assert ws.row_dimensions[1].height == 20, (
        f"banner row height: expected 20, got {ws.row_dimensions[1].height!r}"
    )
