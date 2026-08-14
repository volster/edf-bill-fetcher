import openpyxl
import pandas as pd

from edf_bill_fetcher.io.writers.superseded import write_superseded_reconciliation_sheet


def _bb() -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "Invoice #": "B",
                "Bill Date": "01 Jan 2023",
                "Period From": "01 Jan 2022",
                "Period To": "01 Jan 2023",
                "Days Billed": 365,
                "Period Charge (£)": 500.0,
                "Unlawful Charge (£)": 250.0,
                "Excess Days": 30,
                "Value Source": "PDF",
                "Cancel/Rebill Admitted": True,
                "Reason Assessment": "test",
            },
            {
                "Invoice #": "D",
                "Bill Date": "01 Jan 2024",
                "Period From": "01 Jan 2023",
                "Period To": "01 Jan 2024",
                "Days Billed": 365,
                "Period Charge (£)": 600.0,
                "Unlawful Charge (£)": 300.0,
                "Excess Days": 40,
                "Value Source": "PDF",
                "Cancel/Rebill Admitted": False,
                "Reason Assessment": "test",
            },
        ]
    )


def test_write_superseded_reconciliation_sheet_groups_by_killer() -> None:
    ws = openpyxl.Workbook().active
    domination_map = {"B": ("A", False), "D": ("A", False)}
    write_superseded_reconciliation_sheet(ws, _bb(), domination_map)
    col_a = [ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)]
    assert any(v == "KILLER: A" for v in col_a)
    assert "B" in col_a
    assert "D" in col_a


def test_write_superseded_reconciliation_sheet_links() -> None:
    ws = openpyxl.Workbook().active
    domination_map = {"B": ("A", False)}
    write_superseded_reconciliation_sheet(
        ws,
        _bb(),
        domination_map,
        evidence_index={"inv:B": 10},
        invoice_pdf_paths={"B": "evidence_files/B.pdf", "A": "evidence_files/A.pdf"},
        live_row_map={"A": 8},
    )
    # find row for B
    b_row = next(r for r in range(1, ws.max_row + 1) if ws.cell(row=r, column=1).value == "B")
    hdrs = [c.value for c in ws[7]]
    cols = {h: i + 1 for i, h in enumerate(hdrs)}
    assert ws.cell(row=b_row, column=cols["Killer on spreadsheet"]).hyperlink is not None
    assert (
        "'Back-billing Analysis'!A8"
        in ws.cell(row=b_row, column=cols["Killer on spreadsheet"]).hyperlink.location
    )
    assert (
        ws.cell(row=b_row, column=cols["Original invoice on spreadsheet"]).hyperlink.location
        == "'EDF Evidence Report'!A10"
    )
    assert (
        ws.cell(row=b_row, column=cols["Original invoice PDF"]).hyperlink.target
        == "evidence_files/B.pdf"
    )
    assert (
        ws.cell(row=b_row, column=cols["Killer invoice PDF"]).hyperlink.target
        == "evidence_files/A.pdf"
    )


def test_write_superseded_reconciliation_sheet_returns_survivor_row_map() -> None:
    ws = openpyxl.Workbook().active
    domination_map = {"B": ("A", False), "D": ("A", False)}
    recon_row_map = write_superseded_reconciliation_sheet(ws, _bb(), domination_map)
    assert recon_row_map == {"A": 8}


def test_write_superseded_reconciliation_sheet_returns_row_map_with_multiple_groups() -> None:
    ws = openpyxl.Workbook().active
    domination_map = {"B": ("A", False), "D": ("C", False)}
    recon_row_map = write_superseded_reconciliation_sheet(ws, _bb(), domination_map)
    # Row 8 = KILLER: A, row 9 = B data, row 10 = KILLER: C, row 11 = D data.
    assert recon_row_map == {"A": 8, "C": 10}


def test_write_superseded_reconciliation_sheet_chain_note_on_reason_assessment() -> None:
    ws = openpyxl.Workbook().active
    domination_map = {"B": ("A", False), "D": ("A", False)}
    write_superseded_reconciliation_sheet(ws, _bb(), domination_map)
    b_row = next(r for r in range(1, ws.max_row + 1) if ws.cell(row=r, column=1).value == "B")
    col_10 = ws.cell(row=b_row, column=10).value
    assert "Superseded by A" in col_10
    assert "Superseded" in col_10


def test_write_superseded_reconciliation_sheet_total_and_empty() -> None:
    ws = openpyxl.Workbook().active
    domination_map = {"B": ("A", False), "D": ("A", False)}
    write_superseded_reconciliation_sheet(ws, _bb(), domination_map)
    labels = [ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)]
    total_row = next(r for r, v in enumerate(labels, 1) if v and "TOTAL SUPERSEDED" in str(v))
    # unlawful col (7)
    assert ws.cell(row=total_row, column=7).value == 550.0
    ws2 = openpyxl.Workbook().active
    write_superseded_reconciliation_sheet(ws2, _bb(), {})
    assert (
        any(
            v == "KILLER: A"
            for v in [ws2.cell(row=r, column=1).value for r in range(1, ws2.max_row + 1)]
        )
        is False
    )
    labels2 = [ws2.cell(row=r, column=1).value for r in range(1, ws2.max_row + 1)]
    total_row2 = next(r for r, v in enumerate(labels2, 1) if v and "TOTAL SUPERSEDED" in str(v))
    assert ws2.cell(row=total_row2, column=7).value == 0.0
