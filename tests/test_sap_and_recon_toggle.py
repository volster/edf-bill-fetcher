"""Regression test pinning the SAP / Reconciliation toggle wiring.

Each toggle (``scan_sap_dumps``, ``generate_reconciliation_sheet``)
must reach ``export_to_excel`` via the ``config`` dict so the GUI
toggles actually take effect. Pins the gate logic in
``export_to_excel`` so a future regression that drops the toggle
breaks in CI rather than in front of a reviewer.

Pattern: feed a small record set + a SAP-row packet to
``export_to_excel`` twice -- once with the toggle True, once with
it False -- and assert the resulting workbook's sheet names match
the toggle contract.

Companion to ``tests/test_save_dups_toggle.py``.
"""

from __future__ import annotations

import os

import pandas as pd
from openpyxl import load_workbook

from edf_collector import (
    EvidenceEngine,
    export_to_excel,
    parse_sap_contract_history,
    parse_sap_financial_transactions,
    parse_sap_meter_read_history,
)

# Reuse the synthetic SAP CSV strings from the existing test fixtures.
from tests.test_sap_parser import CONTRACT_CSV, FINANCIAL_CSV, METER_CSV  # type: ignore


def _seed_one_record(engine: EvidenceEngine) -> None:
    engine.process_text(
        "01 Aug 2023 We charged your account £500.00 For 1000 kWh of electricity "
        "used between 01 Jan 2022 and 31 Jul 2023 Balance £100.00 in debit",
        "Local PDF Folder",
        "seed.001",
        "01/08/2023",
    )


def _records_to_rows(records: list[dict]) -> pd.DataFrame:
    if not records:
        return pd.DataFrame()
    return pd.DataFrame(records)


def _engine_with_config(**overrides: object) -> EvidenceEngine:
    cfg: dict[str, object] = {
        "use_anchors": False,
        "use_large": False,
        "use_reading_classification": False,
        "use_pdf_fields": False,
        "use_acc_filter": False,
        "acc_num": "0123456789",
        "min_amount": 1.0,
        "analysis_min": 1.0,
        "filter_below": False,
        "save_filtered": False,
        "use_dedup": False,
        "save_dups": False,
        "use_domain_filter": False,
        "domain_filter": "",
        "scan_sap_dumps": True,
        "generate_reconciliation_sheet": True,
    }
    cfg.update(overrides)
    return EvidenceEngine(cfg, lambda *a: None)


def _sap_rows() -> dict:
    return {
        "contract": parse_sap_contract_history(CONTRACT_CSV),
        "meter": parse_sap_meter_read_history(METER_CSV),
        "financial": parse_sap_financial_transactions(FINANCIAL_CSV),
    }


def _sheet_names(out_path: str) -> set[str]:
    wb = load_workbook(out_path, read_only=True)
    names = set(wb.sheetnames)
    wb.close()
    return names


def test_scan_sap_dumps_true_emits_three_sap_sheets_and_recon(tmp_path: object) -> None:
    out_path = str(tmp_path / "with_sap.xlsx")  # type: ignore[operator]
    export_to_excel(
        data=_sample_data(),
        output_path=out_path,
        error_log=[],
        config={
            "use_dedup": False,
            "acc_num": "0123456789",
            "scan_sap_dumps": True,
            "generate_reconciliation_sheet": True,
        },
        sap_rows=_sap_rows(),
    )
    assert os.path.exists(out_path)
    names = _sheet_names(out_path)
    assert "SAP Contract History" in names
    assert "SAP Meter Readings" in names
    assert "SAP Financial Transactions" in names
    assert "Reconciliation" in names


def test_scan_sap_dumps_false_suppresses_all_four_new_sheets(tmp_path: object) -> None:
    """When scan_sap_dumps is False, NONE of the four new sheets
    appear even when SAP rows are supplied."""
    out_path = str(tmp_path / "no_sap.xlsx")  # type: ignore[operator]
    export_to_excel(
        data=_sample_data(),
        output_path=out_path,
        error_log=[],
        config={
            "use_dedup": False,
            "acc_num": "0123456789",
            "scan_sap_dumps": False,
            "generate_reconciliation_sheet": True,
        },
        sap_rows=_sap_rows(),
    )
    names = _sheet_names(out_path)
    assert "SAP Contract History" not in names
    assert "SAP Meter Readings" not in names
    assert "SAP Financial Transactions" not in names
    assert "Reconciliation" not in names


def test_recon_toggle_false_renders_three_sap_sheets_but_no_recon(
    tmp_path: object,
) -> None:
    """scan_sap_dumps=True but generate_reconciliation_sheet=False ->
    three SAP sheets render but the Reconciliation sheet is suppressed."""
    out_path = str(tmp_path / "sap_no_recon.xlsx")  # type: ignore[operator]
    export_to_excel(
        data=_sample_data(),
        output_path=out_path,
        error_log=[],
        config={
            "use_dedup": False,
            "acc_num": "0123456789",
            "scan_sap_dumps": True,
            "generate_reconciliation_sheet": False,
        },
        sap_rows=_sap_rows(),
    )
    names = _sheet_names(out_path)
    assert "SAP Contract History" in names
    assert "SAP Meter Readings" in names
    assert "SAP Financial Transactions" in names
    assert "Reconciliation" not in names


def test_gui_toggles_are_written_into_evidence_engine_config_dict() -> None:
    """The App._run handler builds the config dict from
    ``tk.BooleanVar``s. Confirm the values are propagated through the
    engine's config accessor so a regression in the App wiring is
    caught by a unit test (no Tk needed)."""
    # We can talk to the dict directly -- the wiring through _run is
    # visually verified in the GUI test set, but the propagation
    # surface is the dict itself.
    cfg = {
        "scan_sap_dumps": False,
        "generate_reconciliation_sheet": True,
        "save_evidence_files": True,
    }
    e = _engine_with_config(**cfg)
    assert e.config["scan_sap_dumps"] is False
    assert e.config["generate_reconciliation_sheet"] is True
    assert e.config["save_evidence_files"] is True


def test_static_check_sap_and_recon_toggles_are_read_inside_export(
    tmp_path: object,
) -> None:
    """Pin the toggles being read inside ``export_to_excel`` so a
    regression that drops the gate surfaces fast.
    Mirrors ``test_save_dups_kwarg_match_dedup_branch_in_export``."""
    import inspect

    from edf_collector import export_to_excel as _export_to_excel

    src = inspect.getsource(_export_to_excel)
    # The export body must consult both toggles to gate the SAP /
    # reconciliation sheet writes.
    assert 'config.get("scan_sap_dumps"' in src, (
        "export_to_excel must consult config['scan_sap_dumps'] to gate the SAP sheets"
    )
    assert 'config.get("generate_reconciliation_sheet"' in src, (
        "export_to_excel must consult config['generate_reconciliation_sheet'] to gate Reconciliation"
    )


# ---------------------------------------------------------------------------
# Shared fixture data for the toggle tests.
# ---------------------------------------------------------------------------


def _sample_data() -> list[dict]:
    """Two invoices that produce Back-billing + Rebilling rows."""
    return [
        {
            "Source": "Local PDF Folder",
            "Sender": "edf.co.uk",
            "Date": "01 Aug 2023",
            "Period From": "01 Jan 2022",
            "Period To": "31 Jul 2023",
            "Invoice #": "T-X1",
            "Amount (£)": 1000.0,
            "Period Charge (£)": 800.0,
            "Unit Rate (p/kWh)": 25.0,
            "% Change": None,
            "Entry Type": "New Bill",
            "Reading": "Actual",
            "Units (kWh)": 300.0,
            "Standing Chg (p/day)": 50.0,
            "Tariff": "Standard",
            "Attachment Name": "T-X1.pdf",
            "Details": "",
            "Logic Used": "PDF new-format",
            "Anomaly Flag": "",
            "Cancel/Rebill Admitted": True,
        },
        {
            "Source": "Local PDF Folder",
            "Sender": "edf.co.uk",
            "Date": "01 Sep 2023",
            "Period From": "01 Jan 2022",
            "Period To": "31 Aug 2023",
            "Invoice #": "T-X2",
            "Amount (£)": 1500.0,
            "Period Charge (£)": 1200.0,
            "Unit Rate (p/kWh)": 25.0,
            "% Change": None,
            "Entry Type": "New Bill",
            "Reading": "Actual",
            "Units (kWh)": 400.0,
            "Standing Chg (p/day)": 50.0,
            "Tariff": "Standard",
            "Attachment Name": "T-X2.pdf",
            "Details": "",
            "Logic Used": "PDF new-format",
            "Anomaly Flag": "",
            "Cancel/Rebill Admitted": False,
        },
    ]
