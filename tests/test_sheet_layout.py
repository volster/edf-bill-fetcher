"""Unit tests for the shared row-layout helpers in sheet_layout.py."""

from __future__ import annotations

from openpyxl import Workbook

from edf_bill_fetcher.io.writers.sheet_layout import (
    freeze_at,
    write_banner,
    write_header_row,
    write_merged_text,
    write_section_label,
    write_trailing_total,
)


def _open_ws() -> object:
    wb = Workbook()
    return wb.active


def _rgb(cell) -> str:
    return cell.fill.start_color.rgb


def _merged(ws) -> list[str]:
    return [str(r) for r in ws.merged_cells.ranges]


def test_write_banner_fills_all_columns() -> None:
    ws = _open_ws()
    write_banner(ws, "TITLE", ncols=4)
    assert ws.cell(row=1, column=1).value == "TITLE"
    assert _rgb(ws.cell(row=1, column=1)) == "00FE5716"
    assert ws.cell(row=1, column=1).font.bold is True
    assert ws.cell(row=1, column=1).font.color.rgb == "00FFFFFF"
    assert ws.cell(row=1, column=1).font.size == 13
    for c in range(2, 5):
        assert ws.cell(row=1, column=c).value is None
        assert _rgb(ws.cell(row=1, column=c)) == "00FE5716"
        assert ws.cell(row=1, column=c).border is not None


def test_write_banner_custom_color_and_height() -> None:
    ws = _open_ws()
    write_banner(ws, "TITLE", ncols=3, color="123456", height=30)
    assert _rgb(ws.cell(row=1, column=1)) == "00123456"
    assert _rgb(ws.cell(row=1, column=2)) == "00123456"
    assert ws.row_dimensions[1].height == 30


def test_write_section_label_spans_and_fills() -> None:
    ws = _open_ws()
    write_section_label(ws, row=2, label="LEGAL CONTEXT", ncols=5)
    assert ws.cell(row=2, column=1).value == "LEGAL CONTEXT"
    assert ws.cell(row=2, column=1).font.bold is True
    assert ws.cell(row=2, column=1).font.color.rgb == "00FFFFFF"
    assert ws.cell(row=2, column=1).font.size == 11
    assert _rgb(ws.cell(row=2, column=1)) == "0010367A"
    for c in range(2, 6):
        assert ws.cell(row=2, column=c).value is None
        assert _rgb(ws.cell(row=2, column=c)) == "0010367A"
        assert ws.cell(row=2, column=c).border.left.style == "thin"


def test_write_merged_text_merges_and_wraps() -> None:
    ws = _open_ws()
    write_merged_text(ws, row=3, text_value="some paragraph", ncols=6, height=90)
    assert ws.cell(row=3, column=1).value == "some paragraph"
    assert ws.cell(row=3, column=1).alignment.wrap_text is True
    assert ws.cell(row=3, column=1).font.italic is False
    assert ws.cell(row=3, column=1).border.left.style == "thin"
    assert _merged(ws) == ["A3:F3"]
    assert ws.row_dimensions[3].height == 90


def test_write_merged_text_italic_flag() -> None:
    ws = _open_ws()
    write_merged_text(ws, row=5, text_value="instruction", ncols=4, italic=True)
    assert ws.cell(row=5, column=1).font.italic is True
    assert _merged(ws) == ["A5:D5"]


def test_write_merged_text_border_false_leaves_no_border() -> None:
    ws = _open_ws()
    write_merged_text(ws, row=5, text_value="instruction", ncols=4, border=False)
    assert ws.cell(row=5, column=1).border.left.style is None
    assert ws.cell(row=5, column=1).border.top.style is None
    assert _merged(ws) == ["A5:D5"]


def test_write_header_row_default_center_alignment() -> None:
    ws = _open_ws()
    write_header_row(ws, row=7, headers=["Invoice #", "Bill Date", "Period From"])
    assert ws.cell(row=7, column=1).value == "Invoice #"
    assert ws.cell(row=7, column=1).font.bold is True
    assert ws.cell(row=7, column=1).font.color.rgb == "00FFFFFF"
    assert ws.cell(row=7, column=1).font.size == 10
    assert _rgb(ws.cell(row=7, column=1)) == "0010367A"
    assert ws.cell(row=7, column=1).alignment.horizontal == "center"
    assert ws.cell(row=7, column=2).value == "Bill Date"
    assert ws.cell(row=7, column=3).value == "Period From"


def test_write_header_row_left_alignment_and_height() -> None:
    ws = _open_ws()
    write_header_row(ws, row=7, headers=["A", "B"], align="left", height=28)
    assert ws.cell(row=7, column=1).alignment.horizontal == "left"
    assert ws.cell(row=7, column=2).alignment.horizontal == "left"
    assert ws.row_dimensions[7].height == 28


def test_write_trailing_total_merges_label_and_values() -> None:
    ws = _open_ws()
    write_trailing_total(
        ws,
        row=10,
        label="TOTAL",
        values=[(6, 1234.5), (10, 99.99)],
        label_span=5,
        ncols=12,
    )
    assert ws.cell(row=10, column=1).value == "TOTAL"
    assert _merged(ws) == ["A10:E10"]
    assert ws.cell(row=10, column=6).value == 1234.5
    assert ws.cell(row=10, column=6).number_format == "#,##0.00"
    assert ws.cell(row=10, column=10).value == 99.99
    assert ws.cell(row=10, column=10).number_format == "#,##0.00"
    assert ws.row_dimensions[10].height == 22


def test_write_trailing_total_fills_every_non_value_column() -> None:
    ws = _open_ws()
    write_trailing_total(
        ws,
        row=10,
        label="TOTAL",
        values=[(6, 1234.5)],
        label_span=5,
        ncols=12,
    )
    for c in range(1, 13):
        if c == 6:
            continue
        assert _rgb(ws.cell(row=10, column=c)) == "0010367A"
        assert ws.cell(row=10, column=c).border is not None
    assert _rgb(ws.cell(row=10, column=2)) == "0010367A"
    assert _rgb(ws.cell(row=10, column=5)) == "0010367A"


def test_freeze_at_sets_freeze_panes() -> None:
    ws = _open_ws()
    freeze_at(ws, "A8")
    assert ws.freeze_panes == "A8"
