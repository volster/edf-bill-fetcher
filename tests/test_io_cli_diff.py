"""Tests for the ``--diff`` CLI subcommand (``edf_bill_fetcher.io.cli``).

Covers ``run_cli_diff`` — the headless run-diff surface that ``main()``
dispatches to: it reads two ``records.json`` files, calls the
``processors.run_diff.diff_records`` processor, prints a human-readable
summary (counts + per-row ``+ ADDED`` / ``- REMOVED`` / ``~ CHANGED``
lines), and optionally writes a 3-sheet Excel workbook via
``io.writers.diff`` when ``--diff-output`` is given.

Every test drives the real CLI function with on-disk fixture JSON and
asserts observable behaviour — stdout content, workbook structure, exit
codes — never call counts.
"""

from __future__ import annotations

import json
import sys
from pathlib import Path
from typing import Any

import pytest

from edf_bill_fetcher.io.cli import main, run_cli_diff

# ---------------------------------------------------------------------------
# Fixture records — fully synthetic, no real customer data
# ---------------------------------------------------------------------------


def _write_records_json(path: Path, records: list[dict[str, Any]]) -> None:
    """Write a bare-list records JSON file (the ``--records-json`` shape)."""
    path.write_text(json.dumps(records), encoding="utf-8")


def _old_records() -> list[dict[str, Any]]:
    return [
        {
            "Date": "2026-03-01",
            "Amount (£)": 120.50,
            "Source": "Local PDF",
            "Details": "Automatic estimate",
        },
        {
            "Date": "2026-04-01",
            "Amount (£)": 130.00,
            "Source": "Local PDF",
            "Details": "Actual reading",
        },
        {
            "Date": "2026-05-01",
            "Amount (£)": 140.00,
            "Source": "Local PDF",
            "Details": "Dropped in the new run",
        },
    ]


def _new_records() -> list[dict[str, Any]]:
    return [
        {
            "Date": "2026-03-01",
            "Amount (£)": 120.50,
            "Source": "Local PDF",
            "Details": "Automatic estimate",
        },
        {
            "Date": "2026-04-01",
            "Amount (£)": 155.00,
            "Source": "Local PDF",
            "Details": "Actual reading, re-billed",
        },
        {
            "Date": "2026-06-01",
            "Amount (£)": 160.00,
            "Source": "HTM Export",
            "Details": "Brand new row",
        },
    ]


@pytest.fixture
def records_pair(tmp_path: Path) -> tuple[Path, Path]:
    """Write the old/new fixture pair and return both paths."""
    old_path = tmp_path / "old_records.json"
    new_path = tmp_path / "new_records.json"
    _write_records_json(old_path, _old_records())
    _write_records_json(new_path, _new_records())
    return old_path, new_path


# ---------------------------------------------------------------------------
# run_cli_diff
# ---------------------------------------------------------------------------


class TestRunCliDiff:
    """Cover the ``run_cli_diff`` headless run-diff entry point."""

    def test_summary_counts_and_per_row_lines(
        self,
        records_pair: tuple[Path, Path],
        capsys: pytest.CaptureFixture[str],
    ) -> None:
        """``--diff OLD NEW`` prints counts plus one line per delta row.

        The old/new fixture pair produces exactly one added row
        (2026-06-01), one removed row (2026-05-01) and one changed row
        (2026-04-01: amount + details).  The unchanged 2026-03-01 row
        must not appear.
        """
        old_path, new_path = records_pair

        run_cli_diff([str(old_path), str(new_path)])

        out = capsys.readouterr().out
        assert "Added:   1" in out
        assert "Removed: 1" in out
        assert "Changed: 1" in out
        assert "+ ADDED 2026-06-01 £160.00 HTM Export" in out
        assert "- REMOVED 2026-05-01 £140.00 Local PDF" in out
        assert "~ CHANGED 2026-04-01 £155.00 Local PDF" in out
        assert "[Amount (£): £130.00 → £155.00" in out
        assert "Details: Actual reading → Actual reading, re-billed]" in out
        # The unchanged row is neither added, removed nor changed: exactly
        # one line per delta row.
        assert out.count("+ ADDED") == 1
        assert out.count("- REMOVED") == 1
        assert out.count("~ CHANGED") == 1

    def test_diff_output_writes_three_sheet_workbook(
        self,
        records_pair: tuple[Path, Path],
        tmp_path: Path,
    ) -> None:
        """``--diff-output`` writes an xlsx with added/removed/changed sheets."""
        import openpyxl

        old_path, new_path = records_pair
        out_xlsx = tmp_path / "diff.xlsx"

        run_cli_diff([str(old_path), str(new_path), "--diff-output", str(out_xlsx)])

        assert out_xlsx.exists()
        wb = openpyxl.load_workbook(str(out_xlsx))
        assert wb.sheetnames == ["Added Records", "Removed Records", "Changed Records"]

        added = wb["Added Records"]
        removed = wb["Removed Records"]
        changed = wb["Changed Records"]

        # Added/removed sheets: plain headers, one data row each.
        assert added.cell(row=1, column=1).value == "Date"
        assert added.cell(row=2, column=1).value == "2026-06-01"
        assert added.max_row == 2
        assert removed.cell(row=2, column=1).value == "2026-05-01"

        # Changed sheet: paired old/new columns + a Changed Fields column.
        headers = [changed.cell(row=1, column=c).value for c in range(1, changed.max_column + 1)]
        assert "Amount (£) (old)" in headers
        assert "Amount (£) (new)" in headers
        assert "Details (old)" in headers
        assert "Details (new)" in headers
        assert headers[-1] == "Changed Fields"
        amount_new_col = headers.index("Amount (£) (new)") + 1
        details_new_col = headers.index("Details (new)") + 1
        changed_fields_col = headers.index("Changed Fields") + 1
        assert changed.cell(row=2, column=amount_new_col).value == 155.00
        assert changed.cell(row=2, column=details_new_col).value == "Actual reading, re-billed"
        assert "Amount (£): £130.00 → £155.00" in str(
            changed.cell(row=2, column=changed_fields_col).value
        )

    def test_missing_file_exits_1_cleanly(
        self,
        tmp_path: Path,
        capsys: pytest.CaptureFixture[str],
    ) -> None:
        """A missing input file writes a single ERROR line and exits 1."""
        with pytest.raises(SystemExit) as exc:
            run_cli_diff([str(tmp_path / "missing.json"), str(tmp_path / "also_missing.json")])
        assert exc.value.code == 1
        err = capsys.readouterr().err
        assert err.strip().startswith("ERROR:")
        assert "Traceback" not in err

    def test_wrapper_dict_records_unwrapped(
        self,
        tmp_path: Path,
        capsys: pytest.CaptureFixture[str],
    ) -> None:
        """A wrapper ``{"records": [...]}`` JSON is unwrapped like the report CLIs."""
        old_wrapper = tmp_path / "old.json"
        new_wrapper = tmp_path / "new.json"
        old_wrapper.write_text(
            json.dumps({"records": _old_records(), "extracted_at": "2026-01-01"}),
            encoding="utf-8",
        )
        new_wrapper.write_text(
            json.dumps({"records": _new_records(), "extracted_at": "2026-01-02"}),
            encoding="utf-8",
        )

        run_cli_diff([str(old_wrapper), str(new_wrapper)])

        out = capsys.readouterr().out
        assert "Added:   1" in out
        assert "Removed: 1" in out
        assert "Changed: 1" in out


class TestMainDiffDispatch:
    """Cover the ``main()`` ``--diff`` dispatch branch."""

    def test_diff_dispatch(
        self,
        monkeypatch: pytest.MonkeyPatch,
        records_pair: tuple[Path, Path],
        capsys: pytest.CaptureFixture[str],
    ) -> None:
        """``main()`` with ``--diff`` dispatches to ``run_cli_diff``."""
        old_path, new_path = records_pair
        monkeypatch.setattr(
            sys,
            "argv",
            ["edf-collector", "--diff", str(old_path), str(new_path)],
        )

        main()  # Should complete without raising SystemExit.
        assert "Changed: 1" in capsys.readouterr().out
