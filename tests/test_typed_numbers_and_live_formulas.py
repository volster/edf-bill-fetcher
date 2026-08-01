"""Spec §3.4 + §4 acceptance: typed-numbers pass, live-formulas pass,
post-save ignoredErrors injection, column-width helper, currency/int formatters.
PR #8: workbook-wide polish."""
from __future__ import annotations

import os
import re
import zipfile

from openpyxl import Workbook

from edf_bill_fetcher.helpers.excel_utils import (
    set_column_widths_from_spec,
    suppress_text_warning,
    suppress_text_warnings_post_save,
)
from edf_bill_fetcher.helpers.formatting import (
    apply_currency_format,
    apply_int_format,
)


def test_set_column_widths_from_spec_applies_all() -> None:
    """set_column_widths_from_spec writes column widths correctly."""
    wb = Workbook()
    ws = wb.active
    set_column_widths_from_spec(ws, {"A": 18, "B": 12, "C": 14})
    assert ws.column_dimensions["A"].width == 18
    assert ws.column_dimensions["B"].width == 12
    assert ws.column_dimensions["C"].width == 14


def test_apply_currency_format_sets_format_and_coerces_float() -> None:
    """apply_currency_format coerces string to float and sets currency format."""
    wb = Workbook()
    ws = wb.active
    cell = ws.cell(row=1, column=1, value="1234.56")  # string
    apply_currency_format(cell)
    assert cell.value == 1234.56, cell.value
    assert cell.number_format == "\u00a3#,##0.00", cell.number_format


def test_apply_int_format_sets_int_with_no_decimal() -> None:
    """apply_int_format coerces string to int and sets integer format."""
    wb = Workbook()
    ws = wb.active
    cell = ws.cell(row=1, column=1, value="1234")  # string
    apply_int_format(cell)
    assert cell.value == 1234
    assert cell.number_format == "#,##0", cell.number_format


def test_suppress_text_warning_adds_to_queue() -> None:
    """suppress_text_warning queues a suppression entry keyed by sheet title."""
    from edf_bill_fetcher.helpers.excel_utils import _TEXT_SUPPRESSION_QUEUE
    _TEXT_SUPPRESSION_QUEUE.clear()
    wb = Workbook()
    ws = wb.active
    ws.title = "TestSheet"
    suppress_text_warning(ws, "F", 2, 100)
    queue = _TEXT_SUPPRESSION_QUEUE.get("TestSheet")
    assert queue == [("F", 2, 100)], queue


def test_suppress_text_warnings_post_save_rewrites_zip_with_block() -> None:
    """Post-save zip injection rounds cleanly: ignoredErrors block appears in sheet XML."""
    from edf_bill_fetcher.helpers.excel_utils import _TEXT_SUPPRESSION_QUEUE
    _TEXT_SUPPRESSION_QUEUE.clear()

    out = "/tmp/opencode/test_post_save.xlsx"  # type: ignore[operator]
    os.makedirs("/tmp/opencode", exist_ok=True)
    if os.path.exists(out):
        os.remove(out)

    wb = Workbook()
    ws = wb.active
    ws.title = "TestSheet"
    for r in range(1, 6):
        ws.cell(row=r, column=1, value=f"INV-00{r}")
    _TEXT_SUPPRESSION_QUEUE["TestSheet"] = [("A", 1, 5)]
    wb.save(out)
    suppress_text_warnings_post_save(out)

    with zipfile.ZipFile(out, "r") as z:
        for n in z.namelist():
            if not n.startswith("xl/worksheets/sheet") or not n.endswith(".xml"):
                continue
            xml = z.read(n).decode("utf-8", errors="replace")
            pattern = (
                r"<ignoredErrors><ignoredError sqref=\"A1:A5\" "
                r"numberStoredAsText=\"1\"/></ignoredErrors>"
            )
            assert re.search(pattern, xml), (
                f"{n} lacks ignoredErrors block for A1:A5 text-ID suppression"
            )
            return

    raise AssertionError("no worksheet XML found in zipped xlsx")
