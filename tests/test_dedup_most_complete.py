"""Spec 2 regression: dedup must keep the *most complete* version of a
duplicate cluster, not just the highest-precedence source.

Pre-fix the dedup walker sorts by ``["_src_pri", "_sort"]`` and runs
``df.duplicated(..., keep="first")`` — so source precedence wins, ties
broken by earliest date.  A record whose source is lower-precedence
but carries richer fields (Invoice #, Reading, Tariff, Standing
Charge, etc.) would lose to a higher-precedence-but-sparser twin.

This pins the spec'd behaviour:

    "duplicates should be assessed and the most complete version of
     the information presented"

After the fix the kept row is whichever duplicate has the most
populated substantive fields.  Source precedence becomes the
tie-breaker when two siblings are equally complete, and date remains
the final tie.

The companion ``test_pass1_dedup_cross_source.py`` fixture keeps
its HTM-wins assertion because across sources there the rows are
*identical* in field-count — only ``Date`` differs — so completeness
is tied and precedence falls through as before.

This file adds the missing complement: a same-period collision
where the lower-precedence row carries MORE data, and asserts the
lower-precedence-but-richer row survives.
"""

from __future__ import annotations

import os
from pathlib import Path

import pytest
from openpyxl import load_workbook

from edf_bill_fetcher.io.writers import export_to_excel


@pytest.fixture
def workdir() -> Path:
    scratch = Path(os.environ.get("USERPROFILE", "/tmp")) / f".edf_dedup_mc_{os.getpid()}"
    scratch.mkdir(parents=True, exist_ok=True)
    return scratch


def _make_records() -> list[dict]:
    """Two duplicates on the same Period-To + Amount, where the
    lower-precedence (PST) row carries richer data than the higher-
    precedence (HTM) row.

    Field-by-field comparison of what the spec treats as substantive:

    | Field                | HTM (high-prec) | PST (low-prec) |
    | -------------------- | --------------- | -------------- |
    | Date                 | 04/04/2024      | 02/04/2024     |
    | Period From          | N/A             | 01/03/2024     |
    | Period To            | 01/04/2024      | 01/04/2024     |
    | Invoice #            | ""              | INV-777        |
    | Period Charge (£)    | 0.0             | 80.0           |
    | Units (kWh)          | ""              | "100"          |
    | Reading              | ""              | "Actual"       |
    | Tariff               | "N/A"           | "Standard Var" |
    | Standing Chg (p/day) | ""              | "45.5"         |

    HTM has 0 substantive fields populated; PST has 6.  Per the spec
    the dedup must keep PST.
    """
    base = {
        "Date": "01/04/2024",
        "Amount (£)": 100.0,
        "Entry Type": "New Bill",
        "Logic Used": "Period",
        "Details": "",
        "Attachment Name": "",
        "Anomaly Flag": "",
        "Sender": "edfenergy.com",
    }
    htm = dict(base)
    htm.update(
        {
            "Date": "04/04/2024",
            "Source": "HTM Account History",
            "Period From": "N/A",
            "Period To": "01/04/2024",
            "Invoice #": "",
            "Money": 100.0,
            "Period Charge (£)": 0.0,
            "Units (kWh)": "",
            "Reading": "",
            "Tariff": "N/A",
            "Standing Charge": "",
        }
    )
    pst = dict(base)
    pst.update(
        {
            "Date": "02/04/2024",
            "Source": "PST PDF Attachment",
            "Period From": "01/03/2024",
            "Period To": "01/04/2024",
            "Invoice #": "INV-777",
            "Money": 100.0,
            "Period Charge (£)": 80.0,
            "Units (kWh)": "100",
            "Reading": "Actual",
            "Tariff": "Standard Var",
            "Standing Charge": "45.5",
        }
    )
    return [htm, pst]


def _config() -> dict:
    return {
        "use_dedup": True,
        "save_dups": True,
        "use_anchors": False,
        "use_large": False,
        "min_amount": 0.0,
        "filter_below": False,
        "use_dedup_period": True,
        "expanded_columns": True,
        "include_charts": False,
        "include_forecast": False,
    }


class TestMostCompleteWins:
    """When two duplicate rows collide, the dedup walker must keep the
    row whose substantive fields are most populated — not the row
    whose source carries the highest precedence.  Source precedence
    becomes the tie-breaker for equally-complete twins.
    """

    def test_richer_lower_precedence_row_survives_dedup(self, workdir: Path) -> None:
        out = workdir / "out_mc.xlsx"
        export_to_excel(_make_records(), str(out), [], config=_config())
        assert out.exists()

        wb = load_workbook(str(out))
        ws = wb["EDF Evidence Report"]
        kept_sources = [
            ws.cell(row=r, column=1).value
            for r in range(2, ws.max_row + 1)
            if ws.cell(row=r, column=1).value
        ]
        # Pre-fix the walker keeps HTM (highest precedence).
        # Post-fix it keeps PST (richest).  Only one survivor.
        assert kept_sources == ["PST PDF Attachment"], (
            f"expected richer PST to survive; kept sources: {kept_sources}"
        )

    def test_dup_sheet_records_kept_against_lower_priority_sibling(self, workdir: Path) -> None:
        """The dropped row must surface in the Duplicate Entries sheet
        so the user (and ombudsman) can audit what was removed —
        per-spec 'never drop without recording'.
        """
        out = workdir / "out_dup.xlsx"
        export_to_excel(_make_records(), str(out), [], config=_config())
        wb = load_workbook(str(out))
        assert "Duplicate Entries" in wb.sheetnames
        ws = wb["Duplicate Entries"]
        dropped_sources = [
            ws.cell(row=r, column=1).value
            for r in range(2, ws.max_row + 1)
            if ws.cell(row=r, column=1).value
        ]
        assert "HTM Account History" in dropped_sources, (
            f"the sparser HTM row must be recorded as a duplicate; got: {dropped_sources}"
        )
