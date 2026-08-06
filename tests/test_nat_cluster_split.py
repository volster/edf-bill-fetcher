"""Spec 2 edge-case regression: rows with unparseable Period To must
NOT silently cluster via NaT-as-equal in pd.duplicated.

Pre-fix the dedup walker built ``_dedup_date`` as
``period_to_dt.where(period_to_dt.notna(), df["_sort"])`` — so when
Period To was unparseable, the fallback was the parsed source Date.
For a no-period PDF row whose ``Date`` is also unparseable (``_sort``
becomes NaT), ``df.duplicated(keep="first")`` collapses NaT-as-equal
across rows, merging unrelated same-Amount events that share the
absence of a parseable date.

Per the spec ('most complete version presentation'), distinct bills
must not be silently merged just because they share an Amount and
both lack a parseable Period To.

Post-fix ``_dedup_date`` stays NaT for unparseable rows; rows with
NaT dedup dates avoid Pass-1's cluster key entirely and route
through Pass-2's no-period bucket logic, which views them as
distinct unless their own (date, amount) overlaps a 60d window.
"""

from __future__ import annotations

import os
from pathlib import Path

import pytest
from openpyxl import load_workbook

from edf_bill_fetcher.io.writers import export_to_excel


@pytest.fixture
def workdir() -> Path:
    scratch = Path(os.environ.get("USERPROFILE", "/tmp")) / f".edf_nat_split_{os.getpid()}"
    scratch.mkdir(parents=True, exist_ok=True)
    return scratch


def _make_records() -> list[dict]:
    """Two distinct bills, both Period-To-unparseable, same Amount.

    Both rows have a place-holder Period To ("garbage" that won't
    parse, not "N/A").  Pre-fix, the fallback to ``_sort`` would
    have collapsed both into a NaT cluster and ``duplicated`` would
    have flagged the second as a duplicate of the first.
    Post-fix, Pass-1 leaves them distinct because ``_dedup_date``
    is NaT for both rows; Pass-2's 60-day window catches them as
    distinct because the row dates themselves differ by 30 days.
    Either way, both rows survive in the kept set.

    Returns ``(records, dup_marker_records)``.  ``dup_marker_records``
    is a separate dict so callers can sanity-check that the two
    rows we're testing are visually distinct (different Sources +
    different Invoice #s).
    """
    base = {
        "Source": "Local PDF Folder",
        "Amount (£)": 100.0,
        "Entry Type": "New Bill",
        "Logic Used": "Amount",
        "Details": "",
        "Attachment Name": "",
        "Anomaly Flag": "",
        "Sender": "",
    }
    rec_a = dict(base)
    rec_a.update(
        {
            "Source": "Local PDF Folder",
            "Date": "01/03/2024",
            "Period From": "",
            "Period To": "nope",  # unparseable, NOT "N/A"
            "Invoice #": "INV-A",
            "Money": 100.0,
            "Period Charge (£)": 0.0,
            "Units (kWh)": "",
            "Reading": "",
            "Tariff": "",
            "Standing Charge": "",
        }
    )
    rec_b = dict(base)
    rec_b.update(
        {
            "Source": "PST PDF Attachment",
            "Date": "01/02/2024",
            "Period From": "",
            "Period To": "definitely-not-a-date",
            "Invoice #": "INV-B",
            "Money": 100.0,
            "Period Charge (£)": 80.0,
            "Units (kWh)": "",
            "Reading": "",
            "Tariff": "",
            "Standing Charge": "",
        }
    )
    return [rec_a, rec_b]


def _config() -> dict:
    return {
        "use_dedup": True,
        "save_dups": False,  # we want both rows kept cleanly
        "use_anchors": False,
        "use_large": False,
        "min_amount": 0.0,
        "filter_below": False,
        "use_dedup_period": True,
        "expanded_columns": True,
        "include_charts": False,
        "include_forecast": False,
    }


class TestNatClusterSplit:
    """Two unrelated same-Amount records with unparseable Period To
    must NOT be silently merged via NaT-as-equal.
    """

    def test_unparseable_period_to_records_survive_separately(self, workdir: Path) -> None:
        out = workdir / "out_nat.xlsx"
        export_to_excel(_make_records(), str(out), [], config=_config())
        wb = load_workbook(out)  # type: ignore[arg-type]
        ws = wb["EDF Evidence Report"]
        sources = [
            ws.cell(row=r, column=1).value
            for r in range(2, ws.max_row + 1)
            if ws.cell(row=r, column=1).value
        ]
        assert sorted(sources) == ["Local PDF Folder", "PST PDF Attachment"], (
            f"two distinct bills with unparseable Period To must not collapse; got {sources}"
        )

    def test_unparseable_records_kept_when_save_dups_true(self, workdir: Path) -> None:
        """Spec: 'never drop without being recorded'.  Either the row
        survives, or it surfaces on the Duplicate Entries sheet with
        its source label intact.
        """
        cfg = _config()
        cfg["save_dups"] = True
        out = workdir / "out_nat_dups.xlsx"
        export_to_excel(_make_records(), str(out), [], config=cfg)
        wb = load_workbook(out)  # type: ignore[arg-type]
        ws_main = wb["EDF Evidence Report"]
        main_sources = [
            ws_main.cell(row=r, column=1).value
            for r in range(2, ws_main.max_row + 1)
            if ws_main.cell(row=r, column=1).value
        ]
        dup_sources: list[str] = []
        if "Duplicate Entries" in wb.sheetnames:
            ws_dup = wb["Duplicate Entries"]
            dup_sources = [
                ws_dup.cell(row=r, column=1).value
                for r in range(2, ws_dup.max_row + 1)
                if ws_dup.cell(row=r, column=1).value
            ]
        # If Pass-1 collapsed to 1 and Pass-2 didn't merge either,
        # then both rows survived.  If Pass-2 collapsed them (e.g. via
        # 60-day window), then exactly one of them would be on the
        # dup sheet.  Either way: total main + dup covers BOTH sources.
        surviving = set(main_sources) | set(dup_sources)
        assert surviving == {"Local PDF Folder", "PST PDF Attachment"}, (
            f"both sources must survive (kept OR dup-sheet); main={main_sources}, dup={dup_sources}"
        )
