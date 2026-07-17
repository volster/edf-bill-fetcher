"""Tests for the SAP sheet writers (Contract History, Meter Readings,
Financial Transactions).
"""

from openpyxl import Workbook

from edf_collector import (
    write_sap_contract_history_sheet,
    write_sap_financial_transactions_sheet,
    write_sap_meter_readings_sheet,
)

# ---------------------------------------------------------------------------
# SAP Contract History writer
# ---------------------------------------------------------------------------

CONTRACT_ROWS = [
    {
        "Contract From": "2016-06-03",
        "Contract To": "2017-09-30",
        "Product Code": "ESC1_FIXED_1B",
        "Product Description": "Fixed for Business 59 to Sep 17",
        "Contract Reason": "Acquisition (New)",
        "Set Up By": "KNOTT1G",
        "Notes": "",
        "Cancelled Flag": "",
        "Source File": "contract.pdf",
    },
    {
        "Contract From": "2017-10-01",
        "Contract To": "2018-12-31",
        "Product Code": "ESC1_EXTENDED SUPPLY",
        "Product Description": "Extended Supply Elec Product SME",
        "Contract Reason": "Renewal",
        "Set Up By": "PRADH1S",
        "Notes": "",
        "Cancelled Flag": "",
        "Source File": "contract.pdf",
    },
]


def test_contract_writer_title_banner_sap_mention() -> None:
    wb = Workbook()
    ws = wb.active
    write_sap_contract_history_sheet(ws, CONTRACT_ROWS)
    val = str(ws.cell(row=1, column=1).value or "")
    assert "SAP" in val or "Kraken" in val, val


def test_contract_writer_account_in_banner() -> None:
    wb = Workbook()
    ws = wb.active
    write_sap_contract_history_sheet(ws, CONTRACT_ROWS, account="A-123")
    val = str(ws.cell(row=1, column=1).value or "")
    assert "A-123" in val


def test_contract_writer_column_header_count() -> None:
    wb = Workbook()
    ws = wb.active
    write_sap_contract_history_sheet(ws, CONTRACT_ROWS)
    nc = ws.max_column
    assert nc == len(CONTRACT_ROWS[0]), f"expected {len(CONTRACT_ROWS[0])} cols, got {nc}"


def test_contract_writer_row_count() -> None:
    wb = Workbook()
    ws = wb.active
    write_sap_contract_history_sheet(ws, CONTRACT_ROWS)
    expected = 1 + 1 + 1 + len(CONTRACT_ROWS)
    assert ws.max_row == expected, f"expected {expected} rows, got {ws.max_row}"


def test_contract_writer_first_body_row_values() -> None:
    wb = Workbook()
    ws = wb.active
    write_sap_contract_history_sheet(ws, CONTRACT_ROWS)
    assert ws.cell(row=4, column=1).value == "2016-06-03"
    assert ws.cell(row=4, column=3).value == "ESC1_FIXED_1B"


def test_contract_writer_empty_rows_still_banner() -> None:
    wb = Workbook()
    ws = wb.active
    write_sap_contract_history_sheet(ws, [])
    assert ws.cell(row=1, column=1).value is not None
    assert ws.max_row >= 3


# ---------------------------------------------------------------------------
# SAP Meter Readings writer
# ---------------------------------------------------------------------------

METER_ROWS = [
    {
        "Scheduled Read Date": "2016-06-03",
        "Meter Read Date": "2016-06-03",
        "Reading (kWh)": "31264",
        "Read Type": "A",
        "Read Source": "Meter reading by utility company",
        "Read Status": "Released by Agent",
        "Meter Read Reason": "Meter reading at move-in",
        "Register": "001",
        "Source File": "meter.pdf",
    },
    {
        "Scheduled Read Date": "2016-07-14",
        "Meter Read Date": "2016-07-14",
        "Reading (kWh)": "34732",
        "Read Type": "E",
        "Read Source": "Automatic estimation - SAP",
        "Read Status": "Billable",
        "Meter Read Reason": "Periodic Meter Reading",
        "Register": "001",
        "Source File": "meter.pdf",
    },
]


def test_meter_writer_title_banner_sap_mention() -> None:
    wb = Workbook()
    ws = wb.active
    write_sap_meter_readings_sheet(ws, METER_ROWS)
    val = str(ws.cell(row=1, column=1).value or "")
    assert "SAP" in val or "Kraken" in val, val


def test_meter_writer_legend_row_present() -> None:
    wb = Workbook()
    ws = wb.active
    write_sap_meter_readings_sheet(ws, METER_ROWS)
    legend_text = str(ws.cell(row=2, column=1).value or "")
    assert "A" in legend_text
    assert "E" in legend_text


def test_meter_writer_row_count() -> None:
    wb = Workbook()
    ws = wb.active
    write_sap_meter_readings_sheet(ws, METER_ROWS)
    expected = 1 + 1 + 1 + len(METER_ROWS)
    assert ws.max_row == expected, f"expected {expected} rows, got {ws.max_row}"
    expected_cols = len(METER_ROWS[0])
    assert ws.max_column == expected_cols


def test_meter_writer_first_body_row_type_is_A() -> None:
    wb = Workbook()
    ws = wb.active
    write_sap_meter_readings_sheet(ws, METER_ROWS)
    assert ws.cell(row=4, column=1).value == "2016-06-03"
    assert ws.cell(row=4, column=4).value == "31264"
    assert ws.cell(row=4, column=5).value == "A"


# ---------------------------------------------------------------------------
# SAP Financial Transactions writer
# ---------------------------------------------------------------------------

FIN_ROWS = [
    {
        "Document No.": "551000421040",
        "Item": "1",
        "Document Date": "2016-07-18",
        "Posting Date": "2016-07-18",
        "Net Due Date": "2016-07-21",
        "Main Transaction": "0100",
        "Sub Transaction": "0020",
        "Transaction Text": "Dr- Consum Billing Receivable",
        "Amount": "436",
        "Clearing Status": "Cleared Item",
        "Clearing Document": "376001212905",
        "Clearing Date": "2020-03-26",
        "Clearing Reason": "Automatic Clearing",
        "Document Type": "IN",
        "Document Type Description": "Energy Invoicing",
        "Source File": "fin.pdf",
    },
]


def test_fin_writer_title_banner() -> None:
    wb = Workbook()
    ws = wb.active
    write_sap_financial_transactions_sheet(ws, FIN_ROWS)
    val = str(ws.cell(row=1, column=1).value or "")
    assert "SAP" in val or "Kraken" in val or "LEDGER" in val.upper(), val


def test_fin_writer_column_count_matches_species() -> None:
    wb = Workbook()
    ws = wb.active
    write_sap_financial_transactions_sheet(ws, FIN_ROWS)
    assert ws.max_column == len(FIN_ROWS[0])


def test_fin_writer_first_body_row_amount_present() -> None:
    wb = Workbook()
    ws = wb.active
    write_sap_financial_transactions_sheet(ws, FIN_ROWS)
    row = 4
    assert ws.cell(row=row, column=1).value == "551000421040"
    assert ws.cell(row=row, column=9).value == "436"
