"""Provenance sheet tests for the evidence workbook.

The workbook carries a "Provenance" sheet documenting how the
evidence was produced: tool version, generation timestamp, account
reference, and the key configuration thresholds used.  This is what
makes a filed submission self-documenting — an ombudsman (or a
reviewer years later) can see exactly which settings produced the
numbers without reverse-engineering the run.

Two behavioural contracts are pinned here:

1.  The sheet exists on the FULL path (analysis sheets rendered).
2.  The sheet exists on the SHORT path (fewer than 2 analysis rows
    triggers the early-exit ``wb.save``) — provenance must not be
    skipped there, because a degraded run is exactly when the
    record of what happened matters most.
"""

from __future__ import annotations

import re
from datetime import datetime
from pathlib import Path
from typing import cast

import pandas as pd
from openpyxl import load_workbook

from edf_bill_fetcher.io.writers import export_to_excel
from edf_bill_fetcher.models.config import ConfigDict


def _read_declared_version() -> str:
    """Return ``[project] version`` from the repo-root pyproject.toml."""
    text = (Path(__file__).resolve().parents[1] / "pyproject.toml").read_text()
    m = re.search(r'^version\s*=\s*"([^"]+)"', text, re.MULTILINE)
    assert m is not None, "pyproject.toml missing [project] version"
    return m.group(1)


def _sample_df() -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "Date": "2024-05-14",
                "Amount (£)": 1200.00,
                "Entry Type": "New Bill",
                "Invoice #": "INV-001",
                "Period From": "01/04/2024",
                "Period To": "30/04/2024",
                "Source": "HTM Account History",
                "Period Charge (£)": 100.00,
                "Units (kWh)": 500,
            },
            {
                "Date": "2024-05-15",
                "Amount (£)": 800.00,
                "Entry Type": "Payment",
                "Invoice #": "INV-002",
                "Period From": "01/04/2024",
                "Period To": "30/04/2024",
                "Source": "PST PDF Attachment",
                "Period Charge (£)": 80.00,
                "Units (kWh)": 400,
            },
            {
                "Date": "2024-05-16",
                "Amount (£)": 1500.00,
                "Entry Type": "New Bill",
                "Invoice #": "INV-003",
                "Period From": "01/05/2024",
                "Period To": "31/05/2024",
                "Source": "HTM Account History",
                "Period Charge (£)": 120.00,
                "Units (kWh)": 600,
            },
        ]
    )


def _config() -> ConfigDict:
    return cast(
        ConfigDict,
        {
            "use_dedup": True,
            "use_back_billing": False,
            "use_reconciliation": False,
            "analysis_min": 500.0,
            "min_amount": 25.0,
            "save_filtered": True,
            "use_sap": False,
            "report_account_ref": "ACC-PROV",
        },
    )


def _provenance_sheet_rows(tmp_path: Path) -> dict[str, str]:
    """Run a full export and return Provenance sheet rows as {label: value}."""
    out = tmp_path / "prov.xlsx"
    df = _sample_df()
    export_to_excel(df, str(out), error_log=[], config=_config())
    wb = load_workbook(out)
    ws = wb["Provenance"]
    rows: dict[str, str] = {}
    for row in ws.iter_rows(values_only=True):
        if row and row[0] not in (None, ""):
            rows[str(row[0])] = "" if len(row) < 2 or row[1] is None else str(row[1])
    return rows


def test_provenance_sheet_exists_and_is_first(tmp_path: object) -> None:
    """Full path: Provenance is present and is the opening tab."""
    out = tmp_path / "prov.xlsx"  # type: ignore[operator]
    export_to_excel(_sample_df(), str(out), error_log=[], config=_config())
    wb = load_workbook(out)
    assert wb.sheetnames[0] == "Provenance", wb.sheetnames


def test_provenance_records_tool_version(tmp_path: object) -> None:
    """The declared package version (not a fallback) is stamped."""
    rows = _provenance_sheet_rows(tmp_path)  # type: ignore[arg-type]
    assert rows.get("Version") == _read_declared_version()


def test_provenance_records_generation_timestamp_utc(tmp_path: object) -> None:
    """The generation time is a parseable UTC ISO timestamp."""
    rows = _provenance_sheet_rows(tmp_path)  # type: ignore[arg-type]
    raw = rows.get("Generated (UTC)")
    assert raw is not None, "missing Generated (UTC) row"
    parsed = datetime.fromisoformat(raw)
    assert parsed.tzinfo is not None
    assert parsed.utcoffset().total_seconds() == 0  # type: ignore[union-attr]


def test_provenance_records_account_reference(tmp_path: object) -> None:
    rows = _provenance_sheet_rows(tmp_path)  # type: ignore[arg-type]
    assert rows.get("Account Reference") == "ACC-PROV"


def test_provenance_records_key_thresholds(tmp_path: object) -> None:
    """Configuration thresholds used for the run are visible."""
    rows = _provenance_sheet_rows(tmp_path)  # type: ignore[arg-type]
    assert rows.get("analysis_min") == "500.0"
    assert rows.get("min_amount") == "25.0"
    assert rows.get("use_dedup") == "True"
    assert rows.get("save_filtered") == "True"


def test_provenance_sheet_present_on_short_data_path(tmp_path: object) -> None:
    """Even when analysis rows < 2 (early-exit save), provenance exists."""
    out = tmp_path / "short.xlsx"  # type: ignore[operator]
    # Single tiny bill → dfc has 1 row → early-exit path.
    df = pd.DataFrame(
        [
            {
                "Date": "2024-05-14",
                "Amount (£)": 50.00,
                "Entry Type": "New Bill",
                "Invoice #": "INV-001",
                "Period From": "01/04/2024",
                "Period To": "30/04/2024",
                "Source": "HTM Account History",
                "Period Charge (£)": 50.00,
                "Units (kWh)": 100,
            }
        ]
    )
    export_to_excel(df, str(out), error_log=[], config=_config())
    wb = load_workbook(out)
    assert "Provenance" in wb.sheetnames, wb.sheetnames
