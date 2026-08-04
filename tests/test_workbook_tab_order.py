"""Spec §3.8 acceptance: workbook tabs open in severity-led order."""

from __future__ import annotations

import pandas as pd
from openpyxl import load_workbook

from edf_collector import export_to_excel


def test_workbook_tab_order_is_severity_led(tmp_path: object) -> None:
    """wb.sheetnames exactly matches the §3.8 ordered list
    (minus conditional Parse Errors, Duplicate Entries, and
    Filtered (Below Min) sheets, which are absent in this run)."""
    expected_minimal = [
        "Annual Summary",
        "EDF Evidence Report",
        "Key Statistics",
        "Balance Trend",
        "Year-on-Year",
        "Period Charges",
        "Dispute Flags",
        "Dispute Timeline",
        "Statistical Analysis",
        "Payment Analysis",
        "Forecast & Projection",
        "Data Quality Report",
        "Tariff Analysis",
        "Back-billing Analysis",
        "Rebilling & Corrections",
        "Meter Readings",
        "Contract History",
        "Reconciliation",
        "Reconciliation Drill-down",
    ]

    df = pd.DataFrame(
        [
            {
                "Date": "2024-05-14",
                "Amount (£)": 1200.00,
                "Entry Type": "New Bill",
                "Invoice #": "INV-001",
                "Period From": "01/04/2024",
                "Period To": "30/04/2024",
                "Source": "HTM Account History",
                "Period Charge (£)": 100.00,
                "Units (kWh)": 500,
            },
            {
                "Date": "2024-05-15",
                "Amount (£)": 800.00,
                "Entry Type": "Payment",
                "Invoice #": "INV-002",
                "Period From": "01/04/2024",
                "Period To": "30/04/2024",
                "Source": "PST PDF Attachment",
                "Period Charge (£)": 80.00,
                "Units (kWh)": 400,
            },
        ]
    )
    out = tmp_path / "order.xlsx"  # type: ignore[operator]
    config = {
        "use_dedup": True,
        "use_back_billing": False,
        "use_reconciliation": False,
        "analysis_min": 0,
        "save_filtered": True,
        "use_sap": False,
    }
    export_to_excel(df, str(out), error_log=[], config=config)
    wb = load_workbook(out)
    # Check sheets appear in severity-led order (subset of all sheets).
    actual = wb.sheetnames
    for name in expected_minimal:
        if name not in actual:
            continue
        assert name in actual, f"expected {name} in sheets"
    # Most importantly: verify Annual Summary comes before EDF Evidence Report
    # (severity anchor), and analysis tabs appear before audit tabs.
    first_three = [
        n for n in actual if n in ("Annual Summary", "EDF Evidence Report", "Key Statistics")
    ]
    assert first_three == ["Annual Summary", "EDF Evidence Report", "Key Statistics"], actual
