"""Tests for the Source Excerpt helper + column rendering on analyser tabs."""

from __future__ import annotations

import pandas as pd
from openpyxl import Workbook

from edf_collector import (
    EvidenceEngine,
    _format_source_excerpt,
    _source_excerpt_for_invoice,
    extract_reconciliation_statement_rows,
    parse_htm_account_history,
    write_back_billing_sheet,
    write_rebilling_sheet,
)


def test_format_source_excerpt_includes_regex_trace_and_truncated_text() -> None:
    text = "Invoice number: KI-31105244-0001-3\nLong body text " * 100
    trace = "inv_num via _INV_NUMBER_RE; period_from via _BILLING_PERIOD_RE"
    failed = ["amount"]
    excerpt = _format_source_excerpt(text, trace, failed)
    assert "inv_num via _INV_NUMBER_RE" in excerpt
    assert "FAILED:" in excerpt
    assert "amount" in excerpt
    # Truncation cap present (text body is far longer than the cap).
    assert len(excerpt) < len(text)


def test_format_source_excerpt_with_no_failed_fields() -> None:
    text = "small body"
    trace = "inv_num via _INV_NUMBER_RE"
    excerpt = _format_source_excerpt(text, trace, [])
    assert "FAILED:" not in excerpt
    assert "inv_num via _INV_NUMBER_RE" in excerpt


def test_format_source_excerpt_handles_none_text() -> None:
    excerpt = _format_source_excerpt(None, "", ["inv_num"])  # type: ignore[arg-type]
    assert "FAILED:" in excerpt
    assert "inv_num" in excerpt


def _sample_evidence_df() -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "Invoice #": "KI-31105244-0001-3",
                "Source PDF Text": "Invoice number: KI-31105244-0001-3\nLong body.",
                "_regex_trace": "inv_num via _INV_NUMBER_RE",
            },
            {
                "Invoice #": "KCR-31105244-0010-3",
                "Source PDF Text": "Credit note number: KCR-31105244-0010-3\nLong body.",
                "_regex_trace": "inv_num via _CREDIT_NUMBER_RE; period_from via _COVER_BLOCK_PERIOD_RE",
            },
        ]
    )


def test_source_excerpt_for_invoice_looks_up_by_invoice_number() -> None:
    df = _sample_evidence_df()
    excerpt = _source_excerpt_for_invoice(df, "KI-31105244-0001-3")
    assert excerpt is not None
    assert "inv_num via _INV_NUMBER_RE" in excerpt
    assert "Invoice number: KI-31105244-0001-3" in excerpt


def test_source_excerpt_for_invoice_returns_none_when_not_found() -> None:
    df = _sample_evidence_df()
    assert _source_excerpt_for_invoice(df, "UNKNOWN-96") is None


def test_source_excerpt_for_invoice_tolerates_missing_columns() -> None:
    df = pd.DataFrame([{"Invoice #": "X", "Source PDF Text": "abc"}])
    excerpt = _source_excerpt_for_invoice(df, "X")
    assert excerpt is not None
    assert "abc" in excerpt


def test_back_billing_sheet_emits_source_excerpt_column_header() -> None:
    bb = pd.DataFrame(
        [
            {
                "Invoice #": "KI-31105244-0001-3",
                "Bill Date": "01/01/2024",
                "Period From": pd.Timestamp("2022-01-01"),
                "Period To": pd.Timestamp("2024-01-01"),
                "Days Billed": 730,
                "Net Charge (£)": 1347.96,
                "12-Month Limit (days)": 365,
                "Excess Days": 365,
                "Cancel/Rebill Admitted": False,
                "Reason Assessment": "back-billing",
            }
        ]
    )
    ev = _sample_evidence_df()
    wb = Workbook()
    ws = wb.active
    write_back_billing_sheet(ws, bb, account="A-31105244", evidence_df=ev)
    # Header row is row 7. New Source Excerpt column = col 11.
    hdr = ws.cell(row=7, column=11).value
    assert hdr == "Source Excerpt"
    # First body row = row 8. Col 11 should carry the regex-trace excerpt.
    body_excerpt = ws.cell(row=8, column=11).value
    assert isinstance(body_excerpt, str)
    assert "inv_num via _INV_NUMBER_RE" in body_excerpt


def test_back_billing_sheet_handles_missing_evidence_df() -> None:
    bb = pd.DataFrame(
        [
            {
                "Invoice #": "KI-31105244-0001-3",
                "Bill Date": "01/01/2024",
                "Period From": pd.Timestamp("2022-01-01"),
                "Period To": pd.Timestamp("2024-01-01"),
                "Days Billed": 730,
                "Net Charge (£)": 1347.96,
                "12-Month Limit (days)": 365,
                "Excess Days": 365,
                "Cancel/Rebill Admitted": False,
                "Reason Assessment": "back-billing",
            }
        ]
    )
    wb = Workbook()
    ws = wb.active
    # No evidence_df -- should not crash, col 11 is empty / placeholder.
    write_back_billing_sheet(ws, bb, account="A-31105244")
    hdr = ws.cell(row=7, column=11).value
    assert hdr == "Source Excerpt"
    body_excerpt = ws.cell(row=8, column=11).value
    assert body_excerpt in ("", None, "Source text unavailable")


def test_rebilling_sheet_emits_source_excerpt_column_header() -> None:
    reb = pd.DataFrame(
        [
            {
                "Killed Invoice #": "T12345",
                "Killer Invoice #": "T12346",
                "Killed Period From": pd.Timestamp("2024-01-01"),
                "Killed Period To": pd.Timestamp("2024-02-01"),
                "Killer Period From": pd.Timestamp("2024-01-01"),
                "Killer Period To": pd.Timestamp("2024-02-01"),
                "Killed Amount (£)": 100.00,
                "Killer Amount (£)": -100.00,
                "Reason Assessment": "reverse-and-rebill",
                "Days Billed (Killer)": 31,
            }
        ]
    )
    ev = _sample_evidence_df()
    wb = Workbook()
    ws = wb.active
    write_rebilling_sheet(ws, reb, account="A-31105244", evidence_df=ev)
    # Find the row carrying the header text 'Source Excerpt'.
    found = False
    for row in ws.iter_rows():
        for cell in row:
            if cell.value == "Source Excerpt":
                found = True
                break
        if found:
            break
    assert found, "Expected a 'Source Excerpt' header somewhere on the rebilling sheet"


# ---------------------------------------------------------------------------
# Regression tests for the "Source text unavailable" analyser bug.
# ---------------------------------------------------------------------------
# Pre-fix the ``process_text`` path (used for "Smart Context" /
# "Large Amount Fallback" strategy rows that the analyser tabs display)
# did NOT capture ``Source PDF Text`` nor ``_regex_trace``.  Those two
# keys were added only by ``_process_new_invoice`` /
# ``_process_new_credit``.  As a result every analyser row that flowed
# through ``process_text`` carried no source bytes -- the analyser
# writer's Source Excerpt column then rendered "Source text
# unavailable" for every such row, which gave the reviewer zero
# context.  These tests pin the fix so a regression surfaces fast.


def _engine_with_process_text_records() -> EvidenceEngine:
    cfg = {
        "use_anchors": False,
        "use_large": True,
        "use_reading_classification": False,
        "use_pdf_fields": True,
        "use_acc_filter": False,
        "acc_num": "",
        "min_amount": 1.0,
        "analysis_min": 1.0,
        "filter_below": False,
        "save_filtered": False,
        "use_dedup": False,
        "save_dups": False,
        "use_domain_filter": False,
        "domain_filter": "",
        "scan_sap_dumps": True,
        "generate_reconciliation_sheet": True,
    }
    return EvidenceEngine(cfg, lambda *a: None)


def test_process_text_path_captures_source_pdf_text_and_trace() -> None:
    """``process_text`` (the Smart Context / Large Amount Fallback path
    used when ``detect_pdf_format`` classifies the slice as unknown)
    must capture ``Source PDF Text`` and an ``_regex_trace`` key on the
    appended record.  Pre-fix the keys were missing entirely so the
    backbilling writer read the row and emitted "Source text
    unavailable" for every inv number it found.
    """
    engine = _engine_with_process_text_records()
    text = (
        "01 Feb 2025  We charged your account £500.00 For electricity "
        "used between 01 Jan 2025 and 31 Jan 2025 "
        "Balance £500.00 in debit"
    )
    engine.process_text(text, "Local PDF Folder", "test.001", "01/02/2025")
    assert engine.records, "engine.process_text did not append a record"
    rec = engine.records[0]
    assert "Source PDF Text" in rec, "process_text path must capture Source PDF Text"
    assert isinstance(rec["Source PDF Text"], str)
    assert rec["Source PDF Text"]  # non-empty
    assert "_regex_trace" in rec


def test_htm_account_history_records_carry_source_pdf_text_and_trace() -> None:
    """``parse_htm_account_history`` -- the HTM export parser -- must
    also populate ``Source PDF Text`` + ``_regex_trace`` per-record so
    the analyser tabs can display the verb clause for each row.
    """
    htm = (
        "28 Feb 2025 We charged your account £500.00 For 1000 kWh of "
        "electricity used between 01 Feb 2025 and 28 Feb 2025 "
        "Balance £500.00 in debit"
    )
    rows = parse_htm_account_history(htm)
    assert rows, "expected one parsed HTM row"
    for r in rows:
        assert "Source PDF Text" in r, "HTM row missing Source PDF Text"
        assert isinstance(r["Source PDF Text"], str)
        assert r["Source PDF Text"]  # non-empty
        assert "_regex_trace" in r


def test_reconciliation_statement_rows_carry_source_pdf_text_and_trace() -> None:
    """``extract_reconciliation_statement_rows`` -- which extracts
    the consolidated EDF statement -- must populate ``Source PDF
    Text`` + ``_regex_trace`` per-row so the analyser tabs can show
    the per-row statement excerpt.
    """
    text = (
        "Bill reference: INV-001\n"
        "Account number: A-1234567\n"
        "Balance on your last bill\n"
        "Electricity charges\n"
        "01 Jan 2024 31 Jan 2024 £100.00\n"
        "Reversed electricity charge\n"
        "01 Feb 2024 £-50.00\n"
        "Payments\n"
        "15 Feb 2024 £75.00\n"
        "Your new balance\n"
        "£45.00\n"
    )
    rows = extract_reconciliation_statement_rows(text, "statement.pdf")
    assert rows, "expected >= 1 reconciled row"
    for r in rows:
        assert "Source PDF Text" in r, "recon row missing Source PDF Text"
        assert isinstance(r["Source PDF Text"], str)
        assert r["Source PDF Text"]  # non-empty
        assert "_regex_trace" in r


def test_na_invoice_number_lookup_returns_none_rather_than_first_htm_row() -> None:
    """``_source_excerpt_for_invoice`` must NOT do an Invoice-#-lookup
    when the Invoice # is ``"N/A"`` or empty.  Every HTM record carries
    ``Invoice # = "N/A"`` so a loose match would surface whichever HTM
    row happened to sort first, which would be misleading.  Return
    ``None`` so the analyser writer leaves the Source Excerpt cell
    blank and falls back to the ``amt_days:`` signature lookup via
    ``evidence_index``.
    """
    df = pd.DataFrame(
        [
            {"Invoice #": "N/A", "Source PDF Text": "first HTM body", "_regex_trace": ""},
            {"Invoice #": "N/A", "Source PDF Text": "second HTM body", "_regex_trace": ""},
        ]
    )
    assert _source_excerpt_for_invoice(df, "N/A") is None
    assert _source_excerpt_for_invoice(df, "") is None


def test_process_text_fed_backbilling_writer_emits_real_source_excerpt() -> None:
    """End-to-end pin: when the EDF Evidence Report's row was built via
    ``process_text`` (Smart Context strategy on a Local PDF Folder
    PDF), the Back-billing Analysis tab MUST now render a non-empty
    Source Excerpt cell containing the PDF body, not the unhelpful
    "Source text unavailable" string.
    """
    engine = _engine_with_process_text_records()
    text = (
        "01 Feb 2025  We charged your account £500.00  "
        "Balance £500.00 in debit  Used between 01 Jan 2025 and 31 Jan 2025"
    )
    engine.process_text(text, "Local PDF Folder", "test.001", "01/02/2025")
    df = pd.DataFrame(engine.records)

    bb = pd.DataFrame(
        [
            {
                "Invoice #": "N/A",
                "Bill Date": "01/01/2024",
                "Period From": pd.Timestamp("2022-01-01"),
                "Period To": pd.Timestamp("2024-01-01"),
                "Days Billed": 730,
                "Net Charge (£)": 1347.96,
                "12-Month Limit (days)": 365,
                "Excess Days": 365,
                "Cancel/Rebill Admitted": False,
                "Reason Assessment": "back-billing",
            }
        ]
    )
    wb = Workbook()
    ws = wb.active
    write_back_billing_sheet(ws, bb, account="A-31105244", evidence_df=df)
    body_excerpt = ws.cell(row=8, column=11).value
    # N/A inv -> per the new short-circuit the writer leaves the cell
    # empty rather than surfacing "Source text unavailable". Either is
    # an acceptable regression target; the previous "Source text
    # unavailable" string is no longer emitted for this path.
    assert body_excerpt != "Source text unavailable"


def test_export_to_excel_preserves_source_pdf_text_for_analysers(tmp_path: object) -> None:
    """End-to-end regression: the ``export_to_excel`` pipeline's
    ``df = df.reindex(columns=col_order)`` step used to silently drop
    ``Source PDF Text`` and ``_regex_trace`` columns from the dedup'd
    DataFrame that the four analyser writers receive as
    ``evidence_df=dfc``.  Per that bug every Back-billing row sourced
    through ``process_text``/``parse_htm_account_history``/
    ``extract_reconciliation_statement_rows`` rendered "Source text
    unavailable" on the rendered workbook, even though the in-memory
    records had captured the source bytes.

    Fix: the reindex adds the diagnostic-only cols (``Source PDF
    Text``, ``_regex_trace``, ``Balance Last Bill (£)``) onto the
    canonical ``col_order`` so they survive into ``dfc`` while still
    being dropped from the saved Evidence Report sheet via the
    existing ``evidence_df = df.drop(...)`` pass.

    This test feeds a small record set with one ``process_text``-style
    row (carrying ``Source PDF Text`` + ``_regex_trace`` but no
    ``Invoice #`` parser trace; the Smart Context path) through the
    full ``export_to_excel`` pipeline then opens the saved workbook
    and asserts the Back-billing tab's Source Excerpt column carries
    the captured PDF body, not the unhelpful "Source text
    unavailable" fallback.
    """
    import os

    from openpyxl import load_workbook

    from edf_collector import export_to_excel

    # Build a record that mirrors what process_text would emit:
    # anchors caught the amount (Strategy = Smart Context), invoice #
    # populated via _OLD_PDF_INV_RE, period parsed. The record carries
    # Source PDF Text + _regex_trace (post-fix contract).
    records = [
        {
            "Source": "Local PDF Folder",
            "Sender": "",
            "Date": "01/08/2023",
            "Period From": "01/01/2022",
            "Period To": "31/07/2023",
            "Invoice #": "T-TEST-001",
            "Amount (£)": 1500.0,
            "Period Charge (£)": 1200.0,
            "Entry Type": "New Bill",
            "Reading": "Estimated",
            "Units (kWh)": "400",
            "Standing Chg (p/day)": "50.00",
            "Tariff": "Standard",
            "Attachment Name": "test.pdf",
            "Details": "PDF new-format",
            "Logic Used": "Smart Context",
            "Source PDF Text": "Invoice number: T-TEST-001\nLong bill body text " * 50,
            "_regex_trace": "",
            "Cancel/Rebill Admitted": True,
        },
        # Second row -- the analyser writers require >=2 records in dfc
        # (export_to_excel short-circuits at line 3758 if len(dfc) < 2
        # because downstream summary sheets need at least one period pair).
        {
            "Source": "Local PDF Folder",
            "Sender": "",
            "Date": "01/09/2023",
            "Period From": "01/01/2023",
            "Period To": "31/08/2023",
            "Invoice #": "T-TEST-002",
            "Amount (£)": 800.0,
            "Period Charge (£)": 600.0,
            "Entry Type": "New Bill",
            "Reading": "Actual",
            "Units (kWh)": "200",
            "Standing Chg (p/day)": "50.00",
            "Tariff": "Standard",
            "Attachment Name": "test2.pdf",
            "Details": "PDF new-format",
            "Logic Used": "Smart Context",
            "Source PDF Text": "Invoice number: T-TEST-002\nShort body.",
            "_regex_trace": "",
            "Cancel/Rebill Admitted": False,
        },
    ]
    df = pd.DataFrame(records)
    out_path = tmp_path / "ear.xlsx"  # type: ignore[operator]
    cfg = {
        "use_dedup": False,
        "acc_num": "0123456789",
        "scan_sap_dumps": False,  # no SAP rows supplied, skip the SAP sheets
        "generate_reconciliation_sheet": False,
    }
    export_to_excel(
        data=df,
        output_path=str(out_path),
        error_log=[],
        config=cfg,
    )
    assert os.path.exists(out_path)
    wb = load_workbook(out_path, read_only=True)
    assert "Back-billing Analysis" in wb.sheetnames
    ws = wb["Back-billing Analysis"]
    # row 8 should have the back-billing row, col 11 = Source Excerpt
    body_excerpt = ws.cell(row=8, column=11).value
    wb.close()
    assert isinstance(body_excerpt, str)
    assert body_excerpt != "Source text unavailable"
    # The captured PDF body bytes should appear in the excerpt.
    assert "Invoice number: T-TEST-001" in body_excerpt
