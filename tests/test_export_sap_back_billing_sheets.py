"""End-to-end tests for the new SAP Back-billing sheets (spec §9.1).

Verifies the workbook build wires ``detect_sap_back_billing_events``
+ ``match_sap_events_to_edf`` + ``write_sap_back_billing_sheets`` into
``export_to_excel`` under the existing ``scan_sap_dumps`` toggle.

Sheet-level properties pinned here:
  - both new sheets exist when sap_toggle is on, absent when off
  - Sheet 1 summary rows have outline level 0; underlying rows have
    outline level 1 and are hidden by default
  - Sheet 1 hyperlinks point to specific rows on the source SAP sheet
  - Sheet 2 only contains matched pairs
  - Sheet 2 carries hyperlinks to both Sheet 1 and EDF Evidence Report
"""

from __future__ import annotations

from openpyxl import Workbook, load_workbook

from edf_bill_fetcher.helpers.excel_utils import build_sap_row_index_map
from edf_bill_fetcher.io.writers import export_to_excel, write_sap_back_billing_sheets
from edf_bill_fetcher.processors.matching import match_sap_events_to_edf
from edf_bill_fetcher.processors.sap_parsers import parse_sap_financial_transactions
from edf_bill_fetcher.writers._helpers import detect_sap_back_billing_events

# ---------------------------------------------------------------------------
# Helper builders
# ---------------------------------------------------------------------------


def _sap_csv_with_cluster(
    clear_doc: str = "CL1",
    clear_date: str = "26-03-2020",
    rows: list[tuple[str, str, str, str]] | None = None,
) -> str:
    """Build a synthetic SAP CSV with the 32 source columns; defaults
    to 4 rows clustered on the same Clearing Document (≥ min_cluster_size).
    """
    if rows is None:
        rows = [
            ("DOC1", "1234.56", "Dr- Consum Billing Receivable", ""),
            ("DOC2", "-1234.56", "Cr- Credit for Consum Billing", ""),
            ("DOC3", "0.00", "Dr- Consum Billing Receivable", ""),
            ("DOC4", "0.00", "Dr- Consum Billing Receivable", ""),
        ]
    cols = [
        "Kraken ID",
        "SAP Account Number",
        "Business Partner",
        "Account Determination ID",
        "Contract",
        "Fuel Type",
        "Document No.",
        "Item",
        "Sub Item",
        "Payment Method",
        "Document Date",
        "Posting Date ",
        "Net Due Date",
        "Clearing Status",
        "Main Transactions",
        "Sub Transactions",
        "Transaction Text",
        "Amount",
        "Down Payment Flag",
        "Statistical Key Flag",
        "Clearing Document",
        "Clearing Date",
        "Clearing Reason",
        "Clearing Posting Date",
        "Clearing Amount",
        "Restriction",
        "Document Type",
        "Document Type Description",
        "Tax Code",
        "Tax Code Description",
        "G/L Account",
        "G/L Description",
        "Deferral Date",
    ]
    header = '"' + '","'.join(cols) + '"'
    body_lines = [header]
    for _i, (doc, amt, txt, stat_flag) in enumerate(rows):
        line_vals = [
            "A-31105244",
            "671078701920",
            "0159628206",
            "Non-residential customers",
            "2011040650",
            "Electricity",
            doc,
            "1",
            "0",
            "",
            "01-01-2020",
            "01-01-2020",
            "01-02-2020",
            "Cleared Item",
            "0100",
            "0020",
            txt,
            amt,
            "No",
            stat_flag,  # Statistical Key Flag
            clear_doc,
            clear_date,
            "Automatic Clearing",
            clear_date,
            amt,
            "No restriction",
            "IN",
            "Energy Invoicing",
            "A4",
            "Donations or payment for equity funds",
            "0000210251",
            "Billed Debtor SME Elec",
            "",
        ]
        body_lines.append('"' + '","'.join(str(v) for v in line_vals) + '"')
    return "\n".join(body_lines)


def _sample_data_one_record() -> list[dict]:
    """Two invoices — export_to_excel early-exits with fewer than 2 records
    at edf_collector.py:4304, so we always provide at least two.
    """
    base = {
        "Source": "Local PDF Folder",
        "Sender": "edf.co.uk",
        "Date": "01 Aug 2023",
        "Period From": "02/10/2020",
        "Period To": "09/08/2023",
        "Invoice #": "T78",
        "Amount (£)": 1234.56,
        "Period Charge (£)": 100.0,
        "Unit Rate (p/kWh)": "",
        "% Change": "",
        "Entry Type": "New Bill",
        "Reading": "Actual",
        "Units (kWh)": "",
        "Standing Chg (p/day)": "",
        "Tariff": "Standard",
        "Attachment Name": "test.pdf",
        "Details": "",
        "Logic Used": "PDF new-format",
        "Anomaly Flag": "",
        "Cancel/Rebill Admitted": False,
    }
    second = dict(base)
    second["Invoice #"] = "T79"
    second["Period From"] = "10/08/2023"
    second["Period To"] = "10/09/2023"
    second["Date"] = "10/09/2023"
    return [base, second]


# ---------------------------------------------------------------------------
# Unit tests on the wiring helpers
# ---------------------------------------------------------------------------


def test_build_sap_row_index_map_returns_4_plus_i() -> None:
    rows = parse_sap_financial_transactions(_sap_csv_with_cluster(), source_file="x")
    assert len(rows) == 4
    mp = build_sap_row_index_map(rows)
    assert mp[id(rows[0])] == 4
    assert mp[id(rows[1])] == 5
    assert mp[id(rows[2])] == 6
    assert mp[id(rows[3])] == 7


# ---------------------------------------------------------------------------
# Workbook-level tests: end-to-end via export_to_excel
# ---------------------------------------------------------------------------


def test_export_to_excel_emits_both_new_sheets_when_sap_toggle_on(
    tmp_path: object,
) -> None:
    sap_rows = parse_sap_financial_transactions(_sap_csv_with_cluster(), source_file="test.pdf")
    out = str(tmp_path / "wb.xlsx")  # type: ignore[operator]
    export_to_excel(
        data=_sample_data_one_record(),
        output_path=out,
        error_log=[],
        config={
            "use_dedup": False,
            "acc_num": "0123456789",
            "scan_sap_dumps": True,
            "generate_reconciliation_sheet": True,
        },
        sap_rows={
            "financial": sap_rows,
            "contract": [],
            "meter": [],
        },
    )
    wb = load_workbook(out, read_only=True)
    names = set(wb.sheetnames)
    wb.close()
    assert "SAP Back-billing Events" in names, names
    assert "SAP ↔ EDF Matched Events" in names, names


def test_export_to_excel_omits_both_new_sheets_when_sap_toggle_off(
    tmp_path: object,
) -> None:
    sap_rows = parse_sap_financial_transactions(_sap_csv_with_cluster(), source_file="test.pdf")
    out = str(tmp_path / "wb.xlsx")  # type: ignore[operator]
    export_to_excel(
        data=_sample_data_one_record(),
        output_path=out,
        error_log=[],
        config={
            "use_dedup": False,
            "acc_num": "0123456789",
            "scan_sap_dumps": False,
            "generate_reconciliation_sheet": True,
        },
        sap_rows={
            "financial": sap_rows,
            "contract": [],
            "meter": [],
        },
    )
    wb = load_workbook(out, read_only=True)
    names = set(wb.sheetnames)
    wb.close()
    assert "SAP Back-billing Events" not in names
    assert "SAP ↔ EDF Matched Events" not in names


# ---------------------------------------------------------------------------
# Sheet 1 structure: outline groups + collapsed + summary tinting
# ---------------------------------------------------------------------------


def test_sheet1_summary_rows_have_outline_level_zero(tmp_path: object) -> None:
    wb = Workbook()
    events = detect_sap_back_billing_events(
        parse_sap_financial_transactions(_sap_csv_with_cluster(), source_file="test.pdf")
    )
    assert len(events) == 1, "fixture should produce exactly 1 event"
    write_sap_back_billing_sheets(
        wb,
        events,
        [],
        sap_financial_first_row=4,
        edf_rows=[],
        sap_row_index_map={id(events[0].rows[0]): 4},
    )
    ws1 = wb["SAP Back-billing Events"]
    # Spec §3.3 — header moved to row 3, so the first summary row is row 4
    assert ws1.row_dimensions[4].outline_level == 0
    # rows 5-8 are the 4 underlying rows for the single event
    for r in range(5, 9):
        assert ws1.row_dimensions[r].outline_level == 1, f"row {r}"


def test_sheet1_underlying_rows_are_hidden_by_default(tmp_path: object) -> None:
    wb = Workbook()
    events = detect_sap_back_billing_events(
        parse_sap_financial_transactions(_sap_csv_with_cluster(), source_file="test.pdf")
    )
    write_sap_back_billing_sheets(
        wb,
        events,
        [],
        sap_financial_first_row=4,
        edf_rows=[],
        sap_row_index_map={id(events[0].rows[0]): 4},
    )
    ws1 = wb["SAP Back-billing Events"]
    for r in range(5, 9):
        assert ws1.row_dimensions[r].hidden is True, f"row {r} should be hidden"


def test_sheet1_summary_rows_match_event_count(tmp_path: object) -> None:
    """Spec §9.1: the count of outline-level-0 summary rows in the body
    must equal the count of `events` the writer was given."""
    # Build a two-cluster fixture: two distinct Clearing Documents.
    csv = (
        _sap_csv_with_cluster(
            clear_doc="CL1",
            rows=[
                ("DOC1", "100", "Dr- Consum Billing Receivable", ""),
                ("DOC2", "-100", "Cr- Credit for Consum Billing", ""),
                ("DOC3", "0", "Dr- Consum Billing Receivable", ""),
                ("DOC4", "0", "Dr- Consum Billing Receivable", ""),
            ],
        )
        + "\n"
        + _sap_csv_with_cluster(
            clear_doc="CL2",
            rows=[
                ("DOC5", "50", "Dr- Consum Billing Receivable", ""),
                ("DOC6", "-50", "Cr- Credit for Consum Billing", ""),
                ("DOC7", "0", "Dr- Consum Billing Receivable", ""),
                ("DOC8", "0", "Dr- Consum Billing Receivable", ""),
            ],
        )
    )
    wb = Workbook()
    sap_rows = parse_sap_financial_transactions(csv, source_file="test.pdf")
    events = detect_sap_back_billing_events(sap_rows)
    assert len(events) == 2, f"fixture should produce 2 events, got {len(events)}"
    write_sap_back_billing_sheets(
        wb,
        events,
        [],
        sap_financial_first_row=4,
        edf_rows=[],
        sap_row_index_map={},  # links left blank; we're only counting summary rows
    )
    ws1 = wb["SAP Back-billing Events"]
    # Body summary rows live at outline_level 0; row 3 is the header (also
    # outline_level 0) so exclude it by also requiring col A to be a
    # clearing-doc number (not the header label "Clearing Doc #").
    summary_count = 0
    for r in range(4, ws1.max_row + 1):
        if ws1.row_dimensions[r].outline_level == 0:
            v = ws1.cell(row=r, column=1).value
            if v and str(v).strip() not in ("", "Clearing Doc #"):
                summary_count += 1
    assert summary_count == len(events), (
        f"expected {len(events)} summary rows, found {summary_count}"
    )


def test_sheet1_summary_row_hyperlink_to_specific_sap_row(tmp_path: object) -> None:
    wb = Workbook()
    sap_rows = parse_sap_financial_transactions(_sap_csv_with_cluster(), source_file="test.pdf")
    events = detect_sap_back_billing_events(sap_rows)
    mp = build_sap_row_index_map(sap_rows)
    write_sap_back_billing_sheets(
        wb,
        events,
        [],
        sap_financial_first_row=4,
        edf_rows=[],
        sap_row_index_map=mp,
    )
    ws1 = wb["SAP Back-billing Events"]
    c12 = ws1.cell(row=4, column=12)
    assert c12.hyperlink is not None
    loc = c12.hyperlink.location
    assert loc == "'SAP Financial Transactions'!A4", loc


# ---------------------------------------------------------------------------
# PR #3 — Spec §3.3 (issue 3): SAP BB Events legal block deletion
# ---------------------------------------------------------------------------


def test_sap_bb_events_no_legal_context_row_present() -> None:
    wb = Workbook()
    events = detect_sap_back_billing_events(
        parse_sap_financial_transactions(_sap_csv_with_cluster(), source_file="test.pdf")
    )
    write_sap_back_billing_sheets(
        wb,
        events,
        [],
        sap_financial_first_row=4,
        edf_rows=[],
        sap_row_index_map={},
    )
    ws1 = wb["SAP Back-billing Events"]
    for r in range(1, ws1.max_row + 1):
        v = ws1.cell(row=r, column=1).value
        if v is None:
            continue
        s = str(v).upper()
        assert "LEGAL CONTEXT" not in s, f"row {r} still contains LEGAL CONTEXT"
        assert "Back-billing protections" not in s, f"row {r} has the legal blurb"


def test_sap_bb_events_no_intro_paragraph_on_row_5() -> None:
    wb = Workbook()
    events = detect_sap_back_billing_events(
        parse_sap_financial_transactions(_sap_csv_with_cluster(), source_file="test.pdf")
    )
    write_sap_back_billing_sheets(
        wb,
        events,
        [],
        sap_financial_first_row=4,
        edf_rows=[],
        sap_row_index_map={},
    )
    ws1 = wb["SAP Back-billing Events"]
    v = ws1.cell(row=5, column=1).value
    assert v is None or not str(v).startswith("Each row below identifies"), (
        f"row 5 still has the intro paragraph starting with 'Each row below': {v!r}"
    )


def test_sap_bb_events_title_row_contains_event_count_summary() -> None:
    import re

    wb = Workbook()
    events = detect_sap_back_billing_events(
        parse_sap_financial_transactions(_sap_csv_with_cluster(), source_file="test.pdf")
    )
    # Reach the writer directly so the test is hermetic — the public
    # write_sap_back_billing_sheets wrapper doesn't yet forward account.
    from edf_bill_fetcher.io.writers.sap import _write_sap_bb_events_sheet

    ws = wb.create_sheet("SAP Back-billing Events")
    _write_sap_bb_events_sheet(ws, events, sap_financial_first_row=4, account="A-31105244")
    title = str(ws.cell(row=1, column=1).value)
    assert re.search(r"\d+ events \(\d+ net-zero, \d+ with credit\)", title), title


def test_sap_bb_events_header_row_moved_to_row_3() -> None:
    wb = Workbook()
    events = detect_sap_back_billing_events(
        parse_sap_financial_transactions(_sap_csv_with_cluster(), source_file="test.pdf")
    )
    write_sap_back_billing_sheets(
        wb,
        events,
        [],
        sap_financial_first_row=4,
        edf_rows=[],
        sap_row_index_map={},
    )
    ws1 = wb["SAP Back-billing Events"]
    assert ws1.cell(row=3, column=1).value == "Clearing Doc #"
    assert ws1.cell(row=4, column=1).value == "CL1"  # single-cluster fixture summary row


# ---------------------------------------------------------------------------
# Sheet 2 structure: matched pairs only + both-target hyperlinks
# ---------------------------------------------------------------------------


def test_sheet2_only_includes_matched_pairs(tmp_path: object) -> None:
    # Cluster A matches EDF inv "T78" (period-end match + amount within 5%);
    # cluster B is far in time from any EDF invoice and yields no match.
    csv_a = _sap_csv_with_cluster(
        clear_doc="CA",
        clear_date="09-08-2023",
        rows=[
            ("DA1", "28000", "Dr- Consum Billing Receivable", ""),
            ("DA2", "0", "Dr- Consum Billing Receivable", ""),
            ("DA3", "0", "Dr- Consum Billing Receivable", ""),
            ("DA4", "0", "Dr- Consum Billing Receivable", ""),
        ],
    )
    csv_b = _sap_csv_with_cluster(
        clear_doc="CB",
        clear_date="01-01-2010",
        rows=[
            ("DB1", "1", "Dr- Consum Billing Receivable", ""),
            ("DB2", "0", "Dr- Consum Billing Receivable", ""),
            ("DB3", "0", "Dr- Consum Billing Receivable", ""),
            ("DB4", "0", "Dr- Consum Billing Receivable", ""),
        ],
    )
    rows_a = parse_sap_financial_transactions(csv_a, source_file="a")
    rows_b = parse_sap_financial_transactions(csv_b, source_file="b")
    rows = rows_a + rows_b
    events = detect_sap_back_billing_events(rows)
    edf = _sample_data_one_record()
    matches = match_sap_events_to_edf(events, edf)
    # cluster B should produce NO matches; cluster A produces at least one
    matched_docs = {m.event.clearing_doc for m in matches}
    assert "CA" in matched_docs
    assert "CB" not in matched_docs


def test_sheet2_hyperlinks_to_sheet1_and_evidence_report(tmp_path: object) -> None:
    wb = Workbook()
    csv_a = _sap_csv_with_cluster(
        clear_doc="CA",
        clear_date="09-08-2023",
        rows=[
            ("DA1", "28000", "Dr- Consum Billing Receivable", ""),
            ("DA2", "0", "Dr- Consum Billing Receivable", ""),
            ("DA3", "0", "Dr- Consum Billing Receivable", ""),
            ("DA4", "0", "Dr- Consum Billing Receivable", ""),
        ],
    )
    sap_rows = parse_sap_financial_transactions(csv_a, source_file="a")
    events = detect_sap_back_billing_events(sap_rows)
    edf = _sample_data_one_record()
    matches = match_sap_events_to_edf(events, edf)
    mp = build_sap_row_index_map(sap_rows)
    write_sap_back_billing_sheets(
        wb,
        events,
        matches,
        sap_financial_first_row=4,
        edf_rows=edf,
        sap_row_index_map=mp,
    )
    ws2 = wb["SAP ↔ EDF Matched Events"]
    # row 6 is the first matched row on sheet 2
    c1 = ws2.cell(row=6, column=1)
    c4 = ws2.cell(row=6, column=4)
    assert c1.hyperlink is not None
    assert "SAP Back-billing Events" in (c1.hyperlink.location or "")
    assert c4.hyperlink is not None
    assert "EDF Evidence Report" in (c4.hyperlink.location or "")


# ---------------------------------------------------------------------------
# Sheet ordering
# ---------------------------------------------------------------------------


def test_export_to_excel_orders_new_sheets_before_reconciliation(
    tmp_path: object,
) -> None:
    sap_rows = parse_sap_financial_transactions(_sap_csv_with_cluster(), source_file="test.pdf")
    out = str(tmp_path / "wb.xlsx")  # type: ignore[operator]
    export_to_excel(
        data=_sample_data_one_record(),
        output_path=out,
        error_log=[],
        config={
            "use_dedup": False,
            "acc_num": "0123456789",
            "scan_sap_dumps": True,
            "generate_reconciliation_sheet": True,
        },
        sap_rows={"financial": sap_rows, "contract": [], "meter": []},
    )
    wb = load_workbook(out, read_only=True)
    sheets = wb.sheetnames
    wb.close()
    i_ft = sheets.index("SAP Financial Transactions")
    i_bb = sheets.index("SAP Back-billing Events")
    i_m = sheets.index("SAP ↔ EDF Matched Events")
    i_r = sheets.index("Reconciliation")
    assert i_ft < i_bb < i_m < i_r, f"order wrong: ft={i_ft}, bb={i_bb}, m={i_m}, r={i_r}"


# ---------------------------------------------------------------------------
# Cluster-unmatched wiring: unmatched SAP events tagged as internal mechanism
# ---------------------------------------------------------------------------


def test_export_tags_unmatched_event_as_cluster_internal_mechanism(
    tmp_path: object,
) -> None:
    """End-to-end: a SAP event whose posting date falls inside a
    back-billing cluster window but whose amount disagrees with the
    cluster invoice gets tagged 'internal mechanism' in the Matched EDF
    Invoice # column on Sheet 1."""
    # EDF invoice: Date 01 Aug 2023, Period From 01/01/2019 → >365 days gap
    # → detect_back_billing flags it. Period window 01/01/2019..09/08/2023
    # contains the SAP fixture's posting date 01-01-2020. Period Charge 100.0.
    edf_base = {
        "Source": "Local PDF Folder",
        "Sender": "edf.co.uk",
        "Date": "01 Aug 2023",
        "Period From": "01/01/2019",
        "Period To": "09/08/2023",
        "Invoice #": "T-TAG",
        "Amount (£)": 1234.56,
        "Period Charge (£)": 100.0,
        "Unit Rate (p/kWh)": "",
        "% Change": "",
        "Entry Type": "New Bill",
        "Reading": "Actual",
        "Units (kWh)": "",
        "Standing Chg (p/day)": "",
        "Tariff": "Standard",
        "Attachment Name": "test.pdf",
        "Details": "",
        "Logic Used": "PDF new-format",
        "Anomaly Flag": "",
        "Cancel/Rebill Admitted": False,
    }
    edf_second = dict(edf_base)
    edf_second["Invoice #"] = "T-TAG2"
    edf_second["Period From"] = "10/08/2023"
    edf_second["Period To"] = "10/09/2023"
    edf_second["Date"] = "10/09/2023"
    edf = [edf_base, edf_second]
    # SAP cluster: posting date 01-01-2020 (inside the EDF period window),
    # net amount 0.0 (net-zero cluster, disagrees with EDF Period Charge 100.0
    # → no amount-band match → cluster-unmatched).
    sap_rows = parse_sap_financial_transactions(
        _sap_csv_with_cluster(
            clear_doc="CL-TAG",
            clear_date="15-03-2021",
            rows=[
                ("D1", "999.00", "Dr- Consum Billing Receivable", ""),
                ("D2", "-999.00", "Cr- Credit for Consum Billing", ""),
                ("D3", "0.00", "Dr- Consum Billing Receivable", ""),
                ("D4", "0.00", "Dr- Consum Billing Receivable", ""),
            ],
        ),
        source_file="test.pdf",
    )
    out = str(tmp_path / "wb_tag.xlsx")  # type: ignore[operator]
    export_to_excel(
        data=edf,
        output_path=out,
        error_log=[],
        config={
            "use_dedup": False,
            "acc_num": "0123456789",
            "scan_sap_dumps": True,
            "generate_reconciliation_sheet": False,
        },
        sap_rows={"financial": sap_rows, "contract": [], "meter": []},
    )
    wb = load_workbook(out)
    ws1 = wb["SAP Back-billing Events"]
    # Find the summary row for CL-TAG (column 1 = Clearing Doc #).
    tagged_value = None
    for r in range(4, ws1.max_row + 1):
        if str(ws1.cell(row=r, column=1).value or "") == "CL-TAG":
            tagged_value = ws1.cell(row=r, column=11).value
            break
    wb.close()
    assert tagged_value is not None, "CL-TAG summary row not found"
    assert "internal mechanism" in str(tagged_value), (
        f"expected 'internal mechanism' tag, got {tagged_value!r}"
    )


# ---------------------------------------------------------------------------
# Task 8 — cross-referenced 'Backbilling According to SAP' position sheet
# ---------------------------------------------------------------------------


def test_write_sap_back_billing_position_sheet() -> None:
    from edf_bill_fetcher.io.writers.sap import write_sap_back_billing_position_sheet

    wb = Workbook()
    result = {
        "events": [
            {
                "Clearing Doc #": "CLR-100",
                "Clearing Date": "2023-08-01",
                "Clearing Reason": "Reversal",
                "# Rows": 2,
                "Net Amount (£)": 0.0,
                "Has Credit for Consum Billing": "Yes",
                "Period(s)": "—",
                "Matched EDF Invoice #": "T-001",
            }
        ],
        "reconciliation": [
            {
                "SAP Event": "CLR-100",
                "EDF Invoice #": "T-001",
                "EDF Unlawful Charge (£)": 0.0,
                "SAP Net (£)": 0.0,
                "Verdict": "Reconciled",
            }
        ],
        "summary": {"sap_events": 1, "sap_net_total": 0.0, "reconciled": 1},
    }
    ws = write_sap_back_billing_position_sheet(wb, result)
    assert ws.title == "Backbilling According to SAP"
    # The brief's col-a-only scan cannot see the "Reconciled" Verdict (col 5),
    # so scan the full used grid for both sentinel values.
    grid = [
        ws.cell(row=r, column=c).value
        for r in range(1, ws.max_row + 1)
        for c in range(1, ws.max_column + 1)
    ]
    assert any("CLR-100" in str(v) for v in grid)
    assert any("Reconciled" in str(v) for v in grid)
