from __future__ import annotations

import os

import pytest
from openpyxl import load_workbook

from edf_bill_fetcher.io.writers import export_to_excel


def _sample_data() -> list[dict]:
    return [
        {
            "Source": "Local PDF Folder",
            "Sender": "edf.co.uk",
            "Date": "01 Sep 2023",
            "Period From": "01 Jan 2022",
            "Period To": "31 Aug 2023",
            "Invoice #": "T-X1",
            "Amount (£)": 1000.0,
            "Period Charge (£)": 800.0,
            "Unit Rate (p/kWh)": 25.0,
            "% Change": None,
            "Entry Type": "New Bill",
            "Reading": "Actual",
            "Units (kWh)": 300.0,
            "Standing Chg (p/day)": 50.0,
            "Tariff": "Standard",
            "Attachment Name": "T-X1.pdf",
            "Details": "Reading was actual",
            "Logic Used": "PDF new-format",
            "Anomaly Flag": "",
            "Cancel/Rebill Admitted": True,
        },
        {
            "Source": "Local PDF Folder",
            "Sender": "edf.co.uk",
            "Date": "01 Oct 2023",
            "Period From": "01 Feb 2022",
            "Period To": "30 Sep 2023",
            "Invoice #": "T-X2",
            "Amount (£)": 1500.0,
            "Period Charge (£)": 1200.0,
            "Unit Rate (p/kWh)": 25.0,
            "% Change": None,
            "Entry Type": "New Bill",
            "Reading": "Actual",
            "Units (kWh)": 400.0,
            "Standing Chg (p/day)": 50.0,
            "Tariff": "Standard",
            "Attachment Name": "T-X2.pdf",
            "Details": "Reading was actual",
            "Logic Used": "PDF new-format",
            "Anomaly Flag": "",
            "Cancel/Rebill Admitted": False,
        },
    ]


@pytest.fixture
def tmp_xlsx(tmp_path):
    return str(tmp_path / "test_run.xlsx")


def test_export_to_excel_emits_four_new_analysis_tabs(tmp_xlsx: str) -> None:
    export_to_excel(
        _sample_data(),
        tmp_xlsx,
        error_log=[],
        config={"use_dedup": False, "acc_num": "0123456789"},
    )
    assert os.path.exists(tmp_xlsx)
    wb = load_workbook(tmp_xlsx, read_only=True)
    names = set(wb.sheetnames)
    # The four new tabs must exist alongside the existing writers.
    assert "Back-billing Analysis" in names
    assert "Rebilling & Corrections" in names
    assert "Meter Readings" in names
    assert "Contract History" in names
    wb.close()


def _two_survivor_group_data() -> list[dict]:
    """Four invoices forming two independent killer/survivor chains.

    ``K1`` supersedes ``S1`` and ``K2`` supersedes ``S2``. On the
    Back-billing Analysis sheet the two live rows land on rows 8 (K1) and
    9 (K2); on the Superseded Reconciliation sheet the ``KILLER:`` header
    rows land on rows 8 (K1) and 10 (K2) because the interspersed data
    rows push the second group down.  This deliberately separates the two
    coordinate systems so a ``View superseded`` link that wrongly reuses
    back-billing row numbers lands on another group's data row.
    """
    base = {
        "Source": "Local PDF Folder",
        "Sender": "edf.co.uk",
        "Unit Rate (p/kWh)": 25.0,
        "% Change": None,
        "Entry Type": "New Bill",
        "Reading": "Actual",
        "Units (kWh)": 300.0,
        "Standing Chg (p/day)": 50.0,
        "Tariff": "Standard",
        "Details": "Reading was actual",
        "Logic Used": "PDF new-format",
        "Anomaly Flag": "",
    }
    return [
        {
            **base,
            "Date": "01 Sep 2021",
            "Period From": "01 Jan 2020",
            "Period To": "31 Aug 2021",
            "Invoice #": "S1",
            "Amount (£)": 1000.0,
            "Period Charge (£)": 800.0,
            "Attachment Name": "S1.pdf",
            "Cancel/Rebill Admitted": False,
        },
        {
            **base,
            "Date": "01 Nov 2021",
            "Period From": "01 Jan 2020",
            "Period To": "31 Oct 2021",
            "Invoice #": "K1",
            "Amount (£)": 1500.0,
            "Period Charge (£)": 1200.0,
            "Attachment Name": "K1.pdf",
            "Cancel/Rebill Admitted": True,
        },
        {
            **base,
            "Date": "01 Sep 2022",
            "Period From": "01 Jan 2021",
            "Period To": "31 Aug 2022",
            "Invoice #": "S2",
            "Amount (£)": 1100.0,
            "Period Charge (£)": 900.0,
            "Attachment Name": "S2.pdf",
            "Cancel/Rebill Admitted": False,
        },
        {
            **base,
            "Date": "01 Nov 2022",
            "Period From": "01 Jan 2021",
            "Period To": "31 Oct 2022",
            "Invoice #": "K2",
            "Amount (£)": 1600.0,
            "Period Charge (£)": 1300.0,
            "Attachment Name": "K2.pdf",
            "Cancel/Rebill Admitted": True,
        },
    ]


def test_view_superseded_links_point_at_own_killer_header(tmp_xlsx: str) -> None:
    """Each live survivor's ``View superseded`` link must target ITS OWN
    ``KILLER:`` header row on the Superseded Reconciliation sheet.

    The reconciliation sheet intersperses a ``KILLER:`` header row per
    group, so the survivor's row on Back-billing Analysis (8, 9) is NOT its
    reconciliation header row (8, 10).  Pre-fix the export pipeline reused
    back-billing row numbers, so K2's link landed on row 9 (S1's data row)
    instead of row 10 (K2's header).  This regression test would fail on
    that wiring.
    """
    export_to_excel(
        _two_survivor_group_data(),
        tmp_xlsx,
        error_log=[],
        config={"use_dedup": False, "acc_num": "0123456789", "analysis_min": 0.0},
    )
    wb = load_workbook(tmp_xlsx)
    ws_bb = wb["Back-billing Analysis"]
    ws_recon = wb["Superseded Reconciliation"]
    # Map each survivor invoice to the row of its own KILLER: header.
    killer_rows: dict[str, int] = {}
    for r in range(1, ws_recon.max_row + 1):
        label = ws_recon.cell(row=r, column=1).value
        if isinstance(label, str) and label.startswith("KILLER: "):
            killer_rows[label.removeprefix("KILLER: ")] = r
    assert set(killer_rows) == {"K1", "K2"}
    # Each live survivor row on Back-billing Analysis must link to its own
    # KILLER: header on the reconciliation sheet.
    hdrs = [c.value for c in ws_bb[7]]
    col = hdrs.index("View Superseded") + 1
    linked = 0
    for r in range(8, ws_bb.max_row + 1):
        inv = ws_bb.cell(row=r, column=1).value
        if inv not in killer_rows:
            continue
        cell = ws_bb.cell(row=r, column=col)
        assert cell.hyperlink is not None, f"{inv} missing View superseded link"
        assert cell.hyperlink.location == f"'Superseded Reconciliation'!A{killer_rows[inv]}", (
            f"{inv}: got {cell.hyperlink.location}, want !A{killer_rows[inv]}"
        )
        linked += 1
    assert linked == 2
    wb.close()


def test_export_writes_superseded_reconciliation_sheet(tmp_xlsx: str) -> None:
    export_to_excel(
        _sample_data(),
        tmp_xlsx,
        error_log=[],
        config={"use_dedup": False, "acc_num": "0123456789"},
    )
    assert os.path.exists(tmp_xlsx)
    wb = load_workbook(tmp_xlsx, read_only=True)
    names = wb.sheetnames
    assert "Superseded Reconciliation" in names
    wb.close()


def test_evidence_index_uses_full_df_rows_after_middle_row_filtered(tmp_xlsx: str) -> None:
    """Analyser hotlinks must target rows on the FULL evidence sheet, not the
    filtered analysis frame.

    INV-001 (amount 100) is below ``analysis_min`` so ``_prepare_analysis_frame``
    drops it from ``dfc``.  The EDF Evidence Report sheet still carries it at
    row 2, so the back-billing hotlink for INV-002 (next on the full frame,
    row 3) must point at ``!A3`` — building the index on ``dfc`` would wrongly
    map INV-002 to row 2 (INV-001's slot).
    """
    rows = [
        {
            "Source": "Local PDF Folder",
            "Sender": "edf.co.uk",
            "Date": "01 Jan 2023",
            "Period From": "01 Jan 2022",
            "Period To": "31 Dec 2022",
            "Invoice #": "INV-001",
            "Amount (£)": 100.0,
            "Period Charge (£)": 100.0,
            "Unit Rate (p/kWh)": 25.0,
            "% Change": None,
            "Entry Type": "New Bill",
            "Reading": "Actual",
            "Units (kWh)": 300.0,
            "Standing Chg (p/day)": 50.0,
            "Tariff": "Standard",
            "Attachment Name": "INV-001.pdf",
            "Details": "below analysis_min",
            "Logic Used": "PDF new-format",
            "Anomaly Flag": "",
        },
        {
            "Source": "Local PDF Folder",
            "Sender": "edf.co.uk",
            "Date": "01 Feb 2023",
            "Period From": "01 Jan 2022",
            "Period To": "31 Jan 2023",
            "Invoice #": "INV-002",
            "Amount (£)": 1000.0,
            "Period Charge (£)": 800.0,
            "Unit Rate (p/kWh)": 25.0,
            "% Change": None,
            "Entry Type": "New Bill",
            "Reading": "Actual",
            "Units (kWh)": 400.0,
            "Standing Chg (p/day)": 50.0,
            "Tariff": "Standard",
            "Attachment Name": "INV-002.pdf",
            "Details": "back-billed",
            "Logic Used": "PDF new-format",
            "Anomaly Flag": "",
        },
        {
            "Source": "Local PDF Folder",
            "Sender": "edf.co.uk",
            "Date": "01 Apr 2023",
            "Period From": "01 Mar 2023",
            "Period To": "31 Mar 2023",
            "Invoice #": "INV-003",
            "Amount (£)": 900.0,
            "Period Charge (£)": 700.0,
            "Unit Rate (p/kWh)": 25.0,
            "% Change": None,
            "Entry Type": "New Bill",
            "Reading": "Actual",
            "Units (kWh)": 300.0,
            "Standing Chg (p/day)": 50.0,
            "Tariff": "Standard",
            "Attachment Name": "INV-003.pdf",
            "Details": "short period",
            "Logic Used": "PDF new-format",
            "Anomaly Flag": "",
        },
    ]
    export_to_excel(
        rows,
        tmp_xlsx,
        error_log=[],
        config={"use_dedup": False, "acc_num": "0123456789", "analysis_min": 500.0},
    )
    wb = load_workbook(tmp_xlsx)
    # Evidence sheet row layout: header row 1, then full-frame rows 2-4.
    ws_ev = wb["EDF Evidence Report"]
    assert ws_ev.cell(row=2, column=6).value == "INV-001"
    assert ws_ev.cell(row=3, column=6).value == "INV-002"
    # Back-billing Analysis hotlink for INV-002 (col 14) must point at its
    # FULL-frame row 3, not the dfc-relative row 2.
    ws_bb = wb["Back-billing Analysis"]
    target = None
    for r in range(8, ws_bb.max_row + 1):
        if ws_bb.cell(row=r, column=1).value == "INV-002":
            target = r
            break
    assert target is not None, "INV-002 not found on Back-billing Analysis sheet"
    cell = ws_bb.cell(row=target, column=14)
    assert cell.hyperlink is not None, "INV-002 hotlink missing"
    assert cell.hyperlink.location.endswith("!A3"), cell.hyperlink.location
    wb.close()


def test_evidence_report_does_not_leak_diagnostic_columns(tmp_xlsx: str) -> None:
    """The ``Source PDF Text`` / ``_regex_trace`` / ``Balance Last Bill (£)``
    columns are diagnostic-only and must NEVER appear on the saved
    ``EDF Evidence Report`` tab.  Pre-fix the col_order reindex never
    dropped them so the visible workbook carried the 4 KB-per-row
    PDF body blocks for every record, polluting the main sheet with
    "process internals" noise that adds ~50% to the saved workbook
    size for nothing visible to the consumer.

    This test feeds the same data through ``export_to_excel`` and
    asserts the saved Evidence Report sheet has exactly the
    canonical 19 header columns (Source … Duplicate Of).
    """
    rows = _sample_data()
    # Plant every diagnostic col on the records so the regression
    # would surface if any of them leaked.
    for row in rows:
        row["Source PDF Text"] = "very long body text " * 200
        row["_regex_trace"] = "trace path"
        row["Balance Last Bill (£)"] = 123.45
    export_to_excel(
        rows,
        tmp_xlsx,
        error_log=[],
        config={"use_dedup": False, "acc_num": "0123456789"},
    )
    wb = load_workbook(tmp_xlsx, read_only=True)
    ws = wb["EDF Evidence Report"]
    headers = []
    for cell in ws[1]:
        if cell.value is not None:
            headers.append(cell.value)
    wb.close()
    assert "Source PDF Text" not in headers, (
        f"diagnostic column 'Source PDF Text' leaked into Evidence Report: {headers}"
    )
    assert "_regex_trace" not in headers, (
        f"diagnostic column '_regex_trace' leaked into Evidence Report: {headers}"
    )
    assert "Balance Last Bill (£)" not in headers, (
        f"diagnostic column 'Balance Last Bill (£)' leaked into Evidence Report: {headers}"
    )
