"""Phase 2 — payment-analysis chart layout regression.

The ``write_payment_analysis_sheet`` function in ``edf_collector``
historically anchored the in-sheet chart at column H, row
``r + 2`` — the default chart width of 28 chart-units ~ 28 Excel
columns rendered the chart far to the right of the visible data
table (which only spans columns A-E), so a user /
ombudsman reviewer opening the file saw the chart title render
*off-screen* (half-hidden).  This regression test pins the new
contract: chart sits at column B, below the data table, on a
single helper mini-table that makes the chart's data references
explicit.

The test reads back the saved .xlsx with ``openpyxl`` and asserts
the new chart anchor + helper-table layout.  No screenshotting
required — the chart-meta data is what Excel/LibreOffice uses to
position the chart when the file is opened.
"""

from __future__ import annotations

from collections.abc import Iterator
from pathlib import Path
from typing import Any

import pandas as pd
import pytest

from edf_bill_fetcher.writers import write_payment_analysis_sheet


@pytest.fixture
def tmp_dir() -> Iterator[Path]:
    """A writable temp directory (avoids the Windows ``tmp_path``
    ACL ``PermissionError`` that's pre-existing on this host).
    """
    import tempfile

    d = Path(tempfile.mkdtemp(prefix="edf_payment_chart_"))
    yield d
    try:
        for f in d.iterdir():
            f.unlink(missing_ok=True)
        d.rmdir()
    except OSError:
        pass


def _synthetic_payment_records(monthly_dates: list[str]) -> list[dict]:
    """Build a list of Payment / Credit records ordered by date.

    The dates are passed as dd/mm/yyyy strings to match the
    primary-key format ``write_payment_analysis_sheet`` expects
    (it calls ``parse_to_sort_date`` which accepts dd/mm/yyyy).
    """
    return [
        {
            "Date": monthly_dates[i],
            "Entry Type": "Payment",
            "Amount (£)": 50.0 + i * 5.0,
            "Balance After (£)": 100.0 - i * 10.0,
            "Details": f"auto-debit {i}",
        }
        for i in range(len(monthly_dates))
    ]


def _read_sheet(xlsx_path: Path) -> Any:
    """Return the Payment Analysis worksheet via openpyxl.

    Returns ``Any`` because openpyxl's ``Worksheet`` is not in
    the project's ``mypy ..ignore_missing_imports`` allow-list
    and the test module is intentionally tolerant of openpyxl's
    internal vs. public attribute drift (e.g. ``._charts`` is a
    private attribute — openpyxl-Worksheet type signatures don't
    promise it).
    """
    from openpyxl import load_workbook

    wb = load_workbook(xlsx_path, read_only=False, data_only=True)
    try:
        ws = wb["Payment Analysis"]
        return ws
    finally:
        wb.close()


class TestPaymentAnalysisChartLayout:
    """New contract — chart anchored below the table, in-viewport."""

    def _build(self, tmp_dir: Path, count: int = 12) -> Path:
        """Build a synthetic payment workbook with ``count`` rows."""
        from openpyxl import Workbook

        xlsx_path = tmp_dir / "payment.xlsx"
        dates = [f"{i + 1:02d}/02/2024" for i in range(count)]
        records = _synthetic_payment_records(dates)
        # Add a debit record so the engine doesn't try to balance
        # payments-only — the chart code must still fire.
        records.append(
            {
                "Date": "01/01/2024",
                "Entry Type": "New Bill",
                "Amount (£)": 250.0,
                "Balance After (£)": 250.0,
                "Details": "Q1 bill",
            }
        )
        wb = Workbook()
        try:
            ws = wb.active
            write_payment_analysis_sheet(ws, pd.DataFrame(records))
            wb.save(xlsx_path)
        finally:
            wb.close()
        return xlsx_path

    def test_chart_present_and_anchored_in_column_b(self, tmp_dir: Path) -> None:
        """The chart's anchor cell must NOT be column H (the
        off-screen position the old layout used).  It must be
        in column B (or earlier) so a default Excel viewport
        sees the chart without panning.
        """
        xlsx_path = self._build(tmp_dir)
        ws = _read_sheet(xlsx_path)
        assert len(ws._charts) == 1, (
            "Payment Analysis sheet expected exactly one chart; "
            "got either zero (chart code didn't fire) or more "
            "(payment data interpreted as multiple series)."
        )
        chart_anchor = ws._charts[0].anchor  # type: ignore[attr-defined]
        # ``anchor.``_from is a ``OneCellAnchor``-style object with
        # ``col``/``row`` zero-indexed.  Translate to Excel's
        # 1-indexed A=1 .. look-up table:
        #   col 0 → A, col 1 → B, col 7 → H.
        col_idx = chart_anchor._from.col  # type: ignore[attr-defined]
        # Column B is index 1.  Anything ≤ 1 keeps the chart on
        # the left side of the sheet; the previous layout used 7
        # (column H), which is what we want to *prevent* regressing.
        assert col_idx <= 1, (
            f"Chart anchored at column index {col_idx} (Excel "
            f"column {chr(ord('A') + col_idx)}); the new contract "
            f"demands ≤ column B so the chart title sits inside "
            f"the default viewport.  Likely a regression of the "
            f"off-screen H-anchor removed in Phase 2."
        )

    def test_chart_data_blocks_are_inside_visible_table(self, tmp_dir: Path) -> None:
        """The chart-data helper cells must live in columns A-B
        only (not column F as the legacy layout did), so the
        helper mini-table is drawn next to the chart label
        rather than scattered behind it.
        """
        xlsx_path = self._build(tmp_dir)
        ws = _read_sheet(xlsx_path)
        chart = ws._charts[0]  # type: ignore[attr-defined]
        # ``chart.series`` carries the openpyxl ``Series``
        # objects.  Pull the data-reference for the first
        # series, which is the payment-amount column.
        ser = chart.series[0]
        numref = ser.val.numRef.f if ser.val and ser.val.numRef else None
        assert numref is not None, (
            "BarChart series[0] carries no numeric reference; "
            "the chart isn't pointing at the helper mini-table."
        )
        # openpyxl renders ``min``-columns as letters — extract
        # the leftmost column from the range string.
        # e.g. ``'Sheet'!$B$5:$B$16`` → "B".
        import re

        leaf = numref.split("!")[-1]
        m = re.match(r"\$?([A-Z]+)\$?\d+", leaf)
        assert m, f"Cannot parse numeric reference {numref!r}"
        leftmost_col_letter = m.group(1)
        assert leftmost_col_letter in ("A", "B"), (
            f"Chart's leftmost data column is {leftmost_col_letter}; "
            f"the new layout contract keeps the helper mini-table "
            f"in column A (dates) / B (amounts) so it sits next to "
            f"the chart label.  Found anything later — likely a "
            f"regression of the on-disc 'chart-data in column F' "
            f"idiom that landed helper cells behind the chart."
        )

    def test_chart_width_fits_viewport(self, tmp_dir: Path) -> None:
        """The chart's openpyxl ``width`` attribute must be
        capped at 16 (the new contract) rather than the legacy
        28 — past 16 the chart visibly extends past the
        data-table's column-F boundary for a default
        Excel-on-Windows display.
        """
        xlsx_path = self._build(tmp_dir)
        ws = _read_sheet(xlsx_path)
        chart = ws._charts[0]  # type: ignore[attr-defined]
        assert chart.width <= 16, (
            f"Chart width is {chart.width} (openpyxl chart units); "
            f"the new contract caps it at 16.  Larger widths push "
            f"the chart off the visible portion of the workbook "
            f"for default Excel viewports — the exact complaint "
            f"the Phase-2 fix was resolving."
        )
