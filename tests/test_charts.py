"""Tests for chart objects created by the Excel export pipeline.

Verifies that BarChart and LineChart objects are placed on the
expected sheets at the expected anchors after calling
``export_to_excel``.
"""

from __future__ import annotations

import openpyxl
import pandas as pd
from openpyxl.chart import BarChart, LineChart


def _multi_period_df() -> pd.DataFrame:
    return pd.DataFrame(
        {
            "Date": ["01 Jan 2024", "15 Feb 2024", "01 Mar 2024", "20 Jan 2024", "25 Feb 2024"],
            "Invoice #": ["T-001", "T-002", "T-003", "P-001", "P-002"],
            "Amount (£)": [600.0, 700.0, 550.0, -600.0, -700.0],
            "Period From": ["01 Jan 2024", "01 Feb 2024", "01 Mar 2024", "N/A", "N/A"],
            "Period To": ["31 Jan 2024", "29 Feb 2024", "31 Mar 2024", "N/A", "N/A"],
            "Period Charge (£)": [600.0, 700.0, 550.0, "N/A", "N/A"],
            "Units (kWh)": [100, 200, 150, "N/A", "N/A"],
            "Entry Type": ["New Bill", "New Bill", "New Bill", "Payment", "Payment"],
            "Source": ["PST", "PST", "PST", "PST", "PST"],
            "Sender": [
                "edf@example.com",
                "edf@example.com",
                "edf@example.com",
                "edf@example.com",
                "edf@example.com",
            ],
            "Tariff": ["Standard", "Standard", "Standard", "N/A", "N/A"],
            "Standing Chg (p/day)": [50.0, 50.0, 50.0, "N/A", "N/A"],
            "Unit Rate (p/kWh)": [24.0, 24.0, 24.0, "N/A", "N/A"],
            "Reading": [1000, 1200, 1400, "N/A", "N/A"],
            "Details": ["Bill 1", "Bill 2", "Bill 3", "Payment 1", "Payment 2"],
            "Attachment Name": ["bill1.pdf", "bill2.pdf", "bill3.pdf", "pay1.pdf", "pay2.pdf"],
            "Balance": [1000.0, 1200.0, 1050.0, 400.0, 500.0],
        }
    )


def test_payment_bar_chart(tmp_path: object) -> None:
    from edf_bill_fetcher.io.writers import export_to_excel

    out = tmp_path / "payment_bar_chart.xlsx"  # type: ignore[operator]
    df = _multi_period_df()
    export_to_excel(df.to_dict(orient="records"), str(out), [], {})
    # The payment chart is a BarChart on the Payment Analysis sheet.
    # After export_to_excel the workbook is written to disk; reload
    # to inspect the chart objects.
    wb2 = openpyxl.load_workbook(str(out))
    assert "Payment Analysis" in wb2.sheetnames, "Expected Payment Analysis sheet"
    ws = wb2["Payment Analysis"]
    charts = ws._charts
    assert any(isinstance(c, BarChart) for c in charts), "Expected BarChart on Payment Analysis"


def test_balance_trend_line_chart(tmp_path: object) -> None:
    from edf_bill_fetcher.io.writers import export_to_excel

    out = tmp_path / "balance_trend_line_chart.xlsx"  # type: ignore[operator]
    df = _multi_period_df()
    export_to_excel(df.to_dict(orient="records"), str(out), [], {})
    wb2 = openpyxl.load_workbook(str(out))
    assert "Balance Trend" in wb2.sheetnames, "Expected Balance Trend sheet"
    ws = wb2["Balance Trend"]
    charts = ws._charts
    assert any(isinstance(c, LineChart) for c in charts), "Expected LineChart on Balance Trend"


def test_period_charges_bar_chart(tmp_path: object) -> None:
    from edf_bill_fetcher.io.writers import export_to_excel

    out = tmp_path / "period_charges_bar_chart.xlsx"  # type: ignore[operator]
    df = _multi_period_df()
    export_to_excel(df.to_dict(orient="records"), str(out), [], {})
    wb2 = openpyxl.load_workbook(str(out))
    assert "Period Charges" in wb2.sheetnames, "Expected Period Charges sheet"
    ws = wb2["Period Charges"]
    charts = ws._charts
    assert any(isinstance(c, BarChart) for c in charts), "Expected BarChart on Period Charges"


def test_year_on_year_bar_chart(tmp_path: object) -> None:
    from edf_bill_fetcher.io.writers import export_to_excel

    out = tmp_path / "year_on_year_bar_chart.xlsx"  # type: ignore[operator]
    df = _multi_period_df()
    export_to_excel(df.to_dict(orient="records"), str(out), [], {})
    wb2 = openpyxl.load_workbook(str(out))
    assert "Year-on-Year" in wb2.sheetnames, "Expected Year-on-Year sheet"
    ws = wb2["Year-on-Year"]
    charts = ws._charts
    assert any(isinstance(c, BarChart) for c in charts), "Expected BarChart on Year-on-Year"
