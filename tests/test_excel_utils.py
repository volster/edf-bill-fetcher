"""Tests for edf_bill_fetcher.helpers.excel_utils cell primitives.

Covers the relative-path PDF hyperlink helper (``pdf_hyperlink_cell``)
added in the superseded-reconciliation cleanup.
"""

from edf_bill_fetcher.helpers.excel_utils import pdf_hyperlink_cell


def test_pdf_hyperlink_cell_emits_relative_target() -> None:
    import openpyxl

    ws = openpyxl.Workbook().active
    pdf_hyperlink_cell(ws, 1, 1, "evidence_files/T78701920034.pdf")
    c = ws.cell(row=1, column=1)
    assert c.hyperlink is not None
    assert c.hyperlink.target == "evidence_files/T78701920034.pdf"
    assert c.value == "T78701920034.pdf"


def test_pdf_hyperlink_cell_guards_formula_lead() -> None:
    import openpyxl

    ws = openpyxl.Workbook().active
    pdf_hyperlink_cell(ws, 1, 1, "evidence_files/=evil.pdf")
    assert ws.cell(row=1, column=1).data_type == "s"
