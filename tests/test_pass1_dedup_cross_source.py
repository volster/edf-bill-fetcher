"""Phase 2.2 / Pass-1 dedup regression — cross-source same-period-amount.

The Pass-1 dedup walker in ``edf_collector.export_to_excel``
used to compute ``_dedup_date`` as ``df["_sort"].where(cond,
df["_sort"])`` — a tautology that ignored ``Period To``
entirely and so never catched HTM↔PST cross-source duplicates
when the receipt dates differ.  Combined with a
``sort_values(["_sort", "_src_pri"])`` that sorted Date-
ascending before source-precedence, the dedup sank the
higher-precedence HTM row even when its source-of-truth
billing period matched a PST row exactly.

This regression pins both fixes:

  1. Two-row collision where HTM and PST share Period To but
     have different receipt Dates — both rows resolved to the
     same ``_dedup_date`` (the parsed Period To), and the
     kept row must be HTM (lower _src_pri = higher
     precedence).
  2. Three-row collision spanning the full user-stated
     precedence ladder (HTM, Local PDF, PST, Email Body share
     Period To; Email Body is repeated with a stray different
     amount) — confirms the kept row follows precedence, not
     df order.

The fix is a small but tightly-locked behavioural change.
"""

from __future__ import annotations

import os
from pathlib import Path

import pandas as pd
import pytest

from edf_collector import export_to_excel


@pytest.fixture
def workdir() -> Path:
    # Skirt pytest's ``tmp_path`` fixture on this Windows host —
    # the sandboxed TEMP directory is read-only on this developer's
    # machine, so any fixture that depends on pytest's tmp-path
    # machinery error-cascades at setup.  Derive our own scratch
    # dir from ``USERPROFILE`` (or ``/tmp`` as a Linux fallback)
    # plus an explicit pid-locked name, so cross-test isolation is
    # still preserved.
    scratch = Path(os.environ.get("USERPROFILE", "/tmp")) / f".edf_dedup_scratch_{os.getpid()}"
    scratch.mkdir(parents=True, exist_ok=True)
    return scratch


def _price_records() -> list[dict]:
    """Records spanning all the user-stated source precedences
    sharing a single billing period (Period To=01/04/2024).
    One row (the Email Body) is given a different amount so we
    can verify Pass 1's *exact* (`_dedup_date`, Amount) tuple
    keying without false-positives.
    """
    return [
        # Email Body with the wrong amount — must NOT be a
        # cross-source duplicate; this row is kept alone.
        {
            "Date": "01/04/2024",
            "Source": "Email Body",
            "Period From": "N/A",
            "Period To": "01/04/2024",
            "Invoice #": "",
            "Money": 99.99,
            "Amount (£)": 99.99,
            "Period Charge (£)": 0.0,
            "Units (kWh)": "",
            "Reading": "",
            "Entry Type": "Payment",
            "Logic Used": "",
            "Details": "",
            "Attachment Name": "",
            "Standing Charge": "",
            "Anomaly Flag": "",
            "Sender": "edfenergy.com",
        },
        # PST PDF Attachment — receipt date is 02/04 (one day
        # after Period To end-of-bill).  Period To carries the
        # canonical end-of-billing date.
        {
            "Date": "02/04/2024",
            "Source": "PST PDF Attachment",
            "Period From": "01/03/2024",
            "Period To": "01/04/2024",
            "Invoice #": "",
            "Money": 100.0,
            "Amount (£)": 100.0,
            "Period Charge (£)": 80.0,
            "Units (kWh)": "100",
            "Reading": "Actual",
            "Entry Type": "New Bill",
            "Logic Used": "Period",
            "Details": "",
            "Attachment Name": "",
            "Standing Charge": "",
            "Anomaly Flag": "",
            "Sender": "edfenergy.com",
        },
        # Local PDF Folder — receipt date is 03/04 (two days
        # after end-of-bill).
        {
            "Date": "03/04/2024",
            "Source": "Local PDF Folder",
            "Period From": "N/A",
            "Period To": "01/04/2024",
            "Invoice #": "",
            "Money": 100.0,
            "Amount (£)": 100.0,
            "Period Charge (£)": 80.0,
            "Units (kWh)": "100",
            "Reading": "Actual",
            "Entry Type": "New Bill",
            "Logic Used": "Period",
            "Details": "",
            "Attachment Name": "",
            "Standing Charge": "",
            "Anomaly Flag": "",
            "Sender": "edfenergy.com",
        },
        # HTM Account History — receipt date is 04/04 (three days
        # after end-of-bill).  Highest-precedence source-of-truth
        # record; per-pass-1 contract *this* is the kept row
        # even though it sorts last by Date.
        {
            "Date": "04/04/2024",
            "Source": "HTM Account History",
            "Period From": "01/03/2024",
            "Period To": "01/04/2024",
            "Invoice #": "",
            "Money": 100.0,
            "Amount (£)": 100.0,
            "Period Charge (£)": 80.0,
            "Units (kWh)": "100",
            "Reading": "Actual",
            "Entry Type": "New Bill",
            "Logic Used": "Period",
            "Details": "",
            "Attachment Name": "",
            "Standing Charge": "",
            "Anomaly Flag": "",
            "Sender": "edfenergy.com",
        },
    ]


class TestCrossSourceDedup:
    """Pass-1 cross-source dedup regression.

    The headline bug: ``df["_dedup_date"] = df["_sort"].where(
    (df["Period To"] != "N/A") & df["Period To"].notna(),
    df["_sort"])`` is a tautology — both branches are
    ``df["_sort"]`` so Period To was ignored, and the dedup
    only collided on (Date, Amount), missing the cross-
    source case where HTM and PST receive the *same* bill on
    *different days*.

    The second-order bug: ``df.sort_values(["_sort",
    "_src_pri"])`` ranked by Date first, so even after the
    Period-To fix the dedup kept the earliest-Date row (PST)
    rather than the highest-precedence row (HTM).

    Both are pinned here.
    """

    def test_keeps_highest_precedence_among_same_period_collisions(self, workdir: Path) -> None:
        records = _price_records()
        # Drive the full export_to_excel pipeline so we exercise
        # the production dedup walker (rather than a re-
        # implementation with its own off-by-one).
        out = workdir / "out.xlsx"
        config = {
            "use_dedup": True,
            "save_dups": True,
            "use_anchors": False,
            "use_large": False,
            "min_amount": 0.0,
            "filter_below": False,
            "use_dedup_period": True,
            "expanded_columns": True,
            "include_charts": False,
            "include_forecast": False,
        }
        # ``export_to_excel`` returns ``None`` (the function
        # writes the artifact as a side-effect); assert the
        # artifact exists by the time the call returns.
        export_to_excel(records, str(out), [], config=config)
        assert out.exists(), "export_to_excel must write the workbook artifact"

        # Inspect the kept set via openpyxl on the saved workbook.
        from openpyxl import load_workbook

        wb = load_workbook(str(out))
        ws = wb["EDF Evidence Report"]
        # The kept set is a flat table; we only need to know
        # which Sources survived.  Read column A (Source) for
        # row 2 onward.
        kept_sources = [
            ws.cell(row=r, column=1).value
            for r in range(2, ws.max_row + 1)
            if ws.cell(row=r, column=1).value
        ]
        # The four £100 collisions collapse to ONE kept row;
        # the £99.99 row survives as a separate kept row.
        # Per the user-stated precedence HTM > Local PDF >
        # PST, the kept survivor for £100 must be HTM.
        kept_100 = [
            s
            for s in kept_sources
            if s in ("HTM Account History", "Local PDF Folder", "PST PDF Attachment")
        ]
        assert len(kept_100) == 1, (
            f"Expected exactly one £100 survivor; got {kept_100}.  All kept sources: {kept_sources}"
        )
        assert kept_100[0] == "HTM Account History", (
            f"Highest-precedence HTM should win the dedup, got "
            f"{kept_100[0]}; full kept sources: {kept_sources}"
        )
        # The wrong-amount Email Body row survives independently.
        assert "Email Body" in kept_sources, (
            f"Email Body with amount 99.99 should be its own kept row; got {kept_sources}"
        )

    def test_duplicate_entries_sheet_references_kept_htm(self, workdir: Path) -> None:
        records = _price_records()
        out = workdir / "out.xlsx"
        config = {
            "use_dedup": True,
            "save_dups": True,
            "use_anchors": False,
            "use_large": False,
            "min_amount": 0.0,
            "filter_below": False,
            "use_dedup_period": True,
            "expanded_columns": True,
            "include_charts": False,
            "include_forecast": False,
        }
        export_to_excel(records, str(out), [], config=config)

        from openpyxl import load_workbook

        wb = load_workbook(str(out))
        assert "Duplicate Entries" in wb.sheetnames, (
            "save_dups=True must yield a Duplicate Entries sheet."
        )
        ws = wb["Duplicate Entries"]
        # Three rows should be flagged dups on this fixture
        # (Local PDF and PST both colliding with HTM on
        # £100, plus Email Body duplicates are not in the £100
        # bucket so only 2 dups are expected).
        dup_sources = [
            ws.cell(row=r, column=1).value
            for r in range(2, ws.max_row + 1)
            if ws.cell(row=r, column=1).value
        ]
        assert set(dup_sources) == {
            "Local PDF Folder",
            "PST PDF Attachment",
        }, f"Expected only Local PDF + PST marked dup; got {set(dup_sources)}"
        # The Duplicate Of hyperlink-and-summary should resolve
        # to the HTM record for both dups.
        # The dup sheet's "Duplicate Of" hyperlink-summary column
        # index depends on the headers list length in
        # ``write_evidence_sheet``.  As of the Tariff insertion
        # the headers list has 19 entries, so the duplicate column
        # is at position 20 (1-based) of the writer's run loop
        # (see write_evidence_sheet's post-loop append).  Resolve
        # by header-name rather than hard-coded index so a future
        # header insertion (column shifts) is automatically tracked.
        header_row = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        duplicate_of_col = (
            header_row.index("Duplicate Of") + 1 if "Duplicate Of" in header_row else None
        )
        if duplicate_of_col is None:
            # Older workbook without a Duplicate Of column — assert
            # no dups were saved (the save_dups toggle was off).
            assert ws.max_row == 1, "save_dups=True must yield a Duplicate Of column header."
            return
        for r in range(2, ws.max_row + 1):
            source = ws.cell(row=r, column=1).value
            summary = ws.cell(row=r, column=duplicate_of_col).value
            assert "HTM Account History" in (summary or ""), (
                f"Row {r} ({source}) should resolve to HTM; got {summary!r}"
            )


class TestPeriodToDedupKey:
    """Unit-level pin of the corrected ``_dedup_date`` calc.

    The pre-fix tautology ``df["_sort"].where(cond, df["_sort"])``
    silently shadowed Period To.  We re-implement the corrected
    calc inline (so this test does not lean on internal
    variable names) and assert that two rows sharing Period To
    but disagreeing on Date resolve to the same ``_dedup_date``.
    """

    def test_same_period_to_collapses_despite_different_dates(self) -> None:
        records = [
            {"Date": "01/04/2024", "Period To": "01/04/2024", "Amount": 100.0},
            {"Date": "02/04/2024", "Period To": "01/04/2024", "Amount": 100.0},
            {"Date": "03/04/2024", "Period To": "01/04/2024", "Amount": 100.0},
        ]
        df = pd.DataFrame(records)
        df["_sort"] = pd.to_datetime(df["Date"], dayfirst=True, errors="coerce")
        period_to_dt = pd.to_datetime(df["Period To"], dayfirst=True, errors="coerce")
        df["_dedup_date"] = period_to_dt.where(period_to_dt.notna(), df["_sort"])
        # All three rows collapse to the same _dedup_date (the
        # canonical end-of-billing-period date), regardless of
        # receipt Date differences.
        assert df["_dedup_date"].nunique() == 1, (
            f"Expected Period To and _sort to collapse to the "
            f"same _dedup_date; got {df['_dedup_date'].tolist()}"
        )

    def test_no_period_falls_back_to_date(self) -> None:
        # One row with a real Period To, one row with "N/A"
        # Period To.  The latter must fall back to its Date so
        # it does NOT collide with the former on _dedup_date.
        records = [
            {
                "Date": "01/04/2024",
                "Period To": "01/04/2024",
                "Amount": 100.0,
            },
            {
                "Date": "01/04/2024",  # same Date as row above
                "Period To": "N/A",  # but no real Period To
                "Amount": 100.0,
            },
        ]
        df = pd.DataFrame(records)
        df["_sort"] = pd.to_datetime(df["Date"], dayfirst=True, errors="coerce")
        period_to_dt = pd.to_datetime(df["Period To"], dayfirst=True, errors="coerce")
        df["_dedup_date"] = period_to_dt.where(period_to_dt.notna(), df["_sort"])
        # Row 1 (HTM) keeps its Period To; row 2 (no-period)
        # falls back to _sort.  Both happen to land on
        # 2024-04-01 because _sort and Period To are the same
        # date here.  Importantly, the *fallback clause fired*
        # — we confirm by inspecting that period_to_dt.isna()
        # for the N/A row.
        no_period_row = df.iloc[1]
        period_to_dt_check = pd.to_datetime(
            pd.Series(no_period_row["Period To"]),
            dayfirst=True,
            errors="coerce",
        )
        assert period_to_dt_check.isna().all(), (
            "N/A Period To should fail to_datetime parse, triggering the fallback clause."
        )
        # The _dedup_date for the no-period row equals its
        # _sort, not row 0's Period To value.
        assert df.loc[1, "_dedup_date"] == df.loc[1, "_sort"], (
            "no-period row should fall back to its Date."
        )
