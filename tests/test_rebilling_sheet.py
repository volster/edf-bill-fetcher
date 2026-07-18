from __future__ import annotations

import pandas as pd
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from edf_collector import detect_rebilling, write_rebilling_sheet


def _sample_df() -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "Invoice #": "T67",
                "Date": "01 Aug 2023",
                "Period From": "01 Apr 2023",
                "Period To": "31 Jul 2023",
                "Amount (£)": 4401.07,
                "Cancel/Rebill Admitted": True,
                "Attachment Name": "T67.pdf",
            },
            {
                "Invoice #": "T68",
                "Date": "01 Oct 2023",
                "Period From": "01 Jan 2022",
                "Period To": "30 Sep 2023",
                "Amount (£)": 1525.13,
                "Cancel/Rebill Admitted": True,
                "Attachment Name": "T68.pdf",
            },
        ]
    )


def _open_ws(title: str = "Rebilling Analysis") -> Worksheet:
    wb = Workbook()
    ws = wb.active
    ws.title = title
    return ws


def test_write_rebilling_sheet_renders_title_and_subheader() -> None:
    ws = _open_ws()
    rb = detect_rebilling(_sample_df())
    write_rebilling_sheet(ws, rb, account="ACC1")
    # Row 1: title banner.
    a1 = ws.cell(row=1, column=1).value
    assert isinstance(a1, str)
    assert "REBILLING" in a1.upper() or "CORRECTION" in a1.upper()
    assert "ACC1" in a1
    # Row 2 or 3: subheader explains what each row identifies (mention
    # words like "cancel" / "rebill" / "later invoice").
    sub_found = False
    for r in range(2, 5):
        v = ws.cell(row=r, column=1).value
        if isinstance(v, str) and (
            "cancel" in v.lower() or "rebill" in v.lower() or "later invoice" in v.lower()
        ):
            sub_found = True
            break
    assert sub_found


def test_write_rebilling_sheet_writes_table_headers() -> None:
    ws = _open_ws()
    rb = detect_rebilling(_sample_df())
    write_rebilling_sheet(ws, rb, account="ACC1")
    # Header row sits at row 7 (matching Back-billing sheet layout).
    headers = [ws.cell(row=7, column=c).value for c in range(1, 8)]
    expected = [
        "Killer Invoice",
        "Killed Invoice",
        "Killer Date",
        "Killed Date",
        "Period Overlap (days)",
        "Jump-back (days)",
        "Trigger Reason",
    ]
    assert headers == expected


def test_write_rebilling_sheet_emits_one_row_per_pair() -> None:
    ws = _open_ws()
    rb = detect_rebilling(_sample_df())
    write_rebilling_sheet(ws, rb, account="ACC1")
    # Sample has exactly one (T68, T67) pair.
    a8 = ws.cell(row=8, column=1).value
    assert a8 == "T68"
    b8 = ws.cell(row=8, column=2).value
    assert b8 == "T67"
    # No extra data rows.
    assert ws.cell(row=9, column=1).value in (None, "")


def test_write_rebilling_sheet_empty_df_renders_headers_only() -> None:
    ws = _open_ws()
    empty = pd.DataFrame(
        columns=[
            "Killer Invoice",
            "Killed Invoice",
            "Killer Date",
            "Killed Date",
            "Period Overlap (days)",
            "Jump-back (days)",
            "Trigger Reason",
            "Cancel/Rebill Admitted (Killer)",
        ]
    )
    write_rebilling_sheet(ws, empty, account="ACC1")
    # Title still rendered.
    a1 = ws.cell(row=1, column=1).value
    assert isinstance(a1, str)
    assert "REBILLING" in a1.upper()
    # Table headers still rendered.
    headers = [ws.cell(row=7, column=c).value for c in range(1, 8)]
    assert headers[0] == "Killer Invoice"
    # No data row.
    assert ws.cell(row=8, column=1).value in (None, "")


def test_write_rebilling_sheet_freeze_panes_at_data_start() -> None:
    ws = _open_ws()
    rb = detect_rebilling(_sample_df())
    write_rebilling_sheet(ws, rb, account="ACC1")
    # Freeze panes should sit just above the data region (row 8 in 1-indexed).
    assert ws.freeze_panes == "A8"


def test_write_rebilling_sheet_cascade_renders_all_pairs() -> None:
    ws = _open_ws()
    # Build a 4-invoice cascade (similar to spec test case 5).
    df = pd.DataFrame(
        [
            {
                "Invoice #": "T65",
                "Date": "01 Apr 2023",
                "Period From": "01 Feb 2023",
                "Period To": "31 Mar 2023",
                "Amount (£)": 100.0,
                "Cancel/Rebill Admitted": False,
            },
            {
                "Invoice #": "T66",
                "Date": "01 Jun 2023",
                "Period From": "01 Mar 2023",
                "Period To": "31 May 2023",
                "Amount (£)": 100.0,
                "Cancel/Rebill Admitted": False,
            },
            {
                "Invoice #": "T67",
                "Date": "01 Aug 2023",
                "Period From": "01 Apr 2023",
                "Period To": "31 Jul 2023",
                "Amount (£)": 100.0,
                "Cancel/Rebill Admitted": False,
            },
            {
                "Invoice #": "T68",
                "Date": "01 Oct 2023",
                "Period From": "01 Feb 2023",  # extends back; contains T67 fully
                "Period To": "30 Sep 2023",
                "Amount (£)": 100.0,
                "Cancel/Rebill Admitted": True,  # admit-phrase on the killer
            },
        ]
    )
    rb = detect_rebilling(df)
    write_rebilling_sheet(ws, rb, account="ACC1")
    # The table has at least 2 rows (cascade).
    n_rows = sum(1 for r in range(8, 50) if ws.cell(row=r, column=1).value not in (None, ""))
    assert n_rows >= 2
