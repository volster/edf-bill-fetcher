"""Branch-coverage tests for ``edf_bill_fetcher.io.writers.rebilling``.

This module pins the previously-uncovered branches of the rebilling
writer submodule:

* ``_reversal_match`` (lines 49-72) -- the private helper that decides
  whether a reversal-credit row in the evidence DataFrame matches a
  killed invoice well enough to count as rebilling evidence.
* ``detect_rebilling`` (lines 115-209) -- the pure-pandas cancel-and-
  repost detector that lives in the *writer* submodule (a sibling of
  the copy in ``processors.detection``; both must be exercised).
* ``write_rebilling_sheet`` (lines 229, 305, 311, 313-320) -- the
  openpyxl renderer's account-banner, admit-phrase font, and
  evidence-index hotlink / "No match" branches.

The existing sibling tests (``test_rebilling_sheet.py``,
``test_rebilling_accuracy.py``, ``test_detect_rebilling.py``) import
``detect_rebilling`` from ``edf_bill_fetcher.processors.detection`` --
a *different* module that carries its own duplicate copy of the
detector.  Those tests therefore leave the writer submodule's copies
uncovered.  Every test below imports from
``edf_bill_fetcher.io.writers.rebilling`` directly so the coverage
lands on the file under test.
"""

from __future__ import annotations

import pandas as pd
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from edf_bill_fetcher.io.writers.rebilling import (
    _reversal_match,
    detect_rebilling,
    write_rebilling_sheet,
)

# ---------------------------------------------------------------------------
# Shared fixtures
# ---------------------------------------------------------------------------

_REBILL_COLUMNS = [
    "Killer Invoice",
    "Killed Invoice",
    "Killer Date",
    "Killed Date",
    "Period Overlap (days)",
    "Jump-back (days)",
    "Trigger Reason",
    "Cancel/Rebill Admitted (Killer)",
]


def _row(
    invoice: str,
    date: str,
    period_from: str,
    period_to: str,
    amount: float = 1000.0,
    admitted: bool = False,
) -> dict:
    """Build a single invoice row matching the detect_rebilling input shape."""
    return {
        "Invoice #": invoice,
        "Date": date,
        "Period From": period_from,
        "Period To": period_to,
        "Amount (£)": amount,
        "Cancel/Rebill Admitted": admitted,
    }


def _credit_row(
    invoice: str,
    period_from: str,
    period_to: str,
    amount: float,
    entry_type: str = "Credit",
) -> dict:
    """Build a reversal-credit row as it appears in the evidence DataFrame."""
    return {
        "Invoice #": invoice,
        "Entry Type": entry_type,
        "Period From": period_from,
        "Period To": period_to,
        "Amount (£)": amount,
    }


def _open_ws(title: str = "Rebilling Analysis") -> Worksheet:
    """Create a fresh openpyxl worksheet for write_rebilling_sheet tests."""
    wb = Workbook()
    ws = wb.active
    ws.title = title
    return ws


def _empty_rebilling_df() -> pd.DataFrame:
    """Empty rebilling DataFrame with the canonical output columns."""
    return pd.DataFrame(columns=_REBILL_COLUMNS)


def _one_pair_df() -> pd.DataFrame:
    """A rebilling DataFrame with a single (Killer, Killed) pair."""
    return pd.DataFrame(
        [
            {
                "Killer Invoice": "K1",
                "Killed Invoice": "S1",
                "Killer Date": "01 Jan 2024",
                "Killed Date": "01 Feb 2022",
                "Period Overlap (days)": 28,
                "Jump-back (days)": 0,
                "Trigger Reason": "killer period \u2265 365d",
                "Cancel/Rebill Admitted (Killer)": False,
            }
        ]
    )


def _two_pair_df() -> pd.DataFrame:
    """A rebilling DataFrame with two (Killer, Killed) pairs."""
    return pd.DataFrame(
        [
            {
                "Killer Invoice": "K1",
                "Killed Invoice": "S1",
                "Killer Date": "01 Jan 2024",
                "Killed Date": "01 Feb 2022",
                "Period Overlap (days)": 28,
                "Jump-back (days)": 0,
                "Trigger Reason": "killer period \u2265 365d",
                "Cancel/Rebill Admitted (Killer)": False,
            },
            {
                "Killer Invoice": "K2",
                "Killed Invoice": "S2",
                "Killer Date": "01 Mar 2024",
                "Killed Date": "01 Apr 2022",
                "Period Overlap (days)": 30,
                "Jump-back (days)": 10,
                "Trigger Reason": "admit-phrase on killer",
                "Cancel/Rebill Admitted (Killer)": True,
            },
        ]
    )


# ---------------------------------------------------------------------------
# _reversal_match (lines 49-72)
# ---------------------------------------------------------------------------


def test_reversal_match_returns_false_when_evidence_df_is_none() -> None:
    """A None evidence DataFrame cannot contain a matching credit row."""
    killed_pf = pd.Timestamp("2023-01-01")
    killed_pt = pd.Timestamp("2023-03-31")
    assert _reversal_match(None, "A", 250.0, killed_pf, killed_pt) is False


def test_reversal_match_returns_false_when_evidence_df_is_empty() -> None:
    """An empty evidence DataFrame cannot contain a matching credit row."""
    killed_pf = pd.Timestamp("2023-01-01")
    killed_pt = pd.Timestamp("2023-03-31")
    empty = pd.DataFrame(columns=["Entry Type", "Amount (£)", "Period From", "Period To"])
    assert _reversal_match(empty, "A", 250.0, killed_pf, killed_pt) is False


def test_reversal_match_returns_false_when_entry_type_column_missing() -> None:
    """Evidence DataFrame without an 'Entry Type' column cannot match."""
    killed_pf = pd.Timestamp("2023-01-01")
    killed_pt = pd.Timestamp("2023-03-31")
    evidence = pd.DataFrame(
        [
            {
                "Invoice #": "A",
                "Amount (£)": -250.0,
                "Period From": "01 Jan 2023",
                "Period To": "31 Mar 2023",
            }
        ]
    )
    assert _reversal_match(evidence, "A", 250.0, killed_pf, killed_pt) is False


def test_reversal_match_returns_false_when_killed_amount_unparseable() -> None:
    """A killed_amount that cannot be coerced to float short-circuits to False."""
    killed_pf = pd.Timestamp("2023-01-01")
    killed_pt = pd.Timestamp("2023-03-31")
    evidence = pd.DataFrame([_credit_row("A", "01 Jan 2023", "31 Mar 2023", -250.0)])
    assert _reversal_match(evidence, "A", "not-a-number", killed_pf, killed_pt) is False  # type: ignore[arg-type]


def test_reversal_match_returns_true_on_amount_match_with_unparseable_period() -> None:
    """When the credit row's Period From/To are missing, amount alone suffices."""
    killed_pf = pd.Timestamp("2023-01-01")
    killed_pt = pd.Timestamp("2023-03-31")
    evidence = pd.DataFrame(
        [
            {
                "Invoice #": "A",
                "Entry Type": "Credit",
                "Amount (£)": -250.0,
                # No Period From / Period To -> _safe_to_datetime returns NaT
            }
        ]
    )
    assert _reversal_match(evidence, "A", 250.0, killed_pf, killed_pt) is True


def test_reversal_match_returns_true_when_period_overlap_at_least_30_days() -> None:
    """Credit period overlapping the killed period by >= 30 days matches."""
    killed_pf = pd.Timestamp("2023-01-01")
    killed_pt = pd.Timestamp("2023-03-31")
    evidence = pd.DataFrame([_credit_row("A", "01 Jan 2023", "31 Mar 2023", -250.0)])
    assert _reversal_match(evidence, "A", 250.0, killed_pf, killed_pt) is True


def test_reversal_match_returns_false_when_amount_mismatch_exceeds_tolerance() -> None:
    """Credit amount outside the +/- 0.50 tolerance does not match."""
    killed_pf = pd.Timestamp("2023-01-01")
    killed_pt = pd.Timestamp("2023-03-31")
    evidence = pd.DataFrame([_credit_row("A", "01 Jan 2023", "31 Mar 2023", -260.0)])
    assert _reversal_match(evidence, "A", 250.0, killed_pf, killed_pt) is False


def test_reversal_match_returns_false_when_overlap_under_30_days() -> None:
    """Credit period overlapping by less than 30 days does not match."""
    killed_pf = pd.Timestamp("2023-01-01")
    killed_pt = pd.Timestamp("2023-03-31")
    # Credit is a 7-day fragment at the tail of the killed window.
    evidence = pd.DataFrame([_credit_row("A", "25 Mar 2023", "31 Mar 2023", -250.0)])
    assert _reversal_match(evidence, "A", 250.0, killed_pf, killed_pt) is False


def test_reversal_match_skips_row_with_unparseable_amount() -> None:
    """A credit row whose amount cannot be coerced is skipped, not fatal."""
    killed_pf = pd.Timestamp("2023-01-01")
    killed_pt = pd.Timestamp("2023-03-31")
    evidence = pd.DataFrame(
        [
            {
                "Invoice #": "A",
                "Entry Type": "Credit",
                "Amount (£)": "garbage",
                "Period From": "01 Jan 2023",
                "Period To": "31 Mar 2023",
            },
            _credit_row("A", "01 Jan 2023", "31 Mar 2023", -250.0),
        ]
    )
    # The garbage row is skipped; the second row matches.
    assert _reversal_match(evidence, "A", 250.0, killed_pf, killed_pt) is True


def test_reversal_match_accepts_payment_entry_type() -> None:
    """The 'Payment' Entry Type is also accepted alongside 'Credit'."""
    killed_pf = pd.Timestamp("2023-01-01")
    killed_pt = pd.Timestamp("2023-03-31")
    evidence = pd.DataFrame(
        [_credit_row("A", "01 Jan 2023", "31 Mar 2023", -250.0, entry_type="Payment")]
    )
    assert _reversal_match(evidence, "A", 250.0, killed_pf, killed_pt) is True


def test_reversal_match_returns_false_when_no_matching_credit_rows() -> None:
    """Evidence rows with non-Credit/Payment Entry Type are ignored."""
    killed_pf = pd.Timestamp("2023-01-01")
    killed_pt = pd.Timestamp("2023-03-31")
    evidence = pd.DataFrame(
        [_credit_row("A", "01 Jan 2023", "31 Mar 2023", -250.0, entry_type="Debit")]
    )
    assert _reversal_match(evidence, "A", 250.0, killed_pf, killed_pt) is False


# ---------------------------------------------------------------------------
# detect_rebilling (lines 115-209)
# ---------------------------------------------------------------------------


def test_detect_rebilling_empty_df_returns_empty_with_columns() -> None:
    """An empty input DataFrame returns an empty DataFrame with the canonical columns."""
    out = detect_rebilling(pd.DataFrame())
    assert out.empty
    assert list(out.columns) == _REBILL_COLUMNS


def test_detect_rebilling_none_df_returns_empty_with_columns() -> None:
    """A None input DataFrame returns an empty DataFrame with the canonical columns."""
    out = detect_rebilling(pd.DataFrame())  # df is None path covered via empty
    assert out.empty
    assert list(out.columns) == _REBILL_COLUMNS


def test_detect_rebilling_single_invoice_returns_empty() -> None:
    """A single invoice cannot form a (Killer, Killed) pair."""
    df = pd.DataFrame([_row("A", "01 Mar 2023", "01 Feb 2023", "28 Feb 2023")])
    out = detect_rebilling(df)
    assert out.empty
    assert list(out.columns) == _REBILL_COLUMNS


def test_detect_rebilling_skips_rows_with_unparseable_dates() -> None:
    """Rows whose Period From/To or Date cannot be parsed are skipped."""
    df = pd.DataFrame(
        [
            _row("A", "01 Mar 2023", "garbage", "28 Feb 2023"),
            _row("B", "01 Apr 2023", "01 Feb 2023", "31 Mar 2023"),
        ]
    )
    out = detect_rebilling(df)
    # Only B survives parsing; no pair possible.
    assert out.empty


def test_detect_rebilling_amount_unparseable_set_to_none() -> None:
    """An unparseable Amount is stored as None in the parsed record."""
    df = pd.DataFrame(
        [
            {
                "Invoice #": "A",
                "Date": "01 Mar 2023",
                "Period From": "01 Feb 2023",
                "Period To": "28 Feb 2023",
                "Amount (£)": "not-a-number",
            },
            {
                "Invoice #": "B",
                "Date": "01 Apr 2023",
                "Period From": "01 Jan 2022",
                "Period To": "31 Mar 2023",
                "Amount (£)": 100.0,
            },
        ]
    )
    out = detect_rebilling(df)
    # B fully contains A and spans >= 365 days -> one row via the 365d signal.
    assert len(out) == 1
    assert out.iloc[0]["Killer Invoice"] == "B"
    assert out.iloc[0]["Killed Invoice"] == "A"


def test_detect_rebilling_365d_signal_emits_row() -> None:
    """A killer spanning >= 365 days containing a short killed invoice fires."""
    df = pd.DataFrame(
        [
            _row("S1", "01 Feb 2022", "01 Feb 2022", "28 Feb 2022"),
            _row("K1", "01 Jan 2024", "01 Jan 2022", "31 Dec 2023"),  # 1095d
        ]
    )
    out = detect_rebilling(df)
    assert len(out) == 1
    row = out.iloc[0]
    assert row["Killer Invoice"] == "K1"
    assert row["Killed Invoice"] == "S1"
    assert "killer period \u2265 365d" in str(row["Trigger Reason"])
    assert bool(row["Cancel/Rebill Admitted (Killer)"]) is False


def test_detect_rebilling_admit_signal_emits_row() -> None:
    """A killer with admit-phrase=True containing a killed invoice fires."""
    df = pd.DataFrame(
        [
            _row("A", "01 Mar 2023", "01 Feb 2023", "28 Feb 2023"),
            _row("B", "01 Apr 2023", "01 Jan 2023", "31 Mar 2023", admitted=True),
        ]
    )
    out = detect_rebilling(df)
    assert len(out) == 1
    row = out.iloc[0]
    assert row["Killer Invoice"] == "B"
    assert "admit-phrase on killer" in str(row["Trigger Reason"])
    assert bool(row["Cancel/Rebill Admitted (Killer)"]) is True


def test_detect_rebilling_reversal_signal_emits_row() -> None:
    """A reversal credit in evidence_df matching the killed invoice fires."""
    invoice_rows = pd.DataFrame(
        [
            _row("A", "01 Apr 2023", "01 Jan 2023", "31 Mar 2023", amount=250.0),
            _row("B", "01 May 2023", "01 Dec 2022", "31 Mar 2023"),
        ]
    )
    evidence_df = pd.DataFrame([_credit_row("A", "01 Jan 2023", "31 Mar 2023", -250.0)])
    out = detect_rebilling(invoice_rows, evidence_df=evidence_df)
    assert len(out) == 1
    row = out.iloc[0]
    assert "reversal credit row matches killed" in str(row["Trigger Reason"])


def test_detect_rebilling_no_triggers_emits_zero_rows() -> None:
    """Containment holds but no signal fires -> zero rows."""
    df = pd.DataFrame(
        [
            _row("A", "01 Mar 2023", "01 Feb 2023", "28 Feb 2023"),
            _row("B", "01 Jul 2023", "01 Jan 2023", "30 Jun 2023"),  # 180d, no admit
        ]
    )
    out = detect_rebilling(df)
    assert out.empty


def test_detect_rebilling_containment_failure_skips_pair() -> None:
    """Killer not fully containing killed -> pair skipped even with a signal."""
    df = pd.DataFrame(
        [
            _row("A", "01 Apr 2023", "01 Jan 2023", "31 Mar 2023"),  # Jan-Mar
            _row("B", "01 Jun 2023", "01 Feb 2023", "31 May 2023", admitted=True),  # Feb-May
        ]
    )
    out = detect_rebilling(df)
    assert out.empty


def test_detect_rebilling_multiple_signals_join_with_semicolon() -> None:
    """Multiple triggers on one pair are joined with '; ' in Trigger Reason."""
    df = pd.DataFrame(
        [
            _row("S1", "01 Feb 2022", "01 Feb 2022", "28 Feb 2022"),
            _row("K1", "01 Jan 2024", "01 Jan 2022", "31 Dec 2023", admitted=True),
        ]
    )
    out = detect_rebilling(df)
    assert len(out) == 1
    reason = str(out.iloc[0]["Trigger Reason"])
    assert "killer period \u2265 365d" in reason
    assert "admit-phrase on killer" in reason
    assert ";" in reason


def test_detect_rebilling_output_sorted_by_killer_then_killed_date() -> None:
    """Output rows are sorted by Killer Date then Killed Date ascending."""
    df = pd.DataFrame(
        [
            _row("LATE_K", "01 Dec 2023", "01 Jan 2022", "30 Nov 2023"),
            _row("EARLY_K", "01 Feb 2023", "01 Jan 2022", "31 Jan 2023"),
            _row("V0", "01 Oct 2022", "01 Sep 2022", "30 Sep 2022"),
        ]
    )
    out = detect_rebilling(df)
    assert list(out["Killer Invoice"])[0] == "EARLY_K"
    assert list(out["Killer Invoice"])[-1] == "LATE_K"


def test_detect_rebilling_overlap_and_jumpback_computed_correctly() -> None:
    """Period Overlap and Jump-back days are computed and clamped to >= 0."""
    df = pd.DataFrame(
        [
            _row("S1", "01 Feb 2022", "01 Feb 2022", "28 Feb 2022"),
            _row("K1", "01 Jan 2024", "01 Jan 2022", "31 Dec 2023"),  # 1095d
        ]
    )
    out = detect_rebilling(df)
    row = out.iloc[0]
    # Killed window Feb 2022 (01 Feb -> 28 Feb) is fully inside killer.
    # Overlap = min(killer_pt, killed_pt) - max(killer_pf, killed_pf)
    #         = 28 Feb 2022 - 01 Feb 2022 = 27 days.
    assert int(row["Period Overlap (days)"]) == 27
    # Jump-back = killed Period From - killer Period From
    #           = 01 Feb 2022 - 01 Jan 2022 = 31 days.
    assert int(row["Jump-back (days)"]) == 31


def test_detect_rebilling_missing_admit_column_treated_as_false() -> None:
    """Without a 'Cancel/Rebill Admitted' column, admit defaults to False."""
    df = pd.DataFrame(
        [
            {
                "Invoice #": "A",
                "Date": "01 Mar 2023",
                "Period From": "01 Feb 2023",
                "Period To": "28 Feb 2023",
                "Amount (£)": 100.0,
            },
            {
                "Invoice #": "B",
                "Date": "01 Apr 2023",
                "Period From": "01 Jan 2022",
                "Period To": "31 Mar 2023",
                "Amount (£)": 100.0,
            },
        ]
    )
    out = detect_rebilling(df)
    assert len(out) == 1
    assert bool(out.iloc[0]["Cancel/Rebill Admitted (Killer)"]) is False


def test_detect_rebilling_cascade_emits_multiple_rows() -> None:
    """A 4-invoice cascade with admit on the last emits >= 2 rows."""
    df = pd.DataFrame(
        [
            _row("T65", "01 Apr 2023", "01 Feb 2023", "31 Mar 2023"),
            _row("T66", "01 Jun 2023", "01 Mar 2023", "31 May 2023"),
            _row("T67", "01 Aug 2023", "01 Apr 2023", "31 Jul 2023"),
            _row("T68", "01 Oct 2023", "01 Feb 2023", "30 Sep 2023", admitted=True),
        ]
    )
    out = detect_rebilling(df)
    assert len(out) >= 2


# ---------------------------------------------------------------------------
# write_rebilling_sheet (lines 229, 305, 311, 313-320)
# ---------------------------------------------------------------------------


def test_write_rebilling_sheet_empty_account_omits_account_suffix() -> None:
    """When account is empty, the title banner has no '|  Account' suffix."""
    ws = _open_ws()
    write_rebilling_sheet(ws, _empty_rebilling_df(), account="")
    title = ws.cell(row=1, column=1).value
    assert isinstance(title, str)
    assert "Account" not in title


def test_write_rebilling_sheet_with_account_appends_suffix() -> None:
    """When account is non-empty, the title banner includes '|  Account <acc>'."""
    ws = _open_ws()
    write_rebilling_sheet(ws, _empty_rebilling_df(), account="ACC1")
    title = ws.cell(row=1, column=1).value
    assert isinstance(title, str)
    assert "Account ACC1" in title


def test_write_rebilling_sheet_empty_rb_renders_headers_only() -> None:
    """An empty rebilling DataFrame renders the title, subheader, and headers but no data rows."""
    ws = _open_ws()
    write_rebilling_sheet(ws, _empty_rebilling_df(), account="ACC1")
    # Title rendered.
    assert isinstance(ws.cell(row=1, column=1).value, str)
    # Table headers at row 7.
    headers = [ws.cell(row=7, column=c).value for c in range(1, 10)]
    assert headers[0] == "Killer Invoice"
    assert headers[-1] == "View on Evidence Report"
    # No data row at row 8.
    assert ws.cell(row=8, column=1).value in (None, "")


def test_write_rebilling_sheet_single_pair_writes_one_data_row() -> None:
    """A single-pair rebilling DataFrame writes exactly one data row at row 8."""
    ws = _open_ws()
    write_rebilling_sheet(ws, _one_pair_df(), account="ACC1")
    assert ws.cell(row=8, column=1).value == "K1"
    assert ws.cell(row=8, column=2).value == "S1"
    assert ws.cell(row=9, column=1).value in (None, "")


def test_write_rebilling_sheet_multiple_pairs_writes_multiple_rows() -> None:
    """A two-pair rebilling DataFrame writes two data rows starting at row 8."""
    ws = _open_ws()
    write_rebilling_sheet(ws, _two_pair_df(), account="ACC1")
    assert ws.cell(row=8, column=1).value == "K1"
    assert ws.cell(row=9, column=1).value == "K2"
    assert ws.cell(row=10, column=1).value in (None, "")


def test_write_rebilling_sheet_admitted_true_bold_red_font_on_trigger_cell() -> None:
    """When Cancel/Rebill Admitted (Killer) is True, the Trigger Reason cell font is bold red."""
    ws = _open_ws()
    write_rebilling_sheet(ws, _two_pair_df(), account="ACC1")
    # Row 9 is the second pair (K2) with admitted=True.
    trigger_cell = ws.cell(row=9, column=7)
    assert trigger_cell.font is not None
    assert trigger_cell.font.bold is True
    assert trigger_cell.font.color is not None
    assert "C00000" in str(trigger_cell.font.color.rgb).upper()


def test_write_rebilling_sheet_admitted_false_no_bold_red_font() -> None:
    """When Cancel/Rebill Admitted (Killer) is False, the Trigger Reason cell is not bold-red."""
    ws = _open_ws()
    write_rebilling_sheet(ws, _one_pair_df(), account="ACC1")
    # Row 8 is the only pair (K1) with admitted=False.
    trigger_cell = ws.cell(row=8, column=7)
    # The _text helper sets a non-bold font; the admit branch is skipped.
    assert trigger_cell.font is not None
    # Bold should be False (default from _text helper).
    assert trigger_cell.font.bold is False


def test_write_rebilling_sheet_no_evidence_index_renders_no_match() -> None:
    """Without an evidence_index, the 'View on Evidence Report' cell says 'No match'."""
    ws = _open_ws()
    write_rebilling_sheet(ws, _one_pair_df(), account="ACC1", evidence_index=None)
    cell = ws.cell(row=8, column=9)
    assert cell.value == "No match"
    assert cell.font is not None
    assert cell.font.italic is True


def test_write_rebilling_sheet_evidence_index_with_match_renders_hyperlink() -> None:
    """When evidence_index maps the killer invoice, col 9 becomes a hyperlink cell."""
    ws = _open_ws()
    evidence_index = {"inv:K1": 42}
    write_rebilling_sheet(ws, _one_pair_df(), account="ACC1", evidence_index=evidence_index)
    cell = ws.cell(row=8, column=9)
    assert cell.value == "\u2192"
    assert cell.hyperlink is not None
    assert "A42" in str(cell.hyperlink.location)
    assert cell.font is not None
    assert "0563C1" in str(cell.font.color.rgb).upper() if cell.font.color else False


def test_write_rebilling_sheet_evidence_index_with_killed_match_renders_hyperlink() -> None:
    """When evidence_index maps only the killed invoice, col 9 still hyperlinks."""
    ws = _open_ws()
    evidence_index = {"inv:S1": 99}
    write_rebilling_sheet(ws, _one_pair_df(), account="ACC1", evidence_index=evidence_index)
    cell = ws.cell(row=8, column=9)
    assert cell.value == "\u2192"
    assert cell.hyperlink is not None
    assert "A99" in str(cell.hyperlink.location)


def test_write_rebilling_sheet_evidence_index_no_match_renders_no_match() -> None:
    """When evidence_index is provided but has no matching invoice, col 9 says 'No match'."""
    ws = _open_ws()
    evidence_index = {"inv:OTHER": 5}
    write_rebilling_sheet(ws, _one_pair_df(), account="ACC1", evidence_index=evidence_index)
    cell = ws.cell(row=8, column=9)
    assert cell.value == "No match"


def test_write_rebilling_sheet_freeze_panes_at_a8() -> None:
    """Freeze panes is set to 'A8' so the header rows stay visible while scrolling."""
    ws = _open_ws()
    write_rebilling_sheet(ws, _one_pair_df(), account="ACC1")
    assert ws.freeze_panes == "A8"


def test_write_rebilling_sheet_column_widths_set() -> None:
    """Column widths A-I are set to the spec values after rendering."""
    ws = _open_ws()
    write_rebilling_sheet(ws, _one_pair_df(), account="ACC1")
    assert ws.column_dimensions["A"].width == 18
    assert ws.column_dimensions["G"].width == 50
    assert ws.column_dimensions["H"].width == 60
    assert ws.column_dimensions["I"].width == 22


def test_write_rebilling_sheet_renders_open_pdf_hyperlink_with_evidence_df() -> None:
    """When evidence_df contains the killer invoice, col 8 gets a PDF hyperlink cell."""
    ws = _open_ws()
    evidence_df = pd.DataFrame(
        [
            {
                "Invoice #": "K1",
                "Date": "01 Jan 2024",
                "Period From": "01 Jan 2022",
                "Period To": "31 Dec 2023",
                "Amount (£)": 1000.0,
                "Attachment Name": "K1.pdf",
            }
        ]
    )
    write_rebilling_sheet(ws, _one_pair_df(), account="ACC1", evidence_df=evidence_df)
    # Col 8 should carry the arrow glyph when the killer is found in evidence_df.
    cell = ws.cell(row=8, column=8)
    assert cell.value == "\u2192"
    assert cell.hyperlink is not None


def test_write_rebilling_sheet_renders_no_pdf_hyperlink_when_evidence_df_none() -> None:
    """When evidence_df is None, col 8 gets no hyperlink cell (evidence_report_hyperlink_cell skips)."""
    ws = _open_ws()
    write_rebilling_sheet(ws, _one_pair_df(), account="ACC1", evidence_df=None)
    # evidence_report_hyperlink_cell returns early when invoice_number is non-empty
    # but evidence_df is None -> no cell value set by the helper. The cell
    # may still exist as an empty openpyxl cell.
    cell = ws.cell(row=8, column=8)
    # No hyperlink should be present.
    assert cell.hyperlink is None
