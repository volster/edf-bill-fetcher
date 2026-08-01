"""Spec 3 (stretch) regression: amalgamated hybrid duplicate rows.

When the optional ``amalgamate_duplicates`` toggle is True, the dedup
walker keeps a single *hybrid* row per duplicate cluster instead of
just the most-complete row.  Non-empty fields are merged column-wise
across all siblings: each column takes the first non-empty / non-N/A
value from the cluster, walking siblings in completeness-descending
order.

With the toggle OFF (default), the prior Spec 2 (most-complete-wins
verbatim) contract holds unchanged.

Pinning:

  * Toggle OFF: HTM (most-complete by populated-field count) wins;
    PST's Invoice # and Local PDF's Tariff are lost.
  * Toggle ON: the kept row carries HTM's fields PLUS PST's Invoice #
    PLUS Local PDF's Tariff — the other two siblings still surface
    on Duplicate Entries.
"""

from __future__ import annotations

import os
from pathlib import Path

import pytest
from openpyxl import load_workbook

from edf_bill_fetcher.writers import export_to_excel


@pytest.fixture
def workdir() -> Path:
    scratch = Path(os.environ.get("USERPROFILE", "/tmp")) / f".edf_amalg_{os.getpid()}"
    scratch.mkdir(parents=True, exist_ok=True)
    return scratch


def _make_records() -> list[dict]:
    """Three duplicates across the same Period To + Amount, with
    distinct column-wise strengths so the amalgam toggle's hybrid
    column-merge contract differs from the bare completeness winner.

    Row construction:

    - HTM has the *most* populated fields (Per From, Period To, Date
      all present + Units) — completeness winner.
    - PST carries Invoice # and Period Charge (missing Period From
      and Units) — richer on some columns, sparser overall.
    - Local PDF carries Tariff and Reading (missing Period From,
      Units, and most other fields) — sparsest overall.
    """
    base = {
        "Amount (£)": 100.0,
        "Entry Type": "New Bill",
        "Logic Used": "Period",
        "Details": "",
        "Attachment Name": "",
        "Anomaly Flag": "",
        "Sender": "",
    }

    def _row(src, date, pf, pt, inv, pc, units, reading, tariff):
        r = dict(base)
        r.update(
            {
                "Source": src,
                "Date": date,
                "Period From": pf,
                "Period To": pt,
                "Invoice #": inv,
                "Period Charge (£)": pc,
                "Units (kWh)": units,
                "Reading": reading,
                "Tariff": tariff,
                "Standing Chg (p/day)": "",
            }
        )
        return r

    htm = _row(
        src="HTM Account History",
        date="04/04/2024",
        pf="01/03/2024",
        pt="01/04/2024",
        inv="",
        pc=0.0,
        units="100",
        reading="N/A",
        tariff="N/A",
    )
    pst = _row(
        src="PST PDF Attachment",
        date="02/04/2024",
        pf="01/03/2024",
        pt="01/04/2024",
        inv="INV-777",
        pc=80.0,
        units="",
        reading="N/A",
        tariff="N/A",
    )
    pdf = _row(
        src="Local PDF Folder",
        date="03/04/2024",
        pf="N/A",
        pt="01/04/2024",
        inv="",
        pc=0.0,
        units="",
        reading="Actual",
        tariff="Standard Var",
    )
    return [htm, pst, pdf]


def _config(amalgamate: bool) -> dict:
    return {
        "use_dedup": True,
        "save_dups": True,
        "use_anchors": False,
        "use_large": False,
        "min_amount": 0.0,
        "filter_below": False,
        "use_dedup_period": True,
        "expanded_columns": True,
        "include_charts": False,
        "include_forecast": False,
        "amalgamate_duplicates": amalgamate,
    }


class TestAmalgamateToggle:
    """Hybrid row contract when ``amalgamate_duplicates=True``."""

    def test_hybrid_merges_invoice_and_tariff_from_sparser_siblings(
        self,
        workdir: Path,
    ) -> None:
        out = workdir / "out_amalg.xlsx"
        export_to_excel(_make_records(), str(out), [], config=_config(amalgamate=True))
        wb = load_workbook(str(out))
        ws = wb["EDF Evidence Report"]
        kept_rows = list(ws.iter_rows(min_row=2, values_only=True))
        assert len(kept_rows) == 1, (
            f"amalgamation of one cluster should leave one kept row; got {len(kept_rows)}"
        )
        # Column positions (openpyxl 1-based, list 0-based):
        # A=Source B=Sender C=Date D=PeriodFrom E=PeriodTo F=Invoice#
        # G=Amount H=PeriodCharge I=UnitRate J=%Change K=EntryType
        # L=Reading M=Units N=StandingChg O=Tariff P=AttachmentName
        # Q=Details R=LogicUsed S=AnomalyFlag
        row = kept_rows[0]
        assert row[5] == "INV-777", f"Invoice # missing in hybrid; row[5]={row[5]!r}"
        assert row[14] == "Standard Var", f"Tariff missing in hybrid; row[14]={row[14]!r}"
        assert row[11] == "Actual", f"Reading missing in hybrid; row[11]={row[11]!r}"
        assert row[0] == "HTM Account History", f"Source {row[0]!r} should be HTM"

    def test_dup_sheet_still_records_every_dropped_sibling(
        self,
        workdir: Path,
    ) -> None:
        out = workdir / "out_dup_amalg.xlsx"
        export_to_excel(_make_records(), str(out), [], config=_config(amalgamate=True))
        wb = load_workbook(str(out))
        assert "Duplicate Entries" in wb.sheetnames
        ws = wb["Duplicate Entries"]
        dropped = [
            ws.cell(row=r, column=1).value
            for r in range(2, ws.max_row + 1)
            if ws.cell(row=r, column=1).value
        ]
        # Two siblings dropped: PST and Local PDF.
        assert "PST PDF Attachment" in dropped, (
            f"PST sibling must surface on dup sheet; got {dropped}"
        )
        assert "Local PDF Folder" in dropped, f"PDF sibling must surface; got {dropped}"

    def test_toggle_off_keeps_bare_completeness_winner_only(
        self,
        workdir: Path,
    ) -> None:
        """With the toggle OFF, no amalgamation — verbatim
        completeness-wins (Spec 2)."""
        out = workdir / "out_off.xlsx"
        export_to_excel(_make_records(), str(out), [], config=_config(amalgamate=False))
        wb = load_workbook(str(out))
        ws = wb["EDF Evidence Report"]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        assert len(rows) == 1
        row = rows[0]
        # The surviving row is HTM — the completeness winner (most populated
        # fields).
        assert row[0] == "HTM Account History", f"Source {row[0]!r} should be HTM"
        # PST's Invoice # was NOT merged — HTM had no Invoice #, so
        # the verbatim HTM row has an empty Invoice # column.
        assert row[5] in (None, "", "N/A"), (
            f"Invoice # {row[5]!r} should be empty: no amalgamation picks PST's value"
        )
        # PDF's Tariff was NOT merged — HTM had N/A.
        assert row[14] in (None, "", "N/A"), f"Tariff {row[14]!r} should be N/A: no amalgamation"
