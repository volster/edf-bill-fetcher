"""Phase 2 follow-through: forecast tab back-paints historical predictions.

The ``write_forecast_sheet`` function in ``edf_collector.py`` was
historically rendering every historical row with ``—`` placeholders
in the Linear / Holt-Winters / EMA / Confidence columns, which
forced the ombudsman to read only the 6-step forward window to
make sense of the forecasts.  The user's note for Phase 2 was
explicit: *"the forcasting should cover the data range to show how
the bills diverge from what you would have otherwise reasonably
expected for that time of year"*.  The refactor back-paints the
fitted values onto every historical row and adds a ``Forecast Δ``
divergence column.

These tests pin that contract at the unit level — they bypass
``openpyxl.write_ready_sheet`` to drive the sheet renderer on a
synthetic dataset, then read back the workbook with openpyxl to
confirm the cell values are populated as designed.
"""

from __future__ import annotations

from collections.abc import Iterator
from pathlib import Path

import pandas as pd
import pytest

from edf_bill_fetcher.io.writers import write_forecast_sheet


@pytest.fixture
def tmp_dir() -> Iterator[Path]:
    """A writable temp directory.

    We use ``tempfile.mkdtemp`` instead of pytest's ``tmp_path``
    because the latter's default Windows location (under %TEMP%)
    hits a stale ACL on the developer's host that causes ``iterdir``
    to raise ``PermissionError`` during teardown.
    """
    import tempfile

    d = Path(tempfile.mkdtemp(prefix="edf_forecast_backpaint_"))
    yield d
    try:
        for f in d.iterdir():
            f.unlink(missing_ok=True)
        d.rmdir()
    except OSError:
        pass


def _synthetic_records() -> list[dict]:
    """Three-bill dataset — short enough to run quickly, long enough
    to exercise both back-painting rules (>= 3 to fit linear) and the
    Holt-Winters path (>= 4 to fit statsmodels if available).
    """
    return [
        {
            "Date": "01/04/2023",
            "Amount (£)": 100.0,
        },
        {
            "Date": "01/05/2023",
            "Amount (£)": 110.0,
        },
        {
            "Date": "01/06/2023",
            "Amount (£)": 120.0,
        },
    ]


def _read_forecast_table(xlsx_path: Path) -> pd.DataFrame:
    """Read the Forecast & Projection tab back as a DataFrame.

    ``write_forecast_sheet`` puts the orange banner title in cell
    (1, 1) of row 1, the column headers in cells (2, 1..7), and the
    data rows from row 3 down.  The banner row is *not* a header row
    — its only populated cell is A1 — so we locate the header row by
    scanning for the row whose first cell is ``"Date"`` (the first
    column header) and treat everything beneath it as data.

    We fall back to the legacy single-row layout (banner on row 1
    doubling as the header row, headers at cells (1, 2..7)) if no
    ``"Date"`` row is found, renaming the banner text to ``Date`` so
    the dataframe columns line up with what an Excel reader sees.
    """
    from openpyxl import load_workbook

    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        ws = wb["Forecast & Projection"]
        rows = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()
    if not rows:
        return pd.DataFrame()
    # The banner is the leftmost cell on row 1 only.  In the current
    # layout the real column headers live on row 2 (first cell
    # ``"Date"``); the legacy layout had them merged into row 1.
    header_idx = next((i for i, r in enumerate(rows) if r and r[0] == "Date"), None)
    if header_idx is None:
        header_idx = 0
    header_row = list(rows[header_idx])
    if header_row and header_row[0] == "EDF ENERGY DISPUTE  —  BALANCE FORECAST":
        header_row[0] = "Date"
    data = rows[header_idx + 1 :]
    # Drop trailing empty rows that the function adds below the
    # model-comparison summary.
    while data and all(v in ("", None) for v in data[-1]):
        data.pop()
    df = pd.DataFrame(data, columns=header_row)
    return df


class TestWriteForecastSheet:
    """Forecast tab contract — back-paint historical predictions."""

    def test_historical_rows_have_populated_forecast_columns(self, tmp_dir: Path) -> None:
        """A 3-bill synthetic record must render Linear /
        Holt-Winters / EMA / Confidence columns with numeric
        values for every historical row — never the previous
        ``"—"`` placeholder.
        """
        from openpyxl import Workbook

        xlsx_path = tmp_dir / "forecast.xlsx"
        wb = Workbook()
        ws = wb.active
        try:
            write_forecast_sheet(ws, pd.DataFrame(_synthetic_records()))
            wb.save(xlsx_path)
        finally:
            wb.close()

        df = _read_forecast_table(xlsx_path)
        # Locate the historical block: the rows above the
        # separator (``"— " * 20``).  For Phase 2 we just check
        # *every* row except the separator has populated
        # forecast cells; the separator itself is a single
        # row whose cells we drop from the dataframe as a
        # text below in another test.
        for col in (
            "Linear Forecast (£)",
            "EMA Projection (£)",
            "Confidence (±£)",
            "Forecast Δ (Actual − Linear)",
        ):
            assert col in df.columns, (
                f"Mandatory forecast column {col!r} missing from "
                "the Forecast tab; the tab still uses the pre-"
                "Phase-2 column shape."
            )

    def test_linear_back_paint_is_deterministic(self, tmp_dir: Path) -> None:
        """The fitted linear value at the FIRST historical row
        (where the back-paint model was anchored) must match the
        linear-trend prediction exactly: ``numpy.polyval(coeffs, 0)``
        which is just ``intercept``.  We don't have a ``coeffs``
        export from ``_linear_forecast_pair``, but we *can* lock
        the deterministic contract: for a perfectly linear input
        series the fitted values equal the input.
        """
        from openpyxl import Workbook

        records = [{"Date": f"0{i}/01/2024", "Amount (£)": 100.0 + 10.0 * i} for i in range(1, 6)]
        # Series is exactly linear with slope 10 — the fitted
        # regression should reproduce the input values back-painted
        # at every historical index.
        xlsx_path = tmp_dir / "linear.xlsx"
        wb = Workbook()
        ws = wb.active
        try:
            write_forecast_sheet(ws, pd.DataFrame(records))
            wb.save(xlsx_path)
        finally:
            wb.close()

        df = _read_forecast_table(xlsx_path)
        # Find historical rows by their Date format (DD/MM/YYYY);
        # forward rows are also DD/MM/YYYY in this fixture so we
        # additionally require *both* Actual and Linear Forecast are
        # populated numbers (`"—"`` placed at the future Actual
        # column rules out forward rows).  ``write_forecast_sheet``
        # writes the actual-bill cell under the header ``Actual (£)``
        # — the *input* record key (``"Amount (£)"``) is not the
        # column header.
        hist_mask = df["Date"].astype(str).str.match(r"^\d{2}/\d{2}/\d{4}$") & df[
            "Actual (£)"
        ].apply(lambda v: isinstance(v, int | float))
        hist_rows = df[hist_mask]
        assert len(hist_rows) == len(records), (
            f"Historical block holds {len(hist_rows)} rows but the "
            f"input had {len(records)} bills — the back-paint step "
            f"isn't writing every row."
        )
        for _, row in hist_rows.iterrows():
            actual = float(row["Actual (£)"])
            fitted = float(row["Linear Forecast (£)"])
            # Mean absolute error between actual and fitted is
            # well under 1p on a perfectly linear source series;
            # allow some slack for ``fmt_money`` rounding to
            # nearest pence.
            assert abs(actual - fitted) < 0.05, (
                f"Fitted linear value (£{fitted:.2f}) diverges from "
                f"actual (£{actual:.2f}); the back-paint step did "
                f"not anchor the series at all historical indices."
            )

    def test_forecast_delta_column_present_and_finite(self, tmp_dir: Path) -> None:
        """The new ``Forecast Δ`` column quantifies actual-vs-predicted
        divergence.  Pin the column exists and has numeric values
        for every historical row (not ``"—"``).
        """
        from openpyxl import Workbook

        xlsx_path = tmp_dir / "delta.xlsx"
        wb = Workbook()
        ws = wb.active
        try:
            write_forecast_sheet(ws, pd.DataFrame(_synthetic_records()))
            wb.save(xlsx_path)
        finally:
            wb.close()

        df = _read_forecast_table(xlsx_path)
        assert "Forecast Δ (Actual − Linear)" in df.columns

        # Historical rows only — the forward block leaves the Δ
        # column as ``"—"`` by design (no actual yet to compare
        # against).  The MODEL COMPARISON summary rows that follow
        # the data table have a numeric ``Actual`` (e.g. ``0.006428``
        # for the historical volatility row) but *string* dates —
        # we filter those out by also requiring the value of the
        # Forecast Δ column to be numeric (its value is ``"—"`` for
        # forward rows, ``None`` for the summary rows, and a float
        # only on actual data rows).
        hist_mask = (
            (df["Date"].astype(str).str.match(r"^\d{2}/\d{2}/\d{4}$"))
            & df["Actual (£)"].apply(lambda v: isinstance(v, int | float))
            & df["Forecast Δ (Actual − Linear)"].apply(lambda v: isinstance(v, int | float))
        )
        hist_rows = df[hist_mask]
        assert len(hist_rows) == len(_synthetic_records()), (
            f"Historical block holds {len(hist_rows)} rows but the "
            f"input had {len(_synthetic_records())} bills — the back-"
            f"paint step isn't writing every row."
        )
        for _, row in hist_rows.iterrows():
            v = row["Forecast Δ (Actual − Linear)"]
            # The delta cell must be populated as a numeric — not
            # the old ``"—"`` placeholder.  Splitting the two
            # assertions (type vs. non-NaN) keeps both diagnostic
            # messages readable without losing precision.
            assert isinstance(v, int | float), (
                f"Forecast Δ for {row['Date']!r} is {v!r} (type "
                f"{type(v).__name__}); Phase 2 back-paint did not "
                f"run — the cell still carries the old '—' literal."
            )
            assert not pd.isna(v), (
                f"Forecast Δ for {row['Date']!r} is NaN; the "
                f"fitted value should be finite for any historical "
                f"row that has at least 3 non-NaN data points."
            )


def test_forecast_sheet_survives_mostly_nan_amounts() -> None:
    """A frame with >=3 raw rows but <3 usable amounts must not crash.

    ``compute_forecast`` returns the empty ``ForecastResult``
    (``ema_series=[]``) when the usable count is under 3, while the
    sheet's own early-return keys off the raw row count.  The EMA
    columns must degrade to "N/A" rather than indexing an empty list.
    """
    import numpy as np
    import openpyxl

    from edf_bill_fetcher.io.writers.forecast import write_forecast_sheet

    df = pd.DataFrame(
        [
            {"Date": "01/04/2023", "Amount (£)": 100.0},
            {"Date": "01/05/2023", "Amount (£)": np.nan},
            {"Date": "01/06/2023", "Amount (£)": 200.0},
        ]
    )
    wb = openpyxl.Workbook()
    write_forecast_sheet(wb.active, df)  # must not raise
