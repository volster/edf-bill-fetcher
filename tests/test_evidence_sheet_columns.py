"""Column-letter invariants for the evidence sheet writer (F1 / SEV-1).

Every COL_* in ``edf_bill_fetcher/io/writers/evidence.py`` is derived from
the ``EVIDENCE_HEADERS`` list, and live formulas must reference the derived
letters — not hard-coded columns.  In particular the Amount column is **G**
(not E, which is Period To), so the ``% Change`` / ``Anomaly Flag`` formulas
and the summary sheet's cross-sheet ``amt_col`` range must all use G.
"""

from __future__ import annotations

import pandas as pd
from openpyxl import Workbook

from edf_bill_fetcher.io.writers.evidence import (
    COL_AMOUNT,
    EVIDENCE_HEADERS,
    write_evidence_sheet,
    write_summary_sheet,
)


def _fixture_df() -> pd.DataFrame:
    """Minimal DataFrame with the columns the evidence writer expects."""
    return pd.DataFrame(
        {
            "Date": ["01 Jan 2024", "15 Feb 2024", "20 Mar 2024"],
            "Invoice #": ["T-001", "T-002", "T-003"],
            "Amount (£)": [100.0, 200.0, 300.0],
            "Period From": ["01 Jan 2024", "01 Feb 2024", "01 Mar 2024"],
            "Period To": ["31 Jan 2024", "29 Feb 2024", "31 Mar 2024"],
            "Period Charge (£)": [100.0, 200.0, 300.0],
            "Units (kWh)": [100, 200, 300],
            "Entry Type": ["New Bill", "New Bill", "New Bill"],
            "Source": ["PST", "PST", "PST"],
            "Sender": ["edf@example.com", "edf@example.com", "edf@example.com"],
            "Tariff": ["Standard", "Standard", "Standard"],
            "Standing Chg (p/day)": [50.0, 50.0, 50.0],
            "Unit Rate (p/kWh)": [24.0, 24.0, 24.0],
            "Reading": [1000, 1200, 1400],
            "Details": ["Bill 1", "Bill 2", "Bill 3"],
            "Attachment Name": ["bill1.pdf", "bill2.pdf", "bill3.pdf"],
        }
    )


def _col_letter(header: str) -> str:
    idx = EVIDENCE_HEADERS.index(header) + 1
    from openpyxl.utils import get_column_letter

    return str(get_column_letter(idx))


def test_percent_change_formula_references_amount_column_g() -> None:
    """% Change must divide Amount (col G) deltas, not Period To (col E)."""
    assert _col_letter("Amount (£)") == "G", "fixture assumption: Amount is G"
    assert _col_letter("Period To") == "E", "fixture assumption: Period To is E"
    assert COL_AMOUNT == 7

    wb = Workbook()
    ws = wb.active
    write_evidence_sheet(ws, _fixture_df())

    pct_col = _col_letter("% Change")
    # Row 3 compares amounts G3 vs G2.
    formula = ws[f"{pct_col}3"].value
    assert isinstance(formula, str), f"% Change cell must be a formula, got: {formula!r}"
    assert formula.startswith("=IFERROR(("), formula
    assert "(G3-G2)" in formula, f"% Change formula must use G, got: {formula}"
    assert "E2" not in formula, f"% Change formula must not reference E2, got: {formula}"
    assert "E3" not in formula, f"% Change formula must not reference E3, got: {formula}"


def test_anomaly_flag_formula_references_amount_column_g() -> None:
    """Anomaly Flag must compare Amount (col G) values, not Period To (col E)."""
    wb = Workbook()
    ws = wb.active
    write_evidence_sheet(ws, _fixture_df())

    anomaly_col = _col_letter("Anomaly Flag")
    formula = ws[f"{anomaly_col}3"].value
    assert isinstance(formula, str), f"Anomaly cell must be a formula, got: {formula!r}"
    assert formula.startswith("=IF(AND("), formula
    assert "G2>0" in formula, f"Anomaly Flag formula must use G, got: {formula}"
    assert "G3>G2*2" in formula, f"Anomaly Flag formula must use G, got: {formula}"
    assert "E2" not in formula, f"Anomaly Flag formula must not reference E2, got: {formula}"
    assert "E3" not in formula, f"Anomaly Flag formula must not reference E3, got: {formula}"


def test_summary_amt_col_references_amount_column_g() -> None:
    """Annual Summary cross-sheet formulas must aggregate evidence col G."""
    wb = Workbook()
    ws = wb.active
    write_summary_sheet(ws, years=[2024], evidence_sheet_name="EDF Evidence Report")

    # B2 holds the balance-range formula built from amt_col (MAXIFS-MINIFS).
    b2 = ws["B2"].value
    assert isinstance(b2, str), f"summary B2 must be a formula, got: {b2!r}"
    assert "$G$2:$G$" in b2, f"summary amt_col must reference G, got: {b2}"
    assert "$E$2:$E$" not in b2, f"summary amt_col must not reference col E, got: {b2}"


def test_anomaly_conditional_formatting_uses_derived_column() -> None:
    """Conditional formatting range must track COL_ANOMALY, not hard-coded S."""
    wb = Workbook()
    ws = wb.active
    write_evidence_sheet(ws, _fixture_df())

    anomaly_col = _col_letter("Anomaly Flag")
    ranges = [str(cf.sqref) for cf in ws.conditional_formatting]
    assert any(rng.startswith(f"{anomaly_col}2:") for rng in ranges), (
        f"no conditional-formatting range on {anomaly_col}2...; got {ranges}"
    )
    assert not any(rng.startswith("S2:") and anomaly_col != "S" for rng in ranges), (
        f"hard-coded S range present; got {ranges}"
    )
