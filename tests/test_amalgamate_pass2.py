"""Spec 3 (stretch) edge-case regression: amalgamation when Pass-2 is the
only dedup path.

When the dedup walker fires Pass-2 (no-period same-amount ±60d windows)
for a cluster that has no Pass-1 match, the amalgamate path's lookup
via ``kept_pass1_index`` silently skips and falls back to bare-dedup
behaviour (each sibling kept verbatim, no per-column merge).

Per the spec toggle, amalgamation should produce a single hybrid row
per cluster *regardless* of which pass detected the duplicates; the
column-merging contract is identical to Pass-1's.

Pin the contract: two no-period same-amount records that Pass-2
collapses should hybridize when ``amalgamate_duplicates=True``.
"""

from __future__ import annotations

import os
from pathlib import Path

import pytest
from openpyxl import load_workbook

from edf_bill_fetcher.writers import export_to_excel


@pytest.fixture
def workdir() -> Path:
    scratch = Path(os.environ.get("USERPROFILE", "/tmp")) / f".edf_pass2_amalg_{os.getpid()}"
    scratch.mkdir(parents=True, exist_ok=True)
    return scratch


def _make_records() -> list[dict]:
    """Two no-period Pass-2 records, same amount, within 60d window:

    - Local PDF Folder, Date 01/03/2024, Amount 100.0, Invoice # "INV-001",
      Period Charge 0.0 (sparser).
    - PST PDF Attachment, Date 15/02/2024, Amount 100.0, Invoice # "INV-002",
      Period Charge 80.0, Tariff "Standard" (richer).

    Both records carry ``Period To = "N/A"`` so Pass-1 ignores them.
    Pass-2 collapses onto the kept anchor (whichever has the later
    bucket position under reverse-iteration).  With
    ``amalgamate_duplicates=True`` the kept set must contain exactly
    one row — the hybrid — and the dropped sibling must surface on
    the Duplicate Entries sheet.
    """
    rec_a = {
        "Date": "01/03/2024",
        "Source": "Local PDF Folder",
        "Period From": "",
        "Period To": "N/A",
        "Invoice #": "INV-001",
        "Money": 100.0,
        "Amount (£)": 100.0,
        "Period Charge (£)": 0.0,
        "Units (kWh)": "",
        "Reading": "",
        "Tariff": "N/A",
        "Standing Charge": "",
        "Entry Type": "New Bill",
        "Logic Used": "Amount",
        "Details": "",
        "Attachment Name": "",
        "Anomaly Flag": "",
        "Sender": "",
    }
    rec_b = {
        "Date": "15/02/2024",
        "Source": "PST PDF Attachment",
        "Period From": "",
        "Period To": "N/A",
        "Invoice #": "INV-002",
        "Money": 100.0,
        "Amount (£)": 100.0,
        "Period Charge (£)": 80.0,
        "Units (kWh)": "",
        "Reading": "",
        "Tariff": "Standard",
        "Standing Charge": "",
        "Entry Type": "New Bill",
        "Logic Used": "Amount",
        "Details": "",
        "Attachment Name": "",
        "Anomaly Flag": "",
        "Sender": "",
    }
    return [rec_a, rec_b]


def _config(amalgamate: bool) -> dict:
    return {
        "use_dedup": True,
        "save_dups": True,
        "use_anchors": False,
        "use_large": False,
        "min_amount": 0.0,
        "filter_below": False,
        "use_dedup_period": True,
        "expanded_columns": True,
        "include_charts": False,
        "include_forecast": False,
        "amalgamate_duplicates": amalgamate,
    }


class TestAmalgamatePass2Coverage:
    """Amalgamation must produce a single hybrid row per cluster even
    when only Pass-2 is the dedup path.
    """

    def test_pass2_cluster_hybridizes_under_amalgamate(self, workdir: Path) -> None:
        out = workdir / "out_pass2.xlsx"
        export_to_excel(_make_records(), str(out), [], config=_config(amalgamate=True))
        wb = load_workbook(str(out))
        ws = wb["EDF Evidence Report"]
        rows = [
            {cell.column_letter: cell.value for cell in row if cell.value is not None}
            for row in ws.iter_rows(min_row=2)
        ]
        # Filter out fully-empty rows (post-Dropna).
        nonempty = [r for r in rows if r]
        assert len(nonempty) == 1, (
            f"Pass-2 cluster under amalgamate must produce exactly 1 hybrid; "
            f"got {len(nonempty)} rows: {nonempty}"
        )

    def test_pass2_hybrid_carries_columns_from_dropped_sibling(self, workdir: Path) -> None:
        """Per-column value of the hybrid must come from the dropped
        sibling when the kept anchor's column is empty/N/A."""
        out = workdir / "out_pass2_cols.xlsx"
        export_to_excel(_make_records(), str(out), [], config=_config(amalgamate=True))
        wb = load_workbook(str(out))
        ws = wb["EDF Evidence Report"]
        cell_by_col = {cell.column_letter: cell.value for cell in ws[2] if cell.value is not None}
        # Period Charge (£) = column H; rec_b's 80.0 must win.
        assert cell_by_col.get("H") == 80.0, (
            f"Pass-2 amalgamate must carry dropped sibling's Period Charge; "
            f"got H={cell_by_col.get('H')!r}"
        )

    def test_pass2_dropped_sibling_surface_on_dup_sheet(self, workdir: Path) -> None:
        """Spec: 'never drop without being recorded'.  Pass-2 dup must
        still appear on the Duplicate Entries sheet.
        """
        out = workdir / "out_pass2_dup.xlsx"
        export_to_excel(_make_records(), str(out), [], config=_config(amalgamate=True))
        wb = load_workbook(str(out))
        assert "Duplicate Entries" in wb.sheetnames
        ws = wb["Duplicate Entries"]
        dropped = [
            ws.cell(row=r, column=1).value
            for r in range(2, ws.max_row + 1)
            if ws.cell(row=r, column=1).value
        ]
        assert len(dropped) == 1, f"Pass-2 dup must surface on dup sheet; got {dropped}"
        # Either Local PDF or PST was the dropped sibling (which one
        # depends on reverse-iteration bucket choice).  Either is fine
        # — we just need exactly one of them dropped.
        assert dropped[0] in ("Local PDF Folder", "PST PDF Attachment"), (
            f"dropped source {dropped[0]!r} is not a known source label"
        )
