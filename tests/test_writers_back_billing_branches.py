"""Branch-coverage tests for edf_bill_fetcher/io/writers/back_billing.py.

Targets the missed-line inventory reported by coverage:
  * _assess_reason (lines 49-65) — both admitted=True and admitted=False narratives.
  * detect_back_billing (lines 119-173) — empty/None df, no events, single event,
    multi events, with and without the Cancel/Rebill Admitted column, unparseable
    Period From/To, unparseable Amount, Bill Date sort key, output column shape.
  * write_back_billing_sheet (lines 297, 330-331, 333-340) — Bill Date as a
    pandas Timestamp, the evidence_index amt_days fallback path, the
    TypeError/ValueError guard around that fallback, and the hyperlink cell
    emitted when a target row is resolved.

The tests exercise the writer module's OWN public surface (imported from
``edf_bill_fetcher.io.writers.back_billing``) rather than the re-exported
``processors.detection.detect_back_billing`` copy the sibling suites use, so
the writer module's lines are the ones recorded as covered.
"""

from __future__ import annotations

from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from edf_bill_fetcher.io.writers.back_billing import (
    _assess_reason,
    detect_back_billing,
    write_back_billing_sheet,
)

# Canonical output column set produced by detect_back_billing. Kept here as a
# module constant so every shape assertion names the same contract.
_EXPECTED_COLUMNS = [
    "Invoice #",
    "Bill Date",
    "Period From",
    "Period To",
    "Days Billed",
    "Period Charge (£)",
    "Value Source",
    "12-Month Limit (days)",
    "Excess Days",
    "Unlawful Charge (£)",
    "Cancel/Rebill Admitted",
    "Reason Assessment",
]


def _row(
    invoice: str = "T-001",
    date: str = "01 Jan 2025",
    period_from: str = "01 Jan 2023",
    period_to: str = "31 Dec 2023",
    amount: float = 1000.0,
    admitted: bool | None = None,
    attachment: str = "T-001.pdf",
) -> dict:
    """Build a single evidence-row dict matching the detector's input schema."""
    out: dict = {
        "Invoice #": invoice,
        "Date": date,
        "Period From": period_from,
        "Period To": period_to,
        "Amount (£)": amount,
        "Attachment Name": attachment,
    }
    if admitted is not None:
        out["Cancel/Rebill Admitted"] = admitted
    return out


def _open_ws(title: str = "Back-billing Analysis") -> Worksheet:
    """Create a fresh openpyxl worksheet, mirroring the sibling writer tests."""
    wb = Workbook()
    ws = wb.active
    ws.title = title
    return ws


# ---------------------------------------------------------------------------
# _assess_reason (lines 49-65)
# ---------------------------------------------------------------------------


def test_assess_reason_admitted_branch_mentions_admission() -> None:
    """Cover the admitted=True branch (lines 53-58).

    The narrative for an admitted cancellation/reversal must name the cover-page
    admission as direct evidence and embed the invoice number, bill date,
    formatted period endpoints, and excess day count.
    """
    pf = pd.Timestamp("2022-04-04")
    pt = pd.Timestamp("2023-07-26")
    bd = pd.Timestamp("2024-08-09")
    text = _assess_reason("T-6715690", bd, 113, True, pf, pt)
    assert isinstance(text, str)
    assert "T-6715690" in text
    assert "09 Aug 2024" in text
    assert "04 Apr 2022" in text
    assert "26 Jul 2023" in text
    assert "113" in text
    assert "admits a cancellation/reversal" in text


def test_assess_reason_not_admitted_branch_mentions_missing_phrase() -> None:
    """Cover the admitted=False branch (lines 60-64).

    Without an admit phrase the narrative must say so explicitly and still carry
    the invoice, bill date, and formatted period span.
    """
    pf = pd.Timestamp("2020-10-01")
    pt = pd.Timestamp("2023-07-07")
    bd = pd.Timestamp("2024-08-09")
    text = _assess_reason("KI-0001", bd, 600, False, pf, pt)
    assert isinstance(text, str)
    assert "KI-0001" in text
    assert "09 Aug 2024" in text
    assert "01 Oct 2020" in text
    assert "07 Jul 2023" in text
    assert "600" in text
    assert "No admit-phrase was found on the cover page." in text
    # And must NOT carry the admission wording.
    assert "admits a cancellation/reversal" not in text


# ---------------------------------------------------------------------------
# detect_back_billing (lines 119-173)
# ---------------------------------------------------------------------------


def test_detect_back_billing_none_df_returns_empty_with_columns() -> None:
    """Cover line 131-132 (df is None early return)."""
    out = detect_back_billing(None)  # type: ignore[arg-type]
    assert out.empty
    assert list(out.columns) == _EXPECTED_COLUMNS


def test_detect_back_billing_empty_df_returns_empty_with_columns() -> None:
    """Cover line 131-132 (df.empty early return)."""
    out = detect_back_billing(pd.DataFrame())
    assert out.empty
    assert list(out.columns) == _EXPECTED_COLUMNS


def test_detect_back_billing_no_long_periods_returns_empty_with_columns() -> None:
    """Cover lines 134-142, 166-168 (loop runs, no row exceeds 365, out.empty)."""
    df = pd.DataFrame(
        [
            _row(
                invoice="A", date="01 Jan 2024", period_from="01 Dec 2023", period_to="28 Dec 2023"
            ),
            _row(
                invoice="B", date="01 Jan 2024", period_from="01 Jan 2023", period_to="31 Dec 2023"
            ),
        ]
    )
    out = detect_back_billing(df)
    assert out.empty
    assert list(out.columns) == _EXPECTED_COLUMNS


def test_detect_back_billing_single_long_period_with_admit_column() -> None:
    """Cover lines 143-165 for a single admitted back-bill (has_admit=True, admitted=True)."""
    df = pd.DataFrame(
        [
            _row(
                invoice="T-6715690",
                date="09 Aug 2024",
                period_from="04 Apr 2022",
                period_to="26 Jul 2023",  # 478 days span, bill >365 days after Period To
                amount=4401.07,
                admitted=True,
            )
        ]
    )
    out = detect_back_billing(df)
    assert len(out) == 1
    row = out.iloc[0]
    assert row["Invoice #"] == "T-6715690"
    assert int(row["Days Billed"]) == 478
    assert int(row["Excess Days"]) > 0
    assert int(row["12-Month Limit (days)"]) == 365
    assert float(row["Period Charge (£)"]) == 4401.07
    assert row["Value Source"] == "Amount (fallback)"
    assert bool(row["Cancel/Rebill Admitted"]) is True
    assert isinstance(row["Reason Assessment"], str)
    assert "admits a cancellation/reversal" in row["Reason Assessment"]
    assert list(out.columns) == _EXPECTED_COLUMNS


def test_detect_back_billing_single_long_period_without_admit_column() -> None:
    """Cover line 148 (has_admit=False -> admitted defaults to False)."""
    df = pd.DataFrame(
        [
            _row(
                invoice="NO-ADMIT-COL",
                date="09 Aug 2024",
                period_from="04 Apr 2022",
                period_to="26 Jul 2023",
                amount=100.0,
            )
        ]
    )
    out = detect_back_billing(df)
    assert len(out) == 1
    assert bool(out.iloc[0]["Cancel/Rebill Admitted"]) is False
    assert "No admit-phrase was found" in out.iloc[0]["Reason Assessment"]


def test_detect_back_billing_multiple_long_periods_sorted_by_bill_date() -> None:
    """Cover lines 166-173 (out non-empty, sort by parsed Bill Date, reindex)."""
    df = pd.DataFrame(
        [
            _row(
                invoice="LATE",
                date="01 Dec 2024",
                period_from="01 Jan 2021",
                period_to="30 Nov 2023",
                amount=5000.0,
                admitted=False,
            ),
            _row(
                invoice="EARLY",
                date="01 Jan 2024",
                period_from="01 Jan 2021",
                period_to="31 Dec 2022",
                amount=3000.0,
                admitted=True,
            ),
        ]
    )
    out = detect_back_billing(df)
    assert list(out["Invoice #"]) == ["EARLY", "LATE"]
    assert out.index.tolist() == [0, 1]  # reset_index applied
    assert list(out.columns) == _EXPECTED_COLUMNS


def test_detect_back_billing_unparseable_period_rows_skipped() -> None:
    """Cover line 138-139 (pd.isna(pf) or pd.isna(pt) -> continue)."""
    df = pd.DataFrame(
        [
            _row(invoice="BAD-PF", date="01 Jan 2025", period_from="N/A", period_to="31 Dec 2023"),
            _row(
                invoice="BAD-PT", date="01 Jan 2025", period_from="01 Jan 2022", period_to="garbage"
            ),
            _row(
                invoice="GOOD",
                date="01 Jan 2025",
                period_from="01 Jan 2022",
                period_to="31 Dec 2023",
                amount=200.0,
            ),
        ]
    )
    out = detect_back_billing(df)
    assert set(out["Invoice #"]) == {"GOOD"}
    assert len(out) == 1


def test_detect_back_billing_unparseable_amount_falls_back_to_zero() -> None:
    """Cover lines 144-147 (float(amt_raw) raises -> charge = 0.0)."""
    df = pd.DataFrame(
        [
            _row(
                invoice="BAD-AMT",
                date="09 Aug 2024",
                period_from="04 Apr 2022",
                period_to="26 Jul 2023",
                amount="not-a-number",  # type: ignore[arg-type]
                admitted=False,
            )
        ]
    )
    out = detect_back_billing(df)
    assert len(out) == 1
    assert float(out.iloc[0]["Period Charge (£)"]) == 0.0


def test_detect_back_billing_unparseable_bill_date_skips_row() -> None:
    """Cover the bill_date NaT guard (bill_date_dt is NaT -> continue).

    Under the new legal rule, the bill Date is required to compute the
    eligibility gate (Date - Period To > 365).  A row with an unparseable
    bill Date cannot be evaluated and is skipped.
    """
    df = pd.DataFrame(
        [
            _row(
                invoice="NODATE",
                date="garbage-date",
                period_from="01 Jan 2022",
                period_to="31 Dec 2023",
                amount=100.0,
                admitted=False,
            ),
            _row(
                invoice="DATED",
                date="01 Jan 2025",
                period_from="01 Jan 2022",
                period_to="31 Dec 2023",
                amount=200.0,
                admitted=False,
            ),
        ]
    )
    out = detect_back_billing(df)
    assert list(out["Invoice #"]) == ["DATED"]


# ---------------------------------------------------------------------------
# write_back_billing_sheet (lines 297, 330-331, 333-340)
# ---------------------------------------------------------------------------


def _bb_row(
    invoice: str = "KI-0001",
    bill_date: object = "01 Jan 2024",
    period_from: object = "01 Jan 2022",
    period_to: object = "31 Dec 2023",
    days_billed: int = 730,
    period_charge: object = 1347.96,
    value_source: str = "Period Charge",
    excess_days: int = 365,
    unlawful_charge: object = 673.98,
    admitted: bool = False,
    reason: str = "back-billing",
) -> dict:
    """Build a single back-billing output row dict (the writer's input shape)."""
    return {
        "Invoice #": invoice,
        "Bill Date": bill_date,
        "Period From": period_from,
        "Period To": period_to,
        "Days Billed": days_billed,
        "Period Charge (£)": period_charge,
        "Value Source": value_source,
        "12-Month Limit (days)": 365,
        "Excess Days": excess_days,
        "Unlawful Charge (£)": unlawful_charge,
        "Cancel/Rebill Admitted": admitted,
        "Reason Assessment": reason,
    }


def test_write_back_billing_sheet_bill_date_as_timestamp_is_formatted() -> None:
    """Cover line 297 (bill_date_val is pd.Timestamp -> strftime).

    When the Bill Date column already holds a pandas Timestamp, the writer must
    render it as a 'DD Mon YYYY' string in column 2 rather than the raw repr.
    """
    bb = pd.DataFrame(
        [
            _bb_row(
                invoice="KI-TS",
                bill_date=pd.Timestamp("2024-01-15"),
                period_from=pd.Timestamp("2022-01-01"),
                period_to=pd.Timestamp("2024-01-01"),
            )
        ]
    )
    ws = _open_ws()
    write_back_billing_sheet(ws, bb, account="A1")
    # Column 2 of row 8 is the Bill Date cell.
    bill_date_cell = ws.cell(row=8, column=2).value
    assert bill_date_cell == "15 Jan 2024"
    # Period From / Period To are also Timestamps -> formatted (lines 299-300,
    # 302-303 exercised as a side effect).
    assert ws.cell(row=8, column=3).value == "01 Jan 2022"
    assert ws.cell(row=8, column=4).value == "01 Jan 2024"


def test_write_back_billing_sheet_bill_date_as_datetime_is_formatted() -> None:
    """Cover line 297 with a stdlib datetime instance (the | datetime branch)."""
    bb = pd.DataFrame(
        [
            _bb_row(
                invoice="KI-DT",
                bill_date=datetime(2024, 2, 29),
                period_from=datetime(2022, 1, 1),
                period_to=datetime(2024, 1, 1),
            )
        ]
    )
    ws = _open_ws()
    write_back_billing_sheet(ws, bb, account="A1")
    assert ws.cell(row=8, column=2).value == "29 Feb 2024"


def test_write_back_billing_sheet_evidence_index_amt_days_fallback_path_exercised() -> None:
    """Exercise the evidence_index amt_days fallback path (lines 323-329).

    The inv: lookup misses, so the writer computes the amt_days signature and
    looks it up. Here the signature matches a target row, so the hyperlink cell
    is emitted.
    """
    bb = pd.DataFrame(
        [
            _bb_row(
                invoice="UNKNOWN-INVOICE",
                bill_date="01 Jan 2024",
                period_from="01 Jan 2022",
                period_to="31 Dec 2023",
                days_billed=730,
                period_charge=1347.96,
                excess_days=365,
            )
        ]
    )
    evidence_index = {"amt_days:1347.96|730": 42}
    ws = _open_ws()
    write_back_billing_sheet(ws, bb, account="A1", evidence_index=evidence_index)
    cell = ws.cell(row=8, column=14)
    assert cell.value == "\u2192"
    assert cell.hyperlink is not None
    assert cell.hyperlink.location == "'EDF Evidence Report'!A42"


def test_write_back_billing_sheet_evidence_index_inv_lookup_resolves_directly() -> None:
    """Cover lines 322-323, 332-340 (inv: lookup hits on the first try).

    A direct inv: match must short-circuit the amt_days fallback and emit the
    hyperlink pointing at the resolved evidence-report row.
    """
    bb = pd.DataFrame(
        [
            _bb_row(
                invoice="KI-DIRECT",
                bill_date="01 Jan 2024",
                period_from="01 Jan 2022",
                period_to="31 Dec 2023",
                days_billed=730,
                period_charge=100.0,
                excess_days=365,
            )
        ]
    )
    evidence_index = {"inv:KI-DIRECT": 7}
    ws = _open_ws()
    write_back_billing_sheet(ws, bb, account="A1", evidence_index=evidence_index)
    cell = ws.cell(row=8, column=14)
    assert cell.value == "\u2192"
    assert cell.hyperlink is not None
    assert cell.hyperlink.location == "'EDF Evidence Report'!A7"


def test_write_back_billing_sheet_no_evidence_index_emits_no_match() -> None:
    """Cover line 322 false branch (evidence_index is None -> 'No match')."""
    bb = pd.DataFrame(
        [
            _bb_row(
                invoice="KI-NOIDX",
                bill_date="01 Jan 2024",
                period_from="01 Jan 2022",
                period_to="31 Dec 2023",
            )
        ]
    )
    ws = _open_ws()
    write_back_billing_sheet(ws, bb, account="A1", evidence_index=None)
    assert ws.cell(row=8, column=14).value == "No match"


def test_write_back_billing_sheet_excess_days_over_30_highlights_red() -> None:
    """Cover lines 313-314 (Excess Days > 30 -> bold red font on column 9)."""
    bb = pd.DataFrame(
        [
            _bb_row(
                invoice="KI-BIG",
                bill_date="01 Jan 2024",
                period_from="01 Jan 2022",
                period_to="31 Dec 2023",
                days_billed=730,
                period_charge=100.0,
                excess_days=365,  # > 30
            )
        ]
    )
    ws = _open_ws()
    write_back_billing_sheet(ws, bb, account="A1")
    font = ws.cell(row=8, column=9).font
    assert font.bold is True
    assert font.color is not None
    # openpyxl renders the C00000 colour as an 8-char ARGB string ending in
    # C00000 (alpha-prefixed). Assert the suffix so the test is robust to the
    # alpha-channel formatting openpyxl applies.
    assert font.color.rgb.endswith("C00000")


def test_write_back_billing_sheet_excess_days_under_30_no_red_highlight() -> None:
    """Cover the false branch of line 313 (Excess Days <= 30 -> no red font)."""
    bb = pd.DataFrame(
        [
            _bb_row(
                invoice="KI-SMALL",
                bill_date="01 Jan 2024",
                period_from="01 Jan 2023",
                period_to="31 Jan 2024",
                days_billed=395,
                period_charge=100.0,
                excess_days=30,  # exactly 30, not > 30
            )
        ]
    )
    ws = _open_ws()
    write_back_billing_sheet(ws, bb, account="A1")
    font = ws.cell(row=8, column=9).font
    # The default _num cell font is not the red bold one.
    assert not (
        font.bold is True and font.color is not None and font.color.rgb in ("FFC00000", "C00000")
    )


def test_write_back_billing_sheet_account_in_title_banner() -> None:
    """Cover lines 218-219 (account truthy -> title carries the account)."""
    bb = pd.DataFrame(columns=_EXPECTED_COLUMNS)
    ws = _open_ws()
    write_back_billing_sheet(ws, bb, account="A-9999")
    title = ws.cell(row=1, column=1).value
    assert isinstance(title, str)
    assert "A-9999" in title
    assert "BACK-BILLING" in title.upper()


def test_write_back_billing_sheet_no_account_keeps_plain_title() -> None:
    """Cover line 218 false branch (account empty -> plain title)."""
    bb = pd.DataFrame(columns=_EXPECTED_COLUMNS)
    ws = _open_ws()
    write_back_billing_sheet(ws, bb, account="")
    title = ws.cell(row=1, column=1).value
    assert isinstance(title, str)
    assert title == "BACK-BILLING EVENTS ANALYSIS"
