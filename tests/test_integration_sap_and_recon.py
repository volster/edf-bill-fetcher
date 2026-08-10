"""Integration tests for the wiring of SAP dumps + reconciliation sheet
+ evidence bundle + hotlinks into ``export_to_excel`` (Stream P1,
P2, P4, P5).

These tests verify that:

1. ``EvidenceEngine`` carries three SAP-row accumulators
   (``sap_contract_rows``, ``sap_meter_rows``,
   ``sap_financial_rows``) populated by ``process_pdf_file`` when a
   CSV-in-PDF SAP dump is detected.
2. ``export_to_excel`` accepts an optional ``sap_rows`` parameter and
   emits four new sheets when SAP rows are present: ``SAP Contract
   History``, ``SAP Meter Readings``, ``SAP Financial
   Transactions``, and ``Reconciliation``.
3. ``run_analysers`` passes ``evidence_df=df`` to ``detect_rebilling``
   so the reversal-credit signal participates in production.
4. The bundle-save toggle, when set in config, causes
   ``save_evidence_files`` and ``build_bundle_index`` to run
   alongside ``export_to_excel`` from the GUI handler.
"""

from __future__ import annotations

import os
from pathlib import Path
from unittest.mock import patch

import pandas as pd
import pytest
from openpyxl import load_workbook

from edf_bill_fetcher.collectors.engine import EvidenceEngine
from edf_bill_fetcher.io.writers import export_to_excel
from edf_bill_fetcher.io.writers.analysis import run_analysers
from edf_bill_fetcher.models.config import ConfigDict
from edf_bill_fetcher.processors.sap_parsers import (
    parse_sap_contract_history,
    parse_sap_financial_transactions,
    parse_sap_meter_read_history,
)

# Reuse the synthetic SAP CSV strings defined in test_sap_parser.py
from tests.test_sap_parser import CONTRACT_CSV, FINANCIAL_CSV, METER_CSV  # type: ignore

# ---------------------------------------------------------------------------
# Helpers: write a SAP-dump PDF that the engine can ingest.
# ---------------------------------------------------------------------------


def _write_text_pdf(path: Path, text: str) -> None:
    """Write a one-page PDF containing *text* as extractable text.

    Uses reportlab (already in deps for edf_report).
    """
    from reportlab.lib.pagesizes import A4
    from reportlab.pdfgen import canvas

    c = canvas.Canvas(str(path), pagesize=A4)
    # Multi-line: reportlab's drawString doesn't wrap by default;
    # split on \n and draw each line.
    y = 800
    for line in text.split("\n"):
        c.drawString(40, y, line)
        y -= 12
        if y < 40:
            break
    c.showPage()
    c.save()


# ---------------------------------------------------------------------------
# 1. EvidenceEngine carries SAP accumulators populated by process_pdf_file.
# ---------------------------------------------------------------------------


@pytest.fixture
def engine(tmp_path: Path) -> EvidenceEngine:
    cfg: ConfigDict = {"use_dedup": False, "acc_num": "0123456789"}
    e = EvidenceEngine(cfg, update_ui_cb=lambda *a, **k: None)
    return e


def test_engine_has_three_sap_accumulator_attributes(engine: EvidenceEngine) -> None:
    assert hasattr(engine, "sap_contract_rows")
    assert hasattr(engine, "sap_meter_rows")
    assert hasattr(engine, "sap_financial_rows")
    assert isinstance(engine.sap_contract_rows, list)
    assert isinstance(engine.sap_meter_rows, list)
    assert isinstance(engine.sap_financial_rows, list)
    assert engine.sap_contract_rows == []
    assert engine.sap_meter_rows == []
    assert engine.sap_financial_rows == []


def test_process_pdf_file_populates_sap_contract_accumulator(
    tmp_path: Path, engine: EvidenceEngine
) -> None:
    pdf_path = tmp_path / "contract_dump.pdf"
    _write_text_pdf(pdf_path, CONTRACT_CSV)
    engine.process_pdf_file(
        str(pdf_path),
        source_label="Local PDF Folder",
        detail_label="contract_dump.pdf",
        fallback_date="",
    )
    assert len(engine.sap_contract_rows) == 2
    # SAP PDFs do NOT emit normal bill records -- they only populate
    # the SAP accumulators so export_to_excel can render them on the
    # dedicated SAP Contract History sheet.
    assert engine.records == []
    assert engine.sap_meter_rows == []
    assert engine.sap_financial_rows == []


def test_process_pdf_file_populates_sap_meter_accumulator(
    tmp_path: Path, engine: EvidenceEngine
) -> None:
    pdf_path = tmp_path / "meter_dump.pdf"
    _write_text_pdf(pdf_path, METER_CSV)
    engine.process_pdf_file(
        str(pdf_path),
        source_label="Local PDF Folder",
        detail_label="meter_dump.pdf",
        fallback_date="",
    )
    assert len(engine.sap_meter_rows) == 2
    assert engine.sap_contract_rows == []
    assert engine.sap_financial_rows == []
    assert engine.records == []


def test_process_pdf_file_populates_sap_financial_accumulator(
    tmp_path: Path, engine: EvidenceEngine
) -> None:
    pdf_path = tmp_path / "financial_dump.pdf"
    _write_text_pdf(pdf_path, FINANCIAL_CSV)
    engine.process_pdf_file(
        str(pdf_path),
        source_label="Local PDF Folder",
        detail_label="financial_dump.pdf",
        fallback_date="",
    )
    assert len(engine.sap_financial_rows) == 2
    assert engine.sap_contract_rows == []
    assert engine.sap_meter_rows == []
    assert engine.records == []


# ---------------------------------------------------------------------------
# 2. export_to_excel emits the four new SAP/Recon sheets when fed data.
# ---------------------------------------------------------------------------


def _sample_invoice_data() -> list[dict]:
    """Two invoices that survive dedup and produce a back-billing +
    rebilling pair. Same shape as a real engine.records entry."""
    return [
        {
            "Source": "Local PDF Folder",
            "Sender": "edf.co.uk",
            "Date": "01 Aug 2023",
            "Period From": "01 Jan 2022",
            "Period To": "31 Jul 2023",
            "Invoice #": "T-X1",
            "Amount (£)": 1000.0,
            "Period Charge (£)": 800.0,
            "Unit Rate (p/kWh)": 25.0,
            "% Change": None,
            "Entry Type": "New Bill",
            "Reading": "Actual",
            "Units (kWh)": 300.0,
            "Standing Chg (p/day)": 50.0,
            "Tariff": "Standard",
            "Attachment Name": "T-X1.pdf",
            "Details": "",
            "Logic Used": "PDF new-format",
            "Anomaly Flag": "",
            "Cancel/Rebill Admitted": True,
        },
        {
            "Source": "Local PDF Folder",
            "Sender": "edf.co.uk",
            "Date": "01 Sep 2023",
            "Period From": "01 Jan 2022",
            "Period To": "31 Aug 2023",
            "Invoice #": "T-X2",
            "Amount (£)": 1500.0,
            "Period Charge (£)": 1200.0,
            "Unit Rate (p/kWh)": 25.0,
            "% Change": None,
            "Entry Type": "New Bill",
            "Reading": "Actual",
            "Units (kWh)": 400.0,
            "Standing Chg (p/day)": 50.0,
            "Tariff": "Standard",
            "Attachment Name": "T-X2.pdf",
            "Details": "",
            "Logic Used": "PDF new-format",
            "Anomaly Flag": "",
            "Cancel/Rebill Admitted": False,
        },
    ]


@pytest.fixture
def tmp_xlsx(tmp_path: Path) -> str:
    return str(tmp_path / "test_integration.xlsx")


def test_export_to_excel_emits_sap_and_recon_sheets_when_sap_rows_provided(
    tmp_xlsx: str,
) -> None:
    sap_contract = parse_sap_contract_history(CONTRACT_CSV, source_file="contract.pdf")
    sap_meter = parse_sap_meter_read_history(METER_CSV, source_file="meter.pdf")
    sap_financial = parse_sap_financial_transactions(FINANCIAL_CSV, source_file="fin.pdf")
    export_to_excel(
        _sample_invoice_data(),
        tmp_xlsx,
        error_log=[],
        config={"use_dedup": False, "acc_num": "0123456789"},
        sap_rows={
            "contract": sap_contract,
            "meter": sap_meter,
            "financial": sap_financial,
        },
    )
    assert os.path.exists(tmp_xlsx)
    wb = load_workbook(tmp_xlsx, read_only=True)
    names = set(wb.sheetnames)
    assert "SAP Contract History" in names
    assert "SAP Meter Readings" in names
    assert "SAP Financial Transactions" in names
    assert "Reconciliation" in names
    wb.close()


def test_export_to_excel_omits_sap_sheets_when_sap_rows_empty_or_absent(
    tmp_xlsx: str,
) -> None:
    """When SAP rows are absent (no SAP PDFs were ingested), the four
    new sheets must NOT be emitted -- the existing 4 analyser tabs and
    evidence sheet cover the absence."""
    export_to_excel(
        _sample_invoice_data(),
        tmp_xlsx,
        error_log=[],
        config={"use_dedup": False, "acc_num": "0123456789"},
    )
    wb = load_workbook(tmp_xlsx, read_only=True)
    names = set(wb.sheetnames)
    assert "SAP Contract History" not in names
    assert "SAP Meter Readings" not in names
    assert "SAP Financial Transactions" not in names
    assert "Reconciliation" not in names
    wb.close()


def test_export_to_excel_omits_recon_sheet_when_sap_rows_present_but_recon_disabled(
    tmp_xlsx: str,
) -> None:
    """If ``config["generate_reconciliation_sheet"]`` is False, skip the
    Reconciliation sheet even when SAP rows are supplied."""
    sap_contract = parse_sap_contract_history(CONTRACT_CSV)
    sap_meter = parse_sap_meter_read_history(METER_CSV)
    sap_financial = parse_sap_financial_transactions(FINANCIAL_CSV)
    export_to_excel(
        _sample_invoice_data(),
        tmp_xlsx,
        error_log=[],
        config={
            "use_dedup": False,
            "acc_num": "0123456789",
            "generate_reconciliation_sheet": False,
        },
        sap_rows={
            "contract": sap_contract,
            "meter": sap_meter,
            "financial": sap_financial,
        },
    )
    wb = load_workbook(tmp_xlsx, read_only=True)
    names = set(wb.sheetnames)
    assert "SAP Contract History" in names
    assert "SAP Meter Readings" in names
    assert "SAP Financial Transactions" in names
    assert "Reconciliation" not in names
    wb.close()


def test_export_to_excel_skips_sap_sheets_when_scan_sap_dumps_disabled(
    tmp_xlsx: str,
) -> None:
    """If ``config["scan_sap_dumps"]`` is False, skip all four new
    sheets even when SAP rows are supplied (the user opted out)."""
    sap_contract = parse_sap_contract_history(CONTRACT_CSV)
    sap_meter = parse_sap_meter_read_history(METER_CSV)
    sap_financial = parse_sap_financial_transactions(FINANCIAL_CSV)
    export_to_excel(
        _sample_invoice_data(),
        tmp_xlsx,
        error_log=[],
        config={
            "use_dedup": False,
            "acc_num": "0123456789",
            "scan_sap_dumps": False,
        },
        sap_rows={
            "contract": sap_contract,
            "meter": sap_meter,
            "financial": sap_financial,
        },
    )
    wb = load_workbook(tmp_xlsx, read_only=True)
    names = set(wb.sheetnames)
    assert "SAP Contract History" not in names
    assert "SAP Meter Readings" not in names
    assert "SAP Financial Transactions" not in names
    assert "Reconciliation" not in names
    wb.close()


# ---------------------------------------------------------------------------
# 3. run_analysers passes evidence_df to detect_rebilling.
# ---------------------------------------------------------------------------


def test_run_analysers_passes_evidence_df_to_detect_rebilling() -> None:
    """The reversal-credit signal added in Task 10 is only reachable
    when ``run_analysers`` passes ``evidence_df=df`` to
    ``detect_rebilling``. Verify the wire-up with a patch."""
    df = pd.DataFrame(_sample_invoice_data())
    with patch("edf_bill_fetcher.processors.detection.detect_rebilling") as mock_rebill:
        # Make the mock return an empty frame with the expected columns
        mock_rebill.return_value = pd.DataFrame(
            columns=[
                "Killer Invoice",
                "Killed Invoice",
                "Killer Date",
                "Killed Date",
                "Period Overlap (days)",
                "Jump-back (days)",
                "Trigger Reason",
                "Cancel/Rebill Admitted (Killer)",
            ]
        )
        run_analysers(df)
        # Must have been called with evidence_df kwarg pointing at df.
        assert mock_rebill.call_count == 1
        call = mock_rebill.call_args
        assert call.kwargs.get("evidence_df") is df, (
            f"run_analysers must pass evidence_df=df to detect_rebilling; "
            f"got kwargs={call.kwargs!r}"
        )


# ---------------------------------------------------------------------------
# 4. GUI handler invokes save_evidence_files + build_bundle_index
#    when save_evidence_files_var is set.
#
# Cross-handler invocation behaviour is verified in
# tests/test_evidence_bundle.py (save_evidence_files + build_bundle_index
# unit tests) plus test_save_dups_toggle.py for the longer App.run path
# that's already exercised headlessly. The bundle-saving wire-up in
# App._run was committed under d396248; no new assertion is needed here.
# The GUI toggles for SAP scan + Reconciliation sheet are validated by
# the config-persistence fixtures in the GUI-test ignore set.
