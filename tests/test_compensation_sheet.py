"""TDD tests for the Compensation evidence sheet writer.

Covers ``write_compensation_sheet`` in ``io/writers/compensation.py``:
one row per claim grouped by category under label rows, a trailing total
of ``indicative_amount``, a DISCLAIMER row, no-op on empty/None rows, and
the ``export_to_excel`` wiring that emits a ``Compensation`` tab when the
estimator yields rows.
"""

from __future__ import annotations

from typing import Any, cast

import pandas as pd
from openpyxl import Workbook, load_workbook

from edf_bill_fetcher.io.writers.compensation import write_compensation_sheet
from edf_bill_fetcher.io.writers.export import export_to_excel
from edf_bill_fetcher.models.config import ConfigDict
from edf_bill_fetcher.processors.compensation import DISCLAIMER

CATEGORY_LABELS = {
    "back_billing_excess": "BACK-BILLING EXCESS",
    "credit_hold_interest": "CREDIT-HOLD INTEREST",
    "late_credit_interest": "LATE-CREDIT INTEREST",
}


def _row(
    category: str,
    invoice_ref: str,
    date: str,
    base_amount: float,
    days: int,
    rate: float | None,
    indicative_amount: float,
) -> dict[str, Any]:
    return {
        "category": category,
        "invoice_ref": invoice_ref,
        "date": date,
        "base_amount": base_amount,
        "days": days,
        "rate": rate,
        "indicative_amount": indicative_amount,
        "legal_basis": "Legal basis",
        "disclaimer": DISCLAIMER,
    }


def _sample_rows() -> list[dict[str, Any]]:
    return [
        _row("back_billing_excess", "KI-0001", "2024-03-01", 1200.00, 425, None, 647.21),
        _row("credit_hold_interest", "KCR-0001", "2026-02-01", 100.00, 30, 0.02, 0.16),
        _row("late_credit_interest", "KCR-0001", "2026-02-01", 100.00, 120, 0.02, 0.66),
    ]


def _collect_text(ws: Any) -> list[str]:
    """Flatten every non-empty cell value in the sheet to a list of strings."""
    out: list[str] = []
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                out.append(str(cell.value))
    return out


# ---------- one row per claim, grouped by category ----------


def test_renders_one_row_per_claim_grouped_by_category() -> None:
    """Each claim becomes a data row under its category's label row."""
    wb = Workbook()
    ws = wb.active
    write_compensation_sheet(ws, _sample_rows())

    texts = _collect_text(ws)
    # Category label rows present.
    for label in CATEGORY_LABELS.values():
        assert label in texts
    # Each invoice_ref appears exactly once (one row per claim).
    assert texts.count("KI-0001") == 1
    assert texts.count("KCR-0001") == 2  # hold + late rows share the invoice
    # Header row present.
    assert "Indicative Amount (£)" in texts


def test_trailing_total_equals_sum_of_indicative_amount() -> None:
    """The trailing total row sums every claim's indicative_amount."""
    wb = Workbook()
    ws = wb.active
    write_compensation_sheet(ws, _sample_rows())

    expected = round(sum(r["indicative_amount"] for r in _sample_rows()), 2)
    texts = _collect_text(ws)
    assert any("TOTAL" in t for t in texts)
    # The total value cell is present as a numeric cell.
    total_cells = [
        cell.value
        for row in ws.iter_rows()
        for cell in row
        if isinstance(cell.value, (int, float)) and cell.value == expected
    ]
    assert total_cells, f"expected a cell holding the total {expected}"


def test_disclaimer_row_present() -> None:
    """The estimator's DISCLAIMER constant is written verbatim."""
    wb = Workbook()
    ws = wb.active
    write_compensation_sheet(ws, _sample_rows())
    assert DISCLAIMER in _collect_text(ws)


# ---------- empty / None rows -> no-op ----------


def test_empty_rows_no_op() -> None:
    """An empty row list writes no data rows and does not crash."""
    wb = Workbook()
    ws = wb.active
    write_compensation_sheet(ws, [])
    # No header row, no total, no disclaimer.
    assert _collect_text(ws) == []


def test_none_rows_no_op() -> None:
    """A None row list writes no data rows and does not crash."""
    wb = Workbook()
    ws = wb.active
    write_compensation_sheet(ws, None)  # type: ignore[arg-type]
    assert _collect_text(ws) == []


# ---------- export wiring ----------


def test_export_to_excel_emits_compensation_tab(tmp_path: object) -> None:
    """A synthetic record set that yields compensation rows emits a
    Compensation tab in the exported workbook."""
    df = pd.DataFrame(
        [
            {
                "Date": "2024-03-01",
                "Amount (£)": 1200.00,
                "Entry Type": "New Bill",
                "Invoice #": "KI-0001",
                "Period From": "01/01/2022",
                "Period To": "28/02/2024",
                "Source": "HTM Account History",
                "Period Charge (£)": 1200.00,
                "Units (kWh)": 500,
            },
            {
                "Date": "2026-02-01",
                "Amount (£)": -100.00,
                "Entry Type": "Credit",
                "Invoice #": "KCR-0001",
                "Period From": "01/01/2026",
                "Period To": "31/01/2026",
                "Source": "HTM Account History",
                "Period Charge (£)": 0.00,
                "Units (kWh)": 0,
            },
        ]
    )
    out = tmp_path / "comp.xlsx"  # type: ignore[operator]
    config = cast(
        ConfigDict,
        {
            "use_dedup": True,
            "use_back_billing": False,
            "use_reconciliation": False,
            "analysis_min": 0,
            "save_filtered": True,
            "use_sap": False,
            "as_of": "2026-06-01",
        },
    )
    export_to_excel(df, str(out), error_log=[], config=config)
    wb = load_workbook(out)
    assert "Compensation" in wb.sheetnames
    ws = wb["Compensation"]
    texts = _collect_text(ws)
    assert DISCLAIMER in texts
    assert any("TOTAL" in t for t in texts)
